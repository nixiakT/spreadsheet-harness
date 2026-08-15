from __future__ import annotations

import hashlib
import os
import warnings
from collections.abc import Callable
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

import pytest
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as SpreadsheetImage
from openpyxl.formatting.rule import ColorScale, FormatObject, Rule
from openpyxl.styles import Color
from openpyxl.worksheet.table import Table
from PIL import Image as PILImage

import spreadsheet_harness.ooxml_formula_scan as formula_scan_module
from spreadsheet_harness.ooxml_formula_scan import (
    OOXMLFormulaScanError,
    scan_ooxml_formulas,
)
from spreadsheet_harness.render import find_libreoffice, recalculate_workbook


def _workbook(path: Path, *, formula: str | None = None) -> None:
    workbook = Workbook()
    workbook.active.title = "Data"
    workbook.active["A1"] = 1
    workbook.active["B1"] = formula
    workbook.save(path)
    workbook.close()


def _rewrite_package(
    source: Path,
    destination: Path,
    transform: Callable[[str, bytes], bytes],
) -> None:
    with ZipFile(source) as archive, ZipFile(destination, "w", ZIP_DEFLATED) as output:
        for info in archive.infolist():
            payload = archive.read(info)
            rewritten = ZipInfo(info.filename, date_time=info.date_time)
            rewritten.compress_type = ZIP_DEFLATED
            rewritten.external_attr = info.external_attr
            rewritten.flag_bits = info.flag_bits
            output.writestr(rewritten, transform(info.filename, payload))


def test_formula_scan_certifies_plain_xlsx_without_changing_bytes(tmp_path: Path) -> None:
    path = tmp_path / "plain.xlsx"
    _workbook(path)
    before = path.read_bytes()

    result = scan_ooxml_formulas(path)

    assert path.read_bytes() == before
    assert result.package_sha256 == hashlib.sha256(before).hexdigest()
    assert result.workbook_format == "xlsx"
    assert result.worksheet_count == 1
    assert result.scanned_cell_count == 1
    assert result.formula_marker_count == 0
    assert result.formula_kinds == ()
    assert result.has_formulas is False


def test_formula_scan_accepts_exact_table_comment_drawing_and_image_contracts(
    tmp_path: Path,
) -> None:
    path = tmp_path / "rich-but-static.xlsx"
    image_path = tmp_path / "pixel.png"
    PILImage.new("RGB", (2, 2), "white").save(image_path)
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Name", "Value"])
    sheet.append(["A", 1])
    sheet.add_table(Table(displayName="StaticTable", ref="A1:B2"))
    sheet["A1"].comment = Comment("static note", "reviewer")
    sheet.add_image(SpreadsheetImage(image_path), "D1")
    workbook.save(path)
    workbook.close()

    result = scan_ooxml_formulas(path)

    assert result.has_formulas is False


@pytest.mark.parametrize("formula", ["=1+1", "=SUM(A1:A2)"])
def test_formula_scan_detects_cell_formulas(tmp_path: Path, formula: str) -> None:
    path = tmp_path / "formula.xlsx"
    _workbook(path, formula=formula)

    result = scan_ooxml_formulas(path)

    assert result.has_formulas is True
    assert result.formula_marker_count == 1
    assert result.formula_kinds == ("f",)


def test_formula_scan_treats_defined_name_expression_as_recalculation_bearing(
    tmp_path: Path,
) -> None:
    original = tmp_path / "original.xlsx"
    rewritten = tmp_path / "defined-name.xlsx"
    _workbook(original)

    def add_defined_name(name: str, payload: bytes) -> bytes:
        if name != "xl/workbook.xml":
            return payload
        declaration = (
            b'<definedNames><definedName name="Result">'
            b"SUM(Data!$A$1:$A$2)"
            b"</definedName></definedNames>"
        )
        return payload.replace(b"<calcPr", declaration + b"<calcPr", 1)

    _rewrite_package(original, rewritten, add_defined_name)

    result = scan_ooxml_formulas(rewritten)

    assert result.has_formulas is True
    assert "definedname-expression" in result.formula_kinds


@pytest.mark.parametrize("formula_type", ["shared", "array", "dataTable"])
def test_formula_scan_detects_shared_array_and_data_table_formulas(
    tmp_path: Path,
    formula_type: str,
) -> None:
    original = tmp_path / "original.xlsx"
    rewritten = tmp_path / f"{formula_type}.xlsx"
    _workbook(original, formula="=1+1")

    def set_formula_type(name: str, payload: bytes) -> bytes:
        if name != "xl/worksheets/sheet1.xml":
            return payload
        return payload.replace(b"<f>", f'<f t="{formula_type}">'.encode("ascii"), 1)

    _rewrite_package(original, rewritten, set_formula_type)

    result = scan_ooxml_formulas(rewritten)

    assert result.has_formulas is True
    assert result.formula_kinds == ("f",)


@pytest.mark.parametrize(
    ("declaration", "formula_kind"),
    [
        (
            b'<conditionalFormatting sqref="A1"><cfRule type="expression" '
            b'priority="1"><formula>A1&gt;0</formula></cfRule></conditionalFormatting>',
            "formula",
        ),
        (
            b'<dataValidations count="1"><dataValidation type="custom" sqref="A1">'
            b"<formula1>A1&gt;0</formula1></dataValidation></dataValidations>",
            "formula1",
        ),
    ],
)
def test_formula_scan_detects_conditional_formatting_and_validation_formulas(
    tmp_path: Path,
    declaration: bytes,
    formula_kind: str,
) -> None:
    original = tmp_path / "original.xlsx"
    rewritten = tmp_path / f"{formula_kind}.xlsx"
    _workbook(original)

    _rewrite_package(
        original,
        rewritten,
        lambda name, payload: (
            payload.replace(b"</worksheet>", declaration + b"</worksheet>", 1)
            if name == "xl/worksheets/sheet1.xml"
            else payload
        ),
    )

    result = scan_ooxml_formulas(rewritten)

    assert result.has_formulas is True
    assert formula_kind in result.formula_kinds


def test_formula_scan_detects_conditional_format_threshold_formula(
    tmp_path: Path,
) -> None:
    path = tmp_path / "conditional-threshold-formula.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.conditional_formatting.add(
        "A1:A5",
        Rule(
            type="colorScale",
            colorScale=ColorScale(
                cfvo=[
                    FormatObject(type="formula", val="A1"),
                    FormatObject(type="max"),
                ],
                color=[Color("FF0000"), Color("00FF00")],
            ),
        ),
    )
    workbook.save(path)
    workbook.close()

    result = scan_ooxml_formulas(path)

    assert result.has_formulas is True
    assert "cfvo-formula-value" in result.formula_kinds


def test_formula_scan_detects_table_calculated_column_formula(tmp_path: Path) -> None:
    original = tmp_path / "original.xlsx"
    rewritten = tmp_path / "table-formula.xlsx"
    table_part = "xl/tables/table1.xml"
    _workbook(original)

    def declare_table(name: str, payload: bytes) -> bytes:
        if name != "[Content_Types].xml":
            return payload
        declaration = (
            f'<Override PartName="/{table_part}" ContentType="application/vnd.'
            'openxmlformats-officedocument.spreadsheetml.table+xml"/>'
        ).encode("ascii")
        return payload.replace(b"</Types>", declaration + b"</Types>", 1)

    _rewrite_package(original, rewritten, declare_table)
    table_xml = (
        b'<table xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        b'id="1" name="Table1" displayName="Table1" ref="A1:A1">'
        b'<tableColumns count="1"><tableColumn id="1" name="Value">'
        b"<calculatedColumnFormula>1+1</calculatedColumnFormula>"
        b"</tableColumn></tableColumns></table>"
    )
    with ZipFile(rewritten, "a", ZIP_DEFLATED) as archive:
        archive.writestr(table_part, table_xml)

    result = scan_ooxml_formulas(rewritten)

    assert result.has_formulas is True
    assert "calculatedcolumnformula" in result.formula_kinds


def test_formula_scan_accepts_byte_preserving_macro_enabled_container(tmp_path: Path) -> None:
    original = tmp_path / "original.xlsx"
    macro = tmp_path / "plain.xlsm"
    _workbook(original)

    def mark_macro_enabled(name: str, payload: bytes) -> bytes:
        if name != "[Content_Types].xml":
            return payload
        return payload.replace(
            b"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml",
            b"application/vnd.ms-excel.sheet.macroEnabled.main+xml",
            1,
        )

    _rewrite_package(original, macro, mark_macro_enabled)

    result = scan_ooxml_formulas(macro)

    assert result.workbook_format == "xlsm"
    assert result.has_formulas is False


def test_formula_scan_allows_explicit_media_payloads(tmp_path: Path) -> None:
    original = tmp_path / "original.xlsx"
    macro = tmp_path / "plain-with-binaries.xlsm"
    _workbook(original)

    def declare_allowed_parts(name: str, payload: bytes) -> bytes:
        if name != "[Content_Types].xml":
            return payload
        payload = payload.replace(
            b"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml",
            b"application/vnd.ms-excel.sheet.macroEnabled.main+xml",
            1,
        )
        declarations = b'<Override PartName="/xl/media/image1.png" ContentType="image/png"/>'
        return payload.replace(b"</Types>", declarations + b"</Types>", 1)

    _rewrite_package(original, macro, declare_allowed_parts)
    with ZipFile(macro, "a", ZIP_DEFLATED) as archive:
        archive.writestr("xl/media/image1.png", b"test-image")

    result = scan_ooxml_formulas(macro)

    assert result.workbook_format == "xlsm"
    assert result.has_formulas is False


def test_formula_scan_rejects_vba_project_even_with_exact_content_type(tmp_path: Path) -> None:
    original = tmp_path / "original.xlsx"
    macro = tmp_path / "macro.xlsm"
    _workbook(original)

    def declare_vba(name: str, payload: bytes) -> bytes:
        if name != "[Content_Types].xml":
            return payload
        payload = payload.replace(
            b"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml",
            b"application/vnd.ms-excel.sheet.macroEnabled.main+xml",
            1,
        )
        declaration = (
            b'<Override PartName="/xl/vbaProject.bin" '
            b'ContentType="application/vnd.ms-office.vbaProject"/>'
        )
        return payload.replace(b"</Types>", declaration + b"</Types>", 1)

    _rewrite_package(original, macro, declare_vba)
    with ZipFile(macro, "a", ZIP_DEFLATED) as archive:
        archive.writestr("xl/vbaProject.bin", b"opaque executable project")

    with pytest.raises(OOXMLFormulaScanError, match="formula-capable part"):
        scan_ooxml_formulas(macro)


def test_formula_scan_rejects_table_relationship_disguised_as_png(tmp_path: Path) -> None:
    original = tmp_path / "original.xlsx"
    rewritten = tmp_path / "fake-table-image.xlsx"
    _workbook(original)
    relationship_part = "xl/worksheets/_rels/sheet1.xml.rels"
    fake_part = "xl/media/fake.png"
    relationships = (
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rIdFake" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/table" '
        'Target="../media/fake.png"/>'
        "</Relationships>"
    ).encode("ascii")

    def declare_fake_table(name: str, payload: bytes) -> bytes:
        if name == "[Content_Types].xml":
            declaration = (f'<Override PartName="/{fake_part}" ContentType="image/png"/>').encode(
                "ascii"
            )
            return payload.replace(b"</Types>", declaration + b"</Types>", 1)
        return payload

    _rewrite_package(original, rewritten, declare_fake_table)
    with ZipFile(rewritten, "a", ZIP_DEFLATED) as archive:
        archive.writestr(relationship_part, relationships)
        archive.writestr(
            fake_part,
            b"<table><calculatedColumnFormula>1+1</calculatedColumnFormula></table>",
        )

    with pytest.raises(OOXMLFormulaScanError, match="target path|content type"):
        scan_ooxml_formulas(rewritten)


@pytest.mark.parametrize(
    ("variant", "expected_error"),
    [
        ("content-type", "target content type"),
        ("xml-root", "target XML root"),
        ("external-mode", "external target mode"),
    ],
)
def test_formula_scan_binds_table_uri_to_mode_content_type_and_root(
    tmp_path: Path,
    variant: str,
    expected_error: str,
) -> None:
    original = tmp_path / "original.xlsx"
    rewritten = tmp_path / f"bad-table-{variant}.xlsx"
    _workbook(original)
    relationship_part = "xl/worksheets/_rels/sheet1.xml.rels"
    target_part = "xl/tables/table1.xml"
    target = (
        "https://invalid.example/table.xml"
        if variant == "external-mode"
        else "../tables/table1.xml"
    )
    target_mode = ' TargetMode="External"' if variant == "external-mode" else ""
    relationships = (
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rIdTable" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/table" '
        f'Target="{target}"{target_mode}/>'
        "</Relationships>"
    ).encode("ascii")

    def declare_target(name: str, payload: bytes) -> bytes:
        if name != "[Content_Types].xml" or variant == "external-mode":
            return payload
        content_type = (
            "image/png"
            if variant == "content-type"
            else "application/vnd.openxmlformats-officedocument.spreadsheetml.table+xml"
        )
        declaration = (
            f'<Override PartName="/{target_part}" ContentType="{content_type}"/>'
        ).encode("ascii")
        return payload.replace(b"</Types>", declaration + b"</Types>", 1)

    _rewrite_package(original, rewritten, declare_target)
    with ZipFile(rewritten, "a", ZIP_DEFLATED) as archive:
        archive.writestr(relationship_part, relationships)
        if variant != "external-mode":
            root = "worksheet" if variant == "xml-root" else "table"
            archive.writestr(
                target_part,
                (
                    f'<{root} xmlns="http://schemas.openxmlformats.org/'
                    f'spreadsheetml/2006/main"></{root}>'
                ).encode("ascii"),
            )

    with pytest.raises(OOXMLFormulaScanError, match=expected_error):
        scan_ooxml_formulas(rewritten)


def test_formula_scan_rejects_duplicate_parts_as_ambiguous(tmp_path: Path) -> None:
    path = tmp_path / "duplicate.xlsx"
    _workbook(path)
    with ZipFile(path) as archive:
        duplicate_payload = archive.read("xl/worksheets/sheet1.xml")
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        with ZipFile(path, "a", ZIP_DEFLATED) as archive:
            archive.writestr("xl/worksheets/sheet1.xml", duplicate_payload)

    with pytest.raises(OOXMLFormulaScanError, match="duplicate or aliased"):
        scan_ooxml_formulas(path)


def test_formula_scan_rejects_duplicate_cell_identity_as_ambiguous(tmp_path: Path) -> None:
    original = tmp_path / "original.xlsx"
    duplicate = tmp_path / "duplicate-cell.xlsx"
    _workbook(original)

    def duplicate_cell(name: str, payload: bytes) -> bytes:
        if name != "xl/worksheets/sheet1.xml":
            return payload
        return payload.replace(b"</row>", b'<c r="A1" t="n"><v>2</v></c></row>', 1)

    _rewrite_package(original, duplicate, duplicate_cell)

    with pytest.raises(OOXMLFormulaScanError, match="cell identity is ambiguous"):
        scan_ooxml_formulas(duplicate)


def test_formula_scan_rejects_non_zip_payload(tmp_path: Path) -> None:
    path = tmp_path / "not-a-package.xlsx"
    path.write_bytes(b"not an OOXML ZIP package")

    with pytest.raises(OOXMLFormulaScanError, match="Invalid OOXML ZIP"):
        scan_ooxml_formulas(path)


@pytest.mark.parametrize(
    "payload",
    [
        b"<worksheet>",
        b'<!DOCTYPE worksheet [<!ENTITY x "formula">]><worksheet>&x;</worksheet>',
    ],
)
def test_formula_scan_rejects_malformed_or_entity_bearing_xml(
    tmp_path: Path,
    payload: bytes,
) -> None:
    original = tmp_path / "original.xlsx"
    rewritten = tmp_path / "unsafe.xlsx"
    _workbook(original)
    _rewrite_package(
        original,
        rewritten,
        lambda name, current: payload if name == "xl/worksheets/sheet1.xml" else current,
    )

    with pytest.raises(OOXMLFormulaScanError, match="malformed|forbidden DTD"):
        scan_ooxml_formulas(rewritten)


def test_formula_scan_rejects_extension_content_type_mismatch(tmp_path: Path) -> None:
    xlsx = tmp_path / "plain.xlsx"
    mismatched = tmp_path / "plain.xlsm"
    _workbook(xlsx)
    mismatched.write_bytes(xlsx.read_bytes())

    with pytest.raises(OOXMLFormulaScanError, match="does not match"):
        scan_ooxml_formulas(mismatched)


def test_formula_scan_allows_static_defined_names_and_formula_ui_flags(
    tmp_path: Path,
) -> None:
    original = tmp_path / "original.xlsx"
    rewritten = tmp_path / "ui-flags.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet["A1"] = 1
    sheet.print_area = "A1:B3"
    sheet.print_title_rows = "1:1"
    workbook.save(original)
    workbook.close()

    def add_ui_flags(name: str, payload: bytes) -> bytes:
        if name == "xl/workbook.xml":
            return payload.replace(b"<workbookView ", b'<workbookView showFormulaBar="1" ', 1)
        if name == "xl/worksheets/sheet1.xml":
            payload = payload.replace(b"<sheetView ", b'<sheetView showFormulas="0" ', 1)
            ignored = (
                b'<ignoredErrors><ignoredError sqref="A1" formula="1" '
                b'formulaRange="1" unlockedFormula="1"/></ignoredErrors>'
            )
            return payload.replace(b"</worksheet>", ignored + b"</worksheet>", 1)
        return payload

    _rewrite_package(original, rewritten, add_ui_flags)

    result = scan_ooxml_formulas(rewritten)

    assert result.has_formulas is False
    assert result.formula_kinds == ()


def test_formula_attribute_ui_exemptions_require_spreadsheetml_context(
    tmp_path: Path,
) -> None:
    original = tmp_path / "original.xlsx"
    rewritten = tmp_path / "foreign-sheet-view.xlsx"
    _workbook(original)
    extension = (
        b'<extLst><ext uri="test"><x:sheetView xmlns:x="urn:foreign" formula="1+1"/></ext></extLst>'
    )

    _rewrite_package(
        original,
        rewritten,
        lambda name, payload: (
            payload.replace(b"</worksheet>", extension + b"</worksheet>", 1)
            if name == "xl/worksheets/sheet1.xml"
            else payload
        ),
    )

    result = scan_ooxml_formulas(rewritten)

    assert result.has_formulas is True
    assert "attribute:formula" in result.formula_kinds


def test_formula_scan_accepts_real_libreoffice_origin_formula_free_workbook(
    tmp_path: Path,
) -> None:
    if find_libreoffice() is None:
        pytest.skip("LibreOffice is unavailable")
    source = tmp_path / "source.xlsx"
    converted = tmp_path / "libreoffice-origin.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet["A1"] = "plain"
    sheet.print_area = "A1:B3"
    sheet.print_title_rows = "1:1"
    workbook.save(source)
    workbook.close()

    metadata = recalculate_workbook(source, converted, timeout_seconds=30.0)
    result = scan_ooxml_formulas(converted)

    assert metadata["backend"] == "libreoffice-headless"
    assert result.has_formulas is False


@pytest.mark.parametrize("encoding", ["utf-16", "utf-32"])
def test_formula_scan_rejects_encoding_independent_dtd_and_entities(
    tmp_path: Path,
    encoding: str,
) -> None:
    original = tmp_path / "original.xlsx"
    rewritten = tmp_path / f"unsafe-{encoding}.xlsx"
    _workbook(original)
    xml = (
        f'<?xml version="1.0" encoding="{encoding.upper()}"?>'
        '<!DOCTYPE worksheet [<!ENTITY x "formula">]>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        "&x;</worksheet>"
    ).encode(encoding)
    _rewrite_package(
        original,
        rewritten,
        lambda name, payload: xml if name == "xl/worksheets/sheet1.xml" else payload,
    )

    with pytest.raises(OOXMLFormulaScanError, match="forbidden DTD|malformed"):
        scan_ooxml_formulas(rewritten)


@pytest.mark.parametrize(
    ("part_name", "old_type"),
    [
        ("_rels/.rels", b"officeDocument"),
        ("xl/_rels/workbook.xml.rels", b"worksheet"),
    ],
)
def test_formula_scan_rejects_spoofed_relationship_type_uri(
    tmp_path: Path,
    part_name: str,
    old_type: bytes,
) -> None:
    original = tmp_path / "original.xlsx"
    rewritten = tmp_path / "spoofed-relationship.xlsx"
    _workbook(original)

    def spoof(name: str, payload: bytes) -> bytes:
        if name != part_name:
            return payload
        marker = b"http://schemas.openxmlformats.org/officeDocument/2006/relationships/" + old_type
        return payload.replace(marker, b"https://invalid.example/" + old_type, 1)

    _rewrite_package(original, rewritten, spoof)

    with pytest.raises(OOXMLFormulaScanError, match="office-document|relationship URI"):
        scan_ooxml_formulas(rewritten)


@pytest.mark.parametrize(
    ("relationship_part", "expected_error"),
    [("root", "root relationships"), ("workbook", "workbook relationships")],
)
def test_formula_scan_requires_exact_relationship_content_types(
    tmp_path: Path,
    relationship_part: str,
    expected_error: str,
) -> None:
    original = tmp_path / "original.xlsx"
    rewritten = tmp_path / f"bad-{relationship_part}-rels.xlsx"
    _workbook(original)

    def spoof(name: str, payload: bytes) -> bytes:
        if name != "[Content_Types].xml":
            return payload
        payload = payload.replace(
            b"application/vnd.openxmlformats-package.relationships+xml",
            b"application/xml",
            1,
        )
        if relationship_part == "workbook":
            override = (
                b'<Override PartName="/_rels/.rels" ContentType="application/vnd.'
                b'openxmlformats-package.relationships+xml"/>'
            )
            payload = payload.replace(b"</Types>", override + b"</Types>", 1)
        return payload

    _rewrite_package(original, rewritten, spoof)

    with pytest.raises(OOXMLFormulaScanError, match=expected_error):
        scan_ooxml_formulas(rewritten)


def test_formula_scan_rejects_worksheet_content_type_spoofing(tmp_path: Path) -> None:
    original = tmp_path / "original.xlsx"
    rewritten = tmp_path / "spoofed-content-type.xlsx"
    _workbook(original)

    def spoof(name: str, payload: bytes) -> bytes:
        if name != "[Content_Types].xml":
            return payload
        return payload.replace(
            b"application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml",
            b"application/octet-stream",
            1,
        )

    _rewrite_package(original, rewritten, spoof)

    with pytest.raises(OOXMLFormulaScanError, match="sheet content type"):
        scan_ooxml_formulas(rewritten)


@pytest.mark.parametrize("sheet_kind", ["chartsheet", "dialogsheet"])
def test_formula_scan_matches_nonworksheet_relationship_root_and_content_type(
    tmp_path: Path,
    sheet_kind: str,
) -> None:
    original = tmp_path / "original.xlsx"
    converted = tmp_path / f"clean-{sheet_kind}.xlsx"
    spoofed = tmp_path / f"spoofed-{sheet_kind}.xlsx"
    _workbook(original)

    worksheet_type = b"application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"
    replacement_type = (
        f"application/vnd.openxmlformats-officedocument.spreadsheetml.{sheet_kind}+xml"
    ).encode("ascii")
    source_part = "xl/worksheets/sheet1.xml"
    target_part = f"xl/{sheet_kind}s/sheet1.xml"

    def convert_sheet(name: str, payload: bytes) -> bytes:
        if name == "xl/_rels/workbook.xml.rels":
            return payload.replace(b'/worksheet"', f'/{sheet_kind}"'.encode("ascii"), 1).replace(
                source_part.encode("ascii"), target_part.encode("ascii"), 1
            )
        if name == "[Content_Types].xml":
            return payload.replace(worksheet_type, replacement_type, 1).replace(
                source_part.encode("ascii"), target_part.encode("ascii"), 1
            )
        if name == source_part:
            return payload.replace(b"<worksheet ", f"<{sheet_kind} ".encode("ascii"), 1).replace(
                b"</worksheet>", f"</{sheet_kind}>".encode("ascii"), 1
            )
        return payload

    with ZipFile(original) as archive, ZipFile(converted, "w", ZIP_DEFLATED) as output:
        for info in archive.infolist():
            output.writestr(
                target_part if info.filename == source_part else info.filename,
                convert_sheet(info.filename, archive.read(info)),
            )
    assert scan_ooxml_formulas(converted).has_formulas is False

    _rewrite_package(
        converted,
        spoofed,
        lambda name, payload: (
            payload.replace(replacement_type, b"application/octet-stream", 1)
            if name == "[Content_Types].xml"
            else payload
        ),
    )
    with pytest.raises(OOXMLFormulaScanError, match="sheet content type"):
        scan_ooxml_formulas(spoofed)


@pytest.mark.parametrize(
    ("part_name", "content_type"),
    [
        (
            "xl/embeddings/embeddedWorkbook.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ),
        ("xl/model/item.data", "application/vnd.ms-excel.model+data"),
        ("xl/externalLinks/externalLink1.xml", "application/xml"),
        (
            "xl/pivotTables/pivotTable1.xml",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.pivotTable+xml",
        ),
        ("xl/ctrlProps/ctrlProp1.xml", "application/xml"),
        ("XL/CTRLPROPS/case-variant.xml", "application/xml"),
        (
            "xl/misc/renamed-control.xml",
            "application/vnd.ms-excel.controlproperties+xml",
        ),
        ("xl/misc/renamed-vba.xml", "application/vnd.ms-office.vbaProject"),
    ],
)
def test_formula_scan_rejects_opaque_formula_capable_parts(
    tmp_path: Path,
    part_name: str,
    content_type: str,
) -> None:
    original = tmp_path / "original.xlsx"
    rewritten = tmp_path / "opaque.xlsx"
    _workbook(original)

    def declare_part(name: str, payload: bytes) -> bytes:
        if name != "[Content_Types].xml":
            return payload
        declaration = (f'<Override PartName="/{part_name}" ContentType="{content_type}"/>').encode(
            "ascii"
        )
        return payload.replace(b"</Types>", declaration + b"</Types>", 1)

    _rewrite_package(original, rewritten, declare_part)
    nested_payload = original.read_bytes() if part_name.endswith(".xlsx") else b"opaque"
    with ZipFile(rewritten, "a", ZIP_DEFLATED) as archive:
        archive.writestr(part_name, nested_payload)

    with pytest.raises(OOXMLFormulaScanError, match="formula-capable"):
        scan_ooxml_formulas(rewritten)


@pytest.mark.parametrize(
    "relationship_kind",
    ["calcChain", "pivotCacheDefinition", "externalLink", "package", "ctrlProp"],
)
def test_formula_scan_rejects_renamed_formula_capable_relationship_targets(
    tmp_path: Path,
    relationship_kind: str,
) -> None:
    original = tmp_path / "original.xlsx"
    rewritten = tmp_path / "renamed-opaque.xlsx"
    _workbook(original)
    part_name = "xl/misc/innocent.xml"
    relationship = (
        '<Relationship Id="rIdOpaque" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/'
        f'{relationship_kind}" Target="misc/innocent.xml"/>'
    ).encode("ascii")

    def declare_renamed_target(name: str, payload: bytes) -> bytes:
        if name == "[Content_Types].xml":
            declaration = (
                f'<Override PartName="/{part_name}" ContentType="application/xml"/>'
            ).encode("ascii")
            return payload.replace(b"</Types>", declaration + b"</Types>", 1)
        if name == "xl/_rels/workbook.xml.rels":
            return payload.replace(b"</Relationships>", relationship + b"</Relationships>", 1)
        return payload

    _rewrite_package(original, rewritten, declare_renamed_target)
    with ZipFile(rewritten, "a", ZIP_DEFLATED) as archive:
        archive.writestr(part_name, b'<opaque xmlns="urn:innocent">cached data</opaque>')

    with pytest.raises(OOXMLFormulaScanError, match="formula-capable relationship"):
        scan_ooxml_formulas(rewritten)


@pytest.mark.parametrize("relationship_kind", ["oleObject", "package"])
def test_formula_scan_rejects_embedded_formula_workbook_renamed_outside_embeddings(
    tmp_path: Path,
    relationship_kind: str,
) -> None:
    original = tmp_path / "original.xlsx"
    embedded = tmp_path / "embedded.xlsx"
    rewritten = tmp_path / "renamed-embedded.xlsx"
    _workbook(original)
    _workbook(embedded, formula="=SUM(A1:A2)")
    embedded_part = "xl/misc/object.bin"
    sheet_relationships = "xl/worksheets/_rels/sheet1.xml.rels"
    relationship = (
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rIdOpaque" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/'
        f'{relationship_kind}" Target="../misc/object.bin"/>'
        "</Relationships>"
    ).encode("ascii")

    def declare_embedded_target(name: str, payload: bytes) -> bytes:
        if name == "[Content_Types].xml":
            declaration = (
                f'<Override PartName="/{embedded_part}" '
                'ContentType="application/vnd.openxmlformats-officedocument.oleObject"/>'
            ).encode("ascii")
            return payload.replace(b"</Types>", declaration + b"</Types>", 1)
        if name == "xl/worksheets/sheet1.xml":
            declaration = (
                '<oleObjects><oleObject progId="Excel.Sheet.12" shapeId="1" '
                'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" '
                'r:id="rIdOpaque"/></oleObjects>'
            ).encode("ascii")
            return payload.replace(b"</worksheet>", declaration + b"</worksheet>", 1)
        return payload

    _rewrite_package(original, rewritten, declare_embedded_target)
    with ZipFile(rewritten, "a", ZIP_DEFLATED) as archive:
        archive.writestr(sheet_relationships, relationship)
        archive.writestr(embedded_part, embedded.read_bytes())

    with pytest.raises(
        OOXMLFormulaScanError,
        match="formula-capable relationship|opaque binary",
    ):
        scan_ooxml_formulas(rewritten)


def test_formula_scan_rejects_unknown_extension_relationship_as_ambiguous(
    tmp_path: Path,
) -> None:
    original = tmp_path / "original.xlsx"
    rewritten = tmp_path / "unknown-extension.xlsx"
    _workbook(original)
    embedded_part = "xl/misc/payload.xml"
    sheet_relationships = "xl/worksheets/_rels/sheet1.xml.rels"
    relationship = (
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rIdOpaque" Type="urn:vendor:calc/opaquePayload" '
        'Target="../misc/payload.xml"/>'
        "</Relationships>"
    ).encode("ascii")

    def declare_extension(name: str, payload: bytes) -> bytes:
        if name == "[Content_Types].xml":
            declaration = (
                f'<Override PartName="/{embedded_part}" ContentType="application/xml"/>'
            ).encode("ascii")
            return payload.replace(b"</Types>", declaration + b"</Types>", 1)
        if name == "xl/worksheets/sheet1.xml":
            extension = (
                '<extLst><ext uri="urn:vendor:calc"><v:payload '
                'xmlns:v="urn:vendor:calc" '
                'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" '
                'r:id="rIdOpaque"/></ext></extLst>'
            ).encode("ascii")
            return payload.replace(b"</worksheet>", extension + b"</worksheet>", 1)
        return payload

    _rewrite_package(original, rewritten, declare_extension)
    with ZipFile(rewritten, "a", ZIP_DEFLATED) as archive:
        archive.writestr(sheet_relationships, relationship)
        archive.writestr(embedded_part, b'<payload xmlns="urn:vendor:calc">opaque</payload>')

    with pytest.raises(OOXMLFormulaScanError, match="unsupported or ambiguous relationship"):
        scan_ooxml_formulas(rewritten)


def test_formula_scan_rejects_unclassified_opaque_binary_part(tmp_path: Path) -> None:
    original = tmp_path / "original.xlsx"
    embedded = tmp_path / "embedded.xlsx"
    rewritten = tmp_path / "opaque-binary.xlsx"
    _workbook(original)
    _workbook(embedded, formula="=1+1")
    embedded_part = "xl/misc/payload.bin"

    def declare_binary(name: str, payload: bytes) -> bytes:
        if name != "[Content_Types].xml":
            return payload
        declaration = (
            f'<Override PartName="/{embedded_part}" ContentType="application/octet-stream"/>'
        ).encode("ascii")
        return payload.replace(b"</Types>", declaration + b"</Types>", 1)

    _rewrite_package(original, rewritten, declare_binary)
    with ZipFile(rewritten, "a", ZIP_DEFLATED) as archive:
        archive.writestr(embedded_part, embedded.read_bytes())

    with pytest.raises(OOXMLFormulaScanError, match="opaque binary"):
        scan_ooxml_formulas(rewritten)


def test_formula_scan_rejects_symlink_hardlink_and_symbolic_parent(
    tmp_path: Path,
) -> None:
    real = tmp_path / "real.xlsx"
    _workbook(real)
    symlink = tmp_path / "symlink.xlsx"
    symlink.symlink_to(real.name)
    with pytest.raises(OOXMLFormulaScanError, match="regular non-symbolic"):
        scan_ooxml_formulas(symlink)

    hardlink = tmp_path / "hardlink.xlsx"
    os.link(real, hardlink)
    with pytest.raises(OOXMLFormulaScanError, match="one-link"):
        scan_ooxml_formulas(real)
    hardlink.unlink()

    real_parent = tmp_path / "real-parent"
    real_parent.mkdir()
    nested = real_parent / "nested.xlsx"
    _workbook(nested)
    symbolic_parent = tmp_path / "symbolic-parent"
    symbolic_parent.symlink_to(real_parent, target_is_directory=True)
    with pytest.raises(OOXMLFormulaScanError, match="parent chain"):
        scan_ooxml_formulas(symbolic_parent / "nested.xlsx")

    nonregular = tmp_path / "directory.xlsx"
    nonregular.mkdir()
    with pytest.raises(OOXMLFormulaScanError, match="regular non-symbolic"):
        scan_ooxml_formulas(nonregular)


def test_formula_scan_rejects_parent_replaced_by_symlink_during_open(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    parent = tmp_path / "bound-parent"
    displaced = tmp_path / "displaced-parent"
    parent.mkdir()
    path = parent / "plain.xlsx"
    _workbook(path)
    replaced = False

    def replace_parent_before_file_open(stage: str, _: Path) -> None:
        nonlocal replaced
        if stage == "parent_chain_opened" and not replaced:
            replaced = True
            parent.rename(displaced)
            parent.symlink_to(displaced.name, target_is_directory=True)

    monkeypatch.setattr(
        formula_scan_module,
        "_formula_scan_lease_hook",
        replace_parent_before_file_open,
    )

    with pytest.raises(OOXMLFormulaScanError, match="parent chain changed identity"):
        scan_ooxml_formulas(path)

    assert replaced


def test_formula_scan_enforces_explicit_file_size_limit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = tmp_path / "plain.xlsx"
    _workbook(path)
    monkeypatch.setattr(
        formula_scan_module,
        "OOXML_FORMULA_SCAN_MAX_FILE_BYTES",
        path.stat().st_size - 1,
    )

    with pytest.raises(OOXMLFormulaScanError, match="file size"):
        scan_ooxml_formulas(path)


def test_formula_scan_rejects_live_path_rewrite_after_snapshot_even_if_bytes_return(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    formula = tmp_path / "formula.xlsx"
    clean = tmp_path / "clean.xlsx"
    _workbook(formula, formula="=1+1")
    _workbook(clean)
    formula_bytes = formula.read_bytes()
    clean_bytes = clean.read_bytes()
    original_scan = formula_scan_module._scan_ooxml_snapshot

    def swap_live_path(snapshot: object, **kwargs: object):
        formula.write_bytes(clean_bytes)
        formula.write_bytes(formula_bytes)
        return original_scan(snapshot, **kwargs)

    monkeypatch.setattr(formula_scan_module, "_scan_ooxml_snapshot", swap_live_path)

    with pytest.raises(OOXMLFormulaScanError, match="changed after snapshotting"):
        scan_ooxml_formulas(formula)

    assert formula.read_bytes() == formula_bytes


@pytest.mark.parametrize("mutation", ["replace", "rewrite"])
def test_formula_scan_rejects_source_mutation_during_snapshot(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    mutation: str,
) -> None:
    path = tmp_path / "source.xlsx"
    replacement = tmp_path / "replacement.xlsx"
    _workbook(path)
    _workbook(replacement, formula="=1+1")
    mutated = False

    def mutate_after_snapshot(stage: str, _: Path) -> None:
        nonlocal mutated
        if stage == "snapshot_copied" and not mutated:
            mutated = True
            if mutation == "replace":
                replacement.replace(path)
            else:
                path.write_bytes(replacement.read_bytes())

    monkeypatch.setattr(
        formula_scan_module,
        "_formula_scan_lease_hook",
        mutate_after_snapshot,
    )

    with pytest.raises(OOXMLFormulaScanError, match="changed during snapshotting"):
        scan_ooxml_formulas(path)

    assert mutated


def test_formula_scan_closes_all_capabilities_when_package_scan_fails(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = tmp_path / "plain.xlsx"
    _workbook(path)
    original_open = os.open
    opened: list[int] = []

    def track_open(*args: object, **kwargs: object) -> int:
        descriptor = original_open(*args, **kwargs)
        opened.append(descriptor)
        return descriptor

    def fail_scan(*_: object, **__: object) -> None:
        raise OOXMLFormulaScanError("injected package scan failure")

    monkeypatch.setattr(formula_scan_module.os, "open", track_open)
    monkeypatch.setattr(formula_scan_module, "_scan_ooxml_snapshot", fail_scan)

    with pytest.raises(OOXMLFormulaScanError, match="package scan failure"):
        scan_ooxml_formulas(path)

    assert opened
    for descriptor in opened:
        with pytest.raises(OSError):
            os.fstat(descriptor)


def test_formula_scan_closes_child_descriptor_when_initial_fstat_fails(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = tmp_path / "nested" / "plain.xlsx"
    path.parent.mkdir()
    _workbook(path)
    original_open = os.open
    original_fstat = os.fstat
    opened: list[int] = []

    def track_open(*args: object, **kwargs: object) -> int:
        descriptor = original_open(*args, **kwargs)
        opened.append(descriptor)
        return descriptor

    def fail_first_child_fstat(descriptor: int) -> os.stat_result:
        if len(opened) >= 2 and descriptor == opened[1]:
            raise OSError("injected child-directory fstat failure")
        return original_fstat(descriptor)

    monkeypatch.setattr(formula_scan_module.os, "open", track_open)
    monkeypatch.setattr(formula_scan_module.os, "fstat", fail_first_child_fstat)

    with pytest.raises(OOXMLFormulaScanError, match="could not be safely opened"):
        scan_ooxml_formulas(path)

    assert len(opened) == 2
    for descriptor in opened:
        with pytest.raises(OSError):
            original_fstat(descriptor)


def test_formula_scan_lease_holds_complete_parent_chain_valid_control(tmp_path: Path) -> None:
    parent = tmp_path / "parent" / "nested"
    parent.mkdir(parents=True)
    path = parent / "plain.xlsx"
    _workbook(path)

    with formula_scan_module.OOXMLFormulaScanLease.open(path) as lease:
        assert lease.closed is False
        assert lease.scan.package_sha256 == hashlib.sha256(path.read_bytes()).hexdigest()
        assert len(lease.parent_identities) == len(path.absolute().parent.parts)
        lease.verify_binding(checkpoint="valid_control")

    assert lease.closed is True


def test_formula_scan_rejects_parent_swap_during_scan(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    parent = tmp_path / "parent"
    displaced = tmp_path / "displaced"
    parent.mkdir()
    path = parent / "plain.xlsx"
    _workbook(path)
    swapped = False

    def swap_after_scan(stage: str, _: Path) -> None:
        nonlocal swapped
        if stage == "scan_complete" and not swapped:
            swapped = True
            parent.rename(displaced)
            parent.mkdir()
            _workbook(parent / path.name)

    monkeypatch.setattr(
        formula_scan_module,
        "_formula_scan_lease_hook",
        swap_after_scan,
    )

    with pytest.raises(OOXMLFormulaScanError, match="parent chain changed identity"):
        scan_ooxml_formulas(path)

    assert swapped
    assert (displaced / path.name).is_file()
    assert path.is_file()


def test_formula_scan_rejects_file_replacement_immediately_before_lease_exit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = tmp_path / "plain.xlsx"
    replacement = tmp_path / "replacement.xlsx"
    displaced = tmp_path / "displaced.xlsx"
    _workbook(path)
    _workbook(replacement, formula="=1+1")
    replacement_identity = replacement.stat().st_dev, replacement.stat().st_ino
    replaced = False

    def replace_before_exit(stage: str, _: Path) -> None:
        nonlocal replaced
        if stage == "before_lease_exit" and not replaced:
            replaced = True
            path.rename(displaced)
            replacement.rename(path)

    monkeypatch.setattr(
        formula_scan_module,
        "_formula_scan_lease_hook",
        replace_before_exit,
    )

    with pytest.raises(OOXMLFormulaScanError, match="changed after snapshotting"):
        scan_ooxml_formulas(path)

    observed = path.stat()
    assert replaced
    assert (observed.st_dev, observed.st_ino) == replacement_identity
    assert path.is_file()
    assert displaced.is_file()


def test_formula_scan_rebinds_lexical_name_after_hashing_held_descriptor(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = tmp_path / "plain.xlsx"
    replacement = tmp_path / "replacement.xlsx"
    displaced = tmp_path / "displaced.xlsx"
    _workbook(path)
    _workbook(replacement, formula="=1+1")
    original_snapshot = formula_scan_module._descriptor_snapshot
    snapshot_calls = 0
    replacement_identity = replacement.stat().st_dev, replacement.stat().st_ino

    def replace_after_verification_hash(descriptor: int, *, expected_size: int) -> bytes:
        nonlocal snapshot_calls
        payload = original_snapshot(descriptor, expected_size=expected_size)
        snapshot_calls += 1
        if snapshot_calls == 2:
            path.rename(displaced)
            replacement.rename(path)
        return payload

    monkeypatch.setattr(
        formula_scan_module,
        "_descriptor_snapshot",
        replace_after_verification_hash,
    )

    with pytest.raises(OOXMLFormulaScanError, match="changed after snapshotting"):
        scan_ooxml_formulas(path)

    observed = path.stat()
    assert snapshot_calls == 2
    assert (observed.st_dev, observed.st_ino) == replacement_identity
    assert path.is_file()
    assert displaced.is_file()


def test_formula_scan_convenience_result_claims_only_closed_lease_snapshot(
    tmp_path: Path,
) -> None:
    path = tmp_path / "plain.xlsx"
    replacement = tmp_path / "replacement.xlsx"
    _workbook(path)
    original_bytes = path.read_bytes()
    _workbook(replacement, formula="=1+1")

    result = scan_ooxml_formulas(path)
    replacement.replace(path)

    assert result.package_sha256 == hashlib.sha256(original_bytes).hexdigest()
    assert result.has_formulas is False
    assert hashlib.sha256(path.read_bytes()).hexdigest() != result.package_sha256
