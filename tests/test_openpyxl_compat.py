from __future__ import annotations

import zipfile
from pathlib import Path
from xml.etree import ElementTree

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.reader.excel import ExcelReader

from spreadsheet_harness.openpyxl_compat import load_workbook
from spreadsheet_harness.render import sheet_inventory_identity


def _sheet_signature(path: Path) -> list[dict[str, object]]:
    return sheet_inventory_identity(path)["sheets"]


def _add_fake_vba_project(source: Path, destination: Path, payload: bytes) -> None:
    content_types_namespace = (
        "http://schemas.openxmlformats.org/package/2006/content-types"
    )
    relationships_namespace = (
        "http://schemas.openxmlformats.org/package/2006/relationships"
    )
    with zipfile.ZipFile(source) as package, zipfile.ZipFile(destination, "w") as output:
        for member in package.infolist():
            data = package.read(member.filename)
            if member.filename == "[Content_Types].xml":
                root = ElementTree.fromstring(data)
                for override in root.findall(
                    f"{{{content_types_namespace}}}Override"
                ):
                    if override.get("PartName") == "/xl/workbook.xml":
                        override.set(
                            "ContentType",
                            "application/vnd.ms-excel.sheet.macroEnabled.main+xml",
                        )
                root.append(
                    ElementTree.Element(
                        f"{{{content_types_namespace}}}Default",
                        {
                            "Extension": "bin",
                            "ContentType": "application/vnd.ms-office.vbaProject",
                        },
                    )
                )
                data = ElementTree.tostring(
                    root,
                    encoding="utf-8",
                    xml_declaration=True,
                )
            elif member.filename == "xl/_rels/workbook.xml.rels":
                root = ElementTree.fromstring(data)
                root.append(
                    ElementTree.Element(
                        f"{{{relationships_namespace}}}Relationship",
                        {
                            "Id": "rIdSheetHarnessVba",
                            "Type": (
                                "http://schemas.microsoft.com/office/2006/"
                                "relationships/vbaProject"
                            ),
                            "Target": "vbaProject.bin",
                        },
                    )
                )
                data = ElementTree.tostring(
                    root,
                    encoding="utf-8",
                    xml_declaration=True,
                )
            output.writestr(member, data)
        output.writestr("xl/vbaProject.bin", payload)


def test_load_workbook_repeatedly_preserves_empty_chartsheet(
    empty_chartsheet_workbook: Path,
) -> None:
    expected = _sheet_signature(empty_chartsheet_workbook)
    original_reader = ExcelReader.read_chartsheet

    for value in ("first", "second"):
        workbook = load_workbook(empty_chartsheet_workbook)
        assert workbook.sheetnames == ["Data", "Chart"]
        assert [type(sheet).__name__ for sheet in workbook._sheets] == [
            "Worksheet",
            "Chartsheet",
        ]
        assert workbook["Chart"].sheet_state == "hidden"
        workbook["Data"]["A1"] = value
        workbook.save(empty_chartsheet_workbook)
        workbook.close()

        reopened = load_workbook(empty_chartsheet_workbook, read_only=True)
        assert reopened.sheetnames == ["Data", "Chart"]
        assert reopened["Chart"].sheet_state == "hidden"
        reopened.close()
        assert _sheet_signature(empty_chartsheet_workbook) == expected
        assert ExcelReader.read_chartsheet is original_reader


def test_load_workbook_delegates_chartsheet_drawing_handling(tmp_path: Path) -> None:
    source = tmp_path / "drawing-chartsheet.xlsx"
    saved = tmp_path / "drawing-chartsheet-saved.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    for row, value in enumerate((1, 2, 3), start=1):
        worksheet.cell(row, 1, value)
    chart = BarChart()
    chart.add_data(Reference(worksheet, min_col=1, min_row=1, max_row=3))
    chartsheet = workbook.create_chartsheet("Chart")
    chartsheet.add_chart(chart)
    chartsheet.sheet_state = "hidden"
    workbook.save(source)
    workbook.close()

    loaded = load_workbook(source)
    assert len(loaded["Chart"]._charts) == 1
    assert loaded["Chart"].sheet_state == "hidden"
    loaded.save(saved)
    loaded.close()

    reopened = load_workbook(saved)
    assert len(reopened["Chart"]._charts) == 1
    assert reopened["Chart"].sheet_state == "hidden"
    reopened.close()
    assert _sheet_signature(saved) == _sheet_signature(source)


def test_load_workbook_preserves_vba_archive_with_empty_chartsheet(
    empty_chartsheet_workbook: Path,
    tmp_path: Path,
) -> None:
    macro_workbook = tmp_path / "empty-chartsheet.xlsm"
    saved = tmp_path / "empty-chartsheet-saved.xlsm"
    payload = b"sheet-harness fake VBA payload"
    _add_fake_vba_project(empty_chartsheet_workbook, macro_workbook, payload)
    expected = _sheet_signature(macro_workbook)

    workbook = load_workbook(macro_workbook, keep_vba=True)
    assert workbook.vba_archive is not None
    assert workbook["Chart"].sheet_state == "hidden"
    workbook["Data"]["A1"] = "macro-safe edit"
    workbook.save(saved)
    workbook.close()

    with zipfile.ZipFile(saved) as package:
        assert package.read("xl/vbaProject.bin") == payload
    assert _sheet_signature(saved) == expected
