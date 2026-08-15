from __future__ import annotations

import json
import zipfile
from pathlib import Path

from openpyxl import Workbook, load_workbook

from spreadsheet_harness.formula_runtime import (
    FormulaCellState,
    FormulaInventory,
    formula_coordinate_sha256,
    formula_inventory,
    formula_runtime_report,
)


def _save_formula_workbook(path: Path, formula: str | None) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    if formula is not None:
        sheet["A1"] = formula
    workbook.save(path)
    workbook.close()


def test_formula_inventory_tracks_add_change_delete_and_exact_undo(tmp_path: Path) -> None:
    workbook_path = tmp_path / "formula-state.xlsx"
    _save_formula_workbook(workbook_path, None)
    initial = formula_inventory(workbook_path)

    _save_formula_workbook(workbook_path, "=1+1")
    added = formula_inventory(workbook_path)
    _save_formula_workbook(workbook_path, "=1+2")
    changed = formula_inventory(workbook_path)
    _save_formula_workbook(workbook_path, None)
    deleted = formula_inventory(workbook_path)
    _save_formula_workbook(workbook_path, "=1+1")
    restored = formula_inventory(workbook_path)

    assert initial.cells == {}
    assert set(added.cells) == {("Data", "A1")}
    assert added.cells[("Data", "A1")].formula_sha256 != changed.cells[
        ("Data", "A1")
    ].formula_sha256
    assert deleted.cells == {}
    assert restored.state_sha256 == added.state_sha256


def test_formula_runtime_report_is_exact_bounded_and_contains_no_formula_text(
    tmp_path: Path,
) -> None:
    secret_formula = "EVALUATOR_FORMULA_MUST_NOT_LEAK"
    workbook_path = tmp_path / "formula-no-leak.xlsx"
    _save_formula_workbook(workbook_path, f"={secret_formula}()")
    extracted = formula_inventory(workbook_path)
    extracted_report = formula_runtime_report(extracted, extracted.cells)
    assert secret_formula not in repr(extracted)
    assert secret_formula not in json.dumps(extracted_report, sort_keys=True)

    cells = {
        ("Data", f"A{row}"): FormulaCellState(
            formula_sha256=f"{row:064x}",
            cached_type="e",
            cached_value=("Err:509" if row == 1 else "#DIV/0!"),
        )
        for row in range(1, 602)
    }
    inventory = FormulaInventory(
        workbook_sha256="a" * 64,
        state_sha256="b" * 64,
        cells=cells,
    )

    report = formula_runtime_report(inventory, cells)
    serialized = json.dumps(report, sort_keys=True)

    assert report["coordinate_count"] == 601
    assert report["coordinate_sha256"] == formula_coordinate_sha256(cells)
    assert report["formula_cells_present"] == 601
    assert report["formula_cells_absent"] == 0
    assert report["calculation_errors"]["count"] == 601
    assert len(report["calculation_errors"]["coordinates"]) == 32
    assert report["calculation_errors"]["coordinates_truncated"] is True
    assert report["calculation_errors"]["coordinates"][0]["error"] == "Err:509"
    assert secret_formula not in serialized


def test_formula_inventory_includes_hidden_worksheets_and_ignores_chartsheets(
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "mixed-sheets.xlsx"
    workbook = Workbook()
    workbook.active.title = "Visible"
    workbook["Visible"]["A1"] = "=1+1"
    hidden = workbook.create_sheet("Hidden")
    hidden.sheet_state = "hidden"
    hidden["B2"] = "=2+2"
    workbook.create_chartsheet("Chart")
    workbook.save(workbook_path)
    workbook.close()

    inventory = formula_inventory(workbook_path)

    assert set(inventory.cells) == {("Visible", "A1"), ("Hidden", "B2")}


def test_formula_inventory_binds_shared_formula_followers_to_the_master(
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "shared-formula.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet["A1"] = 1
    sheet["A2"] = "=A1+1"
    sheet["A3"] = "=A2+1"
    workbook.save(workbook_path)
    workbook.close()

    rewritten = tmp_path / "shared-formula-rewritten.xlsx"
    with zipfile.ZipFile(workbook_path) as source, zipfile.ZipFile(
        rewritten, "w"
    ) as destination:
        for member in source.infolist():
            payload = source.read(member)
            if member.filename == "xl/worksheets/sheet1.xml":
                payload = payload.replace(
                    b"<f>A1+1</f>",
                    b'<f t="shared" ref="A2:A3" si="0">A1+1</f>',
                ).replace(
                    b"<f>A2+1</f>",
                    b'<f t="shared" si="0"></f>',
                )
            destination.writestr(member, payload)
    rewritten.replace(workbook_path)

    inventory = formula_inventory(workbook_path)

    assert set(inventory.cells) == {("Data", "A2"), ("Data", "A3")}
    assert inventory.cells[("Data", "A2")].formula_sha256 != inventory.cells[
        ("Data", "A3")
    ].formula_sha256


def test_formula_runtime_report_marks_deleted_formula_coordinates_absent(
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "deleted-formula.xlsx"
    _save_formula_workbook(workbook_path, "=1+1")
    before = formula_inventory(workbook_path)
    workbook = load_workbook(workbook_path)
    workbook["Data"]["A1"] = None
    workbook.save(workbook_path)
    workbook.close()
    after = formula_inventory(workbook_path)

    report = formula_runtime_report(after, before.cells)

    assert report["formula_cells_present"] == 0
    assert report["formula_cells_absent"] == 1
    assert report["coverage_complete"] is True


def test_formula_runtime_report_does_not_treat_error_looking_text_as_an_error() -> None:
    inventory = FormulaInventory(
        workbook_sha256="a" * 64,
        state_sha256="b" * 64,
        cells={
            ("Data", "A1"): FormulaCellState(
                formula_sha256="c" * 64,
                cached_type="str",
                cached_value="#REF!",
            )
        },
    )

    report = formula_runtime_report(inventory, inventory.cells)

    assert report["calculation_errors"]["count"] == 0
