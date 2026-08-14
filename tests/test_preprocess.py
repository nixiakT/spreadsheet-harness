from __future__ import annotations

import csv
import json
from pathlib import Path

import pytest
import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, GradientFill, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from spreadsheet_harness.preprocess import (
    DETERMINISTIC_PROFILE_BOUNDS,
    DETERMINISTIC_PROFILE_SCHEMA_VERSION,
    PreprocessError,
    build_deterministic_profile,
    build_preprocess_view,
    preprocess_workbook,
    render_deterministic_profile,
    render_markdown,
)
from spreadsheet_harness.render import (
    convert_spreadsheet_copy,
    find_libreoffice,
    sha256_file,
)


def _save_feature_workbook(path: Path) -> None:
    workbook = Workbook()
    data = workbook.active
    data.title = "Data"
    data.append(["Name", "Amount", "Category"])
    data.append(["Alpha", 10, "A"])
    data.append(["Beta", 20, "B"])
    data["B4"] = "=SUM(B2:B3)"
    data["A1"].font = Font(bold=True, color="FFFFFFFF")
    data["A1"].fill = PatternFill("solid", fgColor="FF336699")
    data["B2"].number_format = "$#,##0.00"
    data.merge_cells("E1:F1")
    data["E1"] = "Merged heading"
    table = Table(displayName="Sales", ref="A1:C3")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    data.add_table(table)

    hidden = workbook.create_sheet("Hidden")
    hidden.sheet_state = "hidden"
    hidden["A1"] = "secret-ish metadata"
    workbook.save(path)
    workbook.close()


def _long_number_format(sheet_index: int, format_index: int) -> str:
    prefix = f'"fmt-{sheet_index}-{format_index}-'
    suffix = '"$#,##0.00'
    return prefix + ("0" * (250 - len(prefix) - len(suffix))) + suffix


def _sheet(data: dict[str, object], name: str) -> dict[str, object]:
    sheets = data["sheets"]
    assert isinstance(sheets, list)
    return next(sheet for sheet in sheets if sheet["name"] == name)


def _cell(sheet: dict[str, object], coordinate: str) -> dict[str, object]:
    cells = sheet["cells"]
    assert isinstance(cells, list)
    return next(cell for cell in cells if cell["coordinate"] == coordinate)


def test_preprocess_writes_three_views_and_preserves_source(tmp_path: Path) -> None:
    source = tmp_path / "features.xlsx"
    _save_feature_workbook(source)
    before = source.read_bytes()

    result = preprocess_workbook(source, tmp_path / "views")

    assert source.read_bytes() == before
    assert result.source_sha256 == sha256_file(source)
    assert result.json_path.is_file()
    assert result.yaml_path.is_file()
    assert result.markdown_path.is_file()

    json_view = json.loads(result.json_path.read_text(encoding="utf-8"))
    yaml_view = yaml.safe_load(result.yaml_path.read_text(encoding="utf-8"))
    markdown = result.markdown_path.read_text(encoding="utf-8")
    assert json_view == yaml_view == result.data
    assert "## Inventory" in markdown
    assert "### Formulas and values" in markdown
    assert "### Regions" in markdown
    assert "### Merged ranges" in markdown
    assert "### Tables" in markdown
    assert "### Style summary" in markdown


def test_preprocess_inventory_formula_region_merge_table_and_style(tmp_path: Path) -> None:
    source = tmp_path / "features.xlsx"
    _save_feature_workbook(source)

    view = build_preprocess_view(source)
    inventory = view["inventory"]
    assert inventory["sheet_count"] == 2
    data_inventory = inventory["sheets"][0]
    assert data_inventory["name"] == "Data"
    assert data_inventory["used_region"] == "A1:F4"
    assert data_inventory["formula_count"] == 1
    assert data_inventory["merged_range_count"] == 1
    assert data_inventory["table_count"] == 1

    data = _sheet(view, "Data")
    formula = _cell(data, "B4")
    assert formula["formula"] == "=SUM(B2:B3)"
    # openpyxl-authored formulas have no cached result until a calc engine runs.
    assert formula["value"] is None
    assert data["merged_ranges"] == ["E1:F1"]
    assert data["tables"][0]["name"] == "Sales"
    assert data["tables"][0]["range"] == "A1:C3"
    assert data["styles"]
    assert any(style["fill"]["type"] == "solid" for style in data["styles"])
    assert data["regions"][0]["range"] == "A1:F4"
    region_cells = {cell["coordinate"] for cell in data["regions"][0]["cells"]}
    assert {"A1", "B4", "E1"}.issubset(region_cells)


def test_requested_region_is_used_in_all_views(tmp_path: Path) -> None:
    source = tmp_path / "features.xlsx"
    _save_feature_workbook(source)

    view = build_preprocess_view(source, regions={"Data": ["A1:B2"]})
    data = _sheet(view, "Data")

    assert [region["range"] for region in data["regions"]] == ["A1:B2"]
    assert {cell["coordinate"] for cell in data["regions"][0]["cells"]} == {
        "A1",
        "B1",
        "A2",
        "B2",
    }
    markdown = render_markdown(view)
    assert "#### Region `A1:B2`" in markdown


def test_unknown_sheet_region_is_rejected(tmp_path: Path) -> None:
    source = tmp_path / "features.xlsx"
    _save_feature_workbook(source)

    with pytest.raises(PreprocessError, match="unknown sheets"):
        build_preprocess_view(source, regions={"Missing": "A1:B2"})


def test_cell_budget_records_truncation(tmp_path: Path) -> None:
    source = tmp_path / "features.xlsx"
    _save_feature_workbook(source)

    view = build_preprocess_view(source, max_cells_per_sheet=2)
    data = _sheet(view, "Data")

    assert len(data["cells"]) == 2
    assert data["cells_truncated"] is True
    assert len(data["regions"][0]["cells"]) == 2
    assert data["regions"][0]["truncated"] is True


def test_deterministic_profile_has_provenance_confidence_and_truncation(
    tmp_path: Path,
) -> None:
    source = tmp_path / "features.xlsx"
    _save_feature_workbook(source)

    profile = build_deterministic_profile(
        source,
        bounds={
            "max_sheets": 1,
            "max_cells_per_sheet": 3,
            "max_regions_per_sheet": 1,
            "max_formula_clusters_per_sheet": 1,
        },
    )
    rendered = render_deterministic_profile(profile)
    parsed = json.loads(rendered)

    assert parsed["schema_version"] == DETERMINISTIC_PROFILE_SCHEMA_VERSION
    assert len(parsed["profile_sha256"]) == 64
    assert parsed["task_independent"] is True
    assert parsed["truncation"]["sheets"] is True
    sheet = parsed["sheets"][0]
    assert sheet["provenance"] == {
        "method": "openpyxl-read-only-profile",
        "sheet": "Data",
        "range": "A1:F4",
    }
    assert sheet["confidence"]["regions"] == "heuristic"
    assert sheet["truncation"]["cells"] is True
    assert sheet["regions"][0]["provenance"]["sheet"] == "Data"
    assert len(rendered) <= parsed["bounds"]["max_rendered_chars"]


def test_deterministic_profile_is_stable_and_preserves_source(tmp_path: Path) -> None:
    source = tmp_path / "features.xlsx"
    _save_feature_workbook(source)
    before = source.read_bytes()

    first = build_deterministic_profile(source)
    second = build_deterministic_profile(source)

    assert first == second
    assert source.read_bytes() == before


def test_deterministic_profile_defaults_are_compact_but_keep_key_evidence(
    tmp_path: Path,
) -> None:
    assert DETERMINISTIC_PROFILE_BOUNDS == {
        "max_sheets": 8,
        "max_cells_per_sheet": 384,
        "max_regions_per_sheet": 3,
        "max_header_rows": 2,
        "max_columns_per_region": 8,
        "max_sample_rows_per_region": 3,
        "max_number_formats_per_region": 6,
        "max_formula_clusters_per_sheet": 6,
        "max_provenance_cells_per_claim": 3,
        "max_scalar_chars": 96,
        "max_rendered_chars": 12_000,
    }

    source = tmp_path / "representative.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Evidence"
    for row in range(1, 11):
        for column in range(1, 11):
            worksheet.cell(row, column, row * column)
    worksheet["B2"].number_format = "$#,##0.00"
    worksheet["J10"] = "=$A10+B$1+$C$2"
    workbook.save(source)
    workbook.close()

    profile = build_deterministic_profile(source)
    sheet = profile["sheets"][0]
    region = sheet["regions"][0]
    sample_cells = {item["cell"] for item in region["sample"]}

    assert {"A1", "J1", "A10", "J10"}.issubset(sample_cells)
    assert any(1 < int("".join(filter(str.isdigit, cell))) < 10 for cell in sample_cells)
    assert region["number_formats"] == {"$#,##0.00": 1}
    assert region["provenance"]["sheet"] == "Evidence"
    assert region["provenance"]["sample_cells"] == ["A1", "F5", "J10"]
    assert sheet["formula_clusters"][0]["sample_formulas"] == [
        {
            "cell": "J10",
            "formula": "=$A10+B$1+$C$2",
            "truncated": False,
        }
    ]
    assert profile["source"]["format"] == "xlsx"
    assert profile["backend"]["reader"] == "openpyxl"
    assert len(render_deterministic_profile(profile)) <= 12_000


def test_profile_rendering_compacts_repetition_without_dropping_formula_evidence(
    tmp_path: Path,
) -> None:
    source = tmp_path / "repeated.xlsx"
    workbook = Workbook()
    for sheet_index in range(8):
        worksheet = workbook.active if sheet_index == 0 else workbook.create_sheet()
        worksheet.title = f"Month {sheet_index + 1}"
        worksheet.append(["Input", "Output"])
        for row in range(2, 22):
            worksheet.cell(row, 1, row)
            worksheet.cell(row, 2, f"=$A{row}*B$1")
            worksheet.cell(row, 2).number_format = "0.00"
    workbook.save(source)
    workbook.close()

    rendered = render_deterministic_profile(build_deterministic_profile(source))
    parsed = json.loads(rendered)

    assert len(rendered) <= DETERMINISTIC_PROFILE_BOUNDS["max_rendered_chars"]
    assert parsed["truncation"]["rendered"] is True
    assert parsed["source"]["format"] == "xlsx"
    assert parsed["backend"]["reader"] == "openpyxl"
    assert all(sheet["formula_clusters"] for sheet in parsed["sheets"])
    assert all(
        sheet["formula_clusters"][0]["sample_formulas"]
        for sheet in parsed["sheets"]
    )
    assert all(sheet["regions"][0]["number_formats"] for sheet in parsed["sheets"])
    assert all(sheet["provenance"]["sheet"] for sheet in parsed["sheets"])


def test_profile_rendering_hard_caps_long_number_formats(tmp_path: Path) -> None:
    source = tmp_path / "long-formats.xlsx"
    workbook = Workbook()
    for sheet_index in range(8):
        worksheet = workbook.active if sheet_index == 0 else workbook.create_sheet()
        worksheet.title = f"Formats {sheet_index + 1}"
        for format_index in range(6):
            number_format = _long_number_format(sheet_index, format_index)
            assert len(number_format) == 250
            cell = worksheet.cell(row=1, column=format_index + 1, value=format_index + 1)
            cell.number_format = number_format
    workbook.save(source)
    workbook.close()

    profile = build_deterministic_profile(source)
    assert len(profile["sheets"]) == 8
    for sheet in profile["sheets"]:
        region = sheet["regions"][0]
        assert len(region["number_formats"]) == 6
        assert region["number_formats_truncated"] is True
        assert all(
            len(number_format) <= DETERMINISTIC_PROFILE_BOUNDS["max_scalar_chars"]
            for number_format in region["number_formats"]
        )
        assert region["unit_hints"]
        assert all(
            hint["provenance"]["number_format_truncated"] is True
            and len(hint["provenance"]["number_format"])
            <= DETERMINISTIC_PROFILE_BOUNDS["max_scalar_chars"]
            for hint in region["unit_hints"]
        )

    rendered = render_deterministic_profile(profile)

    assert len(rendered) <= DETERMINISTIC_PROFILE_BOUNDS["max_rendered_chars"]
    assert render_deterministic_profile(profile) == rendered
    parsed = json.loads(rendered)
    assert len(parsed["sheets"]) == 8
    assert parsed["truncation"]["rendered"] is True


def test_preprocess_summarizes_gradient_fills(tmp_path: Path) -> None:
    source = tmp_path / "gradient.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"] = "gradient"
    worksheet["A1"].fill = GradientFill(
        type="linear",
        degree=45,
        stop=("FFFF0000", "FF00FF00"),
    )
    workbook.save(source)
    workbook.close()

    view = build_preprocess_view(source)
    fill = view["sheets"][0]["styles"][0]["fill"]

    assert fill["type"] == "linear"
    assert fill["gradient"]["degree"] == 45.0
    assert [stop["color"]["rgb"] for stop in fill["gradient"]["stops"]] == [
        "FFFF0000",
        "FF00FF00",
    ]
    assert fill["gradient"]["truncated"] is False


def test_csv_preprocessing_uses_stdlib_and_preserves_formula(tmp_path: Path) -> None:
    source = tmp_path / "input.csv"
    with source.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["Item", "Value"])
        writer.writerow(["One", "4"])
        writer.writerow(["Total", "=SUM(B2:B2)"])

    view = build_preprocess_view(source)
    sheet = view["sheets"][0]

    assert view["source"]["format"] == "csv"
    assert view["backend"]["reader"] == "python-csv+openpyxl"
    assert sheet["used_region"] == "A1:B3"
    assert _cell(sheet, "B3")["formula"] == "=SUM(B2:B2)"


def test_preprocess_refuses_to_replace_existing_artifacts(tmp_path: Path) -> None:
    source = tmp_path / "features.xlsx"
    _save_feature_workbook(source)
    output = tmp_path / "views"
    output.mkdir()
    (output / "preprocess.json").write_text("keep", encoding="utf-8")

    with pytest.raises(PreprocessError, match="Refusing to overwrite"):
        preprocess_workbook(source, output)

    assert (output / "preprocess.json").read_text(encoding="utf-8") == "keep"


def test_unsupported_format_is_rejected(tmp_path: Path) -> None:
    source = tmp_path / "workbook.txt"
    source.write_text("not a workbook", encoding="utf-8")

    with pytest.raises(PreprocessError, match="Unsupported spreadsheet format"):
        build_preprocess_view(source)


@pytest.mark.parametrize("target_format", ["ods", "xls"])
@pytest.mark.skipif(find_libreoffice() is None, reason="LibreOffice is not installed")
def test_legacy_and_ods_preprocessing_integration(tmp_path: Path, target_format: str) -> None:
    source = tmp_path / "features.xlsx"
    _save_feature_workbook(source)
    converted = convert_spreadsheet_copy(
        source,
        tmp_path / f"converted-{target_format}",
        target_format=target_format,
    )
    before_hash = sha256_file(converted)

    view = build_preprocess_view(converted)

    assert sha256_file(converted) == before_hash
    assert view["source"]["format"] == target_format
    assert view["backend"]["conversion"]["backend"] == "libreoffice-headless"
    assert view["inventory"]["sheet_count"] == 2
