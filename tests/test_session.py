from __future__ import annotations

import hashlib
import re
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from spreadsheet_harness.errors import ToolInputError, WorkbookValidationError
from spreadsheet_harness.evidence_contract import EvidenceScope
from spreadsheet_harness.session import WorkbookSession
from spreadsheet_harness.target_grounding import (
    TargetGroundingMode,
    TargetGroundingRejected,
)


def test_inspect_write_fill_format_and_undo(sample_workbook: Path, tmp_path: Path) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "run")

    listed = session.list_sheets()
    assert [item["name"] for item in listed["sheets"]] == ["Sales", "Lookup"]

    before = session.inspect_range("Sales", "A1:D3")
    assert before["matrix"][1][3] == "=B2*C2"
    assert before["cells"][0]["style"]["font"]["bold"] is True

    result = session.write_range("Sales", "B4", [[5, 1.25, "=B4*C4"]])
    assert result["cells_written"] == 3
    session.fill_formula("Sales", "D2", "D2:D4")
    session.format_range("Sales", "B2:B4", {"number_format": "0", "fill_color": "FFF2CC"})

    changed = session.inspect_range("Sales", "B4:D4")
    assert changed["matrix"][0] == [5, 1.25, "=B4*C4"]

    session.clear_range("Sales", "C4")
    assert session.inspect_range("Sales", "C4:C4")["matrix"] == [[None]]
    session.undo_last()
    assert session.inspect_range("Sales", "C4:C4")["matrix"] == [[1.25]]

    workbook = load_workbook(session.workbook_path, data_only=False)
    assert workbook["Sales"]["D4"].value == "=B4*C4"
    workbook.close()
    assert session.paths.trajectory.is_file()
    assert len(list(session.paths.snapshots.glob("*.xlsx"))) == 2


def test_failed_mutation_keeps_workbook_valid(sample_workbook: Path, tmp_path: Path) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "run")
    original = session.workbook_path.read_bytes()
    with pytest.raises(ToolInputError):
        session.manage_sheet("delete", "missing")
    assert session.workbook_path.read_bytes() == original
    assert session.list_sheets()["sheets"][0]["name"] == "Sales"


def test_failed_mutation_does_not_shadow_the_last_successful_undo_snapshot(
    sample_workbook: Path,
    tmp_path: Path,
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "run")
    original = session.inspect_range("Sales", "B4:B4")["matrix"][0][0]
    session.write_range("Sales", "B4", [[5]])
    snapshots_after_success = tuple(session.paths.snapshots.glob("*.xlsx"))

    with pytest.raises(ToolInputError):
        session.manage_sheet("delete", "missing")

    assert tuple(session.paths.snapshots.glob("*.xlsx")) == snapshots_after_success
    session.undo_last()
    assert session.inspect_range("Sales", "B4:B4")["matrix"] == [[original]]


def test_semantic_noop_does_not_add_an_undo_snapshot(
    sample_workbook: Path,
    tmp_path: Path,
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "run")
    session.write_range("Sales", "B4", [[5]])
    snapshots_after_edit = tuple(session.paths.snapshots.glob("*.xlsx"))

    result = session.write_range("Sales", "B4", [[5]])

    assert result["workbook_effects"]["semantic_changed"] is False
    assert tuple(session.paths.snapshots.glob("*.xlsx")) == snapshots_after_edit


def test_artifact_revisions_bind_reads_and_committed_mutations(
    sample_workbook: Path, tmp_path: Path
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "run")
    initial = session.artifact_ref()
    assert initial.revision == 0
    assert initial.sha256 == hashlib.sha256(session.workbook_path.read_bytes()).hexdigest()

    result = session.write_range("Sales", "B4", [[5]])
    changed = session.artifact_ref()
    inspection = session.inspect_range("Sales", "B4:B4")

    assert changed.revision == 1
    assert result["artifact_revision_before"] == 0
    assert result["artifact_revision_after"] == 1
    assert result["artifact_transition_id"] == 1
    assert inspection["artifact_revision"] == 1
    assert inspection["artifact_sha256"] == changed.sha256
    assert [item.to_dict() for item in session.artifact_transitions] == [
        {
            "transition_id": 1,
            "operation": "write_range",
            "kind": "mutation",
            "before": initial.to_dict(),
            "after": changed.to_dict(),
        }
    ]


def test_failed_mutation_does_not_advance_artifact_revision(
    sample_workbook: Path, tmp_path: Path
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "run")
    initial = session.artifact_ref()

    with pytest.raises(ToolInputError):
        session.manage_sheet("delete", "missing")

    assert session.artifact_ref() == initial
    assert session.artifact_transitions == ()


def test_transition_recorder_fault_restores_bytes_and_keeps_lineage_unchanged(
    sample_workbook: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "run")
    session.enable_target_grounding()
    bytes_before = session.workbook_path.read_bytes()
    artifact_before = session.artifact_ref()
    transitions_before = session.artifact_transitions
    observation = session.record_target_observation(
        artifact=artifact_before,
        scope=EvidenceScope.one("Sales", "B4"),
    )
    declaration = session.declare_edit_target(
        target_scope=EvidenceScope.one("Sales", "B4"),
        observation_ids=[observation["observation_id"]],
    )
    original_record = session.recorder.record

    def fail_transition_record(event: str, payload: dict | None = None) -> None:
        if event == "artifact.transition":
            raise RuntimeError("transition recorder fault")
        original_record(event, payload)

    monkeypatch.setattr(session.recorder, "record", fail_transition_record)

    with pytest.raises(RuntimeError, match="transition recorder fault"):
        session.write_range(
            "Sales",
            "B4",
            [[5]],
            declaration_id=declaration["declaration_id"],
        )

    assert session.workbook_path.read_bytes() == bytes_before
    assert session.artifact_ref() == artifact_before
    assert session.artifact_transitions == transitions_before
    assert session.committed_target_authorizations == ()
    replay = session.write_range(
        "Sales",
        "B4",
        [[6]],
        declaration_id=declaration["declaration_id"],
    )
    assert replay["ok"] is False
    assert replay["type"] == "TargetGroundingError"
    assert replay["target_grounding"]["decision"] == "rejected.replayed_declaration"


def test_commit_recorder_fault_keeps_published_bytes_and_lineage_consistent(
    sample_workbook: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "run")
    session.enable_target_grounding()
    bytes_before = session.workbook_path.read_bytes()
    artifact_before = session.artifact_ref()
    transitions_before = session.artifact_transitions
    observation = session.record_target_observation(
        artifact=artifact_before,
        scope=EvidenceScope.one("Sales", "B4"),
    )
    declaration = session.declare_edit_target(
        target_scope=EvidenceScope.one("Sales", "B4"),
        observation_ids=[observation["observation_id"]],
    )
    original_record = session.recorder.record

    def fail_commit_record(event: str, payload: dict | None = None) -> None:
        if event == "workbook.mutation.committed":
            raise RuntimeError("commit recorder fault")
        original_record(event, payload)

    monkeypatch.setattr(session.recorder, "record", fail_commit_record)

    with pytest.raises(RuntimeError, match="commit recorder fault"):
        session.write_range(
            "Sales",
            "B4",
            [[5]],
            declaration_id=declaration["declaration_id"],
        )

    bytes_after = session.workbook_path.read_bytes()
    artifact_after = session.artifact_ref()
    transitions_after = session.artifact_transitions
    assert bytes_after != bytes_before
    assert artifact_after.revision == artifact_before.revision + 1
    assert artifact_after.sha256 == hashlib.sha256(bytes_after).hexdigest()
    assert len(transitions_after) == len(transitions_before) + 1
    assert transitions_after[-1].before == artifact_before
    assert transitions_after[-1].after == artifact_after
    committed = session.committed_target_authorizations
    assert len(committed) == 1
    assert committed[0].transition == transitions_after[-1]
    assert committed[0].staged_artifact == artifact_after


def test_strict_noop_commit_record_fault_keeps_all_ledgers_unchanged(
    sample_workbook: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "run")
    session.enable_target_grounding()
    artifact_before = session.artifact_ref()
    bytes_before = session.workbook_path.read_bytes()
    observation = session.record_target_observation(
        artifact=artifact_before,
        scope=EvidenceScope.one("Sales", "B4"),
    )
    declaration = session.declare_edit_target(
        target_scope=EvidenceScope.one("Sales", "B4"),
        observation_ids=[observation["observation_id"]],
    )
    original_record = session.recorder.record

    def fail_noop_commit_record(event: str, payload: dict | None = None) -> None:
        if event == "target_grounding.authorization.committed":
            raise RuntimeError("no-op authorization recorder fault")
        original_record(event, payload)

    monkeypatch.setattr(session.recorder, "record", fail_noop_commit_record)

    with pytest.raises(RuntimeError, match="no-op authorization recorder fault"):
        session.run_staged_external_mutation(
            operation="code_interpreter",
            declaration_id=declaration["declaration_id"],
            runner=lambda _path: {"ok": True},
        )

    assert session.workbook_path.read_bytes() == bytes_before
    assert session.artifact_ref() == artifact_before
    assert session.artifact_transitions == ()
    assert session.committed_target_authorizations == ()


def test_advisory_strict_noop_recorder_fault_keeps_all_ledgers_unchanged(
    sample_workbook: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "run")
    session.enable_target_grounding(TargetGroundingMode.ADVISORY)
    artifact_before = session.artifact_ref()
    bytes_before = session.workbook_path.read_bytes()
    transitions_before = session.artifact_transitions
    assessments_before = session.committed_advisory_target_assessments
    original_record = session.recorder.record

    def fail_noop_commit_record(event: str, payload: dict | None = None) -> None:
        if event == "target_grounding.advisory_assessment.committed":
            raise RuntimeError("advisory no-op recorder fault")
        original_record(event, payload)

    monkeypatch.setattr(session.recorder, "record", fail_noop_commit_record)

    with pytest.raises(RuntimeError, match="advisory no-op recorder fault"):
        session.run_staged_external_mutation(
            operation="code_interpreter",
            declaration_id=None,
            runner=lambda _path: {"ok": True},
        )

    assert session.workbook_path.read_bytes() == bytes_before
    assert session.artifact_ref() == artifact_before
    assert session.artifact_transitions == transitions_before
    assert session.committed_advisory_target_assessments == assessments_before
    assert session.committed_target_authorizations == ()


def test_advisory_opaque_staging_publishes_missing_declaration_with_real_footprint(
    sample_workbook: Path, tmp_path: Path
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "run")
    session.enable_target_grounding(TargetGroundingMode.ADVISORY)

    def mutate(staged: Path) -> dict[str, object]:
        workbook = load_workbook(staged)
        workbook["Sales"]["B4"] = 17
        workbook.save(staged)
        workbook.close()
        return {"ok": True, "stdout": "edited staged copy"}

    result = session.run_staged_external_mutation(
        operation="code_interpreter",
        declaration_id=None,
        runner=mutate,
    )

    assert result["ok"] is True
    assert result["mutation_published"] is True
    assert session.inspect_range("Sales", "B4")["matrix"] == [[17]]
    assert session.committed_target_authorizations == ()
    committed = session.committed_advisory_target_assessments[-1]
    assert committed.assessment.decision.value == "rejected.missing_declaration"
    assert committed.assessment.footprint.scope == EvidenceScope.one("Sales", "B4")


def test_advisory_undo_publishes_without_declaration_and_preserves_lineage(
    sample_workbook: Path, tmp_path: Path
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "run")
    initial = session.artifact_ref()
    session.enable_target_grounding(TargetGroundingMode.ADVISORY)
    changed = session.write_range("Sales", "B4", [[17]])

    undone = session.undo_last()

    assert changed["ok"] is True
    assert undone["ok"] is True
    assert session.artifact_ref().revision == 2
    assert session.artifact_ref().sha256 == initial.sha256
    assert len(session.artifact_transitions) == 2
    assert [
        item.assessment.decision.value
        for item in session.committed_advisory_target_assessments
    ] == ["rejected.missing_declaration", "rejected.missing_declaration"]
    assert session.committed_target_authorizations == ()


def test_external_reconcile_is_noninterfering_in_advisory_and_enforced_stays_closed(
    sample_workbook: Path,
    tmp_path: Path,
) -> None:
    off = WorkbookSession.create(sample_workbook, tmp_path / "off")
    advisory = WorkbookSession.create(sample_workbook, tmp_path / "advisory")
    enforced = WorkbookSession.create(sample_workbook, tmp_path / "enforced")
    advisory.enable_target_grounding(TargetGroundingMode.ADVISORY)
    enforced.enable_target_grounding(TargetGroundingMode.ENFORCE)
    off_initial = off.artifact_ref()
    advisory_initial = advisory.artifact_ref()
    enforced_initial = enforced.artifact_ref()

    for session in (off, advisory, enforced):
        workbook = load_workbook(session.workbook_path)
        workbook["Sales"]["B4"] = 73
        workbook.save(session.workbook_path)
        workbook.close()

    off_transition = off.reconcile_external_artifact(off_initial, operation="external")
    advisory_transition = advisory.reconcile_external_artifact(
        advisory_initial,
        operation="external",
    )
    with pytest.raises(TargetGroundingRejected):
        enforced.reconcile_external_artifact(
            enforced_initial,
            operation="external",
        )

    assert off_transition is not None
    assert advisory_transition is not None
    assert off.inspect_range("Sales", "B4")["matrix"] == [[73]]
    assert advisory.inspect_range("Sales", "B4")["matrix"] == [[73]]
    assert enforced.inspect_range("Sales", "B4")["matrix"] != [[73]]
    assessment = advisory.committed_advisory_target_assessments[-1]
    assert assessment.assessment.decision.value == "rejected.missing_declaration"
    assert assessment.transition == advisory_transition


@pytest.mark.parametrize(
    "mode",
    [
        TargetGroundingMode.OFF,
        TargetGroundingMode.ADVISORY,
        TargetGroundingMode.ENFORCE,
    ],
)
def test_nonprotected_recalculation_reconcile_never_requires_target_assessment(
    sample_workbook: Path,
    tmp_path: Path,
    mode: TargetGroundingMode,
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / mode.value)
    if mode is not TargetGroundingMode.OFF:
        session.enable_target_grounding(mode)
    before = session.artifact_ref()
    workbook = load_workbook(session.workbook_path)
    workbook["Sales"]["B4"] = 74
    workbook.save(session.workbook_path)
    workbook.close()

    transition = session.reconcile_external_artifact(
        before,
        operation="recalculate",
        kind="derived_recalculation",
    )

    assert transition is not None
    assert transition.kind == "derived_recalculation"
    assert session.inspect_range("Sales", "B4")["matrix"] == [[74]]
    assert session.committed_target_authorizations == ()
    assert session.committed_advisory_target_assessments == ()
    if mode is TargetGroundingMode.ADVISORY:
        assert [
            event.event_type for event in session.advisory_target_lifecycle_events
        ] == ["transition"]


def test_advisory_reconcile_transition_recorder_fault_restores_all_ledgers(
    sample_workbook: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "run")
    session.enable_target_grounding(TargetGroundingMode.ADVISORY)
    before = session.target_grounding_initial_artifact
    assert before is not None
    original_bytes = session.workbook_path.read_bytes()
    workbook = load_workbook(session.workbook_path)
    workbook["Sales"]["B4"] = 73
    workbook.save(session.workbook_path)
    workbook.close()
    original_record = session.recorder.record

    def fail_transition(event: str, payload: dict | None = None) -> None:
        if event == "artifact.transition":
            raise RuntimeError("advisory transition recorder fault")
        original_record(event, payload)

    monkeypatch.setattr(session.recorder, "record", fail_transition)

    with pytest.raises(RuntimeError, match="advisory transition recorder fault"):
        session.reconcile_external_artifact(before, operation="external")

    assert session.workbook_path.read_bytes() == original_bytes
    assert session.artifact_ref() == before
    assert session.artifact_transitions == ()
    assert session.committed_advisory_target_assessments == ()


def test_undo_transition_recorder_fault_restores_pre_undo_artifact(
    sample_workbook: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "run")
    session.write_range("Sales", "B4", [[5]])
    bytes_before = session.workbook_path.read_bytes()
    artifact_before = session.artifact_ref()
    transitions_before = session.artifact_transitions
    original_record = session.recorder.record

    def fail_transition_record(event: str, payload: dict | None = None) -> None:
        if event == "artifact.transition":
            raise RuntimeError("undo transition recorder fault")
        original_record(event, payload)

    monkeypatch.setattr(session.recorder, "record", fail_transition_record)

    with pytest.raises(RuntimeError, match="undo transition recorder fault"):
        session.undo_last()

    assert session.workbook_path.read_bytes() == bytes_before
    assert session.artifact_ref() == artifact_before
    assert session.artifact_transitions == transitions_before


def test_undo_publishes_new_lineage_revision_even_when_bytes_match_initial(
    sample_workbook: Path, tmp_path: Path
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "run")
    initial = session.artifact_ref()
    session.write_range("Sales", "B4", [[5]])

    result = session.undo_last()
    restored = session.artifact_ref()

    assert restored.revision == 2
    assert restored.sha256 == initial.sha256
    assert result["artifact_revision_before"] == 1
    assert result["artifact_revision_after"] == 2
    assert session.artifact_transitions[-1].kind == "undo"


def test_external_write_must_be_reconciled_into_artifact_lineage(
    sample_workbook: Path, tmp_path: Path
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "run")
    initial = session.artifact_ref()
    workbook = load_workbook(session.workbook_path)
    workbook["Sales"]["A2"] = "external"
    workbook.save(session.workbook_path)
    workbook.close()

    with pytest.raises(WorkbookValidationError, match="outside a recorded"):
        session.artifact_ref()

    transition = session.reconcile_external_artifact(
        initial,
        operation="code_interpreter",
    )
    assert transition is not None
    assert transition.before == initial
    assert transition.after.revision == 1
    assert session.artifact_ref() == transition.after


def test_reconcile_rejects_and_restores_invalid_external_bytes(
    sample_workbook: Path,
    tmp_path: Path,
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "run")
    bytes_before = session.workbook_path.read_bytes()
    artifact_before = session.artifact_ref()
    transitions_before = session.artifact_transitions
    session.workbook_path.write_bytes(b"not an OOXML workbook")

    with pytest.raises(WorkbookValidationError, match="validation failed"):
        session.reconcile_external_artifact(
            artifact_before,
            operation="code_interpreter",
        )

    assert session.workbook_path.read_bytes() == bytes_before
    assert session.artifact_ref() == artifact_before
    assert session.artifact_transitions == transitions_before


def test_reconcile_recorder_fault_restores_external_bytes_and_lineage(
    sample_workbook: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "run")
    bytes_before = session.workbook_path.read_bytes()
    artifact_before = session.artifact_ref()
    transitions_before = session.artifact_transitions
    workbook = load_workbook(session.workbook_path)
    workbook["Sales"]["A2"] = "uncommitted external edit"
    workbook.save(session.workbook_path)
    workbook.close()
    original_record = session.recorder.record

    def fail_transition_record(event: str, payload: dict | None = None) -> None:
        if event == "artifact.transition":
            raise RuntimeError("external transition recorder fault")
        original_record(event, payload)

    monkeypatch.setattr(session.recorder, "record", fail_transition_record)

    with pytest.raises(RuntimeError, match="external transition recorder fault"):
        session.reconcile_external_artifact(
            artifact_before,
            operation="code_interpreter",
        )

    assert session.workbook_path.read_bytes() == bytes_before
    assert session.artifact_ref() == artifact_before
    assert session.artifact_transitions == transitions_before


def test_reconcile_cache_advances_only_after_successful_publication(
    sample_workbook: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "run")
    initial = session.artifact_ref()
    workbook = load_workbook(session.workbook_path)
    workbook["Sales"]["A2"] = "first committed external edit"
    workbook.save(session.workbook_path)
    workbook.close()
    first_transition = session.reconcile_external_artifact(
        initial,
        operation="code_interpreter",
    )
    assert first_transition is not None
    committed_bytes = session.workbook_path.read_bytes()
    committed_artifact = session.artifact_ref()
    committed_transitions = session.artifact_transitions

    workbook = load_workbook(session.workbook_path)
    workbook["Sales"]["A3"] = "second uncommitted external edit"
    workbook.save(session.workbook_path)
    workbook.close()
    original_record = session.recorder.record

    def fail_transition_record(event: str, payload: dict | None = None) -> None:
        if event == "artifact.transition":
            raise RuntimeError("second transition recorder fault")
        original_record(event, payload)

    monkeypatch.setattr(session.recorder, "record", fail_transition_record)

    with pytest.raises(RuntimeError, match="second transition recorder fault"):
        session.reconcile_external_artifact(
            committed_artifact,
            operation="code_interpreter",
        )

    assert session.workbook_path.read_bytes() == committed_bytes
    assert session.artifact_ref() == committed_artifact
    assert session.artifact_transitions == committed_transitions


def test_recalculate_passes_exact_timeout_to_backend(
    sample_workbook: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "run")
    observed: list[float] = []

    def fake_recalculate(
        source: Path,
        destination: Path,
        *,
        timeout_seconds: float,
    ) -> dict[str, object]:
        assert source == destination == session.workbook_path
        observed.append(timeout_seconds)
        digest = hashlib.sha256(source.read_bytes()).hexdigest()
        return {
            "backend": "test-recalculator",
            "source_sha256": digest,
            "output_sha256": digest,
            "atomic_replace": True,
        }

    monkeypatch.setattr(
        "spreadsheet_harness.render.recalculate_workbook",
        fake_recalculate,
    )

    result = session.recalculate(timeout_seconds=3.25)

    assert observed == [3.25]
    assert result["workbook_changed"] is False
    assert result["artifact_transition_id"] is None


def test_range_limits(sample_workbook: Path, tmp_path: Path) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "run")
    with pytest.raises(ToolInputError, match="limit"):
        session.inspect_range("Sales", "A1:Z1000")
    with pytest.raises(ToolInputError, match="rectangular"):
        session.write_range("Sales", "A1", [[1], [2, 3]])


def test_write_range_rejects_high_confidence_unprefixed_formula_text(
    sample_workbook: Path,
    tmp_path: Path,
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "run")
    original = session.workbook_path.read_bytes()

    with pytest.raises(ToolInputError, match="strings beginning with '='"):
        session.write_range("Sales", "H6", [["AVERAGE($B2:$D2)+$F$1"]])

    assert session.workbook_path.read_bytes() == original


def test_write_range_accepts_formula_and_explicit_or_weak_text(
    sample_workbook: Path,
    tmp_path: Path,
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "run")

    result = session.write_range(
        "Sales",
        "H6",
        [["=AVERAGE($B2:$D2)+$F$1", "SUM(A1:A3)", "SUM of actuals"]],
    )

    assert result["ok"] is True
    workbook = load_workbook(session.workbook_path, data_only=False)
    try:
        assert workbook["Sales"]["H6"].data_type == "f"
        assert workbook["Sales"]["I6"].value == "SUM(A1:A3)"
        assert workbook["Sales"]["J6"].value == "SUM of actuals"
    finally:
        workbook.close()


def test_inspect_range_reports_tables(sample_workbook: Path, tmp_path: Path) -> None:
    workbook = load_workbook(sample_workbook)
    sheet = workbook["Sales"]
    table = Table(displayName="SalesTable", ref="A1:D3")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    sheet.add_table(table)
    workbook.save(sample_workbook)
    workbook.close()

    session = WorkbookSession.create(sample_workbook, tmp_path / "run")
    inspected = session.inspect_range("Sales", "A1:D3")

    assert inspected["tables"] == [{"name": "SalesTable", "ref": "A1:D3"}]


def test_fill_formula_reports_sample_formulas_and_drifting_ranges(
    sample_workbook: Path, tmp_path: Path
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "run")

    session.write_range("Sales", "H6", [["=SUM($E6:G6)"]])
    result = session.fill_formula("Sales", "H6", "H6:J7")

    assert result["sample_formulas"] == [
        {"cell": "H6", "formula": "=SUM($E6:G6)"},
        {"cell": "I6", "formula": "=SUM($E6:H6)"},
        {"cell": "H7", "formula": "=SUM($E7:G7)"},
        {"cell": "J7", "formula": "=SUM($E7:I7)"},
    ]
    assert result["warnings"][0]["type"] == "possible_expanding_or_drifting_range"
    assert result["warnings"][0]["source_range"] == "$E6:G6"
    assert "$E6:$G6" in result["warnings"][0]["message"]


@pytest.mark.parametrize(
    ("source_value", "expected_error"),
    [
        (None, "does not contain a formula"),
        (
            "SUM(A1:A3)",
            "formula-like text without a leading '='; assign an Excel formula string",
        ),
    ],
)
def test_fill_formula_explains_invalid_source(
    sample_workbook: Path,
    tmp_path: Path,
    source_value: str | None,
    expected_error: str,
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "run")
    if source_value is not None:
        session.write_range("Sales", "H6", [[source_value]])

    with pytest.raises(ToolInputError, match=re.escape(expected_error)):
        session.fill_formula("Sales", "H6", "H6:J6")


def test_fill_formula_warns_on_relative_horizontal_range_drift(
    sample_workbook: Path, tmp_path: Path
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "run")

    session.write_range("Sales", "H6", [["=SUM(E6:G6)"]])
    result = session.fill_formula("Sales", "H6", "H6:J6")

    assert result["sample_formulas"] == [
        {"cell": "H6", "formula": "=SUM(E6:G6)"},
        {"cell": "I6", "formula": "=SUM(F6:H6)"},
        {"cell": "J6", "formula": "=SUM(G6:I6)"},
    ]
    assert result["warnings"][0]["source_range"] == "E6:G6"
    assert result["warnings"][0]["examples"][0] == {
        "cell": "I6",
        "translated_range": "F6:H6",
    }


def test_fill_formula_expands_single_target_cell_to_source_to_endpoint(
    sample_workbook: Path, tmp_path: Path
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "run")

    session.write_range("Sales", "H6", [["=SUM($E6:$G6)"]])
    result = session.fill_formula("Sales", "H6", "J6")

    assert result["range"] == "H6:J6"
    assert result["requested_range"] == "J6"
    assert result["target_range_expanded_from_endpoint"] is True
    assert result["cells_filled"] == 3
    assert session.inspect_range("Sales", "H6:J6")["matrix"][0] == [
        "=SUM($E6:$G6)",
        "=SUM($E6:$G6)",
        "=SUM($E6:$G6)",
    ]


def test_fill_formula_does_not_warn_when_range_endpoints_are_anchored(
    sample_workbook: Path, tmp_path: Path
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "run")

    session.write_range("Sales", "H6", [["=SUM($E6:$G6)"]])
    result = session.fill_formula("Sales", "H6", "H6:J7")

    assert result["warnings"] == []
    assert result["sample_formulas"][1] == {"cell": "I6", "formula": "=SUM($E6:$G6)"}
