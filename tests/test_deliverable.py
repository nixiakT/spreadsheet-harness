from __future__ import annotations

import hashlib
import json
import os
import stat
from collections.abc import Mapping
from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook, load_workbook

import spreadsheet_harness.deliverable as deliverable_module
import spreadsheet_harness.ooxml_formula_scan as formula_scan_module
import spreadsheet_harness.trajectory as trajectory_module
from spreadsheet_harness.deliverable import (
    ADVISORY_TARGET_GROUNDING_CERTIFICATE_SCHEMA_VERSION,
    COMPARISON_RESULT_SCHEMA_VERSION,
    DELIVERABLE_CERTIFICATE_SCHEMA_VERSION,
    LEGACY_DELIVERABLE_CERTIFICATE_SCHEMA_VERSION,
    NO_FORMULA_ATTESTATION_SCHEMA_VERSION,
    TARGET_GROUNDING_CERTIFICATE_SCHEMA_VERSION,
    DeliverableValidationError,
    audit_deliverable_certificate,
    finalize_deliverable,
    score_read_only,
    validate_evidence_certificate,
)
from spreadsheet_harness.errors import WorkbookValidationError
from spreadsheet_harness.evidence_contract import (
    PIXEL_SHA256_ALGORITHM,
    ContractSpec,
    EffectKind,
    EventKind,
    EvidenceContractMonitor,
    EvidenceEvent,
    EvidenceScope,
)
from spreadsheet_harness.session import WorkbookSession
from spreadsheet_harness.target_grounding import TargetGroundingMode
from spreadsheet_harness.trajectory import read_trajectory


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _workspace_snapshot(root: Path) -> dict[str, tuple[Any, ...]]:
    snapshot: dict[str, tuple[Any, ...]] = {}
    for path in sorted(root.rglob("*")):
        relative = path.relative_to(root).as_posix()
        metadata = path.lstat()
        if path.is_symlink():
            snapshot[relative] = ("symlink", path.readlink().as_posix())
        elif path.is_file():
            snapshot[relative] = ("file", metadata.st_mode & 0o777, path.read_bytes())
        else:
            snapshot[relative] = ("directory", metadata.st_mode & 0o777)
    return snapshot


def _book(path: Path, *, include_formula: bool = False) -> None:
    workbook = Workbook()
    workbook.active.title = "Sheet1"
    workbook.active["A1"] = 0
    if include_formula:
        workbook.active["B1"] = "=1+1"
    workbook.save(path)
    workbook.close()


def _candidate(
    tmp_path: Path,
    *,
    target_grounding: bool | TargetGroundingMode = False,
    include_formula: bool = False,
) -> tuple[WorkbookSession, dict[str, Any]]:
    source = tmp_path / "source.xlsx"
    _book(source, include_formula=include_formula)
    session = WorkbookSession.create(source, tmp_path / "run", run_id="deliverable-test")
    initial = session.artifact_ref()
    if target_grounding:
        mode = (
            target_grounding
            if isinstance(target_grounding, TargetGroundingMode)
            else TargetGroundingMode.ENFORCE
        )
        session.enable_target_grounding(mode)
        observation = session.record_target_observation(
            artifact=initial,
            scope=EvidenceScope.one("Sheet1", "A1"),
        )
        declaration = session.declare_edit_target(
            target_scope=EvidenceScope.one("Sheet1", "A1"),
            observation_ids=[observation["observation_id"]],
        )
        session.write_range(
            "Sheet1",
            "A1",
            [[42]],
            declaration_id=declaration["declaration_id"],
        )
    else:
        session.write_range("Sheet1", "A1", [[42]])
    candidate = session.artifact_ref()

    spec = ContractSpec.from_mapping(
        {
            "schema_version": 1,
            "rules": [
                {
                    "id": "readback",
                    "trigger": "mutation.committed",
                    "require": {"event": "range.inspected", "artifact": "current"},
                }
            ],
        }
    )
    monitor = EvidenceContractMonitor(spec, initial.sha256)
    monitor.observe(
        EvidenceEvent(
            EventKind.MUTATION_COMMITTED,
            initial.sha256,
            candidate.sha256,
            effects=frozenset({EffectKind.VALUE}),
            scope=EvidenceScope.one("Sheet1", "A1:A1"),
        )
    )
    monitor.observe(
        EvidenceEvent(
            EventKind.RANGE_INSPECTED,
            candidate.sha256,
            scope=EvidenceScope.one("Sheet1", "A1:A1"),
        )
    )
    decision = monitor.submission_decision().to_dict()
    session.enable_completion_attempt_capture()
    attempt = session.capture_completion_attempt(
        stage="solve",
        turn=1,
        response_id="response-1",
        call_id="submit-call-1",
    )
    agent = {
        "final_text": "done",
        "turns": 1,
        "stage": "solve",
        "response_id": "response-1",
        "terminal_tool": "submit_result",
        "observed_terminal_tool": "submit_result",
        "terminal_submissions": 1,
        "terminal_response": {
            "status": "accepted",
            "response_id": "response-1",
            "acknowledgement": {},
            "completion_attempt_id": attempt.attempt_id,
        },
        "completion_attempts": [attempt.to_dict()],
        "evidence_contract": {"status": monitor.status(), "decision": decision},
    }
    return session, agent


def _visual_candidate(
    tmp_path: Path,
    *,
    page_sha256: str,
) -> tuple[WorkbookSession, dict[str, Any]]:
    source = tmp_path / "visual-source.xlsx"
    _book(source)
    session = WorkbookSession.create(
        source,
        tmp_path / "visual-run",
        run_id="visual-deliverable-test",
    )
    initial = session.artifact_ref()
    session.format_range("Sheet1", "A1:A1", {"font": {"bold": True}})
    candidate = session.artifact_ref()
    spec = ContractSpec.from_mapping(
        {
            "schema_version": 1,
            "rules": [
                {
                    "id": "visual",
                    "trigger": "effects.visual_changed",
                    "require_sequence": [
                        {"event": "workbook.rendered", "artifact": "current"},
                        {
                            "event": "rendered_page.viewed",
                            "artifact": "same_render",
                            "scope": "changed_visual_scope",
                        },
                    ],
                }
            ],
        }
    )
    monitor = EvidenceContractMonitor(spec, initial.sha256)
    scope = EvidenceScope.one("Sheet1", "A1:A1")
    render_page = {
        "page_id": "Sheet1:1",
        "page_index": 1,
        "file_sha256": page_sha256,
        "width": 1,
        "height": 1,
        "sheet": "Sheet1",
        "sheet_page": 1,
        "cell_scope": scope.to_dict(),
    }
    monitor.observe(
        EvidenceEvent(
            EventKind.MUTATION_COMMITTED,
            initial.sha256,
            candidate.sha256,
            effects=frozenset({EffectKind.STYLE, EffectKind.VISUAL}),
            scope=scope,
        )
    )
    monitor.observe(
        EvidenceEvent(
            EventKind.WORKBOOK_RENDERED,
            candidate.sha256,
            render_id="candidate-render",
            render_manifest_sha256="a" * 64,
            metadata={
                "producer_tool": "render_workbook",
                "backend": "test-renderer",
                "version": {"renderer": "1"},
                "mode": "per_sheet",
                "dpi": 144,
                "page_count": 1,
                "pages": [render_page],
            },
        )
    )
    monitor.observe(
        EvidenceEvent(
            EventKind.RENDERED_PAGE_VIEWED,
            candidate.sha256,
            scope=scope,
            related_render_id="candidate-render",
            related_render_manifest_sha256="a" * 64,
            page_id="Sheet1:1",
            page_sha256=page_sha256,
            metadata={
                "producer_tool": "view_image",
                "delivery_status": "provider_response_confirmed",
                "confirmation_id": "candidate-confirmation",
                "provider_response_id": "response-visual-view",
                "attachment_file_sha256": page_sha256,
                "page_file_sha256": page_sha256,
                "page_pixel_sha256": "b" * 64,
                "pixel_sha256_algorithm": PIXEL_SHA256_ALGORITHM,
                "width": 1,
                "height": 1,
                "image_mode": "RGB",
                "render_mode": "per_sheet",
                "page_index": 1,
                "sheet": "Sheet1",
                "sheet_page": 1,
                "cell_scope": scope.to_dict(),
            },
        )
    )
    session.enable_completion_attempt_capture()
    attempt = session.capture_completion_attempt(
        stage="solve",
        turn=1,
        response_id="response-visual",
        call_id="submit-call-visual",
    )
    agent = {
        "final_text": "visual done",
        "turns": 1,
        "stage": "solve",
        "response_id": "response-visual",
        "terminal_tool": "submit_result",
        "observed_terminal_tool": "submit_result",
        "terminal_submissions": 1,
        "terminal_response": {
            "status": "accepted",
            "response_id": "response-visual",
            "acknowledgement": {},
            "completion_attempt_id": attempt.attempt_id,
        },
        "completion_attempts": [attempt.to_dict()],
        "evidence_contract": {
            "status": monitor.status(),
            "decision": monitor.submission_decision().to_dict(),
        },
    }
    return session, agent


def _fake_render(
    page_bytes: bytes,
):
    def render(
        workbook_path: Path,
        output_dir: Path,
        *,
        workspace: Path,
        artifact: Any,
        timeout_seconds: float,
        dpi: int,
    ) -> dict[str, Any]:
        del workbook_path, timeout_seconds
        output_dir.mkdir(parents=True)
        page = output_dir / "page-001.png"
        page.write_bytes(page_bytes)
        pages = [
            {
                "index": 1,
                "relative_path": page.relative_to(workspace).as_posix(),
                "sha256": _sha256(page),
                "width": 1,
                "height": 1,
                "sheet": "Sheet1",
                "sheet_page": 1,
            }
        ]
        portable_manifest = {
            "schema_version": "spreadsheet-portable-render-manifest-v1",
            "artifact": artifact.to_dict(),
            "backend": "test-renderer",
            "version": {"renderer": "1"},
            "mode": "per_sheet",
            "dpi": dpi,
            "page_count": 1,
            "pages": pages,
        }
        manifest = output_dir / "portable-render-manifest.json"
        manifest.write_text(
            json.dumps(
                portable_manifest,
                ensure_ascii=True,
                sort_keys=True,
                separators=(",", ":"),
            )
            + "\n",
            encoding="ascii",
        )
        record = {
            **portable_manifest,
            "schema_version": "spreadsheet-final-render-witness-v1",
            "manifest_relative_path": manifest.relative_to(workspace).as_posix(),
            "manifest_sha256": _sha256(manifest),
        }
        return {
            **record,
            "witness_sha256": hashlib.sha256(
                json.dumps(
                    record,
                    ensure_ascii=True,
                    sort_keys=True,
                    separators=(",", ":"),
                ).encode("ascii")
            ).hexdigest(),
        }

    return render


def _unchanged_recalculation(session: WorkbookSession) -> dict[str, Any]:
    from spreadsheet_harness.ooxml_formula_scan import (
        OOXML_FORMULA_SCAN_SCHEMA_VERSION,
        OOXML_NO_FORMULA_BACKEND,
        OOXML_NO_FORMULA_PROFILE,
        scan_ooxml_formulas,
    )

    artifact = session.artifact_ref()
    scan = scan_ooxml_formulas(session.workbook_path)
    return {
        "backend": OOXML_NO_FORMULA_BACKEND,
        "version": OOXML_FORMULA_SCAN_SCHEMA_VERSION,
        "profile": OOXML_NO_FORMULA_PROFILE,
        "format": scan.workbook_format,
        "source_sha256": artifact.sha256,
        "output_sha256": artifact.sha256,
        "atomic_replace": False,
        "publication": "verified_no_write",
        "workbook_sha256_before": artifact.sha256,
        "workbook_sha256_after": artifact.sha256,
        "workbook_changed": False,
        "artifact_revision_before": artifact.revision,
        "artifact_revision_after": artifact.revision,
        "artifact_transition_id": None,
        "formula_scan": scan.to_dict(),
        "workbook_effects": {
            "schema_version": "workbook-effect-diff-v1",
            "semantic_changed": False,
            "complete": True,
            "effects": [],
            "scope": EvidenceScope().to_dict(),
            "formula_scope": EvidenceScope().to_dict(),
            "changed_cell_count": 0,
            "scanned_cell_count": scan.scanned_cell_count,
            "reasons": [],
        },
    }


def _changed_recalculation(session: WorkbookSession) -> dict[str, Any]:
    before = session.artifact_ref()
    workbook = load_workbook(session.workbook_path)
    try:
        workbook.calculation.fullCalcOnLoad = not bool(workbook.calculation.fullCalcOnLoad)
        workbook.save(session.workbook_path)
    finally:
        workbook.close()
    transition = session.reconcile_external_artifact(
        before,
        operation="recalculate",
        kind="derived_recalculation",
    )
    assert transition is not None
    after = session.artifact_ref()
    return {
        "backend": "test-recalculator",
        "version": "1",
        "profile": "isolated",
        "format": "xlsx",
        "source_sha256": before.sha256,
        "output_sha256": after.sha256,
        "atomic_replace": True,
        "workbook_sha256_before": before.sha256,
        "workbook_sha256_after": after.sha256,
        "workbook_changed": True,
        "artifact_revision_before": before.revision,
        "artifact_revision_after": after.revision,
        "artifact_transition_id": transition.transition_id,
        "workbook_effects": {
            "schema_version": "workbook-effect-diff-v1",
            "semantic_changed": False,
            "complete": True,
            "effects": [],
            "scope": EvidenceScope().to_dict(),
            "formula_scope": EvidenceScope().to_dict(),
            "changed_cell_count": 0,
            "scanned_cell_count": 1,
            "reasons": [],
        },
    }


def _semantic_recalculation(session: WorkbookSession) -> dict[str, Any]:
    before = session.artifact_ref()
    workbook = load_workbook(session.workbook_path)
    try:
        workbook.active["A1"] = 99
        workbook.save(session.workbook_path)
    finally:
        workbook.close()
    transition = session.reconcile_external_artifact(
        before,
        operation="recalculate",
        kind="derived_recalculation",
    )
    assert transition is not None
    after = session.artifact_ref()
    return {
        "backend": "test-recalculator",
        "version": "1",
        "profile": "isolated",
        "format": "xlsx",
        "source_sha256": before.sha256,
        "output_sha256": after.sha256,
        "atomic_replace": True,
        "workbook_sha256_before": before.sha256,
        "workbook_sha256_after": after.sha256,
        "workbook_changed": True,
        "artifact_revision_before": before.revision,
        "artifact_revision_after": after.revision,
        "artifact_transition_id": transition.transition_id,
        "workbook_effects": {
            "schema_version": "workbook-effect-diff-v1",
            "semantic_changed": True,
            "complete": True,
            "effects": [EffectKind.VALUE.value],
            "scope": EvidenceScope.one("Sheet1", "A1:A1").to_dict(),
            "formula_scope": EvidenceScope().to_dict(),
            "changed_cell_count": 1,
            "scanned_cell_count": 1,
            "reasons": [],
        },
    }


def _unknown_recalculation(session: WorkbookSession) -> dict[str, Any]:
    metadata = _semantic_recalculation(session)
    metadata["workbook_effects"] = {
        "schema_version": "workbook-effect-diff-v1",
        "semantic_changed": True,
        "complete": False,
        "effects": [EffectKind.UNKNOWN.value],
        "scope": EvidenceScope.workbook().to_dict(),
        "formula_scope": EvidenceScope().to_dict(),
        "changed_cell_count": 0,
        "scanned_cell_count": 0,
        "reasons": ["unsupported OOXML extension content changed"],
    }
    return metadata


def _certificate_digest(document: dict[str, Any]) -> str:
    payload = {key: value for key, value in document.items() if key != "certificate_sha256"}
    encoded = json.dumps(
        payload,
        ensure_ascii=True,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("ascii")
    return hashlib.sha256(encoded).hexdigest()


def _nested_digest(document: dict[str, Any], digest_field: str) -> str:
    payload = {key: value for key, value in document.items() if key != digest_field}
    encoded = json.dumps(
        payload,
        ensure_ascii=True,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("ascii")
    return hashlib.sha256(encoded).hexdigest()


def _witness_digest(document: dict[str, Any]) -> str:
    payload = {key: value for key, value in document.items() if key != "witness_sha256"}
    encoded = json.dumps(
        payload,
        ensure_ascii=True,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("ascii")
    return hashlib.sha256(encoded).hexdigest()


def test_finalization_binds_candidate_final_copy_and_read_only_score(
    tmp_path: Path,
) -> None:
    session, agent = _candidate(tmp_path)

    bundle = finalize_deliverable(
        session,
        agent,
    )

    assert COMPARISON_RESULT_SCHEMA_VERSION == "spreadsheet-comparison-result-v28"
    assert bundle.candidate_artifact == bundle.final_artifact
    assert bundle.certificate["final_artifact"] == bundle.final_artifact.to_dict()
    assert bundle.certificate["scoring_copy"]["sha256"] == bundle.final_artifact.sha256
    assert bundle.certificate["scoring_copy"]["artifact_role"] == "same_revision_replica"
    assert bundle.certificate["scoring_copy"]["creates_artifact_transition"] is False
    assert bundle.certificate["lineage"]["transition_count"] == len(session.artifact_transitions)
    submission = bundle.certificate["candidate"]["submission"]
    assert submission["completion_attempt_id"] == 1
    assert submission["completion_attempt_count"] == 1
    assert (
        submission["completion_attempt_record_sha256"]
        == agent["completion_attempts"][0]["record_sha256"]
    )
    assert str(tmp_path) not in json.dumps(bundle.certificate, ensure_ascii=True)
    assert _sha256(bundle.scoring_copy) == _sha256(session.workbook_path)
    assert bundle.scoring_copy.stat().st_mode & 0o222 == 0
    assert score_read_only(bundle, lambda path: load_workbook(path).active["A1"].value) == 42

    audit = audit_deliverable_certificate(
        bundle.certificate,
        agent_evidence=agent,
        run_root=session.workspace,
        output_workbook=session.workbook_path,
    )
    assert audit.valid is True
    assert audit.final_artifact == bundle.final_artifact
    assert audit.scoring_copy == bundle.scoring_copy


def test_score_read_only_uses_bound_snapshot_across_source_rename_restore(
    tmp_path: Path,
) -> None:
    session, agent = _candidate(tmp_path)
    bundle = finalize_deliverable(session, agent)
    original_sha256 = _sha256(bundle.scoring_copy)
    original_identity = (
        bundle.scoring_copy.stat().st_dev,
        bundle.scoring_copy.stat().st_ino,
    )
    alternate = session.workspace / "alternate-scoring.xlsx"
    displaced = session.workspace / "displaced-scoring.xlsx"
    _book(alternate)
    workbook = load_workbook(alternate)
    try:
        workbook.active["A1"] = -1
        workbook.save(alternate)
    finally:
        workbook.close()
    alternate.chmod(0o400)
    observed: dict[str, Any] = {}

    def score_bound_snapshot(scoring_input: Path) -> Any:
        bundle.scoring_copy.replace(displaced)
        alternate.replace(bundle.scoring_copy)
        try:
            observed["input_path"] = scoring_input
            observed["input_sha256"] = _sha256(scoring_input)
            scored = load_workbook(scoring_input, read_only=True, data_only=True)
            try:
                return scored.active["A1"].value
            finally:
                scored.close()
        finally:
            bundle.scoring_copy.replace(alternate)
            displaced.replace(bundle.scoring_copy)

    assert score_read_only(bundle, score_bound_snapshot) == 42
    assert observed["input_path"] != bundle.scoring_copy
    assert observed["input_sha256"] == original_sha256
    restored = bundle.scoring_copy.stat()
    assert (restored.st_dev, restored.st_ino) == original_identity
    assert _sha256(bundle.scoring_copy) == original_sha256


def test_v28_rejects_missing_completion_attempt_id(tmp_path: Path) -> None:
    session, agent = _candidate(tmp_path)
    del agent["terminal_response"]["completion_attempt_id"]

    with pytest.raises(DeliverableValidationError, match="accepted terminal response"):
        finalize_deliverable(
            session,
            agent,
        )


def test_v28_certifies_noncompletion_without_claiming_accepted_evidence(
    tmp_path: Path,
) -> None:
    session, agent = _candidate(tmp_path)
    agent["final_text"] = "Evidence obligations remained unsatisfied."
    agent["terminal_response"] = None
    agent["evidence_contract"] = {
        "status": {"submission_ready": False},
        "decision": {"allowed": False, "contract_satisfied": False},
    }

    bundle = finalize_deliverable(
        session,
        agent,
    )

    candidate = bundle.certificate["candidate"]
    assert candidate["outcome"] == "audited_noncompletion"
    assert candidate["evidence_certificate"] is None
    assert candidate["submission"]["accepted_submission"] is False
    assert candidate["submission"]["accepted_evidence_certificate"] is False
    assert candidate["submission"]["completion_attempt_count"] == 1
    assert bundle.certificate["evidence_policy"]["accepted_candidate_evidence"] is False
    assert bundle.certificate["evidence_policy"]["candidate_evidence_carried_forward"] is False
    assert audit_deliverable_certificate(
        bundle.certificate,
        agent_evidence=agent,
        run_root=session.workspace,
        output_workbook=session.workbook_path,
    ).valid


def test_v28_certifies_early_stage_noncompletion_with_no_captured_attempts(
    tmp_path: Path,
) -> None:
    session, agent = _candidate(tmp_path)
    agent.update(
        {
            "final_text": "The read-only preprocessing stage failed.",
            "response_id": None,
            "terminal_response": None,
            "terminal_submissions": 0,
            "completion_attempts": [],
            "evidence_contract": None,
            "stages": [
                {
                    "name": "extract",
                    "agent": {
                        "completion_attempts": None,
                        "terminal_response": None,
                    },
                }
            ],
        }
    )

    bundle = finalize_deliverable(
        session,
        agent,
    )

    assert bundle.certificate["candidate"]["outcome"] == "audited_noncompletion"
    assert bundle.certificate["candidate"]["submission"]["completion_attempt_count"] == 0
    assert audit_deliverable_certificate(
        bundle.certificate,
        agent_evidence=agent,
        run_root=session.workspace,
        output_workbook=session.workbook_path,
    ).valid


def test_v28_rejects_unknown_completion_attempt_id(tmp_path: Path) -> None:
    session, agent = _candidate(tmp_path)
    agent["terminal_response"]["completion_attempt_id"] = 99

    with pytest.raises(DeliverableValidationError, match="not bound"):
        finalize_deliverable(
            session,
            agent,
        )


def test_v28_rejects_completion_attempt_response_id_mismatch(tmp_path: Path) -> None:
    session, agent = _candidate(tmp_path)
    attempt = agent["completion_attempts"][0]
    attempt["response_id"] = "different-response"
    attempt["record_sha256"] = _nested_digest(attempt, "record_sha256")

    with pytest.raises(DeliverableValidationError, match="not bound"):
        finalize_deliverable(
            session,
            agent,
        )


def test_v28_rejects_attempt_that_was_not_the_accepted_terminal_call(
    tmp_path: Path,
) -> None:
    session, agent = _candidate(tmp_path)
    second = session.capture_completion_attempt(
        stage="solve",
        turn=2,
        response_id="response-1",
        call_id="submit-call-2",
    )
    agent["completion_attempts"].append(second.to_dict())
    agent["terminal_submissions"] = 2
    agent["turns"] = 2
    agent["terminal_response"]["completion_attempt_id"] = 1

    with pytest.raises(DeliverableValidationError, match="not bound"):
        finalize_deliverable(
            session,
            agent,
        )


@pytest.mark.parametrize(
    ("render_timeout", "recalculation_timeout", "expected_timeout"),
    [
        (7.25, None, 7.25),
        (7.25, 2.5, 2.5),
    ],
)
def test_finalization_passes_exact_recalculation_timeout(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    render_timeout: float,
    recalculation_timeout: float | None,
    expected_timeout: float,
) -> None:
    session, agent = _candidate(tmp_path)
    observed: list[float] = []

    def fake_recalculate(*, timeout_seconds: float) -> dict[str, Any]:
        observed.append(timeout_seconds)
        return _unchanged_recalculation(session)

    monkeypatch.setattr(session, "recalculate_for_finalization", fake_recalculate)

    bundle = finalize_deliverable(
        session,
        agent,
        render_timeout_seconds=render_timeout,
        recalculation_timeout_seconds=recalculation_timeout,
    )

    assert observed == [expected_timeout]
    assert bundle.certificate["postprocess"]["timeout_seconds"] == expected_timeout


def test_default_finalization_uses_hash_bound_no_formula_noop_without_libreoffice(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session, agent = _candidate(tmp_path)
    candidate = session.artifact_ref()
    candidate_bytes = session.workbook_path.read_bytes()
    transitions = session.artifact_transitions

    def forbidden_recalculation(*_: Any, **__: Any) -> dict[str, Any]:
        raise AssertionError("LibreOffice recalculation must not run for a formula-free workbook")

    monkeypatch.setattr(session, "recalculate", forbidden_recalculation)
    monkeypatch.setattr(
        "spreadsheet_harness.render.recalculate_workbook",
        forbidden_recalculation,
    )

    bundle = finalize_deliverable(session, agent)

    assert session.workbook_path.read_bytes() == candidate_bytes
    assert session.artifact_ref() == candidate
    assert session.artifact_transitions == transitions
    assert bundle.candidate_artifact == bundle.final_artifact == candidate
    assert bundle.recalculation == bundle.certificate["postprocess"]["recalculation"]
    assert bundle.recalculation["backend"] == "sheetledger-ooxml-noop"
    assert bundle.recalculation["version"] == "ooxml-formula-scan-v1"
    assert bundle.recalculation["profile"] == "verified-no-formula-byte-identical-v1"
    assert bundle.recalculation["source_sha256"] == candidate.sha256
    assert bundle.recalculation["output_sha256"] == candidate.sha256
    assert bundle.recalculation["workbook_changed"] is False
    assert bundle.recalculation["artifact_transition_id"] is None
    assert audit_deliverable_certificate(
        bundle.certificate,
        agent_evidence=agent,
        run_root=session.workspace,
        output_workbook=session.workbook_path,
    ).valid


def test_no_formula_attestation_is_certificate_bound_and_offline_replayed(
    tmp_path: Path,
) -> None:
    session, agent = _candidate(tmp_path)

    bundle = finalize_deliverable(session, agent)

    certificate = bundle.certificate
    attestation = certificate["recalculation_attestation"]
    recalculation = certificate["postprocess"]["recalculation"]
    assert certificate["schema_version"] == DELIVERABLE_CERTIFICATE_SCHEMA_VERSION
    assert attestation["schema_version"] == NO_FORMULA_ATTESTATION_SCHEMA_VERSION
    assert attestation["artifact"] == certificate["final_artifact"]
    assert attestation["backend"] == recalculation["backend"]
    assert attestation["version"] == recalculation["version"]
    assert attestation["profile"] == recalculation["profile"]
    assert attestation["publication"] == "verified_no_write"
    assert attestation["formula_scan"] == recalculation["formula_scan"]
    assert attestation["formula_scan"]["formula_marker_count"] == 0
    assert attestation["witness_sha256"] == _witness_digest(attestation)
    assert audit_deliverable_certificate(
        certificate,
        agent_evidence=agent,
        run_root=session.workspace,
        output_workbook=session.workbook_path,
    ).valid


def test_offline_audit_executes_independent_formula_rescan(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session, agent = _candidate(tmp_path)
    bundle = finalize_deliverable(session, agent)
    scan_called = False

    def fail_independent_scan(stage: str, _: Path) -> None:
        nonlocal scan_called
        if stage == "certificate_audit":
            scan_called = True
            raise formula_scan_module.OOXMLFormulaScanError("injected rescan failure")

    monkeypatch.setattr(
        formula_scan_module,
        "_formula_scan_lease_hook",
        fail_independent_scan,
    )

    audit = audit_deliverable_certificate(
        bundle.certificate,
        agent_evidence=agent,
        run_root=session.workspace,
        output_workbook=session.workbook_path,
    )

    assert scan_called
    assert not audit.valid


def test_session_finalization_rejects_parent_symlink_swap_while_lease_is_open(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session, agent = _candidate(tmp_path)
    parent = session.workbook_path.parent
    displaced = session.workspace / "displaced-artifacts"
    swapped = False

    def swap_parent(stage: str, _: Path) -> None:
        nonlocal swapped
        if stage == "session_finalization" and not swapped:
            swapped = True
            parent.rename(displaced)
            parent.symlink_to(displaced.name, target_is_directory=True)

    monkeypatch.setattr(
        formula_scan_module,
        "_formula_scan_lease_hook",
        swap_parent,
    )

    with pytest.raises(WorkbookValidationError, match="cannot certify"):
        finalize_deliverable(session, agent)

    assert swapped
    assert parent.is_symlink()
    assert (displaced / session.workbook_path.name).is_file()
    assert not (session.workspace / "scoring-output.xlsx").exists()


def test_deliverable_finalization_rejects_real_parent_replacement_during_lease(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session, agent = _candidate(tmp_path)
    parent = session.workbook_path.parent
    displaced = session.workspace / "displaced-artifacts"
    original_bytes = session.workbook_path.read_bytes()
    replacement_identity: tuple[int, int] | None = None

    def replace_parent(stage: str, _: Path) -> None:
        nonlocal replacement_identity
        if stage == "deliverable_finalization" and replacement_identity is None:
            parent.rename(displaced)
            parent.mkdir()
            session.workbook_path.write_bytes(original_bytes)
            metadata = parent.lstat()
            replacement_identity = (metadata.st_dev, metadata.st_ino)

    monkeypatch.setattr(
        formula_scan_module,
        "_formula_scan_lease_hook",
        replace_parent,
    )

    with pytest.raises(DeliverableValidationError, match="descriptor-bound inspection"):
        finalize_deliverable(session, agent)

    observed = parent.lstat()
    assert replacement_identity == (observed.st_dev, observed.st_ino)
    assert session.workbook_path.read_bytes() == original_bytes
    assert (displaced / session.workbook_path.name).read_bytes() == original_bytes
    assert not (session.workspace / "scoring-output.xlsx").exists()


def test_formula_scan_lease_finalization_checkpoints_valid_control(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session, agent = _candidate(tmp_path)
    observed_stages: list[str] = []

    def observe(stage: str, _: Path) -> None:
        observed_stages.append(stage)

    monkeypatch.setattr(
        formula_scan_module,
        "_formula_scan_lease_hook",
        observe,
    )

    bundle = finalize_deliverable(session, agent)
    audit = audit_deliverable_certificate(
        bundle.certificate,
        agent_evidence=agent,
        run_root=session.workspace,
        output_workbook=session.workbook_path,
    )

    assert audit.valid
    assert "session_finalization" in observed_stages
    assert "deliverable_finalization" in observed_stages
    assert "certificate_audit" in observed_stages
    assert observed_stages.count("before_lease_exit") >= 3


@pytest.mark.parametrize(
    "mutation",
    [
        "backend",
        "version",
        "profile",
        "atomic_replace",
        "publication",
        "transition",
        "effects",
        "formula_scan",
    ],
)
def test_offline_audit_rejects_rehashed_no_formula_invariant_tampering(
    tmp_path: Path,
    mutation: str,
) -> None:
    session, agent = _candidate(tmp_path)
    bundle = finalize_deliverable(session, agent)
    tampered = json.loads(json.dumps(bundle.certificate))
    recalculation = tampered["postprocess"]["recalculation"]
    attestation = tampered["recalculation_attestation"]
    if mutation in {"backend", "version", "profile"}:
        recalculation[mutation] = f"forged-{mutation}"
        attestation[mutation] = f"forged-{mutation}"
    elif mutation == "atomic_replace":
        recalculation["atomic_replace"] = True
    elif mutation == "publication":
        recalculation["publication"] = "atomic_replace"
        attestation["publication"] = "atomic_replace"
    elif mutation == "transition":
        recalculation["artifact_transition_id"] = 999
    elif mutation == "effects":
        effects = recalculation["workbook_effects"]
        effects["semantic_changed"] = True
        effects["effects"] = [EffectKind.VALUE.value]
        effects["scope"] = EvidenceScope.one("Sheet1", "A1:A1").to_dict()
        effects["changed_cell_count"] = 1
    else:
        for scan in (recalculation["formula_scan"], attestation["formula_scan"]):
            scan["formula_marker_count"] = 1
            scan["formula_kinds"] = ["f"]
    attestation["witness_sha256"] = _witness_digest(attestation)
    tampered["certificate_sha256"] = _certificate_digest(tampered)

    assert not audit_deliverable_certificate(
        tampered,
        agent_evidence=agent,
        run_root=session.workspace,
        output_workbook=session.workbook_path,
    ).valid


def test_new_no_formula_certificate_cannot_be_replayed_as_legacy_v2(
    tmp_path: Path,
) -> None:
    session, agent = _candidate(tmp_path)
    bundle = finalize_deliverable(session, agent)
    downgraded = json.loads(json.dumps(bundle.certificate))
    downgraded["schema_version"] = LEGACY_DELIVERABLE_CERTIFICATE_SCHEMA_VERSION
    del downgraded["recalculation_attestation"]
    downgraded["certificate_sha256"] = _certificate_digest(downgraded)

    assert not audit_deliverable_certificate(
        downgraded,
        agent_evidence=agent,
        run_root=session.workspace,
        output_workbook=session.workbook_path,
    ).valid


def test_offline_audit_keeps_explicit_legacy_v2_certificate_compatible(
    tmp_path: Path,
) -> None:
    session, agent = _candidate(tmp_path)
    bundle = finalize_deliverable(session, agent)
    legacy = json.loads(json.dumps(bundle.certificate))
    legacy["schema_version"] = LEGACY_DELIVERABLE_CERTIFICATE_SCHEMA_VERSION
    del legacy["recalculation_attestation"]
    recalculation = legacy["postprocess"]["recalculation"]
    recalculation.update(
        {
            "backend": "legacy-test-recalculator",
            "version": "1",
            "profile": "isolated",
            "atomic_replace": True,
        }
    )
    del recalculation["publication"]
    del recalculation["formula_scan"]
    legacy["certificate_sha256"] = _certificate_digest(legacy)

    assert audit_deliverable_certificate(
        legacy,
        agent_evidence=agent,
        run_root=session.workspace,
        output_workbook=session.workbook_path,
    ).valid


def test_formula_finalization_fails_before_mutating_bytes_revision_or_trajectory(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session, agent = _candidate(tmp_path, include_formula=True)
    candidate = session.artifact_ref()
    candidate_bytes = session.workbook_path.read_bytes()
    transitions = session.artifact_transitions
    trajectory_bytes = session.paths.trajectory.read_bytes()
    snapshots = tuple(path.name for path in session.paths.snapshots.iterdir())

    def forbidden_recalculation(*_: Any, **__: Any) -> dict[str, Any]:
        raise AssertionError("formula hotfix must fail before invoking LibreOffice")

    monkeypatch.setattr(
        "spreadsheet_harness.render.recalculate_workbook",
        forbidden_recalculation,
    )

    with pytest.raises(
        WorkbookValidationError,
        match="cache-preserving transactional recalculation backend",
    ):
        finalize_deliverable(session, agent)

    assert session.workbook_path.read_bytes() == candidate_bytes
    assert session.artifact_ref() == candidate
    assert session.artifact_transitions == transitions
    assert session.paths.trajectory.read_bytes() == trajectory_bytes
    assert tuple(path.name for path in session.paths.snapshots.iterdir()) == snapshots
    assert not (session.workspace / "scoring").exists()


def test_formula_finalization_does_not_trust_callback_noop_metadata(
    tmp_path: Path,
) -> None:
    session, agent = _candidate(tmp_path, include_formula=True)
    candidate = session.artifact_ref()
    candidate_bytes = session.workbook_path.read_bytes()
    transitions = session.artifact_transitions
    trajectory_bytes = session.paths.trajectory.read_bytes()
    callback_called = False

    def forged_noop() -> dict[str, Any]:
        nonlocal callback_called
        callback_called = True
        return _unchanged_recalculation(session)

    with pytest.raises(
        DeliverableValidationError,
        match="callbacks are disabled",
    ):
        finalize_deliverable(
            session,
            agent,
            recalculation_callback=forged_noop,
        )

    assert callback_called is False
    assert session.workbook_path.read_bytes() == candidate_bytes
    assert session.artifact_ref() == candidate
    assert session.artifact_transitions == transitions
    assert session.paths.trajectory.read_bytes() == trajectory_bytes


def test_final_certificate_exposes_and_replays_committed_target_authorization(
    tmp_path: Path,
) -> None:
    session, agent = _candidate(tmp_path, target_grounding=True)

    bundle = finalize_deliverable(
        session,
        agent,
    )

    grounding = bundle.certificate["target_grounding"]
    assert grounding["enabled"] is True
    assert grounding["initial_artifact"]["revision"] == 0
    assert grounding["initial_transition_count"] == 0
    assert grounding["authorization_count"] == 1
    authorization = grounding["authorizations"][0]
    assert authorization["publication"]["kind"] == "artifact_transition"
    assert (
        authorization["publication"]["transition"]
        == bundle.certificate["lineage"]["transitions"][0]
    )
    assert grounding["authorization_chain_head_sha256"] == authorization["authorization_sha256"]
    trajectory = [json.loads(line) for line in session.paths.trajectory.read_text().splitlines()]
    durable = [
        json.loads(row["payload"]["target_grounding_commit_json"])
        for row in trajectory
        if row["event"] == "artifact.transition"
        and "target_grounding_commit_json" in row["payload"]
    ]
    assert durable == [authorization]
    assert audit_deliverable_certificate(
        bundle.certificate,
        agent_evidence=agent,
        run_root=session.workspace,
        output_workbook=session.workbook_path,
    ).valid


def test_advisory_certificate_exposes_and_freshly_replays_observer_ledger(
    tmp_path: Path,
) -> None:
    session, agent = _candidate(
        tmp_path,
        target_grounding=TargetGroundingMode.ADVISORY,
    )

    bundle = finalize_deliverable(
        session,
        agent,
    )

    grounding = bundle.certificate["target_grounding"]
    assert grounding["schema_version"] == (ADVISORY_TARGET_GROUNDING_CERTIFICATE_SCHEMA_VERSION)
    assert grounding["mode"] == TargetGroundingMode.ADVISORY.value
    assert grounding["active"] is True
    assert grounding["enforced"] is False
    assert grounding["initial_artifact"]["revision"] == 0
    assert grounding["initial_transition_count"] == 0
    assert grounding["assessment_count"] == 1
    assessment = grounding["assessments"][0]
    assert assessment["assessment"]["declaration_status"] == "valid"
    assert assessment["assessment"]["counterfactual_enforcement_decision"] == "authorized"
    assert (
        assessment["publication"]["transition"] == (bundle.certificate["lineage"]["transitions"][0])
    )
    assert grounding["assessment_chain_head_sha256"] == assessment["commitment_sha256"]
    assert grounding["lifecycle_replay_complete"] is True
    assert grounding["lifecycle_event_count"] == len(grounding["lifecycle_events"])
    assert grounding["lifecycle_event_count"] >= 4
    assert grounding["lifecycle_events"][0]["event_type"] == "observation"
    assert grounding["lifecycle_events"][-1]["event_type"] == "commitment"
    assert (
        grounding["lifecycle_chain_head_sha256"]
        == (grounding["lifecycle_events"][-1]["event_sha256"])
    )
    assert grounding["lifecycle_final_counters"]["pending_preparation_count"] == 0
    assert grounding["lifecycle_final_counters"]["commitment_count"] == 1
    assert session.committed_target_authorizations == ()
    assert audit_deliverable_certificate(
        bundle.certificate,
        agent_evidence=agent,
        run_root=session.workspace,
        output_workbook=session.workbook_path,
    ).valid


@pytest.mark.parametrize(
    ("target_grounding", "enabled"),
    [
        (False, False),
        (TargetGroundingMode.ENFORCE, True),
    ],
)
def test_off_and_enforce_certificates_preserve_v1_compatibility(
    tmp_path: Path,
    target_grounding: bool | TargetGroundingMode,
    enabled: bool,
) -> None:
    session, agent = _candidate(tmp_path, target_grounding=target_grounding)

    bundle = finalize_deliverable(
        session,
        agent,
    )

    grounding = bundle.certificate["target_grounding"]
    assert grounding["schema_version"] == TARGET_GROUNDING_CERTIFICATE_SCHEMA_VERSION
    assert grounding["enabled"] is enabled
    assert "mode" not in grounding
    assert "assessments" not in grounding
    assert audit_deliverable_certificate(
        bundle.certificate,
        agent_evidence=agent,
        run_root=session.workspace,
        output_workbook=session.workbook_path,
    ).valid


def test_fresh_audit_rejects_rehashed_missing_advisory_assessment(
    tmp_path: Path,
) -> None:
    session, agent = _candidate(
        tmp_path,
        target_grounding=TargetGroundingMode.ADVISORY,
    )
    bundle = finalize_deliverable(
        session,
        agent,
    )
    tampered = json.loads(json.dumps(bundle.certificate))
    grounding = tampered["target_grounding"]
    grounding["assessments"] = []
    grounding["assessment_count"] = 0
    grounding["assessment_chain_head_sha256"] = "0" * 64
    tampered["certificate_sha256"] = _certificate_digest(tampered)

    assert not audit_deliverable_certificate(
        tampered,
        agent_evidence=agent,
        run_root=session.workspace,
        output_workbook=session.workbook_path,
    ).valid


def test_fresh_audit_rejects_rehashed_advisory_assurance_tampering(
    tmp_path: Path,
) -> None:
    session, agent = _candidate(
        tmp_path,
        target_grounding=TargetGroundingMode.ADVISORY,
    )
    bundle = finalize_deliverable(
        session,
        agent,
    )
    tampered = json.loads(json.dumps(bundle.certificate))
    grounding = tampered["target_grounding"]
    commitment = grounding["assessments"][0]
    assessment = commitment["assessment"]
    assessment["assurance"]["authorized_publication"] = True
    assessment["assessment_sha256"] = _nested_digest(
        assessment,
        "assessment_sha256",
    )
    commitment["commitment_sha256"] = _nested_digest(
        commitment,
        "commitment_sha256",
    )
    grounding["assessment_chain_head_sha256"] = commitment["commitment_sha256"]
    tampered["certificate_sha256"] = _certificate_digest(tampered)

    assert not audit_deliverable_certificate(
        tampered,
        agent_evidence=agent,
        run_root=session.workspace,
        output_workbook=session.workbook_path,
    ).valid


def test_strict_noop_has_one_durable_authorization_without_a_transition(
    tmp_path: Path,
) -> None:
    session, agent = _candidate(tmp_path, target_grounding=True)
    artifact = session.artifact_ref()
    observation = session.record_target_observation(
        artifact=artifact,
        scope=EvidenceScope.one("Sheet1", "A1"),
    )
    declaration = session.declare_edit_target(
        target_scope=EvidenceScope.one("Sheet1", "A1"),
        observation_ids=[observation["observation_id"]],
    )
    transition_count = len(session.artifact_transitions)

    result = session.run_staged_external_mutation(
        operation="code_interpreter",
        declaration_id=declaration["declaration_id"],
        runner=lambda _path: {"ok": True},
    )

    assert result["target_grounding"]["decision"] == "authorized_no_op"
    assert len(session.artifact_transitions) == transition_count
    bundle = finalize_deliverable(
        session,
        agent,
    )
    grounding = bundle.certificate["target_grounding"]
    assert grounding["authorization_count"] == 2
    no_op = grounding["authorizations"][1]
    assert no_op["publication"] == {"kind": "strict_no_op", "transition": None}
    assert no_op["staged_artifact"] == artifact.to_dict()
    assert (
        no_op["previous_authorization_sha256"]
        == grounding["authorizations"][0]["authorization_sha256"]
    )
    assert audit_deliverable_certificate(
        bundle.certificate,
        agent_evidence=agent,
        run_root=session.workspace,
        output_workbook=session.workbook_path,
    ).valid


def test_fresh_audit_rejects_rehashed_target_scope_tampering(tmp_path: Path) -> None:
    session, agent = _candidate(tmp_path, target_grounding=True)
    bundle = finalize_deliverable(
        session,
        agent,
    )
    tampered = json.loads(json.dumps(bundle.certificate))
    authorization = tampered["target_grounding"]["authorizations"][0]
    target_range = authorization["provenance"]["declaration"]["target_scope"]["ranges"][0]
    target_range.update({"range": "B1:B1", "cell_count": 1})
    provenance = authorization["provenance"]
    provenance["provenance_sha256"] = _nested_digest(provenance, "provenance_sha256")
    authorization["authorization_sha256"] = _nested_digest(authorization, "authorization_sha256")
    tampered["target_grounding"]["authorization_chain_head_sha256"] = authorization[
        "authorization_sha256"
    ]
    tampered["certificate_sha256"] = _certificate_digest(tampered)

    assert not audit_deliverable_certificate(
        tampered,
        agent_evidence=agent,
        run_root=session.workspace,
        output_workbook=session.workbook_path,
    ).valid


def test_fresh_audit_rejects_missing_committed_authorization(tmp_path: Path) -> None:
    session, agent = _candidate(tmp_path, target_grounding=True)
    bundle = finalize_deliverable(
        session,
        agent,
    )
    tampered = json.loads(json.dumps(bundle.certificate))
    grounding = tampered["target_grounding"]
    grounding["authorizations"] = []
    grounding["authorization_count"] = 0
    grounding["authorization_chain_head_sha256"] = "0" * 64
    tampered["certificate_sha256"] = _certificate_digest(tampered)

    assert not audit_deliverable_certificate(
        tampered,
        agent_evidence=agent,
        run_root=session.workspace,
        output_workbook=session.workbook_path,
    ).valid


def test_fresh_audit_rejects_duplicate_committed_authorization(tmp_path: Path) -> None:
    session, agent = _candidate(tmp_path, target_grounding=True)
    bundle = finalize_deliverable(
        session,
        agent,
    )
    tampered = json.loads(json.dumps(bundle.certificate))
    grounding = tampered["target_grounding"]
    grounding["authorizations"].append(json.loads(json.dumps(grounding["authorizations"][0])))
    grounding["authorization_count"] = 2
    grounding["authorization_chain_head_sha256"] = grounding["authorizations"][1][
        "authorization_sha256"
    ]
    tampered["certificate_sha256"] = _certificate_digest(tampered)

    assert not audit_deliverable_certificate(
        tampered,
        agent_evidence=agent,
        run_root=session.workspace,
        output_workbook=session.workbook_path,
    ).valid


def test_fresh_audit_rejects_authorization_hash_cycle(tmp_path: Path) -> None:
    session, agent = _candidate(tmp_path, target_grounding=True)
    bundle = finalize_deliverable(
        session,
        agent,
    )
    tampered = json.loads(json.dumps(bundle.certificate))
    authorization = tampered["target_grounding"]["authorizations"][0]
    authorization["previous_authorization_sha256"] = authorization["authorization_sha256"]
    authorization["authorization_sha256"] = _nested_digest(authorization, "authorization_sha256")
    tampered["target_grounding"]["authorization_chain_head_sha256"] = authorization[
        "authorization_sha256"
    ]
    tampered["certificate_sha256"] = _certificate_digest(tampered)

    assert not audit_deliverable_certificate(
        tampered,
        agent_evidence=agent,
        run_root=session.workspace,
        output_workbook=session.workbook_path,
    ).valid


def test_candidate_certificate_cannot_replay_after_artifact_hash_cycle(
    tmp_path: Path,
) -> None:
    session, agent = _candidate(tmp_path)
    certified_bytes = session.workbook_path.read_bytes()
    certified_sha256 = _sha256(session.workbook_path)

    session.write_range("Sheet1", "A1", [[43]])
    before_restore = session.artifact_ref()
    session.workbook_path.write_bytes(certified_bytes)
    transition = session.reconcile_external_artifact(
        before_restore,
        operation="test_restore_old_bytes",
    )

    assert transition is not None
    assert session.artifact_ref().sha256 == certified_sha256
    assert session.artifact_ref().revision == 3
    assert agent["evidence_contract"]["decision"]["certificate"]["revision_index"] == 1

    with pytest.raises(ValueError, match="stale for the managed artifact revision"):
        finalize_deliverable(
            session,
            agent,
        )


@pytest.mark.parametrize(
    ("field", "value", "message"),
    [
        ("terminal_tool", "assistant_text", "terminal submission"),
        ("observed_terminal_tool", "assistant_text", "terminal submission"),
        ("terminal_submissions", 0, "terminal submission"),
        ("response_id", "different-response", "accepted terminal response"),
    ],
)
def test_finalization_rejects_unbound_terminal_evidence(
    tmp_path: Path,
    field: str,
    value: Any,
    message: str,
) -> None:
    session, agent = _candidate(tmp_path)
    agent[field] = value

    with pytest.raises(ValueError, match=message):
        finalize_deliverable(
            session,
            agent,
        )


def test_scoring_replica_creation_cannot_publish_an_artifact_transition(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session, agent = _candidate(tmp_path)
    original_copy = deliverable_module._atomic_scoring_copy

    def copy_with_unexpected_transition(
        source: Path,
        destination: Path,
    ) -> Any:
        publication = original_copy(source, destination)
        with monkeypatch.context() as scoped_patch:
            scoped_patch.setattr(session.recorder, "record", lambda *_args, **_kwargs: None)
            session.write_range("Sheet1", "A1", [[43]])
        return publication

    monkeypatch.setattr(
        deliverable_module,
        "_atomic_scoring_copy",
        copy_with_unexpected_transition,
    )

    with pytest.raises(ValueError, match="must not publish an artifact transition"):
        finalize_deliverable(
            session,
            agent,
        )
    retained = session.workspace / "scoring-output.xlsx"
    assert retained.is_file()
    assert stat.S_IMODE(retained.lstat().st_mode) == 0o400


def test_scoring_copy_is_anonymous_and_read_only_before_publication(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "final.xlsx"
    source.write_bytes(b"workbook-bytes")
    observed: dict[str, Any] = {}
    original_link = deliverable_module._link_anonymous_file_noreplace_at

    def observe_anonymous_inode(
        source_descriptor: int,
        destination_parent_descriptor: int,
        destination_name: str,
    ) -> None:
        metadata = os.fstat(source_descriptor)
        observed["identity"] = (metadata.st_dev, metadata.st_ino)
        observed["nlink"] = metadata.st_nlink
        observed["mode"] = stat.S_IMODE(metadata.st_mode)
        observed["bytes"] = os.pread(source_descriptor, metadata.st_size, 0)
        observed["destination_absent"] = not destination.exists()
        original_link(
            source_descriptor,
            destination_parent_descriptor,
            destination_name,
        )

    monkeypatch.setattr(
        deliverable_module,
        "_link_anonymous_file_noreplace_at",
        observe_anonymous_inode,
    )

    publication = deliverable_module._atomic_scoring_copy(source, destination)

    assert observed == {
        "identity": publication.identity,
        "nlink": 0,
        "mode": 0o400,
        "bytes": source.read_bytes(),
        "destination_absent": True,
    }
    assert destination.read_bytes() == source.read_bytes()
    assert stat.S_IMODE(destination.lstat().st_mode) == 0o400
    publication.commit()


def test_scoring_copy_rejects_same_length_write_corruption_without_residue(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "final.xlsx"
    source.write_bytes(b"workbook-bytes")
    before = _workspace_snapshot(tmp_path)
    original_write = os.write
    corrupted = False

    def corrupt_first_write(descriptor: int, data: bytes) -> int:
        nonlocal corrupted
        if not corrupted and data:
            corrupted = True
            data = bytes([data[0] ^ 1]) + data[1:]
        return original_write(descriptor, data)

    monkeypatch.setattr(deliverable_module.os, "write", corrupt_first_write)

    with pytest.raises(DeliverableValidationError, match="not byte-identical"):
        deliverable_module._atomic_scoring_copy(source, destination)

    assert corrupted
    assert not destination.exists()
    assert _workspace_snapshot(tmp_path) == before


def test_scoring_copy_rejects_source_path_replacement_during_copy(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = tmp_path / "source.xlsx"
    displaced = tmp_path / "displaced-source.xlsx"
    destination = tmp_path / "final.xlsx"
    source.write_bytes(b"original-bytes")
    original_read = os.read
    replaced = False

    def replace_before_first_read(descriptor: int, count: int) -> bytes:
        nonlocal replaced
        if not replaced:
            replaced = True
            source.rename(displaced)
            source.write_bytes(b"replaced-bytes")
        return original_read(descriptor, count)

    monkeypatch.setattr(deliverable_module.os, "read", replace_before_first_read)

    with pytest.raises(DeliverableValidationError, match="Scoring source changed"):
        deliverable_module._atomic_scoring_copy(source, destination)

    assert replaced
    assert source.read_bytes() == b"replaced-bytes"
    assert displaced.read_bytes() == b"original-bytes"
    assert not destination.exists()


def test_competing_destination_immediately_before_link_is_untouched(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "final.xlsx"
    source.write_bytes(b"workbook-bytes")
    competitor = b"competitor-owned-bytes"
    original_link = deliverable_module._link_anonymous_file_noreplace_at
    competitor_identity: tuple[int, int] | None = None

    def compete_then_link(
        source_descriptor: int,
        destination_parent_descriptor: int,
        destination_name: str,
    ) -> None:
        nonlocal competitor_identity
        destination.write_bytes(competitor)
        destination.chmod(0o600)
        metadata = destination.lstat()
        competitor_identity = (metadata.st_dev, metadata.st_ino)
        original_link(
            source_descriptor,
            destination_parent_descriptor,
            destination_name,
        )

    monkeypatch.setattr(
        deliverable_module,
        "_link_anonymous_file_noreplace_at",
        compete_then_link,
    )

    with pytest.raises(DeliverableValidationError, match="already exists"):
        deliverable_module._atomic_scoring_copy(source, destination)

    metadata = destination.lstat()
    assert competitor_identity == (metadata.st_dev, metadata.st_ino)
    assert destination.read_bytes() == competitor
    assert stat.S_IMODE(metadata.st_mode) == 0o600
    assert sorted(path.name for path in tmp_path.iterdir()) == ["final.xlsx", "source.xlsx"]


def test_anonymous_scoring_publication_valid_control(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "final.xlsx"
    source.write_bytes(b"workbook-bytes")

    publication = deliverable_module._atomic_scoring_copy(source, destination)

    assert destination.read_bytes() == source.read_bytes()
    assert (destination.stat().st_dev, destination.stat().st_ino) == publication.identity
    assert stat.S_IMODE(destination.stat().st_mode) == 0o400
    publication.commit()


def test_post_link_exception_retains_exact_recoverable_inode(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "final.xlsx"
    source.write_bytes(b"workbook-bytes")
    original_link = deliverable_module._link_anonymous_file_noreplace_at
    linked_identity: tuple[int, int] | None = None

    def link_then_fail(*args: Any) -> None:
        nonlocal linked_identity
        original_link(*args)
        metadata = destination.lstat()
        linked_identity = (metadata.st_dev, metadata.st_ino)
        raise OSError("injected post-link exception")

    monkeypatch.setattr(
        deliverable_module,
        "_link_anonymous_file_noreplace_at",
        link_then_fail,
    )

    with pytest.raises(DeliverableValidationError, match="Failed to publish"):
        deliverable_module._atomic_scoring_copy(source, destination)

    metadata = destination.lstat()
    assert linked_identity == (metadata.st_dev, metadata.st_ino)
    assert destination.read_bytes() == source.read_bytes()
    assert stat.S_IMODE(metadata.st_mode) == 0o400
    recovered = deliverable_module._existing_scoring_copy(source, destination)
    recovered.commit()


def test_post_link_competitor_replacement_is_never_moved_deleted_or_chmoded(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "final.xlsx"
    displaced = tmp_path / "linked-residue.xlsx"
    source.write_bytes(b"workbook-bytes")
    competitor = b"competitor-owned-bytes"
    original_link = deliverable_module._link_anonymous_file_noreplace_at
    competitor_identity: tuple[int, int] | None = None

    def replace_after_link(*args: Any) -> None:
        nonlocal competitor_identity
        original_link(*args)
        destination.rename(displaced)
        destination.write_bytes(competitor)
        destination.chmod(0o600)
        metadata = destination.lstat()
        competitor_identity = (metadata.st_dev, metadata.st_ino)
        raise OSError("injected replacement after link")

    monkeypatch.setattr(
        deliverable_module,
        "_link_anonymous_file_noreplace_at",
        replace_after_link,
    )

    with pytest.raises(DeliverableValidationError, match="Failed to publish"):
        deliverable_module._atomic_scoring_copy(source, destination)

    competitor_metadata = destination.lstat()
    assert competitor_identity == (competitor_metadata.st_dev, competitor_metadata.st_ino)
    assert destination.read_bytes() == competitor
    assert stat.S_IMODE(competitor_metadata.st_mode) == 0o600
    assert displaced.read_bytes() == source.read_bytes()
    assert stat.S_IMODE(displaced.lstat().st_mode) == 0o400


def test_scoring_verify_rebinds_name_after_hashing_held_inode(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "final.xlsx"
    displaced = tmp_path / "linked-residue.xlsx"
    source.write_bytes(b"workbook-bytes")
    competitor = b"competitor-owned-bytes"
    original_hash = deliverable_module._sha256_descriptor
    replacement_identity: tuple[int, int] | None = None

    def replace_after_scoring_hash(descriptor: int, *, expected_size: int) -> str:
        nonlocal replacement_identity
        digest = original_hash(descriptor, expected_size=expected_size)
        if destination.exists() and replacement_identity is None:
            descriptor_metadata = os.fstat(descriptor)
            destination_metadata = destination.lstat()
            if (descriptor_metadata.st_dev, descriptor_metadata.st_ino) == (
                destination_metadata.st_dev,
                destination_metadata.st_ino,
            ):
                destination.rename(displaced)
                destination.write_bytes(competitor)
                destination.chmod(0o600)
                replacement_metadata = destination.lstat()
                replacement_identity = (
                    replacement_metadata.st_dev,
                    replacement_metadata.st_ino,
                )
        return digest

    monkeypatch.setattr(
        deliverable_module,
        "_sha256_descriptor",
        replace_after_scoring_hash,
    )

    with pytest.raises(DeliverableValidationError, match="Scoring replica changed"):
        deliverable_module._atomic_scoring_copy(source, destination)

    observed = destination.lstat()
    assert replacement_identity == (observed.st_dev, observed.st_ino)
    assert destination.read_bytes() == competitor
    assert stat.S_IMODE(observed.st_mode) == 0o600
    assert displaced.read_bytes() == source.read_bytes()
    assert stat.S_IMODE(displaced.lstat().st_mode) == 0o400


def test_workspace_replacement_before_link_cannot_redirect_held_inode(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = tmp_path / "source.xlsx"
    workspace = tmp_path / "workspace"
    displaced_workspace = tmp_path / "bound-workspace"
    workspace.mkdir()
    destination = workspace / "final.xlsx"
    source.write_bytes(b"workbook-bytes")
    competitor = b"competitor-owned-bytes"
    original_link = deliverable_module._link_anonymous_file_noreplace_at
    competitor_identity: tuple[int, int] | None = None

    def replace_workspace_then_link(
        source_descriptor: int,
        destination_parent_descriptor: int,
        destination_name: str,
    ) -> None:
        nonlocal competitor_identity
        workspace.rename(displaced_workspace)
        workspace.mkdir()
        destination.write_bytes(competitor)
        destination.chmod(0o600)
        metadata = destination.lstat()
        competitor_identity = (metadata.st_dev, metadata.st_ino)
        original_link(
            source_descriptor,
            destination_parent_descriptor,
            destination_name,
        )

    monkeypatch.setattr(
        deliverable_module,
        "_link_anonymous_file_noreplace_at",
        replace_workspace_then_link,
    )

    with pytest.raises(DeliverableValidationError, match="workspace changed identity"):
        deliverable_module._atomic_scoring_copy(source, destination)

    competitor_metadata = destination.lstat()
    assert competitor_identity == (competitor_metadata.st_dev, competitor_metadata.st_ino)
    assert destination.read_bytes() == competitor
    assert stat.S_IMODE(competitor_metadata.st_mode) == 0o600
    residue = displaced_workspace / destination.name
    assert residue.read_bytes() == source.read_bytes()
    assert stat.S_IMODE(residue.lstat().st_mode) == 0o400


def test_publication_rollback_is_namespace_inert(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "final.xlsx"
    source.write_bytes(b"workbook-bytes")
    publication = deliverable_module._atomic_scoring_copy(source, destination)

    def reject_mutation(*_: Any, **__: Any) -> None:
        raise AssertionError("published rollback must not mutate a namespace")

    monkeypatch.setattr(deliverable_module.os, "rename", reject_mutation)
    monkeypatch.setattr(deliverable_module.os, "replace", reject_mutation)
    monkeypatch.setattr(deliverable_module.os, "unlink", reject_mutation)
    monkeypatch.setattr(deliverable_module.os, "rmdir", reject_mutation)

    publication.rollback()

    assert destination.read_bytes() == source.read_bytes()
    assert stat.S_IMODE(destination.lstat().st_mode) == 0o400


def test_publication_verify_rejects_symbolic_workspace_parent(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    workspace = tmp_path / "workspace"
    displaced_workspace = tmp_path / "displaced-workspace"
    workspace.mkdir()
    destination = workspace / "final.xlsx"
    source.write_bytes(b"workbook-bytes")
    publication = deliverable_module._atomic_scoring_copy(source, destination)
    workspace.rename(displaced_workspace)
    workspace.symlink_to(displaced_workspace.name, target_is_directory=True)

    with pytest.raises(DeliverableValidationError, match="workspace path contains a symbolic"):
        publication.verify()

    publication.rollback()
    assert (displaced_workspace / destination.name).read_bytes() == source.read_bytes()


def test_precommit_fsync_failure_leaves_no_named_residue(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "final.xlsx"
    source.write_bytes(b"workbook-bytes")
    before = _workspace_snapshot(tmp_path)
    original_fsync = os.fsync
    calls = 0

    def fail_first_fsync(descriptor: int) -> None:
        nonlocal calls
        calls += 1
        if calls == 1:
            raise OSError("injected anonymous-inode fsync failure")
        original_fsync(descriptor)

    monkeypatch.setattr(deliverable_module.os, "fsync", fail_first_fsync)

    with pytest.raises(DeliverableValidationError, match="Failed to publish"):
        deliverable_module._atomic_scoring_copy(source, destination)

    assert calls == 1
    assert not destination.exists()
    assert _workspace_snapshot(tmp_path) == before


def test_post_link_fsync_failure_retains_recoverable_publication(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "final.xlsx"
    source.write_bytes(b"workbook-bytes")
    original_fsync = os.fsync
    calls = 0

    def fail_third_fsync(descriptor: int) -> None:
        nonlocal calls
        calls += 1
        if calls == 3:
            raise OSError("injected post-link fsync failure")
        original_fsync(descriptor)

    monkeypatch.setattr(deliverable_module.os, "fsync", fail_third_fsync)

    with pytest.raises(DeliverableValidationError, match="Failed to publish"):
        deliverable_module._atomic_scoring_copy(source, destination)

    assert calls == 3
    assert destination.read_bytes() == source.read_bytes()
    recovered = deliverable_module._existing_scoring_copy(source, destination)
    recovered.commit()


def test_scoring_failure_close_error_still_closes_every_open_descriptor(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "final.xlsx"
    source.write_bytes(b"workbook-bytes")
    original_open = os.open
    original_close = os.close
    opened: list[int] = []
    failed = False

    def track_open(*args: Any, **kwargs: Any) -> int:
        descriptor = original_open(*args, **kwargs)
        opened.append(descriptor)
        return descriptor

    def close_then_fail_once(descriptor: int) -> None:
        nonlocal failed
        original_close(descriptor)
        if not failed:
            failed = True
            raise OSError("injected descriptor close failure")

    def fail_link(*_: Any) -> None:
        raise DeliverableValidationError("injected publication failure")

    monkeypatch.setattr(deliverable_module.os, "open", track_open)
    monkeypatch.setattr(deliverable_module.os, "close", close_then_fail_once)
    monkeypatch.setattr(
        deliverable_module,
        "_link_anonymous_file_noreplace_at",
        fail_link,
    )

    with pytest.raises(
        DeliverableValidationError,
        match="Failed to close scoring-publication descriptors",
    ):
        deliverable_module._atomic_scoring_copy(source, destination)

    assert failed
    assert opened
    for descriptor in opened:
        with pytest.raises(OSError):
            os.fstat(descriptor)


def test_late_validation_failure_never_publishes_and_retry_succeeds(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session, agent = _candidate(tmp_path)
    scoring_copy = session.workspace / "scoring-output.xlsx"
    before = _workspace_snapshot(session.workspace)
    original_audit = deliverable_module._audit_target_grounding_commit_chain

    def fail_validation(*_: Any, **__: Any) -> None:
        raise DeliverableValidationError("injected target-chain failure")

    monkeypatch.setattr(
        deliverable_module,
        "_audit_target_grounding_commit_chain",
        fail_validation,
    )

    with pytest.raises(DeliverableValidationError, match="injected target-chain"):
        finalize_deliverable(session, agent)
    assert not scoring_copy.exists()
    assert _workspace_snapshot(session.workspace) == before

    monkeypatch.setattr(
        deliverable_module,
        "_audit_target_grounding_commit_chain",
        original_audit,
    )
    bundle = finalize_deliverable(session, agent)
    assert bundle.scoring_copy == scoring_copy
    assert scoring_copy.is_file()


def test_event_record_failure_retains_replica_and_retry_succeeds(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session, agent = _candidate(tmp_path)
    scoring_copy = session.workspace / "scoring-output.xlsx"
    before = _workspace_snapshot(session.workspace)
    original_record = trajectory_module._TrajectoryTransaction.record

    def fail_finalization_event(
        transaction: Any,
        event: str,
        payload: dict[str, Any] | None = None,
    ) -> None:
        original_record(transaction, event, payload)
        raise OSError("injected trajectory failure")

    monkeypatch.setattr(
        trajectory_module._TrajectoryTransaction,
        "record",
        fail_finalization_event,
    )

    with pytest.raises(OSError, match="injected trajectory"):
        finalize_deliverable(session, agent)
    assert scoring_copy.is_file()
    assert _workspace_snapshot(session.workspace) != before

    monkeypatch.setattr(
        trajectory_module._TrajectoryTransaction,
        "record",
        original_record,
    )
    bundle = finalize_deliverable(session, agent)
    assert bundle.scoring_copy == scoring_copy
    assert scoring_copy.is_file()


def test_event_failure_does_not_delete_or_authorize_a_replaced_scoring_file(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session, agent = _candidate(tmp_path)
    scoring_copy = session.workspace / "scoring-output.xlsx"
    displaced_copy = session.workspace / "displaced-scoring-output.xlsx"
    replacement = b"replacement-scoring-bytes"
    original_record = trajectory_module._TrajectoryTransaction.record
    identities: dict[str, tuple[int, int]] = {}

    def replace_then_fail(
        transaction: Any,
        event: str,
        payload: dict[str, Any] | None = None,
    ) -> None:
        owned_file_metadata = scoring_copy.lstat()
        identities["owned_file"] = (
            owned_file_metadata.st_dev,
            owned_file_metadata.st_ino,
        )
        scoring_copy.rename(displaced_copy)
        scoring_copy.write_bytes(replacement)
        scoring_copy.chmod(0o600)
        replacement_file_metadata = scoring_copy.lstat()
        identities["replacement_file"] = (
            replacement_file_metadata.st_dev,
            replacement_file_metadata.st_ino,
        )
        original_record(transaction, event, payload)
        raise OSError("injected post-replacement event failure")

    monkeypatch.setattr(
        trajectory_module._TrajectoryTransaction,
        "record",
        replace_then_fail,
    )

    with pytest.raises(OSError, match="post-replacement event failure"):
        finalize_deliverable(session, agent)

    assert identities["owned_file"] != identities["replacement_file"]
    assert scoring_copy.read_bytes() == replacement
    assert stat.S_IMODE(scoring_copy.lstat().st_mode) == 0o600
    assert displaced_copy.read_bytes() != replacement
    assert stat.S_IMODE(displaced_copy.lstat().st_mode) == 0o400


def test_source_replacement_before_event_commit_retains_immutable_publication(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session, agent = _candidate(tmp_path)
    scoring_copy = session.workspace / "scoring-output.xlsx"
    displaced_source = session.workspace / "displaced-output.xlsx"
    original_bytes = session.workbook_path.read_bytes()
    original_record = trajectory_module._TrajectoryTransaction.record
    replaced = False

    def replace_source_after_append(
        transaction: Any,
        event: str,
        payload: dict[str, Any] | None = None,
    ) -> None:
        nonlocal replaced
        original_record(transaction, event, payload)
        session.workbook_path.rename(displaced_source)
        session.workbook_path.write_bytes(original_bytes)
        replaced = True

    monkeypatch.setattr(
        trajectory_module._TrajectoryTransaction,
        "record",
        replace_source_after_append,
    )

    with pytest.raises(DeliverableValidationError, match="Scoring source changed identity"):
        finalize_deliverable(session, agent)

    assert replaced
    assert session.workbook_path.read_bytes() == original_bytes
    assert displaced_source.read_bytes() == original_bytes
    assert scoring_copy.read_bytes() == original_bytes
    assert stat.S_IMODE(scoring_copy.lstat().st_mode) == 0o400


def test_replacement_before_trajectory_durability_never_records_authorization(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session, agent = _candidate(tmp_path)
    scoring_copy = session.workspace / "scoring-output.xlsx"
    displaced_copy = session.workspace / "displaced-scoring-output.xlsx"
    replacement = b"replacement-before-trajectory-durability"
    original_exit = trajectory_module._TrajectoryTransaction.__exit__
    replaced = False

    def replace_before_append(transaction: Any, *args: Any) -> bool:
        nonlocal replaced
        if (
            not replaced
            and b'"event":"observer.finalization_recorded"' in transaction._encoded_record_bytes
        ):
            scoring_copy.rename(displaced_copy)
            scoring_copy.write_bytes(replacement)
            scoring_copy.chmod(0o600)
            replaced = True
        return original_exit(transaction, *args)

    monkeypatch.setattr(
        trajectory_module._TrajectoryTransaction,
        "__exit__",
        replace_before_append,
    )

    with pytest.raises(DeliverableValidationError, match="Scoring replica changed"):
        finalize_deliverable(session, agent)

    assert replaced
    assert scoring_copy.read_bytes() == replacement
    finalization_events = [
        row
        for row in read_trajectory(session.paths.trajectory)
        if row["event"] == "observer.finalization_recorded"
    ]
    assert len(finalization_events) == 1
    payload = finalization_events[0]["payload"]
    assert payload["candidate_outcome"] == "accepted_candidate"
    assert payload["accepted_deliverable"] is False
    assert payload["record_role"] == "observer_only_fresh_audit_required"


def test_atomic_scoring_copy_rejects_existing_replica_without_removing_it(
    tmp_path: Path,
) -> None:
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "final.xlsx"
    source.write_bytes(b"workbook-bytes")

    first = deliverable_module._atomic_scoring_copy(source, destination)
    first.commit()

    with pytest.raises(DeliverableValidationError, match="already exists"):
        deliverable_module._atomic_scoring_copy(source, destination)

    assert _sha256(destination) == first.sha256
    assert destination.lstat().st_mode & 0o777 == 0o400


def test_second_finalization_is_idempotent_and_preserves_first_certificate(
    tmp_path: Path,
) -> None:
    session, agent = _candidate(tmp_path)
    first = finalize_deliverable(session, agent)
    trajectory_before = session.paths.trajectory.read_bytes()

    second = finalize_deliverable(session, agent)

    assert second.certificate == first.certificate
    assert second.scoring_copy == first.scoring_copy
    assert first.scoring_copy.is_file()
    assert session.paths.trajectory.read_bytes() == trajectory_before
    assert audit_deliverable_certificate(
        first.certificate,
        agent_evidence=agent,
        run_root=session.workspace,
        output_workbook=session.workbook_path,
    ).valid


def test_failed_idempotent_finalization_preserves_existing_replica(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session, agent = _candidate(tmp_path)
    first = finalize_deliverable(session, agent)
    before = _workspace_snapshot(session.workspace)
    original_fsync = os.fsync
    calls = 0

    def fail_directory_fsync(descriptor: int) -> None:
        nonlocal calls
        calls += 1
        if calls == 2:
            raise OSError("injected trajectory directory fsync failure")
        original_fsync(descriptor)

    monkeypatch.setattr(trajectory_module.os, "fsync", fail_directory_fsync)

    with pytest.raises(OSError, match="trajectory directory fsync"):
        finalize_deliverable(session, agent)

    assert first.scoring_copy.is_file()
    assert _workspace_snapshot(session.workspace) == before


def test_durable_trajectory_close_failure_preserves_recoverable_publication(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session, agent = _candidate(tmp_path)
    scoring_copy = session.workspace / "scoring-output.xlsx"
    original_exit = trajectory_module._TrajectoryTransaction.__exit__
    original_close = os.close
    injected = False

    def exit_with_one_close_failure(transaction: Any, *args: Any) -> bool:
        def close_then_fail(descriptor: int) -> None:
            nonlocal injected
            original_close(descriptor)
            if not injected:
                injected = True
                raise OSError("injected durable trajectory close failure")

        trajectory_module.os.close = close_then_fail
        try:
            return original_exit(transaction, *args)
        finally:
            trajectory_module.os.close = original_close

    monkeypatch.setattr(
        trajectory_module._TrajectoryTransaction,
        "__exit__",
        exit_with_one_close_failure,
    )

    with pytest.raises(OSError, match="durable trajectory close failure"):
        finalize_deliverable(session, agent)

    assert injected
    assert scoring_copy.is_file()
    finalization_events = [
        row
        for row in read_trajectory(session.paths.trajectory)
        if row["event"] == "observer.finalization_recorded"
    ]
    assert len(finalization_events) == 1

    monkeypatch.setattr(
        trajectory_module._TrajectoryTransaction,
        "__exit__",
        original_exit,
    )
    recovered = finalize_deliverable(session, agent)
    assert recovered.scoring_copy == scoring_copy


def test_finalization_recovers_an_exact_replica_without_an_event(
    tmp_path: Path,
) -> None:
    session, agent = _candidate(tmp_path)
    scoring_copy = session.workspace / "scoring-output.xlsx"
    residue = deliverable_module._atomic_scoring_copy(
        session.workbook_path,
        scoring_copy,
    )
    residue.commit()

    bundle = finalize_deliverable(session, agent)

    assert bundle.scoring_copy == scoring_copy
    assert _sha256(scoring_copy) == residue.sha256
    finalization_events = [
        json.loads(line)
        for line in session.paths.trajectory.read_text(encoding="utf-8").splitlines()
        if json.loads(line)["event"] == "observer.finalization_recorded"
    ]
    assert len(finalization_events) == 1
    assert audit_deliverable_certificate(
        bundle.certificate,
        agent_evidence=agent,
        run_root=session.workspace,
        output_workbook=session.workbook_path,
    ).valid


def test_failed_residue_recovery_preserves_existing_replica(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session, agent = _candidate(tmp_path)
    scoring_copy = session.workspace / "scoring-output.xlsx"
    residue = deliverable_module._atomic_scoring_copy(
        session.workbook_path,
        scoring_copy,
    )
    residue.commit()
    before = _workspace_snapshot(session.workspace)
    original_record = trajectory_module._TrajectoryTransaction.record

    def append_then_fail(
        transaction: Any,
        event: str,
        payload: dict[str, Any] | None = None,
    ) -> None:
        original_record(transaction, event, payload)
        raise OSError("injected residue event failure")

    monkeypatch.setattr(
        trajectory_module._TrajectoryTransaction,
        "record",
        append_then_fail,
    )

    with pytest.raises(OSError, match="residue event failure"):
        finalize_deliverable(session, agent)

    assert scoring_copy.is_file()
    assert _workspace_snapshot(session.workspace) == before


@pytest.mark.parametrize(
    ("field", "value"),
    [
        ("artifact_role", "new_artifact_revision"),
        ("creates_artifact_transition", True),
    ],
)
def test_audit_rejects_scoring_replica_revision_semantics_tampering(
    tmp_path: Path,
    field: str,
    value: Any,
) -> None:
    session, agent = _candidate(tmp_path)
    bundle = finalize_deliverable(
        session,
        agent,
    )
    tampered = json.loads(json.dumps(bundle.certificate))
    tampered["scoring_copy"][field] = value
    tampered["certificate_sha256"] = _certificate_digest(tampered)

    assert not audit_deliverable_certificate(
        tampered,
        agent_evidence=agent,
        run_root=session.workspace,
        output_workbook=session.workbook_path,
    ).valid


def test_audit_rejects_rehashed_alternate_scoring_path(tmp_path: Path) -> None:
    session, agent = _candidate(tmp_path)
    bundle = finalize_deliverable(session, agent)
    alternate = session.workspace / "alternate-scoring.xlsx"
    alternate.write_bytes(bundle.scoring_copy.read_bytes())
    alternate.chmod(0o400)
    tampered = json.loads(json.dumps(bundle.certificate))
    tampered["scoring_copy"]["relative_path"] = alternate.name
    tampered["certificate_sha256"] = _certificate_digest(tampered)

    assert not audit_deliverable_certificate(
        tampered,
        agent_evidence=agent,
        run_root=session.workspace,
        output_workbook=session.workbook_path,
    ).valid


@pytest.mark.parametrize(
    ("field", "value"),
    [
        ("backend", 7),
        ("version", {"local": "1"}),
        ("profile", "/tmp/unsafe-profile"),
        ("format", "/tmp/output.xlsx"),
    ],
)
def test_recalculation_certificate_metadata_is_strict_and_path_free(
    tmp_path: Path,
    field: str,
    value: Any,
) -> None:
    session, agent = _candidate(tmp_path)
    bundle = finalize_deliverable(session, agent)
    tampered = json.loads(json.dumps(bundle.certificate))
    tampered["postprocess"]["recalculation"][field] = value
    if field in {"backend", "version", "profile"}:
        tampered["recalculation_attestation"][field] = value
        tampered["recalculation_attestation"]["witness_sha256"] = _witness_digest(
            tampered["recalculation_attestation"]
        )
    tampered["certificate_sha256"] = _certificate_digest(tampered)

    assert not audit_deliverable_certificate(
        tampered,
        agent_evidence=agent,
        run_root=session.workspace,
        output_workbook=session.workbook_path,
    ).valid


def test_recalculation_effects_reject_absolute_diagnostic_paths(
    tmp_path: Path,
) -> None:
    session, agent = _candidate(tmp_path)
    bundle = finalize_deliverable(session, agent)
    tampered = json.loads(json.dumps(bundle.certificate))
    tampered["postprocess"]["recalculation"]["workbook_effects"]["reasons"] = [
        "ValueError: /tmp/private/workbook.xlsx"
    ]
    tampered["certificate_sha256"] = _certificate_digest(tampered)

    assert not audit_deliverable_certificate(
        tampered,
        agent_evidence=agent,
        run_root=session.workspace,
        output_workbook=session.workbook_path,
    ).valid


@pytest.mark.parametrize("callback_kind", ["changed", "semantic", "unknown", "error"])
def test_reportable_finalization_rejects_callbacks_before_any_side_effect(
    tmp_path: Path,
    callback_kind: str,
) -> None:
    session, agent = _candidate(tmp_path)
    artifact = session.artifact_ref()
    transitions = session.artifact_transitions
    workspace = _workspace_snapshot(session.workspace)
    callback_called = False

    callbacks = {
        "changed": lambda: _changed_recalculation(session),
        "semantic": lambda: _semantic_recalculation(session),
        "unknown": lambda: _unknown_recalculation(session),
        "error": lambda: (_ for _ in ()).throw(RuntimeError("callback failure")),
    }

    def callback() -> Mapping[str, Any]:
        nonlocal callback_called
        callback_called = True
        return callbacks[callback_kind]()

    with pytest.raises(DeliverableValidationError, match="callbacks are disabled"):
        finalize_deliverable(session, agent, recalculation_callback=callback)

    assert callback_called is False
    assert session.artifact_ref() == artifact
    assert session.artifact_transitions == transitions
    assert _workspace_snapshot(session.workspace) == workspace


def test_evidence_certificate_replay_rejects_scope_and_witness_tampering(
    tmp_path: Path,
) -> None:
    _, agent = _candidate(tmp_path)
    certificate = json.loads(json.dumps(agent["evidence_contract"]["decision"]["certificate"]))
    certificate["events"][-1]["scope"] = EvidenceScope.one("Other", "A1:A1").to_dict()
    event_payload = {
        key: value
        for key, value in certificate["events"][-1].items()
        if key != "event_chain_sha256"
    }
    previous_chain = certificate["events"][-2]["event_chain_sha256"]
    event_bytes = json.dumps(
        event_payload,
        ensure_ascii=True,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("ascii")
    new_chain = hashlib.sha256(bytes.fromhex(previous_chain) + event_bytes).hexdigest()
    certificate["events"][-1]["event_chain_sha256"] = new_chain
    certificate["event_chain_sha256"] = new_chain
    certificate["certificate_sha256"] = _certificate_digest(certificate)

    try:
        validate_evidence_certificate(certificate)
    except ValueError as exc:
        assert "scope, witnesses, or obligations" in str(exc)
    else:
        raise AssertionError("tampered scope unexpectedly passed certificate replay")


def test_audit_rejects_tampered_scoring_copy_and_outer_certificate(
    tmp_path: Path,
) -> None:
    session, agent = _candidate(tmp_path)
    bundle = finalize_deliverable(
        session,
        agent,
    )

    tampered_certificate = json.loads(json.dumps(bundle.certificate))
    tampered_certificate["final_artifact"]["sha256"] = "f" * 64
    assert not audit_deliverable_certificate(
        tampered_certificate,
        agent_evidence=agent,
        run_root=session.workspace,
        output_workbook=session.workbook_path,
    ).valid

    os.chmod(bundle.scoring_copy, 0o600)
    workbook = load_workbook(bundle.scoring_copy)
    try:
        workbook.active["A1"] = 7
        workbook.save(bundle.scoring_copy)
    finally:
        workbook.close()
    assert not audit_deliverable_certificate(
        bundle.certificate,
        agent_evidence=agent,
        run_root=session.workspace,
        output_workbook=session.workbook_path,
    ).valid


def test_scoring_copy_must_remain_read_only_even_when_bytes_match(
    tmp_path: Path,
) -> None:
    session, agent = _candidate(tmp_path)
    bundle = finalize_deliverable(
        session,
        agent,
    )
    bundle.scoring_copy.chmod(0o600)

    with pytest.raises(ValueError, match="read-only regular non-symbolic"):
        score_read_only(bundle, lambda _: None)
    assert not audit_deliverable_certificate(
        bundle.certificate,
        agent_evidence=agent,
        run_root=session.workspace,
        output_workbook=session.workbook_path,
    ).valid


def test_score_read_only_preserves_scorer_exception_when_snapshot_mutates(
    tmp_path: Path,
) -> None:
    session, agent = _candidate(tmp_path)
    bundle = finalize_deliverable(
        session,
        agent,
    )
    scorer_error = RuntimeError("scorer failed after writing")

    def mutate_then_raise(path: Path) -> None:
        path.chmod(0o600)
        path.write_bytes(b"tampered scoring bytes")
        raise scorer_error

    with pytest.raises(RuntimeError, match="failed after writing") as caught:
        score_read_only(bundle, mutate_then_raise)

    assert caught.value is scorer_error
    assert isinstance(caught.value.__cause__, DeliverableValidationError)
    assert any(
        "Scoring input verification also failed" in note
        for note in getattr(caught.value, "__notes__", ())
    )


def test_score_read_only_preserves_unmutated_scorer_exception(tmp_path: Path) -> None:
    session, agent = _candidate(tmp_path)
    bundle = finalize_deliverable(
        session,
        agent,
    )
    scorer_error = RuntimeError("scorer failed without writing")

    def raise_without_mutation(_: Path) -> None:
        raise scorer_error

    with pytest.raises(RuntimeError, match="without writing") as caught:
        score_read_only(bundle, raise_without_mutation)

    assert caught.value is scorer_error


def test_score_read_only_preserves_primary_exception_across_cleanup_failures(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session, agent = _candidate(tmp_path)
    bundle = finalize_deliverable(session, agent)
    scorer_error = RuntimeError("primary scorer failure")
    original_open = os.open
    original_fstat = os.fstat
    original_fchmod = os.fchmod
    original_publication_close = deliverable_module._ScoringCopyPublication.close
    opened: list[int] = []
    observed_root: Path | None = None

    def track_open(*args: object, **kwargs: object) -> int:
        descriptor = original_open(*args, **kwargs)
        opened.append(descriptor)
        return descriptor

    def fail_cleanup_fchmod(descriptor: int, mode: int) -> None:
        if mode == 0o700:
            raise OSError("injected cleanup fchmod failure")
        original_fchmod(descriptor, mode)

    def close_then_fail(
        publication: deliverable_module._ScoringCopyPublication,
    ) -> None:
        original_publication_close(publication)
        raise DeliverableValidationError("injected publication close failure")

    def raise_from_scorer(path: Path) -> None:
        nonlocal observed_root
        observed_root = path.parent
        raise scorer_error

    monkeypatch.setattr(deliverable_module.os, "open", track_open)
    monkeypatch.setattr(deliverable_module.os, "fchmod", fail_cleanup_fchmod)
    monkeypatch.setattr(
        deliverable_module._ScoringCopyPublication,
        "close",
        close_then_fail,
    )

    with pytest.raises(RuntimeError, match="primary scorer failure") as caught:
        score_read_only(bundle, raise_from_scorer)

    assert caught.value is scorer_error
    notes = getattr(caught.value, "__notes__", ())
    assert any("injected cleanup fchmod failure" in note for note in notes)
    assert any("injected publication close failure" in note for note in notes)
    assert opened
    for descriptor in opened:
        with pytest.raises(OSError):
            original_fstat(descriptor)

    assert observed_root is not None
    observed_root.chmod(0o700)
    (observed_root / "input.xlsx").unlink()
    observed_root.rmdir()


def test_score_read_only_raises_cleanup_failure_without_primary_exception(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session, agent = _candidate(tmp_path)
    bundle = finalize_deliverable(session, agent)
    original_publication_close = deliverable_module._ScoringCopyPublication.close

    def close_then_fail(
        publication: deliverable_module._ScoringCopyPublication,
    ) -> None:
        original_publication_close(publication)
        raise DeliverableValidationError("injected publication close failure")

    monkeypatch.setattr(
        deliverable_module._ScoringCopyPublication,
        "close",
        close_then_fail,
    )

    with pytest.raises(DeliverableValidationError, match="Failed to clean up") as caught:
        score_read_only(bundle, lambda _: 42)

    assert isinstance(caught.value.__cause__, DeliverableValidationError)
    assert "injected publication close failure" in str(caught.value.__cause__)


def test_score_read_only_never_cleans_replacement_root_by_path(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session, agent = _candidate(tmp_path)
    bundle = finalize_deliverable(session, agent)
    scorer_error = RuntimeError("scorer failed after replacing snapshot root")
    original_open = os.open
    original_fstat = os.fstat
    opened: list[int] = []
    observed: dict[str, Path] = {}
    replacement_payload = b"competitor-owned-content"

    def track_open(*args: object, **kwargs: object) -> int:
        descriptor = original_open(*args, **kwargs)
        opened.append(descriptor)
        return descriptor

    def replace_root_then_raise(path: Path) -> None:
        root = path.parent
        displaced = root.with_name(f"{root.name}-displaced")
        root.replace(displaced)
        root.mkdir(mode=0o711)
        replacement = root / "competitor.txt"
        replacement.write_bytes(replacement_payload)
        root.chmod(0o711)
        observed.update(root=root, displaced=displaced, replacement=replacement)
        raise scorer_error

    monkeypatch.setattr(deliverable_module.os, "open", track_open)

    with pytest.raises(RuntimeError, match="replacing snapshot root") as caught:
        score_read_only(bundle, replace_root_then_raise)

    assert caught.value is scorer_error
    notes = getattr(caught.value, "__notes__", ())
    assert any("Scoring input verification also failed" in note for note in notes)
    assert any("replacement was left untouched" in note for note in notes)
    root = observed["root"]
    displaced = observed["displaced"]
    replacement = observed["replacement"]
    assert replacement.read_bytes() == replacement_payload
    assert stat.S_IMODE(root.stat().st_mode) == 0o711
    assert list(displaced.iterdir()) == []
    assert stat.S_IMODE(displaced.stat().st_mode) == 0o700
    assert opened
    for descriptor in opened:
        with pytest.raises(OSError):
            original_fstat(descriptor)

    replacement.unlink()
    root.rmdir()
    displaced.rmdir()


def test_fresh_audit_rejects_internal_scoring_copy_symlink(
    tmp_path: Path,
) -> None:
    session, agent = _candidate(tmp_path)
    bundle = finalize_deliverable(
        session,
        agent,
    )
    relocated = bundle.scoring_copy.parent / "relocated-final.xlsx"
    bundle.scoring_copy.replace(relocated)
    bundle.scoring_copy.symlink_to(relocated.name)

    audit = audit_deliverable_certificate(
        bundle.certificate,
        agent_evidence=agent,
        run_root=session.workspace,
        output_workbook=session.workbook_path,
    )

    assert audit.valid is False
    assert audit.reasons == ("deliverable_lineage_invalid:DeliverableValidationError",)


def test_formula_free_visual_candidate_needs_no_postprocess_pixel_carry(
    tmp_path: Path,
) -> None:
    candidate_page = b"\x89PNG\r\n\x1a\ncandidate-pixels"
    candidate_sha256 = hashlib.sha256(candidate_page).hexdigest()
    session, agent = _visual_candidate(tmp_path, page_sha256=candidate_sha256)

    bundle = finalize_deliverable(session, agent)

    assert bundle.candidate_artifact == bundle.final_artifact
    assert bundle.certificate["visual_equivalence_witness"] is None
    assert bundle.certificate["evidence_policy"]["candidate_visual_evidence_present"] is True
    assert (
        bundle.certificate["evidence_policy"]["pixel_equivalence_required_for_visual_carry"]
        is False
    )
    assert audit_deliverable_certificate(
        bundle.certificate,
        agent_evidence=agent,
        run_root=session.workspace,
        output_workbook=session.workbook_path,
    ).valid
