from __future__ import annotations

import json
import platform
import shutil
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook

from spreadsheet_harness.agent import ResponseTurn
from spreadsheet_harness.arms import run_arm
from spreadsheet_harness.budget import RunBudget
from spreadsheet_harness.code_interpreter import ensure_strict_code_isolation
from spreadsheet_harness.completion_attempt import audit_completion_attempt
from spreadsheet_harness.config import ProviderConfig
from spreadsheet_harness.errors import CodeIsolationError
from spreadsheet_harness.evidence_contract import (
    ContractMode,
    ContractSpec,
    audit_evidence_certificate,
)
from spreadsheet_harness.render import find_libreoffice
from spreadsheet_harness.session import WorkbookSession
from spreadsheet_harness.skills import SkillRegistry

ROOT = Path(__file__).resolve().parents[1]
EVIDENCE_CONTRACT = ROOT / "contracts" / "spreadsheet-evidence-v1.yaml"

_EDIT_CODE = """import sheet_harness

wb = sheet_harness.load_workbook()
ws = wb["Sales"]
assert ws["A2"].value == "Apple"
assert ws["B2"].value == 2
ws["A2"] = "Banana"
sheet_harness.save_workbook(wb)
wb.close()

verified = sheet_harness.load_workbook(data_only=False)
assert verified["Sales"]["A2"].value == "Banana"
assert verified["Sales"]["B2"].value == 2
verified.close()
print("verified Sales!A2=Banana; Sales!B2 unchanged")
"""


@pytest.fixture(scope="module", autouse=True)
def _require_real_strict_code_isolation() -> None:
    if platform.system() != "Linux" or shutil.which("bwrap") is None:
        pytest.skip("v28 run_arm integration requires Linux and Bubblewrap")
    try:
        ensure_strict_code_isolation()
    except CodeIsolationError as exc:
        pytest.skip(f"Bubblewrap namespaces are unavailable: {exc}")


def _scripted_client(arm: str) -> type:
    class ScriptedV28Client:
        requests: list[dict[str, Any]] = []

        def __init__(self, _: ProviderConfig) -> None:
            self.turn = 0

        def __enter__(self) -> ScriptedV28Client:
            return self

        def __exit__(self, *_: Any) -> None:
            return None

        def create(self, payload: dict[str, Any], **__: Any) -> ResponseTurn:
            self.requests.append(payload)
            self.turn += 1
            script = (
                (
                    "inspect_range",
                    {"sheet": "Sales", "range_ref": "A2"},
                ),
                (
                    "declare_edit_target",
                    {
                        "targets": [{"sheet": "Sales", "range_ref": "A2"}],
                        "observation_ids": [1],
                    },
                ),
                (
                    "code_interpreter",
                    {"code": _EDIT_CODE, "declaration_id": 1},
                ),
                (
                    "inspect_range",
                    {"sheet": "Sales", "range_ref": "A1:B3"},
                ),
                ("submit_result", {}),
            )
            name, arguments = script[self.turn - 1]
            return ResponseTurn(
                response_id=f"{arm}-response-{self.turn}",
                output=[
                    {
                        "type": "function_call",
                        "id": f"{arm}-function-{self.turn}",
                        "call_id": f"{arm}-call-{self.turn}",
                        "name": name,
                        "arguments": json.dumps(arguments),
                    }
                ],
                text="",
                usage={"input_tokens": 2, "output_tokens": 1, "total_tokens": 3},
            )

    return ScriptedV28Client


def _paper_scripted_client() -> type:
    class ScriptedPaperClient:
        stage_order = ("extract", "vision_verify", "latex_verify", "reconcile", "solve")
        instance_count = 0
        requests: dict[str, list[dict[str, Any]]] = {stage: [] for stage in stage_order}

        def __init__(self, _: ProviderConfig) -> None:
            self.stage = self.stage_order[type(self).instance_count]
            type(self).instance_count += 1
            self.turn = 0

        def __enter__(self) -> ScriptedPaperClient:
            return self

        def __exit__(self, *_: Any) -> None:
            return None

        @staticmethod
        def _latest_tool_output(payload: dict[str, Any]) -> dict[str, Any]:
            raw = next(
                item["output"]
                for item in reversed(payload["input"])
                if item.get("type") == "function_call_output"
            )
            return json.loads(raw)

        def create(self, payload: dict[str, Any], **__: Any) -> ResponseTurn:
            self.requests[self.stage].append(payload)
            self.turn += 1
            evidence = (
                "summary: verified workbook evidence\n"
                "provenance:\n"
                "  sheet: Sales\n"
                "  range: A1:D3\n"
                "  page: 1\n"
            )
            if self.stage == "extract":
                script = (
                    ("list_sheets", {}),
                    ("inspect_range", {"sheet": "Sales", "range_ref": "A1:D3"}),
                    ("submit_result", {"result": evidence}),
                )
                name, arguments = script[self.turn - 1]
            elif self.stage == "vision_verify":
                if self.turn == 1:
                    name, arguments = "render_workbook", {"dpi": 72}
                elif self.turn == 2:
                    rendered = self._latest_tool_output(payload)
                    name = "view_image"
                    arguments = {"image_path": rendered["pages"][0]["image_path"]}
                else:
                    name, arguments = "submit_result", {"result": evidence}
            elif self.stage == "latex_verify":
                script = (
                    ("range_to_latex", {"sheet": "Sales", "range_ref": "A1:D3"}),
                    ("submit_result", {"result": evidence}),
                )
                name, arguments = script[self.turn - 1]
            elif self.stage == "reconcile":
                return ResponseTurn(
                    response_id="paper-reconcile-response-1",
                    output=[
                        {
                            "type": "message",
                            "role": "assistant",
                            "content": [{"type": "output_text", "text": evidence}],
                        }
                    ],
                    text=evidence,
                    usage={"input_tokens": 2, "output_tokens": 1, "total_tokens": 3},
                )
            else:
                solve_script = (
                    ("inspect_range", {"sheet": "Sales", "range_ref": "A2"}),
                    (
                        "declare_edit_target",
                        {
                            "targets": [{"sheet": "Sales", "range_ref": "A2"}],
                            "observation_ids": [1],
                        },
                    ),
                    ("code_interpreter", {"code": _EDIT_CODE, "declaration_id": 1}),
                    ("inspect_range", {"sheet": "Sales", "range_ref": "A1:B3"}),
                    ("submit_result", {}),
                )
                name, arguments = solve_script[self.turn - 1]
            return ResponseTurn(
                response_id=f"paper-{self.stage}-response-{self.turn}",
                output=[
                    {
                        "type": "function_call",
                        "id": f"paper-{self.stage}-function-{self.turn}",
                        "call_id": f"paper-{self.stage}-call-{self.turn}",
                        "name": name,
                        "arguments": json.dumps(arguments),
                    }
                ],
                text="",
                usage={"input_tokens": 2, "output_tokens": 1, "total_tokens": 3},
            )

    return ScriptedPaperClient


def _trajectory(session: WorkbookSession) -> list[dict[str, Any]]:
    return [
        json.loads(line)
        for line in session.paths.trajectory.read_text(encoding="utf-8").splitlines()
    ]


def _assert_recent_output_contains(payload: dict[str, Any], expected: str) -> None:
    outputs = [
        item["output"] for item in payload["input"] if item.get("type") == "function_call_output"
    ]
    assert any(expected in output for output in outputs)


@pytest.mark.parametrize("arm", ["bare", "profile", "native", "ours"])
def test_v28_editing_arms_run_real_grounded_contract_chain(
    arm: str,
    sample_workbook: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    client = _scripted_client(arm)
    monkeypatch.setattr("spreadsheet_harness.agent.ResponsesClient", client)
    session = WorkbookSession.create(sample_workbook, tmp_path / f"v28-{arm}")
    initial_artifact = session.artifact_ref()
    contract = ContractSpec.load(EVIDENCE_CONTRACT)

    result = run_arm(
        arm,  # type: ignore[arg-type]
        ProviderConfig("https://example.test/v1", "not-a-real-key", "test-model"),
        session,
        SkillRegistry(()),
        "Replace Sales!A2 with Banana and preserve all other cells.",
        max_output_tokens=512,
        max_elapsed_seconds=120,
        budget=RunBudget(
            max_model_calls=5,
            max_total_tokens=100,
            max_elapsed_seconds=120,
        ),
        max_turns_per_arm=5,
        evidence_contract=contract,
        contract_mode=ContractMode.ENFORCE,
        enable_target_grounding=True,
        capture_completion_attempts=True,
    )

    assert len(client.requests) == 5
    assert [request["tool_choice"] for request in client.requests] == [
        {"type": "function", "name": "inspect_range"},
        {"type": "function", "name": "declare_edit_target"},
        "auto",
        {"type": "function", "name": "inspect_range"},
        {"type": "function", "name": "submit_result"},
    ]
    assert [[tool["name"] for tool in request["tools"]] for request in client.requests[:2]] == [
        ["inspect_range"],
        ["declare_edit_target"],
    ]
    mutation_schema = next(
        tool for tool in client.requests[2]["tools"] if tool["name"] == "code_interpreter"
    )
    assert set(mutation_schema["parameters"]["required"]) == {"code", "declaration_id"}
    assert [tool["name"] for tool in client.requests[3]["tools"]] == ["inspect_range"]
    assert [tool["name"] for tool in client.requests[4]["tools"]] == ["submit_result"]
    _assert_recent_output_contains(client.requests[1], '"observation_id":1')
    _assert_recent_output_contains(client.requests[2], '"declaration_id":1')
    _assert_recent_output_contains(client.requests[3], '"decision":"authorized"')

    assert result.turns == 5
    assert result.tool_calls == 4
    assert result.observed_forced_tool_prefix == [
        "inspect_range",
        "declare_edit_target",
    ]
    assert [item["name"] for item in result.tool_trace] == [
        "inspect_range",
        "declare_edit_target",
        "code_interpreter",
        "inspect_range",
    ]
    assert all(item["ok"] is True for item in result.tool_trace)
    assert result.terminal_submissions == 1
    assert result.terminal_response == {
        "status": "accepted",
        "response_id": f"{arm}-response-5",
        "acknowledgement": {},
        "completion_attempt_id": 1,
    }

    final_artifact = session.artifact_ref()
    assert final_artifact.revision == initial_artifact.revision + 1
    assert final_artifact.sha256 != initial_artifact.sha256
    workbook = load_workbook(session.workbook_path, data_only=False)
    try:
        assert workbook["Sales"]["A2"].value == "Banana"
        assert workbook["Sales"]["B2"].value == 2
        assert workbook["Sales"]["D2"].value == "=B2*C2"
        assert workbook["Lookup"]["B2"].value == 0.1
    finally:
        workbook.close()

    assert len(session.artifact_transitions) == 1
    transition = session.artifact_transitions[0]
    assert transition.operation == "code_interpreter"
    assert transition.kind == "external_mutation"
    assert transition.before == initial_artifact
    assert transition.after == final_artifact

    assert len(session.committed_target_authorizations) == 1
    authorization = session.committed_target_authorizations[0]
    assert authorization.provenance.accepted is True
    assert authorization.provenance.declaration.declaration_id == 1
    assert authorization.provenance.declaration.observation_ids == (1,)
    assert authorization.provenance.declaration.target_scope.to_dict()["ranges"] == [
        {"sheet": "Sales", "range": "A2:A2", "cell_count": 1}
    ]
    assert authorization.transition == transition
    assert len(authorization.canonical_sha256) == 64

    assert result.evidence_contract is not None
    decision = result.evidence_contract["decision"]
    assert decision["allowed"] is True
    assert decision["contract_satisfied"] is True
    assert decision["artifact_changed"] is True
    assert decision["certificate"]["accepted_revision_sha256"] == final_artifact.sha256
    certificate_audit = audit_evidence_certificate(decision["certificate"])
    assert certificate_audit["valid"] is True
    assert certificate_audit["accepted_revision_sha256"] == final_artifact.sha256

    assert result.completion_attempts is not None
    assert len(result.completion_attempts) == 1
    attempt = result.completion_attempts[0]
    assert attempt["turn"] == 5
    assert attempt["stage"] == "solve"
    assert attempt["response_id"] == f"{arm}-response-5"
    assert attempt["call_id"] == f"{arm}-call-5"
    assert attempt["artifact"] == final_artifact.to_dict()
    assert audit_completion_attempt(session.workspace, attempt).valid is True

    events = _trajectory(session)
    responded = [event for event in events if event["event"] == "model.responded"]
    assert [event["payload"]["function_calls"][0]["name"] for event in responded] == [
        "inspect_range",
        "declare_edit_target",
        "code_interpreter",
        "inspect_range",
        "submit_result",
    ]
    assert all(
        len(event["payload"]["function_calls"][0]["arguments_sha256"]) == 64 for event in responded
    )
    returned = [event for event in events if event["event"] == "tool.returned"]
    assert returned[2]["payload"]["result"]["target_grounding"]["decision"] == "authorized"
    assert returned[2]["payload"]["result"]["artifact_revision_after"] == 1
    assert returned[3]["payload"]["result"]["artifact_revision"] == 1
    assert returned[3]["payload"]["result"]["matrix"] == [
        ["Item", "Qty"],
        ["Banana", 2],
        ["Pear", 4],
    ]

    lifecycle = [
        event["event"]
        for event in events
        if event["event"]
        in {
            "agent.completion_attempt_captured",
            "evidence_contract.submission_checked",
            "agent.terminal_submitted",
            "agent.completed",
        }
    ]
    assert lifecycle == [
        "agent.completion_attempt_captured",
        "evidence_contract.submission_checked",
        "agent.terminal_submitted",
        "agent.completed",
    ]
    completion_event = next(
        event for event in events if event["event"] == "agent.completion_attempt_captured"
    )
    submission_event = next(
        event for event in events if event["event"] == "evidence_contract.submission_checked"
    )
    terminal_event = next(event for event in events if event["event"] == "agent.terminal_submitted")
    assert completion_event["payload"]["record"] == attempt
    assert submission_event["payload"]["completion_attempt_id"] == attempt["attempt_id"]
    assert submission_event["payload"]["decision"]["allowed"] is True
    assert terminal_event["payload"]["completion_attempt_call_id"] == attempt["call_id"]


@pytest.mark.skipif(find_libreoffice() is None, reason="LibreOffice is not installed")
def test_v28_paper_arm_runs_read_only_stages_then_real_grounded_solve(
    sample_workbook: Path,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    client = _paper_scripted_client()
    monkeypatch.setattr("spreadsheet_harness.agent.ResponsesClient", client)
    session = WorkbookSession.create(sample_workbook, tmp_path / "v28-paper")
    initial_artifact = session.artifact_ref()

    result = run_arm(
        "paper",
        ProviderConfig("https://example.test/v1", "not-a-real-key", "test-model"),
        session,
        SkillRegistry(()),
        "Replace Sales!A2 with Banana and preserve all other cells.",
        max_output_tokens=512,
        max_elapsed_seconds=180,
        budget=RunBudget(
            max_model_calls=14,
            max_total_tokens=200,
            max_elapsed_seconds=180,
        ),
        max_turns_per_arm=20,
        evidence_contract=ContractSpec.load(EVIDENCE_CONTRACT),
        contract_mode=ContractMode.ENFORCE,
        enable_target_grounding=True,
        capture_completion_attempts=True,
    )

    assert client.instance_count == 5
    assert {stage: len(requests) for stage, requests in client.requests.items()} == {
        "extract": 3,
        "vision_verify": 3,
        "latex_verify": 2,
        "reconcile": 1,
        "solve": 5,
    }
    assert [request["tool_choice"] for request in client.requests["solve"]] == [
        {"type": "function", "name": "inspect_range"},
        {"type": "function", "name": "declare_edit_target"},
        "auto",
        {"type": "function", "name": "inspect_range"},
        {"type": "function", "name": "submit_result"},
    ]

    assert result.turns == 14
    assert result.tool_calls == 9
    assert [stage["name"] for stage in result.stages] == [
        "extract",
        "vision_verify",
        "latex_verify",
        "reconcile",
        "solve",
    ]
    for stage in result.stages[:4]:
        assert stage["read_only_verified"] is True
        assert stage["workbook_sha256_before"] == initial_artifact.sha256
        assert stage["workbook_sha256_after"] == initial_artifact.sha256
        assert len(stage["evidence_sha256"]) == 64
    vision_trace = result.stages[1]["tool_trace"]
    assert [item["name"] for item in vision_trace] == [
        "render_workbook",
        "view_image",
    ]
    assert vision_trace[1]["image_attached"] is True
    assert vision_trace[1]["image_delivery_confirmed"] is True

    solve = result.stages[-1]
    assert solve["read_only_verified"] is False
    assert solve["observed_forced_tool_prefix"] == [
        "inspect_range",
        "declare_edit_target",
    ]
    assert solve["tool_name_trace"] == [
        "inspect_range",
        "declare_edit_target",
        "code_interpreter",
        "inspect_range",
    ]

    final_artifact = session.artifact_ref()
    assert final_artifact.revision == 1
    assert final_artifact.sha256 != initial_artifact.sha256
    assert len(session.artifact_transitions) == 1
    assert len(session.committed_target_authorizations) == 1
    assert session.committed_target_authorizations[0].transition == session.artifact_transitions[0]
    workbook = load_workbook(session.workbook_path, data_only=False)
    try:
        assert workbook["Sales"]["A2"].value == "Banana"
        assert workbook["Sales"]["B2"].value == 2
        assert workbook["Sales"]["D2"].value == "=B2*C2"
    finally:
        workbook.close()

    assert result.evidence_contract is not None
    decision = result.evidence_contract["decision"]
    assert decision["allowed"] is True
    assert decision["contract_satisfied"] is True
    assert audit_evidence_certificate(decision["certificate"])["valid"] is True
    assert result.completion_attempts is not None
    assert len(result.completion_attempts) == 1
    attempt = result.completion_attempts[0]
    assert attempt["stage"] == "solve"
    assert attempt["turn"] == 5
    assert attempt["artifact"] == final_artifact.to_dict()
    assert audit_completion_attempt(session.workspace, attempt).valid is True

    events = _trajectory(session)
    completed_stages = [
        event["payload"]["stage"] for event in events if event["event"] == "agent.completed"
    ]
    assert completed_stages == [
        "extract",
        "vision_verify",
        "latex_verify",
        "reconcile",
        "solve",
    ]
    target_events = [
        event["event"] for event in events if event["event"].startswith("target_grounding.")
    ]
    assert target_events == [
        "target_grounding.enabled",
        "target_grounding.observation",
        "target_grounding.declaration",
        "target_grounding.observation",
    ]
