from __future__ import annotations

import shutil
import warnings
from collections.abc import Callable
from pathlib import Path
from zipfile import ZipFile

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image as SpreadsheetImage
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from PIL import Image as PillowImage

import spreadsheet_harness.workbook_diff as workbook_diff
from spreadsheet_harness.evidence_contract import EffectKind, EvidenceScope
from spreadsheet_harness.render import find_libreoffice, recalculate_workbook
from spreadsheet_harness.workbook_diff import diff_workbooks


def _copies(sample_workbook: Path, tmp_path: Path) -> tuple[Path, Path]:
    before = tmp_path / "before.xlsx"
    after = tmp_path / "after.xlsx"
    shutil.copy2(sample_workbook, before)
    shutil.copy2(sample_workbook, after)
    return before, after


def _replace_or_add_zip_part(path: Path, part_name: str, payload: bytes) -> None:
    replacement = path.with_name(f".{path.stem}-rewrite{path.suffix}")
    with ZipFile(path) as source, ZipFile(replacement, "w") as target:
        replaced = False
        for info in source.infolist():
            if info.filename == part_name:
                target.writestr(info, payload)
                replaced = True
            else:
                target.writestr(info, source.read(info))
        if not replaced:
            target.writestr(part_name, payload)
        target.comment = source.comment
    replacement.replace(path)


def _add_empty_custom_properties_plumbing(
    path: Path,
    *,
    add_part: bool = True,
    add_relationship: bool = True,
    add_content_type: bool = True,
    content_type: bytes = (
        b"application/vnd.openxmlformats-officedocument.custom-properties+xml"
    ),
    relationship_count: int = 1,
) -> None:
    if add_part:
        _replace_or_add_zip_part(
            path,
            "docProps/custom.xml",
            (
                b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                b'<Properties xmlns="http://schemas.openxmlformats.org/'
                b'officeDocument/2006/custom-properties" '
                b'xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/'
                b'docPropsVTypes"/>'
            ),
        )
    if add_relationship:
        with ZipFile(path) as package:
            relationships = package.read("_rels/.rels")
        additions = b"".join(
            (
                b'<Relationship Id="rIdCustom'
                + str(index).encode("ascii")
                + b'" Type="http://schemas.openxmlformats.org/officeDocument/'
                b'2006/relationships/custom-properties" '
                b'Target="docProps/custom.xml"/>'
            )
            for index in range(1, relationship_count + 1)
        )
        changed = relationships.replace(
            b"</Relationships>",
            additions + b"</Relationships>",
            1,
        )
        assert changed != relationships
        _replace_or_add_zip_part(path, "_rels/.rels", changed)
    if add_content_type:
        with ZipFile(path) as package:
            declarations = package.read("[Content_Types].xml")
        addition = (
            b'<Override PartName="/docProps/custom.xml" ContentType="'
            + content_type
            + b'"/>'
        )
        changed = declarations.replace(
            b"</Types>",
            addition + b"</Types>",
            1,
        )
        assert changed != declarations
        _replace_or_add_zip_part(path, "[Content_Types].xml", changed)


def _rich_workbook(tmp_path: Path) -> Path:
    image_path = tmp_path / "source.png"
    PillowImage.new("RGB", (12, 8), color=(190, 30, 45)).save(image_path)

    path = tmp_path / "rich.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Label", "Actual", "Alternative"])
    for row, value in enumerate((3, 7, 4, 9), start=2):
        sheet.cell(row, 1, f"Item {row - 1}")
        sheet.cell(row, 2, value)
        sheet.cell(row, 3, value + 10)

    chart = BarChart()
    chart.title = "Actuals"
    chart.style = 10
    chart.add_data(Reference(sheet, min_col=2, min_row=1, max_row=5), titles_from_data=True)
    chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=5))
    sheet.add_chart(chart, "E2")

    comparison_chart = BarChart()
    comparison_chart.title = "Alternatives"
    comparison_chart.style = 15
    comparison_chart.add_data(
        Reference(sheet, min_col=3, min_row=1, max_row=5), titles_from_data=True
    )
    sheet.add_chart(comparison_chart, "E18")

    sheet.add_image(SpreadsheetImage(image_path), "J2")
    sheet.conditional_formatting.add(
        "B2:B5",
        CellIsRule(
            operator="greaterThan",
            formula=["5"],
            fill=PatternFill("solid", fgColor="FFF2CC"),
        ),
    )
    validation = DataValidation(type="whole", operator="between", formula1="0", formula2="20")
    validation.add("C2:C5")
    sheet.add_data_validation(validation)
    sheet.page_margins.left = 0.6
    sheet.print_options.gridLines = True
    sheet.page_setup.orientation = "landscape"
    workbook.save(path)
    workbook.close()
    return path


def _rich_copies(tmp_path: Path) -> tuple[Path, Path]:
    source = _rich_workbook(tmp_path)
    before = tmp_path / "rich-before.xlsx"
    after = tmp_path / "rich-after.xlsx"
    shutil.copy2(source, before)
    shutil.copy2(source, after)
    return before, after


def test_identical_workbooks_have_no_semantic_effects(
    sample_workbook: Path, tmp_path: Path
) -> None:
    before, after = _copies(sample_workbook, tmp_path)

    result = diff_workbooks(before, after)

    assert result.complete is True
    assert result.semantic_changed is False
    assert result.effects == frozenset()
    assert result.scope.empty
    assert result.formula_scope.empty


def test_resaving_without_semantic_changes_is_not_a_user_mutation(
    sample_workbook: Path, tmp_path: Path
) -> None:
    before, after = _copies(sample_workbook, tmp_path)
    workbook = load_workbook(after)
    workbook.save(after)
    workbook.close()

    result = diff_workbooks(before, after)

    assert result.complete is True
    assert result.semantic_changed is False


def test_rich_workbook_byte_rewrite_remains_a_semantic_noop(tmp_path: Path) -> None:
    before, after = _rich_copies(tmp_path)
    original_bytes = after.read_bytes()
    with ZipFile(after, "a") as archive:
        archive.comment = b"different package bytes, identical workbook semantics"

    assert after.read_bytes() != original_bytes
    result = diff_workbooks(before, after)

    assert result.complete is True
    assert result.semantic_changed is False
    assert result.effects == frozenset()


def test_lossy_openpyxl_chart_resave_is_not_misclassified_as_noop(tmp_path: Path) -> None:
    before, after = _rich_copies(tmp_path)
    workbook = load_workbook(after)
    workbook.save(after)
    workbook.close()

    result = diff_workbooks(before, after)

    assert result.complete is True
    assert {EffectKind.STRUCTURE, EffectKind.VISUAL} <= result.effects
    assert result.scope.covers(EvidenceScope.worksheet("Data"))


def test_raw_chart_part_mapping_detects_second_chart_style_change(tmp_path: Path) -> None:
    source = tmp_path / "two-charts.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    for value in range(1, 6):
        sheet.append([value, value * 2])
    first_chart = BarChart()
    first_chart.add_data(Reference(sheet, min_col=1, min_row=1, max_row=5))
    sheet.add_chart(first_chart, "D2")
    second_chart = BarChart()
    second_chart.style = 15
    second_chart.add_data(Reference(sheet, min_col=2, min_row=1, max_row=5))
    sheet.add_chart(second_chart, "D18")
    workbook.save(source)
    workbook.close()

    before = tmp_path / "two-charts-before.xlsx"
    after = tmp_path / "two-charts-after.xlsx"
    shutil.copy2(source, before)
    shutil.copy2(source, after)
    workbook = load_workbook(after)
    workbook["Data"]._charts[1].style = 22
    workbook.save(after)
    workbook.close()

    result = diff_workbooks(before, after)

    assert result.complete is True
    assert {EffectKind.STRUCTURE, EffectKind.VISUAL} <= result.effects
    assert result.scope.covers(EvidenceScope.worksheet("Data"))


def test_openpyxl_loss_warning_fails_closed(
    sample_workbook: Path, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    before, after = _copies(sample_workbook, tmp_path)
    real_load_workbook = workbook_diff.load_workbook

    def warned_load_workbook(*args: object, **kwargs: object) -> object:
        warnings.warn("unsupported extension will be removed", UserWarning, stacklevel=2)
        return real_load_workbook(*args, **kwargs)

    monkeypatch.setattr(workbook_diff, "load_workbook", warned_load_workbook)

    result = diff_workbooks(before, after)

    assert result.complete is False
    assert result.effects == {EffectKind.UNKNOWN}
    assert result.scope == EvidenceScope.workbook()
    assert "potentially lossy" in result.reasons[0]


def test_changed_unsupported_package_part_fails_closed(
    sample_workbook: Path, tmp_path: Path
) -> None:
    before, after = _copies(sample_workbook, tmp_path)
    part_name = "xl/embeddings/oleObject1.bin"
    with ZipFile(before, "a") as archive:
        archive.writestr(part_name, b"opaque-before")
    with ZipFile(after, "a") as archive:
        archive.writestr(part_name, b"opaque-after")

    result = diff_workbooks(before, after)

    assert result.complete is False
    assert result.effects == {EffectKind.UNKNOWN}
    assert result.scope == EvidenceScope.workbook()
    assert "unsupported OOXML part set or content changed" in result.reasons[0]
    assert part_name in result.reasons[0]


@pytest.mark.parametrize(
    "part_name",
    [
        "xl/unknownFeature1.xml",
        "docProps/custom.xml",
        "xl/calcChain.xml",
        "xl/persons/person.xml",
        "xl/comments1.xml",
    ],
    ids=["unknown-custom", "custom-properties", "calc-chain", "people", "comments"],
)
def test_changed_residual_package_part_fails_closed(
    sample_workbook: Path,
    tmp_path: Path,
    part_name: str,
) -> None:
    before, after = _copies(sample_workbook, tmp_path)
    _replace_or_add_zip_part(before, part_name, b'<part state="before"/>')
    _replace_or_add_zip_part(after, part_name, b'<part state="after"/>')

    result = diff_workbooks(before, after)

    assert result.complete is False
    assert result.effects == {EffectKind.UNKNOWN}
    assert result.scope == EvidenceScope.workbook()
    assert part_name in result.reasons[0]


@pytest.mark.parametrize(
    ("part_name", "old", "new"),
    [
        ("docProps/core.xml", b">openpyxl</dc:creator>", b">changed</dc:creator>"),
    ],
    ids=["core-creator"],
)
def test_changed_document_property_part_fails_closed(
    sample_workbook: Path,
    tmp_path: Path,
    part_name: str,
    old: bytes,
    new: bytes,
) -> None:
    before, after = _copies(sample_workbook, tmp_path)
    with ZipFile(after) as package:
        original = package.read(part_name)
    changed = original.replace(old, new, 1)
    assert changed != original
    _replace_or_add_zip_part(after, part_name, changed)

    result = diff_workbooks(before, after)

    assert result.complete is False
    assert result.effects == {EffectKind.UNKNOWN}
    assert result.scope == EvidenceScope.workbook()
    assert part_name in result.reasons[0]


def test_producer_only_application_property_change_is_semantic_noop(
    sample_workbook: Path,
    tmp_path: Path,
) -> None:
    before, after = _copies(sample_workbook, tmp_path)
    part_name = "docProps/app.xml"
    with ZipFile(after) as package:
        original = package.read(part_name)
    changed = original.replace(
        b">Microsoft Excel Compatible / Openpyxl 3.1.5</Application>",
        b">LibreOffice/26.2</Application>",
        1,
    )
    assert changed != original
    _replace_or_add_zip_part(after, part_name, changed)

    result = diff_workbooks(before, after)

    assert result.complete is True
    assert result.semantic_changed is False
    assert result.effects == frozenset()


def test_material_custom_document_property_change_fails_closed(
    sample_workbook: Path,
    tmp_path: Path,
) -> None:
    before, after = _copies(sample_workbook, tmp_path)
    payload = b"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/custom-properties"
 xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">
 <property fmtid="{D5CDD505-2E9C-101B-9397-08002B2CF9AE}" pid="2" name="ReviewState">
  <vt:lpwstr>material</vt:lpwstr>
 </property>
</Properties>"""
    _replace_or_add_zip_part(after, "docProps/custom.xml", payload)

    result = diff_workbooks(before, after)

    assert result.complete is False
    assert result.effects == {EffectKind.UNKNOWN}
    assert "docProps/custom.xml" in result.reasons[0]


def test_complete_empty_custom_properties_plumbing_is_a_semantic_noop(
    sample_workbook: Path,
    tmp_path: Path,
) -> None:
    before, after = _copies(sample_workbook, tmp_path)
    _add_empty_custom_properties_plumbing(after)

    result = diff_workbooks(before, after)

    assert result.complete is True, result.reasons
    assert result.semantic_changed is False
    assert result.effects == frozenset()


@pytest.mark.parametrize(
    ("add_part", "add_relationship", "add_content_type", "content_type", "relationship_count"),
    [
        (False, True, False, b"unused", 1),
        (True, True, False, b"unused", 1),
        (True, False, True, b"application/vnd.openxmlformats-officedocument.custom-properties+xml", 1),
        (True, True, True, b"application/x-wrong-custom-properties", 1),
        (True, True, True, b"application/vnd.openxmlformats-officedocument.custom-properties+xml", 2),
    ],
    ids=[
        "dangling-relationship",
        "missing-content-type",
        "missing-relationship",
        "wrong-content-type",
        "duplicate-relationship",
    ],
)
def test_incomplete_empty_custom_properties_plumbing_fails_closed(
    sample_workbook: Path,
    tmp_path: Path,
    add_part: bool,
    add_relationship: bool,
    add_content_type: bool,
    content_type: bytes,
    relationship_count: int,
) -> None:
    before, after = _copies(sample_workbook, tmp_path)
    _add_empty_custom_properties_plumbing(
        after,
        add_part=add_part,
        add_relationship=add_relationship,
        add_content_type=add_content_type,
        content_type=content_type,
        relationship_count=relationship_count,
    )

    result = diff_workbooks(before, after)

    assert result.complete is False
    assert result.effects == {EffectKind.UNKNOWN}
    assert result.scope == EvidenceScope.workbook()
    assert "docProps/custom.xml" in result.reasons[0]


def test_unexplained_change_in_accounted_package_plumbing_fails_closed(
    sample_workbook: Path,
    tmp_path: Path,
) -> None:
    before, after = _copies(sample_workbook, tmp_path)
    part_name = "[Content_Types].xml"
    with ZipFile(after) as package:
        original = package.read(part_name)
    changed = original.replace(
        b"</Types>",
        (
            b'<Default Extension="mystery" '
            b'ContentType="application/x-spreadsheet-harness-test"/></Types>'
        ),
        1,
    )
    assert changed != original
    _replace_or_add_zip_part(after, part_name, changed)

    result = diff_workbooks(before, after)

    assert result.complete is False
    assert result.effects == {EffectKind.UNKNOWN}
    assert result.scope == EvidenceScope.workbook()
    assert "content types" in result.reasons[0]
    assert part_name in result.reasons[0]


def test_content_types_root_attribute_fails_closed(
    sample_workbook: Path,
    tmp_path: Path,
) -> None:
    before, after = _copies(sample_workbook, tmp_path)
    part_name = "[Content_Types].xml"
    with ZipFile(after) as package:
        original = package.read(part_name)
    changed = original.replace(b"<Types ", b'<Types suspicious="true" ', 1)
    assert changed != original
    _replace_or_add_zip_part(after, part_name, changed)

    result = diff_workbooks(before, after)

    assert result.complete is False
    assert result.effects == {EffectKind.UNKNOWN}
    assert part_name in result.reasons[0]


def test_material_root_relationship_change_fails_closed(
    sample_workbook: Path,
    tmp_path: Path,
) -> None:
    before, after = _copies(sample_workbook, tmp_path)
    part_name = "_rels/.rels"
    with ZipFile(after) as package:
        original = package.read(part_name)
    changed = original.replace(
        b"</Relationships>",
        (
            b'<Relationship Id="rId999" '
            b'Type="https://example.test/material-relation" '
            b'Target="material.xml"/></Relationships>'
        ),
        1,
    )
    assert changed != original
    _replace_or_add_zip_part(after, part_name, changed)

    result = diff_workbooks(before, after)

    assert result.complete is False
    assert result.effects == {EffectKind.UNKNOWN}
    assert "root OOXML relationships" in result.reasons[0]


@pytest.mark.skipif(find_libreoffice() is None, reason="LibreOffice is not installed")
def test_libreoffice_recalculation_round_trip_has_complete_typed_effects(
    sample_workbook: Path,
    tmp_path: Path,
) -> None:
    before, after = _copies(sample_workbook, tmp_path)

    recalculate_workbook(after, after, timeout_seconds=30.0)
    result = diff_workbooks(before, after)

    assert result.complete is True, result.reasons
    assert result.semantic_changed is True
    assert EffectKind.UNKNOWN not in result.effects
    assert {EffectKind.STRUCTURE, EffectKind.STYLE, EffectKind.VISUAL} <= result.effects


def test_non_excel_libreoffice_formula_syntax_marker_fails_closed(
    sample_workbook: Path,
    tmp_path: Path,
) -> None:
    before, after = _copies(sample_workbook, tmp_path)
    part_name = "xl/workbook.xml"
    with ZipFile(after) as package:
        payload = package.read(part_name)
    extension = (
        b'<extLst><ext uri="{7626C862-2A13-11E5-B345-FEFF819CDC9F}">'
        b'<loext:extCalcPr xmlns:loext="http://schemas.libreoffice.org/" '
        b'stringRefSyntax="CalcA1"/></ext></extLst>'
    )
    changed = payload.replace(b"</workbook>", extension + b"</workbook>", 1)
    assert changed != payload
    _replace_or_add_zip_part(after, part_name, changed)

    result = diff_workbooks(before, after)

    assert result.complete is False
    assert result.effects == {EffectKind.UNKNOWN}
    assert part_name in result.reasons[0]


def test_manual_calculation_policy_cannot_hide_beside_a_cell_edit(
    sample_workbook: Path,
    tmp_path: Path,
) -> None:
    before, after = _copies(sample_workbook, tmp_path)
    workbook = load_workbook(after)
    workbook["Sales"]["A2"] = "edited"
    workbook.save(after)
    workbook.close()
    part_name = "xl/workbook.xml"
    with ZipFile(after) as package:
        original = package.read(part_name)
    changed = original.replace(
        b"<calcPr ",
        b'<calcPr calcMode="manual" ',
        1,
    )
    assert changed != original
    _replace_or_add_zip_part(after, part_name, changed)

    result = diff_workbooks(before, after)

    assert result.complete is False
    assert result.effects == {EffectKind.UNKNOWN}
    assert result.scope == EvidenceScope.workbook()
    assert part_name in result.reasons[0]


def test_unclassified_workbook_xml_cannot_hide_beside_a_cell_edit(
    sample_workbook: Path,
    tmp_path: Path,
) -> None:
    before, after = _copies(sample_workbook, tmp_path)
    workbook = load_workbook(after)
    workbook["Sales"]["A2"] = "edited"
    workbook.save(after)
    workbook.close()
    part_name = "xl/workbook.xml"
    with ZipFile(after) as package:
        original = package.read(part_name)
    changed = original.replace(
        b"<sheets>",
        b'<sheets unclassified="true">',
        1,
    )
    assert changed != original
    _replace_or_add_zip_part(after, part_name, changed)

    result = diff_workbooks(before, after)

    assert result.complete is False
    assert result.effects == {EffectKind.UNKNOWN}
    assert result.scope == EvidenceScope.workbook()
    assert part_name in result.reasons[0]


def test_relationship_unreachable_part_cannot_hide_beside_a_cell_edit(
    sample_workbook: Path,
    tmp_path: Path,
) -> None:
    before, after = _copies(sample_workbook, tmp_path)
    workbook = load_workbook(after)
    workbook["Sales"]["B4"] = 17
    workbook.save(after)
    workbook.close()
    disguised_part = "xl/worksheets/sheet999.xml"
    _replace_or_add_zip_part(after, disguised_part, b"<worksheet/>")

    result = diff_workbooks(before, after)

    assert result.complete is False
    assert result.effects == {EffectKind.UNKNOWN}
    assert result.scope == EvidenceScope.workbook()
    assert disguised_part in result.reasons[0]


def test_unknown_worksheet_namespace_cannot_hide_beside_a_cell_edit(
    sample_workbook: Path,
    tmp_path: Path,
) -> None:
    before, after = _copies(sample_workbook, tmp_path)
    part_name = "xl/worksheets/sheet1.xml"
    closing_tag = b"</worksheet>"
    with ZipFile(before) as package:
        before_payload = package.read(part_name)
    before_payload = before_payload.replace(
        closing_tag,
        b'<attack:payload xmlns:attack="urn:attack">before</attack:payload>'
        + closing_tag,
        1,
    )
    _replace_or_add_zip_part(before, part_name, before_payload)
    workbook = load_workbook(after)
    workbook["Sales"]["B4"] = 17
    workbook.save(after)
    workbook.close()
    with ZipFile(after) as package:
        after_payload = package.read(part_name)
    after_payload = after_payload.replace(
        closing_tag,
        b'<attack:payload xmlns:attack="urn:attack">after</attack:payload>'
        + closing_tag,
        1,
    )
    _replace_or_add_zip_part(after, part_name, after_payload)

    result = diff_workbooks(before, after)

    assert result.complete is False
    assert result.effects == {EffectKind.UNKNOWN}
    assert part_name in result.reasons[0]


def test_workbook_theme_change_has_visual_workbook_scope(
    sample_workbook: Path, tmp_path: Path
) -> None:
    before, after = _copies(sample_workbook, tmp_path)
    workbook = load_workbook(after)
    assert workbook.loaded_theme is not None
    workbook.loaded_theme = workbook.loaded_theme.replace(b"4F81BD", b"112233", 1)
    workbook.save(after)
    workbook.close()

    result = diff_workbooks(before, after)

    assert result.complete is True
    assert {EffectKind.STYLE, EffectKind.VISUAL} <= result.effects
    assert result.scope == EvidenceScope.workbook()


def _change_chart_title(chart: BarChart) -> None:
    chart.title = "Revised actuals"


def _change_chart_type(chart: BarChart) -> None:
    chart.type = "bar"


def _change_chart_source(chart: BarChart) -> None:
    chart.series[0].val.numRef.f = "'Data'!$C$2:$C$5"


def _change_chart_anchor(chart: BarChart) -> None:
    chart.anchor._from.col += 1


def _change_chart_style(chart: BarChart) -> None:
    chart.style = 12


@pytest.mark.parametrize(
    "mutation",
    [
        _change_chart_title,
        _change_chart_type,
        _change_chart_source,
        _change_chart_anchor,
        _change_chart_style,
    ],
    ids=["title", "type", "series-source", "anchor", "style"],
)
def test_same_count_chart_semantic_changes_are_visual_and_structural(
    tmp_path: Path,
    mutation: Callable[[BarChart], None],
) -> None:
    before, after = _rich_copies(tmp_path)
    workbook = load_workbook(after)
    mutation(workbook["Data"]._charts[0])
    workbook.save(after)
    workbook.close()

    result = diff_workbooks(before, after)

    assert result.complete is True
    assert {EffectKind.STRUCTURE, EffectKind.VISUAL} <= result.effects
    assert result.scope.covers(EvidenceScope.worksheet("Data"))


@pytest.mark.parametrize("mutation", ["content", "anchor", "size"])
def test_same_count_image_content_anchor_and_size_changes_are_visual(
    tmp_path: Path, mutation: str
) -> None:
    before, after = _rich_copies(tmp_path)
    workbook = load_workbook(after)
    sheet = workbook["Data"]
    image = sheet._images[0]
    if mutation == "content":
        replacement_path = tmp_path / "replacement.png"
        PillowImage.new("RGB", (12, 8), color=(20, 110, 170)).save(replacement_path)
        sheet._images.clear()
        sheet.add_image(SpreadsheetImage(replacement_path), "J2")
    elif mutation == "anchor":
        image.anchor._from.row += 1
    else:
        image.anchor.ext.cx += 100_000
    workbook.save(after)
    workbook.close()

    result = diff_workbooks(before, after)

    assert result.complete is True
    assert {EffectKind.STRUCTURE, EffectKind.VISUAL} <= result.effects
    assert result.scope.covers(EvidenceScope.worksheet("Data"))


@pytest.mark.parametrize("mutation", ["formula", "differential-style"])
def test_conditional_formatting_rule_changes_are_detected(tmp_path: Path, mutation: str) -> None:
    before, after = _rich_copies(tmp_path)
    workbook = load_workbook(after)
    conditional_formatting = next(iter(workbook["Data"].conditional_formatting))
    rule = conditional_formatting.rules[0]
    if mutation == "formula":
        rule.formula = ["8"]
    else:
        rule.dxf.fill.fgColor.rgb = "FF00AA44"
    workbook.save(after)
    workbook.close()

    result = diff_workbooks(before, after)

    assert result.complete is True
    assert {EffectKind.STYLE, EffectKind.VISUAL} <= result.effects
    assert result.scope.covers(EvidenceScope.worksheet("Data"))


def test_data_validation_rule_change_is_detected(tmp_path: Path) -> None:
    before, after = _rich_copies(tmp_path)
    workbook = load_workbook(after)
    workbook["Data"].data_validations.dataValidation[0].formula2 = "50"
    workbook.save(after)
    workbook.close()

    result = diff_workbooks(before, after)

    assert result.complete is True
    assert EffectKind.STRUCTURE in result.effects
    assert result.scope.covers(EvidenceScope.worksheet("Data"))


@pytest.mark.parametrize("mutation", ["margin", "print-options", "page-setup"])
def test_page_and_print_semantic_changes_are_detected(tmp_path: Path, mutation: str) -> None:
    before, after = _rich_copies(tmp_path)
    workbook = load_workbook(after)
    sheet = workbook["Data"]
    if mutation == "margin":
        sheet.page_margins.left = 1.25
    elif mutation == "print-options":
        sheet.print_options.gridLines = False
    else:
        sheet.page_setup.orientation = "portrait"
    workbook.save(after)
    workbook.close()

    result = diff_workbooks(before, after)

    assert result.complete is True
    assert {EffectKind.STRUCTURE, EffectKind.VISUAL} <= result.effects
    assert result.scope.covers(EvidenceScope.worksheet("Data"))


def test_value_formula_and_style_effects_have_typed_scopes(
    sample_workbook: Path, tmp_path: Path
) -> None:
    before, after = _copies(sample_workbook, tmp_path)
    workbook = load_workbook(after)
    sheet = workbook["Sales"]
    sheet["B4"] = 7
    sheet["D4"] = "=B4*C4"
    sheet["F6"].fill = PatternFill("solid", fgColor="FFF2CC")
    workbook.save(after)
    workbook.close()

    result = diff_workbooks(before, after)

    assert result.complete is True
    assert result.semantic_changed is True
    assert result.effects == {
        EffectKind.VALUE,
        EffectKind.FORMULA,
        EffectKind.STYLE,
        EffectKind.VISUAL,
    }
    assert result.scope.covers(EvidenceScope.one("Sales", "B4:B4"))
    assert result.scope.covers(EvidenceScope.one("Sales", "D4:D4"))
    assert result.scope.covers(EvidenceScope.one("Sales", "F6:F6"))
    assert result.formula_scope == EvidenceScope.one("Sales", "D4:D4")
    assert result.changed_cell_count == 3


def test_structure_and_layout_changes_require_sheet_scope(
    sample_workbook: Path, tmp_path: Path
) -> None:
    before, after = _copies(sample_workbook, tmp_path)
    workbook = load_workbook(after)
    sheet = workbook["Sales"]
    sheet.merge_cells("F1:G1")
    sheet.row_dimensions[4].height = 28
    workbook.save(after)
    workbook.close()

    result = diff_workbooks(before, after)

    assert result.complete is True
    assert {EffectKind.STRUCTURE, EffectKind.VISUAL} <= result.effects
    assert result.scope.covers(EvidenceScope.worksheet("Sales"))


def test_worksheet_set_change_uses_workbook_wildcard(sample_workbook: Path, tmp_path: Path) -> None:
    before, after = _copies(sample_workbook, tmp_path)
    workbook = load_workbook(after)
    workbook.create_sheet("Added")
    workbook.save(after)
    workbook.close()

    result = diff_workbooks(before, after)

    assert result.semantic_changed is True
    assert result.scope == EvidenceScope.workbook()
    assert {EffectKind.STRUCTURE, EffectKind.VISUAL} <= result.effects


def test_added_formula_sheet_preserves_formula_trigger_and_scope(
    sample_workbook: Path,
    tmp_path: Path,
) -> None:
    before, after = _copies(sample_workbook, tmp_path)
    workbook = load_workbook(after)
    workbook.create_sheet("Calculated")["A1"] = "=1+1"
    workbook.save(after)
    workbook.close()

    result = diff_workbooks(before, after)

    assert result.complete is True
    assert {EffectKind.FORMULA, EffectKind.STRUCTURE, EffectKind.VISUAL} <= result.effects
    assert result.scope == EvidenceScope.workbook()
    assert result.formula_scope.covers(EvidenceScope.one("Calculated", "A1:A1"))
    assert result.changed_cell_count == 1


def test_scan_limit_fails_closed_to_unknown_workbook_scope(
    sample_workbook: Path, tmp_path: Path
) -> None:
    before, after = _copies(sample_workbook, tmp_path)

    result = diff_workbooks(before, after, max_scanned_cells=1)

    assert result.complete is False
    assert result.semantic_changed is True
    assert result.effects == {EffectKind.UNKNOWN}
    assert result.scope == EvidenceScope.workbook()
    assert "exceeded" in result.reasons[0]
