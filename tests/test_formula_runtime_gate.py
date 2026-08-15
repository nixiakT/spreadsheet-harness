from __future__ import annotations

import json
import zipfile
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook

from spreadsheet_harness.agent import ResponseTurn, SpreadsheetAgent
from spreadsheet_harness.budget import RunBudget
from spreadsheet_harness.config import ProviderConfig
from spreadsheet_harness.errors import AgentExecutionFailure
from spreadsheet_harness.formula_runtime import (
    formula_coordinate_sha256,
    formula_inventory,
)
from spreadsheet_harness.session import WorkbookSession
from spreadsheet_harness.tools import ToolOutcome
from spreadsheet_harness.trajectory import read_trajectory


def _schema(name: str, properties: dict[str, Any]) -> dict[str, Any]:
    return {
        "type": "function",
        "name": name,
        "description": name,
        "strict": False,
        "parameters": {
            "type": "object",
            "properties": properties,
            "required": [],
            "additionalProperties": False,
        },
    }


class FormulaGateTools:
    def __init__(
        self,
        session: WorkbookSession,
        *,
        sparse_valid: bool = True,
        original_formula: str | None = None,
        drop_formula_during_sparse: bool = False,
        normalize_formula_during_recalc: bool = False,
    ) -> None:
        self.session = session
        self.sparse_valid = sparse_valid
        self.original_formula = original_formula
        self.drop_formula_during_sparse = drop_formula_during_sparse
        self.normalize_formula_during_recalc = normalize_formula_during_recalc
        self.normalized_formula = False
        self.pending_scope: tuple[tuple[str, str], ...] = ()
        self.scope_history: list[tuple[tuple[str, str], ...]] = []
        self.calls: list[tuple[str, dict[str, Any]]] = []
        self.schemas = [
            _schema("code_interpreter", {"code": {"type": "string"}}),
            _schema(
                "recalculate_and_read",
                {
                    "validation_scope": {"type": "string"},
                    "sheet": {"type": ["string", "null"]},
                    "range_ref": {"type": ["string", "null"]},
                },
            ),
            _schema("undo_last", {}),
        ]

    def set_pending_formula_validation_scope(
        self, coordinates: set[tuple[str, str]]
    ) -> None:
        self.pending_scope = tuple(sorted(coordinates))
        self.scope_history.append(self.pending_scope)

    def _mutate(self, action: str) -> None:
        workbook = load_workbook(self.session.workbook_path)
        sheet = workbook["Sales"]
        if action == "set_formula":
            sheet["H1"] = "=1+1"
        elif action == "rewrite_formula":
            sheet["H1"] = "=2+2"
        elif action == "set_value":
            sheet["H1"] = 7
        elif action == "delete_formula":
            sheet["H1"] = None
        elif action == "bulk_formulas":
            for row in range(1, 602):
                sheet.cell(row, 10, f"={row}+1")
        else:
            raise AssertionError(f"unknown test action: {action}")
        workbook.save(self.session.workbook_path)
        workbook.close()

    def _normalize_existing_formula_xml(self) -> None:
        rewritten = self.session.workbook_path.with_suffix(".normalized.xlsx")
        with zipfile.ZipFile(self.session.workbook_path) as source, zipfile.ZipFile(
            rewritten, "w"
        ) as destination:
            for member in source.infolist():
                payload = source.read(member)
                if member.filename == "xl/worksheets/sheet1.xml":
                    payload = payload.replace(
                        b"<f>B2*C2</f>",
                        b"<f> B2*C2 </f>",
                    )
                destination.writestr(member, payload)
        rewritten.replace(self.session.workbook_path)
        self.normalized_formula = True

    def _restore(self) -> None:
        workbook = load_workbook(self.session.workbook_path)
        workbook["Sales"]["H1"] = self.original_formula
        workbook.save(self.session.workbook_path)
        workbook.close()

    def invoke(self, name: str, arguments: dict[str, Any]) -> ToolOutcome:
        self.calls.append((name, dict(arguments)))
        if name == "code_interpreter":
            self._mutate(str(arguments["code"]))
            return ToolOutcome({"ok": True, "workbook_changed": True})
        if name == "undo_last":
            self._restore()
            return ToolOutcome({"ok": True, "workbook_changed": True})
        assert name == "recalculate_and_read"
        if self.normalize_formula_during_recalc and not self.normalized_formula:
            self._normalize_existing_formula_xml()
        if arguments.get("validation_scope") == "pending_formula_changes":
            if self.drop_formula_during_sparse:
                workbook = load_workbook(self.session.workbook_path)
                workbook["Sales"]["H1"] = None
                workbook.save(self.session.workbook_path)
                workbook.close()
            inventory = formula_inventory(self.session.workbook_path)
            present = sum(
                coordinate in inventory.cells for coordinate in self.pending_scope
            )
            if self.sparse_valid:
                calculation_errors = {
                    "count": 0,
                    "coordinates": [],
                    "coordinate_limit": 32,
                    "coordinates_truncated": False,
                }
            else:
                sheet, coordinate = self.pending_scope[0]
                calculation_errors = {
                    "count": 1,
                    "coordinates": [
                        {
                            "sheet": sheet,
                            "coordinate": coordinate,
                            "error": "#DIV/0!",
                        }
                    ],
                    "coordinate_limit": 32,
                    "coordinates_truncated": False,
                }
            return ToolOutcome(
                {
                    "ok": True,
                    "calculation_valid": self.sparse_valid,
                    "calculation_errors": calculation_errors,
                    "validation_scope": {
                        "kind": "pending_formula_changes",
                        "coordinate_count": len(self.pending_scope),
                        "coordinate_sha256": formula_coordinate_sha256(
                            self.pending_scope
                        ),
                        "coverage_complete": True,
                        "formula_cells_present": present,
                        "formula_cells_absent": len(self.pending_scope) - present,
                        "cached_blank_count": present,
                        "cached_blank_coordinates": [],
                        "cached_blank_coordinates_truncated": present > 0,
                    },
                    "calculation": {"backend": "fake-libreoffice"},
                }
            )
        sheet = str(arguments.get("sheet", "Sales"))
        range_ref = str(arguments.get("range_ref", "A1"))
        return ToolOutcome(
            {
                "ok": True,
                "calculation_valid": True,
                "calculation_errors": {
                    "sheet": sheet,
                    "range": range_ref,
                    "count": 0,
                    "coordinates": [],
                    "coordinate_limit": 32,
                    "coordinates_truncated": False,
                },
                "inspection": {"sheet": sheet, "range": range_ref},
                "calculation": {"backend": "fake-libreoffice"},
            }
        )


def _tool_step(name: str, arguments: dict[str, Any] | None = None) -> tuple[str, Any]:
    return "tool", (name, arguments or {})


def _text_step(text: str = "done") -> tuple[str, Any]:
    return "text", text


def _patch_sequence_client(
    monkeypatch: Any,
    steps: list[tuple[str, Any]],
) -> list[dict[str, Any]]:
    requests: list[dict[str, Any]] = []

    class SequenceClient:
        def __init__(self, *_: Any, **__: Any) -> None:
            self.turn = 0

        def __enter__(self) -> SequenceClient:
            return self

        def __exit__(self, *_: Any) -> None:
            return None

        def create(self, payload: dict[str, Any], **_: Any) -> ResponseTurn:
            requests.append(payload)
            kind, value = steps[self.turn]
            self.turn += 1
            if kind == "text":
                output = [
                    {
                        "type": "message",
                        "role": "assistant",
                        "content": [{"type": "output_text", "text": value}],
                    }
                ]
                text = str(value)
            else:
                name, arguments = value
                output = [
                    {
                        "type": "function_call",
                        "id": f"fc-{self.turn}",
                        "call_id": f"call-{self.turn}",
                        "name": name,
                        "arguments": json.dumps(arguments),
                    }
                ]
                text = ""
            return ResponseTurn(
                f"response-{self.turn}",
                output,
                text,
                {"input_tokens": 10, "output_tokens": 2, "total_tokens": 12},
            )

    monkeypatch.setattr("spreadsheet_harness.agent.ResponsesClient", SequenceClient)
    return requests


def _agent(
    tools: FormulaGateTools,
    *,
    max_turns: int,
    require_workbook_change: bool = True,
    force_code_on_stalled_edit: bool = False,
    budget: RunBudget | None = None,
) -> SpreadsheetAgent:
    return SpreadsheetAgent(
        ProviderConfig("https://example.test/v1", "test-key", "test-model"),
        tools,  # type: ignore[arg-type]
        max_turns=max_turns,
        max_output_tokens=2_000,
        required_tool_termination=True,
        require_workbook_change=require_workbook_change,
        require_formula_runtime_validation=True,
        force_code_on_stalled_edit=force_code_on_stalled_edit,
        budget=budget,
    )


def _session(sample_workbook: Path, tmp_path: Path, name: str) -> WorkbookSession:
    return WorkbookSession.create(sample_workbook, tmp_path / name)


def test_formula_edit_cannot_submit_until_sparse_runtime_validation(
    sample_workbook: Path, tmp_path: Path, monkeypatch: Any
) -> None:
    session = _session(sample_workbook, tmp_path, "formula-submit-blocked")
    tools = FormulaGateTools(session)
    requests = _patch_sequence_client(
        monkeypatch,
        [
            _tool_step("code_interpreter", {"code": "set_formula"}),
            _tool_step("submit_result"),
            _tool_step(
                "recalculate_and_read",
                {"validation_scope": "pending_formula_changes"},
            ),
            _tool_step("submit_result"),
        ],
    )

    result = _agent(tools, max_turns=5).run("Add the requested formula.")

    assert result.observed_terminal_tool == "submit_result"
    assert [name for name, _ in tools.calls] == [
        "code_interpreter",
        "recalculate_and_read",
    ]
    assert len(requests) == 4
    events = read_trajectory(session.paths.trajectory)
    assert any(
        event["event"] == "agent.pending_formula_terminal_reprompted"
        for event in events
    )
    formula_events = [
        event for event in events if "formula" in str(event.get("event", ""))
    ]
    serialized = json.dumps(formula_events, sort_keys=True)
    assert "=1+1" not in serialized
    assert "answer_position" not in serialized


def test_clean_recalculation_before_formula_edit_does_not_satisfy_gate(
    sample_workbook: Path, tmp_path: Path, monkeypatch: Any
) -> None:
    session = _session(sample_workbook, tmp_path, "early-clean-recalc")
    tools = FormulaGateTools(session)
    _patch_sequence_client(
        monkeypatch,
        [
            _tool_step(
                "recalculate_and_read", {"sheet": "Sales", "range_ref": "H1"}
            ),
            _tool_step("code_interpreter", {"code": "set_formula"}),
            _tool_step("submit_result"),
            _tool_step(
                "recalculate_and_read",
                {"validation_scope": "pending_formula_changes"},
            ),
            _tool_step("submit_result"),
        ],
    )

    result = _agent(tools, max_turns=6).run("Add the requested formula.")

    assert result.observed_terminal_tool == "submit_result"
    assert [name for name, _ in tools.calls].count("recalculate_and_read") == 2


@pytest.mark.parametrize(
    ("range_ref", "expected_recalculation_calls"),
    [("A1", 2), ("H1", 1)],
)
def test_range_validation_clears_only_covered_dirty_formula_coordinates(
    sample_workbook: Path,
    tmp_path: Path,
    monkeypatch: Any,
    range_ref: str,
    expected_recalculation_calls: int,
) -> None:
    session = _session(sample_workbook, tmp_path, f"range-cover-{range_ref}")
    tools = FormulaGateTools(session)
    steps = [
        _tool_step("code_interpreter", {"code": "set_formula"}),
        _tool_step(
            "recalculate_and_read",
            {"sheet": "Sales", "range_ref": range_ref},
        ),
        _tool_step("submit_result"),
    ]
    if range_ref == "A1":
        steps[3:3] = [
            _tool_step(
                "recalculate_and_read",
                {"validation_scope": "pending_formula_changes"},
            ),
            _tool_step("submit_result"),
        ]
    _patch_sequence_client(monkeypatch, steps)

    result = _agent(tools, max_turns=len(steps) + 1).run("Add the formula.")

    assert result.observed_terminal_tool == "submit_result"
    assert [name for name, _ in tools.calls].count("recalculate_and_read") == (
        expected_recalculation_calls
    )


def test_sparse_runtime_validation_handles_more_than_500_dirty_formulas_once(
    sample_workbook: Path, tmp_path: Path, monkeypatch: Any
) -> None:
    session = _session(sample_workbook, tmp_path, "bulk-formulas")
    tools = FormulaGateTools(session)
    _patch_sequence_client(
        monkeypatch,
        [
            _tool_step("code_interpreter", {"code": "bulk_formulas"}),
            _tool_step(
                "recalculate_and_read",
                {"validation_scope": "pending_formula_changes"},
            ),
            _tool_step("submit_result"),
        ],
    )

    result = _agent(tools, max_turns=4).run("Fill all formulas.")

    assert result.observed_terminal_tool == "submit_result"
    nonempty_scopes = [scope for scope in tools.scope_history if scope]
    assert len(nonempty_scopes[-1]) == 601
    assert [name for name, _ in tools.calls].count("recalculate_and_read") == 1


def test_invalid_sparse_runtime_validation_remains_blocked(
    sample_workbook: Path, tmp_path: Path, monkeypatch: Any
) -> None:
    session = _session(sample_workbook, tmp_path, "invalid-sparse")
    tools = FormulaGateTools(session, sparse_valid=False)
    _patch_sequence_client(
        monkeypatch,
        [
            _tool_step("code_interpreter", {"code": "set_formula"}),
            _tool_step(
                "recalculate_and_read",
                {"validation_scope": "pending_formula_changes"},
            ),
            _tool_step("submit_result"),
        ],
    )

    with pytest.raises(AgentExecutionFailure, match="runtime validation") as caught:
        _agent(tools, max_turns=3).run("Add the formula.")

    assert caught.value.reason == "edit_recovery_exhausted"


def test_sparse_validation_rejects_unexpected_formula_disappearance(
    sample_workbook: Path, tmp_path: Path, monkeypatch: Any
) -> None:
    session = _session(sample_workbook, tmp_path, "unexpected-formula-drop")
    tools = FormulaGateTools(session, drop_formula_during_sparse=True)
    _patch_sequence_client(
        monkeypatch,
        [
            _tool_step("code_interpreter", {"code": "set_formula"}),
            _tool_step(
                "recalculate_and_read",
                {"validation_scope": "pending_formula_changes"},
            ),
            _tool_step("submit_result"),
        ],
    )

    with pytest.raises(AgentExecutionFailure, match="runtime validation"):
        _agent(tools, max_turns=3).run("Add the formula.")


def test_sparse_validation_accepts_an_intentional_formula_deletion(
    sample_workbook: Path, tmp_path: Path, monkeypatch: Any
) -> None:
    workbook = load_workbook(sample_workbook)
    workbook["Sales"]["H1"] = "=9+9"
    workbook.save(sample_workbook)
    workbook.close()
    session = _session(sample_workbook, tmp_path, "intentional-formula-delete")
    tools = FormulaGateTools(session, original_formula="=9+9")
    _patch_sequence_client(
        monkeypatch,
        [
            _tool_step("code_interpreter", {"code": "delete_formula"}),
            _tool_step(
                "recalculate_and_read",
                {"validation_scope": "pending_formula_changes"},
            ),
            _tool_step("submit_result"),
        ],
    )

    result = _agent(tools, max_turns=4).run("Delete the obsolete formula.")

    assert result.observed_terminal_tool == "submit_result"
    assert [name for name, _ in tools.calls].count("recalculate_and_read") == 1


def test_recalculation_formula_normalization_is_not_classified_as_model_edit(
    sample_workbook: Path, tmp_path: Path, monkeypatch: Any
) -> None:
    session = _session(sample_workbook, tmp_path, "formula-normalization")
    tools = FormulaGateTools(session, normalize_formula_during_recalc=True)
    _patch_sequence_client(
        monkeypatch,
        [
            _tool_step("code_interpreter", {"code": "set_value"}),
            _tool_step(
                "recalculate_and_read", {"sheet": "Sales", "range_ref": "H1"}
            ),
            _tool_step("submit_result"),
        ],
    )

    result = _agent(tools, max_turns=4).run("Write a non-formula value.")

    assert tools.normalized_formula is True
    assert result.observed_terminal_tool == "submit_result"
    assert [name for name, _ in tools.calls].count("recalculate_and_read") == 1


def test_formula_rewrite_after_clean_validation_becomes_dirty_again(
    sample_workbook: Path, tmp_path: Path, monkeypatch: Any
) -> None:
    session = _session(sample_workbook, tmp_path, "formula-rewrite")
    tools = FormulaGateTools(session)
    _patch_sequence_client(
        monkeypatch,
        [
            _tool_step("code_interpreter", {"code": "set_formula"}),
            _tool_step(
                "recalculate_and_read",
                {"validation_scope": "pending_formula_changes"},
            ),
            _tool_step("code_interpreter", {"code": "rewrite_formula"}),
            _tool_step("submit_result"),
            _tool_step(
                "recalculate_and_read",
                {"validation_scope": "pending_formula_changes"},
            ),
            _tool_step("submit_result"),
        ],
    )

    result = _agent(tools, max_turns=7).run("Add and refine the formula.")

    assert result.observed_terminal_tool == "submit_result"
    assert [name for name, _ in tools.calls].count("recalculate_and_read") == 2


def test_undo_to_exact_formula_baseline_clears_dirty_state_without_recalc(
    sample_workbook: Path, tmp_path: Path, monkeypatch: Any
) -> None:
    workbook = load_workbook(sample_workbook)
    workbook["Sales"]["H1"] = "=9+9"
    workbook.save(sample_workbook)
    workbook.close()
    session = _session(sample_workbook, tmp_path, "formula-undo")
    tools = FormulaGateTools(session, original_formula="=9+9")
    _patch_sequence_client(
        monkeypatch,
        [
            _tool_step("code_interpreter", {"code": "set_formula"}),
            _tool_step("undo_last"),
            _tool_step("submit_result"),
        ],
    )

    result = _agent(
        tools, max_turns=4, require_workbook_change=False
    ).run("Try and undo a formula edit.")

    assert result.observed_terminal_tool == "submit_result"
    assert [name for name, _ in tools.calls] == ["code_interpreter", "undo_last"]


def test_nonformula_edit_submits_without_formula_recalculation(
    sample_workbook: Path, tmp_path: Path, monkeypatch: Any
) -> None:
    session = _session(sample_workbook, tmp_path, "nonformula-edit")
    tools = FormulaGateTools(session)
    _patch_sequence_client(
        monkeypatch,
        [
            _tool_step("code_interpreter", {"code": "set_value"}),
            _tool_step("submit_result"),
        ],
    )

    result = _agent(tools, max_turns=3).run("Write a value.")

    assert result.observed_terminal_tool == "submit_result"
    assert [name for name, _ in tools.calls] == ["code_interpreter"]


def test_text_completion_is_blocked_while_formula_validation_is_pending(
    sample_workbook: Path, tmp_path: Path, monkeypatch: Any
) -> None:
    session = _session(sample_workbook, tmp_path, "formula-text")
    tools = FormulaGateTools(session)
    _patch_sequence_client(
        monkeypatch,
        [
            _tool_step("code_interpreter", {"code": "set_formula"}),
            _text_step(),
            _tool_step(
                "recalculate_and_read",
                {"validation_scope": "pending_formula_changes"},
            ),
            _tool_step("submit_result"),
        ],
    )

    result = _agent(tools, max_turns=5).run("Add the formula.")

    assert result.observed_terminal_tool == "submit_result"
    assert any(
        event["event"] == "agent.pending_formula_text_reprompted"
        for event in read_trajectory(session.paths.trajectory)
    )


@pytest.mark.parametrize("shared_budget", [False, True])
def test_late_forced_code_edit_preserves_recalc_and_submit_slots(
    sample_workbook: Path,
    tmp_path: Path,
    monkeypatch: Any,
    shared_budget: bool,
) -> None:
    session = _session(sample_workbook, tmp_path, f"late-code-{shared_budget}")
    tools = FormulaGateTools(session)
    requests = _patch_sequence_client(
        monkeypatch,
        [
            _tool_step(
                "recalculate_and_read", {"sheet": "Sales", "range_ref": "A1"}
            ),
            _tool_step(
                "recalculate_and_read", {"sheet": "Sales", "range_ref": "A1"}
            ),
            _tool_step("code_interpreter", {"code": "set_formula"}),
            _tool_step(
                "recalculate_and_read",
                {"validation_scope": "pending_formula_changes"},
            ),
            _tool_step("submit_result"),
        ],
    )
    budget = (
        RunBudget(max_model_calls=5, max_total_tokens=1_000, max_elapsed_seconds=60)
        if shared_budget
        else None
    )

    result = _agent(
        tools,
        max_turns=(10 if shared_budget else 5),
        force_code_on_stalled_edit=True,
        budget=budget,
    ).run("Make the requested edit.")

    assert result.observed_terminal_tool == "submit_result"
    forced_names = [
        request["tool_choice"].get("name")
        if isinstance(request.get("tool_choice"), dict)
        else None
        for request in requests
    ]
    assert forced_names[-3:] == [
        "code_interpreter",
        "recalculate_and_read",
        "submit_result",
    ]
