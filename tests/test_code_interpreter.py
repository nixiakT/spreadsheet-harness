from __future__ import annotations

import json
import platform
import shutil
import subprocess
import sys
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook

from spreadsheet_harness import code_interpreter
from spreadsheet_harness.agent import ResponseTurn, SpreadsheetAgent
from spreadsheet_harness.code_interpreter import LocalCodeInterpreter
from spreadsheet_harness.config import ProviderConfig
from spreadsheet_harness.errors import CodeIsolationError, ToolInputError
from spreadsheet_harness.openpyxl_compat import load_workbook as compat_load_workbook
from spreadsheet_harness.render import sheet_inventory_identity
from spreadsheet_harness.session import WorkbookSession
from spreadsheet_harness.tools import SpreadsheetToolRegistry


def test_outer_sandbox_limits_defer_host_uid_process_limit(monkeypatch: Any) -> None:
    calls: list[tuple[int, tuple[int, int]]] = []
    monkeypatch.setattr(
        code_interpreter.resource,
        "setrlimit",
        lambda kind, value: calls.append((kind, value)),
    )
    monkeypatch.setattr(code_interpreter.platform, "system", lambda: "Linux")

    code_interpreter._outer_sandbox_limits()

    assert all(
        kind != code_interpreter.resource.RLIMIT_NPROC for kind, _ in calls
    )
    calls.clear()
    code_interpreter._limits()
    assert (
        code_interpreter.resource.RLIMIT_NPROC,
        (
            code_interpreter._MAX_SANDBOX_PROCESSES,
            code_interpreter._MAX_SANDBOX_PROCESSES,
        ),
    ) in calls


def test_strict_command_mounts_only_workspace_and_runtime_allowlist(tmp_path: Path) -> None:
    workspace = (tmp_path / "runs" / "task" / "bare").resolve()
    workspace.mkdir(parents=True)
    script = workspace / "probe.py"
    script.write_text("print('ok')", encoding="utf-8")

    command = code_interpreter._strict_command(
        workspace,
        [sys.executable, "-I", str(script)],
        bubblewrap=Path("/usr/bin/bwrap"),
    )

    triples = [command[index : index + 3] for index in range(len(command) - 2)]
    assert ["--ro-bind", "/", "/"] not in triples
    assert ["--bind", str(workspace), str(workspace)] in triples
    writable_binds = [
        triple for triple in triples if triple[0] == "--bind"
    ]
    assert writable_binds == [["--bind", str(workspace), str(workspace)]]
    assert str(tmp_path / "golden.xlsx") not in command
    assert "--unshare-net" in command
    assert "--unshare-pid" in command
    assert command.index("--tmpfs") < command.index("--bind")


def test_venv_mount_is_exact_and_does_not_expose_its_repository(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    repository = tmp_path / "repository"
    venv = repository / ".venv"
    executable = venv / "bin" / "python"
    executable.parent.mkdir(parents=True)
    executable.touch()
    workspace = repository / "benchmarks" / "results" / "task" / "bare"
    workspace.mkdir(parents=True)
    script = workspace / "script.py"
    script.touch()
    monkeypatch.setattr(code_interpreter.sys, "prefix", str(venv))
    monkeypatch.setattr(code_interpreter.sys, "base_prefix", "/usr")
    monkeypatch.setattr(code_interpreter.sys, "executable", str(executable))

    command = code_interpreter._strict_command(
        workspace,
        [str(executable), "-I", str(script)],
        bubblewrap=Path("/usr/bin/bwrap"),
    )
    triples = [command[index : index + 3] for index in range(len(command) - 2)]

    assert ["--ro-bind", str(venv), str(venv)] in triples
    assert ["--ro-bind", str(repository), str(repository)] not in triples
    assert ["--bind", str(workspace), str(workspace)] in triples


def test_required_isolation_fails_when_bubblewrap_is_missing(monkeypatch: Any) -> None:
    code_interpreter._reset_isolation_probe_cache()
    monkeypatch.setattr(code_interpreter.platform, "system", lambda: "Linux")
    monkeypatch.setattr(code_interpreter.shutil, "which", lambda _: None)

    with pytest.raises(CodeIsolationError, match="bwrap"):
        code_interpreter.ensure_strict_code_isolation()


def test_required_isolation_fails_when_bubblewrap_probe_cannot_start(
    monkeypatch: Any,
) -> None:
    code_interpreter._reset_isolation_probe_cache()
    monkeypatch.setattr(code_interpreter.platform, "system", lambda: "Linux")
    monkeypatch.setattr(
        code_interpreter.shutil, "which", lambda _: "/usr/bin/bwrap"
    )

    def failed_run(*_: Any, **__: Any) -> subprocess.CompletedProcess[str]:
        return subprocess.CompletedProcess(
            ["bwrap"],
            1,
            stdout="",
            stderr="bwrap: Creating new namespace failed: Operation not permitted",
        )

    monkeypatch.setattr(code_interpreter.subprocess, "run", failed_run)

    with pytest.raises(CodeIsolationError, match="probe failed"):
        code_interpreter.ensure_strict_code_isolation()


def test_strict_preflight_redacts_configured_secret_before_diagnostic_truncation(
    monkeypatch: Any,
) -> None:
    code_interpreter._reset_isolation_probe_cache()
    secret = "key://tenant+spreadsheet?signature=" + "Q" * 256 + "&scope=%2Fall"
    diagnostic = "x" * 950 + secret + " probe failed"
    monkeypatch.setattr(code_interpreter.platform, "system", lambda: "Linux")
    monkeypatch.setattr(
        code_interpreter.shutil, "which", lambda _: "/usr/bin/bwrap"
    )
    monkeypatch.setattr(
        code_interpreter.subprocess,
        "run",
        lambda *_args, **_kwargs: subprocess.CompletedProcess(
            ["bwrap"],
            1,
            stdout="",
            stderr=diagnostic,
        ),
    )

    with pytest.raises(CodeIsolationError, match="probe failed") as caught:
        code_interpreter.ensure_strict_code_isolation((secret,))

    message = str(caught.value)
    assert secret not in message
    assert secret[:32] not in message
    assert "[REDACTED]" in message


def test_strict_execution_never_falls_back_when_launcher_did_not_start(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    workspace = tmp_path / "workspace"
    workspace.mkdir()
    workbook = workspace / "book.xlsx"
    workbook.touch()
    monkeypatch.setattr(
        code_interpreter,
        "ensure_strict_code_isolation",
        lambda *_args, **_kwargs: {},
    )
    monkeypatch.setattr(
        code_interpreter,
        "_strict_command",
        lambda *_args, **_kwargs: ["bwrap", "strict"],
    )
    interpreter = LocalCodeInterpreter(workspace, workbook, require_isolation=True)
    calls = 0

    def failed_execute(*_: Any, **__: Any) -> subprocess.CompletedProcess[str]:
        nonlocal calls
        calls += 1
        return subprocess.CompletedProcess(
            ["bwrap"],
            1,
            stdout="",
            stderr="bwrap: Permission denied",
        )

    monkeypatch.setattr(interpreter, "_execute", failed_execute)

    with pytest.raises(CodeIsolationError, match="refusing unsandboxed fallback"):
        interpreter.run("print('must not run')")
    assert calls == 1


def test_code_interpreter_preloads_openpyxl_helper_and_reports_workbook_change(
    sample_workbook: Path,
    tmp_path: Path,
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "helper-run")
    interpreter = LocalCodeInterpreter(
        session.workspace,
        session.workbook_path,
        require_isolation=False,
    )

    result = interpreter.run(
        """
from openpyxl import load_workbook

wb = sheet_harness.load_workbook()
ws = wb["Sales"]
overview = sheet_harness.workbook_overview(wb)
assert isinstance(overview, list)
assert all(isinstance(item, dict) for item in overview)
print(overview[0]["name"])
print(sheet_harness.table_refs(ws))
print(sheet_harness.defined_name_refs(wb))
ws["E1"] = "helper wrote"
saved_path = sheet_harness.save_workbook(wb)
assert saved_path == sheet_harness.workbook_path()
wb.close()
"""
    )

    assert result["ok"] is True, result
    assert result["workbook_changed"] is True
    assert result["helper_module"] == "sheet_harness.py"
    assert "Sales" in result["stdout"]


def test_code_interpreter_helper_preserves_empty_chartsheet_across_reopens(
    empty_chartsheet_workbook: Path,
    tmp_path: Path,
) -> None:
    expected_sheets = sheet_inventory_identity(empty_chartsheet_workbook)["sheets"]
    session = WorkbookSession.create(
        empty_chartsheet_workbook,
        tmp_path / "chartsheet-helper-run",
    )
    interpreter = LocalCodeInterpreter(
        session.workspace,
        session.workbook_path,
        require_isolation=False,
    )

    result = interpreter.run(
        """
wb = sheet_harness.load_workbook()
assert wb.sheetnames == ["Data", "Chart"]
assert wb["Chart"].sheet_state == "hidden"
wb["Data"]["A1"] = "first helper edit"
sheet_harness.save_workbook(wb)
wb.close()

reopened = sheet_harness.load_workbook()
assert reopened["Chart"].sheet_state == "hidden"
reopened["Data"]["A2"] = "second helper edit"
sheet_harness.save_workbook(reopened)
reopened.close()

verified = sheet_harness.load_workbook(data_only=False)
assert verified["Chart"].sheet_state == "hidden"
assert verified["Data"]["A2"].value == "second helper edit"
verified.close()
"""
    )

    assert result["ok"] is True, result
    assert result["workbook_changed"] is True
    assert sheet_inventory_identity(session.workbook_path)["sheets"] == expected_sheets
    workbook = compat_load_workbook(session.workbook_path)
    assert workbook["Data"]["A1"].value == "first helper edit"
    assert workbook["Data"]["A2"].value == "second helper edit"
    assert workbook["Chart"].sheet_state == "hidden"
    workbook.close()


def test_code_interpreter_openpyxl_compat_shim_exposes_read_only_formula_and_merge_aliases(
    sample_workbook: Path,
    tmp_path: Path,
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "compat-alias-run")
    interpreter = LocalCodeInterpreter(
        session.workspace,
        session.workbook_path,
        require_isolation=False,
    )

    result = interpreter.run(
        """
wb = sheet_harness.load_workbook()
ws = wb["Sales"]
assert ws.merged_ranges is ws.merged_cells.ranges
assert sorted(str(item) for item in ws.merged_ranges) == ["A5:B5"]
assert ws["D2"].formula == "=B2*C2"
assert ws["A2"].formula is None

for owner, attribute, replacement in (
    (ws, "merged_ranges", ()),
    (ws["D2"], "formula", "=1+1"),
):
    try:
        setattr(owner, attribute, replacement)
    except AttributeError:
        pass
    else:
        raise AssertionError(f"{attribute} compatibility alias must be read-only")

print("compat aliases ok")
wb.close()
"""
    )

    assert result["ok"] is True, result
    assert result["workbook_changed"] is False
    assert "compat aliases ok" in result["stdout"]


def test_code_interpreter_reports_failed_managed_save_attempt(
    sample_workbook: Path,
    tmp_path: Path,
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "failed-save-signal-run")
    interpreter = LocalCodeInterpreter(
        session.workspace,
        session.workbook_path,
        require_isolation=False,
    )

    result = interpreter.run(
        """
wb = sheet_harness.load_workbook()
wb.save = lambda *_args, **_kwargs: (_ for _ in ()).throw(RuntimeError("save failed"))
sheet_harness.save_workbook(wb)
"""
    )

    assert result["ok"] is False
    assert result["workbook_changed"] is False
    assert result["managed_mutation_attempted"] is True


def test_code_interpreter_failed_inspection_has_no_mutation_signal(
    sample_workbook: Path,
    tmp_path: Path,
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "failed-inspection-signal-run")
    interpreter = LocalCodeInterpreter(
        session.workspace,
        session.workbook_path,
        require_isolation=False,
    )

    result = interpreter.run("raise RuntimeError('inspection failed')")

    assert result["ok"] is False
    assert result["workbook_changed"] is False
    assert result["managed_mutation_attempted"] is False


def test_code_interpreter_starts_each_call_in_a_fresh_process(
    sample_workbook: Path,
    tmp_path: Path,
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "fresh-process-run")
    interpreter = LocalCodeInterpreter(
        session.workspace,
        session.workbook_path,
        require_isolation=False,
    )

    first = interpreter.run("transient_name = 42")
    second = interpreter.run("print(transient_name)")

    assert first["ok"] is True
    assert second["ok"] is False
    assert "NameError" in second["stderr"]


@pytest.mark.skipif(
    platform.system() != "Linux" or shutil.which("bwrap") is None,
    reason="strict Bubblewrap integration requires Linux and bwrap",
)
def test_strict_code_interpreter_starts_each_call_in_a_fresh_process(
    sample_workbook: Path,
    tmp_path: Path,
) -> None:
    code_interpreter._reset_isolation_probe_cache()
    session = WorkbookSession.create(sample_workbook, tmp_path / "strict-fresh-run")
    try:
        interpreter = LocalCodeInterpreter(
            session.workspace,
            session.workbook_path,
            require_isolation=True,
        )
    except CodeIsolationError as exc:
        pytest.skip(f"Bubblewrap installed but namespaces unavailable: {exc}")

    first = interpreter.run("transient_name = 42")
    second = interpreter.run("print(transient_name)")

    assert first["ok"] is True
    assert second["ok"] is False
    assert "NameError" in second["stderr"]
    assert first["sandbox"].startswith(code_interpreter.STRICT_ISOLATION_POLICY)


def test_code_interpreter_rolls_back_new_invalid_formula_references(
    sample_workbook: Path,
    tmp_path: Path,
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "formula-gate-run")
    interpreter = LocalCodeInterpreter(
        session.workspace,
        session.workbook_path,
        require_isolation=False,
    )

    result = interpreter.run(
        """
import sheet_harness

wb = sheet_harness.load_workbook()
ws = wb["Sales"]
ws["E2"] = "=SUMIFS(B:B,A:A,$55)"
sheet_harness.save_workbook(wb)
wb.close()
"""
    )

    assert result["ok"] is False
    assert result["workbook_changed"] is False
    assert result["workbook_rolled_back"] is True
    assert result["formula_validation"]["introduced_invalid_reference_count"] == 1
    assert result["formula_validation"]["introduced_formula_text_count"] == 0
    assert result["formula_validation"]["issues"] == [
        {
            "type": "invalid_a1_reference",
            "sheet": "Sales",
            "cell": "E2",
            "invalid_reference": "$55",
            "formula": "=SUMIFS(B:B,A:A,$55)",
        }
    ]
    assert "E$5" in result["error"]
    workbook = load_workbook(session.workbook_path, data_only=False)
    try:
        assert workbook["Sales"]["E2"].value is None
    finally:
        workbook.close()


def test_code_interpreter_preserves_preexisting_invalid_formula_reference(
    sample_workbook: Path,
    tmp_path: Path,
) -> None:
    workbook = load_workbook(sample_workbook)
    workbook["Sales"]["E2"] = "=SUMIFS(B:B,A:A,$55)"
    workbook.save(sample_workbook)
    workbook.close()
    session = WorkbookSession.create(sample_workbook, tmp_path / "formula-gate-existing-run")
    interpreter = LocalCodeInterpreter(
        session.workspace,
        session.workbook_path,
        require_isolation=False,
    )

    result = interpreter.run(
        """
import sheet_harness

wb = sheet_harness.load_workbook()
wb["Sales"]["F2"] = 42
sheet_harness.save_workbook(wb)
wb.close()
"""
    )

    assert result["ok"] is True
    assert result["workbook_changed"] is True
    workbook = load_workbook(session.workbook_path, data_only=False)
    try:
        assert workbook["Sales"]["E2"].value == "=SUMIFS(B:B,A:A,$55)"
        assert workbook["Sales"]["F2"].value == 42
    finally:
        workbook.close()


def test_code_interpreter_rolls_back_unprefixed_formula_text_atomically(
    sample_workbook: Path,
    tmp_path: Path,
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "formula-text-gate-run")
    interpreter = LocalCodeInterpreter(
        session.workspace,
        session.workbook_path,
        require_isolation=False,
    )
    before = session.workbook_path.read_bytes()

    result = interpreter.run(
        """
import sheet_harness

wb = sheet_harness.load_workbook()
ws = wb["Sales"]
ws["H6"] = "AVERAGE($B2:$D2)+$F$1"
ws["I6"] = 99
sheet_harness.save_workbook(wb)
wb.close()
"""
    )

    assert result["ok"] is False
    assert result["workbook_changed"] is False
    assert result["workbook_rolled_back"] is True
    assert result["workbook_sha256_before"] == result["workbook_sha256_after"]
    assert result["workbook_sha256_rejected"] != result["workbook_sha256_after"]
    assert result["formula_validation"] == {
        "ok": False,
        "issue_count": 1,
        "introduced_invalid_reference_count": 0,
        "introduced_formula_text_count": 1,
        "issues": [
            {
                "type": "missing_formula_prefix",
                "sheet": "Sales",
                "cell": "H6",
                "value": "AVERAGE($B2:$D2)+$F$1",
            }
        ],
        "truncated": False,
    }
    assert "leading '='" in result["error"]
    assert session.workbook_path.read_bytes() == before
    workbook = load_workbook(session.workbook_path, data_only=False)
    try:
        assert workbook["Sales"]["H6"].value is None
        assert workbook["Sales"]["I6"].value is None
    finally:
        workbook.close()


def test_code_interpreter_rolls_back_formula_text_after_script_failure(
    sample_workbook: Path,
    tmp_path: Path,
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "failed-formula-text-run")
    interpreter = LocalCodeInterpreter(
        session.workspace,
        session.workbook_path,
        require_isolation=False,
    )
    before = session.workbook_path.read_bytes()

    result = interpreter.run(
        """
import sheet_harness

wb = sheet_harness.load_workbook()
wb["Sales"]["H6"] = "AVERAGE($B2:$D2)+$F$1"
sheet_harness.save_workbook(wb)
wb.close()
raise RuntimeError("verification failed after save")
"""
    )

    assert result["ok"] is False
    assert result["exit_code"] != 0
    assert result["workbook_changed"] is False
    assert result["workbook_rolled_back"] is True
    assert "all partial edits were rolled back" in result["error"]
    assert "verification failed after save" in result["stderr"]
    assert session.workbook_path.read_bytes() == before


def test_code_interpreter_rolls_back_any_partial_edit_after_script_failure(
    sample_workbook: Path,
    tmp_path: Path,
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "failed-partial-edit-run")
    interpreter = LocalCodeInterpreter(
        session.workspace,
        session.workbook_path,
        require_isolation=False,
    )
    before = session.workbook_path.read_bytes()

    result = interpreter.run(
        """
import sheet_harness

wb = sheet_harness.load_workbook()
wb["Sales"]["A1"] = "partial edit"
sheet_harness.save_workbook(wb)
wb.close()
raise RuntimeError("failure after save")
"""
    )

    assert result["ok"] is False
    assert result["workbook_changed"] is False
    assert result["workbook_rolled_back"] is True
    assert "all partial edits were rolled back" in result["error"]
    assert session.workbook_path.read_bytes() == before


def test_code_interpreter_rolls_back_partial_workbook_after_timeout(
    sample_workbook: Path,
    tmp_path: Path,
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "timed-out-workbook-run")
    interpreter = LocalCodeInterpreter(
        session.workspace,
        session.workbook_path,
        require_isolation=False,
    )
    before = session.workbook_path.read_bytes()

    result = interpreter.run(
        """
import time
import sheet_harness

wb = sheet_harness.load_workbook()
wb["Sales"]["H6"] = "partial edit"
sheet_harness.save_workbook(wb)
wb.close()
time.sleep(5)
""",
        timeout_seconds=1,
    )

    assert result["ok"] is False
    assert result["workbook_changed"] is False
    assert result["workbook_rolled_back"] is True
    assert result["workbook_sha256_before"] == result["workbook_sha256_after"]
    assert result["workbook_sha256_rejected"] != result["workbook_sha256_after"]
    assert "partial workbook edits were rolled back" in result["error"]
    assert session.workbook_path.read_bytes() == before


@pytest.mark.parametrize("termination", ["return", "timeout"])
def test_code_interpreter_restores_deleted_workbook(
    sample_workbook: Path,
    tmp_path: Path,
    termination: str,
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "deleted-workbook-run")
    interpreter = LocalCodeInterpreter(
        session.workspace,
        session.workbook_path,
        require_isolation=False,
    )
    before = session.workbook_path.read_bytes()
    suffix = "\nimport time; time.sleep(5)" if termination == "timeout" else ""

    result = interpreter.run(
        f"""
import sheet_harness

sheet_harness.workbook_path().unlink()
{suffix}
""",
        timeout_seconds=1,
    )

    assert result["ok"] is False
    assert result["workbook_changed"] is False
    assert result["workbook_rolled_back"] is True
    assert session.workbook_path.read_bytes() == before


def test_code_interpreter_accepts_prefixed_formula(
    sample_workbook: Path,
    tmp_path: Path,
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "valid-formula-run")
    interpreter = LocalCodeInterpreter(
        session.workspace,
        session.workbook_path,
        require_isolation=False,
    )

    result = interpreter.run(
        """
import sheet_harness

wb = sheet_harness.load_workbook()
wb["Sales"]["H6"] = "=AVERAGE($B2:$D2)+$F$1"
sheet_harness.save_workbook(wb)
wb.close()
"""
    )

    assert result["ok"] is True, result
    workbook = load_workbook(session.workbook_path, data_only=False)
    try:
        assert workbook["Sales"]["H6"].value == "=AVERAGE($B2:$D2)+$F$1"
        assert workbook["Sales"]["H6"].data_type == "f"
    finally:
        workbook.close()


def test_code_interpreter_preserves_preexisting_formula_text(
    sample_workbook: Path,
    tmp_path: Path,
) -> None:
    workbook = load_workbook(sample_workbook)
    workbook["Sales"]["H6"] = "AVERAGE($B2:$D2)+$F$1"
    workbook.save(sample_workbook)
    workbook.close()
    session = WorkbookSession.create(sample_workbook, tmp_path / "formula-text-existing-run")
    interpreter = LocalCodeInterpreter(
        session.workspace,
        session.workbook_path,
        require_isolation=False,
    )

    result = interpreter.run(
        """
import sheet_harness

wb = sheet_harness.load_workbook()
wb["Sales"]["F2"] = 42
sheet_harness.save_workbook(wb)
wb.close()
"""
    )

    assert result["ok"] is True, result
    workbook = load_workbook(session.workbook_path, data_only=False)
    try:
        assert workbook["Sales"]["H6"].value == "AVERAGE($B2:$D2)+$F$1"
        assert workbook["Sales"]["F2"].value == 42
    finally:
        workbook.close()


@pytest.mark.parametrize(
    "value",
    [
        "SUM(A1:A3)",
        "SUM of actuals",
        "SUM (Actuals by month)",
        "Forecast = SUM(A1:A3)",
        "SUM(A1:A3) formula",
        "SUM(Actuals)",
        "TODAY()",
        "UNKNOWNMODERN(A1:A3)",
        "SUM(Table1[Amount])",
        "SUM(A1,Rate)",
        " SUM(A1:A3,B1:B3)",
        "SUM(A1:A3,B1:B3) ",
        "SUM(A1)/",
        "SUM(A1)+(B1",
    ],
)
def test_code_interpreter_accepts_weak_or_non_formula_text(
    sample_workbook: Path,
    tmp_path: Path,
    value: str,
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "formula-text-negative-run")
    interpreter = LocalCodeInterpreter(
        session.workspace,
        session.workbook_path,
        require_isolation=False,
    )

    result = interpreter.run(
        f"""
import sheet_harness

wb = sheet_harness.load_workbook()
wb["Sales"]["H6"] = {value!r}
sheet_harness.save_workbook(wb)
wb.close()
"""
    )

    assert result["ok"] is True, result
    workbook = load_workbook(session.workbook_path, data_only=False)
    try:
        assert workbook["Sales"]["H6"].value == value
    finally:
        workbook.close()


@pytest.mark.parametrize("explicit_text", ["apostrophe", "quote_prefix", "text_format"])
def test_code_interpreter_accepts_explicit_formula_text(
    sample_workbook: Path,
    tmp_path: Path,
    explicit_text: str,
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "explicit-formula-text-run")
    interpreter = LocalCodeInterpreter(
        session.workspace,
        session.workbook_path,
        require_isolation=False,
    )
    assignment = {
        "apostrophe": 'cell.value = "\'SUM(A1:A3,B1:B3)"',
        "quote_prefix": (
            'cell.value = "SUM(A1:A3,B1:B3)"\ncell.quotePrefix = True'
        ),
        "text_format": (
            'cell.value = "SUM(A1:A3,B1:B3)"\ncell.number_format = "@"'
        ),
    }[explicit_text]

    result = interpreter.run(
        f"""
import sheet_harness

wb = sheet_harness.load_workbook()
cell = wb["Sales"]["H6"]
{assignment}
sheet_harness.save_workbook(wb)
wb.close()
"""
    )

    assert result["ok"] is True, result


def test_code_interpreter_rolls_back_adjacent_same_shape_formula_text(
    sample_workbook: Path,
    tmp_path: Path,
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "formula-text-batch-run")
    interpreter = LocalCodeInterpreter(
        session.workspace,
        session.workbook_path,
        require_isolation=False,
    )

    result = interpreter.run(
        """
import sheet_harness

wb = sheet_harness.load_workbook()
ws = wb["Sales"]
ws["H6"] = "SUM(A1:A3)"
ws["I6"] = "SUM(B1:B3)"
ws["J6"] = "SUM(C1:C3)"
sheet_harness.save_workbook(wb)
wb.close()
"""
    )

    assert result["ok"] is False
    assert result["formula_validation"]["introduced_formula_text_count"] == 3
    assert all(
        issue["type"] == "missing_formula_prefix"
        for issue in result["formula_validation"]["issues"]
    )


def test_code_interpreter_rolls_back_valid_conditional_formula_text(
    sample_workbook: Path,
    tmp_path: Path,
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "formula-text-conditional-run")
    interpreter = LocalCodeInterpreter(
        session.workspace,
        session.workbook_path,
        require_isolation=False,
    )

    result = interpreter.run(
        """
import sheet_harness

wb = sheet_harness.load_workbook()
wb["Sales"]["H6"] = 'IF(A1="",,B1)'
sheet_harness.save_workbook(wb)
wb.close()
"""
    )

    assert result["ok"] is False
    assert result["formula_validation"]["introduced_formula_text_count"] == 1
    assert result["formula_validation"]["issues"][0]["type"] == (
        "missing_formula_prefix"
    )


def test_code_interpreter_reports_all_formula_issues_in_one_rollback(
    sample_workbook: Path,
    tmp_path: Path,
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "combined-formula-gate-run")
    interpreter = LocalCodeInterpreter(
        session.workspace,
        session.workbook_path,
        require_isolation=False,
    )

    result = interpreter.run(
        """
import sheet_harness

wb = sheet_harness.load_workbook()
ws = wb["Sales"]
ws["E2"] = "=SUMIFS(B:B,A:A,$55)"
ws["H6"] = "AVERAGE($B2:$D2)+$F$1"
sheet_harness.save_workbook(wb)
wb.close()
"""
    )

    assert result["ok"] is False
    assert result["workbook_rolled_back"] is True
    assert result["formula_validation"]["issue_count"] == 2
    assert result["formula_validation"]["introduced_invalid_reference_count"] == 1
    assert result["formula_validation"]["introduced_formula_text_count"] == 1
    assert [issue["type"] for issue in result["formula_validation"]["issues"]] == [
        "invalid_a1_reference",
        "missing_formula_prefix",
    ]


def test_code_interpreter_helper_fills_formula(
    sample_workbook: Path,
    tmp_path: Path,
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "fill-helper-run")
    interpreter = LocalCodeInterpreter(
        session.workspace,
        session.workbook_path,
        require_isolation=False,
    )

    result = interpreter.run(
        """
import sheet_harness

wb = sheet_harness.load_workbook()
ws = wb["Sales"]
ws["H6"] = "=SUM($E6:$G6)"
result = sheet_harness.fill_formula(ws, "H6", "H6:J7")
sheet_harness.save_workbook(wb)
wb.close()
print(result)
"""
    )

    assert result["ok"] is True, result
    assert result["workbook_changed"] is True
    assert "'cells_filled': 6" in result["stdout"]
    assert "'warnings': []" in result["stdout"]
    workbook = load_workbook(session.workbook_path, data_only=False)
    try:
        assert workbook["Sales"]["I6"].value == "=SUM($E6:$G6)"
        assert workbook["Sales"]["H7"].value == "=SUM($E7:$G7)"
    finally:
        workbook.close()


@pytest.mark.parametrize(
    ("source_value", "expected_error"),
    [
        (None, "does not contain a formula"),
        (
            "SUM(A1:A3)",
            "formula-like text without a leading '='; assign an Excel formula string",
        ),
    ],
)
def test_code_interpreter_helper_explains_invalid_formula_source(
    sample_workbook: Path,
    tmp_path: Path,
    source_value: str | None,
    expected_error: str,
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "fill-helper-error-run")
    interpreter = LocalCodeInterpreter(
        session.workspace,
        session.workbook_path,
        require_isolation=False,
    )

    result = interpreter.run(
        f"""
import sheet_harness

wb = sheet_harness.load_workbook()
ws = wb["Sales"]
ws["H6"] = {source_value!r}
sheet_harness.fill_formula(ws, "H6", "H6:J6")
"""
    )

    assert result["ok"] is False
    assert expected_error in result["stderr"]


def test_code_interpreter_helper_warns_on_drifting_formula_fill(
    sample_workbook: Path,
    tmp_path: Path,
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "fill-helper-warning-run")
    interpreter = LocalCodeInterpreter(
        session.workspace,
        session.workbook_path,
        require_isolation=False,
    )

    result = interpreter.run(
        """
import sheet_harness

wb = sheet_harness.load_workbook()
ws = wb["Sales"]
ws["H6"] = "=SUM($E6:G6)"
print(sheet_harness.fill_formula(ws, "H6", "H6:J7"))
wb.close()
"""
    )

    assert result["ok"] is True, result
    assert "'possible_expanding_or_drifting_range'" in result["stdout"]
    assert "'translated_range': '$E6:H6'" in result["stdout"]


def test_code_interpreter_helper_expands_endpoint_and_warns_on_relative_drift(
    sample_workbook: Path,
    tmp_path: Path,
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "fill-helper-endpoint-run")
    interpreter = LocalCodeInterpreter(
        session.workspace,
        session.workbook_path,
        require_isolation=False,
    )

    result = interpreter.run(
        """
import sheet_harness

wb = sheet_harness.load_workbook()
ws = wb["Sales"]
ws["H6"] = "=SUM(E6:G6)"
print(sheet_harness.fill_formula(ws, "H6", "J6"))
sheet_harness.save_workbook(wb)
wb.close()
"""
    )

    assert result["ok"] is True, result
    assert result["workbook_changed"] is True
    assert "'range': 'H6:J6'" in result["stdout"]
    assert "'target_range_expanded_from_endpoint': True" in result["stdout"]
    assert "'translated_range': 'F6:H6'" in result["stdout"]


def test_code_interpreter_rejects_compressed_code_placeholder(
    sample_workbook: Path,
    tmp_path: Path,
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "placeholder-run")
    interpreter = LocalCodeInterpreter(
        session.workspace,
        session.workbook_path,
        require_isolation=False,
    )

    with pytest.raises(ToolInputError, match="complete runnable Python"):
        interpreter.run("print('start')\n# ...[compressed]\n")


def test_code_interpreter_openpyxl_compat_shim_supports_legacy_table_access(
    tmp_path: Path,
) -> None:
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo

    workspace = tmp_path / "workspace"
    workspace.mkdir()
    workbook_path = workspace / "table.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["A", "B"])
    sheet.append([1, 2])
    table = Table(displayName="DataTable", ref="A1:B2")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9")
    sheet.add_table(table)
    workbook.save(workbook_path)
    workbook.close()

    interpreter = LocalCodeInterpreter(workspace, workbook_path, require_isolation=False)
    result = interpreter.run(
        """
from openpyxl import load_workbook
wb = load_workbook(sheet_harness.workbook_path())
ws = wb["Data"]
print([table.name for table in ws.tables])
print([table.ref for table in ws._tableparts])
print([(name, table.ref) for name, table in ws.tables.items()])
wb.close()
"""
    )

    assert result["ok"] is True, result
    assert "DataTable" in result["stdout"]
    assert "A1:B2" in result["stdout"]


def test_code_interpreter_helper_accepts_path_and_cell_dtype_alias(
    sample_workbook: Path,
    tmp_path: Path,
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "path-helper-run")
    interpreter = LocalCodeInterpreter(
        session.workspace,
        session.workbook_path,
        require_isolation=False,
    )

    result = interpreter.run(
        """
import os
from openpyxl import load_workbook

path = os.environ["SHEET_WORKBOOK"]
print(sheet_harness.workbook_overview(path)[0]["name"])
wb = load_workbook(path)
cell = wb["Sales"]["D2"]
print(cell.dtype)
wb.close()
"""
    )

    assert result["ok"] is True, result
    assert "Sales" in result["stdout"]
    assert "formula" in result["stdout"]


def test_tool_registry_propagates_required_isolation_failure(
    sample_workbook: Path,
    tmp_path: Path,
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "tool-run")
    tools = SpreadsheetToolRegistry(session, allowed_tools={"code_interpreter"})

    class FailingInterpreter:
        def run(self, *_: Any, **__: Any) -> dict[str, Any]:
            raise CodeIsolationError("sandbox disappeared")

    tools.interpreter = FailingInterpreter()  # type: ignore[assignment]

    with pytest.raises(CodeIsolationError, match="sandbox disappeared"):
        tools.invoke("code_interpreter", {"code": "print('x')"})


@pytest.mark.skipif(
    platform.system() != "Linux" or shutil.which("bwrap") is None,
    reason="strict Bubblewrap integration requires Linux and bwrap",
)
def test_strict_integration_imports_openpyxl_and_hides_sibling(tmp_path: Path) -> None:
    code_interpreter._reset_isolation_probe_cache()
    workspace = tmp_path / "workspace"
    workspace.mkdir()
    workbook = workspace / "book.xlsx"
    outside = tmp_path / "golden.xlsx"
    outside.write_text("must remain hidden", encoding="utf-8")
    try:
        interpreter = LocalCodeInterpreter(
            workspace,
            workbook,
            require_isolation=True,
        )
    except CodeIsolationError as exc:
        pytest.skip(f"Bubblewrap installed but namespaces unavailable: {exc}")

    result = interpreter.run(
        f"""from pathlib import Path
import resource
from openpyxl import Workbook
assert resource.getrlimit(resource.RLIMIT_NPROC) == (64, 64)
try:
    Path({str(outside)!r}).read_text()
except (FileNotFoundError, PermissionError):
    pass
else:
    raise RuntimeError("sibling was readable")
book = Workbook()
book.active["A1"] = "isolated"
book.save(Path({str(workbook)!r}))
book.close()
print("openpyxl-isolated-ok")
"""
    )

    assert result["ok"] is True, result
    assert result["stdout"].strip() == "openpyxl-isolated-ok"
    assert result["sandbox"].startswith(code_interpreter.STRICT_ISOLATION_POLICY)
    assert workbook.is_file()


@pytest.mark.parametrize(
    ("stream", "code_template"),
    [
        ("stdout", "print('x' * 19900 + {secret!r}, end='')"),
        (
            "stderr",
            "import sys\nsys.stderr.write('x' * 19900 + {secret!r})",
        ),
    ],
)
def test_code_interpreter_redacts_secret_before_output_truncation(
    sample_workbook: Path,
    tmp_path: Path,
    stream: str,
    code_template: str,
) -> None:
    secret = "credential-" + "Q" * 256
    session = WorkbookSession.create(sample_workbook, tmp_path / f"{stream}-redaction")
    interpreter = LocalCodeInterpreter(
        session.workspace,
        session.workbook_path,
        max_output_chars=20_000,
        require_isolation=False,
        secrets=(secret,),
    )

    result = interpreter.run(code_template.format(secret=secret))

    assert result["ok"] is True, result
    output = result[stream]
    assert secret not in output
    assert secret[:100] not in output
    assert output.endswith("[REDACTED]")
    assert len(output) <= 20_000


def test_code_interpreter_redacts_timeout_output_before_truncation(
    sample_workbook: Path,
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    secret = "credential-" + "T" * 256
    exposed = "x" * 19_900 + secret
    session = WorkbookSession.create(sample_workbook, tmp_path / "timeout-redaction")
    interpreter = LocalCodeInterpreter(
        session.workspace,
        session.workbook_path,
        max_output_chars=20_000,
        require_isolation=False,
        secrets=(secret,),
    )

    def time_out(
        command: list[str],
        *,
        environment: dict[str, str],
        timeout: int,
    ) -> subprocess.CompletedProcess[str]:
        del environment
        raise subprocess.TimeoutExpired(
            command,
            timeout,
            output=exposed,
            stderr=exposed,
        )

    monkeypatch.setattr(interpreter, "_execute", time_out)

    result = interpreter.run("print('never completes')")

    assert result["ok"] is False
    for stream in ("stdout", "stderr"):
        assert secret not in result[stream]
        assert secret[:100] not in result[stream]
        assert result[stream].endswith("[REDACTED]")
        assert len(result[stream]) <= 20_000


def test_code_interpreter_redacts_bubblewrap_diagnostic_before_truncation(
    sample_workbook: Path,
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    secret = "credential-" + "B" * 256
    diagnostic = "x" * 900 + secret + " permission denied"
    session = WorkbookSession.create(sample_workbook, tmp_path / "bubblewrap-redaction")
    interpreter = LocalCodeInterpreter(
        session.workspace,
        session.workbook_path,
        require_isolation=False,
        secrets=(secret,),
    )
    monkeypatch.setattr(
        interpreter,
        "_command",
        lambda *_args, **_kwargs: (["bwrap", "probe"], "bubblewrap: test"),
    )
    attempts = iter(
        [
            subprocess.CompletedProcess(
                ["bwrap", "probe"],
                1,
                stdout="",
                stderr=diagnostic,
            ),
            subprocess.CompletedProcess(
                [sys.executable],
                0,
                stdout="fallback completed",
                stderr="",
            ),
        ]
    )
    monkeypatch.setattr(interpreter, "_execute", lambda *_args, **_kwargs: next(attempts))

    result = interpreter.run("print('fallback')")

    assert result["ok"] is True, result
    assert secret not in result["bubblewrap_error"]
    assert secret[:100] not in result["bubblewrap_error"]
    assert "[REDACTED]" in result["bubblewrap_error"]


def test_strict_code_interpreter_redacts_launcher_diagnostic_before_truncation(
    sample_workbook: Path,
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    secret = "credential-" + "D" * 256
    diagnostic = "x" * 900 + secret + " permission denied"
    session = WorkbookSession.create(sample_workbook, tmp_path / "strict-redaction")
    monkeypatch.setattr(
        code_interpreter,
        "ensure_strict_code_isolation",
        lambda *_args, **_kwargs: {},
    )
    monkeypatch.setattr(
        code_interpreter,
        "_strict_command",
        lambda *_args, **_kwargs: ["bwrap", "strict"],
    )
    interpreter = LocalCodeInterpreter(
        session.workspace,
        session.workbook_path,
        require_isolation=True,
        secrets=(secret,),
    )
    monkeypatch.setattr(
        interpreter,
        "_execute",
        lambda *_args, **_kwargs: subprocess.CompletedProcess(
            ["bwrap", "strict"],
            1,
            stdout="",
            stderr=diagnostic,
        ),
    )

    with pytest.raises(CodeIsolationError) as caught:
        interpreter.run("print('must not run')")

    message = str(caught.value)
    assert secret not in message
    assert secret[:100] not in message
    assert "[REDACTED]" in message


def test_registry_redaction_reaches_trajectory_and_next_model_turn(
    sample_workbook: Path,
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    secret = "credential-" + "R" * 256
    exposed = "x" * 19_900 + secret
    session = WorkbookSession.create(sample_workbook, tmp_path / "registry-redaction")
    tools = SpreadsheetToolRegistry(
        session,
        enable_code=True,
        allowed_tools={"code_interpreter"},
        redaction_secrets=(secret,),
    )
    assert tools.interpreter is not None
    monkeypatch.setattr(
        tools.interpreter,
        "_execute",
        lambda *_args, **_kwargs: subprocess.CompletedProcess(
            [sys.executable],
            0,
            stdout=exposed,
            stderr="",
        ),
    )

    class BoundaryClient:
        requests: list[dict[str, Any]] = []

        def __init__(self, _: ProviderConfig) -> None:
            self.turn = 0

        def __enter__(self) -> BoundaryClient:
            return self

        def __exit__(self, *_: Any) -> None:
            return None

        def create(self, payload: dict[str, Any], **_: Any) -> ResponseTurn:
            self.requests.append(payload)
            self.turn += 1
            if self.turn == 1:
                return ResponseTurn(
                    "response-1",
                    [
                        {
                            "type": "function_call",
                            "call_id": "call-1",
                            "name": "code_interpreter",
                            "arguments": json.dumps({"code": "print('probe')"}),
                        }
                    ],
                    "",
                    {},
                )
            return ResponseTurn(
                "response-2",
                [
                    {
                        "type": "message",
                        "role": "assistant",
                        "content": [{"type": "output_text", "text": "Done"}],
                    }
                ],
                "Done",
                {},
            )

    BoundaryClient.requests = []
    monkeypatch.setattr("spreadsheet_harness.agent.ResponsesClient", BoundaryClient)
    result = SpreadsheetAgent(
        ProviderConfig("https://example.test/v1", secret, "test-model"),
        tools,
        first_tool_choice="code_interpreter",
        max_turns=2,
    ).run("Inspect the workbook")

    assert result.final_text == "Done"
    assert len(BoundaryClient.requests) == 2
    second_payload = json.dumps(BoundaryClient.requests[1], ensure_ascii=False)
    trajectory = session.paths.trajectory.read_text(encoding="utf-8")
    for rendered in (second_payload, trajectory):
        assert secret not in rendered
        assert secret[:100] not in rendered
        assert "[REDACTED]" in rendered
