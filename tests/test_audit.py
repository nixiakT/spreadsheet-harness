from __future__ import annotations

import hashlib
import json
import shutil
import warnings
import zipfile
from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook, load_workbook

import spreadsheet_harness.audit as audit_module
import spreadsheet_harness.render as render_module
from spreadsheet_harness import benchmark as benchmark_module
from spreadsheet_harness.arms import (
    COMPARISON_FORCED_TOOL_PREFIX_POLICY,
    comparison_stage_turn_caps,
)
from spreadsheet_harness.audit import _valid_sheet_inventory_identity, audit_comparison
from spreadsheet_harness.benchmark import SpreadsheetTask, compare_workbooks
from spreadsheet_harness.comparison import (
    COMPARISON_CONFIGURATION_POLICIES,
    COMPARISON_PROTOCOL_VERSION,
    CONTINUATION_SOURCE_FILENAME,
    INFLIGHT_FILENAME,
    INTERRUPTED_SEALS_FILENAME,
    V24_COMPARISON_CONFIGURATION_POLICIES,
    V24_COMPARISON_MANIFEST_SCHEMA_VERSION,
    V24_COMPARISON_PROTOCOL_VERSION,
    V25_COMPARISON_CONFIGURATION_POLICIES,
    V25_COMPARISON_MANIFEST_SCHEMA_VERSION,
    V25_COMPARISON_PROTOCOL_VERSION,
    V26_COMPARISON_CONFIGURATION_POLICIES,
    V26_COMPARISON_MANIFEST_SCHEMA_VERSION,
    V26_COMPARISON_PROTOCOL_VERSION,
    V27_COMPARISON_CONFIGURATION_POLICIES,
    V27_COMPARISON_MANIFEST_SCHEMA_VERSION,
    V27_COMPARISON_PROTOCOL_VERSION,
    ComparisonBenchmarkRunner,
    _allowed_observed_terminals_policy,
    _stage_allowed_tools_policy,
)
from spreadsheet_harness.config import ProviderConfig
from spreadsheet_harness.errors import (
    RecalculationIntegrityError,
    ScoringInfrastructureError,
)
from spreadsheet_harness.render import recalculate_workbook
from spreadsheet_harness.skills import SkillRegistry


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _text_sha256(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def _book(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.active.title = "Sheet1"
    workbook.active["A1"] = value
    workbook.save(path)
    workbook.close()


def _scoring_metadata_sha256(task: SpreadsheetTask) -> str:
    encoded = json.dumps(
        {
            "answer_position": task.answer_position,
            "answer_sheet": task.answer_sheet,
        },
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )
    return _text_sha256(encoded)


def _fixture(
    tmp_path: Path, *, arm: str = "bare"
) -> tuple[Path, SpreadsheetTask, dict[str, Any]]:
    dataset = tmp_path / "dataset"
    initial = dataset / "initial.xlsx"
    golden = dataset / "golden.xlsx"
    _book(initial, 0)
    _book(golden, 42)
    task = SpreadsheetTask(
        task_id="task-1",
        instruction="Set A1 to 42.",
        input_path=initial,
        golden_path=golden,
        instruction_type="Cell-Level Manipulation",
        answer_position="A1",
        answer_sheet="Sheet1",
    )

    results = tmp_path / "comparison"
    runner = ComparisonBenchmarkRunner(
        ProviderConfig(
            "https://example.test/v1",
            "not-a-real-key",
            "test-model",
            reasoning_effort="none",
        ),
        results,
        skill_registry=SkillRegistry([]),
        arms=(arm,),
        max_model_calls=3,
        max_turns_per_arm=3,
        max_total_tokens=100,
        max_output_tokens=64,
        task_timeout_seconds=30,
        recalculate=False,
    )
    runner._prepare_manifest([task])
    manifest = json.loads(runner.manifest_path.read_text(encoding="utf-8"))
    manifest_sha256 = _sha256(runner.manifest_path)
    run_dir = results / "runs" / task.task_id / arm
    output = run_dir / "artifacts" / "output.xlsx"
    output.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(task.golden_path, output)
    comparison = compare_workbooks(
        task.golden_path,
        output,
        task.answer_position,
        answer_sheet=task.answer_sheet,
    )
    attempt = {"api_protocol": "responses", "endpoint": "/responses"}
    timings = [
        {
            "turn": turn,
            "stage": "solve",
            "attempts": 1,
            "attempt_history": [dict(attempt)],
            "input_tokens": 2 if turn < 3 else 4,
            "output_tokens": 0 if turn < 3 else 2,
            "total_tokens": 2 if turn < 3 else 6,
        }
        for turn in range(1, 4)
    ]
    stage_budget = {
        "limit": {
            "model_calls": 3,
            "total_tokens": 100,
            "elapsed_seconds": 30,
        },
        "used": {
            "model_calls": 3,
            "total_tokens": 10,
            "elapsed_seconds": 0.9,
        },
        "termination": None,
    }
    tool_trace = [
        {"name": "code_interpreter", "ok": True},
        {"name": "code_interpreter", "ok": True},
    ]
    terminal_response = {
        "status": "accepted",
        "response_id": "response-final",
        "acknowledgement": {},
    }
    stage_agent = {
        "final_text": "Spreadsheet task completed.",
        "response_id": "response-final",
        "turns": 3,
        "tool_calls": 2,
        "usage": {"input_tokens": 8, "output_tokens": 2, "total_tokens": 10},
        "request_timings": json.loads(json.dumps(timings)),
        "tool_trace": json.loads(json.dumps(tool_trace)),
        "terminal_submissions": 1,
        "function_calls_total": 3,
        "budget": stage_budget,
        "post_prefix_tool_choice": "auto",
        "terminal_tool": "submit_result",
        "observed_terminal_tool": "submit_result",
        "terminal_response": terminal_response,
    }
    stage = {
        "name": "solve",
        "max_turns": 3,
        "allowed_tools": manifest["stage_allowed_tools"][arm]["solve"],
        "first_tool_choice": "code_interpreter",
        "observed_first_tool": "code_interpreter",
        "forced_tool_prefix": ["code_interpreter", "code_interpreter"],
        "observed_forced_tool_prefix": ["code_interpreter", "code_interpreter"],
        "post_prefix_tool_choice": "auto",
        "terminal_tool": "submit_result",
        "observed_terminal_tool": "submit_result",
        "tool_name_trace": ["code_interpreter", "code_interpreter"],
        "tool_trace": tool_trace,
        "agent": stage_agent,
    }
    row = {
        "task_id": task.task_id,
        "arm": arm,
        "comparison_protocol_version": COMPARISON_PROTOCOL_VERSION,
        "comparison_manifest_sha256": manifest_sha256,
        "instruction_type": task.instruction_type,
        "model": "test-model",
        "api_protocol": "responses",
        "requested_reasoning_effort": "none",
        "reasoning_effort": "none",
        "request_interval_seconds": 0.0,
        "litellm_timeout_seconds": None,
        "generation": manifest["configuration"]["generation"],
        "max_model_calls": 3,
        "max_turns_per_arm": 3,
        "stage_turn_caps": {"solve": 3},
        "calculation_backend": "not_recalculated",
        "status": "completed",
        "outcome_kind": "scored",
        "passed": comparison.passed,
        "artifact_score_passed": comparison.passed,
        "comparison": comparison.to_dict(),
        "agent": {
            "arm": arm,
            "final_text": "Spreadsheet task completed.",
            "response_id": "response-final",
            "turns": 3,
            "tool_calls": 2,
            "usage": {"input_tokens": 8, "output_tokens": 2, "total_tokens": 10},
            "request_timings": json.loads(json.dumps(timings)),
            "tool_trace": [
                {"stage": "solve", **item} for item in tool_trace
            ],
            "terminal_submissions": 1,
            "function_calls_total": 3,
            "post_prefix_tool_choice": "auto",
            "terminal_tool": "submit_result",
            "observed_terminal_tool": "submit_result",
            "terminal_response": json.loads(json.dumps(terminal_response)),
            "stages": [stage],
            "budget": stage_budget,
        },
        "budget": {
            "limit": {
                "model_calls": 3,
                "total_tokens": 100,
                "elapsed_seconds": 30,
            },
            "used": {"model_calls": 3, "total_tokens": 10, "elapsed_seconds": 1.0},
            "termination": None,
        },
        "run_dir": str(run_dir),
        "output_workbook": str(output),
        "output_sha256": _sha256(output),
    }
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")
    return results, task, row


def _set_final_model_execution_failure(row: dict[str, Any], message: str) -> None:
    """Replace a successful final acknowledgement with failure-path evidence."""

    final_agent = row["agent"]["stages"][-1]["agent"]
    for result in (final_agent, row["agent"]):
        result["final_text"] = message
        result.pop("terminal_response", None)


def _set_recalculation_manifest(
    results: Path,
    row: dict[str, Any],
) -> None:
    manifest_path = results / "comparison-manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest["configuration"]["recalculate"] = True
    manifest_path.write_text(json.dumps(manifest) + "\n", encoding="utf-8")
    row["comparison_manifest_sha256"] = _sha256(manifest_path)
    row["calculation_backend"] = "libreoffice"


def _set_agent_tool_recalculation_failure(
    row: dict[str, Any],
    recalculation: dict[str, Any],
) -> None:
    failure_call = {
        "name": "recalculate_and_read",
        "ok": False,
        "error_type": "RecalculationIntegrityError",
        "failure_category": "recalculation_infrastructure",
    }
    stage = row["agent"]["stages"][-1]
    stage_agent = stage["agent"]
    stage_trace = [*stage["tool_trace"], failure_call]
    stage.update(
        {
            "observed_terminal_tool": None,
            "tool_name_trace": [item["name"] for item in stage_trace],
            "tool_trace": stage_trace,
        }
    )
    stage_agent.update(
        {
            "final_text": (
                "Agent interrupted by recalculation infrastructure failure."
            ),
            "tool_calls": 3,
            "tool_trace": json.loads(json.dumps(stage_trace)),
            "terminal_submissions": 0,
            "function_calls_total": 3,
            "observed_terminal_tool": None,
        }
    )
    stage_agent.pop("terminal_response")
    row["agent"].update(
        {
            "final_text": stage_agent["final_text"],
            "tool_calls": 3,
            "tool_trace": [
                {"stage": "solve", **item} for item in stage_trace
            ],
            "terminal_submissions": 0,
            "function_calls_total": 3,
            "observed_terminal_tool": None,
        }
    )
    row["agent"].pop("terminal_response")
    row.update(
        {
            "status": "error",
            "outcome_kind": "infrastructure_failure",
            "passed": False,
            "score_available": False,
            "error": "Recalculation changed sheet identity",
            "error_type": "RecalculationIntegrityError",
            "error_retryable": False,
            "error_category": "recalculation_infrastructure",
            "infrastructure_failure_stage": "agent_tool_recalculation",
            "agent_failure_stage": "solve",
            "infrastructure_failure_tool": "recalculate_and_read",
            "recalculation_failure_reason": "sheet_inventory_changed",
            "recalculation": recalculation,
        }
    )
    row.pop("comparison")
    row.pop("artifact_score_passed")


def _mock_recalculation(
    output: Path,
    monkeypatch: pytest.MonkeyPatch,
    *,
    change_sheet_identity: bool,
) -> dict[str, Any]:
    monkeypatch.setattr(render_module, "find_libreoffice", lambda explicit=None: "/fake/soffice")
    monkeypatch.setattr(render_module, "libreoffice_version", lambda binary: "LibreOffice test")

    def fake_convert(
        source_copy: Path,
        output_dir: Path,
        **kwargs: object,
    ) -> Path:
        output_dir.mkdir(parents=True)
        converted = output_dir / output.name
        shutil.copy2(source_copy, converted)
        if change_sheet_identity:
            workbook = load_workbook(converted)
            try:
                workbook.active.title = "Changed"
                workbook.save(converted)
            finally:
                workbook.close()
        return converted

    monkeypatch.setattr(render_module, "_convert_with_libreoffice", fake_convert)
    if not change_sheet_identity:
        return recalculate_workbook(output, output)
    with pytest.raises(RecalculationIntegrityError) as caught:
        recalculate_workbook(output, output)
    return caught.value.evidence


def _truncated_terminal_fixture(
    tmp_path: Path,
) -> tuple[Path, SpreadsheetTask, dict[str, Any]]:
    results, task, row = _fixture(tmp_path)
    manifest_path = results / "comparison-manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest["configuration"]["api_protocol"] = "chat-completions"
    manifest_path.write_text(json.dumps(manifest) + "\n", encoding="utf-8")

    attempt = {
        "attempt": 1,
        "outcome": "error",
        "error_type": "ProviderOutputLimitError",
        "message": "provider output limit",
        "phase": "response_body",
        "elapsed_seconds": 0.4,
        "first_event_seconds": 0.4,
        "headers_seconds": 0.4,
        "terminal_seconds": 0.4,
        "terminal_event": "chat.completion",
        "status_code": 200,
        "sse_events": 0,
        "transport_exception_type": None,
        "retryable": False,
        "safe_to_retry": False,
        "safe_retry_reason": None,
        "retry_after_seconds": None,
        "backoff_requested_seconds": None,
        "backoff_seconds": None,
        "overload_detected": False,
        "no_header_read_timeout": False,
        "retry_backoff_reason": None,
        "automatic_retry_scheduled": False,
        "automatic_retry_suppressed_reason": "delivery_not_known_safe",
        "logical_request_id": "logical-final",
        "client_request_id": "logical-final-1",
        "request_payload_sha256": "a" * 64,
        "response_headers": {},
        "delivery_state": "terminal_seen",
        "pacing": {"wait_seconds": 0.0},
        "api_protocol": "chat-completions",
        "endpoint": "/chat/completions",
    }
    response_timing = {
        "attempts": 1,
        "elapsed_seconds": 0.4,
        "first_event_seconds": 0.4,
        "headers_seconds": 0.4,
        "terminal_seconds": 0.4,
        "terminal_event": "chat.completion",
        "status_code": 200,
        "sse_events": 0,
        "logical_request_id": "logical-final",
        "client_request_id": "logical-final-1",
        "request_payload_sha256": "a" * 64,
        "response_headers": {},
        "delivery_state": "terminal_seen",
        "pacing_wait_seconds_total": 0.0,
        "attempt_history": [attempt],
    }
    final_timing = {
        "turn": 3,
        "stage": "solve",
        **response_timing,
        "input_serialized_chars": 12,
        "input_serialized_bytes": 12,
        "request_body_chars": 24,
        "request_body_bytes": 24,
        "history_summary_chars": 0,
        "recent_raw_tool_output_chars": 0,
        "recent_image_bytes": 0,
        "recent_image_count": 0,
        "input_tokens": 4,
        "output_tokens": 2,
        "total_tokens": 6,
    }
    terminal_response = {
        "status": "truncated",
        "finish_reason": "length",
        "response_id": "chat-truncated",
        "usage": {"input_tokens": 4, "output_tokens": 2, "total_tokens": 6},
        "timing": response_timing,
        "discarded_message": {
            "sha256": "b" * 64,
            "serialized_chars": 80,
            "serialized_bytes": 80,
            "top_level_field_count": 3,
            "content_item_count": 0,
            "tool_call_count": 1,
        },
    }

    stage = row["agent"]["stages"][-1]
    stage_agent = stage["agent"]
    for timing in stage_agent["request_timings"][:-1]:
        timing["attempt_history"][0].update(
            {
                "api_protocol": "chat-completions",
                "endpoint": "/chat/completions",
            }
        )
    stage_agent["request_timings"][-1] = final_timing
    stage.update(
        {
            "post_prefix_tool_choice": "auto",
            "observed_terminal_tool": "submit_result_length",
        }
    )
    stage_agent.update(
        {
            "final_text": "Terminal response truncated.",
            "response_id": "chat-truncated",
            "first_tool_choice": "code_interpreter",
            "observed_first_tool": "code_interpreter",
            "forced_tool_prefix": ["code_interpreter", "code_interpreter"],
            "observed_forced_tool_prefix": ["code_interpreter", "code_interpreter"],
            "post_prefix_tool_choice": "auto",
            "terminal_tool": "submit_result",
            "observed_terminal_tool": "submit_result_length",
            "terminal_submissions": 0,
            "function_calls_total": 2,
            "terminal_response": terminal_response,
        }
    )
    aggregate_timings = json.loads(json.dumps(stage_agent["request_timings"]))
    row["agent"].update(
        {
            "final_text": "Terminal response truncated.",
            "response_id": "chat-truncated",
            "request_timings": aggregate_timings,
            "first_tool_choice": "code_interpreter",
            "observed_first_tool": "code_interpreter",
            "forced_tool_prefix": ["code_interpreter", "code_interpreter"],
            "observed_forced_tool_prefix": ["code_interpreter", "code_interpreter"],
            "post_prefix_tool_choice": "auto",
            "terminal_tool": "submit_result",
            "observed_terminal_tool": "submit_result_length",
            "terminal_submissions": 0,
            "function_calls_total": 2,
            "terminal_response": json.loads(json.dumps(terminal_response)),
        }
    )
    row.update(
        {
            "comparison_protocol_version": COMPARISON_PROTOCOL_VERSION,
            "comparison_manifest_sha256": _sha256(manifest_path),
            "api_protocol": "chat-completions",
            "status": "completed",
            "outcome_kind": "model_execution_failure",
            "passed": False,
            "error": "Terminal response truncated.",
            "error_type": "AgentExecutionFailure",
            "error_retryable": False,
            "error_category": "model_execution_failure",
            "model_failure_reason": "terminal_submission_truncated",
        }
    )
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")
    return results, task, row


def _budget_truncated_terminal_fixture(
    tmp_path: Path,
) -> tuple[Path, SpreadsheetTask, dict[str, Any]]:
    results, task, row = _truncated_terminal_fixture(tmp_path)
    manifest_path = results / "comparison-manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest["configuration"]["max_total_tokens"] = 5
    manifest_path.write_text(json.dumps(manifest) + "\n", encoding="utf-8")

    termination = {
        "reason": "max_total_tokens",
        "message": "token budget exhausted after the output-limit response",
        "stage": "solve",
        "elapsed_seconds": 1.0,
    }
    budget = {
        "limit": {
            "model_calls": 3,
            "total_tokens": 5,
            "elapsed_seconds": 30,
        },
        "used": {"model_calls": 3, "total_tokens": 10, "elapsed_seconds": 1.0},
        "termination": termination,
    }
    final_stage = row["agent"]["stages"][-1]
    final_stage["observed_terminal_tool"] = "budget_exhausted"
    final_stage["agent"]["observed_terminal_tool"] = "budget_exhausted"
    final_stage["agent"]["budget"] = json.loads(json.dumps(budget))
    row["agent"]["observed_terminal_tool"] = "budget_exhausted"
    row["agent"]["budget"] = json.loads(json.dumps(budget))
    row["budget"] = budget
    row.update(
        {
            "comparison_manifest_sha256": _sha256(manifest_path),
            "error": "token budget exhausted after the output-limit response",
            "model_failure_reason": "budget_exhausted",
        }
    )
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")
    return results, task, row


def _paper_budget_fixture(
    tmp_path: Path,
    *,
    failed_stage: str,
    termination_reason: str = "max_total_tokens",
    failure_turns: int = 0,
) -> tuple[Path, SpreadsheetTask, dict[str, Any]]:
    results, task, row = _fixture(tmp_path)
    manifest_path = results / "comparison-manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    caps = comparison_stage_turn_caps(20, ("paper",))["paper"]
    expected_names = list(caps)
    failed_index = expected_names.index(failed_stage)
    manifest["arms"] = ["paper"]
    manifest["configuration"]["max_model_calls"] = 20
    manifest["configuration"]["max_turns_per_arm"] = 20
    manifest["stage_turn_caps"] = {"paper": caps}
    manifest["forced_tool_prefix_routing"] = {
        "paper": {
            name: list(prefix)
            for name, prefix in COMPARISON_FORCED_TOOL_PREFIX_POLICY["paper"].items()
        }
    }
    manifest["stage_allowed_tools"] = _stage_allowed_tools_policy(("paper",))
    manifest["allowed_observed_terminals"] = _allowed_observed_terminals_policy(
        {"paper": caps}
    )
    manifest_path.write_text(json.dumps(manifest) + "\n", encoding="utf-8")

    old_run_dir = Path(row["run_dir"])
    run_dir = old_run_dir.with_name("paper")
    old_run_dir.rename(run_dir)
    output = run_dir / "artifacts" / "output.xlsx"
    attempt = {"api_protocol": "responses", "endpoint": "/responses"}
    aggregate_timings: list[dict[str, Any]] = []
    stages: list[dict[str, Any]] = []
    for name in expected_names[: failed_index + 1]:
        prefix = list(COMPARISON_FORCED_TOOL_PREFIX_POLICY["paper"][name])
        is_failure = name == failed_stage
        terminal_tool = "assistant_text" if name == "reconcile" else "submit_result"
        turns = failure_turns if is_failure else max(len(prefix) + 1, 1)
        per_turn_tokens = 0
        timings = [
            {
                "turn": turn,
                "stage": name,
                "attempts": 1,
                "attempt_history": [dict(attempt)],
                "input_tokens": per_turn_tokens,
                "output_tokens": 0,
                "total_tokens": per_turn_tokens,
            }
            for turn in range(1, turns + 1)
        ]
        aggregate_timings.extend(json.loads(json.dumps(timings)))
        terminal_submissions = int(not is_failure and terminal_tool == "submit_result")
        tool_trace = [
            {"name": tool_name, "ok": True}
            for tool_name in ([] if is_failure else prefix)
        ]
        stage_usage = {"input_tokens": 0, "output_tokens": 0, "total_tokens": 0}
        stage_budget = {
            "limit": {
                "model_calls": 20,
                "total_tokens": 100,
                "elapsed_seconds": 30,
            },
            "used": {
                "model_calls": len(aggregate_timings),
                "total_tokens": 0,
                "elapsed_seconds": 0.5,
            },
            "termination": None,
        }
        stages.append(
            {
                "name": name,
                "max_turns": caps[name],
                "allowed_tools": manifest["stage_allowed_tools"]["paper"][name],
                "first_tool_choice": prefix[0] if prefix else None,
                "observed_first_tool": prefix[0] if prefix and not is_failure else None,
                "forced_tool_prefix": prefix,
                "observed_forced_tool_prefix": [] if is_failure else prefix,
                "terminal_tool": (
                    ("assistant_text" if turns == 0 else None)
                    if is_failure and name == "reconcile"
                    else terminal_tool
                ),
                "observed_terminal_tool": (
                    "budget_exhausted" if is_failure else terminal_tool
                ),
                "tool_name_trace": [item["name"] for item in tool_trace],
                "tool_trace": json.loads(json.dumps(tool_trace)),
                "agent": {
                    "turns": turns,
                    "tool_calls": len(tool_trace),
                    "usage": stage_usage,
                    "request_timings": timings,
                    "tool_trace": json.loads(json.dumps(tool_trace)),
                    "terminal_submissions": terminal_submissions,
                    "function_calls_total": len(tool_trace) + terminal_submissions,
                    "budget": stage_budget,
                },
            }
        )
        stage = stages[-1]
        stage_agent = stage["agent"]
        if is_failure:
            stage_agent.update(
                {
                    "final_text": "budget exhausted",
                    "response_id": None,
                    "terminal_tool": stage["terminal_tool"],
                    "observed_terminal_tool": "budget_exhausted",
                }
            )
        elif name == "reconcile":
            stage_agent.update(
                {
                    "final_text": "reconciled: true\nprovenance:\n- range: A1",
                    "response_id": "response-reconcile",
                    "terminal_tool": "assistant_text",
                    "observed_terminal_tool": "assistant_text",
                }
            )
        else:
            final_text = f"{name}: verified\nprovenance:\n- range: A1"
            response_id = f"response-{name}"
            stage["post_prefix_tool_choice"] = "auto"
            stage_agent.update(
                {
                    "final_text": final_text,
                    "response_id": response_id,
                    "post_prefix_tool_choice": "auto",
                    "terminal_tool": "submit_result",
                    "observed_terminal_tool": "submit_result",
                    "terminal_response": {
                        "status": "accepted",
                        "response_id": response_id,
                        "acknowledgement": {
                            "mode": "evidence_result",
                            "result_chars": len(final_text),
                            "result_sha256": hashlib.sha256(
                                final_text.encode("utf-8")
                            ).hexdigest(),
                        },
                    },
                }
            )

    used_calls = len(aggregate_timings)
    used_tokens = 100 if termination_reason == "max_total_tokens" else 10
    if aggregate_timings:
        aggregate_timings[-1]["input_tokens"] = used_tokens
        aggregate_timings[-1]["total_tokens"] = used_tokens
        timing_stage = next(
            stage
            for stage in reversed(stages)
            if stage["agent"]["request_timings"]
        )
        timing_stage["agent"]["request_timings"][-1]["input_tokens"] = used_tokens
        timing_stage["agent"]["request_timings"][-1]["total_tokens"] = used_tokens
        timing_stage["agent"]["usage"]["input_tokens"] = used_tokens
        timing_stage["agent"]["usage"]["total_tokens"] = used_tokens
    limit_calls = used_calls if termination_reason == "max_model_calls" else 20
    if termination_reason == "max_model_calls":
        manifest["configuration"]["max_model_calls"] = limit_calls
        manifest_path.write_text(json.dumps(manifest) + "\n", encoding="utf-8")
    termination = {
        "reason": termination_reason,
        "message": "budget exhausted",
        "stage": failed_stage,
        "elapsed_seconds": 1.0,
    }
    budget = {
        "limit": {
            "model_calls": limit_calls,
            "total_tokens": 100,
            "elapsed_seconds": 30,
        },
        "used": {
            "model_calls": used_calls,
            "total_tokens": used_tokens,
            "elapsed_seconds": 1.0,
        },
        "termination": termination,
    }
    for index, stage in enumerate(stages):
        stage_budget = stage["agent"]["budget"]
        stage_budget["limit"]["model_calls"] = limit_calls
        stage_budget["used"]["model_calls"] = sum(
            item["agent"]["turns"] for item in stages[: index + 1]
        )
        stage_budget["used"]["total_tokens"] = sum(
            item["agent"]["usage"]["total_tokens"] for item in stages[: index + 1]
        )
    stages[-1]["agent"]["budget"] = json.loads(json.dumps(budget))
    aggregate_tool_trace = [
        {"stage": stage["name"], **item}
        for stage in stages
        for item in stage["tool_trace"]
    ]
    aggregate_turns = sum(stage["agent"]["turns"] for stage in stages)
    aggregate_tool_calls = sum(stage["agent"]["tool_calls"] for stage in stages)
    aggregate_terminal_submissions = sum(
        stage["agent"]["terminal_submissions"] for stage in stages
    )
    row.update(
        {
            "arm": "paper",
            "comparison_manifest_sha256": _sha256(manifest_path),
            "max_model_calls": limit_calls,
            "max_turns_per_arm": 20,
            "stage_turn_caps": caps,
            "status": "completed",
            "outcome_kind": "model_execution_failure",
            "passed": False,
            "error": "budget exhausted",
            "error_type": "AgentExecutionFailure",
            "error_retryable": False,
            "error_category": "model_execution_failure",
            "model_failure_reason": "budget_exhausted",
            "run_dir": str(run_dir),
            "output_workbook": str(output),
            "agent": {
                "arm": "paper",
                "final_text": stages[-1]["agent"]["final_text"],
                "response_id": stages[-1]["agent"]["response_id"],
                "turns": aggregate_turns,
                "tool_calls": aggregate_tool_calls,
                "usage": {
                    "input_tokens": used_tokens,
                    "output_tokens": 0,
                    "total_tokens": used_tokens,
                },
                "request_timings": aggregate_timings,
                "tool_trace": aggregate_tool_trace,
                "terminal_submissions": aggregate_terminal_submissions,
                "function_calls_total": (
                    aggregate_tool_calls + aggregate_terminal_submissions
                ),
                "stages": stages,
                "budget": budget,
            },
            "budget": budget,
        }
    )
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")
    return results, task, row


def _paper_scored_fixture(
    tmp_path: Path,
) -> tuple[Path, SpreadsheetTask, dict[str, Any]]:
    results, task, row = _paper_budget_fixture(
        tmp_path,
        failed_stage="solve",
        failure_turns=3,
    )
    row = json.loads(json.dumps(row))
    stages = row["agent"]["stages"]

    for budget in [row["budget"], row["agent"]["budget"], *(
        stage["agent"]["budget"] for stage in stages
    )]:
        budget["termination"] = None

    for stage in stages:
        name = stage["name"]
        stage_agent = stage["agent"]
        response_id = f"response-{name}"
        stage_agent.update(
            {
                "response_id": response_id,
                "terminal_tool": stage["terminal_tool"],
                "observed_terminal_tool": stage["observed_terminal_tool"],
            }
        )
        if name in {"extract", "vision_verify", "latex_verify"}:
            final_text = f"{name}: verified\nprovenance:\n- range: A1"
            acknowledgement = {
                "mode": "evidence_result",
                "result_chars": len(final_text),
                "result_sha256": hashlib.sha256(
                    final_text.encode("utf-8")
                ).hexdigest(),
            }
            stage_agent.update(
                {
                    "final_text": final_text,
                    "post_prefix_tool_choice": "auto",
                    "terminal_response": {
                        "status": "accepted",
                        "response_id": response_id,
                        "acknowledgement": acknowledgement,
                    },
                }
            )
            stage["post_prefix_tool_choice"] = "auto"
        elif name == "reconcile":
            stage_agent["final_text"] = (
                "reconciled: true\nprovenance:\n- range: A1"
            )
            stage["post_prefix_tool_choice"] = None
        else:
            prefix = list(COMPARISON_FORCED_TOOL_PREFIX_POLICY["paper"]["solve"])
            tool_trace = [{"name": tool_name, "ok": True} for tool_name in prefix]
            stage.update(
                {
                    "observed_first_tool": prefix[0],
                    "observed_forced_tool_prefix": prefix,
                    "observed_terminal_tool": "submit_result",
                    "post_prefix_tool_choice": "auto",
                    "tool_name_trace": prefix,
                    "tool_trace": json.loads(json.dumps(tool_trace)),
                }
            )
            stage_agent.update(
                {
                    "final_text": "Spreadsheet task completed.",
                    "tool_calls": len(tool_trace),
                    "tool_trace": json.loads(json.dumps(tool_trace)),
                    "terminal_submissions": 1,
                    "function_calls_total": len(tool_trace) + 1,
                    "post_prefix_tool_choice": "auto",
                    "terminal_tool": "submit_result",
                    "observed_terminal_tool": "submit_result",
                    "terminal_response": {
                        "status": "accepted",
                        "response_id": response_id,
                        "acknowledgement": {},
                    },
                }
            )

    aggregate_tool_trace = [
        {"stage": stage["name"], **item}
        for stage in stages
        for item in stage["tool_trace"]
    ]
    aggregate_tool_calls = sum(stage["agent"]["tool_calls"] for stage in stages)
    aggregate_terminal_submissions = sum(
        stage["agent"]["terminal_submissions"] for stage in stages
    )
    final_agent = stages[-1]["agent"]
    row["agent"].update(
        {
            "final_text": final_agent["final_text"],
            "response_id": final_agent["response_id"],
            "tool_calls": aggregate_tool_calls,
            "tool_trace": aggregate_tool_trace,
            "terminal_submissions": aggregate_terminal_submissions,
            "function_calls_total": (
                aggregate_tool_calls + aggregate_terminal_submissions
            ),
            "post_prefix_tool_choice": "auto",
            "terminal_tool": "submit_result",
            "observed_terminal_tool": "submit_result",
            "terminal_response": json.loads(
                json.dumps(final_agent["terminal_response"])
            ),
        }
    )
    row.update({"outcome_kind": "scored", "passed": True})
    for field in (
        "error",
        "error_type",
        "error_retryable",
        "error_category",
        "model_failure_reason",
    ):
        row.pop(field, None)
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")
    return results, task, row


def _tree_hashes(root: Path) -> dict[str, str]:
    return {
        str(path.relative_to(root)): _sha256(path)
        for path in sorted(root.rglob("*"))
        if path.is_file()
    }


def _row_reason(summary: dict[str, Any], reason: str) -> bool:
    return any(reason in row["reasons"] for row in summary["rows"])


def _interrupted_seal(results: Path, task: SpreadsheetTask) -> dict[str, Any]:
    manifest_path = results / "comparison-manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    return {
        "schema_version": 1,
        "task_id": task.task_id,
        "arm": "bare",
        "comparison_protocol_version": COMPARISON_PROTOCOL_VERSION,
        "comparison_manifest_sha256": _sha256(manifest_path),
        "split_provenance": manifest.get("split_provenance"),
        "run_spec_provenance": manifest.get("run_spec_provenance"),
        "status": "interrupted",
        "passed": None,
        "outcome_observed": False,
        "score_available": False,
        "usage_observed": False,
        "replay_permitted": False,
        "error_retryable": False,
        "error_category": "interrupted_unknown_outcome",
        "sealed_at": "2026-08-14T12:00:00+00:00",
        "sealed_from_inflight_marker_sha256": "a" * 64,
    }


def _write_interrupted_seals(results: Path, seals: list[dict[str, Any]]) -> None:
    (results / INTERRUPTED_SEALS_FILENAME).write_text(
        json.dumps({"schema_version": 1, "seals": seals}) + "\n",
        encoding="utf-8",
    )


def _continuation_source(results: Path) -> dict[str, Any]:
    manifest_path = results / "comparison-manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    record: dict[str, Any] = {
        "schema_version": 1,
        "comparison_manifest_sha256": _sha256(manifest_path),
        "repository_source": {
            "schema_version": 1,
            "git_commit": "1" * 40,
            "git_tree": "2" * 40,
            "remote_tracking_ref": "refs/remotes/origin/main",
            "remote_tracking_commit": "1" * 40,
            "remote_name": "origin",
            "remote_ref": "refs/heads/main",
            "remote_observed_commit": "1" * 40,
            "source_fingerprint": manifest["harness_source"],
        },
    }
    unsigned = json.dumps(
        record,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )
    record["record_sha256"] = _text_sha256(unsigned)
    return record


def _write_continuation_source(results: Path, record: dict[str, Any]) -> None:
    (results / CONTINUATION_SOURCE_FILENAME).write_text(
        json.dumps(record) + "\n",
        encoding="utf-8",
    )


def test_audit_comparison_valid_and_read_only(tmp_path: Path) -> None:
    results, task, _ = _fixture(tmp_path)
    before = _tree_hashes(tmp_path)

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is True
    assert summary["reasons"] == []
    assert summary["expected_rows"] == summary["observed_rows"] == 1
    assert summary["valid_rows"] == 1
    assert summary["rows"][0]["audit_valid"] is True
    assert summary["rows"][0]["fresh_comparison"]["passed"] is True
    assert "mcnemar_exact_p" not in summary
    assert "stratified_bootstrap_95" not in summary
    assert "holm_adjusted_p" not in summary
    assert summary["rows"][0]["output_sha256"] == summary["rows"][0][
        "expected_output_sha256"
    ]
    assert _tree_hashes(tmp_path) == before


def test_audit_accepts_verified_recalculation_identity_evidence(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    results, task, row = _fixture(tmp_path)
    output = Path(row["output_workbook"])
    row["recalculation"] = _mock_recalculation(
        output,
        monkeypatch,
        change_sheet_identity=False,
    )
    row["output_sha256"] = _sha256(output)
    _set_recalculation_manifest(results, row)
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is True
    assert summary["study_complete"] is True
    assert summary["inference_valid"] is True
    assert summary["rows"][0]["audit_valid"] is True


def test_v28_audit_reopens_and_scores_hidden_chartsheet_through_worksheet_view(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    results, task, row = _fixture(tmp_path)
    output = Path(row["output_workbook"])
    for workbook_path in (task.golden_path, output):
        workbook = load_workbook(workbook_path)
        try:
            workbook.create_chartsheet("Chart").sheet_state = "hidden"
            workbook.save(workbook_path)
        finally:
            workbook.close()

    comparison = benchmark_module.compare_workbooks_chartsheet_safe(
        task.golden_path,
        output,
        task.answer_position,
        answer_sheet=task.answer_sheet,
    )
    manifest_path = results / "comparison-manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest["tasks"][0]["golden_sha256"] = _sha256(task.golden_path)
    manifest_path.write_text(json.dumps(manifest) + "\n", encoding="utf-8")
    row.update(
        {
            "comparison_manifest_sha256": _sha256(manifest_path),
            "passed": comparison.passed,
            "artifact_score_passed": comparison.passed,
            "comparison": comparison.to_dict(),
            "output_sha256": _sha256(output),
        }
    )
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")

    reopened_paths: list[Path] = []
    original_loader = audit_module.load_workbook

    def tracking_loader(path: str | Path, *args: Any, **kwargs: Any) -> Any:
        reopened_paths.append(Path(path))
        return original_loader(path, *args, **kwargs)

    monkeypatch.setattr(audit_module, "load_workbook", tracking_loader)
    output_sha256 = _sha256(output)

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is True
    assert summary["inference_valid"] is True
    assert reopened_paths and all(path != output for path in reopened_paths)
    audited = summary["rows"][0]
    assert audited["artifact_validation_method"] == (
        "ooxml-inventory-plus-worksheet-only-openpyxl-view-v1"
    )
    assert audited["sheet_names"] == ["Sheet1", "Chart"]
    assert audited["worksheet_names"] == ["Sheet1"]
    assert audited["fresh_comparison"]["passed"] is True
    assert _sha256(output) == output_sha256


def test_v28_audit_classifies_unavailable_fresh_scorer_as_infrastructure_no_score(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    results, task, _ = _fixture(tmp_path)

    def unavailable_scorer(*_: Any, **__: Any) -> Any:
        raise ScoringInfrastructureError("validated worksheet view is unsupported")

    monkeypatch.setattr(
        audit_module,
        "compare_workbooks_chartsheet_safe",
        unavailable_scorer,
    )

    summary = audit_comparison(results, [task])

    audited = summary["rows"][0]
    assert summary["audit_valid"] is False
    assert audited["score_available"] is False
    assert audited["scorer_infrastructure_stage"] == "fresh_score"
    assert audited["scorer_infrastructure_error_type"] == (
        "ScoringInfrastructureError"
    )
    assert "scorer_infrastructure_no_score" in audited["reasons"]
    assert "artifact_reopen_failed" not in audited["reasons"]


def test_v28_audit_accepts_reproduced_scoring_infrastructure_no_score_row(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    results, task, row = _fixture(tmp_path)
    row.update(
        {
            "status": "error",
            "outcome_kind": "infrastructure_failure",
            "passed": False,
            "score_available": False,
            "error": "worksheet scorer unsupported",
            "error_type": "ScoringInfrastructureError",
            "error_retryable": False,
            "error_category": "scoring_infrastructure",
            "infrastructure_failure_stage": "scoring",
            "scoring_failure_reason": "worksheet_scorer_unsupported",
        }
    )
    row.pop("comparison")
    row.pop("artifact_score_passed")
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")

    def unavailable_scorer(*_: Any, **__: Any) -> Any:
        raise ScoringInfrastructureError("worksheet scorer unsupported")

    monkeypatch.setattr(
        audit_module,
        "compare_workbooks_chartsheet_safe",
        unavailable_scorer,
    )

    summary = audit_comparison(results, [task])

    audited = summary["rows"][0]
    assert summary["audit_valid"] is True, (summary["reasons"], audited)
    assert summary["journal_integrity_valid"] is True
    assert summary["study_complete"] is False
    assert summary["inference_valid"] is False
    assert summary["inference_invalid_reasons"] == [
        "scoring_infrastructure_failure"
    ]
    assert summary["known_scoring_infrastructure_failure_rows"] == 1
    assert summary["known_recalculation_infrastructure_failure_rows"] == 0
    assert summary["known_passed_rows"] == 0
    assert summary["known_failed_rows"] == 0
    assert audited["audit_valid"] is True
    assert audited["score_available"] is False
    assert audited["scorer_infrastructure_reproduced"] is True
    assert "artifact_reopen_failed" not in audited["reasons"]


def test_audit_sheet_inventory_schema_rejects_duplicate_names_and_no_visible_sheet() -> None:
    sheets = [
        {"index": 0, "kind": "worksheet", "name": "Data", "visibility": "visible"},
        {"index": 1, "kind": "chartsheet", "name": "Chart", "visibility": "hidden"},
    ]

    def identity(records: list[dict[str, Any]]) -> dict[str, Any]:
        encoded = json.dumps(
            records,
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        )
        return {
            "schema_version": 2,
            "workbook_sha256": "a" * 64,
            "inventory_sha256": _text_sha256(encoded),
            "sheets": records,
        }

    assert _valid_sheet_inventory_identity(identity(sheets)) is True
    duplicate = [dict(sheet) for sheet in sheets]
    duplicate[1]["name"] = "data"
    assert _valid_sheet_inventory_identity(identity(duplicate)) is False
    all_hidden = [{**sheet, "visibility": "hidden"} for sheet in sheets]
    assert _valid_sheet_inventory_identity(identity(all_hidden)) is False


def test_audit_rejects_duplicate_workbook_xml_recalculation_evidence(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    results, task, row = _fixture(tmp_path)
    output = Path(row["output_workbook"])
    recalculation = _mock_recalculation(
        output,
        monkeypatch,
        change_sheet_identity=False,
    )

    with zipfile.ZipFile(output) as package:
        workbook_xml = package.read("xl/workbook.xml")
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        with zipfile.ZipFile(output, "a") as package:
            package.writestr("xl/workbook.xml", workbook_xml)

    output_sha256 = _sha256(output)
    recalculation["source_sha256"] = output_sha256
    recalculation["output_sha256"] = output_sha256
    recalculation["sheet_inventory_integrity"]["pre"][
        "workbook_sha256"
    ] = output_sha256
    recalculation["sheet_inventory_integrity"]["post"][
        "workbook_sha256"
    ] = output_sha256
    comparison = compare_workbooks(
        task.golden_path,
        output,
        task.answer_position,
        answer_sheet=task.answer_sheet,
    )
    row.update(
        {
            "recalculation": recalculation,
            "output_sha256": output_sha256,
            "passed": comparison.passed,
            "artifact_score_passed": comparison.passed,
            "comparison": comparison.to_dict(),
        }
    )
    _set_recalculation_manifest(results, row)
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is False
    assert "artifact_reopen_failed" in summary["rows"][0]["reasons"]


def test_audit_requires_recalculation_identity_evidence_when_enabled(tmp_path: Path) -> None:
    results, task, row = _fixture(tmp_path)
    _set_recalculation_manifest(results, row)
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is False
    assert "recalculation_evidence_missing" in summary["rows"][0]["reasons"]


def test_audit_accepts_recalculation_identity_failure_but_invalidates_inference(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    results, task, row = _fixture(tmp_path)
    output = Path(row["output_workbook"])
    recalculation = _mock_recalculation(
        output,
        monkeypatch,
        change_sheet_identity=True,
    )
    _set_recalculation_manifest(results, row)
    row.update(
        {
            "status": "error",
            "outcome_kind": "infrastructure_failure",
            "passed": False,
            "score_available": False,
            "error": "Recalculation changed sheet identity",
            "error_type": "RecalculationIntegrityError",
            "error_retryable": False,
            "error_category": "recalculation_infrastructure",
            "infrastructure_failure_stage": "recalculation",
            "recalculation_failure_reason": "sheet_inventory_changed",
            "recalculation": recalculation,
            "output_sha256": _sha256(output),
        }
    )
    row.pop("comparison")
    row.pop("artifact_score_passed")
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is True
    assert summary["journal_integrity_valid"] is True
    assert summary["study_complete"] is False
    assert summary["inference_valid"] is False
    assert summary["inference_invalid_reasons"] == [
        "recalculation_infrastructure_failure"
    ]
    assert summary["known_passed_rows"] == 0
    assert summary["known_failed_rows"] == 0
    assert summary["known_recalculation_infrastructure_failure_rows"] == 1
    assert summary["rows"][0]["score_available"] is False
    assert "outcome_passed" not in summary["rows"][0]


def test_audit_accepts_agent_tool_recalculation_failure_without_terminal(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    results, task, row = _fixture(tmp_path, arm="ours")
    output = Path(row["output_workbook"])
    recalculation = _mock_recalculation(
        output,
        monkeypatch,
        change_sheet_identity=True,
    )
    _set_recalculation_manifest(results, row)
    _set_agent_tool_recalculation_failure(row, recalculation)
    row["output_sha256"] = _sha256(output)
    (results / "results.jsonl").write_text(
        json.dumps(row) + "\n", encoding="utf-8"
    )

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is True
    assert summary["journal_integrity_valid"] is True
    assert summary["study_complete"] is False
    assert summary["inference_valid"] is False
    assert summary["inference_invalid_reasons"] == [
        "recalculation_infrastructure_failure"
    ]
    assert summary["known_recalculation_infrastructure_failure_rows"] == 1
    audited = summary["rows"][0]
    assert audited["audit_valid"] is True
    assert audited["infrastructure_failure_stage"] == (
        "agent_tool_recalculation"
    )
    assert audited["agent_failure_stage"] == "solve"
    assert audited["infrastructure_failure_tool"] == "recalculate_and_read"


@pytest.mark.parametrize(
    "tamper",
    [
        "missing_agent_stage",
        "forged_terminal",
        "aggregate_terminal_submission",
        "missing_failure_trace",
    ],
)
def test_audit_rejects_invalid_agent_tool_recalculation_evidence(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    tamper: str,
) -> None:
    results, task, row = _fixture(tmp_path, arm="ours")
    output = Path(row["output_workbook"])
    recalculation = _mock_recalculation(
        output,
        monkeypatch,
        change_sheet_identity=True,
    )
    _set_recalculation_manifest(results, row)
    _set_agent_tool_recalculation_failure(row, recalculation)
    row["output_sha256"] = _sha256(output)
    stage = row["agent"]["stages"][-1]
    if tamper == "missing_agent_stage":
        row.pop("agent_failure_stage")
    elif tamper == "forged_terminal":
        stage["observed_terminal_tool"] = "submit_result"
        stage["agent"]["observed_terminal_tool"] = "submit_result"
        row["agent"]["observed_terminal_tool"] = "submit_result"
    elif tamper == "aggregate_terminal_submission":
        row["agent"]["terminal_submissions"] = 1
    else:
        stage["tool_trace"].pop()
        stage["agent"]["tool_trace"].pop()
        row["agent"]["tool_trace"].pop()
    (results / "results.jsonl").write_text(
        json.dumps(row) + "\n", encoding="utf-8"
    )

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is False
    assert summary["rows"][0]["audit_valid"] is False
    assert summary["rows"][0]["reasons"]


@pytest.mark.parametrize("tamper", ["pre_digest", "match_claim", "post_hash", "score"])
def test_audit_rejects_tampered_recalculation_failure_evidence(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    tamper: str,
) -> None:
    results, task, row = _fixture(tmp_path)
    output = Path(row["output_workbook"])
    recalculation = _mock_recalculation(
        output,
        monkeypatch,
        change_sheet_identity=True,
    )
    _set_recalculation_manifest(results, row)
    row.update(
        {
            "status": "error",
            "outcome_kind": "infrastructure_failure",
            "passed": False,
            "score_available": False,
            "error": "Recalculation changed sheet identity",
            "error_type": "RecalculationIntegrityError",
            "error_retryable": False,
            "error_category": "recalculation_infrastructure",
            "infrastructure_failure_stage": "recalculation",
            "recalculation_failure_reason": "sheet_inventory_changed",
            "recalculation": recalculation,
            "output_sha256": _sha256(output),
        }
    )
    row.pop("comparison")
    row.pop("artifact_score_passed")
    integrity = recalculation["sheet_inventory_integrity"]
    if tamper == "pre_digest":
        integrity["pre"]["inventory_sha256"] = "0" * 64
    elif tamper == "match_claim":
        integrity["matched"] = True
    elif tamper == "post_hash":
        recalculation["output_sha256"] = "0" * 64
    else:
        row["comparison"] = {"passed": False}
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is False
    assert summary["rows"][0]["audit_valid"] is False
    assert summary["rows"][0]["reasons"]


def test_live_v23_pilot_audit_has_no_version_drift_false_positives() -> None:
    results = Path(
        "benchmarks/results/qwen36-local-pilot16-v2-bare-ours-v23-seed41"
    )
    dataset = Path("benchmarks/data/spreadsheetbench_verified_400")
    if not results.is_dir() or not dataset.is_dir():
        pytest.skip("ignored historical v23 pilot artifacts are not available")
    manifest = json.loads(
        (results / "comparison-manifest.json").read_text(encoding="utf-8")
    )
    from spreadsheet_harness.benchmark import load_verified_tasks

    tasks_by_id = {task.task_id: task for task in load_verified_tasks(dataset)}
    summary = audit_comparison(
        results,
        [tasks_by_id[task_id] for task_id in manifest["task_ids"]],
    )
    forbidden_exact = {
        "comparison_manifest_schema_mismatch",
        "comparison_manifest_protocol_mismatch",
        "comparison_manifest_policy_mismatch",
        "continuation_source_invalid",
        "interrupted_arm_task_seals_invalid",
    }
    forbidden_fragments = {
        "result_protocol_mismatch:",
        "result_continuation_source_without_record:",
        "outcome_kind_invalid",
        "stored_artifact_score_passed_mismatch",
    }

    assert not (set(summary["reasons"]) & forbidden_exact)
    assert not any(
        fragment in reason
        for reason in summary["reasons"]
        for fragment in forbidden_fragments
    )
    assert summary["valid_rows"] == 9
    assert summary["known_passed_rows"] == 6
    assert summary["known_failed_rows"] == 3


def test_audit_accepts_known_model_execution_failure_as_completed_false(
    tmp_path: Path,
) -> None:
    results, task, row = _fixture(tmp_path)
    _set_final_model_execution_failure(row, "managed workbook remained unchanged")
    row.update(
        {
            "outcome_kind": "model_execution_failure",
            "passed": False,
            "error": "managed workbook remained unchanged",
            "error_type": "AgentExecutionFailure",
            "error_retryable": False,
            "error_category": "model_execution_failure",
            "model_failure_reason": "workbook_unchanged",
        }
    )
    (results / "results.jsonl").write_text(
        json.dumps(row) + "\n", encoding="utf-8"
    )

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is True
    assert summary["study_complete"] is True
    assert summary["inference_valid"] is True
    assert summary["known_passed_rows"] == 0
    assert summary["known_failed_rows"] == 1
    assert summary["known_model_execution_failure_rows"] == 1
    assert summary["rows"][0]["outcome_passed"] is False
    assert summary["rows"][0]["fresh_comparison"]["passed"] is True
    assert summary["rows"][0]["model_failure_reason"] == "workbook_unchanged"


@pytest.mark.parametrize("target", ["aggregate", "final_stage"])
def test_v26_audit_rejects_accepted_response_on_model_execution_failure(
    tmp_path: Path,
    target: str,
) -> None:
    results, task, row = _fixture(tmp_path)
    accepted_response = json.loads(json.dumps(row["agent"]["terminal_response"]))
    _set_final_model_execution_failure(row, "managed workbook remained unchanged")
    row.update(
        {
            "outcome_kind": "model_execution_failure",
            "passed": False,
            "error": "managed workbook remained unchanged",
            "error_type": "AgentExecutionFailure",
            "error_retryable": False,
            "error_category": "model_execution_failure",
            "model_failure_reason": "workbook_unchanged",
        }
    )
    if target == "aggregate":
        row["agent"]["terminal_response"] = accepted_response
    else:
        row["agent"]["stages"][-1]["agent"][
            "terminal_response"
        ] = accepted_response
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is False
    assert _row_reason(summary, "accepted_terminal_response_evidence_invalid")


def test_audit_accepts_only_one_provable_response_token_overage(
    tmp_path: Path,
) -> None:
    results, task, row = _fixture(tmp_path)
    _set_final_model_execution_failure(row, "token budget exhausted")
    row.update(
        {
            "outcome_kind": "model_execution_failure",
            "passed": False,
            "error": "token budget exhausted",
            "error_type": "AgentExecutionFailure",
            "error_retryable": False,
            "error_category": "model_execution_failure",
            "model_failure_reason": "budget_exhausted",
        }
    )
    row["budget"]["used"]["total_tokens"] = 110
    row["budget"]["termination"] = {
        "reason": "max_total_tokens",
        "message": "token budget exhausted",
        "stage": "solve",
        "elapsed_seconds": 1.0,
    }
    row["agent"]["usage"]["total_tokens"] = 110
    row["agent"]["budget"] = row["budget"]
    row["agent"]["stages"][0]["observed_terminal_tool"] = "budget_exhausted"
    stage_agent = row["agent"]["stages"][0]["agent"]
    stage_agent["budget"] = json.loads(json.dumps(row["budget"]))
    stage_agent["usage"] = {"input_tokens": 108, "output_tokens": 2, "total_tokens": 110}
    stage_agent["terminal_submissions"] = 0
    stage_agent["function_calls_total"] = stage_agent["tool_calls"]
    row["agent"]["usage"] = dict(stage_agent["usage"])
    row["agent"]["terminal_submissions"] = 0
    row["agent"]["function_calls_total"] = row["agent"]["tool_calls"]
    for aggregate_timing, stage_timing, tokens in zip(
        row["agent"]["request_timings"],
        stage_agent["request_timings"],
        (40, 40, 30),
        strict=True,
    ):
        for timing in (aggregate_timing, stage_timing):
            timing["input_tokens"] = tokens
            timing["output_tokens"] = 0
            timing["total_tokens"] = tokens
    row["agent"]["request_timings"][-1]["output_tokens"] = 2
    row["agent"]["request_timings"][-1]["input_tokens"] = 28
    stage_agent["request_timings"][-1]["output_tokens"] = 2
    stage_agent["request_timings"][-1]["input_tokens"] = 28

    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")
    accepted = audit_comparison(results, [task])
    assert accepted["audit_valid"] is True
    assert not _row_reason(accepted, "budget_used_invalid:total_tokens")

    for timing, tokens in zip(
        row["agent"]["request_timings"], (50, 50, 10), strict=True
    ):
        timing["input_tokens"] = tokens
        timing["output_tokens"] = 0
        timing["total_tokens"] = tokens
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")
    rejected = audit_comparison(results, [task])
    assert rejected["audit_valid"] is False
    assert _row_reason(rejected, "budget_used_invalid:total_tokens")


def test_v25_budget_failure_allows_only_an_observed_forced_prefix(
    tmp_path: Path,
) -> None:
    results, task, row = _fixture(tmp_path)
    _set_final_model_execution_failure(row, "token budget exhausted")
    termination = {
        "reason": "max_total_tokens",
        "message": "token budget exhausted",
        "stage": "solve",
        "elapsed_seconds": 1.0,
    }
    timing = json.loads(json.dumps(row["agent"]["request_timings"][0]))
    timing.update({"input_tokens": 98, "output_tokens": 2, "total_tokens": 100})
    row.update(
        {
            "outcome_kind": "model_execution_failure",
            "passed": False,
            "error": "token budget exhausted",
            "error_type": "AgentExecutionFailure",
            "error_retryable": False,
            "error_category": "model_execution_failure",
            "model_failure_reason": "budget_exhausted",
        }
    )
    row["budget"]["used"].update({"model_calls": 1, "total_tokens": 100})
    row["budget"]["termination"] = termination
    row["agent"].update(
        {
            "turns": 1,
            "tool_calls": 0,
            "usage": {"input_tokens": 98, "output_tokens": 2, "total_tokens": 100},
            "request_timings": [json.loads(json.dumps(timing))],
            "tool_trace": [],
            "terminal_submissions": 0,
            "function_calls_total": 0,
        }
    )
    row["agent"]["budget"] = row["budget"]
    stage = row["agent"]["stages"][0]
    stage["observed_forced_tool_prefix"] = ["code_interpreter"]
    stage["observed_terminal_tool"] = "budget_exhausted"
    stage["tool_name_trace"] = []
    stage["tool_trace"] = []
    stage["agent"] = {
        "turns": 1,
        "tool_calls": 0,
        "usage": {"input_tokens": 98, "output_tokens": 2, "total_tokens": 100},
        "request_timings": [json.loads(json.dumps(timing))],
        "tool_trace": [],
        "terminal_submissions": 0,
        "function_calls_total": 0,
        "budget": json.loads(json.dumps(row["budget"])),
    }
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")

    accepted = audit_comparison(results, [task])
    assert accepted["audit_valid"] is True

    stage["observed_forced_tool_prefix"] = ["inspect_range"]
    stage["observed_first_tool"] = "inspect_range"
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")
    rejected = audit_comparison(results, [task])
    assert rejected["audit_valid"] is False
    assert _row_reason(rejected, "agent_observed_prefix_mismatch:solve")


def test_nonbudget_failure_still_requires_exact_forced_prefix(
    tmp_path: Path,
) -> None:
    results, task, row = _fixture(tmp_path)
    _set_final_model_execution_failure(row, "managed workbook remained unchanged")
    row.update(
        {
            "outcome_kind": "model_execution_failure",
            "passed": False,
            "error": "workbook unchanged",
            "error_type": "AgentExecutionFailure",
            "error_retryable": False,
            "error_category": "model_execution_failure",
            "model_failure_reason": "workbook_unchanged",
        }
    )
    row["agent"]["stages"][0]["observed_forced_tool_prefix"] = [
        "code_interpreter"
    ]
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")

    summary = audit_comparison(results, [task])
    assert summary["audit_valid"] is False
    assert _row_reason(summary, "agent_observed_prefix_mismatch:solve")


@pytest.mark.parametrize(
    ("outcome_kind", "failure_reason"),
    [("scored", None), ("model_execution_failure", "workbook_unchanged")],
)
def test_nonbudget_outcome_rejects_budget_terminal(
    tmp_path: Path,
    outcome_kind: str,
    failure_reason: str | None,
) -> None:
    results, task, row = _fixture(tmp_path)
    if failure_reason is not None:
        row.update(
            {
                "outcome_kind": outcome_kind,
                "passed": False,
                "error": "workbook unchanged",
                "error_type": "AgentExecutionFailure",
                "error_retryable": False,
                "error_category": "model_execution_failure",
                "model_failure_reason": failure_reason,
            }
        )
    row["agent"]["stages"][0]["observed_terminal_tool"] = "budget_exhausted"
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is False
    assert _row_reason(summary, "agent_observed_terminal_invalid:solve")


@pytest.mark.parametrize("termination_reason", ["max_model_calls", "max_total_tokens"])
def test_paper_budget_failure_allows_stage_prefix_and_zero_turn_final_stage(
    tmp_path: Path,
    termination_reason: str,
) -> None:
    results, task, _ = _paper_budget_fixture(
        tmp_path,
        failed_stage="reconcile",
        termination_reason=termination_reason,
    )

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is True
    assert summary["rows"][0]["model_failure_reason"] == "budget_exhausted"


def test_paper_reconcile_budget_terminal_depends_on_failure_path(
    tmp_path: Path,
) -> None:
    zero_results, zero_task, zero_row = _paper_budget_fixture(
        tmp_path / "zero",
        failed_stage="reconcile",
        failure_turns=0,
    )
    positive_results, positive_task, positive_row = _paper_budget_fixture(
        tmp_path / "positive",
        failed_stage="reconcile",
        failure_turns=1,
    )

    assert audit_comparison(zero_results, [zero_task])["audit_valid"] is True
    assert audit_comparison(positive_results, [positive_task])["audit_valid"] is True

    zero_row = json.loads(json.dumps(zero_row))
    zero_row["agent"]["stages"][-1]["terminal_tool"] = None
    (zero_results / "results.jsonl").write_text(
        json.dumps(zero_row) + "\n", encoding="utf-8"
    )
    zero_rejected = audit_comparison(zero_results, [zero_task])
    assert zero_rejected["audit_valid"] is False
    assert _row_reason(zero_rejected, "agent_terminal_tool_mismatch:reconcile")

    positive_row = json.loads(json.dumps(positive_row))
    positive_row["agent"]["stages"][-1]["terminal_tool"] = "assistant_text"
    (positive_results / "results.jsonl").write_text(
        json.dumps(positive_row) + "\n", encoding="utf-8"
    )
    positive_rejected = audit_comparison(positive_results, [positive_task])
    assert positive_rejected["audit_valid"] is False
    assert _row_reason(positive_rejected, "agent_terminal_tool_mismatch:reconcile")


def test_paper_zero_turn_budget_stage_requires_matching_termination(
    tmp_path: Path,
) -> None:
    results, task, row = _paper_budget_fixture(
        tmp_path,
        failed_stage="reconcile",
    )
    stage_budget = json.loads(json.dumps(row["agent"]["stages"][-1]["agent"]["budget"]))
    stage_budget["termination"]["stage"] = "latex_verify"
    row["agent"]["stages"][-1]["agent"]["budget"] = stage_budget
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is False
    assert _row_reason(summary, "agent_stage_budget_mismatch:reconcile")


def test_v24_contract_keeps_historical_failure_taxonomy_and_budget_rule(
    tmp_path: Path,
) -> None:
    results, task, row = _fixture(tmp_path)
    manifest_path = results / "comparison-manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest["schema_version"] = V24_COMPARISON_MANIFEST_SCHEMA_VERSION
    manifest["comparison_protocol_version"] = V24_COMPARISON_PROTOCOL_VERSION
    manifest["configuration"].update(V24_COMPARISON_CONFIGURATION_POLICIES)
    manifest["allowed_observed_terminals"]["bare"]["solve"] = [
        "submit_result",
        "assistant_text",
        "final_recovery_code_interpreter",
    ]
    manifest_path.write_text(json.dumps(manifest) + "\n", encoding="utf-8")
    row.update(
        {
            "comparison_protocol_version": V24_COMPARISON_PROTOCOL_VERSION,
            "comparison_manifest_sha256": _sha256(manifest_path),
            "outcome_kind": "model_execution_failure",
            "passed": False,
            "error": "token budget exhausted",
            "error_type": "AgentExecutionFailure",
            "error_retryable": False,
            "error_category": "model_execution_failure",
            "model_failure_reason": "budget_exhausted",
        }
    )
    row["budget"]["used"]["total_tokens"] = 110
    row["budget"]["termination"] = {
        "reason": "max_total_tokens",
        "message": "token budget exhausted",
        "stage": "solve",
        "elapsed_seconds": 1.0,
    }
    row["agent"]["usage"]["total_tokens"] = 110
    row["agent"]["budget"] = row["budget"]
    row["agent"]["stages"][0]["agent"]["budget"] = row["budget"]
    for timing, tokens in zip(row["agent"]["request_timings"], (40, 40, 30), strict=True):
        timing["total_tokens"] = tokens
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is False
    assert _row_reason(summary, "model_execution_failure_reason_invalid")
    assert _row_reason(summary, "budget_used_invalid:total_tokens")
    assert manifest["configuration"]["model_execution_failure_reasons"] != (
        COMPARISON_CONFIGURATION_POLICIES["model_execution_failure_reasons"]
    )


@pytest.mark.parametrize(
    ("target", "expected_reason"),
    [
        ("aggregate_order", "agent_request_timings_mismatch"),
        ("stage_timing", "agent_request_timings_mismatch"),
    ],
)
def test_v25_audit_binds_aggregate_and_stage_request_timings(
    tmp_path: Path,
    target: str,
    expected_reason: str,
) -> None:
    results, task, row = _fixture(tmp_path)
    row = json.loads(json.dumps(row))
    if target == "aggregate_order":
        row["agent"]["request_timings"].reverse()
    else:
        row["agent"]["stages"][0]["agent"]["request_timings"][0]["turn"] = 99
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is False
    assert _row_reason(summary, expected_reason)


@pytest.mark.parametrize(
    ("target", "expected_reason"),
    [
        ("wrapper", "agent_stage_tool_trace_mismatch:solve"),
        ("nested", "agent_stage_tool_trace_mismatch:solve"),
        ("aggregate", "agent_tool_trace_mismatch"),
        ("names", "agent_stage_tool_names_mismatch:solve"),
        ("disallowed", "agent_stage_tool_not_allowed:solve"),
        ("count", "agent_tool_count_mismatch"),
    ],
)
def test_v25_audit_binds_tool_traces_and_counts(
    tmp_path: Path,
    target: str,
    expected_reason: str,
) -> None:
    results, task, row = _fixture(tmp_path)
    row = json.loads(json.dumps(row))
    stage = row["agent"]["stages"][0]
    if target == "wrapper":
        stage["tool_trace"][0]["ok"] = False
    elif target == "nested":
        stage["agent"]["tool_trace"][0]["ok"] = False
    elif target == "aggregate":
        row["agent"]["tool_trace"][0]["ok"] = False
    elif target == "names":
        stage["tool_name_trace"] = ["delete_rows", "code_interpreter"]
    elif target == "disallowed":
        disallowed = {"name": "delete_rows", "ok": True}
        stage["tool_trace"].append(disallowed)
        stage["agent"]["tool_trace"].append(dict(disallowed))
        stage["tool_name_trace"].append("delete_rows")
        stage["agent"]["tool_calls"] += 1
        stage["agent"]["function_calls_total"] += 1
        row["agent"]["tool_trace"].append({"stage": "solve", **disallowed})
        row["agent"]["tool_calls"] += 1
        row["agent"]["function_calls_total"] += 1
    else:
        row["agent"]["tool_calls"] = 99
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is False
    assert _row_reason(summary, expected_reason)


@pytest.mark.parametrize(
    ("target", "expected_reason"),
    [
        ("aggregate_usage", "agent_stage_usage_mismatch"),
        ("usage_arithmetic", "agent_stage_usage_mismatch"),
        ("stage_usage", "agent_stage_timing_usage_mismatch:solve"),
        ("stage_budget", "agent_stage_budget_mismatch:solve"),
    ],
)
def test_v25_audit_reconciles_usage_and_stage_budgets(
    tmp_path: Path,
    target: str,
    expected_reason: str,
) -> None:
    results, task, row = _fixture(tmp_path)
    row = json.loads(json.dumps(row))
    stage_agent = row["agent"]["stages"][0]["agent"]
    if target == "aggregate_usage":
        row["agent"]["usage"] = {
            "input_tokens": -999,
            "output_tokens": 1009,
            "total_tokens": 10,
        }
    elif target == "usage_arithmetic":
        row["agent"]["usage"] = {
            "input_tokens": 8,
            "output_tokens": 3,
            "total_tokens": 99,
        }
    elif target == "stage_usage":
        stage_agent["usage"]["input_tokens"] += 1
        stage_agent["usage"]["output_tokens"] -= 1
    else:
        stage_agent["budget"]["used"]["model_calls"] = 2
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is False
    assert _row_reason(summary, expected_reason)


def test_v24_audit_does_not_require_v25_exact_agent_evidence(tmp_path: Path) -> None:
    results, task, row = _fixture(tmp_path)
    manifest_path = results / "comparison-manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest["schema_version"] = V24_COMPARISON_MANIFEST_SCHEMA_VERSION
    manifest["comparison_protocol_version"] = V24_COMPARISON_PROTOCOL_VERSION
    manifest["configuration"].update(V24_COMPARISON_CONFIGURATION_POLICIES)
    manifest["allowed_observed_terminals"]["bare"]["solve"] = [
        "submit_result",
        "assistant_text",
        "final_recovery_code_interpreter",
    ]
    manifest_path.write_text(json.dumps(manifest) + "\n", encoding="utf-8")
    row = json.loads(json.dumps(row))
    row["comparison_protocol_version"] = V24_COMPARISON_PROTOCOL_VERSION
    row["comparison_manifest_sha256"] = _sha256(manifest_path)
    row["agent"]["request_timings"].reverse()
    row["agent"]["tool_calls"] = 99
    row["agent"]["stages"][0]["agent"]["budget"]["used"]["model_calls"] = 2
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is True


def test_v25_audit_preserves_schema14_exact_agent_contract(tmp_path: Path) -> None:
    results, task, row = _fixture(tmp_path)
    manifest_path = results / "comparison-manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest["schema_version"] = V25_COMPARISON_MANIFEST_SCHEMA_VERSION
    manifest["comparison_protocol_version"] = V25_COMPARISON_PROTOCOL_VERSION
    manifest["configuration"].update(V25_COMPARISON_CONFIGURATION_POLICIES)
    manifest["allowed_observed_terminals"] = _allowed_observed_terminals_policy(
        manifest["stage_turn_caps"],
        protocol_version=V25_COMPARISON_PROTOCOL_VERSION,
    )
    manifest["harness_source"]["files"][0]["sha256"] = "0" * 64
    combined = hashlib.sha256()
    for entry in manifest["harness_source"]["files"]:
        combined.update(entry["path"].encode("utf-8"))
        combined.update(b"\0")
        combined.update(entry["sha256"].encode("ascii"))
        combined.update(b"\n")
    manifest["harness_source"]["sha256"] = combined.hexdigest()
    manifest_path.write_text(json.dumps(manifest) + "\n", encoding="utf-8")
    row = json.loads(json.dumps(row))
    row["comparison_protocol_version"] = V25_COMPARISON_PROTOCOL_VERSION
    row["comparison_manifest_sha256"] = _sha256(manifest_path)
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")

    accepted = audit_comparison(results, [task])
    assert accepted["audit_valid"] is True

    row["agent"]["request_timings"].reverse()
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")
    rejected = audit_comparison(results, [task])
    assert rejected["audit_valid"] is False
    assert _row_reason(rejected, "agent_request_timings_mismatch")


def test_v26_audit_accepts_exact_truncated_terminal_evidence(tmp_path: Path) -> None:
    results, task, _ = _truncated_terminal_fixture(tmp_path)

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is True
    assert summary["rows"][0]["model_failure_reason"] == (
        "terminal_submission_truncated"
    )


def test_v26_audit_accepts_budget_precedence_with_truncated_response(
    tmp_path: Path,
) -> None:
    results, task, _ = _budget_truncated_terminal_fixture(tmp_path)

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is True
    assert summary["rows"][0]["model_failure_reason"] == "budget_exhausted"


@pytest.mark.parametrize(
    "target",
    [
        "aggregate_response_missing",
        "stage_response_missing",
        "response_usage",
        "termination_ceiling",
        "observed_terminal",
    ],
)
def test_v26_audit_rejects_tampered_budget_precedence_truncation(
    tmp_path: Path,
    target: str,
) -> None:
    results, task, row = _budget_truncated_terminal_fixture(tmp_path)
    row = json.loads(json.dumps(row))
    final_stage = row["agent"]["stages"][-1]
    if target == "aggregate_response_missing":
        row["agent"].pop("terminal_response")
    elif target == "stage_response_missing":
        final_stage["agent"].pop("terminal_response")
    elif target == "response_usage":
        row["agent"]["terminal_response"]["usage"]["total_tokens"] = 7
    elif target == "termination_ceiling":
        for budget in (
            row["budget"],
            row["agent"]["budget"],
            final_stage["agent"]["budget"],
        ):
            budget["limit"]["total_tokens"] = 10
    else:
        final_stage["observed_terminal_tool"] = "submit_result_length"
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is False
    assert _row_reason(summary, "terminal_submission_truncated_evidence_invalid")


def test_v26_audit_accepts_empty_ack_terminal_response(tmp_path: Path) -> None:
    results, task, _ = _fixture(tmp_path)

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is True


def test_v26_audit_rejects_evidence_result_for_bare_solve(tmp_path: Path) -> None:
    results, task, row = _fixture(tmp_path)
    row = json.loads(json.dumps(row))
    final_text = "stage evidence result"
    acknowledgement = {
        "mode": "evidence_result",
        "result_chars": len(final_text),
        "result_sha256": hashlib.sha256(final_text.encode("utf-8")).hexdigest(),
    }
    final_agent = row["agent"]["stages"][-1]["agent"]
    final_agent["final_text"] = final_text
    final_agent["terminal_response"]["acknowledgement"] = acknowledgement
    row["agent"]["final_text"] = final_text
    row["agent"]["terminal_response"] = json.loads(
        json.dumps(final_agent["terminal_response"])
    )
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is False
    assert _row_reason(summary, "accepted_terminal_response_evidence_invalid")


def test_v26_audit_accepts_stage_bound_paper_terminal_responses(
    tmp_path: Path,
) -> None:
    results, task, _ = _paper_scored_fixture(tmp_path)

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is True


@pytest.mark.parametrize(
    ("stage_name", "wrong_mode"),
    [
        ("extract", "empty_ack"),
        ("vision_verify", "empty_ack"),
        ("latex_verify", "empty_ack"),
        ("reconcile", "evidence_result"),
        ("solve", "evidence_result"),
    ],
)
def test_v26_audit_rejects_terminal_response_mode_for_wrong_paper_stage(
    tmp_path: Path,
    stage_name: str,
    wrong_mode: str,
) -> None:
    results, task, row = _paper_scored_fixture(tmp_path)
    row = json.loads(json.dumps(row))
    stage = next(
        item for item in row["agent"]["stages"] if item["name"] == stage_name
    )
    stage_agent = stage["agent"]
    response_id = stage_agent["response_id"]
    if wrong_mode == "empty_ack":
        stage_agent["final_text"] = "Spreadsheet task completed."
        stage_agent["terminal_response"]["acknowledgement"] = {}
    else:
        final_text = f"{stage_name} evidence"
        stage_agent["final_text"] = final_text
        stage_agent["terminal_response"] = {
            "status": "accepted",
            "response_id": response_id,
            "acknowledgement": {
                "mode": "evidence_result",
                "result_chars": len(final_text),
                "result_sha256": hashlib.sha256(
                    final_text.encode("utf-8")
                ).hexdigest(),
            },
        }
    if stage_name == "solve":
        row["agent"]["final_text"] = stage_agent["final_text"]
        row["agent"]["terminal_response"] = json.loads(
            json.dumps(stage_agent["terminal_response"])
        )
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is False
    assert _row_reason(summary, "accepted_terminal_response_evidence_invalid")


@pytest.mark.parametrize(
    "target",
    [
        "aggregate_response_missing",
        "stage_response_missing",
        "aggregate_response_tampered",
        "stage_response_tampered",
        "wrapper_response",
        "ack_text",
    ],
)
def test_v26_audit_rejects_tampered_accepted_terminal_response(
    tmp_path: Path,
    target: str,
) -> None:
    results, task, row = _fixture(tmp_path)
    row = json.loads(json.dumps(row))
    final_stage = row["agent"]["stages"][-1]
    final_agent = final_stage["agent"]
    if target == "aggregate_response_missing":
        row["agent"].pop("terminal_response")
    elif target == "stage_response_missing":
        final_agent.pop("terminal_response")
    elif target == "aggregate_response_tampered":
        row["agent"]["terminal_response"]["response_id"] = "other-response"
    elif target == "stage_response_tampered":
        final_agent["terminal_response"]["status"] = "truncated"
    elif target == "wrapper_response":
        final_stage["terminal_response"] = {"status": "accepted"}
    elif target == "ack_text":
        final_agent["final_text"] = "model-authored text"
        row["agent"]["final_text"] = "model-authored text"
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is False
    assert _row_reason(summary, "accepted_terminal_response_evidence_invalid")


@pytest.mark.parametrize(
    "target",
    ["evidence_chars", "evidence_chars_bool", "evidence_sha256", "evidence_extra_field"],
)
def test_v26_audit_rejects_tampered_paper_evidence_result(
    tmp_path: Path,
    target: str,
) -> None:
    results, task, row = _paper_scored_fixture(tmp_path)
    row = json.loads(json.dumps(row))
    stage_agent = row["agent"]["stages"][0]["agent"]
    acknowledgement = stage_agent["terminal_response"]["acknowledgement"]
    if target == "evidence_chars":
        acknowledgement["result_chars"] += 1
    elif target == "evidence_chars_bool":
        acknowledgement["result_chars"] = True
    elif target == "evidence_sha256":
        acknowledgement["result_sha256"] = "0" * 64
    else:
        acknowledgement["extra"] = True
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is False
    assert _row_reason(summary, "accepted_terminal_response_evidence_invalid")


@pytest.mark.parametrize(
    ("path", "value"),
    [
        (("model_failure_reason",), "terminal_submission_invalid"),
        (("api_protocol",), "responses"),
        (("agent", "terminal_response", "finish_reason"), "stop"),
        (("agent", "terminal_response", "response_id"), "other-response"),
        (("agent", "terminal_response", "timing", "status_code"), 201),
        (
            (
                "agent",
                "terminal_response",
                "timing",
                "attempt_history",
                0,
                "error_type",
            ),
            "ProviderError",
        ),
        (
            (
                "agent",
                "terminal_response",
                "timing",
                "attempt_history",
                0,
                "delivery_state",
            ),
            "headers_seen",
        ),
        (("agent", "terminal_response", "usage", "total_tokens"), 7),
        (("agent", "terminal_response", "discarded_message", "sha256"), "x" * 64),
        (("agent", "stages", 0, "terminal_tool"), None),
        (("agent", "stages", 0, "agent", "terminal_submissions"), 1),
        (("agent", "stages", 0, "agent", "observed_terminal_tool"), "submit_result"),
        (("agent", "stages", 0, "agent", "request_timings", 2, "turn"), 2),
        (
            ("agent", "stages", 0, "agent", "tool_trace"),
            [{"name": "submit_result", "ok": True}],
        ),
        (("agent", "stages", 0, "terminal_response"), {"status": "truncated"}),
    ],
)
def test_v26_audit_rejects_tampered_truncated_terminal_evidence(
    tmp_path: Path,
    path: tuple[str | int, ...],
    value: Any,
) -> None:
    results, task, row = _truncated_terminal_fixture(tmp_path)
    row = json.loads(json.dumps(row))
    target: Any = row
    for component in path[:-1]:
        target = target[component]
    target[path[-1]] = value
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is False
    assert _row_reason(summary, "terminal_submission_truncated_evidence_invalid")


def test_v26_audit_requires_derivable_forced_terminal_route_basis(
    tmp_path: Path,
) -> None:
    results, task, row = _truncated_terminal_fixture(tmp_path)
    row = json.loads(json.dumps(row))
    final_stage = row["agent"]["stages"][-1]
    final_stage["max_turns"] = 4
    final_stage["agent"]["budget"]["limit"]["model_calls"] = 4
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is False
    assert _row_reason(summary, "terminal_submission_truncated_evidence_invalid")


def test_v26_audit_rejects_truncation_before_last_stage(tmp_path: Path) -> None:
    results, task, row = _truncated_terminal_fixture(tmp_path)
    row = json.loads(json.dumps(row))
    trailing = json.loads(json.dumps(row["agent"]["stages"][0]))
    trailing["name"] = "unexpected-trailing-stage"
    trailing["agent"].pop("terminal_response")
    trailing["observed_terminal_tool"] = "submit_result"
    trailing["agent"]["observed_terminal_tool"] = "submit_result"
    row["agent"]["stages"].append(trailing)
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is False
    assert _row_reason(summary, "agent_stage_order_mismatch")


@pytest.mark.parametrize(
    ("field", "value", "expected_reason"),
    [
        ("passed", True, "model_execution_failure_not_failed"),
        ("error_category", "routing_protocol", "model_execution_failure_category_invalid"),
        ("model_failure_reason", "route_failed", "model_execution_failure_reason_invalid"),
        ("error_type", "AgentRoutingError", "model_execution_failure_type_invalid"),
        ("error_retryable", True, "model_execution_failure_retryable_invalid"),
        ("artifact_score_passed", False, "stored_artifact_score_passed_mismatch"),
    ],
)
def test_audit_rejects_tampered_model_execution_failure_taxonomy(
    tmp_path: Path,
    field: str,
    value: Any,
    expected_reason: str,
) -> None:
    results, task, row = _fixture(tmp_path)
    _set_final_model_execution_failure(row, "managed workbook remained unchanged")
    row.update(
        {
            "outcome_kind": "model_execution_failure",
            "passed": False,
            "error": "managed workbook remained unchanged",
            "error_type": "AgentExecutionFailure",
            "error_retryable": False,
            "error_category": "model_execution_failure",
            "model_failure_reason": "workbook_unchanged",
        }
    )
    row[field] = value
    (results / "results.jsonl").write_text(
        json.dumps(row) + "\n", encoding="utf-8"
    )

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is False
    assert _row_reason(summary, expected_reason)


def test_audit_rejects_ambiguous_inflight_marker(tmp_path: Path) -> None:
    results, task, _ = _fixture(tmp_path)
    (results / INFLIGHT_FILENAME).write_text("{}\n", encoding="utf-8")

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is False
    assert "ambiguous_inflight_arm_task" in summary["reasons"]


def test_audit_accepts_exact_interrupted_seal_as_integral_but_incomplete(
    tmp_path: Path,
) -> None:
    results, task, _ = _fixture(tmp_path)
    (results / "results.jsonl").write_text("", encoding="utf-8")
    _write_interrupted_seals(results, [_interrupted_seal(results, task)])
    before = _tree_hashes(tmp_path)

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is True
    assert summary["journal_integrity_valid"] is True
    assert summary["study_complete"] is False
    assert summary["inference_valid"] is False
    assert summary["inference_invalid_reasons"] == ["interrupted_unknown_outcome"]
    assert summary["interrupted_arm_tasks"] == 1
    assert summary["interrupted_arm_task_keys"] == ["task-1::bare"]
    assert summary["known_passed_rows"] == 0
    assert summary["known_failed_rows"] == 0
    assert summary["mcnemar_exact_p"] is None
    assert summary["stratified_bootstrap_95"] is None
    assert summary["holm_adjusted_p"] is None
    assert summary["rows"][0]["journal_integrity_valid"] is True
    assert summary["rows"][0]["audit_valid"] is False
    assert summary["rows"][0]["outcome_observed"] is False
    assert "missing_result_row" not in summary["rows"][0]["reasons"]
    assert summary["reasons"] == ["interrupted_unknown_outcome:task-1::bare"]
    assert _tree_hashes(tmp_path) == before


def test_audit_allows_missing_empty_journal_when_every_key_is_sealed(
    tmp_path: Path,
) -> None:
    results, task, _ = _fixture(tmp_path)
    (results / "results.jsonl").unlink()
    _write_interrupted_seals(results, [_interrupted_seal(results, task)])

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is True
    assert summary["journal_integrity_valid"] is True
    assert summary["study_complete"] is False
    assert "results_file_missing" not in summary["reasons"]


def test_audit_binds_result_to_exact_continuation_source(tmp_path: Path) -> None:
    results, task, row = _fixture(tmp_path)
    continuation = _continuation_source(results)
    manifest = json.loads(
        (results / "comparison-manifest.json").read_text(encoding="utf-8")
    )
    manifest["repository_source"] = continuation["repository_source"]
    (results / "comparison-manifest.json").write_text(
        json.dumps(manifest) + "\n", encoding="utf-8"
    )
    row["comparison_manifest_sha256"] = _sha256(
        results / "comparison-manifest.json"
    )
    continuation = _continuation_source(results)
    _write_continuation_source(results, continuation)
    row["continuation_source"] = continuation
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is True
    assert summary["continuation_source"] == continuation
    assert summary["continuation_source_file_sha256"] == _sha256(
        results / CONTINUATION_SOURCE_FILENAME
    )


def test_audit_rejects_manifest_continuation_repository_source_mismatch(
    tmp_path: Path,
) -> None:
    results, task, row = _fixture(tmp_path)
    manifest_path = results / "comparison-manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    continuation = _continuation_source(results)
    manifest["repository_source"] = {
        **continuation["repository_source"],
        "git_commit": "3" * 40,
        "remote_tracking_commit": "3" * 40,
        "remote_observed_commit": "3" * 40,
    }
    manifest_path.write_text(json.dumps(manifest) + "\n", encoding="utf-8")
    row["comparison_manifest_sha256"] = _sha256(manifest_path)
    continuation = _continuation_source(results)
    _write_continuation_source(results, continuation)
    row["continuation_source"] = continuation
    (results / "results.jsonl").write_text(
        json.dumps(row) + "\n", encoding="utf-8"
    )

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is False
    assert "continuation_source_invalid" in summary["reasons"]


def test_audit_rejects_current_manifest_for_registered_split_not_bound_to_git(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    results, task, _ = _fixture(tmp_path)
    manifest_path = results / "comparison-manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest["split_provenance"] = {
        "manifest_id": "qwen35-trace2skill-local-v27-reserve79-v1"
    }
    manifest["repository_source"] = _continuation_source(results)[
        "repository_source"
    ]
    manifest_path.write_text(json.dumps(manifest) + "\n", encoding="utf-8")
    monkeypatch.setattr(
        "spreadsheet_harness.audit._current_repository_git_identity",
        lambda: {
            "git_commit": "3" * 40,
            "git_tree": "4" * 40,
            "remote_tracking_commit": "3" * 40,
        },
    )

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is False
    assert "comparison_manifest_repository_checkout_mismatch" in summary["reasons"]


@pytest.mark.parametrize(
    ("target", "field", "value"),
    [
        ("record", "comparison_manifest_sha256", "0" * 64),
        ("record", "record_sha256", "0" * 64),
        ("repository", "git_commit", "A" * 40),
        ("repository", "git_tree", "short"),
        ("repository", "remote_tracking_ref", "refs/remotes/origin/dev"),
        ("repository", "remote_tracking_commit", "3" * 40),
        ("repository", "remote_name", "upstream"),
        ("repository", "remote_ref", "refs/heads/dev"),
        ("repository", "remote_observed_commit", "3" * 40),
        ("repository", "source_fingerprint", {"sha256": "0" * 64, "files": []}),
    ],
)
def test_audit_rejects_invalid_continuation_source(
    tmp_path: Path,
    target: str,
    field: str,
    value: Any,
) -> None:
    results, task, row = _fixture(tmp_path)
    continuation = _continuation_source(results)
    if target == "record":
        continuation[field] = value
    else:
        continuation["repository_source"][field] = value
    _write_continuation_source(results, continuation)
    row["continuation_source"] = continuation
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")

    summary = audit_comparison(results, [task])

    assert summary["journal_integrity_valid"] is False
    assert "continuation_source_invalid" in summary["reasons"]


def test_audit_rejects_result_continuation_source_mismatch(tmp_path: Path) -> None:
    results, task, row = _fixture(tmp_path)
    continuation = _continuation_source(results)
    _write_continuation_source(results, continuation)
    row["continuation_source"] = {**continuation, "record_sha256": "0" * 64}
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")

    summary = audit_comparison(results, [task])

    assert summary["journal_integrity_valid"] is False
    assert "result_continuation_source_mismatch:1" in summary["reasons"]


def test_audit_rejects_row_continuation_without_sidecar(tmp_path: Path) -> None:
    results, task, row = _fixture(tmp_path)
    row["continuation_source"] = _continuation_source(results)
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")

    summary = audit_comparison(results, [task])

    assert summary["journal_integrity_valid"] is False
    assert "result_continuation_source_without_record:1" in summary["reasons"]


@pytest.mark.parametrize(
    ("field", "value"),
    [
        ("passed", False),
        ("score_available", True),
        ("replay_permitted", True),
        ("sealed_at", "not-a-timestamp"),
        ("sealed_from_inflight_marker_sha256", "A" * 64),
        ("task_id", "outside-frozen-tasks"),
        ("arm", "ours"),
    ],
)
def test_audit_rejects_malformed_interrupted_seal(
    tmp_path: Path,
    field: str,
    value: Any,
) -> None:
    results, task, _ = _fixture(tmp_path)
    seal = _interrupted_seal(results, task)
    seal[field] = value
    _write_interrupted_seals(results, [seal])

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is False
    assert summary["journal_integrity_valid"] is False
    assert summary["study_complete"] is False
    assert summary["inference_valid"] is False
    assert "interrupted_arm_task_seals_invalid" in summary["reasons"]
    assert summary["interrupted_arm_tasks"] == 0


def test_audit_rejects_duplicate_interrupted_seal_keys(tmp_path: Path) -> None:
    results, task, _ = _fixture(tmp_path)
    seal = _interrupted_seal(results, task)
    _write_interrupted_seals(results, [seal, dict(seal)])

    summary = audit_comparison(results, [task])

    assert summary["journal_integrity_valid"] is False
    assert "interrupted_arm_task_seals_invalid" in summary["reasons"]


def test_audit_rejects_result_row_conflicting_with_interrupted_seal(
    tmp_path: Path,
) -> None:
    results, task, _ = _fixture(tmp_path)
    _write_interrupted_seals(results, [_interrupted_seal(results, task)])

    summary = audit_comparison(results, [task])

    assert summary["journal_integrity_valid"] is False
    assert "result_row_conflicts_with_interrupted_seal:task-1::bare" in summary[
        "reasons"
    ]


def test_audit_rejects_split_provenance_tampering(tmp_path: Path) -> None:
    results, task, row = _fixture(tmp_path)
    manifest_path = results / "comparison-manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    task_ids = manifest["task_ids"]
    provenance = {
        "manifest_id": "qwen35-trace2skill-local-unattempted-pilot16-v2",
        "schema_version": "spreadsheetbench-trace2skill-derivative-v2",
        "manifest_sha256": (
            "f29d6e5627161b355c24acfbda6c5dcc250d12b5f4933d3c3fb0c50a8bac39b3"
        ),
        "task_count": len(task_ids),
        "task_ids_sha256": _text_sha256(
            "".join(f"{task_id}\n" for task_id in task_ids)
        ),
        "dataset_json_sha256": manifest["dataset_manifest_sha256"],
    }
    manifest["split_provenance"] = provenance
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")
    row["comparison_manifest_sha256"] = _sha256(manifest_path)
    row["split_provenance"] = {**provenance, "manifest_sha256": "2" * 64}
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is False
    assert _row_reason(summary, "row_manifest_mismatch:split_provenance")


def test_audit_rejects_split_provenance_task_order_mismatch(tmp_path: Path) -> None:
    results, task, row = _fixture(tmp_path)
    manifest_path = results / "comparison-manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest["split_provenance"] = {
        "manifest_id": "pilot-v2",
        "schema_version": "derivative-v2",
        "manifest_sha256": "1" * 64,
        "task_count": 1,
        "task_ids_sha256": "2" * 64,
        "dataset_json_sha256": manifest["dataset_manifest_sha256"],
    }
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")
    row["comparison_manifest_sha256"] = _sha256(manifest_path)
    row["split_provenance"] = manifest["split_provenance"]
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is False
    assert "comparison_manifest_split_provenance_invalid" in summary["reasons"]


def test_audit_rejects_malformed_split_provenance_without_raising(
    tmp_path: Path,
) -> None:
    results, task, row = _fixture(tmp_path)
    manifest_path = results / "comparison-manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest["split_provenance"] = {
        "manifest_id": [],
        "schema_version": "spreadsheetbench-trace2skill-derivative-v2",
        "manifest_sha256": "1" * 64,
        "task_count": 1,
        "task_ids_sha256": _text_sha256("task-1\n"),
        "dataset_json_sha256": manifest["dataset_manifest_sha256"],
    }
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")
    row["comparison_manifest_sha256"] = _sha256(manifest_path)
    row["split_provenance"] = manifest["split_provenance"]
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is False
    assert "comparison_manifest_split_provenance_invalid" in summary["reasons"]


@pytest.mark.parametrize(
    ("target", "field", "value", "expected_reason"),
    [
        (
            "manifest",
            "schema_version",
            9,
            "comparison_manifest_schema_mismatch",
        ),
        (
            "manifest",
            "comparison_protocol_version",
            "resource_matched_multi_arm_v19",
            "comparison_manifest_protocol_mismatch",
        ),
        (
            "row",
            "comparison_protocol_version",
            "resource_matched_multi_arm_v19",
            "result_protocol_mismatch:1",
        ),
    ],
)
def test_audit_rejects_protocol_provenance_mismatch(
    tmp_path: Path,
    target: str,
    field: str,
    value: Any,
    expected_reason: str,
) -> None:
    results, task, row = _fixture(tmp_path)
    path = (
        results / "comparison-manifest.json"
        if target == "manifest"
        else results / "results.jsonl"
    )
    if target == "manifest":
        payload = json.loads(path.read_text(encoding="utf-8"))
        payload[field] = value
    else:
        payload = dict(row)
        payload[field] = value
    path.write_text(json.dumps(payload) + ("\n" if target == "row" else ""), encoding="utf-8")

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is False
    assert expected_reason in summary["reasons"]


def test_audit_rejects_duplicate_manifest_json_key(tmp_path: Path) -> None:
    results, task, _ = _fixture(tmp_path)
    path = results / "comparison-manifest.json"
    raw = path.read_text(encoding="utf-8")
    path.write_text(
        raw[:-1] + ',"schema_version":12}',
        encoding="utf-8",
    )

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is False
    assert "comparison_manifest_invalid" in summary["reasons"]


def test_audit_rejects_duplicate_result_json_key(tmp_path: Path) -> None:
    results, task, _ = _fixture(tmp_path)
    path = results / "results.jsonl"
    raw = path.read_text(encoding="utf-8").rstrip("\n")
    path.write_text(raw[:-1] + ',"task_id":"task-1"}\n', encoding="utf-8")

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is False
    assert "invalid_jsonl_line:1" in summary["reasons"]


@pytest.mark.parametrize(
    ("filename", "expected_reason"),
    [
        ("comparison-manifest.json", "comparison_manifest_invalid"),
        ("results.jsonl", "results_file_unreadable"),
    ],
)
def test_audit_rejects_symlinked_core_document(
    tmp_path: Path,
    filename: str,
    expected_reason: str,
) -> None:
    results, task, _ = _fixture(tmp_path)
    path = results / filename
    target = tmp_path / f"real-{filename}"
    path.rename(target)
    path.symlink_to(target)

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is False
    assert expected_reason in summary["reasons"]


@pytest.mark.parametrize(
    ("target", "field", "value", "expected_reason"),
    [
        ("row", "comparison_manifest_sha256", "0" * 64, "manifest_sha256_binding_mismatch"),
        ("row", "model", "different-model", "row_manifest_mismatch:model"),
        ("row", "generation", {"temperature": 0.5}, "row_manifest_mismatch:generation"),
        ("row", "stage_turn_caps", {"solve": 99}, "row_manifest_mismatch:stage_turn_caps"),
        ("row", "calculation_backend", "libreoffice", "row_manifest_mismatch:calculation_backend"),
        (
            "row",
            "artifact_score_passed",
            False,
            "stored_artifact_score_passed_mismatch",
        ),
        ("agent", "arm", "ours", "agent_arm_mismatch"),
        ("stage", "observed_forced_tool_prefix", [], "agent_observed_prefix_mismatch:solve"),
        ("stage", "observed_terminal_tool", "unknown", "agent_observed_terminal_invalid:solve"),
        ("timing", "attempt_history", [], "request_attempt_audit_inexact"),
    ],
)
def test_audit_rejects_tampered_resource_and_routing_evidence(
    tmp_path: Path,
    target: str,
    field: str,
    value: Any,
    expected_reason: str,
) -> None:
    results, task, row = _fixture(tmp_path)
    if target == "row":
        row[field] = value
    elif target == "agent":
        row["agent"][field] = value
    elif target == "stage":
        row["agent"]["stages"][0][field] = value
    else:
        row["agent"]["request_timings"][0][field] = value
        row["agent"]["stages"][0]["agent"]["request_timings"][0][field] = value
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is False
    assert _row_reason(summary, expected_reason)


@pytest.mark.parametrize("field", ["configuration", "harness_source", "runtime"])
def test_audit_requires_frozen_manifest_provenance(
    tmp_path: Path,
    field: str,
) -> None:
    results, task, _ = _fixture(tmp_path)
    path = results / "comparison-manifest.json"
    manifest = json.loads(path.read_text(encoding="utf-8"))
    manifest.pop(field)
    path.write_text(json.dumps(manifest) + "\n", encoding="utf-8")

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is False
    expected_fragment = "source_fingerprint" if field == "harness_source" else field
    assert any(expected_fragment in reason for reason in summary["reasons"])


def test_v28_audit_requires_manifest_source_to_match_active_checkout(
    tmp_path: Path,
) -> None:
    results, task, row = _fixture(tmp_path)
    path = results / "comparison-manifest.json"
    manifest = json.loads(path.read_text(encoding="utf-8"))
    manifest["harness_source"]["files"][0]["sha256"] = "0" * 64
    combined = hashlib.sha256()
    for entry in manifest["harness_source"]["files"]:
        combined.update(entry["path"].encode("utf-8"))
        combined.update(b"\0")
        combined.update(entry["sha256"].encode("ascii"))
        combined.update(b"\n")
    manifest["harness_source"]["sha256"] = combined.hexdigest()
    path.write_text(json.dumps(manifest) + "\n", encoding="utf-8")
    row["comparison_manifest_sha256"] = _sha256(path)
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is False
    assert "comparison_manifest_source_checkout_mismatch" in summary["reasons"]


def test_v26_audit_accepts_historical_source_fingerprint(tmp_path: Path) -> None:
    results, task, row = _fixture(tmp_path)
    path = results / "comparison-manifest.json"
    manifest = json.loads(path.read_text(encoding="utf-8"))
    manifest["schema_version"] = V26_COMPARISON_MANIFEST_SCHEMA_VERSION
    manifest["comparison_protocol_version"] = V26_COMPARISON_PROTOCOL_VERSION
    manifest["configuration"].update(V26_COMPARISON_CONFIGURATION_POLICIES)
    manifest["allowed_observed_terminals"] = _allowed_observed_terminals_policy(
        manifest["stage_turn_caps"],
        protocol_version=V26_COMPARISON_PROTOCOL_VERSION,
    )
    manifest["stage_allowed_tools"] = _stage_allowed_tools_policy(
        tuple(manifest["arms"]),
        protocol_version=V26_COMPARISON_PROTOCOL_VERSION,
    )
    manifest["harness_source"]["files"][0]["sha256"] = "0" * 64
    combined = hashlib.sha256()
    for entry in manifest["harness_source"]["files"]:
        combined.update(entry["path"].encode("utf-8"))
        combined.update(b"\0")
        combined.update(entry["sha256"].encode("ascii"))
        combined.update(b"\n")
    manifest["harness_source"]["sha256"] = combined.hexdigest()
    path.write_text(json.dumps(manifest) + "\n", encoding="utf-8")
    row["comparison_protocol_version"] = V26_COMPARISON_PROTOCOL_VERSION
    row["comparison_manifest_sha256"] = _sha256(path)
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is True


def test_v27_audit_accepts_historical_source_fingerprint(tmp_path: Path) -> None:
    results, task, row = _fixture(tmp_path)
    path = results / "comparison-manifest.json"
    manifest = json.loads(path.read_text(encoding="utf-8"))
    manifest["schema_version"] = V27_COMPARISON_MANIFEST_SCHEMA_VERSION
    manifest["comparison_protocol_version"] = V27_COMPARISON_PROTOCOL_VERSION
    manifest["configuration"].update(V27_COMPARISON_CONFIGURATION_POLICIES)
    manifest["allowed_observed_terminals"] = _allowed_observed_terminals_policy(
        manifest["stage_turn_caps"],
        protocol_version=V27_COMPARISON_PROTOCOL_VERSION,
    )
    manifest["stage_allowed_tools"] = _stage_allowed_tools_policy(
        tuple(manifest["arms"]),
        protocol_version=V27_COMPARISON_PROTOCOL_VERSION,
    )
    manifest["harness_source"]["files"][0]["sha256"] = "0" * 64
    combined = hashlib.sha256()
    for entry in manifest["harness_source"]["files"]:
        combined.update(entry["path"].encode("utf-8"))
        combined.update(b"\0")
        combined.update(entry["sha256"].encode("ascii"))
        combined.update(b"\n")
    manifest["harness_source"]["sha256"] = combined.hexdigest()
    path.write_text(json.dumps(manifest) + "\n", encoding="utf-8")
    row["comparison_protocol_version"] = V27_COMPARISON_PROTOCOL_VERSION
    row["comparison_manifest_sha256"] = _sha256(path)
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is True


def test_audit_rejects_tampered_stored_passed(tmp_path: Path) -> None:
    results, task, row = _fixture(tmp_path)
    row["passed"] = False
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is False
    assert _row_reason(summary, "stored_passed_mismatch")


def test_audit_rejects_tampered_workbook(tmp_path: Path) -> None:
    results, task, row = _fixture(tmp_path)
    output = Path(row["output_workbook"])
    workbook = load_workbook(output)
    workbook["Sheet1"]["A1"] = 99
    workbook.save(output)
    workbook.close()

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is False
    assert _row_reason(summary, "artifact_hash_mismatch")
    assert _row_reason(summary, "stored_passed_mismatch")
    assert _row_reason(summary, "stored_comparison_mismatch")


@pytest.mark.parametrize("path_field", ["run_dir", "output_workbook"])
def test_audit_rejects_artifact_path_outside_managed_arm(
    tmp_path: Path, path_field: str
) -> None:
    results, task, row = _fixture(tmp_path)
    outside_run = tmp_path / "outside"
    outside_output = outside_run / "artifacts" / "output.xlsx"
    _book(outside_output, 42)
    if path_field == "run_dir":
        row[path_field] = str(outside_run)
    else:
        row[path_field] = str(outside_output)
    (results / "results.jsonl").write_text(json.dumps(row) + "\n", encoding="utf-8")

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is False
    expected = (
        "run_dir_outside_expected_arm"
        if path_field == "run_dir"
        else "output_path_not_managed_artifact"
    )
    assert _row_reason(summary, expected)


def test_audit_rejects_duplicate_result_row(tmp_path: Path) -> None:
    results, task, row = _fixture(tmp_path)
    (results / "results.jsonl").write_text(
        json.dumps(row) + "\n" + json.dumps(row) + "\n", encoding="utf-8"
    )

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is False
    assert summary["observed_rows"] == 2
    assert _row_reason(summary, "duplicate_result_rows")


def test_audit_rejects_missing_result_row(tmp_path: Path) -> None:
    results, task, _ = _fixture(tmp_path)
    (results / "results.jsonl").write_text("", encoding="utf-8")

    summary = audit_comparison(results, [task])

    assert summary["audit_valid"] is False
    assert summary["expected_rows"] == 1
    assert summary["observed_rows"] == 0
    assert _row_reason(summary, "missing_result_row")
