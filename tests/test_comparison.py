from __future__ import annotations

import hashlib
import json
import shutil
from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook, load_workbook

import spreadsheet_harness.comparison as comparison_module
import spreadsheet_harness.render as render_module
from spreadsheet_harness import benchmark as benchmark_module
from spreadsheet_harness import cli as cli_module
from spreadsheet_harness.agent import AgentResult
from spreadsheet_harness.arms import PaperStageValidationError
from spreadsheet_harness.benchmark import SpreadsheetTask, compare_workbooks
from spreadsheet_harness.comparison import (
    AVAILABLE_COMPARISON_ARMS,
    COMPARISON_ARMS,
    COMPARISON_MANIFEST_SCHEMA_VERSION,
    COMPARISON_PROTOCOL_VERSION,
    LEGACY_COMPARISON_PROTOCOL_VERSION,
    RUN_SPEC_ANCHORS,
    V24_COMPARISON_MANIFEST_SCHEMA_VERSION,
    V24_COMPARISON_PROTOCOL_VERSION,
    V25_COMPARISON_MANIFEST_SCHEMA_VERSION,
    V25_COMPARISON_PROTOCOL_VERSION,
    V25_RUN_SPEC_SOURCE_CONTRACT,
    V26_COMPARISON_MANIFEST_SCHEMA_VERSION,
    V26_COMPARISON_PROTOCOL_VERSION,
    V26_RUN_SPEC_SOURCE_CONTRACT,
    V27_COMPARISON_CONFIGURATION_POLICIES,
    V27_COMPARISON_MANIFEST_SCHEMA_VERSION,
    V27_COMPARISON_PROTOCOL_VERSION,
    V27_RUN_SPEC_SOURCE_CONTRACT,
    ComparisonBenchmarkRunner,
    RunSpecAnchor,
    _allowed_observed_terminals_policy,
    _arm_order,
    _balanced_arm_orders,
    _stage_allowed_tools_policy,
    comparison_summary,
    load_pilot_run_spec,
    manifest_execution_contract,
    require_launchable_run_spec,
    verify_repository_source_state,
)
from spreadsheet_harness.config import ProviderConfig
from spreadsheet_harness.errors import (
    AgentBudgetError,
    AgentExecutionFailure,
    AgentRoutingError,
    CodeIsolationError,
    HarnessError,
    ProviderError,
    ScoringInfrastructureError,
)
from spreadsheet_harness.skills import SkillRegistry
from spreadsheet_harness.trajectory import read_trajectory


def _book(path: Path, value: object = 1) -> None:
    workbook = Workbook()
    workbook.active["A1"] = value
    workbook.save(path)
    workbook.close()


def _tasks(tmp_path: Path) -> list[SpreadsheetTask]:
    initial = tmp_path / "initial.xlsx"
    golden = tmp_path / "golden.xlsx"
    _book(initial)
    _book(golden)
    return [
        SpreadsheetTask(
            "cell-1",
            "cell secret instruction",
            initial,
            golden,
            "Cell-Level Manipulation",
            "TOP_SECRET_CELL_RANGE",
            "TOP_SECRET_CELL_SHEET",
        ),
        SpreadsheetTask(
            "sheet-1",
            "sheet secret instruction",
            initial,
            golden,
            "Sheet-Level Manipulation",
            "TOP_SECRET_SHEET_RANGE",
            "TOP_SECRET_SHEET_NAME",
        ),
    ]


def test_arm_order_is_deterministic_rotation() -> None:
    first = _arm_order("task-a", 7, COMPARISON_ARMS)
    second = _arm_order("task-a", 7, COMPARISON_ARMS)

    assert first == second
    assert set(first) == set(COMPARISON_ARMS)
    assert len(first) == len(COMPARISON_ARMS)


def test_balanced_arm_orders_counterbalance_every_position() -> None:
    task_ids = [f"task-{index}" for index in range(6)]
    orders = _balanced_arm_orders(task_ids, 20260811, COMPARISON_ARMS)

    assert orders == _balanced_arm_orders(task_ids, 20260811, COMPARISON_ARMS)
    for position in range(3):
        counts = {arm: 0 for arm in COMPARISON_ARMS}
        for order in orders.values():
            counts[order[position]] += 1
        assert counts == {arm: 2 for arm in COMPARISON_ARMS}


def test_comparison_manifest_hides_answer_metadata(tmp_path: Path) -> None:
    tasks = _tasks(tmp_path)
    runner = ComparisonBenchmarkRunner(
        ProviderConfig(
            "https://example.test/v1",
            "not-a-real-key",
            "test-model",
            max_retries=3,
            request_interval_seconds=20.0,
            litellm_timeout_seconds=600,
        ),
        tmp_path / "comparison",
        skill_registry=SkillRegistry([]),
    )

    manifest = runner._manifest(tasks)
    encoded = json.dumps(manifest)

    assert manifest["task_count"] == 2
    assert manifest["schema_version"] == 17
    assert COMPARISON_MANIFEST_SCHEMA_VERSION == 17
    assert COMPARISON_PROTOCOL_VERSION == "resource_matched_multi_arm_v28"
    assert manifest["comparison_protocol_version"] == COMPARISON_PROTOCOL_VERSION
    assert manifest["configuration"]["code_workbook_formula_gate"] == (
        "rollback-new-invalid-a1-or-high-confidence-unprefixed-formula-text-v2"
    )
    assert manifest["configuration"]["failed_edit_recovery_policy"] == (
        "shared_state_based_recovery_v1"
    )
    assert manifest["configuration"]["spreadsheet_skill_policy"] == (
        "pre-evaluation-baseline-frozen-v1"
    )
    assert manifest["configuration"]["edit_recovery_prompt_policy"] == (
        "self-contained-request-scoped-verification-v1"
    )
    assert manifest["configuration"]["terminal_submission_policy"] == (
        "empty-ack-harness-final-text-v1"
    )
    assert manifest["configuration"]["edit_recovery_terminal_policy"] == (
        "penultimate-recovery-final-submit-v1"
    )
    assert manifest["configuration"]["ours_tool_policy"] == (
        "fixed-six-code-first-v1"
    )
    assert manifest["configuration"]["deterministic_profile_policy"] == (
        "representative-evidence-12k-v1"
    )
    assert manifest["configuration"]["formula_verification_skill_policy"] == (
        "trajectory-local-transfer-gate-v1"
    )
    assert manifest["configuration"]["recalculation_integrity_policy"] == (
        "exact-ordered-sheet-kind-name-visibility-v2"
    )
    assert manifest["configuration"]["recalculation_failure_policy"] == (
        "audited-infrastructure-error-no-score-v1"
    )
    assert manifest["configuration"]["artifact_reopen_policy"] == (
        "ooxml-inventory-plus-worksheet-only-openpyxl-view-v1"
    )
    assert manifest["configuration"]["scoring_compatibility_policy"] == (
        "worksheet-only-ooxml-view-scorer-infrastructure-no-score-v1"
    )
    assert "artifact_reopen_policy" not in V27_COMPARISON_CONFIGURATION_POLICIES
    assert "scoring_compatibility_policy" not in V27_COMPARISON_CONFIGURATION_POLICIES
    assert manifest["arms"] == list(COMPARISON_ARMS)
    assert manifest["arm_display_names"] == {
        "bare": "bare",
        "paper": "paper-inspired",
        "ours": "ours",
    }
    assert manifest["forced_tool_prefix_routing"] == {
        "bare": {"solve": ["code_interpreter", "code_interpreter"]},
        "paper": {
            "extract": ["list_sheets", "inspect_range"],
            "vision_verify": ["render_workbook", "view_image"],
            "latex_verify": ["range_to_latex"],
            "reconcile": [],
            "solve": ["code_interpreter", "code_interpreter"],
        },
        "ours": {"solve": ["code_interpreter", "code_interpreter"]},
    }
    assert manifest["post_prefix_routing"] == {
        "tool_choice": "auto",
        "terminal_tool": "submit_result",
        "applies_to": "comparison stages with workbook tools after forced prefix",
        "direct_text_stages": ["paper.reconcile"],
    }
    assert manifest["stage_allowed_tools"] == {
        "bare": {"solve": ["code_interpreter"]},
        "paper": {
            "extract": ["inspect_range", "list_sheets"],
            "vision_verify": ["render_workbook", "view_image"],
            "latex_verify": ["range_to_latex"],
            "reconcile": [],
            "solve": ["code_interpreter"],
        },
        "ours": {
            "solve": [
                "code_interpreter",
                "fill_formula",
                "inspect_range",
                "recalculate_and_read",
                "render_workbook",
                "view_image",
            ]
        },
    }
    assert _stage_allowed_tools_policy(
        ("ours",), protocol_version=V25_COMPARISON_PROTOCOL_VERSION
    ) == {"ours": {"solve": "all"}}
    assert manifest["allowed_observed_terminals"]["paper"]["reconcile"] == [
        "assistant_text",
        "budget_exhausted",
    ]
    assert manifest["allowed_observed_terminals"]["ours"]["solve"] == [
        "submit_result",
        "submit_result_length",
        "budget_exhausted",
    ]
    assert manifest["allowed_observed_terminals"]["bare"]["solve"] == [
        "submit_result",
        "submit_result_length",
        "budget_exhausted",
    ]
    assert manifest["allowed_observed_terminals"]["paper"]["solve"] == [
        "submit_result",
        "submit_result_length",
        "budget_exhausted",
    ]
    assert manifest["forced_prefix_wire_policy"] == {
        "tool_choice": "explicit_function",
        "available_tools": "forced tool only",
        "terminal_tool_available": False,
    }
    assert manifest["stage_turn_caps"] == {
        "bare": {"solve": 20},
        "paper": {
            "extract": 6,
            "vision_verify": 3,
            "latex_verify": 3,
            "reconcile": 1,
            "solve": 7,
        },
        "ours": {"solve": 20},
    }
    assert manifest["turn_cap_policy"]["version"] == "per_arm_turn_cap_v2"
    assert manifest["turn_cap_policy"]["max_turns_per_arm"] == 20
    assert manifest["turn_cap_policy"]["paper_scaling_version"] == (
        "constrained_largest_remainder_v1"
    )
    assert manifest["deterministic_profile"]["enabled"] is True
    assert manifest["deterministic_profile"]["consumed_by_arms"] == ["ours"]
    assert set(manifest["deterministic_profile"]["task_profile_sha256"]) == {
        "cell-1",
        "sheet-1",
    }
    assert "TOP_SECRET" not in encoded
    assert all("scoring_metadata_sha256" in task for task in manifest["tasks"])
    assert manifest["configuration"]["circuit_breaker_threshold"] == 3
    assert manifest["configuration"]["circuit_breaker_threshold_categories"] == [
        "provider_transient",
        "routing_protocol",
    ]
    assert manifest["configuration"]["circuit_breaker_immediate_categories"] == [
        "provider_fatal",
        "recalculation_infrastructure",
    ]
    assert manifest["configuration"]["overload_retry_min_seconds"] == 15.0
    assert manifest["configuration"]["connect_retry_min_seconds"] == 30.0
    assert manifest["configuration"]["request_interval_seconds"] == 20.0
    assert manifest["configuration"]["litellm_timeout_seconds"] == 600.0
    assert manifest["configuration"]["max_turns_per_arm"] == 20
    assert manifest["configuration"]["request_pacing_policy"] == (
        "process_local_min_attempt_start_interval_v1"
    )
    assert manifest["configuration"]["request_pacing_scope"] == (
        "comparison_runner_process"
    )
    assert manifest["configuration"]["request_pacing_retries_included"] is True
    assert manifest["configuration"]["request_pacing_first_attempt_immediate"] is True
    assert manifest["configuration"]["automatic_retry_policy"] == (
        "delivery-aware-allowlist-v1"
    )
    assert manifest["configuration"]["safe_retry_http_statuses"] == [425, 429, 503]
    assert manifest["configuration"]["safe_automatic_retry_reasons"] == [
        "connect_error",
        "connect_timeout",
        "explicit_overload",
        "http_425",
        "http_429",
        "http_503",
        "pool_timeout",
    ]
    assert manifest["configuration"]["capacity_retry_delay_policy"] == (
        "max-valid-retry-after-and-overload-min-then-global-cap"
    )
    assert manifest["configuration"]["retry_backoff_max_seconds"] == 60.0
    assert manifest["configuration"]["read_timeout_policy"] == "fail-closed-no-replay"
    assert manifest["configuration"]["http_408_policy"] == "fail-closed-no-replay"
    assert manifest["configuration"]["stream_interruption_policy"] == (
        "fail-closed-no-replay"
    )
    assert manifest["configuration"]["request_attempt_telemetry"] == (
        "delivery-safe-retry-ids-headers-backoff-pacing-v4"
    )
    assert manifest["configuration"]["result_manifest_binding_policy"] == (
        "exact-manifest-sha256-v1"
    )
    assert manifest["configuration"]["resume_journal_policy"] == (
        "durable-inflight-fail-closed-no-replay-v3"
    )
    assert manifest["configuration"]["model_execution_failure_policy"] == (
        "known-false-score-artifact-and-request-audited-nonbreaker-v1"
    )
    assert manifest["configuration"]["model_execution_failure_reasons"] == [
        "budget_exhausted",
        "edit_recovery_exhausted",
        "terminal_submission_invalid",
        "terminal_submission_truncated",
        "workbook_unchanged",
    ]
    assert manifest["hidden_from_models"] == [
        "instruction_type",
        "answer_position",
        "answer_sheet",
        "golden_path",
    ]


@pytest.mark.parametrize(
    "protocol_version",
    [
        LEGACY_COMPARISON_PROTOCOL_VERSION,
        V24_COMPARISON_PROTOCOL_VERSION,
        V25_COMPARISON_PROTOCOL_VERSION,
    ],
)
def test_historical_terminal_policies_retain_tool_stage_text_fallback(
    protocol_version: str,
) -> None:
    policy = _allowed_observed_terminals_policy(
        {
            "bare": {"solve": 2},
            "paper": {"reconcile": 1, "solve": 2},
        },
        protocol_version=protocol_version,
    )

    assert "assistant_text" in policy["bare"]["solve"]
    assert "assistant_text" in policy["paper"]["solve"]
    assert policy["paper"]["reconcile"][0] == "assistant_text"


@pytest.mark.parametrize(
    "protocol_version",
    [
        V26_COMPARISON_PROTOCOL_VERSION,
        V27_COMPARISON_PROTOCOL_VERSION,
        COMPARISON_PROTOCOL_VERSION,
    ],
)
def test_v26_and_later_terminal_policy_require_submit_for_tool_stages(
    protocol_version: str,
) -> None:
    policy = _allowed_observed_terminals_policy(
        {
            "bare": {"solve": 2},
            "paper": {"reconcile": 1, "solve": 2},
        },
        protocol_version=protocol_version,
    )

    assert policy["bare"]["solve"] == [
        "submit_result",
        "submit_result_length",
        "budget_exhausted",
    ]
    assert policy["paper"]["solve"] == [
        "submit_result",
        "submit_result_length",
        "budget_exhausted",
    ]
    assert policy["paper"]["reconcile"] == [
        "assistant_text",
        "budget_exhausted",
    ]


def test_optional_ablation_arms_are_available_without_changing_default_manifest(
    tmp_path: Path,
) -> None:
    tasks = _tasks(tmp_path)
    assert COMPARISON_ARMS == ("bare", "paper", "ours")
    assert set(AVAILABLE_COMPARISON_ARMS) == {
        "bare",
        "profile",
        "native",
        "paper",
        "ours",
    }
    runner = ComparisonBenchmarkRunner(
        ProviderConfig("https://example.test/v1", "key", "model"),
        tmp_path / "ablations",
        skill_registry=SkillRegistry([]),
        arms=("bare", "profile", "native", "ours"),
    )

    manifest = runner._manifest(tasks)

    assert manifest["arms"] == ["bare", "profile", "native", "ours"]
    contract = manifest["deterministic_profile"]
    assert contract["enabled"] is True
    assert contract["consumed_by_arms"] == ["profile", "ours"]
    assert contract["schema_version"] == "deterministic-workbook-profile-v1"
    assert contract["task_independent"] is True
    assert contract["model_calls"] == 0
    assert set(contract["task_profile_sha256"]) == {"cell-1", "sheet-1"}
    assert all(len(value) == 64 for value in contract["task_profile_sha256"].values())
    assert manifest["stage_turn_caps"]["profile"] == {"solve": 20}
    assert manifest["stage_turn_caps"]["native"] == {"solve": 20}


def test_comparison_manifest_records_custom_turn_caps_and_zero_pacing(
    tmp_path: Path,
) -> None:
    runner = ComparisonBenchmarkRunner(
        ProviderConfig(
            "http://localhost:8000/v1",
            "EMPTY",
            "Qwen/Qwen3.5-35B-A3B",
            request_interval_seconds=0,
        ),
        tmp_path / "qwen-100-turns",
        skill_registry=SkillRegistry([]),
        arms=("bare", "paper", "ours"),
        max_model_calls=100,
        max_turns_per_arm=100,
    )

    manifest = runner._manifest(_tasks(tmp_path))

    assert manifest["configuration"]["request_interval_seconds"] == 0.0
    assert manifest["configuration"]["max_model_calls"] == 100
    assert manifest["configuration"]["max_turns_per_arm"] == 100
    assert manifest["stage_turn_caps"] == {
        "bare": {"solve": 100},
        "paper": {
            "extract": 30,
            "vision_verify": 15,
            "latex_verify": 15,
            "reconcile": 5,
            "solve": 35,
        },
        "ours": {"solve": 100},
    }


def test_comparison_manifest_records_and_locks_generation_controls(tmp_path: Path) -> None:
    tasks = _tasks(tmp_path)
    output = tmp_path / "generation-lock"
    first = ComparisonBenchmarkRunner(
        ProviderConfig(
            "https://example.test/v1",
            "not-a-real-key",
            "test-model",
            api_protocol="chat-completions",
            seed=41,
            temperature=1.0,
            litellm_timeout_seconds=600,
        ),
        output,
        skill_registry=SkillRegistry([]),
    )
    first._prepare_manifest(tasks)

    manifest = json.loads(first.manifest_path.read_text(encoding="utf-8"))
    assert manifest["configuration"]["api_protocol"] == "chat-completions"
    assert manifest["configuration"]["litellm_timeout_seconds"] == 600.0
    assert manifest["configuration"]["generation"] == {
        "temperature": 1.0,
        "seed": 41,
    }

    second = ComparisonBenchmarkRunner(
        ProviderConfig(
            "https://example.test/v1",
            "not-a-real-key",
            "test-model",
            api_protocol="chat-completions",
            seed=42,
            temperature=1.0,
            litellm_timeout_seconds=600,
        ),
        output,
        skill_registry=SkillRegistry([]),
    )
    with pytest.raises(HarnessError, match="different frozen config"):
        second._prepare_manifest(tasks)


def test_comparison_manifest_and_rows_bind_split_provenance(tmp_path: Path) -> None:
    provenance = {
        "manifest_id": "test-derivative-split-v2",
        "schema_version": "spreadsheetbench-trace2skill-derivative-v2",
        "manifest_sha256": (
            "f29d6e5627161b355c24acfbda6c5dcc250d12b5f4933d3c3fb0c50a8bac39b3"
        ),
        "task_count": 2,
        "task_ids_sha256": hashlib.sha256(b"cell-1\nsheet-1\n").hexdigest(),
        "dataset_json_sha256": "3" * 64,
    }
    runner = ComparisonBenchmarkRunner(
        ProviderConfig("https://example.test/v1", "key", "model"),
        tmp_path / "split-binding",
        skill_registry=SkillRegistry([]),
        split_provenance=provenance,
    )

    tasks = _tasks(tmp_path)
    monkeypatch_provenance = provenance.copy()
    with pytest.MonkeyPatch.context() as monkeypatch:
        monkeypatch.setattr(
            "spreadsheet_harness.comparison.verify_trace2skill_split_provenance",
            lambda value: value == monkeypatch_provenance,
        )
        monkeypatch.setattr(
            "spreadsheet_harness.comparison._dataset_manifest_sha256",
            lambda _: "3" * 64,
        )
        manifest = runner._manifest(tasks)

    assert manifest["split_provenance"] == provenance
    provenance["manifest_id"] = "mutated-after-construction"
    assert manifest["split_provenance"]["manifest_id"] == "test-derivative-split-v2"


def test_comparison_manifest_rejects_invalid_split_provenance_before_write(
    tmp_path: Path,
) -> None:
    tasks = _tasks(tmp_path)
    runner = ComparisonBenchmarkRunner(
        ProviderConfig("https://example.test/v1", "key", "model"),
        tmp_path / "invalid-split-preflight",
        skill_registry=SkillRegistry([]),
        split_provenance={
            "manifest_id": "attacker-split",
            "schema_version": "attacker-v1",
            "manifest_sha256": "1" * 64,
            "task_count": 2,
            "task_ids_sha256": hashlib.sha256(b"cell-1\nsheet-1\n").hexdigest(),
            "dataset_json_sha256": "3" * 64,
        },
    )

    with pytest.raises(HarnessError, match="frozen split provenance"):
        runner._prepare_manifest(tasks)

    assert not runner.manifest_path.exists()


def test_split_manifest_rejects_offset_or_limit(monkeypatch: Any, tmp_path: Path) -> None:
    split = tmp_path / "split.json"
    split.write_text(json.dumps({"task_ids": []}), encoding="utf-8")
    parser = cli_module.build_parser()
    monkeypatch.setattr(
        cli_module,
        "download_verified",
        lambda _: tmp_path / "dataset",
    )
    monkeypatch.setattr(cli_module, "load_verified_tasks", lambda _: [])
    monkeypatch.setattr(
        cli_module,
        "load_and_verify_trace2skill_split_manifest",
        lambda *_: {"valid": True, "task_ids": []},
    )

    for selector in (["--offset", "1"], ["--limit", "1"]):
        args = parser.parse_args(
            ["benchmark", "compare", "--split-manifest", str(split), *selector]
        )
        with pytest.raises(HarnessError, match="derivative manifest"):
            cli_module.cmd_benchmark_compare(args)


@pytest.mark.parametrize(
    "selector",
    [["--task-id", "cell-1"], ["--task-id-file", "ids.txt"]],
)
def test_split_manifest_rejects_task_id_selectors(
    monkeypatch: Any,
    tmp_path: Path,
    selector: list[str],
) -> None:
    split = tmp_path / "split.json"
    split.write_text(json.dumps({"task_ids": ["cell-1"]}), encoding="utf-8")
    (tmp_path / "ids.txt").write_text("cell-1\n", encoding="utf-8")
    parser = cli_module.build_parser()
    tasks = _tasks(tmp_path)
    monkeypatch.chdir(tmp_path)
    monkeypatch.setattr(cli_module, "download_verified", lambda _: tmp_path / "dataset")
    monkeypatch.setattr(cli_module, "load_verified_tasks", lambda _: tasks)
    monkeypatch.setattr(
        cli_module,
        "load_and_verify_trace2skill_split_manifest",
        lambda *_: {"valid": True, "task_ids": ["cell-1"]},
    )
    args = parser.parse_args(
        ["benchmark", "compare", "--split-manifest", str(split), *selector]
    )

    with pytest.raises(HarnessError, match="derivative manifest"):
        cli_module.cmd_benchmark_compare(args)


def test_comparison_uses_verified_manifest_order_without_rereading(
    monkeypatch: Any,
    tmp_path: Path,
) -> None:
    split = tmp_path / "split.json"
    split.write_text("this is deliberately not JSON", encoding="utf-8")
    tasks = _tasks(tmp_path)
    parser = cli_module.build_parser()
    monkeypatch.setattr(cli_module, "download_verified", lambda _: tmp_path / "dataset")
    monkeypatch.setattr(cli_module, "load_verified_tasks", lambda _: tasks)
    monkeypatch.setattr(
        cli_module,
        "load_and_verify_trace2skill_split_manifest",
        lambda *_: {
            "valid": True,
            "manifest_id": "test-split",
            "schema_version": "test-split-v1",
            "manifest_sha256": "1" * 64,
            "usable_tasks": 2,
            "task_ids": ["sheet-1", "cell-1"],
            "task_ids_sha256": "2" * 64,
            "dataset_json_sha256": "3" * 64,
        },
    )
    monkeypatch.setattr(
        cli_module,
        "trace2skill_split_provenance",
        lambda report: {
            "manifest_id": report["manifest_id"],
            "schema_version": report["schema_version"],
            "manifest_sha256": report["manifest_sha256"],
            "task_count": report["usable_tasks"],
            "task_ids_sha256": report["task_ids_sha256"],
            "dataset_json_sha256": report["dataset_json_sha256"],
        },
    )
    captured: dict[str, Any] = {}

    class FakeRunner:
        def __init__(self, *_: Any, **kwargs: Any) -> None:
            captured["split_provenance"] = kwargs["split_provenance"]

        def run(self, selected: list[SpreadsheetTask]) -> dict[str, Any]:
            captured["task_ids"] = [task.task_id for task in selected]
            return {
                "missing_arm_tasks": 0,
                "arms": {"bare": {"errors": 0}, "ours": {"errors": 0}},
            }

    monkeypatch.setattr(cli_module, "ComparisonBenchmarkRunner", FakeRunner)
    monkeypatch.setattr(
        cli_module,
        "_provider",
        lambda _: ProviderConfig("https://example.test/v1", "key", "model"),
    )
    args = parser.parse_args(
        [
            "benchmark",
            "compare",
            "--split-manifest",
            str(split),
            "--output",
            str(tmp_path / "output"),
        ]
    )

    assert cli_module.cmd_benchmark_compare(args) == 0
    assert captured["task_ids"] == ["sheet-1", "cell-1"]
    assert captured["split_provenance"] == {
        "manifest_id": "test-split",
        "schema_version": "test-split-v1",
        "manifest_sha256": "1" * 64,
        "task_count": 2,
        "task_ids_sha256": "2" * 64,
        "dataset_json_sha256": "3" * 64,
    }


def test_comparison_rejects_unreachable_turn_ceiling(tmp_path: Path) -> None:
    with pytest.raises(ValueError, match="max_model_calls must be at least"):
        ComparisonBenchmarkRunner(
            ProviderConfig("https://example.test/v1", "key", "model"),
            tmp_path / "unreachable-turns",
            skill_registry=SkillRegistry([]),
            max_model_calls=20,
            max_turns_per_arm=100,
        )


def _pilot_run_spec_runner(
    tmp_path: Path,
    monkeypatch: Any,
    *,
    split_provenance: dict[str, Any] | None = None,
    spec_filename: str = "qwen35-trace2skill-local-pilot16-run-spec-v1.json",
) -> tuple[ComparisonBenchmarkRunner, list[SpreadsheetTask]]:
    source = _tasks(tmp_path)[0]
    spec_path = Path("benchmarks/protocols") / spec_filename
    document, provenance, raw = load_pilot_run_spec(spec_path)
    resources = document["execution"]["resources"]
    if split_provenance is None:
        split_provenance = document["execution"]["split_provenance"]
    monkeypatch.setattr(
        "spreadsheet_harness.comparison.verify_trace2skill_split_provenance",
        lambda _: True,
    )
    monkeypatch.setattr(
        "spreadsheet_harness.comparison._dataset_manifest_sha256",
        lambda _: split_provenance["dataset_json_sha256"],
    )
    monkeypatch.setattr(
        "spreadsheet_harness.comparison.verify_repository_source_state",
        lambda: {
            "schema_version": 1,
            "git_commit": "1" * 40,
            "git_tree": "2" * 40,
            "remote_tracking_ref": "refs/remotes/origin/main",
            "remote_tracking_commit": "1" * 40,
            "source_fingerprint": {"sha256": "3" * 64, "files": []},
        },
    )
    split_path = Path(document["repository_relative_paths"]["split_manifest"])
    pilot_ids = json.loads(split_path.read_text(encoding="utf-8"))["task_ids"]
    tasks = [
        SpreadsheetTask(
            task_id,
            source.instruction,
            source.input_path,
            source.golden_path,
            source.instruction_type,
            source.answer_position,
            source.answer_sheet,
        )
        for task_id in pilot_ids
    ]
    runner = ComparisonBenchmarkRunner(
        ProviderConfig(
            "http://101.37.174.109:8010/v1",
            "not-a-real-key",
            "qwen36-35b-a3b",
            reasoning_effort="none",
            timeout_seconds=700,
            max_retries=0,
            request_interval_seconds=0,
            temperature=1,
            top_p=1,
            seed=41,
            presence_penalty=2,
            top_k=40,
            min_p=0,
            repetition_penalty=1,
            enable_thinking=False,
            api_protocol="chat-completions",
            litellm_timeout_seconds=600,
        ),
        tmp_path / "pilot-output",
        skill_registry=SkillRegistry([Path("skills")]),
        arms=("bare", "ours"),
        max_model_calls=resources["max_model_calls"],
        max_turns_per_arm=resources["max_turns_per_arm"],
        max_total_tokens=resources["max_total_tokens"],
        max_output_tokens=resources["max_output_tokens_per_call"],
        task_timeout_seconds=resources["task_timeout_seconds"],
        recalculate=resources["recalculate"],
        arm_order_seed=resources["arm_order_seed"],
        circuit_breaker_threshold=resources["circuit_breaker_threshold"],
        split_provenance=split_provenance,
        run_spec_document=document,
        run_spec_provenance=provenance,
        run_spec_bytes=raw,
    )
    return runner, tasks


def _generic_journal_runner(
    tmp_path: Path,
    monkeypatch: Any,
) -> tuple[ComparisonBenchmarkRunner, list[SpreadsheetTask]]:
    tasks = _tasks(tmp_path)[:1]
    runner = ComparisonBenchmarkRunner(
        ProviderConfig("https://example.test/v1", "not-a-real-key", "test-model"),
        tmp_path / "journal-output",
        skill_registry=SkillRegistry([]),
        arms=("bare", "ours"),
        max_model_calls=8,
        max_turns_per_arm=8,
        max_total_tokens=120000,
        max_output_tokens=4096,
        task_timeout_seconds=1200,
        recalculate=False,
        arm_order_seed=20260812,
    )
    monkeypatch.setattr(
        "spreadsheet_harness.comparison.ensure_strict_code_isolation",
        lambda *_args, **_kwargs: {},
    )
    return runner, tasks


def test_pilot_run_spec_rejects_noncanonical_bytes(tmp_path: Path) -> None:
    source = Path(
        "benchmarks/protocols/qwen35-trace2skill-local-pilot16-run-spec-v1.json"
    )
    target = tmp_path / source.name
    target.write_bytes(source.read_bytes() + b"\n")

    with pytest.raises(HarnessError, match="checksum"):
        load_pilot_run_spec(target)


def test_legacy_run_spec_is_parseable_but_never_launchable() -> None:
    _, provenance, _ = load_pilot_run_spec(
        Path("benchmarks/protocols/qwen35-trace2skill-local-pilot16-run-spec-v1.json")
    )

    with pytest.raises(HarnessError, match="read-only"):
        require_launchable_run_spec(provenance)


def test_v24_run_spec_is_historical_and_read_only() -> None:
    document, provenance, _ = load_pilot_run_spec(
        Path("benchmarks/protocols/qwen35-trace2skill-local-postopt16-run-spec-v1.json")
    )

    with pytest.raises(HarnessError, match="read-only"):
        require_launchable_run_spec(provenance)

    anchor = next(
        candidate
        for candidate in RUN_SPEC_ANCHORS
        if candidate.run_spec_id == document["run_spec_id"]
    )

    assert anchor == RunSpecAnchor(
        run_spec_id=document["run_spec_id"],
        filename="qwen35-trace2skill-local-postopt16-run-spec-v1.json",
        sha256=provenance["run_spec_sha256"],
        schema_version=document["schema_version"],
        phase="post_optimization_evaluation",
        split_manifest_id="qwen35-trace2skill-local-postopt16-v1",
        comparison_protocol_version=V24_COMPARISON_PROTOCOL_VERSION,
        comparison_manifest_schema_version=V24_COMPARISON_MANIFEST_SCHEMA_VERSION,
        launchable=False,
    )


def test_v25_confirmation_run_spec_is_historical_and_read_only() -> None:
    document, provenance, _ = load_pilot_run_spec(
        Path("benchmarks/protocols/qwen35-trace2skill-local-confirm16-run-spec-v1.json")
    )

    with pytest.raises(HarnessError, match="read-only"):
        require_launchable_run_spec(provenance)
    anchor = next(
        candidate
        for candidate in RUN_SPEC_ANCHORS
        if candidate.run_spec_id == document["run_spec_id"]
    )

    assert anchor == RunSpecAnchor(
        run_spec_id=document["run_spec_id"],
        filename="qwen35-trace2skill-local-confirm16-run-spec-v1.json",
        sha256="61ec4d37d0548e1be63ebf8619feb591d98ca78d7dce4d9d573886498ca74984",
        schema_version=document["schema_version"],
        phase="post_optimization_confirmation",
        split_manifest_id="qwen35-trace2skill-local-confirm16-v1",
        comparison_protocol_version=V25_COMPARISON_PROTOCOL_VERSION,
        comparison_manifest_schema_version=V25_COMPARISON_MANIFEST_SCHEMA_VERSION,
        launchable=False,
    )
    with pytest.raises(HarnessError, match="read-only"):
        require_launchable_run_spec(provenance, resume=True)


def test_v25_source_contract_remains_pinned_to_historical_source() -> None:
    document, _, _ = load_pilot_run_spec(
        Path("benchmarks/protocols/qwen35-trace2skill-local-confirm16-run-spec-v1.json")
    )

    assert document["execution"]["source_contract"] == V25_RUN_SPEC_SOURCE_CONTRACT
    historical_manifest = {
        "schema_version": V25_COMPARISON_MANIFEST_SCHEMA_VERSION,
        "comparison_protocol_version": V25_COMPARISON_PROTOCOL_VERSION,
        "configuration": {},
    }
    assert manifest_execution_contract(historical_manifest)["source_contract"] == (
        V25_RUN_SPEC_SOURCE_CONTRACT
    )
    assert document["execution"]["source_contract"] != (
        benchmark_module._run_spec_source_fingerprint()
    )


def test_v26_confirmation_run_spec_is_historical_and_read_only() -> None:
    path = Path(
        "benchmarks/protocols/"
        "qwen35-trace2skill-local-v26-confirm16-run-spec-v1.json"
    )
    document, provenance, _ = load_pilot_run_spec(path)

    with pytest.raises(HarnessError, match="read-only"):
        require_launchable_run_spec(provenance)

    anchor = next(
        candidate
        for candidate in RUN_SPEC_ANCHORS
        if candidate.run_spec_id == document["run_spec_id"]
    )

    assert anchor == RunSpecAnchor(
        run_spec_id="qwen36-local-v26-confirm-eval16-v1-bare-ours-seed41",
        filename=path.name,
        sha256="4bca7fe452c9ba2dadc31c374f29abcda575cb243e5f960789f2f50b4191884a",
        schema_version=document["schema_version"],
        phase="v26_post_optimization_confirmation",
        split_manifest_id="qwen35-trace2skill-local-v26-confirm16-v1",
        comparison_protocol_version=V26_COMPARISON_PROTOCOL_VERSION,
        comparison_manifest_schema_version=V26_COMPARISON_MANIFEST_SCHEMA_VERSION,
        launchable=False,
    )
    assert document["execution"]["source_contract"] == V26_RUN_SPEC_SOURCE_CONTRACT
    assert document["execution"]["resources"] == {
        "max_model_calls": 12,
        "max_turns_per_arm": 12,
        "max_total_tokens": 180000,
        "max_output_tokens_per_call": 4096,
        "task_timeout_seconds": 1800.0,
        "recalculate": True,
        "task_retries": 0,
        "circuit_breaker_threshold": 3,
        "arm_order_seed": 20260812,
    }
    with pytest.raises(HarnessError, match="read-only"):
        require_launchable_run_spec(provenance, resume=True)


def test_v26_source_contract_remains_pinned_to_historical_source() -> None:
    historical_manifest = {
        "schema_version": V26_COMPARISON_MANIFEST_SCHEMA_VERSION,
        "comparison_protocol_version": V26_COMPARISON_PROTOCOL_VERSION,
        "configuration": {},
    }

    assert manifest_execution_contract(historical_manifest)["source_contract"] == (
        V26_RUN_SPEC_SOURCE_CONTRACT
    )
    assert V26_RUN_SPEC_SOURCE_CONTRACT != (
        benchmark_module._run_spec_source_fingerprint()
    )


def test_v27_reserve79_run_spec_anchor_is_historical_and_read_only() -> None:
    path = Path(
        "benchmarks/protocols/"
        "qwen35-trace2skill-local-v27-reserve79-run-spec-v1.json"
    )
    document, provenance, _ = load_pilot_run_spec(path)
    with pytest.raises(HarnessError, match="read-only"):
        require_launchable_run_spec(provenance)

    anchor = next(
        candidate
        for candidate in RUN_SPEC_ANCHORS
        if candidate.run_spec_id == document["run_spec_id"]
    )

    assert anchor == RunSpecAnchor(
        run_spec_id="qwen36-local-v27-reserve79-eval-v1-bare-ours-seed41",
        filename=path.name,
        sha256="748fd0458e9b2c20adf5161fc9471e4f29421faecd5b4e02bdfa6b32b9342371",
        schema_version=document["schema_version"],
        phase="v27_reserve79_evaluation",
        split_manifest_id="qwen35-trace2skill-local-v27-reserve79-v1",
        comparison_protocol_version=V27_COMPARISON_PROTOCOL_VERSION,
        comparison_manifest_schema_version=V27_COMPARISON_MANIFEST_SCHEMA_VERSION,
        launchable=False,
    )
    assert document["execution"]["source_contract"] == V27_RUN_SPEC_SOURCE_CONTRACT
    with pytest.raises(HarnessError, match="read-only"):
        require_launchable_run_spec(provenance, resume=True)


def test_v27_source_contract_remains_pinned_to_historical_source() -> None:
    document, _, _ = load_pilot_run_spec(
        Path(
            "benchmarks/protocols/"
            "qwen35-trace2skill-local-v27-reserve79-run-spec-v1.json"
        )
    )
    historical_manifest = {
        "schema_version": V27_COMPARISON_MANIFEST_SCHEMA_VERSION,
        "comparison_protocol_version": V27_COMPARISON_PROTOCOL_VERSION,
        "configuration": {},
    }

    assert document["execution"]["source_contract"] == V27_RUN_SPEC_SOURCE_CONTRACT
    assert manifest_execution_contract(historical_manifest)["source_contract"] == (
        V27_RUN_SPEC_SOURCE_CONTRACT
    )
    assert V27_RUN_SPEC_SOURCE_CONTRACT != (
        benchmark_module._run_spec_source_fingerprint()
    )


def test_v25_preflight_rejects_historical_run_before_isolation(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    runner, tasks = _pilot_run_spec_runner(
        tmp_path,
        monkeypatch,
        spec_filename="qwen35-trace2skill-local-confirm16-run-spec-v1.json",
    )
    events: list[str] = []
    monkeypatch.setattr(
        "spreadsheet_harness.comparison.ensure_strict_code_isolation",
        lambda *_args, **_kwargs: events.append("isolation") or {},
    )

    with pytest.raises(HarnessError, match="read-only"):
        runner.preflight(tasks)

    assert events == []
    assert not runner.output_dir.exists()


def test_pilot_split_requires_canonical_run_spec_in_runner(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    document, _, _ = load_pilot_run_spec(
        Path("benchmarks/protocols/qwen35-trace2skill-local-pilot16-run-spec-v1.json")
    )
    runner, tasks = _pilot_run_spec_runner(tmp_path, monkeypatch)
    runner.run_spec_document = None
    runner.run_spec_provenance = None
    runner.run_spec_bytes = None

    with pytest.raises(HarnessError, match="requires its code-anchored run spec"):
        runner._manifest(tasks)


def test_comparison_preflight_rejects_direct_protected_task_bypass(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    source = _tasks(tmp_path)[0]
    task = SpreadsheetTask(
        "33157",
        source.instruction,
        source.input_path,
        source.golden_path,
        source.instruction_type,
        source.answer_position,
        source.answer_sheet,
    )
    runner = ComparisonBenchmarkRunner(
        ProviderConfig("https://example.test/v1", "not-a-real-key", "test-model"),
        tmp_path / "direct-protected-bypass",
        skill_registry=SkillRegistry([]),
        arms=("bare", "ours"),
        max_model_calls=8,
        max_turns_per_arm=8,
    )
    monkeypatch.setattr(
        "spreadsheet_harness.comparison.ensure_strict_code_isolation",
        lambda *_args, **_kwargs: {},
    )

    with pytest.raises(HarnessError, match="Protected evaluation task IDs"):
        runner.preflight([task])

    assert not runner.manifest_path.exists()


def test_repository_source_state_requires_clean_head_at_local_origin_main(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    calls: list[tuple[str, ...]] = []
    outputs = iter(
        [
            str(tmp_path),
            "1" * 40,
            "2" * 40,
            "1" * 40,
            "",
            f"{'1' * 40}\trefs/heads/main",
        ]
    )

    def fake_run(command: list[str], **_: Any) -> Any:
        calls.append(tuple(command[-2:]))
        return type("Completed", (), {"returncode": 0, "stdout": next(outputs)})()

    monkeypatch.setattr("spreadsheet_harness.comparison.subprocess.run", fake_run)
    monkeypatch.setattr(
        "spreadsheet_harness.comparison._source_fingerprint",
        lambda: {"sha256": "3" * 64, "files": []},
    )

    state = verify_repository_source_state(tmp_path)

    assert state["git_commit"] == "1" * 40
    assert state["git_tree"] == "2" * 40
    assert state["remote_tracking_commit"] == state["git_commit"]
    assert state["remote_observed_commit"] == state["git_commit"]
    assert calls[-1] == ("origin", "refs/heads/main")


def test_repository_source_state_rejects_remote_head_drift(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    outputs = iter(
        [
            str(tmp_path),
            "1" * 40,
            "2" * 40,
            "1" * 40,
            "",
            f"{'4' * 40}\trefs/heads/main",
        ]
    )
    monkeypatch.setattr(
        "spreadsheet_harness.comparison.subprocess.run",
        lambda *_args, **_kwargs: type(
            "Completed", (), {"returncode": 0, "stdout": next(outputs)}
        )(),
    )

    with pytest.raises(HarnessError, match="observed origin/main"):
        verify_repository_source_state(tmp_path)


def test_repository_source_state_rejects_remote_query_failure(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    outputs = iter([str(tmp_path), "1" * 40, "2" * 40, "1" * 40, ""])

    def fake_run(command: list[str], **kwargs: Any) -> Any:
        assert kwargs["env"]["GIT_TERMINAL_PROMPT"] == "0"
        if "ls-remote" in command:
            assert kwargs["timeout"] == 30
            return type("Completed", (), {"returncode": 2, "stdout": ""})()
        return type("Completed", (), {"returncode": 0, "stdout": next(outputs)})()

    monkeypatch.setattr("spreadsheet_harness.comparison.subprocess.run", fake_run)

    with pytest.raises(HarnessError, match="repository source state"):
        verify_repository_source_state(tmp_path)


def test_historical_v24_preflight_rejects_before_source_or_isolation(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    runner, tasks = _pilot_run_spec_runner(
        tmp_path,
        monkeypatch,
        spec_filename="qwen35-trace2skill-local-postopt16-run-spec-v1.json",
    )
    events: list[str] = []
    monkeypatch.setattr(
        "spreadsheet_harness.comparison.verify_repository_source_state",
        lambda: events.append("source") or {"schema_version": 1},
    )
    monkeypatch.setattr(
        "spreadsheet_harness.comparison.ensure_strict_code_isolation",
        lambda *_args, **_kwargs: events.append("isolation") or {},
    )

    with pytest.raises(HarnessError, match="read-only"):
        runner.preflight(tasks)

    assert events == []
    assert not runner.output_dir.exists()


def test_pilot_run_spec_runner_rejects_resolved_contract_mismatch(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    document, _, _ = load_pilot_run_spec(
        Path("benchmarks/protocols/qwen35-trace2skill-local-pilot16-run-spec-v1.json")
    )
    runner, tasks = _pilot_run_spec_runner(
        tmp_path,
        monkeypatch,
        split_provenance=document["execution"]["split_provenance"],
    )
    monkeypatch.setattr(
        "spreadsheet_harness.comparison.verify_trace2skill_split_provenance",
        lambda _: True,
    )
    monkeypatch.setattr(
        "spreadsheet_harness.comparison._dataset_manifest_sha256",
        lambda _: document["execution"]["split_provenance"]["dataset_json_sha256"],
    )
    pilot_ids = json.loads(
        Path(
            "benchmarks/protocols/"
            "qwen35-trace2skill-local-unattempted-pilot16-v2.json"
        ).read_text(encoding="utf-8")
    )["task_ids"]
    source = tasks[0]
    tasks = [
        SpreadsheetTask(
            task_id,
            source.instruction,
            source.input_path,
            source.golden_path,
            source.instruction_type,
            source.answer_position,
            source.answer_sheet,
        )
        for task_id in pilot_ids
    ]
    runner.max_total_tokens += 1

    with pytest.raises(HarnessError, match="execution contract"):
        runner._manifest(tasks)

    assert not runner.output_dir.exists()


def test_comparison_crash_marker_blocks_ambiguous_replay(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    runner, tasks = _generic_journal_runner(tmp_path, monkeypatch)
    monkeypatch.setattr(
        "spreadsheet_harness.comparison.ensure_strict_code_isolation",
        lambda *_args, **_kwargs: {},
    )
    monkeypatch.setattr(
        runner,
        "_run_one",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(SystemExit("crash")),
    )

    with pytest.raises(SystemExit, match="crash"):
        runner.run(tasks, resume=False)

    marker = json.loads(runner.inflight_path.read_text(encoding="utf-8"))
    assert marker["task_id"] == tasks[0].task_id
    assert marker["run_spec_provenance"] == runner.run_spec_provenance

    with pytest.raises(HarnessError, match="ambiguous in-flight"):
        runner.run(tasks, resume=True)


def test_comparison_seals_interrupted_marker_without_replay(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    runner, tasks = _generic_journal_runner(tmp_path, monkeypatch)
    provider_calls: list[str] = []
    monkeypatch.setattr(
        "spreadsheet_harness.comparison.ensure_strict_code_isolation",
        lambda *_args, **_kwargs: {},
    )
    runner.preflight(tasks)
    runner._prepare_manifest(tasks)
    manifest_sha = hashlib.sha256(runner.manifest_path.read_bytes()).hexdigest()
    runner._write_inflight(
        tasks[0].task_id,
        "ours",
        comparison_manifest_sha256=manifest_sha,
    )
    marker_bytes = runner.inflight_path.read_bytes()
    monkeypatch.setattr(
        "spreadsheet_harness.comparison.run_arm",
        lambda *_args, **_kwargs: provider_calls.append("run_arm"),
    )

    row = runner.seal_interrupted_inflight(tasks)

    assert row["status"] == "interrupted"
    assert row["passed"] is None
    assert row["error_category"] == "interrupted_unknown_outcome"
    assert row["replay_permitted"] is False
    assert not runner.inflight_path.exists()
    seals = json.loads(runner.interrupted_seals_path.read_text(encoding="utf-8"))
    assert seals == {"schema_version": 1, "seals": [row]}
    assert not runner.results_path.exists()
    assert not runner.continuation_source_path.exists()
    assert runner.continuation_source_record is None
    assert provider_calls == []
    assert runner.seal_interrupted_inflight(tasks) == row

    runner.inflight_path.write_bytes(marker_bytes)
    assert runner.seal_interrupted_inflight(tasks) == row
    assert not runner.inflight_path.exists()

    sampled: list[tuple[str, str]] = []

    def stop_after_one(task: SpreadsheetTask, arm: str, **_: Any) -> dict[str, Any]:
        sampled.append((task.task_id, arm))
        return {
            "task_id": task.task_id,
            "arm": arm,
            "comparison_protocol_version": COMPARISON_PROTOCOL_VERSION,
            "comparison_manifest_sha256": manifest_sha,
            "split_provenance": runner.split_provenance,
            "run_spec_provenance": runner.run_spec_provenance,
            "status": "error",
            "passed": False,
            "error_category": "provider_fatal",
            "budget": {"used": {"model_calls": 0, "total_tokens": 0}},
        }

    monkeypatch.setattr(runner, "_run_one", stop_after_one)
    monkeypatch.setattr(
        "spreadsheet_harness.audit.audit_comparison",
        lambda *_args, **_kwargs: {
            "audit_valid": False,
            "reasons": ["expected incomplete test run"],
            "manifest_sha256": manifest_sha,
            "results_sha256": None,
        },
    )
    runner.run(tasks, resume=True)
    assert (tasks[0].task_id, "ours") not in sampled


def test_comparison_resume_clears_marker_after_bound_row_fsync(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    runner, tasks = _generic_journal_runner(tmp_path, monkeypatch)
    monkeypatch.setattr(
        "spreadsheet_harness.comparison.ensure_strict_code_isolation",
        lambda *_args, **_kwargs: {},
    )
    runner.preflight(tasks)
    runner._prepare_manifest(tasks)
    manifest_sha = hashlib.sha256(runner.manifest_path.read_bytes()).hexdigest()
    task = tasks[0]
    arm = "ours"
    row = {
        "task_id": task.task_id,
        "arm": arm,
        "comparison_protocol_version": COMPARISON_PROTOCOL_VERSION,
        "comparison_manifest_sha256": manifest_sha,
        "split_provenance": runner.split_provenance,
        "run_spec_provenance": runner.run_spec_provenance,
        "continuation_source": runner._prepare_continuation_source(
            comparison_manifest_sha256=manifest_sha
        ),
        "status": "error",
        "passed": False,
        "error_category": "provider_fatal",
        "budget": {"used": {"model_calls": 0, "total_tokens": 0}},
    }
    runner._write_inflight(task.task_id, arm, comparison_manifest_sha256=manifest_sha)
    runner._append(row)
    monkeypatch.setattr(
        "spreadsheet_harness.audit.audit_comparison",
        lambda *_args, **_kwargs: {
            "audit_valid": False,
            "reasons": ["expected incomplete test run"],
            "manifest_sha256": manifest_sha,
            "results_sha256": None,
        },
    )

    runner.run(tasks, resume=True)

    assert not runner.inflight_path.exists()
    assert runner.results_path.read_text(encoding="utf-8").count("\n") == 1


def test_interrupted_seals_reject_extra_document_fields(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    runner, tasks = _generic_journal_runner(tmp_path, monkeypatch)
    runner.preflight(tasks)
    runner._prepare_manifest(tasks)
    manifest_sha = hashlib.sha256(runner.manifest_path.read_bytes()).hexdigest()
    runner.interrupted_seals_path.write_text(
        json.dumps({"schema_version": 1, "seals": [], "extra": True}) + "\n",
        encoding="utf-8",
    )

    with pytest.raises(HarnessError, match="document is invalid"):
        runner._validate_interrupted_seals(
            tasks, comparison_manifest_sha256=manifest_sha
        )


def test_comparison_summary_uses_end_to_end_denominator_and_pairing(tmp_path: Path) -> None:
    tasks = _tasks(tmp_path)
    rows = [
        {
            "task_id": task.task_id,
            "arm": arm,
            "comparison_protocol_version": COMPARISON_PROTOCOL_VERSION,
            "status": "completed",
            "passed": (
                arm == "ours" or (arm == "paper" and task.task_id == "cell-1")
            ),
            "elapsed_seconds": 10,
            "budget": {
                "used": {"model_calls": 2, "total_tokens": 120},
            },
            "agent": {
                "usage": {
                    "input_tokens": 100,
                    "output_tokens": 20,
                    "total_tokens": 120,
                },
                "request_timings": [
                    {"turn": 1, "attempts": 1, "attempt_history": [{}]},
                    {"turn": 2, "attempts": 2, "attempt_history": [{}, {}]},
                ],
            },
        }
        for task in tasks
        for arm in COMPARISON_ARMS
        if not (task.task_id == "sheet-1" and arm == "bare")
    ]
    results = tmp_path / "results.jsonl"
    results.write_text(
        "\n".join(json.dumps(row) for row in rows) + "\n",
        encoding="utf-8",
    )

    summary = comparison_summary(results, tasks)

    assert summary["expected_arm_tasks"] == 6
    assert summary["attempted_arm_tasks"] == 5
    assert summary["missing_arm_tasks"] == 1
    assert summary["arms"]["bare"]["end_to_end_accuracy"] == 0
    assert summary["arms"]["paper"]["end_to_end_accuracy"] == 0.5
    assert summary["arms"]["ours"]["end_to_end_accuracy"] == 1
    assert summary["arms"]["ours"]["total_tokens_sum"] == 240
    assert summary["arm_display_names"]["paper"] == "paper-inspired"
    assert summary["arms"]["ours"]["known_http_attempts_sum"] == 6
    assert summary["arms"]["ours"]["known_successful_request_retries_sum"] == 2
    assert summary["arms"]["ours"]["request_attempt_audit_complete"] is True
    assert summary["inference_valid"] is False
    assert "missing_arm_tasks" in summary["inference_invalid_reasons"]
    assert summary["arms"]["ours"]["cell_level"]["end_to_end_accuracy"] == 1
    assert summary["arms"]["ours"]["sheet_level"]["completion_rate"] == 1
    ours_vs_paper = summary["pairwise"]["paper_vs_ours"]
    assert ours_vs_paper["accuracy_delta_right_minus_left"] == 0.5
    assert ours_vs_paper["right_only_passes"] == 1


def test_comparison_summary_counts_known_model_failure_as_complete_nonbreaker(
    tmp_path: Path,
) -> None:
    task = _tasks(tmp_path)[0]
    rows = [
        {
            "task_id": task.task_id,
            "arm": arm,
            "comparison_protocol_version": COMPARISON_PROTOCOL_VERSION,
            "status": "completed",
            "outcome_kind": (
                "model_execution_failure" if arm == "ours" else "scored"
            ),
            "passed": arm == "bare",
            "error_category": (
                "model_execution_failure" if arm == "ours" else None
            ),
            "model_failure_reason": (
                "edit_recovery_exhausted" if arm == "ours" else None
            ),
            "budget": {"used": {"model_calls": 1, "total_tokens": 1}},
            "agent": {
                "usage": {"total_tokens": 1},
                "request_timings": [
                    {"turn": 1, "attempts": 1, "attempt_history": [{}]}
                ],
            },
        }
        for arm in ("bare", "ours")
    ]
    results = tmp_path / "known-model-failure-summary.jsonl"
    results.write_text(
        "".join(json.dumps(row) + "\n" for row in rows), encoding="utf-8"
    )

    summary = comparison_summary(results, [task], arms=("bare", "ours"))

    assert summary["inference_valid"] is True
    assert summary["completed_arm_tasks"] == 2
    assert summary["errored_arm_tasks"] == 0
    assert summary["known_model_execution_failure_arm_tasks"] == 1
    assert summary["arms"]["ours"]["completion_rate"] == 1
    assert summary["arms"]["ours"]["end_to_end_accuracy"] == 0
    assert summary["arms"]["ours"]["known_model_execution_failures"] == 1
    assert summary["arms"]["ours"]["model_execution_failure_reasons"] == {
        "edit_recovery_exhausted": 1
    }
    assert summary["pairwise"]["bare_vs_ours"]["inference_valid"] is True


def test_comparison_summary_does_not_count_recalculation_failure_as_scored_false(
    tmp_path: Path,
) -> None:
    task = _tasks(tmp_path)[0]
    rows = [
        {
            "task_id": task.task_id,
            "arm": arm,
            "comparison_protocol_version": COMPARISON_PROTOCOL_VERSION,
            "status": "error" if arm == "bare" else "completed",
            "outcome_kind": "infrastructure_failure" if arm == "bare" else "scored",
            "score_available": False if arm == "bare" else True,
            "passed": arm == "ours",
            "error_category": (
                "recalculation_infrastructure" if arm == "bare" else None
            ),
            "recalculation_failure_reason": (
                "sheet_inventory_changed" if arm == "bare" else None
            ),
            "budget": {"used": {"model_calls": 1, "total_tokens": 1}},
            "agent": {
                "usage": {"total_tokens": 1},
                "request_timings": [
                    {"turn": 1, "attempts": 1, "attempt_history": [{}]}
                ],
            },
        }
        for arm in ("bare", "ours")
    ]
    results = tmp_path / "recalculation-infrastructure-summary.jsonl"
    results.write_text(
        "".join(json.dumps(row) + "\n" for row in rows), encoding="utf-8"
    )

    summary = comparison_summary(results, [task], arms=("bare", "ours"))

    assert summary["inference_valid"] is False
    assert "recalculation_infrastructure_failures" in summary[
        "inference_invalid_reasons"
    ]
    assert summary["known_infrastructure_failure_arm_tasks"] == 1
    assert summary["arms"]["bare"]["end_to_end_accuracy"] is None
    assert summary["arms"]["bare"]["wilson_95"] is None
    assert summary["arms"]["bare"]["score_unavailable"] is True
    assert summary["arms"]["bare"]["known_outcome_descriptive"]["tasks"] == 0
    pairwise = summary["pairwise"]["bare_vs_ours"]
    assert pairwise["accuracy_delta_right_minus_left"] is None
    assert pairwise["known_outcome_descriptive"]["pairs"] == 0
    assert "collection_integrity:recalculation_infrastructure_failures" in pairwise[
        "inference_invalid_reasons"
    ]


def test_comparison_summary_keeps_post_breaker_missing_rows_out_of_descriptive_scores(
    tmp_path: Path,
) -> None:
    tasks = _tasks(tmp_path)
    first = tasks[0]
    results = tmp_path / "immediate-recalculation-breaker.jsonl"
    results.write_text(
        json.dumps(
            {
                "task_id": first.task_id,
                "arm": "bare",
                "comparison_protocol_version": COMPARISON_PROTOCOL_VERSION,
                "status": "error",
                "outcome_kind": "infrastructure_failure",
                "score_available": False,
                "passed": False,
                "error_category": "recalculation_infrastructure",
                "recalculation_failure_reason": "sheet_inventory_changed",
                "budget": {"used": {"model_calls": 1, "total_tokens": 1}},
                "agent": {
                    "usage": {"total_tokens": 1},
                    "request_timings": [
                        {"turn": 1, "attempts": 1, "attempt_history": [{}]}
                    ],
                },
            }
        )
        + "\n",
        encoding="utf-8",
    )

    summary = comparison_summary(results, tasks, arms=("bare", "ours"))

    assert summary["inference_valid"] is False
    assert summary["missing_arm_tasks"] == 3
    assert summary["arms"]["bare"]["end_to_end_accuracy"] is None
    assert summary["arms"]["ours"]["end_to_end_accuracy"] is None
    assert summary["arms"]["bare"]["known_outcome_descriptive"]["tasks"] == 0
    assert summary["arms"]["ours"]["known_outcome_descriptive"]["tasks"] == 0
    pair = summary["pairwise"]["bare_vs_ours"]
    assert pair["accuracy_delta_right_minus_left"] is None
    assert pair["known_outcome_descriptive"]["pairs"] == 0


def test_comparison_summary_keeps_scoring_infrastructure_separate_from_recalculation(
    tmp_path: Path,
) -> None:
    task = _tasks(tmp_path)[0]
    rows = [
        {
            "task_id": task.task_id,
            "arm": arm,
            "comparison_protocol_version": COMPARISON_PROTOCOL_VERSION,
            "status": "error" if arm == "bare" else "completed",
            "outcome_kind": "infrastructure_failure" if arm == "bare" else "scored",
            "score_available": False if arm == "bare" else True,
            "passed": arm == "ours",
            "error_category": "scoring_infrastructure" if arm == "bare" else None,
            "scoring_failure_reason": (
                "worksheet_scorer_unsupported" if arm == "bare" else None
            ),
            "budget": {"used": {"model_calls": 1, "total_tokens": 1}},
            "agent": {
                "usage": {"total_tokens": 1},
                "request_timings": [
                    {"turn": 1, "attempts": 1, "attempt_history": [{}]}
                ],
            },
        }
        for arm in ("bare", "ours")
    ]
    results = tmp_path / "scoring-infrastructure-summary.jsonl"
    results.write_text(
        "".join(json.dumps(row) + "\n" for row in rows),
        encoding="utf-8",
    )

    summary = comparison_summary(results, [task], arms=("bare", "ours"))

    assert summary["inference_valid"] is False
    assert "scoring_infrastructure_failures" in summary["inference_invalid_reasons"]
    assert "recalculation_infrastructure_failures" not in summary[
        "inference_invalid_reasons"
    ]
    assert summary["known_scoring_infrastructure_failure_arm_tasks"] == 1
    assert summary["known_recalculation_infrastructure_failure_arm_tasks"] == 0
    bare = summary["arms"]["bare"]
    assert bare["scoring_infrastructure_failures"] == 1
    assert bare["recalculation_infrastructure_failures"] == 0
    assert bare["infrastructure_failure_reasons"] == {
        "worksheet_scorer_unsupported": 1
    }
    assert bare["cell_level"]["scoring_infrastructure_failures"] == 1
    assert bare["cell_level"]["recalculation_infrastructure_failures"] == 0
    pair = summary["pairwise"]["bare_vs_ours"]
    assert "collection_integrity:scoring_infrastructure_failures" in pair[
        "inference_invalid_reasons"
    ]
    assert "collection_integrity:recalculation_infrastructure_failures" not in pair[
        "inference_invalid_reasons"
    ]


def test_comparison_summary_treats_sealed_unknown_as_nonprimary_not_missing(
    tmp_path: Path,
) -> None:
    tasks = _tasks(tmp_path)
    results = tmp_path / "sealed-summary.jsonl"
    rows = [
        {
            "task_id": task.task_id,
            "arm": arm,
            "comparison_protocol_version": COMPARISON_PROTOCOL_VERSION,
            "status": "completed",
            "passed": True,
            "elapsed_seconds": 1,
            "budget": {"used": {"model_calls": 1, "total_tokens": 1}},
            "agent": {"usage": {"total_tokens": 1}, "request_timings": []},
        }
        for task in tasks
        for arm in ("bare", "ours")
        if not (task.task_id == tasks[0].task_id and arm == "ours")
    ]
    results.write_text("".join(json.dumps(row) + "\n" for row in rows), encoding="utf-8")
    interrupted = {f"{tasks[0].task_id}::ours"}

    summary = comparison_summary(
        results,
        tasks,
        arms=("bare", "ours"),
        interrupted_keys=interrupted,
    )

    assert summary["missing_arm_tasks"] == 0
    assert summary["interrupted_unknown_arm_tasks"] == 1
    assert summary["inference_valid"] is False
    assert summary["arms"]["ours"]["end_to_end_accuracy"] is None
    assert summary["arms"]["ours"]["wilson_95"] is None
    assert summary["arms"]["ours"]["cell_level"]["end_to_end_accuracy"] is None
    pair = summary["pairwise"]["bare_vs_ours"]
    assert pair["accuracy_delta_right_minus_left"] is None
    assert pair["mcnemar_exact_p"] is None
    assert pair["stratified_bootstrap_95"] is None
    assert pair["holm_adjusted_p"] is None
    assert pair["known_outcome_descriptive"]["primary"] is False


def test_comparison_summary_attaches_known_descriptive_to_every_pair(
    tmp_path: Path,
) -> None:
    tasks = _tasks(tmp_path)
    results = tmp_path / "multi-pair-sealed.jsonl"
    rows = [
        {
            "task_id": task.task_id,
            "arm": arm,
            "comparison_protocol_version": COMPARISON_PROTOCOL_VERSION,
            "status": "completed",
            "passed": True,
            "budget": {"used": {"model_calls": 0, "total_tokens": 0}},
        }
        for task in tasks
        for arm in COMPARISON_ARMS
        if not (task.task_id == tasks[0].task_id and arm == "ours")
    ]
    results.write_text("".join(json.dumps(row) + "\n" for row in rows), encoding="utf-8")

    summary = comparison_summary(
        results,
        tasks,
        interrupted_keys={f"{tasks[0].task_id}::ours"},
    )

    assert all(
        pair["known_outcome_descriptive"]["primary"] is False
        for pair in summary["pairwise"].values()
    )


def test_comparison_summary_rejects_duplicate_json_keys(tmp_path: Path) -> None:
    task = _tasks(tmp_path)[0]
    results = tmp_path / "duplicate-key.jsonl"
    results.write_text(
        '{"task_id":"cell-1","task_id":"cell-1","arm":"bare"}\n',
        encoding="utf-8",
    )

    summary = comparison_summary(results, [task], arms=("bare",))

    assert summary["invalid_result_rows_ignored"] == 1
    assert "invalid_result_rows" in summary["inference_invalid_reasons"]


def test_comparison_request_attempt_totals_are_known_lower_bounds_for_error_rows(
    tmp_path: Path,
) -> None:
    task = _tasks(tmp_path)[0]
    results = tmp_path / "error-results.jsonl"
    results.write_text(
        json.dumps(
            {
                "task_id": task.task_id,
                "arm": "bare",
                "status": "error",
                "passed": False,
                "error_category": "provider_transient",
                "provider_error": {"attempts": 2},
            }
        )
        + "\n",
        encoding="utf-8",
    )

    summary = comparison_summary(results, [task], arms=("bare",))
    bare = summary["arms"]["bare"]

    assert bare["known_successful_request_retries_sum"] == 0
    assert bare["known_http_attempts_sum"] == 2
    assert bare["known_failed_request_attempts_sum"] == 2
    assert bare["request_attempt_audit_rows"] == 1
    assert bare["request_attempt_audit_complete"] is False


@pytest.mark.parametrize("present_rows", [0, 1])
def test_comparison_request_attempt_audit_is_incomplete_when_expected_rows_are_missing(
    tmp_path: Path,
    present_rows: int,
) -> None:
    tasks = _tasks(tmp_path)
    rows = [
        {
            "task_id": task.task_id,
            "arm": "bare",
            "status": "completed",
            "passed": True,
            "budget": {"used": {"model_calls": 1, "total_tokens": 1}},
            "agent": {
                "usage": {"input_tokens": 1, "output_tokens": 0, "total_tokens": 1},
                "request_timings": [
                    {"turn": 1, "attempts": 1, "attempt_history": [{}]}
                ],
            },
        }
        for task in tasks[:present_rows]
    ]
    results = tmp_path / "missing-audit-rows.jsonl"
    results.write_text(
        "".join(json.dumps(row) + "\n" for row in rows),
        encoding="utf-8",
    )

    summary = comparison_summary(results, tasks, arms=("bare",))
    bare = summary["arms"]["bare"]

    assert bare["expected"] == 2
    assert bare["attempted"] == present_rows
    assert bare["request_attempt_audit_complete"] is False
    assert summary["inference_valid"] is False
    assert "request_attempt_audit_incomplete" in summary["inference_invalid_reasons"]


def test_comparison_summary_rejects_inexact_request_attempt_history(
    tmp_path: Path,
) -> None:
    task = _tasks(tmp_path)[0]
    row = {
        "task_id": task.task_id,
        "arm": "bare",
        "comparison_protocol_version": COMPARISON_PROTOCOL_VERSION,
        "status": "completed",
        "passed": True,
        "budget": {"used": {"model_calls": 1, "total_tokens": 1}},
        "agent": {
            "usage": {"input_tokens": 1, "output_tokens": 0, "total_tokens": 1},
            "request_timings": [
                {"turn": 1, "attempts": 2, "attempt_history": [{}]}
            ],
        },
    }
    results = tmp_path / "inexact-attempt-results.jsonl"
    results.write_text(json.dumps(row) + "\n", encoding="utf-8")

    summary = comparison_summary(results, [task], arms=("bare",))

    assert summary["arms"]["bare"]["request_attempt_audit_complete"] is False
    assert summary["inference_valid"] is False
    assert "request_attempt_audit_incomplete" in summary["inference_invalid_reasons"]


def test_comparison_summary_audits_unexpected_rows_and_disables_inference(
    tmp_path: Path,
) -> None:
    tasks = _tasks(tmp_path)
    expected_rows = [
        {
            "task_id": task.task_id,
            "arm": "bare",
            "status": "completed",
            "passed": True,
            "calculation_backend": backend,
        }
        for task, backend in zip(
            tasks, ("libreoffice", "not_recalculated"), strict=True
        )
    ]
    rows = [
        *expected_rows,
        dict(expected_rows[0]),
        {
            "task_id": "unknown-task",
            "arm": "bare",
            "status": "completed",
            "passed": True,
        },
        {
            "task_id": tasks[0].task_id,
            "arm": "paper",
            "status": "completed",
            "passed": True,
        },
        {
            "task_id": tasks[0].task_id,
            "arm": "alien",
            "status": "completed",
            "passed": True,
        },
    ]
    results = tmp_path / "integrity-results.jsonl"
    results.write_text(
        "".join(json.dumps(row) + "\n" for row in rows),
        encoding="utf-8",
    )

    summary = comparison_summary(results, tasks, arms=("bare",))

    assert summary["attempted_arm_tasks"] == 2
    assert summary["missing_arm_tasks"] == 0
    assert summary["errored_arm_tasks"] == 0
    assert summary["duplicate_arm_tasks"] == 1
    assert summary["duplicate_arm_task_rows"] == 1
    assert summary["duplicate_arm_task_keys"] == [f"{tasks[0].task_id}::bare"]
    assert summary["unknown_task_rows"] == 1
    assert summary["unknown_task_ids"] == ["unknown-task"]
    assert summary["unknown_arm_rows"] == 1
    assert summary["unknown_arms"] == ["alien"]
    assert summary["unexpected_arm_rows"] == 1
    assert summary["unexpected_arms"] == ["paper"]
    assert summary["style_checked"] is False
    assert summary["calculation_backends"] == {
        "libreoffice": 1,
        "not_recalculated": 1,
    }
    assert summary["inference_valid"] is False
    assert set(summary["inference_invalid_reasons"]) >= {
        "duplicate_arm_tasks",
        "unknown_tasks",
        "unknown_arms",
        "unexpected_arms",
    }
    pair = summary["pairwise"]
    assert pair == {}


def test_comparison_summary_clears_pairwise_inference_on_collection_pollution(
    tmp_path: Path,
) -> None:
    tasks = _tasks(tmp_path)
    rows = [
        {
            "task_id": task.task_id,
            "arm": arm,
            "comparison_protocol_version": COMPARISON_PROTOCOL_VERSION,
            "status": "completed",
            "passed": arm == "ours",
        }
        for task in tasks
        for arm in COMPARISON_ARMS
    ]
    rows.append(dict(rows[0]))
    results = tmp_path / "polluted-pairwise.jsonl"
    results.write_text(
        "".join(json.dumps(row) + "\n" for row in rows),
        encoding="utf-8",
    )

    summary = comparison_summary(results, tasks)

    assert summary["inference_valid"] is False
    for pair in summary["pairwise"].values():
        assert pair["inference_valid"] is False
        assert pair["stratified_bootstrap_95"] is None
        assert pair["mcnemar_exact_p"] is None
        assert pair["holm_adjusted_p"] is None
        assert "collection_integrity:duplicate_arm_tasks" in pair[
            "inference_invalid_reasons"
        ]
        for stratum in pair["strata"].values():
            assert stratum["inference_valid"] is False
            assert stratum["stratified_bootstrap_95"] is None
            assert stratum["mcnemar_exact_p"] is None


def test_comparison_summary_supports_preregistered_analysis_subset(
    tmp_path: Path,
) -> None:
    tasks = _tasks(tmp_path)
    rows = [
        {
            "task_id": task.task_id,
            "arm": arm,
            "comparison_protocol_version": COMPARISON_PROTOCOL_VERSION,
            "status": "completed",
            "passed": task.task_id == "cell-1",
            "calculation_backend": "libreoffice",
            "budget": {"used": {"model_calls": 1, "total_tokens": 1}},
            "agent": {
                "usage": {"input_tokens": 1, "output_tokens": 0, "total_tokens": 1},
                "request_timings": [
                    {"turn": 1, "attempts": 1, "attempt_history": [{}]}
                ],
            },
        }
        for task in tasks
        for arm in COMPARISON_ARMS
    ]
    results = tmp_path / "subset-results.jsonl"
    results.write_text(
        "".join(json.dumps(row) + "\n" for row in rows),
        encoding="utf-8",
    )

    full = comparison_summary(results, tasks)
    subset = comparison_summary(results, tasks[:1], collection_tasks=tasks)

    assert full["inference_valid"] is True
    assert subset["inference_valid"] is True
    assert subset["task_count"] == 1
    assert subset["expected_arm_tasks"] == subset["completed_arm_tasks"] == 3
    assert subset["unknown_task_rows"] == 0
    assert subset["arms"]["bare"]["end_to_end_accuracy"] == 1


@pytest.mark.parametrize("protocol", [None, "resource_matched_multi_arm_v19"])
def test_comparison_summary_rejects_missing_or_mismatched_row_protocol(
    tmp_path: Path,
    protocol: str | None,
) -> None:
    task = _tasks(tmp_path)[0]
    row = {
        "task_id": task.task_id,
        "arm": "bare",
        "status": "completed",
        "passed": True,
    }
    if protocol is not None:
        row["comparison_protocol_version"] = protocol
    results = tmp_path / "protocol-mismatch-results.jsonl"
    results.write_text(json.dumps(row) + "\n", encoding="utf-8")

    summary = comparison_summary(results, [task], arms=("bare",))

    assert summary["inference_valid"] is False
    assert summary["protocol_mismatch_rows"] == 1
    assert summary["observed_protocols"] == [protocol or "<missing>"]
    assert "comparison_protocol_mismatch" in summary["inference_invalid_reasons"]


def test_comparison_summary_rejects_analysis_outside_collection(tmp_path: Path) -> None:
    tasks = _tasks(tmp_path)

    with pytest.raises(ValueError, match="subset of collection tasks"):
        comparison_summary(
            tmp_path / "missing.jsonl",
            tasks,
            collection_tasks=tasks[:1],
        )


def test_comparison_runner_refuses_duplicate_arm_task_rows(tmp_path: Path) -> None:
    runner = ComparisonBenchmarkRunner(
        ProviderConfig("https://example.test/v1", "not-a-real-key", "test-model"),
        tmp_path / "duplicate-resume",
        skill_registry=SkillRegistry([]),
    )
    row = {
        "task_id": "cell-1",
        "arm": "bare",
        "status": "completed",
        "passed": True,
    }
    runner.results_path.write_text(
        json.dumps(row) + "\n" + json.dumps(row) + "\n",
        encoding="utf-8",
    )

    with pytest.raises(HarnessError, match="duplicate arm-task rows"):
        runner._latest()


def test_comparison_runner_refuses_protocol_mismatched_resume(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    tasks = _tasks(tmp_path)
    runner = ComparisonBenchmarkRunner(
        ProviderConfig("https://example.test/v1", "not-a-real-key", "test-model"),
        tmp_path / "mismatched-resume",
        skill_registry=SkillRegistry([]),
        arms=("bare",),
    )
    runner._prepare_manifest(tasks)
    manifest_sha256 = hashlib.sha256(runner.manifest_path.read_bytes()).hexdigest()
    runner.results_path.write_text(
        json.dumps(
            {
                "task_id": tasks[0].task_id,
                "arm": "bare",
                "comparison_protocol_version": "resource_matched_multi_arm_v19",
                "comparison_manifest_sha256": manifest_sha256,
                "status": "completed",
                "passed": True,
            }
        )
        + "\n",
        encoding="utf-8",
    )
    monkeypatch.setattr(
        "spreadsheet_harness.comparison.ensure_strict_code_isolation",
        lambda *_args, **_kwargs: {},
    )

    with pytest.raises(HarnessError, match="different or missing protocol"):
        runner.run(tasks)


def test_comparison_runner_refuses_split_provenance_mismatched_resume(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    tasks = _tasks(tmp_path)
    runner = ComparisonBenchmarkRunner(
        ProviderConfig("https://example.test/v1", "not-a-real-key", "test-model"),
        tmp_path / "split-mismatched-resume",
        skill_registry=SkillRegistry([]),
        arms=("bare",),
    )
    runner._prepare_manifest(tasks)
    manifest_sha256 = hashlib.sha256(runner.manifest_path.read_bytes()).hexdigest()
    runner.results_path.write_text(
        json.dumps(
            {
                "task_id": tasks[0].task_id,
                "arm": "bare",
                "comparison_protocol_version": COMPARISON_PROTOCOL_VERSION,
                "comparison_manifest_sha256": manifest_sha256,
                "split_provenance": {"attacker": True},
                "status": "completed",
                "passed": True,
            }
        )
        + "\n",
        encoding="utf-8",
    )
    monkeypatch.setattr(
        "spreadsheet_harness.comparison.ensure_strict_code_isolation",
        lambda *_args, **_kwargs: {},
    )
    monkeypatch.setattr(
        runner,
        "_run_one",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(
            AssertionError("mismatched split provenance must stop sampling")
        ),
    )

    with pytest.raises(HarnessError, match="split provenance"):
        runner.run(tasks)


@pytest.mark.parametrize(
    "damaged",
    [
        '{"task_id":',
        json.dumps(
            {
                "task_id": "cell-1",
                "arm": "bare",
                "comparison_protocol_version": COMPARISON_PROTOCOL_VERSION,
                "status": "completed",
                "passed": True,
            }
        ),
    ],
)
def test_comparison_resume_fails_closed_on_damaged_journal_without_mutation(
    tmp_path: Path,
    monkeypatch: Any,
    damaged: str,
) -> None:
    tasks = _tasks(tmp_path)
    runner = ComparisonBenchmarkRunner(
        ProviderConfig("https://example.test/v1", "not-a-real-key", "test-model"),
        tmp_path / "damaged-resume",
        skill_registry=SkillRegistry([]),
        arms=("bare",),
    )
    runner._prepare_manifest(tasks)
    runner.results_path.write_text(damaged, encoding="utf-8")
    before = runner.results_path.read_bytes()
    monkeypatch.setattr(
        "spreadsheet_harness.comparison.ensure_strict_code_isolation",
        lambda *_args, **_kwargs: {},
    )

    def must_not_run(*_: Any, **__: Any) -> dict[str, Any]:
        raise AssertionError("damaged comparison journal must never be resampled")

    monkeypatch.setattr(runner, "_run_one", must_not_run)

    with pytest.raises(HarnessError, match="damaged or non-terminated"):
        runner.run(tasks)

    assert runner.results_path.read_bytes() == before


def test_comparison_resume_fails_closed_on_invalid_utf8_without_sampling(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    tasks = _tasks(tmp_path)
    runner = ComparisonBenchmarkRunner(
        ProviderConfig("https://example.test/v1", "not-a-real-key", "test-model"),
        tmp_path / "invalid-utf8-resume",
        skill_registry=SkillRegistry([]),
        arms=("bare",),
    )
    runner._prepare_manifest(tasks)
    manifest_sha256 = hashlib.sha256(runner.manifest_path.read_bytes()).hexdigest()
    row = {
        "task_id": tasks[0].task_id,
        "arm": "bare",
        "comparison_protocol_version": COMPARISON_PROTOCOL_VERSION,
        "comparison_manifest_sha256": manifest_sha256,
        "status": "completed",
        "agent": {"final_text": "CORRUPT"},
    }
    encoded = json.dumps(row).encode("utf-8").replace(b"CORRUPT", b"\xff") + b"\n"
    runner.results_path.write_bytes(encoded)
    before = runner.results_path.read_bytes()
    monkeypatch.setattr(
        "spreadsheet_harness.comparison.ensure_strict_code_isolation",
        lambda *_args, **_kwargs: {},
    )

    def must_not_run(*_: Any, **__: Any) -> dict[str, Any]:
        raise AssertionError("invalid UTF-8 comparison journal must stop all sampling")

    monkeypatch.setattr(runner, "_run_one", must_not_run)

    with pytest.raises(HarnessError, match="damaged or non-terminated"):
        runner.run(tasks)

    assert runner.results_path.read_bytes() == before


def test_comparison_runner_calls_arm_without_answer_metadata(
    tmp_path: Path, monkeypatch: Any
) -> None:
    task = _tasks(tmp_path)[0]
    captured: dict[str, Any] = {}
    captured_pacers: list[Any] = []

    def fake_run_arm(**kwargs: Any) -> AgentResult:
        captured.update(kwargs)
        captured_pacers.append(kwargs["pacer"])
        return AgentResult("done", 1, 0, {}, "response")

    monkeypatch.setattr("spreadsheet_harness.comparison.run_arm", fake_run_arm)
    runner = ComparisonBenchmarkRunner(
        ProviderConfig("https://example.test/v1", "not-a-real-key", "test-model"),
        tmp_path / "run-one",
        skill_registry=SkillRegistry([]),
        max_model_calls=100,
        max_turns_per_arm=100,
        recalculate=False,
    )

    row = runner._run_one(task, "bare", comparison_manifest_sha256="a" * 64)
    second_row = runner._run_one(task, "ours", comparison_manifest_sha256="a" * 64)

    assert row["status"] == "completed"
    assert second_row["status"] == "completed"
    assert len(row["output_sha256"]) == 64
    assert captured["instruction"] == task.instruction
    assert captured["max_turns_per_arm"] == 100
    assert captured_pacers == [runner.relay_pacer, runner.relay_pacer]
    assert row["max_model_calls"] == 100
    assert row["max_turns_per_arm"] == 100
    assert row["stage_turn_caps"] == {"solve": 100}
    assert row["comparison_manifest_sha256"] == "a" * 64
    assert row["request_interval_seconds"] == 0.0
    assert row["litellm_timeout_seconds"] is None
    assert "TOP_SECRET" not in json.dumps(captured, default=str)
    trajectory = read_trajectory(Path(row["run_dir"]) / "trajectory.jsonl")
    configured = [item for item in trajectory if item["event"] == "benchmark.configured"]
    assert configured[0]["payload"]["max_model_calls"] == 100
    assert configured[0]["payload"]["max_turns_per_arm"] == 100
    assert configured[0]["payload"]["stage_turn_caps"] == {"solve": 100}
    assert configured[0]["payload"]["litellm_timeout_seconds"] is None
    evaluation = [item for item in trajectory if item["event"] == "benchmark.evaluated"]
    assert len(evaluation) == 1
    assert evaluation[0]["payload"]["passed"] is False
    assert evaluation[0]["payload"]["arm"] == "bare"
    assert evaluation[0]["payload"]["difference_count"] == 1
    assert evaluation[0]["payload"]["difference_categories"] == {
        "metadata_or_structure": 1
    }


def test_comparison_recalculation_sheet_drift_is_persisted_without_scoring(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    task = _tasks(tmp_path)[0]

    monkeypatch.setattr(
        comparison_module,
        "run_arm",
        lambda **_: AgentResult("done", 1, 0, {}, "response"),
    )
    monkeypatch.setattr(render_module, "find_libreoffice", lambda explicit=None: "/fake/soffice")
    monkeypatch.setattr(render_module, "libreoffice_version", lambda binary: "LibreOffice test")

    def fake_convert(
        source_copy: Path,
        output_dir: Path,
        **kwargs: object,
    ) -> Path:
        output_dir.mkdir(parents=True)
        converted = output_dir / source_copy.name
        shutil.copy2(source_copy, converted)
        workbook = load_workbook(converted)
        try:
            workbook.active.title = "Changed"
            workbook.save(converted)
        finally:
            workbook.close()
        return converted

    def must_not_score(*_: object, **__: object) -> object:
        raise AssertionError("sheet identity drift must stop before scoring")

    monkeypatch.setattr(render_module, "_convert_with_libreoffice", fake_convert)
    monkeypatch.setattr(
        comparison_module,
        "compare_workbooks_chartsheet_safe",
        must_not_score,
    )
    runner = ComparisonBenchmarkRunner(
        ProviderConfig("https://example.test/v1", "not-a-real-key", "test-model"),
        tmp_path / "recalculation-drift",
        skill_registry=SkillRegistry([]),
        arms=("bare",),
        max_model_calls=10,
        max_turns_per_arm=10,
        recalculate=True,
    )

    row = runner._run_one(task, "bare", comparison_manifest_sha256="a" * 64)

    assert row["status"] == "error"
    assert row["passed"] is False
    assert row["outcome_kind"] == "infrastructure_failure"
    assert row["score_available"] is False
    assert row["error_category"] == "recalculation_infrastructure"
    assert row["infrastructure_failure_stage"] == "recalculation"
    assert row["recalculation_failure_reason"] == "sheet_inventory_changed"
    assert row["recalculation"]["sheet_inventory_integrity"]["matched"] is False
    assert "comparison" not in row
    assert "artifact_score_passed" not in row
    assert Path(row["recalculation"]["failure_artifact_path"]).is_file()
    run_manifest = json.loads(
        (Path(row["run_dir"]) / "run.json").read_text(encoding="utf-8")
    )
    assert run_manifest["result"]["recalculation"] == row["recalculation"]
    trajectory = read_trajectory(Path(row["run_dir"]) / "trajectory.jsonl")
    assert not any(item["event"] == "benchmark.evaluated" for item in trajectory)
    not_evaluated = [
        item for item in trajectory if item["event"] == "benchmark.not_evaluated"
    ]
    assert not_evaluated[0]["payload"]["error_category"] == (
        "recalculation_infrastructure"
    )
    assert not_evaluated[0]["payload"]["recalculation"] == row["recalculation"]


def test_comparison_classifies_unsupported_scorer_as_infrastructure_no_score(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    task = _tasks(tmp_path)[0]
    monkeypatch.setattr(
        comparison_module,
        "run_arm",
        lambda **_: AgentResult("done", 1, 0, {}, "response"),
    )

    def unsupported(*_: object, **__: object) -> object:
        raise ScoringInfrastructureError("worksheet scorer unsupported")

    monkeypatch.setattr(
        comparison_module,
        "compare_workbooks_chartsheet_safe",
        unsupported,
    )
    runner = ComparisonBenchmarkRunner(
        ProviderConfig("https://example.test/v1", "not-a-real-key", "test-model"),
        tmp_path / "scoring-infrastructure",
        skill_registry=SkillRegistry([]),
        arms=("bare",),
        recalculate=False,
    )

    row = runner._run_one(task, "bare", comparison_manifest_sha256="a" * 64)

    assert row["status"] == "error"
    assert row["outcome_kind"] == "infrastructure_failure"
    assert row["score_available"] is False
    assert row["error_category"] == "scoring_infrastructure"
    assert row["infrastructure_failure_stage"] == "scoring"
    assert row["scoring_failure_reason"] == "worksheet_scorer_unsupported"
    assert "comparison" not in row
    assert "artifact_score_passed" not in row


def test_comparison_run_binds_result_row_to_exact_manifest(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    tasks = _tasks(tmp_path)[:1]
    runner = ComparisonBenchmarkRunner(
        ProviderConfig("https://example.test/v1", "not-a-real-key", "test-model"),
        tmp_path / "manifest-bound",
        skill_registry=SkillRegistry([]),
        arms=("bare",),
        recalculate=False,
    )
    monkeypatch.setattr(
        "spreadsheet_harness.comparison.ensure_strict_code_isolation",
        lambda *_args, **_kwargs: {},
    )
    captured: dict[str, Any] = {}

    def fake_run_one(*_: Any, **kwargs: Any) -> dict[str, Any]:
        captured.update(kwargs)
        task = tasks[0]
        manifest = json.loads(runner.manifest_path.read_text(encoding="utf-8"))
        task_dir = runner.output_dir / "runs" / task.task_id / "bare"
        output = task_dir / "artifacts" / "output.xlsx"
        output.parent.mkdir(parents=True, exist_ok=True)
        output.write_bytes(task.golden_path.read_bytes())
        output_sha256 = hashlib.sha256(output.read_bytes()).hexdigest()
        timings = [
            {
                "turn": 1,
                "stage": "solve",
                "attempts": 1,
                "attempt_history": [
                    {"api_protocol": "responses", "endpoint": "/responses"}
                ],
                "input_tokens": 2,
                "output_tokens": 0,
                "total_tokens": 2,
            },
            {
                "turn": 2,
                "stage": "solve",
                "attempts": 1,
                "attempt_history": [
                    {"api_protocol": "responses", "endpoint": "/responses"}
                ],
                "input_tokens": 2,
                "output_tokens": 0,
                "total_tokens": 2,
            },
            {
                "turn": 3,
                "stage": "solve",
                "attempts": 1,
                "attempt_history": [
                    {"api_protocol": "responses", "endpoint": "/responses"}
                ],
                "input_tokens": 4,
                "output_tokens": 2,
                "total_tokens": 6,
            },
        ]
        tool_trace = [
            {"name": "code_interpreter", "ok": True},
            {"name": "code_interpreter", "ok": True},
        ]
        budget = {
            "limit": {
                "model_calls": 20,
                "total_tokens": 100_000,
                "elapsed_seconds": 900,
            },
            "used": {"model_calls": 3, "total_tokens": 10, "elapsed_seconds": 1.0},
            "termination": None,
        }
        terminal_response = {
            "status": "accepted",
            "response_id": "response-final",
            "acknowledgement": {},
        }
        comparison = compare_workbooks(
            task.golden_path,
            output,
            task.answer_position,
            answer_sheet=task.answer_sheet,
        ).to_dict()
        return {
            "task_id": task.task_id,
            "arm": "bare",
            "comparison_protocol_version": COMPARISON_PROTOCOL_VERSION,
            "comparison_manifest_sha256": kwargs["comparison_manifest_sha256"],
            "instruction_type": task.instruction_type,
            "model": "test-model",
            "api_protocol": "responses",
            "requested_reasoning_effort": manifest["configuration"][
                "requested_reasoning_effort"
            ],
            "reasoning_effort": manifest["configuration"]["reasoning_effort"],
            "request_interval_seconds": 0.0,
            "litellm_timeout_seconds": None,
            "generation": manifest["configuration"]["generation"],
            "max_model_calls": 20,
            "max_turns_per_arm": 20,
            "stage_turn_caps": {"solve": 20},
            "calculation_backend": "not_recalculated",
            "status": "completed",
            "outcome_kind": "scored",
            "passed": comparison["passed"],
            "artifact_score_passed": comparison["passed"],
            "comparison": comparison,
            "run_dir": str(task_dir),
            "output_workbook": str(output),
            "output_sha256": output_sha256,
            "budget": budget,
            "agent": {
                "arm": "bare",
                "final_text": "Spreadsheet task completed.",
                "response_id": "response-final",
                "turns": 3,
                "tool_calls": 2,
                "usage": {"input_tokens": 8, "output_tokens": 2, "total_tokens": 10},
                "request_timings": timings,
                "tool_trace": [
                    {"stage": "solve", **item} for item in tool_trace
                ],
                "terminal_submissions": 1,
                "function_calls_total": 3,
                "post_prefix_tool_choice": "auto",
                "terminal_tool": "submit_result",
                "observed_terminal_tool": "submit_result",
                "terminal_response": terminal_response,
                "budget": budget,
                "stages": [
                    {
                        "name": "solve",
                        "max_turns": 20,
                        "allowed_tools": ["code_interpreter"],
                        "first_tool_choice": "code_interpreter",
                        "observed_first_tool": "code_interpreter",
                        "forced_tool_prefix": ["code_interpreter", "code_interpreter"],
                        "observed_forced_tool_prefix": [
                            "code_interpreter",
                            "code_interpreter",
                        ],
                        "post_prefix_tool_choice": "auto",
                        "terminal_tool": "submit_result",
                        "observed_terminal_tool": "submit_result",
                        "tool_name_trace": [
                            "code_interpreter",
                            "code_interpreter",
                        ],
                        "tool_trace": tool_trace,
                        "agent": {
                            "final_text": "Spreadsheet task completed.",
                            "response_id": "response-final",
                            "turns": 3,
                            "tool_calls": 2,
                            "usage": {
                                "input_tokens": 8,
                                "output_tokens": 2,
                                "total_tokens": 10,
                            },
                            "request_timings": timings,
                            "tool_trace": tool_trace,
                            "terminal_submissions": 1,
                            "function_calls_total": 3,
                            "post_prefix_tool_choice": "auto",
                            "terminal_tool": "submit_result",
                            "observed_terminal_tool": "submit_result",
                            "terminal_response": terminal_response,
                            "budget": budget,
                        },
                    }
                ],
            },
        }

    monkeypatch.setattr(runner, "_run_one", fake_run_one)

    summary = runner.run(tasks)

    expected = hashlib.sha256(runner.manifest_path.read_bytes()).hexdigest()
    assert captured["comparison_manifest_sha256"] == expected
    row = json.loads(runner.results_path.read_text(encoding="utf-8"))
    assert row["comparison_manifest_sha256"] == expected
    assert row["status"] == "completed"
    assert summary["protocol_audit_valid"] is True
    assert summary["inference_valid"] is True

    row["model"] = "tampered-model"
    runner.results_path.write_text(json.dumps(row) + "\n", encoding="utf-8")
    rerun = runner.run(tasks)
    assert rerun["protocol_audit_valid"] is False
    assert rerun["inference_valid"] is False
    assert "comparison_audit_failed" in rerun["inference_invalid_reasons"]


def test_comparison_classifies_paper_stage_validation(
    tmp_path: Path, monkeypatch: Any
) -> None:
    task = _tasks(tmp_path)[0]

    def fail_paper(**_: Any) -> AgentResult:
        raise PaperStageValidationError("vision_verify", "view_image was not attached")

    monkeypatch.setattr("spreadsheet_harness.comparison.run_arm", fail_paper)
    runner = ComparisonBenchmarkRunner(
        ProviderConfig("https://example.test/v1", "not-a-real-key", "test-model"),
        tmp_path / "paper-validation",
        skill_registry=SkillRegistry([]),
        recalculate=False,
    )

    row = runner._run_one(task, "paper", comparison_manifest_sha256="a" * 64)

    assert row["status"] == "error"
    assert row["error_category"] == "paper_stage_validation"
    assert row["paper_stage"] == "vision_verify"
    assert row["paper_stage_reason"] == "view_image was not attached"


def test_comparison_classifies_required_routing_failure(
    tmp_path: Path, monkeypatch: Any
) -> None:
    task = _tasks(tmp_path)[0]

    def fail_routing(**_: Any) -> AgentResult:
        raise AgentRoutingError("required route returned no function call")

    monkeypatch.setattr("spreadsheet_harness.comparison.run_arm", fail_routing)
    runner = ComparisonBenchmarkRunner(
        ProviderConfig("https://example.test/v1", "not-a-real-key", "test-model"),
        tmp_path / "routing-validation",
        skill_registry=SkillRegistry([]),
        recalculate=False,
    )

    row = runner._run_one(task, "bare", comparison_manifest_sha256="a" * 64)

    assert row["status"] == "error"
    assert row["error_category"] == "routing_protocol"


@pytest.mark.parametrize(
    "reason", ["workbook_unchanged", "edit_recovery_exhausted"]
)
def test_comparison_scores_known_model_execution_failure_as_completed_false(
    tmp_path: Path,
    monkeypatch: Any,
    reason: str,
) -> None:
    task = _tasks(tmp_path)[0]
    evidence = AgentResult(
        "unable to finish",
        2,
        0,
        {"input_tokens": 8, "output_tokens": 2, "total_tokens": 10},
        "response-2",
        request_timings=[
            {"turn": turn, "attempts": 1, "attempt_history": [{}]}
            for turn in (1, 2)
        ],
    )

    def fail_execution(**_: Any) -> AgentResult:
        raise AgentExecutionFailure(
            "model did not produce a usable workbook edit",
            reason=reason,
            agent_result=evidence,
        )

    monkeypatch.setattr("spreadsheet_harness.comparison.run_arm", fail_execution)
    runner = ComparisonBenchmarkRunner(
        ProviderConfig("https://example.test/v1", "not-a-real-key", "test-model"),
        tmp_path / f"known-failure-{reason}",
        skill_registry=SkillRegistry([]),
        arms=("bare",),
        max_model_calls=3,
        max_turns_per_arm=3,
        recalculate=False,
    )

    row = runner._run_one(task, "bare", comparison_manifest_sha256="a" * 64)

    assert row["status"] == "completed"
    assert row["passed"] is False
    assert row["artifact_score_passed"] is False
    assert row["comparison"]["passed"] is False
    assert row["outcome_kind"] == "model_execution_failure"
    assert row["error_category"] == "model_execution_failure"
    assert row["model_failure_reason"] == reason
    assert row["error_retryable"] is False
    assert row["agent"] == evidence.to_dict()


def test_comparison_requires_evidence_for_known_model_execution_failure(
    tmp_path: Path, monkeypatch: Any
) -> None:
    task = _tasks(tmp_path)[0]

    def fail_without_evidence(**_: Any) -> AgentResult:
        raise AgentExecutionFailure(
            "model did not edit the workbook",
            reason="workbook_unchanged",
        )

    monkeypatch.setattr("spreadsheet_harness.comparison.run_arm", fail_without_evidence)
    runner = ComparisonBenchmarkRunner(
        ProviderConfig("https://example.test/v1", "not-a-real-key", "test-model"),
        tmp_path / "known-failure-without-evidence",
        skill_registry=SkillRegistry([]),
        arms=("bare",),
        recalculate=False,
    )

    row = runner._run_one(task, "bare", comparison_manifest_sha256="a" * 64)

    assert row["status"] == "error"
    assert row["error_category"] == "harness"
    assert row["error_type"] == "HarnessError"


def test_comparison_records_token_budget_exhaustion_as_completed_false(
    tmp_path: Path, monkeypatch: Any
) -> None:
    task = _tasks(tmp_path)[0]

    class BudgetFailureEvidence:
        def __init__(self, budget: Any) -> None:
            self.budget = budget

        def to_dict(self) -> dict[str, Any]:
            snapshot = self.budget.to_dict()
            return {
                "usage": {"input_tokens": 8, "output_tokens": 3, "total_tokens": 11},
                "request_timings": [{"turn": 1, "total_tokens": 11}],
                "budget": snapshot,
            }

    def exhaust_budget(**kwargs: Any) -> AgentResult:
        budget = kwargs["budget"]
        reservation = budget.begin_model_call(stage="solve")
        with pytest.raises(AgentBudgetError):
            budget.record_response(reservation, {"total_tokens": 11}, stage="solve")
        raise AgentExecutionFailure(
            "token budget exhausted",
            reason="budget_exhausted",
            agent_result=BudgetFailureEvidence(budget),
        )

    monkeypatch.setattr("spreadsheet_harness.comparison.run_arm", exhaust_budget)
    runner = ComparisonBenchmarkRunner(
        ProviderConfig("https://example.test/v1", "not-a-real-key", "test-model"),
        tmp_path / "budget-failure",
        skill_registry=SkillRegistry([]),
        arms=("bare",),
        max_model_calls=3,
        max_turns_per_arm=3,
        max_total_tokens=10,
        recalculate=False,
    )

    row = runner._run_one(task, "bare", comparison_manifest_sha256="a" * 64)

    assert row["status"] == "completed"
    assert row["passed"] is False
    assert row["outcome_kind"] == "model_execution_failure"
    assert row["model_failure_reason"] == "budget_exhausted"
    assert row["budget"]["used"]["total_tokens"] == 11


def test_known_model_execution_failure_row_passes_full_audit(
    tmp_path: Path, monkeypatch: Any
) -> None:
    from spreadsheet_harness.audit import audit_comparison

    task = _tasks(tmp_path)[0]
    runner = ComparisonBenchmarkRunner(
        ProviderConfig("https://example.test/v1", "not-a-real-key", "test-model"),
        tmp_path / "audited-known-failure",
        skill_registry=SkillRegistry([]),
        arms=("bare",),
        max_model_calls=3,
        max_turns_per_arm=3,
        max_total_tokens=100,
        recalculate=False,
    )

    class AuditableFailureEvidence:
        def __init__(self, budget: Any, timings: list[dict[str, Any]]) -> None:
            self.budget = budget
            self.timings = timings

        def to_dict(self) -> dict[str, Any]:
            budget = self.budget.to_dict()
            tool_trace = [
                {"name": "code_interpreter", "ok": True},
                {"name": "code_interpreter", "ok": True},
            ]
            return {
                "arm": "bare",
                "turns": 2,
                "tool_calls": 2,
                "usage": {
                    "input_tokens": 8,
                    "output_tokens": 2,
                    "total_tokens": 10,
                },
                "request_timings": self.timings,
                "tool_trace": [
                    {"stage": "solve", **item} for item in tool_trace
                ],
                "terminal_submissions": 1,
                "function_calls_total": 3,
                "budget": budget,
                "stages": [
                    {
                        "name": "solve",
                        "max_turns": 3,
                        "allowed_tools": ["code_interpreter"],
                        "first_tool_choice": "code_interpreter",
                        "observed_first_tool": "code_interpreter",
                        "forced_tool_prefix": [
                            "code_interpreter",
                            "code_interpreter",
                        ],
                        "observed_forced_tool_prefix": [
                            "code_interpreter",
                            "code_interpreter",
                        ],
                        "terminal_tool": "submit_result",
                        "observed_terminal_tool": "submit_result",
                        "tool_name_trace": [
                            "code_interpreter",
                            "code_interpreter",
                        ],
                        "tool_trace": tool_trace,
                        "agent": {
                            "turns": 2,
                            "tool_calls": 2,
                            "usage": {
                                "input_tokens": 8,
                                "output_tokens": 2,
                                "total_tokens": 10,
                            },
                            "request_timings": self.timings,
                            "tool_trace": tool_trace,
                            "terminal_submissions": 1,
                            "function_calls_total": 3,
                            "budget": budget,
                        },
                    }
                ],
            }

    def fail_with_auditable_evidence(**kwargs: Any) -> AgentResult:
        budget = kwargs["budget"]
        timings = []
        for turn in (1, 2):
            reservation = budget.begin_model_call(stage="solve")
            budget.record_response(
                reservation,
                {
                    "input_tokens": 4,
                    "output_tokens": 1,
                    "total_tokens": 5,
                },
                stage="solve",
            )
            timings.append(
                {
                    "turn": turn,
                    "stage": "solve",
                    "attempts": 1,
                    "attempt_history": [
                        {"api_protocol": "responses", "endpoint": "/responses"}
                    ],
                    "input_tokens": 4,
                    "output_tokens": 1,
                    "total_tokens": 5,
                }
            )
        raise AgentExecutionFailure(
            "final recovery code did not produce a saved edit",
            reason="edit_recovery_exhausted",
            agent_result=AuditableFailureEvidence(budget, timings),
        )

    monkeypatch.setattr(
        "spreadsheet_harness.comparison.run_arm", fail_with_auditable_evidence
    )
    runner._prepare_manifest([task])
    manifest_sha256 = hashlib.sha256(runner.manifest_path.read_bytes()).hexdigest()
    row = runner._run_one(
        task, "bare", comparison_manifest_sha256=manifest_sha256
    )
    runner._append(row)

    report = audit_comparison(runner.output_dir, [task], arms=("bare",))

    assert report["audit_valid"] is True
    assert report["study_complete"] is True
    assert report["known_failed_rows"] == 1
    assert report["known_model_execution_failure_rows"] == 1

    monkeypatch.setattr(
        render_module, "find_libreoffice", lambda explicit=None: "/fake/soffice"
    )
    monkeypatch.setattr(
        render_module, "libreoffice_version", lambda binary: "LibreOffice test"
    )

    def fake_convert(source_copy: Path, output_dir: Path, **_: object) -> Path:
        output_dir.mkdir(parents=True)
        converted = output_dir / source_copy.name
        shutil.copy2(source_copy, converted)
        workbook = load_workbook(converted)
        try:
            workbook.active.title = "Changed"
            workbook.save(converted)
        finally:
            workbook.close()
        return converted

    monkeypatch.setattr(render_module, "_convert_with_libreoffice", fake_convert)
    drift_runner = ComparisonBenchmarkRunner(
        ProviderConfig("https://example.test/v1", "not-a-real-key", "test-model"),
        tmp_path / "audited-known-failure-then-drift",
        skill_registry=SkillRegistry([]),
        arms=("bare",),
        max_model_calls=3,
        max_turns_per_arm=3,
        max_total_tokens=100,
        recalculate=True,
    )
    drift_runner._prepare_manifest([task])
    drift_manifest_sha256 = hashlib.sha256(
        drift_runner.manifest_path.read_bytes()
    ).hexdigest()
    drift_row = drift_runner._run_one(
        task, "bare", comparison_manifest_sha256=drift_manifest_sha256
    )
    drift_runner._append(drift_row)

    drift_report = audit_comparison(
        drift_runner.output_dir, [task], arms=("bare",)
    )

    assert drift_row["outcome_kind"] == "infrastructure_failure"
    assert drift_row["score_available"] is False
    assert drift_row["prior_model_execution_failure"] == {
        "error": "final recovery code did not produce a saved edit",
        "error_type": "AgentExecutionFailure",
        "model_failure_reason": "edit_recovery_exhausted",
    }
    assert drift_report["audit_valid"] is True
    assert drift_report["study_complete"] is False
    assert drift_report["known_recalculation_infrastructure_failure_rows"] == 1


def test_comparison_counts_ambiguous_delivery_as_transient_but_not_retryable(
    tmp_path: Path, monkeypatch: Any
) -> None:
    task = _tasks(tmp_path)[0]

    def fail_ambiguous(**_: Any) -> AgentResult:
        raise ProviderError(
            "Responses request timed out during read",
            retryable=True,
            safe_to_retry=False,
            phase="read",
            delivery_state="ambiguous_post_send",
        )

    monkeypatch.setattr("spreadsheet_harness.comparison.run_arm", fail_ambiguous)
    runner = ComparisonBenchmarkRunner(
        ProviderConfig("https://example.test/v1", "not-a-real-key", "test-model"),
        tmp_path / "ambiguous-delivery",
        skill_registry=SkillRegistry([]),
        recalculate=False,
    )

    row = runner._run_one(task, "bare", comparison_manifest_sha256="a" * 64)

    assert row["status"] == "error"
    assert row["error_category"] == "provider_transient"
    assert row["error_retryable"] is False
    assert row["provider_error"]["safe_to_retry"] is False
    assert row["provider_error"]["delivery_state"] == "ambiguous_post_send"


def test_resume_preserves_historical_provider_circuit_breaker(
    tmp_path: Path, monkeypatch: Any
) -> None:
    tasks = _tasks(tmp_path)
    runner = ComparisonBenchmarkRunner(
        ProviderConfig(
            "https://example.test/v1",
            "not-a-real-key",
            "test-model",
            request_interval_seconds=20.0,
        ),
        tmp_path / "resume-breaker",
        skill_registry=SkillRegistry([]),
    )
    runner._prepare_manifest(tasks)
    manifest_sha256 = hashlib.sha256(runner.manifest_path.read_bytes()).hexdigest()
    runner.results_path.write_text(
        json.dumps(
            {
                "task_id": tasks[0].task_id,
                "arm": "bare",
                "comparison_protocol_version": COMPARISON_PROTOCOL_VERSION,
                "comparison_manifest_sha256": manifest_sha256,
                "status": "error",
                "passed": False,
                "error_category": "provider_fatal",
            }
        )
        + "\n",
        encoding="utf-8",
    )
    monkeypatch.setattr(
        "spreadsheet_harness.comparison.ensure_strict_code_isolation",
        lambda *_args, **_kwargs: {},
    )

    def must_not_run(*_: Any, **__: Any) -> dict[str, Any]:
        raise AssertionError("historical fatal provider error must keep breaker open")

    monkeypatch.setattr(runner, "_run_one", must_not_run)

    summary = runner.run(tasks)

    assert summary["circuit_breaker_tripped"] is True
    assert summary["fatal_provider_arm_tasks"] == 1
    assert summary["attempted_arm_tasks"] == 1


def test_end_to_end_deadline_covers_scoring(
    tmp_path: Path, monkeypatch: Any
) -> None:
    task = _tasks(tmp_path)[0]
    now = [100.0]
    monkeypatch.setattr("spreadsheet_harness.comparison.monotonic", lambda: now[0])
    monkeypatch.setattr("spreadsheet_harness.budget.time.monotonic", lambda: now[0])
    monkeypatch.setattr(
        "spreadsheet_harness.comparison.run_arm",
        lambda **_: AgentResult("done", 1, 0, {}, "response"),
    )

    def slow_score(*_: Any, **__: Any) -> Any:
        now[0] = 102.0
        return type("Comparison", (), {"passed": True, "to_dict": lambda self: {}})()

    monkeypatch.setattr(
        "spreadsheet_harness.comparison.compare_workbooks_chartsheet_safe",
        slow_score,
    )
    runner = ComparisonBenchmarkRunner(
        ProviderConfig("https://example.test/v1", "not-a-real-key", "test-model"),
        tmp_path / "deadline",
        skill_registry=SkillRegistry([]),
        task_timeout_seconds=1,
        recalculate=False,
    )

    row = runner._run_one(task, "bare", comparison_manifest_sha256="a" * 64)

    assert row["status"] == "error"
    assert row["error_category"] == "task_timeout"
    assert row["budget"]["termination"]["reason"] == "max_elapsed_seconds"


def test_deadline_after_token_termination_records_timeout_row(
    tmp_path: Path, monkeypatch: Any
) -> None:
    task = _tasks(tmp_path)[0]
    now = [100.0]
    monkeypatch.setattr("spreadsheet_harness.comparison.monotonic", lambda: now[0])
    monkeypatch.setattr("spreadsheet_harness.budget.time.monotonic", lambda: now[0])

    def token_failure(**kwargs: Any) -> AgentResult:
        budget = kwargs["budget"]
        reservation = budget.begin_model_call(stage="solve")
        with pytest.raises(AgentBudgetError):
            budget.record_response(
                reservation,
                {"total_tokens": 11},
                stage="solve",
            )
        now[0] = 102.0
        raise AgentExecutionFailure(
            "token budget exhausted",
            reason="budget_exhausted",
            agent_result=AgentResult(
                "token budget exhausted",
                1,
                0,
                {"total_tokens": 11},
                "response",
            ),
        )

    monkeypatch.setattr("spreadsheet_harness.comparison.run_arm", token_failure)
    runner = ComparisonBenchmarkRunner(
        ProviderConfig("https://example.test/v1", "not-a-real-key", "test-model"),
        tmp_path / "deadline-after-token-budget",
        skill_registry=SkillRegistry([]),
        arms=("bare",),
        max_model_calls=3,
        max_turns_per_arm=3,
        max_total_tokens=10,
        task_timeout_seconds=1,
        recalculate=False,
    )

    row = runner._run_one(task, "bare", comparison_manifest_sha256="a" * 64)

    assert row["status"] == "error"
    assert row["error_category"] == "task_timeout"
    assert row["error_type"] == "AgentTimeoutError"
    assert row["budget"]["termination"]["reason"] == "max_total_tokens"


def test_comparison_fails_before_writes_when_strict_isolation_is_unavailable(
    tmp_path: Path,
    monkeypatch: Any,
) -> None:
    tasks = _tasks(tmp_path)

    def unavailable(*_: Any, **__: Any) -> dict[str, str]:
        raise CodeIsolationError("bwrap probe failed")

    monkeypatch.setattr(
        "spreadsheet_harness.comparison.ensure_strict_code_isolation",
        unavailable,
    )
    output = tmp_path / "strict-failure"
    runner = ComparisonBenchmarkRunner(
        ProviderConfig(
            "https://example.test/v1",
            "not-a-real-key",
            "test-model",
            request_interval_seconds=20.0,
        ),
        output,
        skill_registry=SkillRegistry([]),
        recalculate=False,
    )

    with pytest.raises(CodeIsolationError, match="bwrap probe failed"):
        runner.run(tasks)

    assert not runner.manifest_path.exists()
    assert not runner.results_path.exists()
