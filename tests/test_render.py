from __future__ import annotations

import json
import warnings
import zipfile
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from PIL import Image

import spreadsheet_harness.render as render_module
from spreadsheet_harness.errors import RecalculationIntegrityError, RenderError
from spreadsheet_harness.render import (
    PNG_SIGNATURE,
    RenderPage,
    find_libreoffice,
    isolated_user_profile,
    libreoffice_command,
    read_png,
    recalculate_workbook,
    render_workbook,
    sha256_file,
    sheet_inventory_identity,
)


def _save_png(path: Path, *, color: str = "white") -> bytes:
    Image.new("RGB", (8, 6), color=color).save(path, format="PNG")
    return path.read_bytes()


def _save_workbook(path: Path, *, two_sheets: bool = True) -> None:
    workbook = Workbook()
    first = workbook.active
    first.title = "Data"
    first["A1"] = "Name"
    first["B1"] = "Value"
    first.append(["one", 1])
    first.append(["two", 2])
    first["B4"] = "=SUM(B2:B3)"
    if two_sheets:
        second = workbook.create_sheet("Summary")
        second["A1"] = "Total"
        second["B1"] = "=Data!B4"
    workbook.save(path)
    workbook.close()


def _minimal_workbook_xml(
    sheets: str,
    *,
    namespace: str = "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    relationship_namespace: str = (
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    ),
) -> bytes:
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<workbook xmlns="{namespace}" xmlns:r="{relationship_namespace}">'
        f"<sheets>{sheets}</sheets>"
        "</workbook>"
    ).encode()


def _minimal_relationships_xml(*, strict: bool = False) -> bytes:
    type_prefix = (
        "http://purl.oclc.org/ooxml/officeDocument/relationships/"
        if strict
        else "http://schemas.openxmlformats.org/officeDocument/2006/relationships/"
    )
    return (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        f'<Relationship Id="rId1" Type="{type_prefix}worksheet" '
        'Target="worksheets/sheet1.xml"/>'
        f'<Relationship Id="rId2" Type="{type_prefix}chartsheet" '
        'Target="chartsheets/sheet1.xml"/>'
        f'<Relationship Id="rel-data" Type="{type_prefix}worksheet" '
        'Target="worksheets/data.xml"/>'
        f'<Relationship Id="rel-chart" Type="{type_prefix}chartsheet" '
        'Target="chartsheets/chart.xml"/>'
        '</Relationships>'
    ).encode()


def _write_minimal_ooxml(
    path: Path,
    workbook_xml: bytes | None,
    *,
    relationships_xml: bytes | None = None,
) -> None:
    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as package:
        package.writestr("[Content_Types].xml", "<Types/>")
        if workbook_xml is not None:
            package.writestr("xl/workbook.xml", workbook_xml)
        package.writestr(
            "xl/_rels/workbook.xml.rels",
            relationships_xml
            if relationships_xml is not None
            else _minimal_relationships_xml(),
        )


def _replace_zip_part(path: Path, part_name: str, replacement: bytes) -> None:
    rewritten = path.with_name(f"{path.stem}-rewritten{path.suffix}")
    with zipfile.ZipFile(path) as source:
        parts = [
            (member.filename, replacement if member.filename == part_name else source.read(member))
            for member in source.infolist()
        ]
    with zipfile.ZipFile(rewritten, "w", compression=zipfile.ZIP_DEFLATED) as target:
        for name, content in parts:
            target.writestr(name, content)
    rewritten.replace(path)


def test_isolated_user_profile_is_unique_and_removed() -> None:
    paths: list[Path] = []
    uris: list[str] = []
    for _ in range(2):
        with isolated_user_profile() as (profile, uri):
            paths.append(profile)
            uris.append(uri)
            assert profile.is_dir()
            assert uri == profile.as_uri()
            assert uri.startswith("file://")
        assert not profile.exists()

    assert paths[0] != paths[1]
    assert uris[0] != uris[1]


def test_libreoffice_command_contains_private_profile(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    source.touch()
    output = tmp_path / "output"
    profile_uri = (tmp_path / "profile with spaces").resolve().as_uri()

    command = libreoffice_command("soffice", source, output, "pdf", profile_uri)

    assert command[0] == "soffice"
    assert "--headless" in command
    assert f"-env:UserInstallation={profile_uri}" in command
    assert command[-1] == str(source)
    assert command[command.index("--outdir") + 1] == str(output)


def test_read_png_returns_original_bytes(tmp_path: Path) -> None:
    png = tmp_path / "original.png"
    original = _save_png(png, color="navy")

    loaded = read_png(png)

    assert loaded == original
    assert loaded.startswith(PNG_SIGNATURE)


def test_read_png_rejects_other_files(tmp_path: Path) -> None:
    text = tmp_path / "not-an-image.png"
    text.write_text("not png", encoding="utf-8")

    with pytest.raises(RenderError, match="Not a PNG"):
        read_png(text)


def test_sheet_inventory_identity_includes_chart_sheets_without_openpyxl_loader(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = tmp_path / "charts.xlsx"
    workbook = Workbook()
    workbook.active.title = "Data"
    chart_sheet = workbook.create_chartsheet("Chart")
    chart_sheet.sheet_state = "hidden"
    workbook.save(source)
    workbook.close()

    def forbidden_loader(*args: object, **kwargs: object) -> object:
        raise AssertionError("sheet inventory must not use the openpyxl workbook loader")

    monkeypatch.setattr(render_module, "load_workbook", forbidden_loader)
    identity = sheet_inventory_identity(source)

    assert identity["sheets"] == [
        {"index": 0, "kind": "worksheet", "name": "Data", "visibility": "visible"},
        {"index": 1, "kind": "chartsheet", "name": "Chart", "visibility": "hidden"},
    ]


def test_sheet_inventory_accepts_strict_namespace_and_defaults_visible(
    tmp_path: Path,
) -> None:
    source = tmp_path / "strict.xlsx"
    xml = _minimal_workbook_xml(
        (
            '<sheet name="Data" sheetId="7" r:id="rel-data"/>'
            '<sheet name="Chart" sheetId="11" state="veryHidden" r:id="rel-chart"/>'
        ),
        namespace="http://purl.oclc.org/ooxml/spreadsheetml/main",
        relationship_namespace="http://purl.oclc.org/ooxml/officeDocument/relationships",
    )
    _write_minimal_ooxml(
        source,
        xml,
        relationships_xml=_minimal_relationships_xml(strict=True),
    )

    identity = sheet_inventory_identity(source)

    assert identity["sheets"] == [
        {"index": 0, "kind": "worksheet", "name": "Data", "visibility": "visible"},
        {
            "index": 1,
            "kind": "chartsheet",
            "name": "Chart",
            "visibility": "veryHidden",
        },
    ]


@pytest.mark.parametrize(
    ("workbook_xml", "message"),
    [
        (b"<workbook", "malformed"),
        (
            _minimal_workbook_xml(
                '<sheet name="Data" sheetId="1" r:id="rId1"/>',
                namespace="urn:not-spreadsheetml",
            ),
            "unsupported SpreadsheetML namespace",
        ),
        (
            _minimal_workbook_xml(
                '<sheet name="Data" sheetId="1" r:id="rId1"/>'
                '<sheet name="data" sheetId="2" r:id="rId2"/>'
            ),
            "duplicate sheet names",
        ),
        (
            _minimal_workbook_xml(
                '<sheet name="Data" sheetId="1" r:id="rId1"/>'
                '<sheet name="Chart" sheetId="1" r:id="rId2"/>'
            ),
            "duplicate sheetId",
        ),
        (
            _minimal_workbook_xml(
                '<sheet name="Data" sheetId="1" r:id="rId1"/>'
                '<sheet name="Chart" sheetId="2" r:id="rId1"/>'
            ),
            "duplicate sheet relationships",
        ),
        (
            _minimal_workbook_xml(
                '<sheet name="Data" sheetId="1" state="sometimes" r:id="rId1"/>'
            ),
            "invalid visibility",
        ),
        (
            _minimal_workbook_xml('<sheet name="Data" sheetId="1"/>'),
            "missing relationship identifier",
        ),
        (
            _minimal_workbook_xml(
                '<sheet name="Data" sheetId="1" r:id="rId1"/>',
                relationship_namespace="urn:not-office-relationships",
            ),
            "missing relationship identifier",
        ),
    ],
)
def test_sheet_inventory_rejects_malformed_namespace_and_duplicate_records(
    tmp_path: Path,
    workbook_xml: bytes,
    message: str,
) -> None:
    source = tmp_path / "invalid.xlsx"
    _write_minimal_ooxml(source, workbook_xml)

    with pytest.raises(RenderError, match=message):
        sheet_inventory_identity(source)


def test_sheet_inventory_rejects_missing_and_duplicate_workbook_parts(
    tmp_path: Path,
) -> None:
    missing = tmp_path / "missing.xlsx"
    _write_minimal_ooxml(missing, None)
    with pytest.raises(RenderError, match="exactly one xl/workbook.xml; found 0"):
        sheet_inventory_identity(missing)

    duplicate = tmp_path / "duplicate.xlsx"
    xml = _minimal_workbook_xml('<sheet name="Data" sheetId="1" r:id="rId1"/>')
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        with zipfile.ZipFile(duplicate, "w") as package:
            package.writestr("xl/workbook.xml", xml)
            package.writestr("xl/workbook.xml", xml)
    with pytest.raises(RenderError, match="exactly one xl/workbook.xml; found 2"):
        sheet_inventory_identity(duplicate)


def test_sheet_inventory_rejects_missing_duplicate_and_malformed_relationship_parts(
    tmp_path: Path,
) -> None:
    xml = _minimal_workbook_xml('<sheet name="Data" sheetId="1" r:id="rId1"/>')
    missing = tmp_path / "missing-rels.xlsx"
    with zipfile.ZipFile(missing, "w") as package:
        package.writestr("xl/workbook.xml", xml)
    with pytest.raises(RenderError, match="exactly one xl/_rels/workbook.xml.rels; found 0"):
        sheet_inventory_identity(missing)

    duplicate = tmp_path / "duplicate-rels.xlsx"
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        with zipfile.ZipFile(duplicate, "w") as package:
            package.writestr("xl/workbook.xml", xml)
            package.writestr("xl/_rels/workbook.xml.rels", _minimal_relationships_xml())
            package.writestr("xl/_rels/workbook.xml.rels", _minimal_relationships_xml())
    with pytest.raises(RenderError, match="exactly one xl/_rels/workbook.xml.rels; found 2"):
        sheet_inventory_identity(duplicate)

    malformed = tmp_path / "malformed-rels.xlsx"
    _write_minimal_ooxml(malformed, xml, relationships_xml=b"<Relationships")
    with pytest.raises(RenderError, match="workbook relationships XML is malformed"):
        sheet_inventory_identity(malformed)


@pytest.mark.parametrize(
    ("relationships", "message"),
    [
        (
            '<Relationships xmlns="urn:not-package-relationships">'
            '<Relationship Id="rId1" Type="urn:worksheet" Target="sheet.xml"/>'
            '</Relationships>',
            "unsupported namespace",
        ),
        (
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="other" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
            'Target="worksheets/sheet1.xml"/>'
            '</Relationships>',
            "missing workbook relationship",
        ),
        (
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
            'Target="https://example.test/sheet.xml" TargetMode="External"/>'
            '</Relationships>',
            "must target an internal",
        ),
        (
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" '
            'Target="styles.xml"/>'
            '</Relationships>',
            "unsupported sheet type",
        ),
    ],
)
def test_sheet_inventory_rejects_invalid_relationship_resolution(
    tmp_path: Path,
    relationships: str,
    message: str,
) -> None:
    source = tmp_path / "invalid-relationship.xlsx"
    xml = _minimal_workbook_xml('<sheet name="Data" sheetId="1" r:id="rId1"/>')
    _write_minimal_ooxml(source, xml, relationships_xml=relationships.encode())

    with pytest.raises(RenderError, match=message):
        sheet_inventory_identity(source)


def test_sheet_inventory_rejects_dtd_entities_and_damaged_zip(tmp_path: Path) -> None:
    malicious = tmp_path / "entity.xlsx"
    xml = _minimal_workbook_xml(
        '<sheet name="&payload;" sheetId="1" r:id="rId1"/>'
    ).replace(
        b'<workbook xmlns=',
        b'<!DOCTYPE workbook [<!ENTITY payload "expanded">]><workbook xmlns=',
        1,
    )
    _write_minimal_ooxml(malicious, xml)
    with pytest.raises(RenderError, match="DTD or entity"):
        sheet_inventory_identity(malicious)

    damaged = tmp_path / "damaged.xlsx"
    damaged.write_bytes(b"PK\x03\x04truncated")
    with pytest.raises(RenderError, match="Could not read OOXML workbook package"):
        sheet_inventory_identity(damaged)


def test_sheet_inventory_rejects_wrong_namespace_sheet_container(tmp_path: Path) -> None:
    source = tmp_path / "mixed-namespace.xlsx"
    xml = (
        b'<?xml version="1.0" encoding="UTF-8"?>'
        b'<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        b'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" '
        b'xmlns:evil="urn:evil">'
        b'<evil:sheets><evil:sheet name="Data" sheetId="1" r:id="rId1"/></evil:sheets>'
        b'</workbook>'
    )
    _write_minimal_ooxml(source, xml)

    with pytest.raises(RenderError, match="exactly one sheets element"):
        sheet_inventory_identity(source)


def test_sheet_inventory_rejects_missing_or_repeated_sheets_elements(tmp_path: Path) -> None:
    namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    relationship_namespace = (
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    )
    missing = tmp_path / "missing-sheets.xlsx"
    _write_minimal_ooxml(
        missing,
        (
            f'<workbook xmlns="{namespace}" xmlns:r="{relationship_namespace}"/>'
        ).encode(),
    )
    with pytest.raises(RenderError, match="exactly one sheets element; found 0"):
        sheet_inventory_identity(missing)

    repeated = tmp_path / "repeated-sheets.xlsx"
    _write_minimal_ooxml(
        repeated,
        (
            f'<workbook xmlns="{namespace}" xmlns:r="{relationship_namespace}">'
            '<sheets><sheet name="Data" sheetId="1" r:id="rId1"/></sheets>'
            '<sheets><sheet name="Chart" sheetId="2" r:id="rId2"/></sheets>'
            '</workbook>'
        ).encode(),
    )
    with pytest.raises(RenderError, match="exactly one sheets element; found 2"):
        sheet_inventory_identity(repeated)


def test_render_page_dict_has_view_image_fields(tmp_path: Path) -> None:
    png = tmp_path / "page.png"
    _save_png(png)
    page = RenderPage(
        index=3,
        path=png,
        sha256=sha256_file(png),
        width=8,
        height=6,
        sheet="Data",
        sheet_page=2,
    )

    payload = page.to_dict(relative_to=tmp_path)

    assert payload["image_path"] == str(png.resolve())
    assert payload["path"] == "page.png"
    assert payload["index"] == 3
    assert payload["page"] == 2
    assert payload["sheet"] == "Data"
    assert payload["sheet_page"] == 2
    assert payload["width"] == 8
    assert payload["height"] == 6
    assert len(payload["sha256"]) == 64


def test_single_sheet_copy_hides_only_disposable_copy(tmp_path: Path) -> None:
    source = tmp_path / "book.xlsx"
    target = tmp_path / "single.xlsx"
    _save_workbook(source)
    before = source.read_bytes()

    render_module._make_single_sheet_copy(source, target, "Summary")

    copied = load_workbook(target)
    try:
        assert copied["Data"].sheet_state == "hidden"
        assert copied["Summary"].sheet_state == "visible"
        assert copied.active.title == "Summary"
    finally:
        copied.close()
    assert source.read_bytes() == before


def test_pymupdf_rasterization_produces_unmodified_png(tmp_path: Path) -> None:
    module = render_module._pymupdf_module()
    pdf = tmp_path / "source.pdf"
    document = module.open()
    page = document.new_page(width=200, height=100)
    page.insert_text((20, 50), "Spreadsheet")
    document.save(str(pdf))
    document.close()

    pages = render_module._rasterize_pdf(
        pdf,
        tmp_path / "png",
        dpi=72,
        filename_prefix="page",
        sheet="Data",
    )

    assert len(pages) == 1
    assert read_png(pages[0].path).startswith(PNG_SIGNATURE)
    assert pages[0].width == 200
    assert pages[0].height == 100
    assert pages[0].sheet == "Data"
    assert pages[0].sheet_page == 1


def test_per_sheet_failure_falls_back_and_records_manifest(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = tmp_path / "book.xlsx"
    _save_workbook(source)
    before = source.read_bytes()

    monkeypatch.setattr(render_module, "find_libreoffice", lambda explicit=None: "/fake/soffice")
    monkeypatch.setattr(render_module, "libreoffice_version", lambda binary: "LibreOffice test")
    monkeypatch.setattr(render_module, "pymupdf_version", lambda: "PyMuPDF test")

    def fail_per_sheet(*args: object, **kwargs: object) -> list[object]:
        raise RenderError("deliberate per-sheet failure")

    def fake_whole(
        source_copy: Path,
        work_dir: Path,
        **kwargs: object,
    ) -> list[render_module._RasterizedPage]:
        assert source_copy != source
        png = work_dir / "png" / "workbook-page-0001.png"
        png.parent.mkdir(parents=True)
        _save_png(png)
        return [
            render_module._RasterizedPage(
                path=png,
                width=8,
                height=6,
                sheet=None,
                sheet_page=None,
            )
        ]

    monkeypatch.setattr(render_module, "_render_per_sheet", fail_per_sheet)
    monkeypatch.setattr(render_module, "_render_whole_workbook", fake_whole)

    result = render_workbook(source, tmp_path / "rendered")
    manifest = json.loads(result.manifest_path.read_text(encoding="utf-8"))

    assert source.read_bytes() == before
    assert result.mode == "whole_workbook"
    assert manifest["backend"] == "libreoffice-headless+pymupdf"
    assert manifest["version"]["libreoffice"] == "LibreOffice test"
    assert manifest["hash"] == sha256_file(source)
    assert manifest["page_count"] == 1
    assert manifest["fallback"]["from"] == "per_sheet"
    assert "deliberate per-sheet failure" in manifest["fallback"]["reason"]
    page_payload = manifest["pages"][0]
    assert Path(page_payload["image_path"]).is_absolute()
    assert page_payload["sheet"] is None
    assert page_payload["page"] == 1
    assert read_png(page_payload["image_path"]).startswith(PNG_SIGNATURE)


def test_recalculation_uses_private_copy_and_atomic_replace(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = tmp_path / "working.xlsx"
    _save_workbook(source)
    original_hash = sha256_file(source)
    seen_sources: list[Path] = []

    monkeypatch.setattr(render_module, "find_libreoffice", lambda explicit=None: "/fake/soffice")
    monkeypatch.setattr(render_module, "libreoffice_version", lambda binary: "LibreOffice test")

    def fake_convert(
        source_copy: Path,
        output_dir: Path,
        **kwargs: object,
    ) -> Path:
        seen_sources.append(source_copy.resolve())
        assert source_copy.resolve() != source.resolve()
        output_dir.mkdir(parents=True)
        converted = output_dir / "working.xlsx"
        converted.write_bytes(source_copy.read_bytes())
        return converted

    monkeypatch.setattr(render_module, "_convert_with_libreoffice", fake_convert)

    metadata = recalculate_workbook(source, source)

    assert seen_sources and seen_sources[0] != source.resolve()
    assert sha256_file(source) == original_hash
    assert metadata["backend"] == "libreoffice-headless"
    assert metadata["version"] == "LibreOffice test"
    assert metadata["source_sha256"] == original_hash
    assert metadata["output_sha256"] == original_hash
    assert metadata["destination_path"] == str(source.resolve())
    assert metadata["atomic_replace"] is True
    assert metadata["published"] is True
    integrity = metadata["sheet_inventory_integrity"]
    assert integrity["matched"] is True
    assert integrity["pre"] == integrity["post"]


def test_recalculation_validates_chartsheet_package_without_openpyxl_loader(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = tmp_path / "charts.xlsx"
    workbook = Workbook()
    workbook.active.title = "Data"
    workbook.create_chartsheet("Chart").sheet_state = "hidden"
    workbook.save(source)
    workbook.close()

    monkeypatch.setattr(render_module, "find_libreoffice", lambda explicit=None: "/fake/soffice")
    monkeypatch.setattr(render_module, "libreoffice_version", lambda binary: "LibreOffice test")

    def fake_convert(source_copy: Path, output_dir: Path, **kwargs: object) -> Path:
        output_dir.mkdir(parents=True)
        converted = output_dir / source.name
        converted.write_bytes(source_copy.read_bytes())
        return converted

    def forbidden_loader(*args: object, **kwargs: object) -> object:
        raise AssertionError("recalculation validation must not load chartsheets with openpyxl")

    monkeypatch.setattr(render_module, "_convert_with_libreoffice", fake_convert)
    monkeypatch.setattr(render_module, "load_workbook", forbidden_loader)

    metadata = recalculate_workbook(source, source)

    assert metadata["sheet_inventory_integrity"]["matched"] is True
    assert metadata["sheet_inventory_integrity"]["pre"]["sheets"] == [
        {"index": 0, "kind": "worksheet", "name": "Data", "visibility": "visible"},
        {"index": 1, "kind": "chartsheet", "name": "Chart", "visibility": "hidden"},
    ]


def test_recalculation_detects_chartsheet_to_worksheet_kind_drift(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = tmp_path / "charts.xlsx"
    workbook = Workbook()
    workbook.active.title = "Data"
    workbook.create_chartsheet("Chart").sheet_state = "hidden"
    workbook.save(source)
    workbook.close()
    before = source.read_bytes()

    monkeypatch.setattr(render_module, "find_libreoffice", lambda explicit=None: "/fake/soffice")
    monkeypatch.setattr(render_module, "libreoffice_version", lambda binary: "LibreOffice test")

    def fake_convert(source_copy: Path, output_dir: Path, **kwargs: object) -> Path:
        output_dir.mkdir(parents=True)
        converted = output_dir / source.name
        converted.write_bytes(source_copy.read_bytes())
        with zipfile.ZipFile(converted) as package:
            relationships = package.read("xl/_rels/workbook.xml.rels")
        relationships = relationships.replace(b"/chartsheet", b"/worksheet")
        _replace_zip_part(converted, "xl/_rels/workbook.xml.rels", relationships)
        return converted

    monkeypatch.setattr(render_module, "_convert_with_libreoffice", fake_convert)

    with pytest.raises(RecalculationIntegrityError) as caught:
        recalculate_workbook(source, source)

    integrity = caught.value.evidence["sheet_inventory_integrity"]
    assert source.read_bytes() == before
    assert integrity["matched"] is False
    assert [sheet["name"] for sheet in integrity["pre"]["sheets"]] == ["Data", "Chart"]
    assert [sheet["name"] for sheet in integrity["post"]["sheets"]] == ["Data", "Chart"]
    assert [sheet["kind"] for sheet in integrity["pre"]["sheets"]] == [
        "worksheet",
        "chartsheet",
    ]
    assert [sheet["kind"] for sheet in integrity["post"]["sheets"]] == [
        "worksheet",
        "worksheet",
    ]


@pytest.mark.parametrize("suffix", [".ods", ".xls", ".csv"])
def test_recalculation_preserves_non_ooxml_format_support(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    suffix: str,
) -> None:
    source = tmp_path / f"working{suffix}"
    source.write_bytes(b"non-ooxml-source")

    monkeypatch.setattr(render_module, "find_libreoffice", lambda explicit=None: "/fake/soffice")
    monkeypatch.setattr(render_module, "libreoffice_version", lambda binary: "LibreOffice test")

    def fake_convert(source_copy: Path, output_dir: Path, **kwargs: object) -> Path:
        output_dir.mkdir(parents=True)
        converted = output_dir / source.name
        converted.write_bytes(b"recalculated-non-ooxml")
        return converted

    monkeypatch.setattr(render_module, "_convert_with_libreoffice", fake_convert)

    metadata = recalculate_workbook(source, source)

    assert source.read_bytes() == b"recalculated-non-ooxml"
    assert metadata["published"] is True
    assert metadata["sheet_inventory_integrity"] == {
        "schema_version": 2,
        "policy": "exact-ordered-sheet-kind-name-visibility-v2",
        "enforced": False,
        "matched": None,
        "pre": None,
        "post": None,
        "reason": "source-or-destination-is-not-ooxml",
    }


@pytest.mark.parametrize("mutation", ["rename", "reorder", "visibility"])
def test_recalculation_fails_closed_when_sheet_identity_changes(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    mutation: str,
) -> None:
    source = tmp_path / "working.xlsx"
    _save_workbook(source)
    before = source.read_bytes()
    before_identity = sheet_inventory_identity(source)

    monkeypatch.setattr(render_module, "find_libreoffice", lambda explicit=None: "/fake/soffice")
    monkeypatch.setattr(render_module, "libreoffice_version", lambda binary: "LibreOffice test")

    def fake_convert(
        source_copy: Path,
        output_dir: Path,
        **kwargs: object,
    ) -> Path:
        output_dir.mkdir(parents=True)
        converted = output_dir / "working.xlsx"
        workbook = load_workbook(source_copy)
        try:
            if mutation == "rename":
                workbook["Summary"].title = "Changed"
            elif mutation == "reorder":
                workbook.move_sheet(workbook["Summary"], offset=-1)
            else:
                workbook["Summary"].sheet_state = "hidden"
            workbook.save(converted)
        finally:
            workbook.close()
        return converted

    monkeypatch.setattr(render_module, "_convert_with_libreoffice", fake_convert)

    with pytest.raises(RecalculationIntegrityError) as caught:
        recalculate_workbook(source, source)

    evidence = caught.value.evidence
    integrity = evidence["sheet_inventory_integrity"]
    failure_artifact = Path(evidence["failure_artifact_path"])
    assert source.read_bytes() == before
    assert evidence["atomic_replace"] is False
    assert evidence["published"] is False
    assert integrity["matched"] is False
    assert integrity["pre"] == before_identity
    assert integrity["post"]["sheets"] != before_identity["sheets"]
    assert failure_artifact.is_file()
    assert sha256_file(failure_artifact) == evidence["output_sha256"]
    assert sheet_inventory_identity(failure_artifact) == integrity["post"]


def test_recalculation_identity_failure_survives_evidence_publish_error(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    source = tmp_path / "working.xlsx"
    _save_workbook(source)
    before = source.read_bytes()

    monkeypatch.setattr(render_module, "find_libreoffice", lambda explicit=None: "/fake/soffice")
    monkeypatch.setattr(render_module, "libreoffice_version", lambda binary: "LibreOffice test")

    def fake_convert(source_copy: Path, output_dir: Path, **kwargs: object) -> Path:
        output_dir.mkdir(parents=True)
        converted = output_dir / "working.xlsx"
        workbook = load_workbook(source_copy)
        try:
            workbook["Summary"].title = "Changed"
            workbook.save(converted)
        finally:
            workbook.close()
        return converted

    monkeypatch.setattr(render_module, "_convert_with_libreoffice", fake_convert)

    def fail_publish(*_: object) -> Path:
        raise OSError("disk unavailable")

    monkeypatch.setattr(
        render_module,
        "_publish_recalculation_failure_artifact",
        fail_publish,
    )

    with pytest.raises(RecalculationIntegrityError) as caught:
        recalculate_workbook(source, source)

    assert source.read_bytes() == before
    assert caught.value.evidence["sheet_inventory_integrity"]["matched"] is False
    assert caught.value.evidence["failure_artifact_path"] is None
    assert caught.value.evidence["failure_artifact_error_type"] == "OSError"


@pytest.mark.skipif(find_libreoffice() is None, reason="LibreOffice is not installed")
def test_libreoffice_recalculation_updates_formula_without_changing_sheet_identity(
    tmp_path: Path,
) -> None:
    source = tmp_path / "formula.xlsx"
    _save_workbook(source, two_sheets=False)

    metadata = recalculate_workbook(source, source)

    recalculated = load_workbook(source, data_only=True, read_only=True)
    try:
        assert recalculated["Data"]["B4"].value == 3
    finally:
        recalculated.close()
    integrity = metadata["sheet_inventory_integrity"]
    assert integrity["matched"] is True
    assert integrity["pre"]["sheets"] == integrity["post"]["sheets"]


@pytest.mark.skipif(find_libreoffice() is None, reason="LibreOffice is not installed")
def test_libreoffice_render_integration_preserves_source(tmp_path: Path) -> None:
    source = tmp_path / "integration.xlsx"
    _save_workbook(source)
    before_hash = sha256_file(source)

    result = render_workbook(source, tmp_path / "rendered", dpi=72)

    assert sha256_file(source) == before_hash
    assert result.pages
    assert result.backend == "libreoffice-headless+pymupdf"
    assert result.version["libreoffice"] != "unknown"
    if result.mode == "per_sheet":
        assert {page.sheet for page in result.pages} == {"Data", "Summary"}
    for page in result.pages:
        assert read_png(page.path).startswith(PNG_SIGNATURE)
