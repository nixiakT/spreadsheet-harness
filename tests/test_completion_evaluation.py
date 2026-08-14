from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from spreadsheet_harness.benchmark import compare_workbooks
from spreadsheet_harness.completion_attempt import CompletionAttemptLedger
from spreadsheet_harness.completion_evaluation import (
    CompletionEvaluationError,
    CompletionEvaluationRecord,
    audit_completion_evaluation,
    evaluate_completion_attempt,
    evaluate_completion_attempts,
)
from spreadsheet_harness.evidence_contract import ArtifactRef


def _book(path: Path, value: int) -> None:
    workbook = Workbook()
    workbook.active["A1"] = value
    workbook.save(path)
    workbook.close()


def _sha256(path: Path) -> str:
    import hashlib

    return hashlib.sha256(path.read_bytes()).hexdigest()


def _evaluator(golden: Path):
    return lambda candidate: compare_workbooks(golden, candidate, "A1")


def test_posthoc_evaluation_is_canonical_and_freshly_auditable(tmp_path: Path) -> None:
    workspace = tmp_path / "workspace"
    workspace.mkdir()
    candidate = workspace / "candidate.xlsx"
    golden = tmp_path / "golden.xlsx"
    _book(candidate, 7)
    _book(golden, 7)
    attempt = CompletionAttemptLedger(workspace).capture(
        candidate,
        ArtifactRef(3, _sha256(candidate)),
        stage="solve",
        turn=9,
        response_id="response-1",
        call_id="call-1",
    )

    evaluation = evaluate_completion_attempt(
        workspace,
        attempt,
        _evaluator(golden),
        evaluator_id="spreadsheetbench-corrected-value-v1",
    )

    assert evaluation.passed is True
    assert evaluation.attempt_id == 1
    assert evaluation.artifact == attempt.artifact
    assert evaluation.completion_record_sha256 == attempt.record_sha256
    assert evaluation.to_dict()["assurance"] == {
        "timing": "posthoc-after-agent-termination",
        "fed_back_to_model": False,
        "contains_evaluator_outcome": True,
        "is_digital_signature": False,
        "proves_task_correctness_beyond_named_evaluator": False,
    }
    assert (
        CompletionEvaluationRecord.from_dict(json.loads(json.dumps(evaluation.to_dict()))).to_dict()
        == evaluation.to_dict()
    )
    assert audit_completion_evaluation(
        workspace,
        attempt.to_dict(),
        evaluation.to_dict(),
        _evaluator(golden),
    ).valid


def test_fresh_audit_rejects_tampered_outcome_or_digest(tmp_path: Path) -> None:
    workspace = tmp_path / "workspace"
    workspace.mkdir()
    candidate = workspace / "candidate.xlsx"
    golden = tmp_path / "golden.xlsx"
    _book(candidate, 1)
    _book(golden, 2)
    attempt = CompletionAttemptLedger(workspace).capture(
        candidate,
        ArtifactRef(1, _sha256(candidate)),
        stage="solve",
        turn=2,
        response_id="response-1",
        call_id="call-1",
    )
    evaluation = evaluate_completion_attempt(
        workspace,
        attempt,
        _evaluator(golden),
        evaluator_id="test-evaluator-v1",
    ).to_dict()
    evaluation["comparison"]["passed"] = True

    audit = audit_completion_evaluation(
        workspace,
        attempt,
        evaluation,
        _evaluator(golden),
    )

    assert audit.valid is False
    assert "digest mismatch" in audit.reasons[0]


def test_evaluator_that_mutates_snapshot_fails_closed(tmp_path: Path) -> None:
    workspace = tmp_path / "workspace"
    workspace.mkdir()
    candidate = workspace / "candidate.xlsx"
    _book(candidate, 1)
    attempt = CompletionAttemptLedger(workspace).capture(
        candidate,
        ArtifactRef(1, _sha256(candidate)),
        stage="solve",
        turn=2,
        response_id="response-1",
        call_id="call-1",
    )

    def mutating_evaluator(path: Path):
        path.chmod(0o600)
        workbook = load_workbook(path)
        workbook.active["A1"] = 9
        workbook.save(path)
        workbook.close()
        return compare_workbooks(candidate, path, "A1")

    with pytest.raises(CompletionEvaluationError, match="mutated"):
        evaluate_completion_attempt(
            workspace,
            attempt,
            mutating_evaluator,
            evaluator_id="bad-evaluator-v1",
        )


def test_evaluator_mutation_is_detected_even_when_evaluator_raises(
    tmp_path: Path,
) -> None:
    workspace = tmp_path / "workspace"
    workspace.mkdir()
    candidate = workspace / "candidate.xlsx"
    _book(candidate, 1)
    attempt = CompletionAttemptLedger(workspace).capture(
        candidate,
        ArtifactRef(1, _sha256(candidate)),
        stage="solve",
        turn=2,
        response_id="response-1",
        call_id="call-1",
    )

    def mutating_failure(path: Path):
        path.chmod(0o600)
        path.write_bytes(b"not-a-workbook")
        raise RuntimeError("scorer failed after writing")

    with pytest.raises(CompletionEvaluationError, match="mutated") as caught:
        evaluate_completion_attempt(
            workspace,
            attempt,
            mutating_failure,
            evaluator_id="bad-evaluator-v1",
        )

    assert isinstance(caught.value.__cause__, RuntimeError)


def test_batch_requires_canonical_attempt_sequence(tmp_path: Path) -> None:
    workspace = tmp_path / "workspace"
    workspace.mkdir()
    candidate = workspace / "candidate.xlsx"
    golden = tmp_path / "golden.xlsx"
    _book(candidate, 1)
    _book(golden, 1)
    ledger = CompletionAttemptLedger(workspace)
    first = ledger.capture(
        candidate,
        ArtifactRef(1, _sha256(candidate)),
        stage="solve",
        turn=2,
        response_id="response-1",
        call_id="call-1",
    )
    second = ledger.capture(
        candidate,
        ArtifactRef(2, _sha256(candidate)),
        stage="solve",
        turn=4,
        response_id="response-2",
        call_id="call-2",
    )

    evaluations = evaluate_completion_attempts(
        workspace,
        [first, second],
        _evaluator(golden),
        evaluator_id="test-evaluator-v1",
    )
    assert [item.attempt_id for item in evaluations] == [1, 2]

    with pytest.raises(CompletionEvaluationError, match="ordered, unique"):
        evaluate_completion_attempts(
            workspace,
            [second],
            _evaluator(golden),
            evaluator_id="test-evaluator-v1",
        )
