from __future__ import annotations

import random
import shutil
import warnings
from collections.abc import Callable
from dataclasses import dataclass
from pathlib import Path
from xml.etree import ElementTree as ET
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table, TableStyleInfo
from oracles import OracleCell, OracleDiff, diff_ooxml

from spreadsheet_harness.evidence_contract import EffectKind, EvidenceScope
from spreadsheet_harness.workbook_diff import WorkbookEffectDiff, diff_workbooks

_ORACLE_SEEDS = (41, 42, 43)
_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_DRAWING = "http://schemas.openxmlformats.org/drawingml/2006/main"
_LIBREOFFICE = "http://schemas.libreoffice.org/"
_CALC_A1_URI = "{7626C862-2A13-11E5-B345-FEFF819CDC9F}"
_CONTENT_TYPES = "http://schemas.openxmlformats.org/package/2006/content-types"
_PACKAGE_RELATIONSHIPS = "http://schemas.openxmlformats.org/package/2006/relationships"
_CUSTOM_PROPERTIES = (
    "http://schemas.openxmlformats.org/officeDocument/2006/custom-properties"
)
_CUSTOM_PROPERTIES_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.custom-properties+xml"
)
_CUSTOM_PROPERTIES_PART = "docProps/custom.xml"
_CUSTOM_PROPERTIES_RELATIONSHIP = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/custom-properties"
)
_STRICT_CUSTOM_PROPERTIES_RELATIONSHIP = (
    "http://purl.oclc.org/ooxml/officeDocument/relationships/custom-properties"
)
_EMPTY_CUSTOM_PROPERTIES = (
    f'<Properties xmlns="{_CUSTOM_PROPERTIES}" '
    'xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes"/>'
).encode()


@dataclass(frozen=True)
class _Fixture:
    path: Path
    sheet: str
    value_cell: str
    formula_cell: str
    style_cell: str
    merge_range: str


def _basic_workbook(tmp_path: Path, seed: int) -> _Fixture:
    rng = random.Random(seed)
    sheet_name = f"Ledger_{seed}_{rng.randrange(1000, 9999)}"
    first_row = rng.randrange(2, 5)
    first_column = rng.randrange(1, 4)
    value_cell = f"{get_column_letter(first_column + 1)}{first_row + 1}"
    formula_cell = f"{get_column_letter(first_column + 3)}{first_row + 1}"
    style_cell = f"{get_column_letter(first_column)}{first_row}"
    merge_range = (
        f"{get_column_letter(first_column + 5)}{first_row + 6}:"
        f"{get_column_letter(first_column + 6)}{first_row + 6}"
    )
    path = tmp_path / f"seed-{seed}-basic.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    headers = ("Item", "Qty", "Price", "Total")
    for offset, header in enumerate(headers):
        sheet.cell(first_row, first_column + offset, header)
    for row_offset in range(1, 4):
        row = first_row + row_offset
        sheet.cell(row, first_column, f"Item {row_offset}")
        sheet.cell(row, first_column + 1, rng.randrange(2, 9))
        sheet.cell(row, first_column + 2, round(rng.uniform(1.5, 8.5), 2))
        quantity = f"{get_column_letter(first_column + 1)}{row}"
        price = f"{get_column_letter(first_column + 2)}{row}"
        sheet.cell(row, first_column + 3, f"={quantity}*{price}")
    workbook.save(path)
    workbook.close()
    return _Fixture(path, sheet_name, value_cell, formula_cell, style_cell, merge_range)


def _table_workbook(tmp_path: Path, seed: int) -> tuple[Path, str, str]:
    path = tmp_path / f"seed-{seed}-table.xlsx"
    sheet_name = f"Table_{seed}"
    table_name = f"Records{seed}"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(["Item", "Qty", "Price"])
    for index in range(1, 6):
        sheet.append([f"Item {index}", index + seed % 3, index * 1.25])
    table = Table(displayName=table_name, ref="A1:C6")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    workbook.create_sheet("Untouched")["A1"] = "scope control"
    workbook.save(path)
    workbook.close()
    return path, sheet_name, table_name


def _chart_workbook(tmp_path: Path, seed: int) -> tuple[Path, str, str]:
    path = tmp_path / f"seed-{seed}-chart.xlsx"
    sheet_name = f"Chart_{seed}"
    title = f"Baseline chart {seed}"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(["Label", "Amount"])
    for index in range(1, 5):
        sheet.append([f"Item {index}", index * (seed % 7 + 1)])
    chart = BarChart()
    chart.title = title
    chart.add_data(Reference(sheet, min_col=2, min_row=1, max_row=5), titles_from_data=True)
    chart.set_categories(Reference(sheet, min_col=1, min_row=2, max_row=5))
    sheet.add_chart(chart, "D2")
    workbook.create_sheet("Untouched")["A1"] = "scope control"
    workbook.save(path)
    workbook.close()
    return path, sheet_name, title


def _copies(source: Path, tmp_path: Path, label: str) -> tuple[Path, Path]:
    before = tmp_path / f"{label}-before.xlsx"
    after = tmp_path / f"{label}-after.xlsx"
    shutil.copy2(source, before)
    shutil.copy2(source, after)
    return before, after


def _mutated_copy(
    fixture: _Fixture,
    tmp_path: Path,
    label: str,
    mutation: Callable[[Workbook], None],
) -> tuple[Path, Path]:
    before, after = _copies(fixture.path, tmp_path, label)
    workbook = load_workbook(after)
    mutation(workbook)
    workbook.save(after)
    workbook.close()
    return before, after


def _replace_or_add_part(path: Path, part_name: str, payload: bytes) -> None:
    replacement = path.with_name(f".{path.stem}-package-rewrite{path.suffix}")
    with ZipFile(path) as source, ZipFile(replacement, "w", compression=ZIP_DEFLATED) as target:
        replaced = False
        for info in source.infolist():
            if info.is_dir():
                continue
            current = payload if info.filename == part_name else source.read(info.filename)
            target.writestr(info.filename, current)
            replaced |= info.filename == part_name
        if not replaced:
            target.writestr(part_name, payload)
        target.comment = source.comment
    replacement.replace(path)


def _add_custom_properties_plumbing(
    path: Path,
    *,
    payload: bytes = _EMPTY_CUSTOM_PROPERTIES,
    add_part: bool = True,
    add_relationship: bool = True,
    add_content_type: bool = True,
    relationship_type: str = _CUSTOM_PROPERTIES_RELATIONSHIP,
    relationship_target: str = _CUSTOM_PROPERTIES_PART,
    relationship_mode: str | None = None,
    relationship_count: int = 1,
    content_type: str = _CUSTOM_PROPERTIES_CONTENT_TYPE,
) -> None:
    if add_part:
        _replace_or_add_part(path, _CUSTOM_PROPERTIES_PART, payload)
    if add_relationship:
        with ZipFile(path) as package:
            root = ET.fromstring(package.read("_rels/.rels"))
        for index in range(relationship_count):
            attributes = {
                "Id": f"rIdOracleCustom{index + 1}",
                "Type": relationship_type,
                "Target": relationship_target,
            }
            if relationship_mode is not None:
                attributes["TargetMode"] = relationship_mode
            ET.SubElement(
                root,
                f"{{{_PACKAGE_RELATIONSHIPS}}}Relationship",
                attributes,
            )
        _replace_or_add_part(
            path,
            "_rels/.rels",
            ET.tostring(root, encoding="utf-8", xml_declaration=True),
        )
    if add_content_type:
        with ZipFile(path) as package:
            root = ET.fromstring(package.read("[Content_Types].xml"))
        ET.SubElement(
            root,
            f"{{{_CONTENT_TYPES}}}Override",
            {"PartName": f"/{_CUSTOM_PROPERTIES_PART}", "ContentType": content_type},
        )
        _replace_or_add_part(
            path,
            "[Content_Types].xml",
            ET.tostring(root, encoding="utf-8", xml_declaration=True),
        )


def _set_xml_default_content_type(path: Path, content_type: str) -> None:
    with ZipFile(path) as package:
        root = ET.fromstring(package.read("[Content_Types].xml"))
    declarations = [
        item
        for item in root.findall(f"{{{_CONTENT_TYPES}}}Default")
        if item.attrib.get("Extension", "").lower() == "xml"
    ]
    assert len(declarations) <= 1
    if declarations:
        declarations[0].set("ContentType", content_type)
    else:
        ET.SubElement(
            root,
            f"{{{_CONTENT_TYPES}}}Default",
            {"Extension": "xml", "ContentType": content_type},
        )
    _replace_or_add_part(
        path,
        "[Content_Types].xml",
        ET.tostring(root, encoding="utf-8", xml_declaration=True),
    )


def _rewrite_zip_container(path: Path) -> None:
    replacement = path.with_name(f".{path.stem}-container-rewrite{path.suffix}")
    with ZipFile(path) as source, ZipFile(replacement, "w", compression=ZIP_DEFLATED) as target:
        for name in reversed(source.namelist()):
            if name.endswith("/"):
                continue
            info = ZipInfo(name, date_time=(2024, 1, 2, 3, 4, 6))
            info.compress_type = ZIP_DEFLATED
            info.external_attr = 0o600 << 16
            target.writestr(info, source.read(name))
        target.comment = b"container metadata is outside workbook semantics"
    replacement.replace(path)


def _rewrite_chart_title(path: Path, old_title: str, new_title: str) -> None:
    with ZipFile(path) as package:
        chart_parts = sorted(name for name in package.namelist() if name.startswith("xl/charts/"))
        assert len(chart_parts) == 1
        part_name = chart_parts[0]
        root = ET.fromstring(package.read(part_name))
    title_nodes = [
        node for node in root.iter() if node.tag == f"{{{_DRAWING}}}t" and node.text == old_title
    ]
    assert len(title_nodes) == 1
    title_nodes[0].text = new_title
    _replace_or_add_part(path, part_name, ET.tostring(root, encoding="utf-8", xml_declaration=True))


def _inject_lossy_extension(path: Path) -> str:
    with ZipFile(path) as package:
        worksheet_parts = sorted(
            name
            for name in package.namelist()
            if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")
        )
        assert len(worksheet_parts) == 1
        part_name = worksheet_parts[0]
        root = ET.fromstring(package.read(part_name))
    extension_list = ET.SubElement(root, f"{{{_MAIN}}}extLst")
    ET.SubElement(extension_list, f"{{{_MAIN}}}ext", {"uri": "{ORACLE-LOSSY-CONTROL}"})
    _replace_or_add_part(path, part_name, ET.tostring(root, encoding="utf-8", xml_declaration=True))
    return part_name


def _inject_unknown_worksheet_node(path: Path) -> None:
    with ZipFile(path) as package:
        worksheet_parts = sorted(
            name
            for name in package.namelist()
            if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")
        )
        assert len(worksheet_parts) == 1
        part_name = worksheet_parts[0]
        root = ET.fromstring(package.read(part_name))
    ET.SubElement(root, "{urn:sheetledger:unknown-control}mystery")
    _replace_or_add_part(path, part_name, ET.tostring(root, encoding="utf-8", xml_declaration=True))


def _inject_content_types_root_attribute(path: Path) -> None:
    part_name = "[Content_Types].xml"
    with ZipFile(path) as package:
        root = ET.fromstring(package.read(part_name))
    root.set("oracleUnknown", "1")
    _replace_or_add_part(path, part_name, ET.tostring(root, encoding="utf-8", xml_declaration=True))


def _worksheet_root(path: Path) -> tuple[str, ET.Element]:
    with ZipFile(path) as package:
        worksheet_parts = sorted(
            name
            for name in package.namelist()
            if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")
        )
        assert len(worksheet_parts) == 1
        part_name = worksheet_parts[0]
        return part_name, ET.fromstring(package.read(part_name))


def _write_worksheet_root(path: Path, part_name: str, root: ET.Element) -> None:
    _replace_or_add_part(path, part_name, ET.tostring(root, encoding="utf-8", xml_declaration=True))


def _set_value_node_attribute(path: Path, coordinate: str) -> None:
    part_name, root = _worksheet_root(path)
    cells = [cell for cell in root.iter(f"{{{_MAIN}}}c") if cell.attrib.get("r") == coordinate]
    assert len(cells) == 1
    value_nodes = list(cells[0].findall(f"{{{_MAIN}}}v"))
    assert len(value_nodes) == 1
    value_nodes[0].set("{urn:sheetledger:unknown-control}opaque", "1")
    _write_worksheet_root(path, part_name, root)


def _set_cell_tail(path: Path, coordinate: str, value: str) -> None:
    part_name, root = _worksheet_root(path)
    cells = [cell for cell in root.iter(f"{{{_MAIN}}}c") if cell.attrib.get("r") == coordinate]
    assert len(cells) == 1
    cells[0].tail = value
    _write_worksheet_root(path, part_name, root)


def _set_cell_style_index(path: Path, coordinate: str, value: str) -> None:
    part_name, root = _worksheet_root(path)
    cells = [cell for cell in root.iter(f"{{{_MAIN}}}c") if cell.attrib.get("r") == coordinate]
    assert len(cells) == 1
    cells[0].set("s", value)
    _write_worksheet_root(path, part_name, root)


def _set_inline_text(path: Path, coordinate: str, value: str | None) -> None:
    part_name, root = _worksheet_root(path)
    cells = [cell for cell in root.iter(f"{{{_MAIN}}}c") if cell.attrib.get("r") == coordinate]
    assert len(cells) == 1
    text_nodes = list(cells[0].iter(f"{{{_MAIN}}}t"))
    assert len(text_nodes) == 1
    text_nodes[0].text = value
    _write_worksheet_root(path, part_name, root)


def _set_shared_formulas(path: Path, master_formula: str) -> None:
    part_name, root = _worksheet_root(path)
    cells = {
        cell.attrib.get("r"): cell
        for cell in root.iter(f"{{{_MAIN}}}c")
        if cell.attrib.get("r") in {"C1", "C2"}
    }
    assert set(cells) == {"C1", "C2"}
    master = cells["C1"].find(f"{{{_MAIN}}}f")
    dependent = cells["C2"].find(f"{{{_MAIN}}}f")
    assert master is not None and dependent is not None
    master.attrib.clear()
    master.attrib.update({"t": "shared", "si": "0", "ref": "C1:C2"})
    master.text = master_formula
    dependent.attrib.clear()
    dependent.attrib.update({"t": "shared", "si": "0"})
    dependent.text = None
    _write_worksheet_root(path, part_name, root)


def _set_array_formula(path: Path, master_formula: str) -> None:
    part_name, root = _worksheet_root(path)
    cells = {
        cell.attrib.get("r"): cell
        for cell in root.iter(f"{{{_MAIN}}}c")
        if cell.attrib.get("r") in {"C1", "C2"}
    }
    assert set(cells) == {"C1", "C2"}
    master = cells["C1"].find(f"{{{_MAIN}}}f")
    dependent = cells["C2"].find(f"{{{_MAIN}}}f")
    assert master is not None and dependent is not None
    master.attrib.clear()
    master.attrib.update({"t": "array", "ref": "C1:C2"})
    master.text = master_formula
    cells["C2"].remove(dependent)
    _write_worksheet_root(path, part_name, root)


def _inject_theme_extension(path: Path) -> str:
    with ZipFile(path) as package:
        theme_parts = sorted(
            name for name in package.namelist() if name.startswith("xl/theme/theme")
        )
        assert len(theme_parts) == 1
        part_name = theme_parts[0]
        root = ET.fromstring(package.read(part_name))
    extension_list = ET.SubElement(root, f"{{{_DRAWING}}}extLst")
    ET.SubElement(extension_list, f"{{{_DRAWING}}}ext", {"uri": "oracle-theme-control"})
    _replace_or_add_part(path, part_name, ET.tostring(root, encoding="utf-8", xml_declaration=True))
    return part_name


def _inject_calc_a1_extension(
    path: Path,
    *,
    string_ref_syntax: str = "ExcelA1",
    extra_attribute: bool = False,
    extra_child: bool = False,
) -> None:
    part_name = "xl/workbook.xml"
    with ZipFile(path) as package:
        root = ET.fromstring(package.read(part_name))
    extension_list = ET.SubElement(root, f"{{{_MAIN}}}extLst")
    extension = ET.SubElement(extension_list, f"{{{_MAIN}}}ext", {"uri": _CALC_A1_URI})
    marker = ET.SubElement(
        extension,
        f"{{{_LIBREOFFICE}}}extCalcPr",
        {"stringRefSyntax": string_ref_syntax},
    )
    if extra_attribute:
        marker.set("unexpected", "1")
    if extra_child:
        ET.SubElement(marker, f"{{{_LIBREOFFICE}}}unexpected")
    _replace_or_add_part(path, part_name, ET.tostring(root, encoding="utf-8", xml_declaration=True))


def _effect_names(result: WorkbookEffectDiff) -> frozenset[str]:
    return frozenset(effect.value for effect in result.effects)


def _expanded_cells(result: WorkbookEffectDiff) -> frozenset[OracleCell]:
    cells = {
        OracleCell(cell_range.sheet, f"{get_column_letter(column)}{row}")
        for cell_range in result.scope.ranges
        for row in range(cell_range.min_row, cell_range.max_row + 1)
        for column in range(cell_range.min_col, cell_range.max_col + 1)
    }
    return frozenset(cells)


def _expanded_formula_cells(result: WorkbookEffectDiff) -> frozenset[OracleCell]:
    cells = {
        OracleCell(cell_range.sheet, f"{get_column_letter(column)}{row}")
        for cell_range in result.formula_scope.ranges
        for row in range(cell_range.min_row, cell_range.max_row + 1)
        for column in range(cell_range.min_col, cell_range.max_col + 1)
    }
    return frozenset(cells)


def _assert_independent_agreement(
    before: Path,
    after: Path,
    *,
    effects: set[str],
    cells: set[OracleCell] | None = None,
    formula_cells: set[OracleCell] | None = None,
    sheets: set[str] | None = None,
    workbook_scope: bool = False,
) -> tuple[OracleDiff, WorkbookEffectDiff]:
    oracle = diff_ooxml(before, after)
    production = diff_workbooks(before, after)
    expected_cells = frozenset(cells or ())
    expected_formula_cells = frozenset(formula_cells or ())
    expected_sheets = frozenset(sheets or ())

    assert oracle.complete is True, oracle.reasons
    assert oracle.semantic_changed is bool(effects)
    assert oracle.effects == effects
    assert oracle.cells == expected_cells
    assert oracle.formula_cells == expected_formula_cells
    assert oracle.sheets == expected_sheets
    assert oracle.workbook_scope is workbook_scope

    assert production.complete is True, production.reasons
    assert production.semantic_changed is bool(effects)
    assert _effect_names(production) == effects
    assert production.scope.wildcard is workbook_scope
    if not workbook_scope:
        assert frozenset(production.scope.sheets) == expected_sheets
        assert _expanded_cells(production) == expected_cells
    if expected_formula_cells:
        assert production.formula_scope.wildcard is False
        assert production.formula_scope.sheets == ()
    else:
        assert production.formula_scope.empty
    assert _expanded_formula_cells(production) == expected_formula_cells
    assert production.changed_cell_count == len(expected_cells)
    return oracle, production


@pytest.mark.parametrize("seed", _ORACLE_SEEDS)
def test_independent_oracle_agrees_on_value_footprint(tmp_path: Path, seed: int) -> None:
    fixture = _basic_workbook(tmp_path, seed)
    before, after = _mutated_copy(
        fixture,
        tmp_path,
        f"value-{seed}",
        lambda workbook: setattr(
            workbook[fixture.sheet][fixture.value_cell], "value", seed * 10 + 7
        ),
    )
    changed = {OracleCell(fixture.sheet, fixture.value_cell)}

    _assert_independent_agreement(before, after, effects={"value"}, cells=changed)


@pytest.mark.parametrize("seed", _ORACLE_SEEDS)
def test_independent_oracle_agrees_on_formula_footprint(tmp_path: Path, seed: int) -> None:
    fixture = _basic_workbook(tmp_path, seed)
    before, after = _mutated_copy(
        fixture,
        tmp_path,
        f"formula-{seed}",
        lambda workbook: setattr(
            workbook[fixture.sheet][fixture.formula_cell],
            "value",
            f"={fixture.value_cell}*{seed % 5 + 3}",
        ),
    )
    changed = {OracleCell(fixture.sheet, fixture.formula_cell)}

    _assert_independent_agreement(
        before,
        after,
        effects={"formula"},
        cells=changed,
        formula_cells=changed,
    )


@pytest.mark.parametrize("seed", _ORACLE_SEEDS)
def test_independent_oracle_agrees_on_style_footprint(tmp_path: Path, seed: int) -> None:
    fixture = _basic_workbook(tmp_path, seed)
    color = f"FF{seed:02X}{(seed * 3) % 256:02X}{(seed * 5) % 256:02X}"

    def change_style(workbook: Workbook) -> None:
        workbook[fixture.sheet][fixture.style_cell].fill = PatternFill("solid", fgColor=color)

    before, after = _mutated_copy(fixture, tmp_path, f"style-{seed}", change_style)
    changed = {OracleCell(fixture.sheet, fixture.style_cell)}

    _assert_independent_agreement(
        before,
        after,
        effects={"style", "visual"},
        cells=changed,
    )


@pytest.mark.parametrize("seed", _ORACLE_SEEDS)
def test_independent_oracle_agrees_on_merge_sheet_scope(tmp_path: Path, seed: int) -> None:
    fixture = _basic_workbook(tmp_path, seed)
    workbook = load_workbook(fixture.path)
    workbook.create_sheet("Untouched")["A1"] = "scope control"
    workbook.save(fixture.path)
    workbook.close()
    before, after = _mutated_copy(
        fixture,
        tmp_path,
        f"merge-{seed}",
        lambda workbook: workbook[fixture.sheet].merge_cells(fixture.merge_range),
    )

    _assert_independent_agreement(
        before,
        after,
        effects={"structure", "visual"},
        sheets={fixture.sheet},
    )


def test_independent_oracle_agrees_on_worksheet_topology(tmp_path: Path) -> None:
    fixture = _basic_workbook(tmp_path, _ORACLE_SEEDS[0])
    before, after = _mutated_copy(
        fixture,
        tmp_path,
        "sheet-topology",
        lambda workbook: workbook.create_sheet("Prospective"),
    )

    _assert_independent_agreement(
        before,
        after,
        effects={"structure", "visual"},
        workbook_scope=True,
    )


def test_independent_oracle_agrees_on_table_structure(tmp_path: Path) -> None:
    source, sheet_name, table_name = _table_workbook(tmp_path, _ORACLE_SEEDS[0])
    before, after = _copies(source, tmp_path, "table")
    workbook = load_workbook(after)
    workbook[sheet_name].tables[table_name].ref = "A1:C5"
    workbook.save(after)
    workbook.close()

    _assert_independent_agreement(
        before,
        after,
        effects={"structure", "visual"},
        sheets={sheet_name},
    )


def test_independent_oracle_agrees_on_defined_name_structure(tmp_path: Path) -> None:
    fixture = _basic_workbook(tmp_path, _ORACLE_SEEDS[1])
    workbook = load_workbook(fixture.path)
    workbook.create_sheet("Untouched")["A1"] = "scope control"
    workbook.save(fixture.path)
    workbook.close()

    def add_defined_name(workbook: Workbook) -> None:
        workbook.defined_names.add(
            DefinedName(
                "AuditedInput",
                attr_text=f"'{fixture.sheet}'!${fixture.value_cell[0]}$1:${fixture.value_cell[0]}$4",
            )
        )

    before, after = _mutated_copy(fixture, tmp_path, "defined-name", add_defined_name)

    _assert_independent_agreement(
        before,
        after,
        effects={"structure"},
        sheets={fixture.sheet, "Untouched"},
    )


def test_independent_oracle_agrees_on_chart_semantics(tmp_path: Path) -> None:
    source, sheet_name, old_title = _chart_workbook(tmp_path, _ORACLE_SEEDS[2])
    before, after = _copies(source, tmp_path, "chart")
    _rewrite_chart_title(after, old_title, "Independently rewritten chart title")

    _assert_independent_agreement(
        before,
        after,
        effects={"structure", "visual"},
        sheets={sheet_name},
    )


def test_resave_and_zip_byte_rewrites_are_valid_noops(tmp_path: Path) -> None:
    fixture = _basic_workbook(tmp_path, _ORACLE_SEEDS[0])
    resave_before, resave_after = _copies(fixture.path, tmp_path, "resave")
    workbook = load_workbook(resave_after)
    workbook.save(resave_after)
    workbook.close()
    rewrite_before, rewrite_after = _copies(fixture.path, tmp_path, "zip-rewrite")
    original_bytes = rewrite_after.read_bytes()
    _rewrite_zip_container(rewrite_after)
    assert rewrite_after.read_bytes() != original_bytes

    for before, after in (
        (resave_before, resave_after),
        (rewrite_before, rewrite_after),
    ):
        _assert_independent_agreement(before, after, effects=set())


@pytest.mark.parametrize(
    "relationship_type",
    [_CUSTOM_PROPERTIES_RELATIONSHIP, _STRICT_CUSTOM_PROPERTIES_RELATIONSHIP],
    ids=["transitional", "strict"],
)
def test_empty_custom_properties_atomic_addition_and_removal_are_noops(
    tmp_path: Path,
    relationship_type: str,
) -> None:
    fixture = _basic_workbook(tmp_path, _ORACLE_SEEDS[0])
    before, after = _copies(fixture.path, tmp_path, "empty-custom-properties")
    _add_custom_properties_plumbing(after, relationship_type=relationship_type)

    _assert_independent_agreement(before, after, effects=set())
    _assert_independent_agreement(after, before, effects=set())


def test_empty_custom_properties_accepts_effective_xml_default(tmp_path: Path) -> None:
    fixture = _basic_workbook(tmp_path, _ORACLE_SEEDS[1])
    _set_xml_default_content_type(fixture.path, _CUSTOM_PROPERTIES_CONTENT_TYPE)
    before, after = _copies(fixture.path, tmp_path, "default-custom-properties")
    _add_custom_properties_plumbing(after, add_content_type=False)

    _assert_independent_agreement(before, after, effects=set())


def test_empty_custom_properties_preserves_a_typed_cell_effect(tmp_path: Path) -> None:
    fixture = _basic_workbook(tmp_path, _ORACLE_SEEDS[2])
    before, after = _copies(fixture.path, tmp_path, "custom-properties-with-cell-edit")
    workbook = load_workbook(after)
    workbook[fixture.sheet][fixture.value_cell] = 2604
    workbook.save(after)
    workbook.close()
    _add_custom_properties_plumbing(after)

    _assert_independent_agreement(
        before,
        after,
        effects={"value"},
        cells={OracleCell(fixture.sheet, fixture.value_cell)},
    )


@pytest.mark.parametrize(
    "case",
    [
        "missing-part",
        "content-only",
        "missing-relationship",
        "missing-content-type",
        "duplicate-relationship",
        "wrong-relationship-type",
        "wrong-target",
        "external-target",
        "wrong-content-type",
    ],
)
def test_incomplete_custom_properties_plumbing_fails_closed(
    tmp_path: Path,
    case: str,
) -> None:
    fixture = _basic_workbook(tmp_path, _ORACLE_SEEDS[0])
    before, after = _copies(fixture.path, tmp_path, f"invalid-custom-{case}")
    options: dict[str, object] = {}
    if case == "missing-part":
        options["add_part"] = False
    elif case == "content-only":
        options.update(add_part=False, add_relationship=False)
    elif case == "missing-relationship":
        options["add_relationship"] = False
    elif case == "missing-content-type":
        options["add_content_type"] = False
    elif case == "duplicate-relationship":
        options["relationship_count"] = 2
    elif case == "wrong-relationship-type":
        options["relationship_type"] = f"{_CUSTOM_PROPERTIES_RELATIONSHIP}-neighbor"
    elif case == "wrong-target":
        options["relationship_target"] = "docProps/custom-neighbor.xml"
    elif case == "external-target":
        options["relationship_mode"] = "External"
    elif case == "wrong-content-type":
        options["content_type"] = "application/x-wrong-custom-properties"
    _add_custom_properties_plumbing(after, **options)

    oracle = diff_ooxml(before, after)
    production = diff_workbooks(before, after)

    assert oracle.complete is False
    assert oracle.effects == {"unknown"}
    assert oracle.workbook_scope is True
    assert "custom-properties" in oracle.reasons[0]
    same_malformed_state = diff_ooxml(after, after)
    assert same_malformed_state.complete is False
    assert same_malformed_state.effects == {"unknown"}
    assert same_malformed_state.workbook_scope is True
    assert production.complete is False
    assert production.effects == {EffectKind.UNKNOWN}
    assert production.scope == EvidenceScope.workbook()


@pytest.mark.parametrize(
    ("case", "payload"),
    [
        (
            "non-empty-property",
            f'<Properties xmlns="{_CUSTOM_PROPERTIES}"><property/></Properties>'.encode(),
        ),
        ("root-attribute", f'<Properties xmlns="{_CUSTOM_PROPERTIES}" opaque="1"/>'.encode()),
        ("wrong-root-namespace", b'<Properties xmlns="urn:near-custom-properties"/>'),
        ("root-text", f'<Properties xmlns="{_CUSTOM_PROPERTIES}">material</Properties>'.encode()),
        ("comment", f'<Properties xmlns="{_CUSTOM_PROPERTIES}"><!--opaque--></Properties>'.encode()),
        (
            "processing-instruction",
            f'<Properties xmlns="{_CUSTOM_PROPERTIES}"><?oracle opaque?></Properties>'.encode(),
        ),
        (
            "dtd-entity",
            (
                f'<!DOCTYPE Properties [<!ENTITY opaque "material">]>'
                f'<Properties xmlns="{_CUSTOM_PROPERTIES}">&opaque;</Properties>'
            ).encode(),
        ),
    ],
)
def test_nonempty_or_unsafe_custom_properties_payload_fails_closed(
    tmp_path: Path,
    case: str,
    payload: bytes,
) -> None:
    fixture = _basic_workbook(tmp_path, _ORACLE_SEEDS[1])
    before, after = _copies(fixture.path, tmp_path, f"invalid-custom-payload-{case}")
    _add_custom_properties_plumbing(after, payload=payload)

    oracle = diff_ooxml(before, after)

    assert oracle.complete is False
    assert oracle.effects == {"unknown"}
    assert oracle.workbook_scope is True
    assert "custom-properties" in oracle.reasons[0]
    same_malformed_state = diff_ooxml(after, after)
    assert same_malformed_state.complete is False
    assert same_malformed_state.effects == {"unknown"}
    assert same_malformed_state.workbook_scope is True


def test_opaque_part_change_fails_closed_in_both_implementations(tmp_path: Path) -> None:
    fixture = _basic_workbook(tmp_path, _ORACLE_SEEDS[0])
    before, after = _copies(fixture.path, tmp_path, "opaque")
    part_name = "xl/embeddings/oleObject1.bin"
    _replace_or_add_part(before, part_name, b"opaque-before")
    _replace_or_add_part(after, part_name, b"opaque-after")

    oracle = diff_ooxml(before, after)
    production = diff_workbooks(before, after)

    assert oracle.complete is False
    assert oracle.effects == {"unknown"}
    assert oracle.workbook_scope is True
    assert part_name in oracle.reasons[0]
    assert production.complete is False
    assert production.effects == {EffectKind.UNKNOWN}
    assert production.scope == EvidenceScope.workbook()


def test_lossy_extension_round_trip_fails_closed_in_both_implementations(
    tmp_path: Path,
) -> None:
    fixture = _basic_workbook(tmp_path, _ORACLE_SEEDS[1])
    before, after = _copies(fixture.path, tmp_path, "lossy-extension")
    _inject_lossy_extension(before)
    shutil.copy2(before, after)
    with warnings.catch_warnings(record=True) as observed:
        warnings.simplefilter("always")
        workbook = load_workbook(after)
        workbook.save(after)
        workbook.close()
    assert any("extension is not supported" in str(item.message) for item in observed)

    oracle = diff_ooxml(before, after)
    production = diff_workbooks(before, after)

    assert oracle.complete is False
    assert oracle.effects == {"unknown"}
    assert "lossy extension" in oracle.reasons[0]
    assert production.complete is False
    assert production.effects == {EffectKind.UNKNOWN}
    assert production.scope == EvidenceScope.workbook()
    assert production.reasons


def test_duplicate_zip_part_is_unknown_not_a_semantic_noop(tmp_path: Path) -> None:
    fixture = _basic_workbook(tmp_path, _ORACLE_SEEDS[2])
    before, after = _copies(fixture.path, tmp_path, "duplicate-part")
    part_name = "xl/worksheets/sheet1.xml"
    with ZipFile(after) as package:
        payload = package.read(part_name)
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        with ZipFile(after, "a") as package:
            package.writestr(part_name, payload)

    oracle = diff_ooxml(before, after)
    production = diff_workbooks(before, after)

    assert oracle.complete is False
    assert oracle.effects == {"unknown"}
    assert "duplicate OOXML part" in oracle.reasons[0]
    assert production.complete is False
    assert production.effects == {EffectKind.UNKNOWN}
    assert production.scope == EvidenceScope.workbook()


def test_oracle_rejects_unknown_namespace_beside_a_known_value_edit(tmp_path: Path) -> None:
    fixture = _basic_workbook(tmp_path, _ORACLE_SEEDS[0])
    before, after = _copies(fixture.path, tmp_path, "unknown-namespace")
    workbook = load_workbook(after)
    workbook[fixture.sheet][fixture.value_cell] = 999
    workbook.save(after)
    workbook.close()
    _inject_unknown_worksheet_node(after)

    oracle = diff_ooxml(before, after)

    assert oracle.complete is False
    assert oracle.effects == {"unknown"}
    assert oracle.workbook_scope is True
    assert "unsupported worksheet XML namespace" in oracle.reasons[0]


def test_oracle_rejects_unknown_content_types_root_attributes(tmp_path: Path) -> None:
    fixture = _basic_workbook(tmp_path, _ORACLE_SEEDS[1])
    before, after = _copies(fixture.path, tmp_path, "content-types-root")
    _inject_content_types_root_attribute(after)

    oracle = diff_ooxml(before, after)

    assert oracle.complete is False
    assert oracle.effects == {"unknown"}
    assert oracle.workbook_scope is True
    assert "unsupported content-types root content" in oracle.reasons[0]


def test_oracle_detects_non_whitespace_tail_in_opaque_xml(tmp_path: Path) -> None:
    fixture = _basic_workbook(tmp_path, _ORACLE_SEEDS[2])
    before, after = _copies(fixture.path, tmp_path, "opaque-tail")
    part_name = "customXml/item1.xml"
    _replace_or_add_part(before, part_name, b"<root><item/>before-tail</root>")
    _replace_or_add_part(after, part_name, b"<root><item/>after-tail</root>")

    oracle = diff_ooxml(before, after)

    assert oracle.complete is False
    assert oracle.effects == {"unknown"}
    assert oracle.workbook_scope is True
    assert part_name in oracle.reasons[0]


def test_oracle_rejects_value_node_attributes_beside_a_known_edit(tmp_path: Path) -> None:
    fixture = _basic_workbook(tmp_path, _ORACLE_SEEDS[0])
    before, after = _copies(fixture.path, tmp_path, "value-node-attribute")
    workbook = load_workbook(after)
    workbook[fixture.sheet][fixture.value_cell] = 321
    workbook.save(after)
    workbook.close()
    _set_value_node_attribute(after, fixture.value_cell)

    oracle = diff_ooxml(before, after)

    assert oracle.complete is False
    assert oracle.effects == {"unknown"}
    assert oracle.workbook_scope is True
    assert "unsupported value-node content" in oracle.reasons[0]
    production = diff_workbooks(before, after)
    assert production.complete is False
    assert _effect_names(production) == {"unknown"}
    assert production.scope.wildcard is True


@pytest.mark.parametrize("corruption", ["cell-tail", "negative-style-index"])
def test_oracle_rejects_invalid_cell_plumbing_beside_a_known_edit(
    tmp_path: Path,
    corruption: str,
) -> None:
    fixture = _basic_workbook(tmp_path, _ORACLE_SEEDS[0])
    before, after = _copies(fixture.path, tmp_path, f"invalid-{corruption}")
    workbook = load_workbook(after)
    workbook[fixture.sheet][fixture.value_cell] = 654
    workbook.save(after)
    workbook.close()
    if corruption == "cell-tail":
        _set_cell_tail(after, fixture.value_cell, "opaque-tail")
    else:
        _set_cell_style_index(after, fixture.value_cell, "-1")

    oracle = diff_ooxml(before, after)

    assert oracle.complete is False
    assert oracle.effects == {"unknown"}
    assert oracle.workbook_scope is True
    expected = "mixed cell content" if corruption == "cell-tail" else "cell style index"
    assert expected in oracle.reasons[0]
    production = diff_workbooks(before, after)
    assert production.complete is False
    assert _effect_names(production) == {"unknown"}
    assert production.scope.wildcard is True


def test_whitespace_only_inline_text_is_semantic_data(tmp_path: Path) -> None:
    fixture = _basic_workbook(tmp_path, _ORACLE_SEEDS[1])
    before, after = _copies(fixture.path, tmp_path, "inline-whitespace")
    _set_inline_text(before, fixture.style_cell, " ")
    _set_inline_text(after, fixture.style_cell, None)
    changed = {OracleCell(fixture.sheet, fixture.style_cell)}

    _assert_independent_agreement(before, after, effects={"value"}, cells=changed)


def test_shared_formula_master_change_covers_all_dependents(tmp_path: Path) -> None:
    source = tmp_path / "shared-formula.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Shared"
    sheet.append([1, 2, "=A1+B1"])
    sheet.append([3, 4, "=A2+B2"])
    workbook.save(source)
    workbook.close()
    before, after = _copies(source, tmp_path, "shared-formula")
    _set_shared_formulas(before, "A1+B1")
    _set_shared_formulas(after, "A1-B1")
    changed = {OracleCell("Shared", "C1"), OracleCell("Shared", "C2")}

    _assert_independent_agreement(
        before,
        after,
        effects={"formula"},
        cells=changed,
        formula_cells=changed,
    )


def test_array_formula_master_change_covers_the_declared_range(tmp_path: Path) -> None:
    source = tmp_path / "array-formula.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Array"
    sheet.append([1, 2, "=A1+B1"])
    sheet.append([3, 4, "=A2+B2"])
    workbook.save(source)
    workbook.close()
    before, after = _copies(source, tmp_path, "array-formula")
    _set_array_formula(before, "A1:A2+B1:B2")
    _set_array_formula(after, "A1:A2-B1:B2")

    oracle = diff_ooxml(before, after)

    changed = frozenset({OracleCell("Array", "C1"), OracleCell("Array", "C2")})
    assert oracle.complete is True, oracle.reasons
    assert oracle.effects == {"formula"}
    assert oracle.cells == changed
    assert oracle.formula_cells == changed
    production = diff_workbooks(before, after)
    assert production.complete is True, production.reasons
    assert _effect_names(production) == {"formula"}
    assert _expanded_cells(production) == changed
    assert _expanded_formula_cells(production) == changed
    assert production.changed_cell_count == 2


def test_opaque_xml_whitespace_tail_change_fails_closed(tmp_path: Path) -> None:
    fixture = _basic_workbook(tmp_path, _ORACLE_SEEDS[2])
    before, after = _copies(fixture.path, tmp_path, "opaque-whitespace-tail")
    part_name = "customXml/item1.xml"
    _replace_or_add_part(before, part_name, b"<root><item/> </root>")
    _replace_or_add_part(after, part_name, b"<root><item/>\t</root>")

    oracle = diff_ooxml(before, after)

    assert oracle.complete is False
    assert oracle.effects == {"unknown"}
    assert part_name in oracle.reasons[0]
    production = diff_workbooks(before, after)
    assert production.complete is False
    assert _effect_names(production) == {"unknown"}
    assert production.scope.wildcard is True
    assert part_name in production.reasons[0]


def test_oracle_rejects_theme_extension_markup(tmp_path: Path) -> None:
    fixture = _basic_workbook(tmp_path, _ORACLE_SEEDS[2])
    before, after = _copies(fixture.path, tmp_path, "theme-extension")
    part_name = _inject_theme_extension(after)

    oracle = diff_ooxml(before, after)

    assert oracle.complete is False
    assert oracle.effects == {"unknown"}
    assert oracle.workbook_scope is True
    assert "lossy extension" in oracle.reasons[0]
    assert part_name in oracle.reasons[0]


def test_oracle_retains_formula_effect_when_a_sheet_is_added(tmp_path: Path) -> None:
    fixture = _basic_workbook(tmp_path, _ORACLE_SEEDS[0])
    before, after = _copies(fixture.path, tmp_path, "new-formula-sheet")
    workbook = load_workbook(after)
    added = workbook.create_sheet("Calculated")
    added["A1"] = "=1+1"
    workbook.save(after)
    workbook.close()

    changed = frozenset({OracleCell("Calculated", "A1")})
    _assert_independent_agreement(
        before,
        after,
        effects={"formula", "structure", "visual"},
        cells=set(changed),
        formula_cells=set(changed),
        workbook_scope=True,
    )


def test_exact_libreoffice_excel_a1_marker_is_a_semantic_noop(tmp_path: Path) -> None:
    fixture = _basic_workbook(tmp_path, _ORACLE_SEEDS[1])
    before, after = _copies(fixture.path, tmp_path, "calc-a1-marker")
    _inject_calc_a1_extension(after)

    _assert_independent_agreement(before, after, effects=set())


def test_oracle_and_production_reject_manual_calc_mode_beside_value_edit(
    tmp_path: Path,
) -> None:
    fixture = _basic_workbook(tmp_path, _ORACLE_SEEDS[1])
    before, after = _copies(fixture.path, tmp_path, "manual-calc-with-value")
    workbook = load_workbook(after)
    workbook[fixture.sheet][fixture.value_cell] = "edited"
    workbook.save(after)
    workbook.close()

    part_name = "xl/workbook.xml"
    with ZipFile(after) as package:
        root = ET.fromstring(package.read(part_name))
    calculation = root.find(f"{{{_MAIN}}}calcPr")
    assert calculation is not None
    calculation.set("calcMode", "manual")
    _replace_or_add_part(
        after,
        part_name,
        ET.tostring(root, encoding="utf-8", xml_declaration=True),
    )

    oracle = diff_ooxml(before, after)
    production = diff_workbooks(before, after)

    assert oracle.complete is False
    assert oracle.effects == {"unknown"}
    assert oracle.workbook_scope is True
    assert "unclassified workbook XML" in oracle.reasons[0]
    assert production.complete is False
    assert production.effects == {EffectKind.UNKNOWN}
    assert production.scope == EvidenceScope.workbook()
    assert part_name in production.reasons[0]


@pytest.mark.parametrize(
    "mutation",
    ["calc-a1-value", "extra-attribute", "extra-child"],
)
def test_calc_a1_marker_variants_fail_closed(tmp_path: Path, mutation: str) -> None:
    fixture = _basic_workbook(tmp_path, _ORACLE_SEEDS[2])
    before, after = _copies(fixture.path, tmp_path, f"invalid-{mutation}")
    _inject_calc_a1_extension(
        after,
        string_ref_syntax="CalcA1" if mutation == "calc-a1-value" else "ExcelA1",
        extra_attribute=mutation == "extra-attribute",
        extra_child=mutation == "extra-child",
    )

    oracle = diff_ooxml(before, after)
    production = diff_workbooks(before, after)

    assert oracle.complete is False
    assert oracle.effects == {"unknown"}
    assert oracle.workbook_scope is True
    assert "lossy extension" in oracle.reasons[0]
    assert production.complete is False
    assert production.effects == {EffectKind.UNKNOWN}
    assert production.scope == EvidenceScope.workbook()
