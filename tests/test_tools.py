from __future__ import annotations

import hashlib
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook
from PIL import Image

from spreadsheet_harness.errors import RecalculationIntegrityError
from spreadsheet_harness.render import find_libreoffice
from spreadsheet_harness.session import WorkbookSession
from spreadsheet_harness.tools import SpreadsheetToolRegistry
from spreadsheet_harness.trajectory import read_trajectory


def test_tool_registry_dispatch_and_errors(sample_workbook: Path, tmp_path: Path) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "run")
    tools = SpreadsheetToolRegistry(session, enable_code=False)
    names = {item["name"] for item in tools.schemas}
    assert {"list_sheets", "inspect_range", "range_to_latex", "view_image"} <= names
    assert "code_interpreter" not in names
    assert tools.invoke("list_sheets", {}).data["ok"] is True
    result = tools.invoke("inspect_range", {"sheet": "Nope", "range_ref": "A1"})
    assert result.data["ok"] is False
    assert result.data["type"] == "ToolInputError"


def test_inspect_range_schema_discloses_limit_and_returns_bounded_evidence(
    sample_workbook: Path, tmp_path: Path
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "run")
    tools = SpreadsheetToolRegistry(
        session, enable_code=False, allowed_tools={"inspect_range"}
    )
    schema = tools.schemas[0]

    assert "runtime maximum of 500 returned cells" in schema["description"]
    assert "limited to 500 cells" in schema["parameters"]["properties"][
        "range_ref"
    ]["description"]

    before = session.workbook_path.read_bytes()
    result = tools.invoke(
        "inspect_range", {"sheet": "Sales", "range_ref": "$A$1:$Z$1000"}
    ).data

    assert result["ok"] is True
    assert result["requested_range"] == "A1:Z1000"
    assert result["requested_cell_count"] == 26_000
    assert result["returned_range"] == "A1:Z19"
    assert result["returned_cell_count"] == 494
    assert result["range"] == "A1:Z19"
    assert result["cell_count"] == 494
    assert result["truncated"] is True
    assert result["limits"] == {"max_cells": 500}
    assert result["truncation"] == {
        "policy": "top_left_rectangle_preserve_columns_when_possible",
        "omitted_cell_count": 25_506,
        "message": (
            "The requested range exceeded 500 cells; this response contains the "
            "reported deterministic top-left rectangle."
        ),
    }
    assert len(result["matrix"]) == 19
    assert all(len(row) == 26 for row in result["matrix"])
    assert session.workbook_path.read_bytes() == before

    returned = [
        event
        for event in read_trajectory(session.paths.trajectory)
        if event["event"] == "tool.returned"
    ]
    assert returned[-1]["payload"]["result"]["truncated"] is True
    assert returned[-1]["payload"]["result"]["range"] == "A1:Z19"


def test_recalculate_and_read_reports_semantic_errors_separately_from_success(
    sample_workbook: Path, tmp_path: Path, monkeypatch: Any
) -> None:
    workbook = load_workbook(sample_workbook)
    error_values = ["#VALUE!", "#REF!", "#NAME?", "#DIV/0!", "#N/A"]
    for row in range(1, 36):
        workbook["Sales"].cell(row, 8, error_values[(row - 1) % len(error_values)])
    workbook["Sales"]["I1"] = "'#REF!"
    workbook.save(sample_workbook)
    workbook.close()

    session = WorkbookSession.create(sample_workbook, tmp_path / "run")
    metadata = {"backend": "libreoffice-headless", "atomic_replace": True}
    monkeypatch.setattr(session, "recalculate", lambda: metadata)
    tools = SpreadsheetToolRegistry(
        session, enable_code=False, allowed_tools={"recalculate_and_read"}
    )
    schema = tools.schemas[0]

    assert "ok=true reports transport/tool success" in schema["description"]
    result = tools.invoke(
        "recalculate_and_read", {"sheet": "Sales", "range_ref": "H1:I35"}
    ).data

    assert result["ok"] is True
    assert result["calculation"] == metadata
    assert result["calculation_valid"] is False
    assert result["calculation_errors"]["sheet"] == "Sales"
    assert result["calculation_errors"]["range"] == "H1:I35"
    assert result["calculation_errors"]["count"] == 35
    assert result["calculation_errors"]["by_error"] == {
        "#DIV/0!": 7,
        "#N/A": 7,
        "#NAME?": 7,
        "#REF!": 7,
        "#VALUE!": 7,
    }
    assert result["calculation_errors"]["coordinate_limit"] == 32
    assert result["calculation_errors"]["coordinates_truncated"] is True
    assert result["calculation_errors"]["coordinates"][0] == {
        "coordinate": "H1",
        "error": "#VALUE!",
    }
    assert result["calculation_errors"]["coordinates"][-1]["coordinate"] == "H32"
    assert "I1" not in {
        item["coordinate"] for item in result["calculation_errors"]["coordinates"]
    }
    assert result["inspection"]["cells"][0]["cached_data_type"] == "e"

    returned = [
        event
        for event in read_trajectory(session.paths.trajectory)
        if event["event"] == "tool.returned"
    ]
    assert returned[-1]["payload"]["result"]["ok"] is True
    assert returned[-1]["payload"]["result"]["calculation_valid"] is False
    assert returned[-1]["payload"]["result"]["calculation_errors"]["count"] == 35


def test_recalculate_and_read_rejects_oversized_target_before_recalculation(
    sample_workbook: Path, tmp_path: Path, monkeypatch: Any
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "oversized-recalc-run")
    recalculate_calls = 0

    def unexpected_recalculation() -> dict[str, Any]:
        nonlocal recalculate_calls
        recalculate_calls += 1
        return {"backend": "should-not-run"}

    monkeypatch.setattr(session, "recalculate", unexpected_recalculation)
    tools = SpreadsheetToolRegistry(
        session, enable_code=False, allowed_tools={"recalculate_and_read"}
    )
    schema = tools.schemas[0]
    before_sha256 = hashlib.sha256(session.workbook_path.read_bytes()).hexdigest()

    result = tools.invoke(
        "recalculate_and_read", {"sheet": "Sales", "range_ref": "$A$1:$Z$100"}
    ).data

    after_sha256 = hashlib.sha256(session.workbook_path.read_bytes()).hexdigest()
    assert "limited to 500 cells" in schema["description"]
    assert "maximum of 500 cells" in schema["parameters"]["properties"][
        "range_ref"
    ]["description"]
    assert result["ok"] is False
    assert result["type"] == "ToolInputError"
    assert result["preflight_rejected"] is True
    assert result["workbook_mutation_attempted"] is False
    assert result["workbook_changed"] is False
    assert "at most 500 cells" in result["error"]
    assert "A1:Z100 contains 2600 cells" in result["error"]
    assert "no recalculation was performed" in result["error"]
    assert recalculate_calls == 0
    assert after_sha256 == before_sha256


def test_registry_propagates_recalculation_integrity_failure(
    sample_workbook: Path, tmp_path: Path, monkeypatch: Any
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "integrity-run")
    evidence = {
        "backend": "libreoffice-headless",
        "sheet_inventory_integrity": {"matched": False},
    }
    failure = RecalculationIntegrityError(
        "Recalculation changed sheet identity",
        evidence=evidence,
    )

    def fail_recalculation() -> dict[str, Any]:
        raise failure

    monkeypatch.setattr(session, "recalculate", fail_recalculation)
    tools = SpreadsheetToolRegistry(
        session, enable_code=False, allowed_tools={"recalculate_and_read"}
    )

    with pytest.raises(RecalculationIntegrityError) as caught:
        tools.invoke(
            "recalculate_and_read",
            {"sheet": "Sales", "range_ref": "A1"},
        )

    assert caught.value is failure
    trajectory = read_trajectory(session.paths.trajectory)
    assert trajectory[-1]["event"] == "tool.failed"
    assert trajectory[-1]["payload"] == {
        "name": "recalculate_and_read",
        "error_type": "RecalculationIntegrityError",
        "failure_category": "recalculation_infrastructure",
        "recalculation": evidence,
    }
    assert not any(
        event["event"] == "tool.returned"
        and event["payload"]["name"] == "recalculate_and_read"
        for event in trajectory
    )


def test_recalculate_and_read_sparse_pending_scope_has_no_500_cell_limit(
    sample_workbook: Path, tmp_path: Path, monkeypatch: Any
) -> None:
    workbook = load_workbook(sample_workbook)
    sheet = workbook["Sales"]
    coordinates = []
    for row in range(1, 602):
        coordinate = f"J{row}"
        sheet[coordinate] = f"={row}+1"
        coordinates.append(("Sales", coordinate))
    workbook.save(sample_workbook)
    workbook.close()

    session = WorkbookSession.create(sample_workbook, tmp_path / "sparse-recalc-run")
    recalculate_calls = 0

    def recalculate() -> dict[str, Any]:
        nonlocal recalculate_calls
        recalculate_calls += 1
        return {"backend": "test-libreoffice", "atomic_replace": True}

    monkeypatch.setattr(session, "recalculate", recalculate)
    tools = SpreadsheetToolRegistry(
        session, enable_code=False, allowed_tools={"recalculate_and_read"}
    )
    tools.set_pending_formula_validation_scope(coordinates)

    result = tools.invoke(
        "recalculate_and_read",
        {"validation_scope": "pending_formula_changes"},
    ).data

    assert recalculate_calls == 1
    assert result["ok"] is True
    assert result["calculation_valid"] is True
    assert result["validation_scope"]["coordinate_count"] == 601
    assert result["validation_scope"]["formula_cells_present"] == 601
    assert result["validation_scope"]["formula_cells_absent"] == 0
    assert result["validation_scope"]["coverage_complete"] is True
    assert result["calculation_errors"] == {
        "count": 0,
        "coordinates": [],
        "coordinate_limit": 32,
        "coordinates_truncated": False,
    }


@pytest.mark.parametrize(
    "arguments",
    [
        {},
        {"validation_scope": "range", "sheet": "Sales"},
        {"validation_scope": "unknown"},
        {"validation_scope": "pending_formula_changes"},
        {
            "validation_scope": "pending_formula_changes",
            "sheet": "Sales",
        },
    ],
)
def test_recalculate_and_read_rejects_invalid_scope_before_mutation(
    sample_workbook: Path,
    tmp_path: Path,
    monkeypatch: Any,
    arguments: dict[str, Any],
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "invalid-recalc-scope")
    recalculate_calls = 0

    def unexpected_recalculation() -> dict[str, Any]:
        nonlocal recalculate_calls
        recalculate_calls += 1
        return {"backend": "should-not-run"}

    monkeypatch.setattr(session, "recalculate", unexpected_recalculation)
    tools = SpreadsheetToolRegistry(
        session, enable_code=False, allowed_tools={"recalculate_and_read"}
    )
    before = session.workbook_path.read_bytes()

    result = tools.invoke("recalculate_and_read", arguments).data

    assert result["ok"] is False
    assert result["type"] == "ToolInputError"
    assert result["preflight_rejected"] is True
    assert result["workbook_mutation_attempted"] is False
    assert result["workbook_changed"] is False
    assert recalculate_calls == 0
    assert session.workbook_path.read_bytes() == before


@pytest.mark.skipif(find_libreoffice() is None, reason="LibreOffice is not installed")
def test_sparse_pending_recalculation_detects_real_libreoffice_formula_error(
    sample_workbook: Path, tmp_path: Path
) -> None:
    workbook = load_workbook(sample_workbook)
    workbook["Sales"]["H1"] = "=1/0"
    workbook["Sales"]["H2"] = "=NO_SUCH_FUNCTION(1)"
    workbook["Sales"]["H3"] = '="#REF!"'
    workbook.save(sample_workbook)
    workbook.close()

    session = WorkbookSession.create(sample_workbook, tmp_path / "sparse-real-recalc-run")
    tools = SpreadsheetToolRegistry(
        session, enable_code=False, allowed_tools={"recalculate_and_read"}
    )
    tools.set_pending_formula_validation_scope(
        {("Sales", "H1"), ("Sales", "H2"), ("Sales", "H3")}
    )

    result = tools.invoke(
        "recalculate_and_read",
        {"validation_scope": "pending_formula_changes"},
    ).data

    assert result["ok"] is True
    assert result["calculation_valid"] is False
    assert result["validation_scope"]["coordinate_count"] == 3
    assert result["calculation_errors"]["count"] == 2
    assert result["calculation_errors"]["coordinates"] == [
        {"sheet": "Sales", "coordinate": "H1", "error": "#DIV/0!"},
        {"sheet": "Sales", "coordinate": "H2", "error": "#NAME?"},
    ]


@pytest.mark.skipif(find_libreoffice() is None, reason="LibreOffice is not installed")
def test_recalculate_and_read_detects_real_libreoffice_formula_errors(
    sample_workbook: Path, tmp_path: Path
) -> None:
    workbook = load_workbook(sample_workbook)
    workbook["Sales"]["H1"] = "=1/0"
    workbook["Sales"]["H2"] = "=NO_SUCH_FUNCTION(1)"
    workbook["Sales"]["H3"] = '="#REF!"'
    workbook.save(sample_workbook)
    workbook.close()

    session = WorkbookSession.create(sample_workbook, tmp_path / "run")
    tools = SpreadsheetToolRegistry(
        session, enable_code=False, allowed_tools={"recalculate_and_read"}
    )
    result = tools.invoke(
        "recalculate_and_read", {"sheet": "Sales", "range_ref": "H1:H3"}
    ).data

    assert result["ok"] is True
    assert result["calculation"]["backend"] == "libreoffice-headless"
    assert result["calculation_valid"] is False
    assert result["calculation_errors"]["count"] == 2
    assert result["calculation_errors"]["by_error"] == {
        "#DIV/0!": 1,
        "#NAME?": 1,
    }
    assert result["calculation_errors"]["coordinates"] == [
        {"coordinate": "H1", "error": "#DIV/0!"},
        {"coordinate": "H2", "error": "#NAME?"},
    ]
    h3 = next(
        cell
        for cell in result["inspection"]["cells"]
        if cell["coordinate"] == "H3"
    )
    assert h3["value"] == "#REF!"
    assert h3["cached_data_type"] == "s"


def test_code_interpreter_schema_requires_self_contained_calls(
    sample_workbook: Path, tmp_path: Path
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "run")
    tools = SpreadsheetToolRegistry(session, enable_code=True)
    schema = next(item for item in tools.schemas if item["name"] == "code_interpreter")
    description = schema["description"]

    assert "fresh Python process" in description
    assert "variables, imports, and workbook objects do not persist" in description
    assert "editing or recovery script self-contained" in description
    workflow = description.split("self-contained:", 1)[1]
    expected_steps = (
        "import",
        "load",
        "re-read the request and inspected workbook state",
        "edit",
        "save",
        "close",
        "reopen",
        "verify the requested change and nearby cells",
        "print compact verification",
    )
    positions = [workflow.index(step) for step in expected_steps]
    assert positions == sorted(positions)


def test_tool_registry_redacts_configured_secret_from_recorded_outcome(
    sample_workbook: Path, tmp_path: Path
) -> None:
    unusual_secret = "credential-with-an-unusual-shape"
    session = WorkbookSession.create(
        sample_workbook,
        tmp_path / "run",
        recorder_secrets=(unusual_secret,),
    )
    tools = SpreadsheetToolRegistry(
        session,
        enable_code=True,
        allowed_tools={"code_interpreter"},
    )

    class LeakingInterpreter:
        def run(self, *_: Any, **__: Any) -> dict[str, Any]:
            return {
                "ok": False,
                "stdout": unusual_secret,
                "nested": {"message": unusual_secret},
            }

    tools.interpreter = LeakingInterpreter()  # type: ignore[assignment]
    result = tools.invoke("code_interpreter", {"code": "print('x')"})

    assert result.data["stdout"] == unusual_secret
    assert unusual_secret not in session.paths.trajectory.read_text(encoding="utf-8")


def test_allowed_tools_filters_schemas_and_dispatch(sample_workbook: Path, tmp_path: Path) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "run")
    allowed = {"list_sheets", "range_to_latex"}
    tools = SpreadsheetToolRegistry(session, enable_code=False, allowed_tools=allowed)
    allowed.clear()

    assert {item["name"] for item in tools.schemas} == {"list_sheets", "range_to_latex"}
    assert tools.invoke("list_sheets", {}).data["ok"] is True

    before = session.workbook_path.read_bytes()
    blocked = tools.invoke(
        "write_range", {"sheet": "Sales", "start_cell": "A1", "values": [["changed"]]}
    )
    assert blocked.data == {
        "ok": False,
        "error": "Unknown tool: write_range",
        "type": "UnknownTool",
    }
    assert session.workbook_path.read_bytes() == before


def test_view_image_only_accepts_page_from_most_recent_render(
    sample_workbook: Path, tmp_path: Path
) -> None:
    session = WorkbookSession.create(sample_workbook, tmp_path / "run")
    tools = SpreadsheetToolRegistry(
        session,
        enable_code=False,
        allowed_tools={"render_workbook", "view_image"},
    )
    render_root = session.paths.artifacts / "render"
    stale = render_root / "render-old" / "page.png"
    latest = render_root / "render-new" / "page.png"
    for path in (stale, latest):
        path.parent.mkdir(parents=True, exist_ok=True)
        Image.new("RGB", (2, 2), "white").save(path)

    before_render = tools.invoke("view_image", {"image_path": str(stale)}).data
    assert before_render["ok"] is False
    assert "render_workbook before" in before_render["error"]

    tools._last_render = {"pages": [{"image_path": str(latest.resolve())}]}
    stale_result = tools.invoke("view_image", {"image_path": str(stale)}).data
    assert stale_result["ok"] is False
    assert "most recent render_workbook" in stale_result["error"]

    latest_result = tools.invoke("view_image", {"image_path": str(latest)})
    assert latest_result.data["ok"] is True
    assert latest_result.image_path == latest.resolve()


def test_range_to_latex_escapes_values_and_reports_structure(
    sample_workbook: Path, tmp_path: Path
) -> None:
    workbook = load_workbook(sample_workbook)
    workbook["Sales"]["A2"] = "unsafe &_%$#{}~^\\\nnext"
    workbook.save(sample_workbook)
    workbook.close()

    session = WorkbookSession.create(sample_workbook, tmp_path / "run")
    tools = SpreadsheetToolRegistry(
        session, enable_code=False, allowed_tools={"range_to_latex"}
    )
    before = session.workbook_path.read_bytes()
    result = tools.invoke("range_to_latex", {"sheet": "Sales", "range_ref": "A1:D5"}).data

    assert result["ok"] is True
    assert (result["rows"], result["columns"], result["cell_count"]) == (5, 4, 20)
    assert result["latex"].startswith(r"\begin{tabular}{llll}")
    assert result["latex"].endswith(r"\end{tabular}")
    assert r"unsafe \&\_\%\$\#\{\}\textasciitilde{}\textasciicircum{}" in result[
        "latex"
    ]
    assert r"\textbackslash{}\newline{}next" in result["latex"]
    assert result["merged_ranges"] == ["A5:B5"]
    heading_style = next(
        style for style in result["style_summary"]["styles"] if "A1" in style["sample_cells"]
    )
    assert heading_style["font"]["bold"] is True
    assert heading_style["fill"] == "FF336699"
    assert session.workbook_path.read_bytes() == before


def test_range_to_latex_enforces_cell_and_output_limits(
    sample_workbook: Path, tmp_path: Path
) -> None:
    workbook = load_workbook(sample_workbook)
    workbook["Sales"]["A2"] = "\\" * 30_000
    workbook.save(sample_workbook)
    workbook.close()

    session = WorkbookSession.create(sample_workbook, tmp_path / "run")
    tools = SpreadsheetToolRegistry(session, enable_code=False)
    bounded = tools.invoke("range_to_latex", {"sheet": "Sales", "range_ref": "A2"}).data

    assert bounded["ok"] is True
    assert bounded["latex_truncated"] is True
    assert bounded["truncated_cell_count"] == 1
    assert len(bounded["latex"]) <= bounded["limits"]["max_latex_chars"] == 65_536

    too_large = tools.invoke(
        "range_to_latex", {"sheet": "Sales", "range_ref": "A1:Z20"}
    ).data
    assert too_large["ok"] is False
    assert too_large["type"] == "ToolInputError"
    assert "limit is 500" in too_large["error"]
