from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


@pytest.fixture
def sample_workbook(tmp_path: Path) -> Path:
    path = tmp_path / "sample.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sales"
    sheet.append(["Item", "Qty", "Price", "Total"])
    sheet.append(["Apple", 2, 3.5, "=B2*C2"])
    sheet.append(["Pear", 4, 2.0, "=B3*C3"])
    sheet["A1"].font = Font(bold=True, color="FFFFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="FF336699")
    sheet.merge_cells("A5:B5")
    sheet["A5"] = "Notes"
    second = workbook.create_sheet("Lookup")
    second.append(["Key", "Value"])
    second.append(["tax", 0.1])
    workbook.save(path)
    return path
