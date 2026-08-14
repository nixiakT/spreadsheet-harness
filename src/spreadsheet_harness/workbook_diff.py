"""Harness-computed semantic effect footprints for workbook revisions."""

from __future__ import annotations

import hashlib
import posixpath
import warnings
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.packaging.relationship import get_dependents, get_rels_path
from openpyxl.xml.functions import fromstring

from .evidence_contract import CellRange, EffectKind, EvidenceScope

_DEFAULT_MAX_SCANNED_CELLS = 250_000
_DEFAULT_MAX_CHANGED_CELLS = 5_000
_DEFAULT_MAX_RANGES = 256
_MAX_BOUNDING_RANGE_CELLS = 10_000
_UNSUPPORTED_PART_PREFIXES = (
    "customUI/",
    "customXml/",
    "xl/activeX/",
    "xl/ctrlProps/",
    "xl/diagrams/",
    "xl/embeddings/",
    "xl/externalLinks/",
    "xl/model/",
    "xl/pivotCache/",
    "xl/pivotTables/",
    "xl/queryTables/",
    "xl/richData/",
    "xl/slicerCaches/",
    "xl/slicers/",
    "xl/threadedComments/",
    "xl/timelines/",
    "xl/webExtensions/",
)
_UNSUPPORTED_PART_NAMES = frozenset(
    {
        "xl/connections.xml",
        "xl/metadata.xml",
        "xl/vbaProject.bin",
        "xl/vbaProjectSignature.bin",
    }
)
_CORE_MODIFIED_TAG = "{http://purl.org/dc/terms/}modified"
_CALC_PR_TAG = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}calcPr"
_CONTENT_TYPES_NAMESPACE = (
    "http://schemas.openxmlformats.org/package/2006/content-types"
)
_PACKAGE_RELATIONSHIPS_NAMESPACE = (
    "http://schemas.openxmlformats.org/package/2006/relationships"
)
_SPREADSHEETML_NAMESPACE = (
    "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
)
_OFFICE_RELATIONSHIPS_NAMESPACE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
)
_XML_NAMESPACE = "http://www.w3.org/XML/1998/namespace"
_ROOT_RELATIONSHIPS_PART = "_rels/.rels"
_CUSTOM_PROPERTIES_PART = "docProps/custom.xml"
_CUSTOM_PROPERTIES_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.custom-properties+xml"
)
_CUSTOM_PROPERTIES_RELATIONSHIP_TYPES = frozenset(
    {
        (
            "http://schemas.openxmlformats.org/officeDocument/2006/relationships/"
            "custom-properties"
        ),
        "http://purl.oclc.org/ooxml/officeDocument/relationships/custom-properties",
    }
)
_DOCUMENT_PROPERTY_PARTS = frozenset(
    {"docProps/app.xml", "docProps/core.xml", _CUSTOM_PROPERTIES_PART}
)
_IGNORED_APP_PROPERTY_TAGS = frozenset(
    {"Application", "AppVersion", "TotalTime"}
)
_CORE_DEFAULT_EQUIVALENTS = {
    "description": frozenset({""}),
    "language": frozenset({"", "en-US"}),
    "lastModifiedBy": frozenset({""}),
    "revision": frozenset({"", "0"}),
    "subject": frozenset({""}),
    "title": frozenset({""}),
}
_IGNORABLE_UNUSED_DEFAULT_CONTENT_TYPES = {
    "jpeg": "image/jpeg",
    "jpg": "image/jpeg",
    "png": "image/png",
}
_DOCUMENT_PROPERTY_ROOT_TAGS = {
    "docProps/app.xml": (
        "{http://schemas.openxmlformats.org/officeDocument/2006/extended-properties}"
        "Properties"
    ),
    "docProps/core.xml": (
        "{http://schemas.openxmlformats.org/package/2006/metadata/core-properties}"
        "coreProperties"
    ),
    "docProps/custom.xml": (
        "{http://schemas.openxmlformats.org/officeDocument/2006/custom-properties}"
        "Properties"
    ),
}
_SUPPORTED_WORKSHEET_CHILDREN = frozenset(
    {
        "autoFilter",
        "colBreaks",
        "cols",
        "conditionalFormatting",
        "dataValidations",
        "dimension",
        "drawing",
        "headerFooter",
        "hyperlinks",
        "mergeCells",
        "pageMargins",
        "pageSetup",
        "phoneticPr",
        "printOptions",
        "rowBreaks",
        "sheetData",
        "sheetFormatPr",
        "sheetPr",
        "sheetProtection",
        "sheetViews",
        "sortState",
        "tableParts",
    }
)
_SUPPORTED_RELATIONSHIP_KINDS = {
    "": frozenset(
        {
            "officeDocument",
            "core-properties",
            "extended-properties",
            "custom-properties",
        }
    ),
    "xl/workbook.xml": frozenset({"worksheet", "styles", "sharedStrings", "theme"}),
}
_LIBREOFFICE_CALC_EXTENSION_URI = "{7626C862-2A13-11E5-B345-FEFF819CDC9F}"


def _scalar(value: Any) -> Any:
    if value is None or isinstance(value, bool | int | float | str):
        return value
    if isinstance(value, datetime | date | time):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return str(value)


def _color(color: Any) -> tuple[Any, ...] | None:
    if color is None:
        return None
    return (
        getattr(color, "type", None),
        getattr(color, "rgb", None),
        getattr(color, "indexed", None),
        getattr(color, "theme", None),
        getattr(color, "tint", None),
        getattr(color, "auto", None),
    )


def _side(side: Any) -> tuple[Any, ...] | None:
    if side is None:
        return None
    return (getattr(side, "style", None), _color(getattr(side, "color", None)))


def _style(cell: Any) -> tuple[Any, ...]:
    font = cell.font
    fill = cell.fill
    border = cell.border
    alignment = cell.alignment
    protection = cell.protection
    return (
        cell.number_format,
        (
            font.name,
            font.sz,
            font.b,
            font.i,
            font.u,
            font.strike,
            font.vertAlign,
            _color(font.color),
        ),
        (
            fill.fill_type,
            _color(fill.fgColor),
            _color(fill.bgColor),
        ),
        (
            _side(border.left),
            _side(border.right),
            _side(border.top),
            _side(border.bottom),
            _side(border.diagonal),
            _side(border.vertical),
            _side(border.horizontal),
            border.diagonalUp,
            border.diagonalDown,
            border.outline,
        ),
        (
            alignment.horizontal,
            alignment.vertical,
            alignment.text_rotation,
            alignment.wrap_text,
            alignment.shrink_to_fit,
            alignment.indent,
            alignment.relativeIndent,
            alignment.justifyLastLine,
            alignment.readingOrder,
        ),
        (protection.locked, protection.hidden),
        cell.quotePrefix,
    )


def _formula_value_signature(cell: Any) -> Any:
    value = cell.value
    if cell.data_type != "f" or isinstance(value, str) or value is None:
        return _scalar(value)
    attributes = vars(value)
    if not attributes or not all(isinstance(key, str) for key in attributes):
        raise ValueError("unsupported formula value representation")
    return (
        f"{type(value).__module__}.{type(value).__qualname__}",
        tuple(sorted((key, _scalar(item)) for key, item in attributes.items())),
    )


def _formula_ranges(cell: Any, sheet_name: str) -> tuple[CellRange, ...]:
    if cell.data_type != "f":
        return ()
    reference = getattr(cell.value, "ref", None)
    if reference is None:
        return (CellRange(sheet_name, cell.column, cell.row, cell.column, cell.row),)
    if not isinstance(reference, str) or not reference:
        raise ValueError("formula range reference is invalid")
    return (CellRange.parse(sheet_name, reference),)


def _cell_content(cell: Any) -> tuple[Any, ...]:
    hyperlink = cell.hyperlink
    comment = cell.comment
    return (
        _formula_value_signature(cell),
        cell.data_type,
        (
            getattr(hyperlink, "target", None),
            getattr(hyperlink, "location", None),
            getattr(hyperlink, "tooltip", None),
        )
        if hyperlink is not None
        else None,
        (comment.text, comment.author) if comment is not None else None,
    )


def _row_dimension(value: Any) -> tuple[Any, ...]:
    return (
        value.height,
        value.hidden,
        value.outlineLevel,
        value.collapsed,
        value.thickTop,
        value.thickBot,
        value.style_id,
    )


def _column_dimension(value: Any) -> tuple[Any, ...]:
    return (
        value.min,
        value.max,
        value.width,
        value.hidden,
        value.bestFit,
        value.outlineLevel,
        value.collapsed,
        value.style_id,
    )


def _xml_signature(element: Any) -> tuple[Any, ...]:
    """Canonicalize generated XML without depending on prefix or attribute order."""

    text = element.text
    if text is not None and not text.strip():
        text = None
    tail = element.tail
    if tail is not None and not tail.strip():
        tail = None
    return (
        str(element.tag),
        tuple(sorted((str(key), str(value)) for key, value in element.attrib.items())),
        text,
        tail,
        tuple(_xml_signature(child) for child in element),
    )


def _residual_xml_signature(
    element: Any,
    *,
    part_name: str,
    allow_benign_libreoffice_calc_extension: bool = False,
) -> tuple[Any, ...]:
    """Canonicalize residual XML while ignoring only proven package churn."""

    text = element.text
    if element.tag == _CORE_MODIFIED_TAG and part_name == "docProps/core.xml":
        # openpyxl refreshes this timestamp on every save. It does not affect
        # workbook calculation, rendering, edit scope, or user-visible cells.
        text = None
    elif text is not None and not text.strip():
        text = None
    tail = element.tail
    if tail is not None and not tail.strip():
        tail = None
    attributes = {str(key): str(value) for key, value in element.attrib.items()}
    if element.tag == _CALC_PR_TAG and part_name == "xl/workbook.xml":
        # Session writes force a future recalculation after every staged save.
        # Normalize only the equivalent default/enable forms introduced by that
        # policy; material settings such as manual calculation remain visible.
        if attributes.get("calcMode") in {None, "auto"}:
            attributes["calcMode"] = "auto"
        for name in ("fullCalcOnLoad", "forceFullCalc"):
            if attributes.get(name) in {None, "1", "true"}:
                attributes[name] = "1"
    return (
        str(element.tag),
        tuple(sorted(attributes.items())),
        text,
        tail,
        tuple(
            _residual_xml_signature(child, part_name=part_name)
            for child in element
            if not (
                allow_benign_libreoffice_calc_extension
                and _benign_libreoffice_calc_extension(child)
            )
        ),
    )


def _part_payload_signature(part_name: str, payload: bytes) -> Any:
    if part_name.endswith((".xml", ".rels", ".vml")):
        return _residual_xml_signature(
            fromstring(payload),
            part_name=part_name,
            allow_benign_libreoffice_calc_extension=(part_name == "xl/workbook.xml"),
        )
    return hashlib.sha256(payload).hexdigest()


def _normalized_workbook_calculation_signature(element: Any) -> tuple[Any, ...] | None:
    """Retain material calculation policy while dropping producer defaults."""

    attributes = {str(key): str(value) for key, value in element.attrib.items()}
    attributes.pop("calcId", None)
    benign_defaults = {
        "calcMode": {"auto"},
        "fullCalcOnLoad": {"1", "true"},
        "forceFullCalc": {"1", "true"},
        "refMode": {"A1"},
        "iterate": {"0", "false"},
        "iterateCount": {"100"},
        # LibreOffice and Excel-family writers use both values as their
        # producer defaults while the harness forces a future recalculation.
        "iterateDelta": {"0.0001", "0.001"},
        "fullPrecision": {"1", "true"},
        "calcCompleted": {"1", "true"},
        "calcOnSave": {"1", "true"},
        "concurrentCalc": {"1", "true"},
    }
    for name, values in benign_defaults.items():
        if attributes.get(name) in values:
            attributes.pop(name)
    text = _normalized_leaf_text(element)
    tail = (element.tail or "").strip()
    children = tuple(_xml_signature(child) for child in element)
    if not attributes and not text and not tail and not children:
        return None
    return (tuple(sorted(attributes.items())), text, tail, children)


def _normalized_workbook_properties_signature(element: Any) -> tuple[Any, ...] | None:
    """Normalize only absent/default producer attributes on workbookPr."""

    attributes = {str(key): str(value) for key, value in element.attrib.items()}
    benign_defaults = {
        "backupFile": {"0", "false"},
        "date1904": {"0", "false"},
        "showObjects": {"all"},
    }
    for name, values in benign_defaults.items():
        if attributes.get(name) in values:
            attributes.pop(name)
    text = _normalized_leaf_text(element)
    tail = (element.tail or "").strip()
    children = tuple(_xml_signature(child) for child in element)
    if not attributes and not text and not tail and not children:
        return None
    return (tuple(sorted(attributes.items())), text, tail, children)


def _workbook_collection_residual(element: Any, *, local_name: str) -> tuple[Any, ...] | None:
    """Retain XML details omitted from sheets/defined-names comparators."""

    container_residual = (
        tuple(sorted((str(key), str(value)) for key, value in element.attrib.items())),
        _normalized_leaf_text(element),
        (element.tail or "").strip(),
    )
    if local_name == "sheets":
        item_tag = f"{{{_SPREADSHEETML_NAMESPACE}}}sheet"
        handled_attributes = {
            "name",
            "sheetId",
            "state",
            f"{{{_OFFICE_RELATIONSHIPS_NAMESPACE}}}id",
        }
        semantic_text = False
    else:
        item_tag = f"{{{_SPREADSHEETML_NAMESPACE}}}definedName"
        handled_attributes = {"name", "localSheetId", "hidden", "function"}
        semantic_text = True
    items: list[tuple[Any, ...]] = []
    for child in element:
        if str(child.tag) != item_tag:
            items.append(("unexpected", _xml_signature(child)))
            continue
        residual_attributes = tuple(
            sorted(
                (str(key), str(value))
                for key, value in child.attrib.items()
                if str(key) not in handled_attributes
            )
        )
        residual = (
            residual_attributes,
            "" if semantic_text else _normalized_leaf_text(child),
            (child.tail or "").strip(),
            tuple(_xml_signature(grandchild) for grandchild in child),
        )
        if any(residual):
            items.append(("item", residual))
    if not any(container_residual) and not items:
        return None
    return (*container_residual, tuple(items))


def _unclassified_workbook_signature(package: ZipFile) -> tuple[Any, ...]:
    """Project workbook.xml down to content not owned by semantic comparators."""

    root = _optional_xml_root(package, "xl/workbook.xml")
    if root is None:
        raise ValueError("workbook OOXML part is missing")
    if str(root.tag) != f"{{{_SPREADSHEETML_NAMESPACE}}}workbook":
        raise ValueError("unsupported workbook OOXML root")
    handled = {"bookViews", "workbookProtection"}
    residual: list[tuple[Any, ...]] = []
    calculation_count = 0
    for child in root:
        local_name = _local_name(child.tag)
        if _namespace(child.tag) == _SPREADSHEETML_NAMESPACE and local_name in handled:
            continue
        if (
            _namespace(child.tag) == _SPREADSHEETML_NAMESPACE
            and local_name in {"definedNames", "sheets"}
        ):
            signature = _workbook_collection_residual(
                child,
                local_name=local_name,
            )
            if signature is not None:
                residual.append((local_name, signature))
            continue
        if child.tag == _CALC_PR_TAG:
            calculation_count += 1
            if calculation_count > 1:
                raise ValueError("ambiguous workbook calculation properties")
            signature = _normalized_workbook_calculation_signature(child)
            if signature is not None:
                residual.append((local_name, signature))
            continue
        if (
            _namespace(child.tag) == _SPREADSHEETML_NAMESPACE
            and local_name == "workbookPr"
        ):
            signature = _normalized_workbook_properties_signature(child)
            if signature is not None:
                residual.append((local_name, signature))
            continue
        if (
            _namespace(child.tag) == _SPREADSHEETML_NAMESPACE
            and local_name == "fileVersion"
            and child.attrib == {"appName": "Calc"}
            and not list(child)
            and not _normalized_leaf_text(child)
            and not (child.tail and child.tail.strip())
        ):
            continue
        if _benign_libreoffice_calc_extension(child):
            continue
        residual.append((local_name, _xml_signature(child)))
    return (
        tuple(sorted((str(key), str(value)) for key, value in root.attrib.items())),
        _normalized_leaf_text(root),
        (root.tail or "").strip(),
        tuple(residual),
    )


def _local_name(tag: Any) -> str:
    return str(tag).rsplit("}", 1)[-1]


def _namespace(tag: Any) -> str:
    value = str(tag)
    return value[1:].split("}", 1)[0] if value.startswith("{") else ""


def _benign_libreoffice_calc_extension(element: Any) -> bool:
    """Recognize only LibreOffice's redundant marker for Excel A1 syntax."""

    if (
        str(element.tag) != f"{{{_SPREADSHEETML_NAMESPACE}}}extLst"
        or element.attrib
        or _normalized_leaf_text(element)
        or (element.tail and element.tail.strip())
        or len(element) != 1
    ):
        return False
    extension = element[0]
    if (
        str(extension.tag) != f"{{{_SPREADSHEETML_NAMESPACE}}}ext"
        or extension.attrib != {"uri": _LIBREOFFICE_CALC_EXTENSION_URI}
        or _normalized_leaf_text(extension)
        or (extension.tail and extension.tail.strip())
        or len(extension) != 1
    ):
        return False
    calc = extension[0]
    return bool(
        str(calc.tag) == "{http://schemas.libreoffice.org/}extCalcPr"
        and calc.attrib == {"stringRefSyntax": "ExcelA1"}
        and not list(calc)
        and not _normalized_leaf_text(calc)
        and not (calc.tail and calc.tail.strip())
    )


def _opaque_extension_signature(package: ZipFile) -> tuple[Any, ...]:
    """Track XML content outside the semantic comparators so it cannot piggyback."""

    records: list[tuple[Any, ...]] = []
    for part_name in sorted(package.namelist()):
        if part_name.endswith("/") or not part_name.endswith((".xml", ".rels")):
            continue
        if part_name.startswith(_UNSUPPORTED_PART_PREFIXES) or part_name in (
            _UNSUPPORTED_PART_NAMES
        ):
            continue
        if part_name in {
            "[Content_Types].xml",
            _ROOT_RELATIONSHIPS_PART,
            *_DOCUMENT_PROPERTY_PARTS,
        }:
            continue
        root = fromstring(package.read(part_name))
        if part_name.endswith(".rels"):
            if (
                str(root.tag) != f"{{{_PACKAGE_RELATIONSHIPS_NAMESPACE}}}Relationships"
                or root.attrib
                or _normalized_leaf_text(root)
                or (root.tail and root.tail.strip())
            ):
                raise ValueError(f"unsupported OOXML relationships root: {part_name}")
            source_part = ""
            if part_name != _ROOT_RELATIONSHIPS_PART:
                prefix, _, filename = part_name.rpartition("/_rels/")
                if not filename.endswith(".rels"):
                    raise ValueError(f"invalid OOXML relationships path: {part_name}")
                source_part = f"{prefix}/{filename.removesuffix('.rels')}".lstrip("/")
            supported_kinds = _SUPPORTED_RELATIONSHIP_KINDS.get(source_part)
            if supported_kinds is None and source_part.startswith("xl/worksheets/"):
                supported_kinds = frozenset({"drawing", "table"})
            elif supported_kinds is None and source_part.startswith("xl/drawings/"):
                supported_kinds = frozenset({"chart", "image"})
            else:
                supported_kinds = supported_kinds or frozenset()
            identifiers: set[str] = set()
            for index, child in enumerate(root):
                if (
                    str(child.tag)
                    != f"{{{_PACKAGE_RELATIONSHIPS_NAMESPACE}}}Relationship"
                    or set(child.attrib) - {"Id", "Type", "Target", "TargetMode"}
                    or list(child)
                    or _normalized_leaf_text(child)
                    or (child.tail and child.tail.strip())
                ):
                    raise ValueError(f"unsupported OOXML relationship content: {part_name}")
                identifier = str(child.attrib.get("Id", ""))
                relation_type = str(child.attrib.get("Type", ""))
                target = str(child.attrib.get("Target", ""))
                target_mode = str(child.attrib.get("TargetMode", "Internal"))
                if (
                    not identifier
                    or identifier in identifiers
                    or not relation_type
                    or not target
                    or target_mode not in {"Internal", "External"}
                ):
                    raise ValueError(f"invalid OOXML relationship: {part_name}")
                identifiers.add(identifier)
                kind = relation_type.rsplit("/", 1)[-1]
                if kind not in supported_kinds:
                    records.append((part_name, index, _xml_signature(child)))
            continue

        worksheet_part = part_name.startswith("xl/worksheets/")

        def visit(
            element: Any,
            path: tuple[int, ...],
            parent_local: str | None,
            *,
            current_part: str = part_name,
            current_worksheet: bool = worksheet_part,
        ) -> None:
            element_namespace = _namespace(element.tag)
            local_name = _local_name(element.tag)
            if (
                current_part == "xl/workbook.xml"
                and _benign_libreoffice_calc_extension(element)
            ):
                return
            unsupported = False
            if current_worksheet:
                unsupported = element_namespace != _SPREADSHEETML_NAMESPACE
                if path == ():
                    unsupported = unsupported or local_name != "worksheet"
                elif parent_local == "worksheet":
                    unsupported = unsupported or local_name not in _SUPPORTED_WORKSHEET_CHILDREN
                elif parent_local == "sheetData":
                    unsupported = unsupported or local_name != "row"
                elif parent_local == "row":
                    unsupported = unsupported or local_name != "c"
                elif parent_local == "c":
                    unsupported = unsupported or local_name not in {"f", "is", "v"}
            elif local_name == "extLst" or element_namespace.endswith(
                "/markup-compatibility/2006"
            ):
                unsupported = True
            if unsupported:
                records.append((current_part, path, _xml_signature(element)))
                return
            if current_worksheet and element.tail and element.tail.strip():
                records.append((current_part, path, "tail", element.tail))
            if current_worksheet and local_name == "c":
                raw_style = element.attrib.get("s")
                if raw_style is not None:
                    try:
                        style_index = int(raw_style)
                    except ValueError as exc:
                        raise ValueError(
                            f"invalid worksheet cell style index: {current_part}"
                        ) from exc
                    if style_index < 0 or str(style_index) != raw_style:
                        raise ValueError(
                            f"invalid worksheet cell style index: {current_part}"
                        )
                for attribute in sorted(set(element.attrib) - {"r", "s", "t"}):
                    records.append(
                        (
                            current_part,
                            path,
                            "cell_attribute",
                            str(attribute),
                            str(element.attrib[attribute]),
                        )
                    )
            if current_worksheet and local_name == "v" and element.attrib:
                records.append(
                    (
                        current_part,
                        path,
                        "value_attributes",
                        tuple(sorted((str(key), str(value)) for key, value in element.attrib.items())),
                    )
                )
            for attribute, value in element.attrib.items():
                attribute_namespace = _namespace(attribute)
                if attribute_namespace not in {
                    "",
                    _OFFICE_RELATIONSHIPS_NAMESPACE,
                    _XML_NAMESPACE,
                }:
                    records.append(
                        (current_part, path, "attribute", str(attribute), str(value))
                    )
            for index, child in enumerate(element):
                visit(child, (*path, index), local_name)

        visit(root, (), None)
    return tuple(records)


def _optional_xml_root(package: ZipFile, part_name: str) -> Any | None:
    try:
        payload = package.read(part_name)
    except KeyError:
        return None
    return fromstring(payload)


def _normalized_leaf_text(element: Any) -> str:
    return (element.text or "").strip()


def _document_property_part_signature(package: ZipFile, part_name: str) -> tuple[Any, ...]:
    """Normalize only documented producer churn in OOXML document metadata."""

    root = _optional_xml_root(package, part_name)
    if root is None:
        return ()
    if str(root.tag) != _DOCUMENT_PROPERTY_ROOT_TAGS[part_name] or root.attrib:
        raise ValueError(f"unsupported {part_name} root element")
    if _normalized_leaf_text(root) or (root.tail and root.tail.strip()):
        raise ValueError(f"unsupported {part_name} root text")
    if part_name == "docProps/custom.xml":
        properties: list[tuple[Any, ...]] = []
        names: set[str] = set()
        for child in root:
            if _local_name(child.tag) != "property":
                raise ValueError("unsupported custom document-property element")
            name = str(child.attrib.get("name", ""))
            if not name or name in names:
                raise ValueError("ambiguous custom document-property name")
            names.add(name)
            attributes = tuple(
                sorted(
                    (str(key), str(value))
                    for key, value in child.attrib.items()
                    if _local_name(key) != "pid"
                )
            )
            properties.append(
                (
                    name,
                    attributes,
                    tuple(_xml_signature(grandchild) for grandchild in child),
                )
            )
        return tuple(sorted(properties, key=repr))

    properties = []
    for child in root:
        local_name = _local_name(child.tag)
        text = _normalized_leaf_text(child)
        if part_name == "docProps/core.xml":
            if child.tag == _CORE_MODIFIED_TAG:
                continue
            if text in _CORE_DEFAULT_EQUIVALENTS.get(local_name, frozenset()):
                continue
        elif part_name == "docProps/app.xml":
            if local_name in _IGNORED_APP_PROPERTY_TAGS:
                continue
            if local_name == "Template" and not text and not list(child):
                continue
        properties.append(_xml_signature(child))
    return tuple(sorted(properties, key=repr))


def _document_properties_signature(package: ZipFile) -> tuple[Any, ...]:
    custom_state = _custom_properties_package_state(package)
    custom_signature = custom_state[1] if custom_state[0] == "material" else ()
    return (
        (
            "docProps/app.xml",
            _document_property_part_signature(package, "docProps/app.xml"),
        ),
        (
            "docProps/core.xml",
            _document_property_part_signature(package, "docProps/core.xml"),
        ),
        (_CUSTOM_PROPERTIES_PART, custom_signature),
    )


def _root_relationship_records(
    package: ZipFile,
) -> tuple[tuple[str, str, str, str], ...]:
    root = _optional_xml_root(package, _ROOT_RELATIONSHIPS_PART)
    if root is None:
        raise ValueError("root OOXML relationships part is missing")
    if (
        str(root.tag) != f"{{{_PACKAGE_RELATIONSHIPS_NAMESPACE}}}Relationships"
        or root.attrib
        or _normalized_leaf_text(root)
        or (root.tail and root.tail.strip())
    ):
        raise ValueError("unsupported root OOXML relationships root")
    relationships: list[tuple[str, str, str, str]] = []
    relationship_ids: set[str] = set()
    for child in root:
        if str(child.tag) != f"{{{_PACKAGE_RELATIONSHIPS_NAMESPACE}}}Relationship":
            raise ValueError("unsupported root OOXML relationship element")
        if (
            set(child.attrib) - {"Id", "Type", "Target", "TargetMode"}
            or list(child)
            or _normalized_leaf_text(child)
            or (child.tail and child.tail.strip())
        ):
            raise ValueError("unsupported root OOXML relationship content")
        relationship_id = str(child.attrib.get("Id", ""))
        relationship_type = str(child.attrib.get("Type", ""))
        target = posixpath.normpath(str(child.attrib.get("Target", "")).lstrip("/"))
        target_mode = str(child.attrib.get("TargetMode", "Internal"))
        if (
            not relationship_id
            or relationship_id in relationship_ids
            or not relationship_type
            or not target
            or target_mode not in {"Internal", "External"}
        ):
            raise ValueError("invalid root OOXML relationship")
        relationship_ids.add(relationship_id)
        relationships.append(
            (relationship_id, relationship_type, target, target_mode)
        )
    semantic_relationships = [item[1:] for item in relationships]
    if len(semantic_relationships) != len(set(semantic_relationships)):
        if any(
            relationship_type in _CUSTOM_PROPERTIES_RELATIONSHIP_TYPES
            or target == _CUSTOM_PROPERTIES_PART
            for relationship_type, target, _ in semantic_relationships
        ):
            raise ValueError(
                f"duplicate {_CUSTOM_PROPERTIES_PART} root relationships are ambiguous"
            )
        raise ValueError("duplicate root OOXML relationships are ambiguous")
    return tuple(relationships)


def _root_relationships_signature(package: ZipFile) -> tuple[Any, ...]:
    relationships = _root_relationship_records(package)
    custom_state = _custom_properties_package_state(
        package,
        root_relationships=relationships,
    )
    normalize_empty_custom = custom_state[0] in {"absent", "empty"}
    return tuple(
        sorted(
            (relationship_type, target, target_mode)
            for _, relationship_type, target, target_mode in relationships
            if not (
                normalize_empty_custom
                and relationship_type in _CUSTOM_PROPERTIES_RELATIONSHIP_TYPES
                and target == _CUSTOM_PROPERTIES_PART
                and target_mode == "Internal"
            )
        )
    )


def _package_part_signatures(
    package: ZipFile,
    *,
    semantic_parts: frozenset[str],
) -> dict[str, Any]:
    """Account for every payload part, excluding ZIP-container-only metadata."""

    part_names = [name for name in package.namelist() if not name.endswith("/")]
    if len(part_names) != len(set(part_names)):
        raise ValueError("duplicate OOXML part names are ambiguous")
    return {
        part_name: (
            _part_payload_signature(part_name, package.read(part_name))
            if part_name in semantic_parts
            else hashlib.sha256(package.read(part_name)).hexdigest()
        )
        for part_name in sorted(part_names)
    }


def _changed_package_parts(
    before_package: ZipFile,
    after_package: ZipFile,
) -> tuple[str, ...]:
    semantic_parts = _accounted_parts(before_package) | _accounted_parts(after_package)
    before_signatures = _package_part_signatures(
        before_package,
        semantic_parts=semantic_parts,
    )
    after_signatures = _package_part_signatures(
        after_package,
        semantic_parts=semantic_parts,
    )
    return tuple(
        sorted(
            part_name
            for part_name in before_signatures.keys() | after_signatures.keys()
            if before_signatures.get(part_name) != after_signatures.get(part_name)
        )
    )


def _relationship_kind(relationship: Any) -> str:
    return str(relationship.Type).rsplit("/", 1)[-1]


def _accounted_parts(package: ZipFile) -> frozenset[str]:
    """Resolve only payload parts reachable through supported OOXML relationships."""

    available = {name for name in package.namelist() if not name.endswith("/")}
    accounted = {
        part_name
        for part_name in (
            "[Content_Types].xml",
            _ROOT_RELATIONSHIPS_PART,
            *_DOCUMENT_PROPERTY_PARTS,
            "xl/workbook.xml",
            "xl/_rels/workbook.xml.rels",
            "xl/styles.xml",
            "xl/sharedStrings.xml",
        )
        if part_name in available
    }
    workbook_relationships_path = "xl/_rels/workbook.xml.rels"
    if workbook_relationships_path not in available:
        return frozenset(accounted)

    workbook_relationships = get_dependents(package, workbook_relationships_path)
    for relationship in workbook_relationships:
        kind = _relationship_kind(relationship)
        target = str(relationship.target).lstrip("/")
        if kind not in {"worksheet", "styles", "sharedStrings", "theme"}:
            continue
        if target in available:
            accounted.add(target)
        if kind != "worksheet":
            continue

        worksheet_relationships_path = get_rels_path(target).lstrip("/")
        if worksheet_relationships_path not in available:
            continue
        accounted.add(worksheet_relationships_path)
        worksheet_relationships = get_dependents(package, worksheet_relationships_path)
        for worksheet_relationship in worksheet_relationships:
            worksheet_kind = _relationship_kind(worksheet_relationship)
            worksheet_target = str(worksheet_relationship.target).lstrip("/")
            if worksheet_kind == "table":
                if worksheet_target in available:
                    accounted.add(worksheet_target)
                continue
            if worksheet_kind != "drawing":
                continue
            if worksheet_target in available:
                accounted.add(worksheet_target)
            drawing_relationships_path = get_rels_path(worksheet_target).lstrip("/")
            if drawing_relationships_path not in available:
                continue
            accounted.add(drawing_relationships_path)
            drawing_relationships = get_dependents(package, drawing_relationships_path)
            for drawing_relationship in drawing_relationships:
                if _relationship_kind(drawing_relationship) not in {"chart", "image"}:
                    continue
                drawing_target = str(drawing_relationship.target).lstrip("/")
                if drawing_target in available:
                    accounted.add(drawing_target)
    return frozenset(accounted)


def _workbook_relationship_targets(
    package: ZipFile,
    relationship_kind: str,
) -> frozenset[str]:
    relationships_path = "xl/_rels/workbook.xml.rels"
    if relationships_path not in package.namelist():
        return frozenset()
    return frozenset(
        str(relationship.target).lstrip("/")
        for relationship in get_dependents(package, relationships_path)
        if _relationship_kind(relationship) == relationship_kind
    )


def _shared_strings_are_fully_referenced(package: ZipFile) -> bool:
    targets = _workbook_relationship_targets(package, "sharedStrings")
    if not targets:
        return True
    if len(targets) != 1:
        raise ValueError("ambiguous shared-strings relationship")
    target = next(iter(targets))
    root = _optional_xml_root(package, target)
    if root is None or str(root.tag) != f"{{{_SPREADSHEETML_NAMESPACE}}}sst":
        raise ValueError("invalid shared-strings OOXML part")
    strings = list(root)
    if any(str(item.tag) != f"{{{_SPREADSHEETML_NAMESPACE}}}si" for item in strings):
        raise ValueError("unsupported shared-strings OOXML content")
    referenced: set[int] = set()
    for worksheet_part in _workbook_relationship_targets(package, "worksheet"):
        worksheet = _optional_xml_root(package, worksheet_part)
        if worksheet is None:
            raise ValueError(f"missing worksheet OOXML part: {worksheet_part}")
        for cell in worksheet.iter(f"{{{_SPREADSHEETML_NAMESPACE}}}c"):
            if cell.attrib.get("t") != "s":
                continue
            values = [
                child
                for child in cell
                if str(child.tag) == f"{{{_SPREADSHEETML_NAMESPACE}}}v"
            ]
            if len(values) != 1 or values[0].text is None:
                raise ValueError("invalid worksheet shared-string reference")
            try:
                index = int(values[0].text)
            except ValueError as exc:
                raise ValueError("invalid worksheet shared-string index") from exc
            if index < 0 or index >= len(strings):
                raise ValueError("worksheet shared-string index is out of range")
            referenced.add(index)
    return referenced == set(range(len(strings)))


def _worksheet_parts_by_name(package: ZipFile) -> dict[str, str]:
    root = _optional_xml_root(package, "xl/workbook.xml")
    relationships_path = "xl/_rels/workbook.xml.rels"
    if root is None or relationships_path not in package.namelist():
        raise ValueError("workbook sheet relationship plumbing is missing")
    relationships = {
        str(relationship.Id): relationship
        for relationship in get_dependents(package, relationships_path)
        if _relationship_kind(relationship) == "worksheet"
    }
    sheets = next(
        (
            child
            for child in root
            if str(child.tag) == f"{{{_SPREADSHEETML_NAMESPACE}}}sheets"
        ),
        None,
    )
    if sheets is None:
        raise ValueError("workbook sheets collection is missing")
    result: dict[str, str] = {}
    relationship_attribute = f"{{{_OFFICE_RELATIONSHIPS_NAMESPACE}}}id"
    for child in sheets:
        if str(child.tag) != f"{{{_SPREADSHEETML_NAMESPACE}}}sheet":
            raise ValueError("unsupported workbook sheet element")
        name = str(child.attrib.get("name", ""))
        relationship_id = str(child.attrib.get(relationship_attribute, ""))
        relationship = relationships.get(relationship_id)
        if not name or name in result or relationship is None:
            raise ValueError("ambiguous workbook sheet relationship")
        result[name] = str(relationship.target).lstrip("/")
    return result


def _worksheet_owned_parts(
    package: ZipFile,
    sheet_name: str,
    *,
    include_related: bool,
) -> frozenset[str]:
    part_name = _worksheet_parts_by_name(package).get(sheet_name)
    if part_name is None:
        return frozenset()
    result = {part_name}
    if not include_related:
        return frozenset(result)
    available = set(package.namelist())
    relationships_path = get_rels_path(part_name).lstrip("/")
    if relationships_path not in available:
        return frozenset(result)
    result.add(relationships_path)
    for relationship in get_dependents(package, relationships_path):
        if _relationship_kind(relationship) not in {"drawing", "table"}:
            continue
        target = str(relationship.target).lstrip("/")
        if target not in available:
            continue
        result.add(target)
        if _relationship_kind(relationship) != "drawing":
            continue
        drawing_relationships_path = get_rels_path(target).lstrip("/")
        if drawing_relationships_path not in available:
            continue
        result.add(drawing_relationships_path)
        for drawing_relationship in get_dependents(
            package,
            drawing_relationships_path,
        ):
            if _relationship_kind(drawing_relationship) not in {"chart", "image"}:
                continue
            drawing_target = str(drawing_relationship.target).lstrip("/")
            if drawing_target in available:
                result.add(drawing_target)
    return frozenset(result)


def _content_types(package: ZipFile) -> tuple[dict[str, str], dict[str, str]]:
    root = fromstring(package.read("[Content_Types].xml"))
    if (
        str(root.tag) != f"{{{_CONTENT_TYPES_NAMESPACE}}}Types"
        or root.attrib
        or _normalized_leaf_text(root)
        or (root.tail and root.tail.strip())
    ):
        raise ValueError("unsupported OOXML content-types root: [Content_Types].xml")
    defaults: dict[str, str] = {}
    overrides: dict[str, str] = {}
    for child in root:
        if list(child) or _normalized_leaf_text(child) or (child.tail and child.tail.strip()):
            raise ValueError("unsupported OOXML content-types child content")
        if str(child.tag) == f"{{{_CONTENT_TYPES_NAMESPACE}}}Default":
            if set(child.attrib) != {"Extension", "ContentType"}:
                raise ValueError("invalid OOXML default content-type declaration")
            key = str(child.attrib.get("Extension", "")).lower()
            value = str(child.attrib.get("ContentType", ""))
            destination = defaults
            if "." in key or "/" in key or "\\" in key:
                raise ValueError("invalid OOXML default content-type extension")
        elif str(child.tag) == f"{{{_CONTENT_TYPES_NAMESPACE}}}Override":
            if set(child.attrib) != {"PartName", "ContentType"}:
                raise ValueError("invalid OOXML override content-type declaration")
            raw_part_name = str(child.attrib.get("PartName", ""))
            if not raw_part_name.startswith("/"):
                raise ValueError("invalid OOXML override content-type part name")
            key = posixpath.normpath(raw_part_name).lstrip("/")
            value = str(child.attrib.get("ContentType", ""))
            destination = overrides
        else:
            raise ValueError("unsupported OOXML content-types element")
        if not key or not value or key in destination:
            raise ValueError("ambiguous OOXML content-types declaration")
        destination[key] = value
    return defaults, overrides


def _custom_properties_package_state(
    package: ZipFile,
    *,
    root_relationships: tuple[tuple[str, str, str, str], ...] | None = None,
) -> tuple[str, tuple[Any, ...]]:
    """Validate custom properties as one atomic OPC part/plumbing state."""

    part_names = [
        name
        for name in package.namelist()
        if not name.endswith("/") and name == _CUSTOM_PROPERTIES_PART
    ]
    if len(part_names) > 1:
        raise ValueError("duplicate custom-properties OOXML parts are ambiguous")
    part_present = bool(part_names)
    relationships = root_relationships or _root_relationship_records(package)
    relevant_relationships = [
        item
        for item in relationships
        if item[1] in _CUSTOM_PROPERTIES_RELATIONSHIP_TYPES
        or item[2] == _CUSTOM_PROPERTIES_PART
    ]
    valid_relationships = [
        item
        for item in relevant_relationships
        if item[1] in _CUSTOM_PROPERTIES_RELATIONSHIP_TYPES
        and item[2] == _CUSTOM_PROPERTIES_PART
        and item[3] == "Internal"
    ]

    defaults, overrides = _content_types(package)
    override_present = _CUSTOM_PROPERTIES_PART in overrides
    effective_content_type = overrides.get(_CUSTOM_PROPERTIES_PART)
    if effective_content_type is None:
        effective_content_type = defaults.get("xml")
    content_type_valid = effective_content_type == _CUSTOM_PROPERTIES_CONTENT_TYPE

    if not part_present:
        if relevant_relationships or override_present:
            raise ValueError(
                f"incomplete {_CUSTOM_PROPERTIES_PART} OOXML package plumbing: "
                "missing part"
            )
        return ("absent", ())
    if len(relevant_relationships) != 1 or len(valid_relationships) != 1:
        raise ValueError(
            f"incomplete {_CUSTOM_PROPERTIES_PART} OOXML package plumbing: "
            "expected one internal root relationship"
        )
    if not content_type_valid:
        raise ValueError(
            f"incomplete {_CUSTOM_PROPERTIES_PART} OOXML package plumbing: "
            "missing or invalid content type"
        )

    signature = _document_property_part_signature(package, _CUSTOM_PROPERTIES_PART)
    return ("material", signature) if signature else ("empty", ())


def _content_types_change_explained(
    before_package: ZipFile,
    after_package: ZipFile,
    *,
    accounted_parts: frozenset[str],
) -> bool:
    """Compare effective types, allowing only known producer declaration churn."""

    def effective(package: ZipFile) -> dict[str, str]:
        defaults, overrides = _content_types(package)
        custom_properties_state = _custom_properties_package_state(package)[0]
        result: dict[str, str] = {}
        for part_name in package.namelist():
            if not part_name or part_name.endswith("/") or part_name == "[Content_Types].xml":
                continue
            if (
                part_name == _CUSTOM_PROPERTIES_PART
                and custom_properties_state in {"absent", "empty"}
            ):
                continue
            content_type = overrides.get(part_name)
            if content_type is None:
                extension = part_name.rpartition(".")[2].lower()
                content_type = defaults.get(extension)
            if not content_type:
                raise ValueError(f"OOXML part has no content type: {part_name}")
            result[part_name] = content_type
        return result

    before_defaults, before_overrides = _content_types(before_package)
    after_defaults, after_overrides = _content_types(after_package)
    before_effective = effective(before_package)
    after_effective = effective(after_package)
    common_parts = before_effective.keys() & after_effective.keys()
    if any(before_effective[name] != after_effective[name] for name in common_parts):
        return False
    changed_effective_parts = before_effective.keys() ^ after_effective.keys()
    if not changed_effective_parts <= accounted_parts:
        return False

    package_part_names = {
        name
        for package in (before_package, after_package)
        for name in package.namelist()
        if name and not name.endswith("/")
    }
    changed_defaults = {
        extension
        for extension in before_defaults.keys() | after_defaults.keys()
        if before_defaults.get(extension) != after_defaults.get(extension)
    }
    for extension in changed_defaults:
        expected = _IGNORABLE_UNUSED_DEFAULT_CONTENT_TYPES.get(extension)
        declarations = {
            value
            for value in (
                before_defaults.get(extension),
                after_defaults.get(extension),
            )
            if value is not None
        }
        if declarations != {expected}:
            return False
        if any(name.rpartition(".")[2].lower() == extension for name in package_part_names):
            return False

    changed_overrides = {
        part_name
        for part_name in before_overrides.keys() | after_overrides.keys()
        if before_overrides.get(part_name) != after_overrides.get(part_name)
    }
    for part_name in changed_overrides:
        if part_name not in package_part_names:
            return False
        if part_name == _CUSTOM_PROPERTIES_PART:
            if any(
                _custom_properties_package_state(package)[0] == "material"
                for package in (before_package, after_package)
            ):
                return False
            continue
        if part_name in changed_effective_parts and part_name in accounted_parts:
            # A reachable worksheet/drawing/etc. can be added or removed. Its
            # semantic effect is classified by the dedicated comparators.
            continue
        before_type = before_effective.get(part_name)
        after_type = after_effective.get(part_name)
        if before_type is None or before_type != after_type:
            return False
    return True


def _explained_package_churn(
    before_package: ZipFile,
    after_package: ZipFile,
    *,
    changed_parts: tuple[str, ...],
    accounted_parts: frozenset[str],
) -> frozenset[str]:
    """Identify package/document metadata rewrites with unchanged meaning."""

    changed = set(changed_parts)
    explained: set[str] = set()
    changed_properties = changed & _DOCUMENT_PROPERTY_PARTS
    if changed_properties:
        if _document_properties_signature(before_package) != _document_properties_signature(
            after_package
        ):
            names = ", ".join(sorted(changed_properties))
            raise ValueError(f"material OOXML document properties changed: {names}")
        explained.update(changed_properties)
    if _ROOT_RELATIONSHIPS_PART in changed:
        if _root_relationships_signature(before_package) != _root_relationships_signature(
            after_package
        ):
            raise ValueError("material root OOXML relationships changed")
        explained.add(_ROOT_RELATIONSHIPS_PART)
    if "[Content_Types].xml" in changed:
        if not _content_types_change_explained(
            before_package,
            after_package,
            accounted_parts=accounted_parts,
        ):
            raise ValueError("material OOXML content types changed: [Content_Types].xml")
        explained.add("[Content_Types].xml")
    return frozenset(explained)


def _serialisable_signature(value: Any) -> tuple[Any, ...] | None:
    if value is None:
        return None
    to_tree = getattr(value, "to_tree", None)
    if not callable(to_tree):
        raise TypeError(f"{type(value).__name__} does not expose deterministic XML")
    return (
        f"{type(value).__module__}.{type(value).__qualname__}",
        _xml_signature(to_tree()),
    )


def _table_signature(sheet: Any) -> tuple[tuple[Any, ...], ...]:
    items: list[tuple[Any, ...]] = []
    for name in sorted(sheet.tables):
        table = sheet.tables[name]
        items.append((name, _serialisable_signature(table)))
    return tuple(items)


def _sheet_structure(sheet: Any) -> tuple[Any, ...]:
    return (
        tuple(sorted(str(item) for item in sheet.merged_cells.ranges)),
        _table_signature(sheet),
        sheet.sheet_state,
        _serialisable_signature(sheet.auto_filter),
        _serialisable_signature(sheet.protection),
        _serialisable_signature(sheet.scenarios),
        _serialisable_signature(sheet.row_breaks),
        _serialisable_signature(sheet.col_breaks),
        _serialisable_signature(sheet.sheet_format),
        tuple(sorted((key, _row_dimension(value)) for key, value in sheet.row_dimensions.items())),
        tuple(
            sorted(
                (key, _column_dimension(value)) for key, value in sheet.column_dimensions.items()
            )
        ),
        str(sheet.print_area),
        sheet.print_title_rows,
        sheet.print_title_cols,
    )


def _page_and_view_signature(sheet: Any) -> tuple[Any, ...]:
    return (
        _serialisable_signature(sheet.sheet_properties),
        _serialisable_signature(sheet.views),
        _serialisable_signature(sheet.page_margins),
        _serialisable_signature(sheet.print_options),
        _serialisable_signature(sheet.page_setup),
        _serialisable_signature(sheet.HeaderFooter),
    )


def _conditional_formatting_signature(sheet: Any) -> tuple[Any, ...]:
    items: list[tuple[Any, ...]] = []
    for conditional_formatting in sheet.conditional_formatting:
        rules = tuple(
            (
                _serialisable_signature(rule.__dict__.get("dxf")),
                _serialisable_signature(rule.__dict__.get("extLst")),
            )
            for rule in conditional_formatting.rules
        )
        items.append((_serialisable_signature(conditional_formatting), rules))
    return tuple(items)


def _data_validation_signature(sheet: Any) -> tuple[Any, ...] | None:
    return _serialisable_signature(sheet.data_validations)


def _package_xml_signature(package: ZipFile, part_name: str) -> tuple[Any, ...]:
    normalized = part_name.lstrip("/")
    try:
        payload = package.read(normalized)
    except KeyError as exc:
        raise ValueError(f"missing OOXML part: {normalized}") from exc
    return _xml_signature(fromstring(payload))


def _unsupported_parts_signature(package: ZipFile) -> tuple[Any, ...]:
    signatures: list[tuple[Any, ...]] = []
    for part_name in sorted(package.namelist()):
        if not (
            part_name in _UNSUPPORTED_PART_NAMES
            or part_name.startswith(_UNSUPPORTED_PART_PREFIXES)
            or part_name.endswith(".vml")
        ):
            continue
        signature = hashlib.sha256(package.read(part_name)).hexdigest()
        signatures.append((part_name, signature))
    return tuple(signatures)


def _theme_signature(workbook: Any) -> tuple[Any, ...] | None:
    theme = workbook.loaded_theme
    if theme is None:
        return None
    if not isinstance(theme, bytes):
        raise TypeError(f"unsupported workbook theme payload: {type(theme).__name__}")
    return _xml_signature(fromstring(theme))


def _workbook_view_signature(workbook: Any) -> tuple[Any, ...]:
    return tuple(_serialisable_signature(view) for view in workbook.views)


def _chart_signature(chart: Any) -> tuple[Any, ...]:
    writer = getattr(chart, "_write", None)
    if not callable(writer):
        raise TypeError(f"unsupported chart type: {type(chart).__name__}")
    return (
        f"{type(chart).__module__}.{type(chart).__qualname__}",
        _xml_signature(writer()),
        _serialisable_signature(chart.anchor),
    )


def _charts_signature(sheet: Any) -> tuple[Any, ...]:
    return tuple(_chart_signature(chart) for chart in getattr(sheet, "_charts", ()))


def _image_signature(image: Any) -> tuple[Any, ...]:
    reader = getattr(image, "_data", None)
    if not callable(reader):
        raise TypeError(f"unsupported image type: {type(image).__name__}")
    data = reader()
    if not isinstance(data, bytes):
        raise TypeError(f"unsupported image payload: {type(data).__name__}")
    return (
        getattr(image, "format", None),
        _scalar(getattr(image, "width", None)),
        _scalar(getattr(image, "height", None)),
        hashlib.sha256(data).hexdigest(),
        _serialisable_signature(image.anchor),
    )


def _images_signature(sheet: Any) -> tuple[Any, ...]:
    return tuple(_image_signature(image) for image in getattr(sheet, "_images", ()))


def _drawing_parts_signature(sheet: Any, package: ZipFile) -> tuple[Any, ...]:
    targets = sorted(
        relationship.Target
        for relationship in sheet._rels
        if str(getattr(relationship, "Type", "")).endswith("/drawing")
    )
    signatures: list[tuple[Any, ...]] = []
    for target in targets:
        drawing_signature = _package_xml_signature(package, target)
        relationships_path = get_rels_path(target)
        relationships = get_dependents(package, relationships_path)
        related_parts: list[tuple[Any, ...]] = []
        for relationship in relationships:
            relationship_kind = str(relationship.Type).rsplit("/", 1)[-1]
            if relationship_kind == "chart":
                part_signature: Any = _package_xml_signature(package, relationship.target)
            elif relationship_kind == "image":
                part_signature = hashlib.sha256(package.read(relationship.target)).hexdigest()
            else:
                raise TypeError(
                    f"unsupported drawing relationship: {relationship_kind or 'unknown'}"
                )
            related_parts.append((relationship.Id, relationship_kind, part_signature))
        signatures.append((drawing_signature, tuple(sorted(related_parts))))
    return tuple(signatures)


def _defined_names(workbook: Any) -> tuple[tuple[Any, ...], ...]:
    values = []
    try:
        names = workbook.defined_names.values()
    except AttributeError:
        names = workbook.defined_names.definedName
    for item in names:
        values.append(
            (
                getattr(item, "name", None),
                getattr(item, "attr_text", None),
                getattr(item, "localSheetId", None),
                getattr(item, "hidden", None),
                getattr(item, "function", None),
            )
        )
    return tuple(sorted(values, key=lambda item: tuple(str(value) for value in item)))


def _compact_ranges(
    cells: dict[str, set[tuple[int, int]]],
    *,
    max_ranges: int,
) -> EvidenceScope:
    ranges: list[CellRange] = []
    fallback_sheets: list[str] = []
    for sheet, coordinates in sorted(cells.items()):
        row_runs: list[tuple[int, int, int]] = []
        by_row: dict[int, list[int]] = defaultdict(list)
        for row, column in coordinates:
            by_row[row].append(column)
        for row, columns in sorted(by_row.items()):
            sorted_columns = sorted(set(columns))
            start = previous = sorted_columns[0]
            for column in sorted_columns[1:]:
                if column == previous + 1:
                    previous = column
                    continue
                row_runs.append((row, start, previous))
                start = previous = column
            row_runs.append((row, start, previous))

        merged: list[CellRange] = []
        for row, min_col, max_col in row_runs:
            if (
                merged
                and merged[-1].sheet == sheet
                and merged[-1].min_col == min_col
                and merged[-1].max_col == max_col
                and merged[-1].max_row + 1 == row
            ):
                previous = merged[-1]
                merged[-1] = CellRange(
                    sheet,
                    min_col,
                    previous.min_row,
                    max_col,
                    row,
                )
            else:
                merged.append(CellRange(sheet, min_col, row, max_col, row))
        if len(ranges) + len(merged) <= max_ranges:
            ranges.extend(merged)
            continue

        min_row = min(row for row, _ in coordinates)
        max_row = max(row for row, _ in coordinates)
        min_col = min(column for _, column in coordinates)
        max_col = max(column for _, column in coordinates)
        bounding = CellRange(sheet, min_col, min_row, max_col, max_row)
        if bounding.cell_count <= _MAX_BOUNDING_RANGE_CELLS:
            ranges.append(bounding)
        else:
            fallback_sheets.append(sheet)
    return EvidenceScope(tuple(ranges), sheets=tuple(fallback_sheets))


@dataclass(frozen=True)
class WorkbookEffectDiff:
    semantic_changed: bool
    complete: bool
    effects: frozenset[EffectKind]
    scope: EvidenceScope
    formula_scope: EvidenceScope
    changed_cell_count: int
    scanned_cell_count: int
    reasons: tuple[str, ...] = ()

    @classmethod
    def unknown(cls, reason: str) -> WorkbookEffectDiff:
        return cls(
            semantic_changed=True,
            complete=False,
            effects=frozenset({EffectKind.UNKNOWN}),
            scope=EvidenceScope.workbook(),
            formula_scope=EvidenceScope(),
            changed_cell_count=0,
            scanned_cell_count=0,
            reasons=(reason,),
        )

    def to_dict(self) -> dict[str, Any]:
        return {
            "schema_version": "workbook-effect-diff-v1",
            "semantic_changed": self.semantic_changed,
            "complete": self.complete,
            "effects": sorted(item.value for item in self.effects),
            "scope": self.scope.to_dict(),
            "formula_scope": self.formula_scope.to_dict(),
            "changed_cell_count": self.changed_cell_count,
            "scanned_cell_count": self.scanned_cell_count,
            "reasons": list(self.reasons),
        }


def diff_workbooks(
    before: str | Path,
    after: str | Path,
    *,
    max_scanned_cells: int = _DEFAULT_MAX_SCANNED_CELLS,
    max_changed_cells: int = _DEFAULT_MAX_CHANGED_CELLS,
    max_ranges: int = _DEFAULT_MAX_RANGES,
) -> WorkbookEffectDiff:
    """Compare workbook semantics and return a conservative typed footprint."""

    before_path = Path(before)
    after_path = Path(after)
    before_book = after_book = None
    before_package = after_package = None
    try:
        before_package = ZipFile(before_path)
        after_package = ZipFile(after_path)
        try:
            _custom_properties_package_state(before_package)
            _custom_properties_package_state(after_package)
            workbook_residual_changed = _unclassified_workbook_signature(
                before_package
            ) != _unclassified_workbook_signature(after_package)
        except ValueError as exc:
            return WorkbookEffectDiff.unknown(str(exc))
        if workbook_residual_changed:
            return WorkbookEffectDiff.unknown(
                "unclassified workbook XML changed: xl/workbook.xml"
            )
        before_opaque_extensions = _opaque_extension_signature(before_package)
        after_opaque_extensions = _opaque_extension_signature(after_package)
        if before_opaque_extensions != after_opaque_extensions:
            changed_opaque_parts = sorted(
                {
                    str(record[0])
                    for record in set(before_opaque_extensions)
                    ^ set(after_opaque_extensions)
                }
            )
            preview = ", ".join(changed_opaque_parts[:8])
            if len(changed_opaque_parts) > 8:
                preview = f"{preview}, ..."
            return WorkbookEffectDiff.unknown(
                f"unsupported OOXML extension or relationship content changed: {preview}"
            )
        before_unsupported_parts = _unsupported_parts_signature(before_package)
        after_unsupported_parts = _unsupported_parts_signature(after_package)
        if before_unsupported_parts != after_unsupported_parts:
            changed_unsupported_parts = sorted(
                {
                    str(record[0])
                    for record in set(before_unsupported_parts)
                    ^ set(after_unsupported_parts)
                }
            )
            preview = ", ".join(changed_unsupported_parts[:8])
            if len(changed_unsupported_parts) > 8:
                preview = f"{preview}, ..."
            return WorkbookEffectDiff.unknown(
                f"unsupported OOXML part set or content changed: {preview}"
            )
        changed_package_parts = _changed_package_parts(before_package, after_package)
        accounted_parts = _accounted_parts(before_package) | _accounted_parts(after_package)
        try:
            explained_package_parts = _explained_package_churn(
                before_package,
                after_package,
                changed_parts=changed_package_parts,
                accounted_parts=accounted_parts,
            )
        except ValueError as exc:
            return WorkbookEffectDiff.unknown(str(exc))
        unaccounted_parts = tuple(
            part_name
            for part_name in changed_package_parts
            if part_name not in accounted_parts and part_name not in explained_package_parts
        )
        if unaccounted_parts:
            preview = ", ".join(unaccounted_parts[:8])
            if len(unaccounted_parts) > 8:
                preview = f"{preview}, ..."
            return WorkbookEffectDiff.unknown(
                f"unaccounted OOXML part set or content changed: {preview}"
            )
        with warnings.catch_warnings(record=True) as load_warnings:
            warnings.simplefilter("always")
            before_book = load_workbook(
                before_path,
                data_only=False,
                keep_vba=before_path.suffix.lower() == ".xlsm",
                keep_links=True,
            )
            after_book = load_workbook(
                after_path,
                data_only=False,
                keep_vba=after_path.suffix.lower() == ".xlsm",
                keep_links=True,
            )
        if load_warnings:
            warning_types = sorted(
                {f"{item.category.__name__}: {item.message}" for item in load_warnings}
            )
            detail = "; ".join(warning_types[:3])
            return WorkbookEffectDiff.unknown(
                f"openpyxl reported potentially lossy workbook content: {detail}"
            )
        effects: set[EffectKind] = set()
        changed_cells: dict[str, set[tuple[int, int]]] = defaultdict(set)
        formula_cells: dict[str, set[tuple[int, int]]] = defaultdict(set)
        structure_sheets: set[str] = set()
        semantic_explained_parts: set[str] = set()
        explained_relationship_kinds: set[str] = set()
        explained_sheet_modes: set[tuple[str, bool]] = set()
        workbook_scope_changed = False
        scanned = 0

        def explain_workbook_part() -> None:
            semantic_explained_parts.add("xl/workbook.xml")

        def explain_workbook_relationships() -> None:
            semantic_explained_parts.add("xl/_rels/workbook.xml.rels")

        def explain_relationship_kind(kind: str) -> None:
            if kind in explained_relationship_kinds:
                return
            explained_relationship_kinds.add(kind)
            for package in (before_package, after_package):
                semantic_explained_parts.update(
                    _workbook_relationship_targets(package, kind)
                )

        def explain_sheet(sheet_name: str, *, include_related: bool = False) -> None:
            key = (sheet_name, include_related)
            if key in explained_sheet_modes:
                return
            explained_sheet_modes.add(key)
            for package in (before_package, after_package):
                semantic_explained_parts.update(
                    _worksheet_owned_parts(
                        package,
                        sheet_name,
                        include_related=include_related,
                    )
                )

        def explain_shared_string_representation() -> None:
            targets = set()
            for package in (before_package, after_package):
                targets.update(
                    _workbook_relationship_targets(package, "sharedStrings")
                )
            if not targets.intersection(changed_package_parts):
                return
            if all(
                _shared_strings_are_fully_referenced(package)
                for package in (before_package, after_package)
            ):
                explain_relationship_kind("sharedStrings")
                explain_workbook_relationships()

        if before_book.sheetnames != after_book.sheetnames:
            effects.update({EffectKind.STRUCTURE, EffectKind.VISUAL})
            explain_workbook_part()
            explain_workbook_relationships()
            before_names = set(before_book.sheetnames)
            after_names = set(after_book.sheetnames)
            removed_formula = False
            for book, names, removed in (
                (before_book, before_names - after_names, True),
                (after_book, after_names - before_names, False),
            ):
                for sheet_name in sorted(names):
                    explain_sheet(sheet_name, include_related=True)
                    sheet = book[sheet_name]
                    candidate_cells = sheet.max_row * sheet.max_column
                    if scanned + candidate_cells > max_scanned_cells:
                        return WorkbookEffectDiff.unknown(
                            f"semantic diff exceeded {max_scanned_cells} scanned cells"
                        )
                    scanned += candidate_cells
                    for row in range(1, sheet.max_row + 1):
                        for column in range(1, sheet.max_column + 1):
                            cell = sheet.cell(row, column)
                            value = cell.value
                            formula = cell.data_type == "f" or (
                                isinstance(value, str) and value.startswith("=")
                            )
                            has_content = any(
                                item is not None
                                for item in (value, cell.hyperlink, cell.comment)
                            )
                            if formula:
                                effects.add(EffectKind.FORMULA)
                                ranges = _formula_ranges(cell, sheet_name)
                                if any(
                                    item.cell_count > max_changed_cells for item in ranges
                                ):
                                    return WorkbookEffectDiff.unknown(
                                        "formula footprint exceeded changed-cell limit"
                                    )
                                for formula_range in ranges:
                                    for formula_row in range(
                                        formula_range.min_row,
                                        formula_range.max_row + 1,
                                    ):
                                        for formula_column in range(
                                            formula_range.min_col,
                                            formula_range.max_col + 1,
                                        ):
                                            coordinate = (formula_row, formula_column)
                                            formula_cells[sheet_name].add(coordinate)
                                            changed_cells[sheet_name].add(coordinate)
                                removed_formula = removed_formula or removed
                            elif has_content:
                                effects.add(EffectKind.VALUE)
                            if cell.has_style:
                                effects.update({EffectKind.STYLE, EffectKind.VISUAL})
                            if (has_content or cell.has_style) and not formula:
                                changed_cells[sheet_name].add((row, column))
                            if (
                                sum(len(items) for items in changed_cells.values())
                                > max_changed_cells
                            ):
                                return WorkbookEffectDiff.unknown(
                                    f"semantic diff exceeded {max_changed_cells} changed cells"
                                )
            if EffectKind.STYLE in effects:
                explain_relationship_kind("styles")
            if EffectKind.VALUE in effects:
                explain_relationship_kind("sharedStrings")
            explain_shared_string_representation()
            unexplained_package_parts = tuple(
                part_name
                for part_name in changed_package_parts
                if part_name
                not in explained_package_parts | frozenset(semantic_explained_parts)
            )
            if unexplained_package_parts:
                preview = ", ".join(unexplained_package_parts[:8])
                if len(unexplained_package_parts) > 8:
                    preview = f"{preview}, ..."
                return WorkbookEffectDiff.unknown(
                    "changed OOXML parts were not explained by the semantic "
                    f"comparators: {preview}"
                )
            compact_formula_scope = _compact_ranges(formula_cells, max_ranges=max_ranges)
            return WorkbookEffectDiff(
                semantic_changed=True,
                complete=True,
                effects=frozenset(effects),
                scope=EvidenceScope.workbook(),
                formula_scope=(
                    EvidenceScope.workbook() if removed_formula else compact_formula_scope
                ),
                changed_cell_count=sum(len(items) for items in changed_cells.values()),
                scanned_cell_count=scanned,
                reasons=("worksheet set or order changed",),
            )
        if _defined_names(before_book) != _defined_names(after_book):
            effects.add(EffectKind.STRUCTURE)
            structure_sheets.update(after_book.sheetnames)
            explain_workbook_part()
        if _theme_signature(before_book) != _theme_signature(after_book):
            effects.update({EffectKind.STYLE, EffectKind.VISUAL})
            workbook_scope_changed = True
            explain_relationship_kind("theme")
            explain_workbook_relationships()
        if _workbook_view_signature(before_book) != _workbook_view_signature(after_book):
            effects.add(EffectKind.VISUAL)
            workbook_scope_changed = True
            explain_workbook_part()
        if _serialisable_signature(before_book.security) != _serialisable_signature(
            after_book.security
        ):
            effects.add(EffectKind.STRUCTURE)
            workbook_scope_changed = True
            explain_workbook_part()

        for sheet_name in after_book.sheetnames:
            before_sheet = before_book[sheet_name]
            after_sheet = after_book[sheet_name]
            if _sheet_structure(before_sheet) != _sheet_structure(after_sheet):
                effects.update({EffectKind.STRUCTURE, EffectKind.VISUAL})
                structure_sheets.add(sheet_name)
                explain_sheet(sheet_name, include_related=True)
                explain_workbook_part()
            if _page_and_view_signature(before_sheet) != _page_and_view_signature(after_sheet):
                effects.update({EffectKind.STRUCTURE, EffectKind.VISUAL})
                structure_sheets.add(sheet_name)
                explain_sheet(sheet_name)
            if _conditional_formatting_signature(before_sheet) != _conditional_formatting_signature(
                after_sheet
            ):
                effects.update({EffectKind.STYLE, EffectKind.VISUAL})
                structure_sheets.add(sheet_name)
                explain_sheet(sheet_name)
                explain_relationship_kind("styles")
            if _data_validation_signature(before_sheet) != _data_validation_signature(after_sheet):
                effects.add(EffectKind.STRUCTURE)
                structure_sheets.add(sheet_name)
                explain_sheet(sheet_name)
            if _charts_signature(before_sheet) != _charts_signature(after_sheet):
                effects.update({EffectKind.STRUCTURE, EffectKind.VISUAL})
                structure_sheets.add(sheet_name)
                explain_sheet(sheet_name, include_related=True)
            if _images_signature(before_sheet) != _images_signature(after_sheet):
                effects.update({EffectKind.STRUCTURE, EffectKind.VISUAL})
                structure_sheets.add(sheet_name)
                explain_sheet(sheet_name, include_related=True)
            if _drawing_parts_signature(before_sheet, before_package) != _drawing_parts_signature(
                after_sheet, after_package
            ):
                effects.update({EffectKind.STRUCTURE, EffectKind.VISUAL})
                structure_sheets.add(sheet_name)
                explain_sheet(sheet_name, include_related=True)

            max_row = max(before_sheet.max_row, after_sheet.max_row)
            max_column = max(before_sheet.max_column, after_sheet.max_column)
            candidate_cells = max_row * max_column
            if scanned + candidate_cells > max_scanned_cells:
                return WorkbookEffectDiff.unknown(
                    f"semantic diff exceeded {max_scanned_cells} scanned cells"
                )
            scanned += candidate_cells
            for row in range(1, max_row + 1):
                for column in range(1, max_column + 1):
                    before_cell = before_sheet.cell(row, column)
                    after_cell = after_sheet.cell(row, column)
                    before_content = _cell_content(before_cell)
                    after_content = _cell_content(after_cell)
                    if before_content != after_content:
                        before_value = before_cell.value
                        after_value = after_cell.value
                        formula_changed = bool(
                            before_cell.data_type == "f"
                            or after_cell.data_type == "f"
                            or any(
                                isinstance(value, str) and value.startswith("=")
                                for value in (before_value, after_value)
                            )
                        )
                        effects.add(EffectKind.FORMULA if formula_changed else EffectKind.VALUE)
                        if formula_changed:
                            ranges = {
                                *_formula_ranges(before_cell, sheet_name),
                                *_formula_ranges(after_cell, sheet_name),
                            }
                            if any(
                                item.cell_count > max_changed_cells for item in ranges
                            ):
                                return WorkbookEffectDiff.unknown(
                                    "formula footprint exceeded changed-cell limit"
                                )
                            for formula_range in ranges:
                                for formula_row in range(
                                    formula_range.min_row,
                                    formula_range.max_row + 1,
                                ):
                                    for formula_column in range(
                                        formula_range.min_col,
                                        formula_range.max_col + 1,
                                    ):
                                        coordinate = (formula_row, formula_column)
                                        changed_cells[sheet_name].add(coordinate)
                                        formula_cells[sheet_name].add(coordinate)
                        else:
                            changed_cells[sheet_name].add((row, column))
                        explain_sheet(sheet_name)
                        explain_relationship_kind("sharedStrings")
                    if _style(before_cell) != _style(after_cell):
                        effects.update({EffectKind.STYLE, EffectKind.VISUAL})
                        changed_cells[sheet_name].add((row, column))
                        explain_sheet(sheet_name)
                        explain_relationship_kind("styles")
                    if sum(len(items) for items in changed_cells.values()) > max_changed_cells:
                        return WorkbookEffectDiff.unknown(
                            f"semantic diff exceeded {max_changed_cells} changed cells"
                        )

        cell_scope = _compact_ranges(changed_cells, max_ranges=max_ranges)
        structure_scope = EvidenceScope(sheets=tuple(sorted(structure_sheets)))
        scope = (
            EvidenceScope.workbook()
            if workbook_scope_changed
            else cell_scope.merged(structure_scope)
        )
        formula_scope = _compact_ranges(formula_cells, max_ranges=max_ranges)
        explain_shared_string_representation()
        unexplained_package_parts = tuple(
            part_name
            for part_name in changed_package_parts
            if part_name
            not in explained_package_parts | frozenset(semantic_explained_parts)
        )
        if unexplained_package_parts:
            preview = ", ".join(unexplained_package_parts[:8])
            if len(unexplained_package_parts) > 8:
                preview = f"{preview}, ..."
            return WorkbookEffectDiff.unknown(
                f"changed OOXML parts were not explained by the semantic comparators: {preview}"
            )
        return WorkbookEffectDiff(
            semantic_changed=bool(effects),
            complete=True,
            effects=frozenset(effects),
            scope=scope,
            formula_scope=formula_scope,
            changed_cell_count=sum(len(items) for items in changed_cells.values()),
            scanned_cell_count=scanned,
        )
    except Exception as exc:
        return WorkbookEffectDiff.unknown(f"{type(exc).__name__}: workbook semantic diff failed")
    finally:
        if before_book is not None:
            before_book.close()
        if after_book is not None:
            after_book.close()
        if before_package is not None:
            before_package.close()
        if after_package is not None:
            after_package.close()


__all__ = ["WorkbookEffectDiff", "diff_workbooks"]
