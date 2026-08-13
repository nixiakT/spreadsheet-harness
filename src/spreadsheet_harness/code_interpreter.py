"""Bounded local Python execution with an optional strict filesystem boundary.

Ordinary local runs retain the trusted-code behavior. Comparison runs require a
Bubblewrap sandbox that exposes only the current run workspace plus allowlisted
Python/runtime files. Required isolation is fail-closed and never falls back to
an unsandboxed process.
"""

from __future__ import annotations

import hashlib
import os
import platform
import re
import resource
import shutil
import subprocess
import sys
import tempfile
import threading
import uuid
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.formula import Tokenizer
from openpyxl.utils import FORMULAE, column_index_from_string, coordinate_to_tuple

from .errors import CodeIsolationError, ToolInputError, redact_sensitive_text

STRICT_ISOLATION_POLICY = "bubblewrap-strict-workspace-v1"
_PROBE_SENTINEL = "SHEET_STRICT_ISOLATION_OK"
_MAX_SANDBOX_PROCESSES = 64
_PROBE_LOCK = threading.Lock()
_PROBE_SUCCESSES: set[tuple[str, ...]] = set()
_RUNTIME_HELPER_NAME = "sheet_harness.py"
_MUTATION_MARKER_ENV = "SHEET_MUTATION_MARKER"
_COMPRESSED_PLACEHOLDER_MARKERS = ("[compressed]", "[truncated]")
_INVALID_ABSOLUTE_ROW_REFERENCE = re.compile(r"^\$\d+$")
_FORMULA_TEXT_PREFIX = re.compile(
    r"^(?P<function>(?:_xlfn\.)?[A-Za-z][A-Za-z0-9_.]*)\(",
    re.IGNORECASE,
)
_A1_ENDPOINT = r"\$?[A-Za-z]{1,3}\$?[1-9]\d*"
_DEFINITE_A1_REFERENCE = re.compile(
    rf"^(?:(?P<sheet>'(?:[^']|'')+'|[A-Za-z_][A-Za-z0-9_.]*)!)?"
    rf"(?P<start>{_A1_ENDPOINT})(?::(?P<end>{_A1_ENDPOINT}))?$"
)
_FORMULA_SIGNAL_OPERATORS = frozenset(
    {"+", "-", "*", "/", "^", "%", "=", "<>", "<", ">", "<=", ">="}
)
_MODERN_FORMULA_FUNCTIONS = {
    "AGGREGATE",
    "CONCAT",
    "FILTER",
    "FORMULATEXT",
    "IFS",
    "ISFORMULA",
    "LAMBDA",
    "LET",
    "MAXIFS",
    "MINIFS",
    "SEQUENCE",
    "SORT",
    "SORTBY",
    "SWITCH",
    "TEXTJOIN",
    "UNIQUE",
    "XLOOKUP",
    "XMATCH",
}
_KNOWN_FORMULA_FUNCTIONS = frozenset(FORMULAE) | _MODERN_FORMULA_FUNCTIONS


@dataclass(frozen=True)
class _FormulaTextCandidate:
    sheet: str
    cell: str
    value: str
    shape: tuple[str, ...]
    reference_operand_count: int
    has_formula_operator: bool
    has_absolute_or_cross_sheet_reference: bool


_OPENPYXL_COMPAT_SHIM = r"""
def _sheet_harness_install_openpyxl_compat():
    try:
        from openpyxl.workbook.defined_name import DefinedNameDict
        if not hasattr(DefinedNameDict, "definedName"):
            DefinedNameDict.definedName = property(lambda self: list(self.values()))

        from openpyxl.worksheet.table import TableList
        if getattr(TableList, "_sheet_harness_iterates_values", False) is not True:
            TableList.__iter__ = lambda self: iter(dict.values(self))
            TableList.items = lambda self: dict.items(self)
            TableList._sheet_harness_iterates_values = True

        from openpyxl.workbook.workbook import Workbook
        if getattr(Workbook, "_sheet_harness_duplicate_name_compat", False) is not True:
            def _duplicate_name(self, name):
                candidate = str(name).lower()
                for sheet in self.worksheets:
                    for table in getattr(sheet, "tables", []):
                        table_name = getattr(table, "name", table)
                        if candidate == str(table_name).lower():
                            return True
                return candidate in getattr(self, "defined_names", {})
            Workbook._duplicate_name = _duplicate_name
            Workbook._sheet_harness_duplicate_name_compat = True

        from openpyxl.worksheet.worksheet import Worksheet
        if not hasattr(Worksheet, "_tableparts"):
            Worksheet._tableparts = property(
                lambda self: list(getattr(self, "tables", {}).values())
            )
        if not hasattr(Worksheet, "merged_ranges"):
            Worksheet.merged_ranges = property(
                lambda self: self.merged_cells.ranges
            )

        from openpyxl.cell.cell import Cell
        if not hasattr(Cell, "dtype"):
            Cell.dtype = property(
                lambda self: "formula" if getattr(self, "data_type", None) == "f" else self.data_type
            )
        if not hasattr(Cell, "formula"):
            Cell.formula = property(
                lambda self: self.value
                if getattr(self, "data_type", None) == "f"
                else None
            )
    except Exception as exc:
        import sys
        print(
            "[sheet_harness] openpyxl compatibility shim failed: "
            f"{type(exc).__name__}: {exc}",
            file=sys.stderr,
        )


_sheet_harness_install_openpyxl_compat()
"""


_RUNTIME_HELPER_SOURCE = (
    _OPENPYXL_COMPAT_SHIM
    + r'''
"""Small runtime helpers available inside spreadsheet-harness code_interpreter."""

import hashlib
import os
import re
from contextlib import contextmanager
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook as _openpyxl_load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter, range_boundaries

_FORMULA_RANGE_RE = re.compile(
    r"(?P<sheet>(?:'[^']+'|[A-Za-z_][A-Za-z0-9_ .]*)!)?"
    r"(?P<start>\$?[A-Za-z]{1,3}\$?\d+):(?P<end>\$?[A-Za-z]{1,3}\$?\d+)"
)
_CELL_REF_RE = re.compile(
    r"(?P<col_abs>\$?)(?P<col>[A-Za-z]{1,3})(?P<row_abs>\$?)(?P<row>\d+)\Z"
)


def workbook_path() -> Path:
    return Path(os.environ["SHEET_WORKBOOK"])


def workbook_sha256(path: str | Path | None = None) -> str:
    target = Path(path) if path is not None else workbook_path()
    digest = hashlib.sha256()
    with target.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _record_managed_mutation_attempt() -> None:
    marker = os.environ.get("SHEET_MUTATION_MARKER")
    if marker:
        Path(marker).touch(exist_ok=True)


def load_workbook(path: str | Path | None = None, *, data_only: bool = False, **kwargs: Any):
    """Load SHEET_WORKBOOK when called without a path."""

    target = Path(path) if path is not None else workbook_path()
    kwargs.setdefault("keep_vba", target.suffix.lower() == ".xlsm")
    kwargs.setdefault("keep_links", True)
    return _openpyxl_load_workbook(target, data_only=data_only, **kwargs)


def table_map(worksheet: Any) -> dict[str, Any]:
    tables = getattr(worksheet, "tables", {})
    if isinstance(tables, dict):
        return {
            str(name): dict.__getitem__(tables, name)
            for name in dict.keys(tables)
        }
    result: dict[str, Any] = {}
    for table in tables or []:
        result[str(getattr(table, "name", len(result) + 1))] = table
    return result


def _is_workbook_like(value: Any) -> bool:
    return hasattr(value, "worksheets") and hasattr(value, "sheetnames")


def _load_if_path(value: Any, *, data_only: bool = False) -> tuple[Any, bool]:
    if value is None:
        return load_workbook(data_only=data_only), True
    if isinstance(value, str | Path):
        return load_workbook(value, data_only=data_only), True
    return value, False


def table_refs(worksheet: Any) -> dict[str, str]:
    return {
        name: str(getattr(table, "ref", table))
        for name, table in table_map(worksheet).items()
    }


def defined_name_refs(workbook: Any | None = None) -> dict[str, str]:
    workbook, should_close = _load_if_path(workbook)
    names = getattr(workbook, "defined_names", {})
    try:
        if isinstance(names, dict):
            return {
                str(name): str(getattr(value, "attr_text", value))
                for name, value in dict.items(names)
            }
        return {
            str(getattr(value, "name", index)): str(getattr(value, "attr_text", value))
            for index, value in enumerate(getattr(names, "definedName", []) or [])
        }
    finally:
        if should_close:
            workbook.close()


def workbook_overview(workbook: Any | None = None) -> list[dict[str, Any]]:
    """Return one structure dictionary per worksheet, in workbook order."""

    workbook, should_close = _load_if_path(workbook)
    overview: list[dict[str, Any]] = []
    try:
        for index, worksheet in enumerate(workbook.worksheets):
            overview.append(
                {
                    "index": index,
                    "name": worksheet.title,
                    "dimension": worksheet.calculate_dimension(),
                    "max_row": worksheet.max_row,
                    "max_column": worksheet.max_column,
                    "tables": table_refs(worksheet),
                    "merged_ranges": [str(item) for item in worksheet.merged_cells.ranges],
                }
            )
        return overview
    finally:
        if should_close:
            workbook.close()


def copy_cell_format(source: Any, target: Any) -> None:
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    if source.alignment:
        target.alignment = copy(source.alignment)
    if source.protection:
        target.protection = copy(source.protection)


def _cell_ref_parts(ref: str) -> dict[str, Any] | None:
    match = _CELL_REF_RE.fullmatch(ref)
    if match is None:
        return None
    return {
        "column_absolute": bool(match.group("col_abs")),
        "row_absolute": bool(match.group("row_abs")),
    }


def _formula_sample_coordinates(
    source_cell: str,
    bounds: tuple[int, int, int, int],
) -> list[str]:
    min_col, min_row, max_col, max_row = bounds
    candidates = [
        source_cell.replace("$", ""),
        f"{get_column_letter(min_col)}{min_row}",
        f"{get_column_letter(min(min_col + 1, max_col))}{min_row}",
        f"{get_column_letter(min_col)}{min(min_row + 1, max_row)}",
        f"{get_column_letter(max_col)}{max_row}",
    ]
    return list(dict.fromkeys(candidates))


def _normalize_fill_target_range(
    source_cell: str,
    target_range: str,
) -> tuple[str, bool]:
    source = source_cell.replace("$", "")
    target = target_range.replace("$", "")
    if ":" in target:
        return target_range, False
    try:
        range_boundaries(target)
        range_boundaries(source)
    except (TypeError, ValueError):
        return target_range, False
    if target.upper() == source.upper():
        return target_range, False
    return f"{source}:{target}", True


def _fill_formula_warnings(
    source_formula: str,
    source_cell: str,
    bounds: tuple[int, int, int, int],
    samples: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    min_col, min_row, max_col, max_row = bounds
    fills_horizontally = max_col > min_col
    fills_vertically = max_row > min_row
    if not fills_horizontally and not fills_vertically:
        return []

    sample_cells = [
        str(sample["cell"])
        for sample in samples
        if sample.get("cell") != source_cell.replace("$", "")
    ]
    warnings: list[dict[str, Any]] = []
    for match in _FORMULA_RANGE_RE.finditer(source_formula):
        start = _cell_ref_parts(match.group("start"))
        end = _cell_ref_parts(match.group("end"))
        if start is None or end is None:
            continue
        issues: list[str] = []
        if fills_horizontally and not (
            start["column_absolute"] and end["column_absolute"]
        ):
            issues.append("column endpoints are not both absolute")
        if fills_vertically and start["row_absolute"] != end["row_absolute"]:
            issues.append("mixed row anchors")
        if not issues:
            continue

        translated_examples: list[dict[str, str]] = []
        for destination in sample_cells:
            translated = Translator(
                "=" + match.group(0),
                origin=source_cell,
            ).translate_formula(destination)[1:]
            if translated != match.group(0):
                translated_examples.append(
                    {"cell": destination, "translated_range": translated}
                )
            if len(translated_examples) >= 3:
                break
        if not translated_examples:
            continue
        warnings.append(
            {
                "type": "possible_expanding_or_drifting_range",
                "source_range": match.group(0),
                "issues": issues,
                "examples": translated_examples,
                "message": (
                    "This range changes during fill_formula. If the range should stay "
                    "fixed across the fill direction, lock both endpoints, e.g. use "
                    "$E6:$G6 instead of E6:G6 or $E6:G6, then refill and verify cached "
                    "values."
                ),
            }
        )
    return warnings


def fill_formula(
    worksheet: Any,
    source_cell: str,
    target_range: str,
    *,
    copy_format: bool = False,
) -> dict[str, Any]:
    formula = worksheet[source_cell].value
    if not isinstance(formula, str) or not formula.startswith("="):
        if (
            isinstance(formula, str)
            and formula == formula.strip()
            and _FORMULA_RANGE_RE.search(formula) is not None
            and re.match(r"[A-Za-z][A-Za-z0-9_.]*\(", formula) is not None
        ):
            raise ValueError(
                f"{source_cell} contains formula-like text without a leading '='; assign an "
                "Excel formula string beginning with '=' before calling fill_formula"
            )
        raise ValueError(f"{source_cell} does not contain a formula")
    normalized_target_range, expanded_from_endpoint = _normalize_fill_target_range(
        source_cell, target_range
    )
    bounds = range_boundaries(normalized_target_range.replace("$", ""))
    min_col, min_row, max_col, max_row = bounds
    count = 0
    samples: list[dict[str, Any]] = []
    sample_coordinates = set(_formula_sample_coordinates(source_cell, bounds))
    for row in range(min_row, max_row + 1):
        for column in range(min_col, max_col + 1):
            destination = worksheet.cell(row=row, column=column)
            destination.value = Translator(
                formula,
                origin=source_cell.replace("$", ""),
            ).translate_formula(destination.coordinate)
            if copy_format:
                copy_cell_format(worksheet[source_cell], destination)
            if destination.coordinate in sample_coordinates:
                samples.append(
                    {"cell": destination.coordinate, "formula": destination.value}
                )
            count += 1
    return {
        "ok": True,
        "worksheet": getattr(worksheet, "title", None),
        "range": normalized_target_range,
        "requested_range": target_range,
        "target_range_expanded_from_endpoint": expanded_from_endpoint,
        "cells_filled": count,
        "source_formula": formula,
        "sample_formulas": samples,
        "warnings": _fill_formula_warnings(
            formula,
            source_cell.replace("$", ""),
            bounds,
            samples,
        ),
    }


def save_workbook(workbook: Any, path: str | Path | None = None) -> Path:
    """Save to SHEET_WORKBOOK when called without a path."""

    target = Path(path) if path is not None else workbook_path()
    if target.resolve() == workbook_path().resolve():
        _record_managed_mutation_attempt()
    calculation = getattr(workbook, "calculation", None)
    if calculation is not None:
        calculation.fullCalcOnLoad = True
        calculation.forceFullCalc = True
        calculation.calcMode = "auto"
    workbook.save(target)
    validator = _openpyxl_load_workbook(
        target,
        read_only=True,
        data_only=False,
        keep_vba=target.suffix.lower() == ".xlsm",
        keep_links=True,
    )
    validator.close()
    return target


@contextmanager
def editable_workbook(path: str | Path | None = None):
    workbook = load_workbook(path, data_only=False)
    try:
        yield workbook
        save_workbook(workbook, path)
    finally:
        workbook.close()


def range_values(worksheet: Any, range_ref: str) -> list[list[Any]]:
    min_col, min_row, max_col, max_row = range_boundaries(range_ref.replace("$", ""))
    return [
        [
            worksheet.cell(row=row, column=column).value
            for column in range(min_col, max_col + 1)
        ]
        for row in range(min_row, max_row + 1)
    ]


def print_workbook_overview(workbook: Any) -> None:
    for item in workbook_overview(workbook):
        print(item)
    refs = defined_name_refs(workbook)
    if refs:
        print({"defined_names": refs})
'''
)


def _limits(*, include_process_limit: bool = True) -> None:
    resource.setrlimit(resource.RLIMIT_CPU, (45, 45))
    resource.setrlimit(resource.RLIMIT_FSIZE, (100 * 1024 * 1024, 100 * 1024 * 1024))
    if include_process_limit and hasattr(resource, "RLIMIT_NPROC"):
        resource.setrlimit(
            resource.RLIMIT_NPROC,
            (_MAX_SANDBOX_PROCESSES, _MAX_SANDBOX_PROCESSES),
        )
    if platform.system() == "Linux" and hasattr(resource, "RLIMIT_AS"):
        resource.setrlimit(resource.RLIMIT_AS, (4 * 1024**3, 4 * 1024**3))


def _outer_sandbox_limits() -> None:
    """Apply limits that cannot prevent Bubblewrap from creating namespaces.

    RLIMIT_NPROC is charged against every process/thread owned by the host UID,
    not just this child. Applying a fixed value before Bubblewrap starts makes
    namespace creation fail on shared servers whose UID already owns more than
    that many threads. The strict launcher lowers the hard process limit after
    Bubblewrap has established the sandbox and before any model code runs.
    """

    _limits(include_process_limit=False)


def _require_bubblewrap() -> Path:
    if platform.system() != "Linux":
        raise CodeIsolationError(
            "Strict comparison code isolation requires Linux and Bubblewrap"
        )
    discovered = shutil.which("bwrap")
    if not discovered:
        raise CodeIsolationError(
            "Strict comparison code isolation requires the bwrap executable"
        )
    return Path(discovered).absolute()


def _runtime_roots(workspace: Path) -> list[Path]:
    """Return the minimum runtime directories needed by the active Python."""

    candidates = [Path("/usr"), Path("/bin"), Path("/lib"), Path("/lib64")]
    candidates.extend([Path(sys.prefix), Path(sys.base_prefix)])
    roots: list[Path] = []
    resolved_workspace = workspace.resolve()
    for candidate in candidates:
        absolute = candidate.absolute()
        if not absolute.exists() or absolute in roots:
            continue
        resolved = absolute.resolve()
        if (
            resolved == Path("/")
            or resolved_workspace == resolved
            or resolved in resolved_workspace.parents
        ):
            raise CodeIsolationError(
                f"Refusing unsafe runtime mount that contains the task workspace: {absolute}"
            )
        # /usr already includes common /usr/local base prefixes. Keep /lib and
        # /lib64 mount points themselves because dynamic loaders use those paths.
        if absolute not in {Path("/bin"), Path("/lib"), Path("/lib64")} and any(
            resolved == root.resolve() or root.resolve() in resolved.parents
            for root in roots
        ):
            continue
        roots.append(absolute)
    executable = Path(sys.executable).absolute()
    if not any(executable == root or root in executable.parents for root in roots):
        raise CodeIsolationError(
            f"Active Python executable is outside the allowlisted runtime roots: {executable}"
        )
    return roots


def _parent_directories(paths: list[Path]) -> list[Path]:
    parents: set[Path] = set()
    for path in paths:
        current = path.absolute().parent
        while current != Path("/"):
            parents.add(current)
            current = current.parent
    return sorted(parents, key=lambda item: (len(item.parts), str(item)))


def _strict_command(
    workspace: Path,
    argv: list[str],
    *,
    bubblewrap: Path | None = None,
) -> list[str]:
    """Build an empty-root Bubblewrap command with explicit runtime mounts."""

    resolved_workspace = workspace.resolve()
    bwrap = bubblewrap or _require_bubblewrap()
    runtime_roots = _runtime_roots(resolved_workspace)
    runtime_files = [
        path
        for path in (
            Path("/etc/ld.so.cache"),
            Path("/etc/localtime"),
        )
        if path.is_file()
    ]
    runtime_parent_paths = _parent_directories([*runtime_roots, *runtime_files])
    workspace_parent_paths = _parent_directories([resolved_workspace])
    command = [
        str(bwrap),
        "--die-with-parent",
        "--new-session",
        "--unshare-net",
        "--unshare-pid",
        "--unshare-ipc",
        "--unshare-uts",
        "--cap-drop",
        "ALL",
    ]
    for parent in runtime_parent_paths:
        command.extend(["--dir", str(parent)])
    for root in runtime_roots:
        command.extend(["--ro-bind", str(root), str(root)])
    for runtime_file in runtime_files:
        command.extend(["--ro-bind", str(runtime_file), str(runtime_file)])
    command.extend(
        [
            "--tmpfs",
            "/tmp",
            "--proc",
            "/proc",
            "--dev",
            "/dev",
        ]
    )
    # A task workspace is commonly below /tmp during the startup probe. Create
    # its mount-point hierarchy only after /tmp is replaced, then bind the
    # exact workspace. Binding an ancestor would expose sibling arms or goldens.
    for parent in workspace_parent_paths:
        if parent != Path("/tmp"):
            command.extend(["--dir", str(parent)])
    command.extend(
        [
            "--bind",
            str(resolved_workspace),
            str(resolved_workspace),
            "--chdir",
            str(resolved_workspace),
            "--",
            *argv,
        ]
    )
    return command


def _environment(
    workspace: Path,
    workbook: Path,
    *,
    mutation_marker: Path | None = None,
) -> dict[str, str]:
    executable_dir = str(Path(sys.executable).absolute().parent)
    environment = {
        "PATH": f"{executable_dir}:/usr/local/bin:/usr/bin:/bin",
        "TMPDIR": "/tmp",
        "LANG": "C.UTF-8",
        "LC_ALL": "C.UTF-8",
        "PYTHONIOENCODING": "utf-8",
        "PYTHONDONTWRITEBYTECODE": "1",
        "OPENBLAS_NUM_THREADS": "1",
        "OMP_NUM_THREADS": "1",
        "MKL_NUM_THREADS": "1",
        "NUMEXPR_NUM_THREADS": "1",
        "VECLIB_MAXIMUM_THREADS": "1",
        "SHEET_WORKSPACE": str(workspace),
        "SHEET_WORKBOOK": str(workbook),
    }
    if mutation_marker is not None:
        environment[_MUTATION_MARKER_ENV] = str(mutation_marker)
    return environment


def _diagnostic(stderr: str, *, secrets: tuple[str, ...] = ()) -> str:
    environment_secrets = tuple(
        secret
        for name in ("OPENAI_API_KEY", "ANTHROPIC_AUTH_TOKEN")
        if (secret := os.environ.get(name))
    )
    redacted = redact_sensitive_text(
        stderr,
        secrets=(*secrets, *environment_secrets),
    )
    cleaned = " ".join(redacted.strip().split())[:1_000]
    return cleaned or "no diagnostic output"


def _file_sha256(path: Path) -> str | None:
    try:
        digest = hashlib.sha256()
        with path.open("rb") as handle:
            for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest()
    except OSError:
        return None


def _valid_a1_endpoint(value: str) -> bool:
    coordinate = value.replace("$", "")
    match = re.fullmatch(r"(?P<column>[A-Za-z]{1,3})(?P<row>[1-9]\d*)", coordinate)
    if match is None:
        return False
    return (
        column_index_from_string(match.group("column")) <= 16_384
        and int(match.group("row")) <= 1_048_576
    )


def _definite_a1_reference(value: str) -> tuple[str, bool, bool] | None:
    match = _DEFINITE_A1_REFERENCE.fullmatch(value)
    if match is None or not all(
        _valid_a1_endpoint(endpoint)
        for endpoint in (match.group("start"), match.group("end"))
        if endpoint is not None
    ):
        return None
    start = match.group("start")
    end = match.group("end")
    anchors = ":".join(
        f"{'C' if part.startswith('$') else 'c'}{'R' if '$' in part[1:] else 'r'}"
        for part in (start, end)
        if part is not None
    )
    shape = f"REF:{'range' if end else 'cell'}:{anchors}:{bool(match.group('sheet'))}"
    return shape, "$" in value, match.group("sheet") is not None


def _balanced_formula_tokens(tokens: list[Any]) -> bool:
    stack: list[str] = []
    expect_operand = True
    previous_kind: str | None = None
    for token in tokens:
        kind = token.type
        subtype = token.subtype
        if kind == "WHITE-SPACE":
            continue
        if kind in {"FUNC", "PAREN", "ARRAY"} and subtype == "OPEN":
            if not expect_operand:
                return False
            stack.append(kind)
            expect_operand = True
        elif kind in {"FUNC", "PAREN", "ARRAY"} and subtype == "CLOSE":
            if (
                not stack
                or stack.pop() != kind
                or (expect_operand and previous_kind not in {"FUNC", "SEP"})
            ):
                return False
            expect_operand = False
        elif kind == "OPERAND":
            if not expect_operand:
                return False
            expect_operand = False
        elif kind == "OPERATOR-PREFIX":
            if not expect_operand:
                return False
        elif kind == "OPERATOR-INFIX":
            if expect_operand:
                return False
            expect_operand = True
        elif kind == "OPERATOR-POSTFIX":
            if expect_operand:
                return False
        elif kind == "SEP" and subtype in {"ARG", "ROW"}:
            if expect_operand and previous_kind not in {"FUNC", "SEP"}:
                return False
            expect_operand = True
        else:
            return False
        previous_kind = kind
    return bool(tokens) and not stack and not expect_operand


def _formula_text_candidate(sheet: str, cell: Any) -> _FormulaTextCandidate | None:
    value = cell.value
    if (
        not isinstance(value, str)
        or value != value.strip()
        or value.startswith(("=", "'"))
        or "\n" in value
        or "\r" in value
        or cell.number_format == "@"
        or bool(
            getattr(
                cell,
                "quotePrefix",
                getattr(getattr(cell, "style_array", None), "quotePrefix", False),
            )
        )
    ):
        return None
    match = _FORMULA_TEXT_PREFIX.match(value)
    if match is None:
        return None
    function = match.group("function").upper().removeprefix("_XLFN.")
    if function not in _KNOWN_FORMULA_FUNCTIONS:
        return None
    try:
        tokens = Tokenizer("=" + value).items
    except Exception:
        return None
    if (
        not tokens
        or tokens[0].type != "FUNC"
        or tokens[0].subtype != "OPEN"
        or not _balanced_formula_tokens(tokens)
    ):
        return None

    shape: list[str] = []
    references = 0
    has_absolute = False
    has_cross_sheet = False
    has_operator = False
    for token in tokens:
        if token.type == "WHITE-SPACE":
            continue
        if token.type == "OPERAND" and token.subtype == "RANGE":
            reference = _definite_a1_reference(token.value)
            if reference is None:
                return None
            reference_shape, absolute, cross_sheet = reference
            shape.append(reference_shape)
            references += 1
            has_absolute = has_absolute or absolute
            has_cross_sheet = has_cross_sheet or cross_sheet
        else:
            shape.append(f"{token.type}:{token.subtype}:{token.value.upper()}")
        has_operator = has_operator or (
            token.type == "OPERATOR-INFIX" and token.value in _FORMULA_SIGNAL_OPERATORS
        )
    if references == 0:
        return None
    return _FormulaTextCandidate(
        sheet=sheet,
        cell=cell.coordinate,
        value=value,
        shape=tuple(shape),
        reference_operand_count=references,
        has_formula_operator=has_operator,
        has_absolute_or_cross_sheet_reference=has_absolute or has_cross_sheet,
    )


def _workbook_formula_state(
    path: Path,
) -> tuple[
    set[tuple[str, str, str, str]],
    dict[tuple[str, str], _FormulaTextCandidate],
]:
    workbook = load_workbook(
        path,
        read_only=True,
        data_only=False,
        keep_vba=path.suffix.lower() == ".xlsm",
        keep_links=True,
    )
    invalid_references: set[tuple[str, str, str, str]] = set()
    formula_text: dict[tuple[str, str], _FormulaTextCandidate] = {}
    try:
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if isinstance(value, str) and value.startswith("="):
                        try:
                            tokens = Tokenizer(value).items
                        except Exception:
                            continue
                        for token in tokens:
                            if (
                                token.type == "OPERAND"
                                and token.subtype == "RANGE"
                                and _INVALID_ABSOLUTE_ROW_REFERENCE.fullmatch(token.value)
                            ):
                                invalid_references.add(
                                    (worksheet.title, cell.coordinate, token.value, value)
                                )
                    else:
                        candidate = _formula_text_candidate(worksheet.title, cell)
                        if candidate is not None:
                            formula_text[(worksheet.title, cell.coordinate)] = candidate
    finally:
        workbook.close()
    return invalid_references, formula_text


def _has_adjacent_cells(candidates: list[_FormulaTextCandidate]) -> bool:
    coordinates = {coordinate_to_tuple(candidate.cell) for candidate in candidates}
    return any(
        (row + row_delta, column + column_delta) in coordinates
        for row, column in coordinates
        for row_delta, column_delta in ((0, 1), (1, 0))
    )


def _high_confidence_formula_text(
    introduced: dict[tuple[str, str], _FormulaTextCandidate],
) -> set[tuple[str, str, str]]:
    batch_groups: dict[tuple[str, tuple[str, ...]], list[_FormulaTextCandidate]] = {}
    for candidate in introduced.values():
        batch_groups.setdefault((candidate.sheet, candidate.shape), []).append(candidate)
    batch_members = {
        (candidate.sheet, candidate.cell)
        for candidates in batch_groups.values()
        if len(candidates) >= 3 and _has_adjacent_cells(candidates)
        for candidate in candidates
    }
    return {
        (candidate.sheet, candidate.cell, candidate.value)
        for key, candidate in introduced.items()
        if candidate.reference_operand_count >= 2
        or candidate.has_formula_operator
        or candidate.has_absolute_or_cross_sheet_reference
        or key in batch_members
    }


def validate_formula_transaction(
    previous_path: Path | None,
    current_path: Path,
) -> tuple[set[tuple[str, str, str, str]], set[tuple[str, str, str]]]:
    """Return high-confidence formula issues introduced between two workbooks."""

    if previous_path is None:
        previous_invalid_references: set[tuple[str, str, str, str]] = set()
        previous_formula_text: dict[tuple[str, str], _FormulaTextCandidate] = {}
    else:
        previous_invalid_references, previous_formula_text = _workbook_formula_state(
            previous_path
        )
    current_invalid_references, current_formula_text = _workbook_formula_state(current_path)
    introduced_formula_candidates = {
        key: candidate
        for key, candidate in current_formula_text.items()
        if previous_formula_text.get(key) != candidate
    }
    return (
        current_invalid_references - previous_invalid_references,
        _high_confidence_formula_text(introduced_formula_candidates),
    )


def _restore_workbook(snapshot: Path | None, workbook: Path) -> None:
    if snapshot is None:
        workbook.unlink(missing_ok=True)
        return
    temporary = workbook.with_name(
        f".{workbook.stem}.code-rollback-{uuid.uuid4().hex}{workbook.suffix}"
    )
    try:
        shutil.copy2(snapshot, temporary)
        temporary.replace(workbook)
    finally:
        temporary.unlink(missing_ok=True)


def _formula_validation_failure(
    invalid_references: set[tuple[str, str, str, str]],
    formula_text: set[tuple[str, str, str]],
) -> tuple[str, dict[str, Any]]:
    invalid_examples = [
        {
            "type": "invalid_a1_reference",
            "sheet": sheet,
            "cell": cell,
            "invalid_reference": reference,
            "formula": formula,
        }
        for sheet, cell, reference, formula in sorted(invalid_references)
    ]
    text_examples = [
        {
            "type": "missing_formula_prefix",
            "sheet": sheet,
            "cell": cell,
            "value": value,
        }
        for sheet, cell, value in sorted(formula_text)
    ]
    issues = invalid_examples + text_examples
    details: list[str] = []
    if invalid_references:
        locations = ", ".join(
            f"{sheet}!{cell} ({reference})"
            for sheet, cell, reference, _ in sorted(invalid_references)[:8]
        )
        details.append(f"invalid A1 references at {locations}")
    if formula_text:
        locations = ", ".join(
            f"{sheet}!{cell}" for sheet, cell, _ in sorted(formula_text)[:8]
        )
        details.append(f"formula-like text without a leading '=' at {locations}")
    guidance = []
    if invalid_references:
        guidance.append("use a column letter in cell references (for example E$5)")
    if formula_text:
        guidance.append("assign intended formulas as strings beginning with '='")
    error = "Workbook edit rolled back because it introduced " + "; ".join(details) + ". "
    error += "; then ".join(guidance) + ", save again, recalculate, and verify the target range."
    return error, {
        "ok": False,
        "issue_count": len(issues),
        "introduced_invalid_reference_count": len(invalid_references),
        "introduced_formula_text_count": len(formula_text),
        "issues": issues[:20],
        "truncated": len(issues) > 20,
    }


def _run_strict_probe(bubblewrap: Path, *, secrets: tuple[str, ...] = ()) -> None:
    with tempfile.TemporaryDirectory(prefix="sheet-code-isolation-") as temporary:
        root = Path(temporary)
        workspace = root / "workspace"
        workspace.mkdir(mode=0o700)
        outside = root / "must-not-be-readable.txt"
        outside.write_text("isolation-probe-secret", encoding="utf-8")
        script = workspace / "probe.py"
        script.write_text(
            f"""from pathlib import Path
import os
import resource
from openpyxl import Workbook

if hasattr(resource, "RLIMIT_NPROC"):
    resource.setrlimit(
        resource.RLIMIT_NPROC,
        ({_MAX_SANDBOX_PROCESSES}, {_MAX_SANDBOX_PROCESSES}),
    )
    if resource.getrlimit(resource.RLIMIT_NPROC) != (
        {_MAX_SANDBOX_PROCESSES},
        {_MAX_SANDBOX_PROCESSES},
    ):
        raise RuntimeError("strict sandbox process limit was not applied")
outside = Path(os.environ["SHEET_PROBE_OUTSIDE"])
try:
    outside.read_bytes()
except (FileNotFoundError, PermissionError):
    pass
else:
    raise RuntimeError("strict sandbox exposed a file outside the workspace")
book = Workbook()
book.active["A1"] = "probe"
book.save(os.environ["SHEET_WORKBOOK"])
book.close()
print("SHEET_STRICT_ISOLATION_OK")
""",
            encoding="utf-8",
        )
        workbook = workspace / "probe.xlsx"
        environment = _environment(workspace, workbook)
        environment["SHEET_PROBE_OUTSIDE"] = str(outside)
        command = _strict_command(
            workspace,
            [sys.executable, "-I", str(script)],
            bubblewrap=bubblewrap,
        )
        try:
            completed = subprocess.run(
                command,
                cwd=workspace,
                env=environment,
                text=True,
                capture_output=True,
                timeout=20,
                check=False,
                preexec_fn=_outer_sandbox_limits if os.name == "posix" else None,
            )
        except (OSError, subprocess.SubprocessError) as exc:
            detail = _diagnostic(str(exc), secrets=secrets)
            raise CodeIsolationError(
                "Strict comparison sandbox probe could not start: "
                f"{type(exc).__name__}: {detail}"
            ) from exc
        if (
            completed.returncode != 0
            or _PROBE_SENTINEL not in completed.stdout
            or not workbook.is_file()
        ):
            raise CodeIsolationError(
                "Strict comparison sandbox probe failed "
                f"(exit {completed.returncode}): "
                f"{_diagnostic(completed.stderr, secrets=secrets)}"
            )


def ensure_strict_code_isolation(secrets: tuple[str, ...] = ()) -> dict[str, str]:
    """Prove that strict isolation and the active venv's openpyxl work."""

    bubblewrap = _require_bubblewrap()
    key = (
        str(bubblewrap),
        sys.executable,
        sys.prefix,
        sys.base_prefix,
        STRICT_ISOLATION_POLICY,
    )
    with _PROBE_LOCK:
        if key not in _PROBE_SUCCESSES:
            _run_strict_probe(bubblewrap, secrets=secrets)
            _PROBE_SUCCESSES.add(key)
    return {
        "policy": STRICT_ISOLATION_POLICY,
        "bubblewrap": str(bubblewrap),
        "python": sys.executable,
    }


def _reset_isolation_probe_cache() -> None:
    """Clear successful probes for deterministic tests."""

    with _PROBE_LOCK:
        _PROBE_SUCCESSES.clear()


class LocalCodeInterpreter:
    def __init__(
        self,
        workspace: Path,
        workbook: Path,
        *,
        default_timeout: int = 30,
        max_output_chars: int = 20_000,
        require_isolation: bool = False,
        secrets: tuple[str, ...] = (),
    ) -> None:
        self.workspace = workspace.resolve()
        self.workbook = workbook.resolve()
        self.default_timeout = default_timeout
        self.max_output_chars = max_output_chars
        self.require_isolation = require_isolation
        self._secrets = tuple(secret for secret in secrets if secret)
        if self.require_isolation:
            if (
                self.workspace != self.workbook
                and self.workspace not in self.workbook.parents
            ):
                raise CodeIsolationError("SHEET_WORKBOOK must be inside the isolated workspace")
            ensure_strict_code_isolation(self._secrets)
        self.code_dir = self.workspace / "code"
        self.code_dir.mkdir(exist_ok=True)
        self._runtime_helper = self.workspace / _RUNTIME_HELPER_NAME
        self._runtime_helper.write_text(_RUNTIME_HELPER_SOURCE, encoding="utf-8")

    def _bounded_output(self, value: str | bytes | None) -> tuple[str, bool]:
        if isinstance(value, bytes):
            value = value.decode("utf-8", errors="replace")
        raw = value or ""
        redacted = redact_sensitive_text(raw, secrets=self._secrets)
        truncated = len(raw) > self.max_output_chars or len(redacted) > self.max_output_chars
        return redacted[: self.max_output_chars], truncated

    def _bounded_diagnostic(self, value: str | bytes | None) -> str:
        if isinstance(value, bytes):
            value = value.decode("utf-8", errors="replace")
        return _diagnostic(value or "", secrets=self._secrets)

    def _limits(self) -> None:
        _limits()

    def _command(
        self,
        script: Path,
        *,
        launcher: Path,
        marker: Path | None = None,
    ) -> tuple[list[str], str]:
        if self.require_isolation:
            if marker is None:
                raise CodeIsolationError("Strict sandbox launcher was not prepared")
            return (
                _strict_command(
                    self.workspace,
                    [
                        sys.executable,
                        "-I",
                        str(launcher),
                        str(script),
                        str(marker),
                    ],
                ),
                f"{STRICT_ISOLATION_POLICY}: writable workspace, runtime allowlist, no network",
            )

        base = [sys.executable, "-I", str(launcher), str(script)]
        bubblewrap = shutil.which("bwrap") if platform.system() == "Linux" else None
        if not bubblewrap:
            return base, "cwd+rlimit (trusted code only)"
        return (
            [
                bubblewrap,
                "--die-with-parent",
                "--new-session",
                "--unshare-net",
                "--ro-bind",
                "/",
                "/",
                "--bind",
                str(self.workspace),
                str(self.workspace),
                "--proc",
                "/proc",
                "--dev",
                "/dev",
                "--chdir",
                str(self.workspace),
                *base,
            ],
            "bubblewrap: read-only host, writable workspace, network disabled",
        )

    def _execute(
        self,
        command: list[str],
        *,
        environment: dict[str, str],
        timeout: int,
    ) -> subprocess.CompletedProcess[str]:
        preexec_fn = _outer_sandbox_limits if self.require_isolation else self._limits
        return subprocess.run(
            command,
            cwd=self.workspace,
            env=environment,
            text=True,
            capture_output=True,
            timeout=timeout,
            check=False,
            preexec_fn=preexec_fn if os.name == "posix" else None,
        )

    def run(self, code: str, *, timeout_seconds: int | None = None) -> dict[str, Any]:
        if not code.strip():
            raise ToolInputError("code must not be empty")
        if len(code) > 30_000:
            raise ToolInputError("code exceeds the 30,000 character limit")
        if any(marker in code.lower() for marker in _COMPRESSED_PLACEHOLDER_MARKERS):
            raise ToolInputError(
                "code contains a compressed/truncated placeholder; send complete runnable Python"
            )
        timeout = timeout_seconds or self.default_timeout
        if timeout < 1 or timeout > 60:
            raise ToolInputError("timeout_seconds must be between 1 and 60")
        before_sha256 = _file_sha256(self.workbook)
        identifier = uuid.uuid4().hex
        script = self.code_dir / f"snippet_{identifier}.py"
        script.write_text(code, encoding="utf-8")
        launcher = self.code_dir / f".launcher_{identifier}.py"
        mutation_marker = self.code_dir / f".managed_mutation_{identifier}"
        rollback_snapshot: Path | None = None
        if self.workbook.is_file():
            rollback_snapshot = (
                self.code_dir / f".workbook_before_{identifier}{self.workbook.suffix}"
            )
            shutil.copy2(self.workbook, rollback_snapshot)
        marker: Path | None = None
        launcher.write_text(
            f"""import resource
import runpy
import sys
from pathlib import Path

script_path = sys.argv[1]
marker_path = sys.argv[2] if len(sys.argv) > 2 else None
workspace = Path(script_path).resolve().parents[1]
sys.path.insert(0, str(workspace))
import {_RUNTIME_HELPER_NAME[:-3]} as sheet_harness

if hasattr(resource, "RLIMIT_NPROC"):
    resource.setrlimit(
        resource.RLIMIT_NPROC,
        ({_MAX_SANDBOX_PROCESSES}, {_MAX_SANDBOX_PROCESSES}),
    )
if marker_path is not None:
    with open(marker_path, "xb"):
        pass
sys.argv = [script_path]
runpy.run_path(
    script_path,
    run_name="__main__",
    init_globals={{"sheet_harness": sheet_harness}},
)
""",
            encoding="utf-8",
        )
        if self.require_isolation:
            marker = self.code_dir / f".sandbox_started_{identifier}"
        command, sandbox = self._command(script, launcher=launcher, marker=marker)
        environment = _environment(
            self.workspace,
            self.workbook,
            mutation_marker=mutation_marker,
        )
        try:
            completed = self._execute(command, environment=environment, timeout=timeout)
            if self.require_isolation and (marker is None or not marker.is_file()):
                raise CodeIsolationError(
                    "Strict comparison sandbox did not start; refusing unsandboxed fallback: "
                    + self._bounded_diagnostic(completed.stderr)
                )
            bubblewrap_error: str | None = None
            namespace_failure = (
                not self.require_isolation
                and sandbox.startswith("bubblewrap:")
                and completed.returncode != 0
                and any(
                    marker_text in completed.stderr.lower()
                    for marker_text in (
                        "creating new namespace failed",
                        "operation not permitted",
                        "permission denied",
                    )
                )
            )
            if namespace_failure:
                bubblewrap_error = self._bounded_diagnostic(completed.stderr)
                completed = self._execute(
                    [sys.executable, "-I", str(launcher), str(script)],
                    environment=environment,
                    timeout=timeout,
                )
                sandbox = "cwd+rlimit fallback (bubblewrap unavailable; trusted code only)"
            stdout, stdout_truncated = self._bounded_output(completed.stdout)
            stderr, stderr_truncated = self._bounded_output(completed.stderr)
            truncated = stdout_truncated or stderr_truncated
            managed_mutation_attempted = mutation_marker.is_file()
            after_sha256 = _file_sha256(self.workbook)
            workbook_changed = before_sha256 != after_sha256
            if completed.returncode != 0 and workbook_changed:
                rejected_sha256 = after_sha256
                _restore_workbook(rollback_snapshot, self.workbook)
                restored_sha256 = _file_sha256(self.workbook)
                return {
                    "ok": False,
                    "exit_code": completed.returncode,
                    "error": (
                        "Code execution failed after changing the workbook; all partial edits "
                        "were rolled back. Correct the script and apply the complete edit again."
                    ),
                    "stdout": stdout,
                    "stderr": stderr,
                    "truncated": truncated,
                    "sandbox": sandbox,
                    "bubblewrap_error": bubblewrap_error,
                    "script": str(script.relative_to(self.workspace)),
                    "workbook_sha256_before": before_sha256,
                    "workbook_sha256_rejected": rejected_sha256,
                    "workbook_sha256_after": restored_sha256,
                    "workbook_changed": False,
                    "workbook_rolled_back": True,
                    "managed_mutation_attempted": managed_mutation_attempted,
                    "helper_module": _RUNTIME_HELPER_NAME,
                    "message": (
                        "The failed code transaction was rolled back. Apply one complete edit, "
                        "save, recalculate, and verify."
                    ),
                }
            if completed.returncode == 0 and workbook_changed:
                try:
                    introduced_invalid_references, introduced_formula_text = (
                        validate_formula_transaction(
                            rollback_snapshot,
                            self.workbook,
                        )
                    )
                except Exception as exc:
                    rejected_sha256 = after_sha256
                    _restore_workbook(rollback_snapshot, self.workbook)
                    restored_sha256 = _file_sha256(self.workbook)
                    return {
                        "ok": False,
                        "exit_code": completed.returncode,
                        "error": (
                            "Workbook edit rolled back because the saved artifact could not "
                            "pass formula validation: "
                            f"{type(exc).__name__}: "
                            f"{redact_sensitive_text(str(exc), secrets=self._secrets)}"
                        ),
                        "stdout": stdout,
                        "stderr": stderr,
                        "truncated": truncated,
                        "sandbox": sandbox,
                        "bubblewrap_error": bubblewrap_error,
                        "script": str(script.relative_to(self.workspace)),
                        "workbook_sha256_before": before_sha256,
                        "workbook_sha256_rejected": rejected_sha256,
                        "workbook_sha256_after": restored_sha256,
                        "workbook_changed": False,
                        "workbook_rolled_back": True,
                        "managed_mutation_attempted": managed_mutation_attempted,
                        "helper_module": _RUNTIME_HELPER_NAME,
                        "message": (
                            "The invalid workbook edit was rolled back. Save a valid workbook "
                            "artifact, reopen it, and verify the target range."
                        ),
                    }
                if introduced_invalid_references or introduced_formula_text:
                    error, validation = _formula_validation_failure(
                        introduced_invalid_references,
                        introduced_formula_text,
                    )
                    rejected_sha256 = after_sha256
                    _restore_workbook(rollback_snapshot, self.workbook)
                    restored_sha256 = _file_sha256(self.workbook)
                    return {
                        "ok": False,
                        "exit_code": completed.returncode,
                        "error": error,
                        "stdout": stdout,
                        "stderr": stderr,
                        "truncated": truncated,
                        "sandbox": sandbox,
                        "bubblewrap_error": bubblewrap_error,
                        "script": str(script.relative_to(self.workspace)),
                        "workbook_sha256_before": before_sha256,
                        "workbook_sha256_rejected": rejected_sha256,
                        "workbook_sha256_after": restored_sha256,
                        "workbook_changed": False,
                        "workbook_rolled_back": True,
                        "managed_mutation_attempted": managed_mutation_attempted,
                        "formula_validation": validation,
                        "helper_module": _RUNTIME_HELPER_NAME,
                        "message": (
                            "The invalid workbook edit was rolled back. Correct every reported "
                            "formula issue in one complete edit, save, recalculate, and verify."
                        ),
                    }
            return {
                "ok": completed.returncode == 0,
                "exit_code": completed.returncode,
                "stdout": stdout,
                "stderr": stderr,
                "truncated": truncated,
                "sandbox": sandbox,
                "bubblewrap_error": bubblewrap_error,
                "script": str(script.relative_to(self.workspace)),
                "workbook_sha256_before": before_sha256,
                "workbook_sha256_after": after_sha256,
                "workbook_changed": workbook_changed,
                "managed_mutation_attempted": managed_mutation_attempted,
                "helper_module": _RUNTIME_HELPER_NAME,
                "message": (
                    "Workbook changed. If your script already reopened or inspected the exact "
                    "target range and stdout shows the expected state, finish now; otherwise run "
                    "one narrow verification or correction."
                    if before_sha256 is not None
                    and after_sha256 is not None
                    and before_sha256 != after_sha256
                    else (
                        "Workbook did not change. If this was meant to edit, save changes "
                        "back to SHEET_WORKBOOK before submitting."
                    )
                ),
            }
        except subprocess.TimeoutExpired as exc:
            after_sha256 = _file_sha256(self.workbook)
            if self.require_isolation and (marker is None or not marker.is_file()):
                raise CodeIsolationError(
                    "Strict comparison sandbox timed out before its launcher started"
                ) from exc
            workbook_changed = before_sha256 != after_sha256
            rejected_sha256: str | None = None
            if workbook_changed:
                rejected_sha256 = after_sha256
                _restore_workbook(rollback_snapshot, self.workbook)
                after_sha256 = _file_sha256(self.workbook)
            stdout, stdout_truncated = self._bounded_output(exc.stdout)
            stderr, stderr_truncated = self._bounded_output(exc.stderr)
            return {
                "ok": False,
                "error": (
                    f"Code execution timed out after {timeout} seconds"
                    + (
                        "; partial workbook edits were rolled back"
                        if workbook_changed
                        else ""
                    )
                ),
                "stdout": stdout,
                "stderr": stderr,
                "truncated": stdout_truncated or stderr_truncated,
                "sandbox": sandbox,
                "script": str(script.relative_to(self.workspace)),
                "workbook_sha256_before": before_sha256,
                "workbook_sha256_rejected": rejected_sha256,
                "workbook_sha256_after": after_sha256,
                "workbook_changed": False,
                "workbook_rolled_back": workbook_changed,
                "managed_mutation_attempted": mutation_marker.is_file(),
                "helper_module": _RUNTIME_HELPER_NAME,
            }
        except OSError as exc:
            if self.require_isolation:
                raise CodeIsolationError(
                    "Strict comparison sandbox could not start: "
                    f"{type(exc).__name__}: {self._bounded_diagnostic(str(exc))}"
                ) from exc
            raise
        finally:
            if launcher is not None:
                launcher.unlink(missing_ok=True)
            if marker is not None:
                marker.unlink(missing_ok=True)
            if rollback_snapshot is not None:
                rollback_snapshot.unlink(missing_ok=True)
            mutation_marker.unlink(missing_ok=True)
