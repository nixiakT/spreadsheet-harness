"""Typed, revision-aware evidence contracts for spreadsheet tool runs.

The monitor deliberately proves only workflow evidence properties. It does not
claim that an edited workbook satisfies the user's semantic intent.
"""

from __future__ import annotations

import hashlib
import json
import math
from collections.abc import Mapping
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path, PureWindowsPath
from typing import Any

import yaml
from openpyxl.utils import get_column_letter, range_boundaries


class ContractValidationError(ValueError):
    """Raised when a contract uses unsupported or ambiguous semantics."""


class ContractStateError(RuntimeError):
    """Raised when a trusted event is inconsistent with artifact lineage."""


class ContractMode(str, Enum):
    SHADOW = "shadow"
    ENFORCE = "enforce"


class EffectKind(str, Enum):
    VALUE = "value"
    FORMULA = "formula"
    STYLE = "style"
    STRUCTURE = "structure"
    VISUAL = "visual"
    UNKNOWN = "unknown"


class EventKind(str, Enum):
    MUTATION_COMMITTED = "mutation.committed"
    MUTATION_ROLLED_BACK = "mutation.rolled_back"
    ARTIFACT_REWRITTEN = "artifact.rewritten"
    RANGE_INSPECTED = "range.inspected"
    WORKBOOK_RECALCULATED = "workbook.recalculated"
    WORKBOOK_RENDERED = "workbook.rendered"
    RENDERED_PAGE_VIEWED = "rendered_page.viewed"


class TriggerKind(str, Enum):
    MUTATION_COMMITTED = "mutation.committed"
    FORMULA_CHANGED = "effects.formula_changed"
    VISUAL_CHANGED = "effects.visual_changed"


class ArtifactConstraint(str, Enum):
    CURRENT = "current"
    RECALCULATED_REVISION = "recalculated_revision"
    SAME_RENDER = "same_render"


class ScopePolicy(str, Enum):
    NONE = "none"
    CHANGED_CELLS_PLUS_BOUNDARY = "changed_cells_plus_boundary"
    CHANGED_FORMULA_CELLS_PLUS_BOUNDARY = "changed_formula_cells_plus_boundary"
    CHANGED_VISUAL_SCOPE = "changed_visual_scope"


_ALLOWED_PREDICATES = frozenset({"no_calc_error"})
_TRUSTED_EVIDENCE_EVENTS = frozenset(
    {
        EventKind.RANGE_INSPECTED,
        EventKind.WORKBOOK_RECALCULATED,
        EventKind.WORKBOOK_RENDERED,
        EventKind.RENDERED_PAGE_VIEWED,
    }
)
_VISUAL_EFFECTS = frozenset(
    {EffectKind.STYLE, EffectKind.STRUCTURE, EffectKind.VISUAL, EffectKind.UNKNOWN}
)
_EVENT_CHAIN_ALGORITHM = "sha256-canonical-json-chain-v1"
_EVENT_CHAIN_GENESIS_SHA256 = "0" * 64
_CERTIFICATE_DIGEST_ALGORITHM = "sha256-canonical-json-v1"
_RECALCULATION_METADATA_KEYS = frozenset(
    {"backend", "version", "source_sha256", "output_sha256", "atomic_replace"}
)
_RENDER_METADATA_KEYS = frozenset(
    {"producer_tool", "backend", "version", "mode", "dpi", "page_count", "pages"}
)
_RENDER_PAGE_METADATA_KEYS = frozenset(
    {
        "page_id",
        "page_index",
        "file_sha256",
        "width",
        "height",
        "sheet",
        "sheet_page",
        "cell_scope",
    }
)
_VIEW_METADATA_KEYS = frozenset(
    {
        "producer_tool",
        "delivery_status",
        "confirmation_id",
        "provider_response_id",
        "attachment_file_sha256",
        "page_file_sha256",
        "page_pixel_sha256",
        "pixel_sha256_algorithm",
        "width",
        "height",
        "image_mode",
        "render_mode",
        "page_index",
        "sheet",
        "sheet_page",
        "cell_scope",
    }
)
PIXEL_SHA256_ALGORITHM = "sha256-pillow-mode-size-pixels-v1"


def _require_keys(
    mapping: Mapping[str, Any],
    *,
    required: frozenset[str],
    optional: frozenset[str],
    context: str,
) -> None:
    if not isinstance(mapping, Mapping):
        raise ContractValidationError(f"{context} must be a mapping")
    if not all(isinstance(key, str) for key in mapping):
        raise ContractValidationError(f"{context} keys must be strings")
    keys = frozenset(str(key) for key in mapping)
    missing = sorted(required - keys)
    unknown = sorted(keys - required - optional)
    if missing:
        raise ContractValidationError(f"{context} is missing keys: {', '.join(missing)}")
    if unknown:
        raise ContractValidationError(f"{context} has unknown keys: {', '.join(unknown)}")


def _validate_sha256(value: str, *, context: str) -> None:
    if not isinstance(value, str) or len(value) != 64 or any(
        character not in "0123456789abcdef" for character in value
    ):
        raise ValueError(f"{context} must be 64 lowercase hexadecimal characters")


def _normalize_json_value(value: Any, *, context: str) -> Any:
    if value is None or type(value) in {str, bool, int}:
        return value
    if type(value) is float:
        if not math.isfinite(value):
            raise ValueError(f"{context} must not contain non-finite numbers")
        return value
    if isinstance(value, list):
        return [
            _normalize_json_value(item, context=f"{context}[{index}]")
            for index, item in enumerate(value)
        ]
    if isinstance(value, Mapping):
        if not all(isinstance(key, str) for key in value):
            raise ValueError(f"{context} object keys must be strings")
        return {
            key: _normalize_json_value(item, context=f"{context}.{key}")
            for key, item in value.items()
        }
    raise ValueError(
        f"{context} contains non-JSON value of type {type(value).__name__}"
    )


def _canonical_json_bytes(value: Any, *, context: str) -> bytes:
    normalized = _normalize_json_value(value, context=context)
    try:
        rendered = json.dumps(
            normalized,
            ensure_ascii=True,
            sort_keys=True,
            separators=(",", ":"),
            allow_nan=False,
        )
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{context} must be canonical JSON data: {exc}") from exc
    return rendered.encode("ascii")


def _validate_portable_label(value: Any, *, context: str) -> None:
    if not isinstance(value, str) or not value:
        raise ValueError(f"{context} must be a non-empty string")
    if Path(value).is_absolute() or PureWindowsPath(value).is_absolute():
        raise ValueError(f"{context} must not contain an absolute path")


@dataclass(frozen=True)
class ArtifactRef:
    """A lineage revision and the exact bytes published at that revision."""

    revision: int
    sha256: str

    def __post_init__(self) -> None:
        if type(self.revision) is not int or self.revision < 0:
            raise ValueError("artifact revision must be non-negative")
        _validate_sha256(self.sha256, context="artifact sha256")

    def to_dict(self) -> dict[str, Any]:
        return {"revision": self.revision, "sha256": self.sha256}


@dataclass(frozen=True)
class ArtifactTransition:
    """One successful publication of new workbook bytes."""

    transition_id: int
    operation: str
    kind: str
    before: ArtifactRef
    after: ArtifactRef

    def __post_init__(self) -> None:
        if type(self.transition_id) is not int or self.transition_id < 1:
            raise ValueError("transition_id must be positive")
        if not isinstance(self.operation, str) or not isinstance(self.kind, str):
            raise TypeError("transition operation and kind must be strings")
        if not self.operation or not self.kind:
            raise ValueError("transition operation and kind must not be empty")
        if self.after.revision != self.before.revision + 1:
            raise ValueError("artifact transitions must advance exactly one revision")
        if self.after.sha256 == self.before.sha256:
            raise ValueError("artifact transitions require new workbook bytes")

    def to_dict(self) -> dict[str, Any]:
        return {
            "transition_id": self.transition_id,
            "operation": self.operation,
            "kind": self.kind,
            "before": self.before.to_dict(),
            "after": self.after.to_dict(),
        }


@dataclass(frozen=True, order=True)
class CellRange:
    """One normalized, bounded worksheet range."""

    sheet: str
    min_col: int
    min_row: int
    max_col: int
    max_row: int

    def __post_init__(self) -> None:
        if not isinstance(self.sheet, str) or not self.sheet:
            raise ValueError("sheet must not be empty")
        if any(
            type(value) is not int
            for value in (self.min_col, self.min_row, self.max_col, self.max_row)
        ):
            raise TypeError("range coordinates must be integers")
        if min(self.min_col, self.min_row, self.max_col, self.max_row) < 1:
            raise ValueError("range coordinates must be positive")
        if self.min_col > self.max_col or self.min_row > self.max_row:
            raise ValueError("range bounds are reversed")

    @classmethod
    def parse(cls, sheet: str, range_ref: str) -> CellRange:
        try:
            min_col, min_row, max_col, max_row = range_boundaries(
                range_ref.replace("$", "")
            )
        except (TypeError, ValueError) as exc:
            raise ValueError(f"Invalid bounded A1 range: {range_ref!r}") from exc
        if not all(isinstance(value, int) for value in (min_col, min_row, max_col, max_row)):
            raise ValueError(f"Range must be bounded: {range_ref!r}")
        return cls(sheet, min_col, min_row, max_col, max_row)

    @property
    def a1(self) -> str:
        return (
            f"{get_column_letter(self.min_col)}{self.min_row}:"
            f"{get_column_letter(self.max_col)}{self.max_row}"
        )

    @property
    def cell_count(self) -> int:
        return (self.max_col - self.min_col + 1) * (self.max_row - self.min_row + 1)

    def expand(self, boundary: int) -> CellRange:
        if boundary < 0:
            raise ValueError("boundary must be non-negative")
        return CellRange(
            self.sheet,
            max(1, self.min_col - boundary),
            max(1, self.min_row - boundary),
            self.max_col + boundary,
            self.max_row + boundary,
        )

    def covers(self, other: CellRange) -> bool:
        return bool(
            self.sheet == other.sheet
            and self.min_col <= other.min_col
            and self.min_row <= other.min_row
            and self.max_col >= other.max_col
            and self.max_row >= other.max_row
        )

    def intersects(self, other: CellRange) -> bool:
        return bool(
            self.sheet == other.sheet
            and self.min_col <= other.max_col
            and other.min_col <= self.max_col
            and self.min_row <= other.max_row
            and other.min_row <= self.max_row
        )

    def to_dict(self) -> dict[str, Any]:
        return {"sheet": self.sheet, "range": self.a1, "cell_count": self.cell_count}


@dataclass(frozen=True)
class EvidenceScope:
    """A finite range union, or an explicitly unbounded workbook scope."""

    ranges: tuple[CellRange, ...] = ()
    sheets: tuple[str, ...] = ()
    wildcard: bool = False

    def __post_init__(self) -> None:
        if type(self.wildcard) is not bool:
            raise TypeError("scope wildcard must be boolean")
        if not isinstance(self.ranges, tuple) or not all(
            isinstance(item, CellRange) for item in self.ranges
        ):
            raise TypeError("scope ranges must contain CellRange values")
        if not isinstance(self.sheets, tuple) or not all(
            isinstance(item, str) for item in self.sheets
        ):
            raise TypeError("scope sheets must contain strings")
        normalized = tuple(sorted(set(self.ranges)))
        if normalized != self.ranges:
            object.__setattr__(self, "ranges", normalized)
        normalized_sheets = tuple(sorted(set(self.sheets)))
        if any(not sheet for sheet in normalized_sheets):
            raise ValueError("scope sheet names must not be empty")
        if normalized_sheets != self.sheets:
            object.__setattr__(self, "sheets", normalized_sheets)
        if self.wildcard and (self.ranges or self.sheets):
            raise ValueError("workbook wildcard scope cannot also list sheets or ranges")

    @classmethod
    def one(cls, sheet: str, range_ref: str) -> EvidenceScope:
        return cls((CellRange.parse(sheet, range_ref),))

    @classmethod
    def workbook(cls) -> EvidenceScope:
        return cls(wildcard=True)

    @classmethod
    def worksheet(cls, sheet: str) -> EvidenceScope:
        return cls(sheets=(sheet,))

    @classmethod
    def from_dict(cls, value: Mapping[str, Any]) -> EvidenceScope:
        _require_keys(
            value,
            required=frozenset({"wildcard", "sheets", "ranges"}),
            optional=frozenset(),
            context="evidence scope",
        )
        if not isinstance(value["wildcard"], bool):
            raise ValueError("evidence scope wildcard must be boolean")
        raw_sheets = value["sheets"]
        raw_ranges = value["ranges"]
        if not isinstance(raw_sheets, list) or not all(
            isinstance(item, str) for item in raw_sheets
        ):
            raise ValueError("evidence scope sheets must be a string list")
        if not isinstance(raw_ranges, list):
            raise ValueError("evidence scope ranges must be a list")
        ranges: list[CellRange] = []
        for index, item in enumerate(raw_ranges):
            if not isinstance(item, Mapping):
                raise ValueError(f"evidence scope ranges[{index}] must be a mapping")
            keys = set(item)
            if keys not in ({"sheet", "range"}, {"sheet", "range", "cell_count"}):
                raise ValueError(
                    f"evidence scope ranges[{index}] has invalid keys: {sorted(keys)}"
                )
            if not isinstance(item.get("sheet"), str) or not isinstance(
                item.get("range"), str
            ):
                raise ValueError(
                    f"evidence scope ranges[{index}] needs string sheet and range"
                )
            parsed = CellRange.parse(item["sheet"], item["range"])
            if "cell_count" in item and (
                type(item["cell_count"]) is not int
                or item["cell_count"] != parsed.cell_count
            ):
                raise ValueError(
                    f"evidence scope ranges[{index}].cell_count does not match its range"
                )
            ranges.append(parsed)
        return cls(
            tuple(ranges),
            sheets=tuple(raw_sheets),
            wildcard=value["wildcard"],
        )

    @property
    def empty(self) -> bool:
        return not self.wildcard and not self.ranges and not self.sheets

    def expand(self, boundary: int = 1) -> EvidenceScope:
        if self.wildcard:
            return self
        return EvidenceScope(
            tuple(item.expand(boundary) for item in self.ranges),
            sheets=self.sheets,
        )

    def covers(self, required: EvidenceScope) -> bool:
        if required.empty:
            return True
        if self.wildcard:
            return True
        if required.wildcard:
            return False
        if not all(sheet in self.sheets for sheet in required.sheets):
            return False
        return all(
            target.sheet in self.sheets
            or any(actual.covers(target) for actual in self.ranges)
            for target in required.ranges
        )

    def merged(self, other: EvidenceScope) -> EvidenceScope:
        if self.wildcard or other.wildcard:
            return EvidenceScope.workbook()
        return EvidenceScope(
            self.ranges + other.ranges,
            sheets=self.sheets + other.sheets,
        )

    def intersects(self, other: EvidenceScope) -> bool:
        if self.empty or other.empty:
            return False
        if self.wildcard or other.wildcard:
            return True
        if set(self.sheets) & set(other.sheets):
            return True
        if any(item.sheet in other.sheets for item in self.ranges):
            return True
        if any(item.sheet in self.sheets for item in other.ranges):
            return True
        return any(
            left.intersects(right)
            for left in self.ranges
            for right in other.ranges
        )

    def to_dict(self) -> dict[str, Any]:
        return {
            "wildcard": self.wildcard,
            "sheets": list(self.sheets),
            "ranges": [item.to_dict() for item in self.ranges],
        }


def _validate_positive_int(value: Any, *, context: str) -> None:
    if type(value) is not int or value < 1:
        raise ValueError(f"{context} must be a positive integer")


def _metadata_scope(value: Any, *, context: str) -> EvidenceScope | None:
    if value is None:
        return None
    if not isinstance(value, Mapping):
        raise ValueError(f"{context} must be null or an evidence scope")
    scope = EvidenceScope.from_dict(value)
    if scope.empty or scope.wildcard or scope.sheets:
        raise ValueError(f"{context} must contain finite cell ranges only")
    return scope


def _validate_render_metadata(metadata: Mapping[str, Any]) -> None:
    if set(metadata) != set(_RENDER_METADATA_KEYS):
        raise ValueError("workbook.rendered requires portable authenticated page metadata")
    if metadata["producer_tool"] != "render_workbook":
        raise ValueError("workbook.rendered producer_tool must be render_workbook")
    _validate_portable_label(metadata["backend"], context="render backend")
    version = metadata["version"]
    if not isinstance(version, Mapping) or not version:
        raise ValueError("render version must be a non-empty mapping")
    for key, value in version.items():
        _validate_portable_label(key, context="render version key")
        _validate_portable_label(value, context=f"render version {key}")
    if metadata["mode"] not in {"per_sheet", "whole_workbook"}:
        raise ValueError("render mode is unsupported")
    _validate_positive_int(metadata["dpi"], context="render dpi")
    _validate_positive_int(metadata["page_count"], context="render page_count")
    pages = metadata["pages"]
    if not isinstance(pages, list) or len(pages) != metadata["page_count"]:
        raise ValueError("render pages must match page_count")
    page_ids: list[str] = []
    for index, page in enumerate(pages, start=1):
        if not isinstance(page, Mapping) or set(page) != set(_RENDER_PAGE_METADATA_KEYS):
            raise ValueError(f"render pages[{index - 1}] has an invalid schema")
        page_id = page["page_id"]
        if not isinstance(page_id, str) or not page_id:
            raise ValueError(f"render pages[{index - 1}].page_id must be non-empty")
        page_ids.append(page_id)
        if page["page_index"] != index:
            raise ValueError("render page indices must be contiguous and 1-based")
        _validate_sha256(
            page["file_sha256"],
            context=f"render pages[{index - 1}].file_sha256",
        )
        _validate_positive_int(page["width"], context="render page width")
        _validate_positive_int(page["height"], context="render page height")
        sheet = page["sheet"]
        sheet_page = page["sheet_page"]
        if metadata["mode"] == "per_sheet":
            if not isinstance(sheet, str) or not sheet:
                raise ValueError("per-sheet render pages require a worksheet name")
            _validate_positive_int(sheet_page, context="render sheet_page")
        elif sheet is not None or sheet_page is not None:
            raise ValueError("whole-workbook render pages must not claim worksheet mapping")
        cell_scope = _metadata_scope(
            page["cell_scope"],
            context=f"render pages[{index - 1}].cell_scope",
        )
        if cell_scope is not None and isinstance(sheet, str) and any(
            item.sheet != sheet for item in cell_scope.ranges
        ):
            raise ValueError("render page cell_scope must stay within its worksheet")
    if len(page_ids) != len(set(page_ids)):
        raise ValueError("render page IDs must be unique")


def _validate_view_metadata(event: EvidenceEvent, metadata: Mapping[str, Any]) -> None:
    if set(metadata) != set(_VIEW_METADATA_KEYS):
        raise ValueError("rendered_page.viewed requires confirmed delivery metadata")
    if metadata["producer_tool"] != "view_image":
        raise ValueError("rendered_page.viewed producer_tool must be view_image")
    if metadata["delivery_status"] != "provider_response_confirmed":
        raise ValueError("rendered_page.viewed requires a successful provider response")
    for key in ("confirmation_id", "provider_response_id", "image_mode"):
        _validate_portable_label(metadata[key], context=f"view {key}")
    for key in (
        "attachment_file_sha256",
        "page_file_sha256",
        "page_pixel_sha256",
    ):
        _validate_sha256(metadata[key], context=f"view {key}")
    if metadata["pixel_sha256_algorithm"] != PIXEL_SHA256_ALGORITHM:
        raise ValueError("rendered_page.viewed pixel hash algorithm is unsupported")
    if not (
        metadata["attachment_file_sha256"]
        == metadata["page_file_sha256"]
        == event.page_sha256
    ):
        raise ValueError("rendered_page.viewed file hashes do not match")
    _validate_positive_int(metadata["width"], context="view width")
    _validate_positive_int(metadata["height"], context="view height")
    _validate_positive_int(metadata["page_index"], context="view page_index")
    if metadata["render_mode"] not in {"per_sheet", "whole_workbook"}:
        raise ValueError("view render_mode is unsupported")
    sheet = metadata["sheet"]
    sheet_page = metadata["sheet_page"]
    if metadata["render_mode"] == "per_sheet":
        if not isinstance(sheet, str) or not sheet:
            raise ValueError("per-sheet views require a worksheet name")
        _validate_positive_int(sheet_page, context="view sheet_page")
    elif sheet is not None or sheet_page is not None:
        raise ValueError("whole-workbook views must not claim worksheet mapping")
    cell_scope = _metadata_scope(metadata["cell_scope"], context="view cell_scope")
    if cell_scope is None:
        if not event.scope.empty:
            raise ValueError("a page without cell mapping must not claim cell or worksheet scope")
    elif event.scope != cell_scope:
        raise ValueError("rendered_page.viewed scope must equal its authenticated cell mapping")


@dataclass(frozen=True)
class Requirement:
    event: EventKind
    artifact: ArtifactConstraint
    scope: ScopePolicy = ScopePolicy.NONE
    predicates: frozenset[str] = frozenset()

    @classmethod
    def from_mapping(cls, value: Mapping[str, Any], *, context: str) -> Requirement:
        _require_keys(
            value,
            required=frozenset({"event"}),
            optional=frozenset({"artifact", "scope", "predicates"}),
            context=context,
        )
        try:
            event = EventKind(str(value["event"]))
        except ValueError as exc:
            raise ContractValidationError(f"{context} has unsupported event {value['event']!r}") from exc
        if event not in _TRUSTED_EVIDENCE_EVENTS:
            raise ContractValidationError(
                f"{context} event {event.value!r} is not a trusted evidence event"
            )
        artifact_value = value.get("artifact", ArtifactConstraint.CURRENT.value)
        try:
            artifact = ArtifactConstraint(str(artifact_value))
        except ValueError as exc:
            raise ContractValidationError(
                f"{context} has unsupported artifact constraint {artifact_value!r}"
            ) from exc
        scope_value = value.get("scope", ScopePolicy.NONE.value)
        try:
            scope = ScopePolicy(str(scope_value))
        except ValueError as exc:
            raise ContractValidationError(
                f"{context} has unsupported scope policy {scope_value!r}"
            ) from exc
        raw_predicates = value.get("predicates", [])
        if not isinstance(raw_predicates, list) or not all(
            isinstance(item, str) for item in raw_predicates
        ):
            raise ContractValidationError(f"{context}.predicates must be a string list")
        predicates = frozenset(raw_predicates)
        unknown_predicates = sorted(predicates - _ALLOWED_PREDICATES)
        if unknown_predicates:
            raise ContractValidationError(
                f"{context} has unsupported predicates: {', '.join(unknown_predicates)}"
            )
        if artifact is ArtifactConstraint.SAME_RENDER and event is not EventKind.RENDERED_PAGE_VIEWED:
            raise ContractValidationError(
                f"{context} may use same_render only with rendered_page.viewed"
            )
        if scope in {
            ScopePolicy.CHANGED_CELLS_PLUS_BOUNDARY,
            ScopePolicy.CHANGED_FORMULA_CELLS_PLUS_BOUNDARY,
        } and event is not EventKind.RANGE_INSPECTED:
            raise ContractValidationError(
                f"{context} may attach a changed-cell scope only to range.inspected"
            )
        if (
            scope is ScopePolicy.CHANGED_VISUAL_SCOPE
            and event is not EventKind.RENDERED_PAGE_VIEWED
        ):
            raise ContractValidationError(
                f"{context} may attach changed_visual_scope only to rendered_page.viewed"
            )
        if predicates and event is not EventKind.RANGE_INSPECTED:
            raise ContractValidationError(
                f"{context} may attach predicates only to range.inspected"
            )
        return cls(event=event, artifact=artifact, scope=scope, predicates=predicates)

    def to_dict(self) -> dict[str, Any]:
        return {
            "event": self.event.value,
            "artifact": self.artifact.value,
            "scope": self.scope.value,
            "predicates": sorted(self.predicates),
        }


@dataclass(frozen=True)
class ContractRule:
    id: str
    trigger: TriggerKind
    requirements: tuple[Requirement, ...]

    @classmethod
    def from_mapping(cls, value: Mapping[str, Any], *, index: int) -> ContractRule:
        context = f"rules[{index}]"
        _require_keys(
            value,
            required=frozenset({"id", "trigger"}),
            optional=frozenset({"require", "require_sequence"}),
            context=context,
        )
        identifier = value["id"]
        if not isinstance(identifier, str) or not identifier.strip():
            raise ContractValidationError(f"{context}.id must be a non-empty string")
        try:
            trigger = TriggerKind(str(value["trigger"]))
        except ValueError as exc:
            raise ContractValidationError(
                f"{context} has unsupported trigger {value['trigger']!r}"
            ) from exc
        has_single = "require" in value
        has_sequence = "require_sequence" in value
        if has_single == has_sequence:
            raise ContractValidationError(
                f"{context} must define exactly one of require or require_sequence"
            )
        raw_requirements: Any
        if has_single:
            raw_requirements = [value["require"]]
        else:
            raw_requirements = value["require_sequence"]
        if not isinstance(raw_requirements, list) or not raw_requirements:
            raise ContractValidationError(f"{context} requirements must be a non-empty list")
        requirements: list[Requirement] = []
        for requirement_index, raw_requirement in enumerate(raw_requirements):
            if not isinstance(raw_requirement, Mapping):
                raise ContractValidationError(
                    f"{context}.requirements[{requirement_index}] must be a mapping"
                )
            requirements.append(
                Requirement.from_mapping(
                    raw_requirement,
                    context=f"{context}.requirements[{requirement_index}]",
                )
            )
        for requirement_index, requirement in enumerate(requirements):
            earlier = requirements[:requirement_index]
            if (
                requirement.artifact is ArtifactConstraint.RECALCULATED_REVISION
                and not any(
                    item.event is EventKind.WORKBOOK_RECALCULATED for item in earlier
                )
            ):
                raise ContractValidationError(
                    f"{context} references recalculated_revision without an earlier "
                    "recalculation step"
                )
            if (
                requirement.artifact is ArtifactConstraint.SAME_RENDER
                and not any(item.event is EventKind.WORKBOOK_RENDERED for item in earlier)
            ):
                raise ContractValidationError(
                    f"{context} references same_render without an earlier render step"
                )
        return cls(id=identifier, trigger=trigger, requirements=tuple(requirements))

    def to_dict(self) -> dict[str, Any]:
        return {
            "id": self.id,
            "trigger": self.trigger.value,
            "requirements": [item.to_dict() for item in self.requirements],
        }


@dataclass(frozen=True)
class ContractSpec:
    schema_version: int
    rules: tuple[ContractRule, ...]
    canonical_sha256: str
    source_sha256: str | None = None
    source_path: str | None = None

    @classmethod
    def from_mapping(
        cls,
        value: Mapping[str, Any],
        *,
        source_bytes: bytes | None = None,
        source_path: str | None = None,
    ) -> ContractSpec:
        _require_keys(
            value,
            required=frozenset({"schema_version", "rules"}),
            optional=frozenset(),
            context="contract",
        )
        if type(value["schema_version"]) is not int or value["schema_version"] != 1:
            raise ContractValidationError("contract.schema_version must be 1")
        raw_rules = value["rules"]
        if not isinstance(raw_rules, list) or not raw_rules:
            raise ContractValidationError("contract.rules must be a non-empty list")
        rules: list[ContractRule] = []
        for index, raw_rule in enumerate(raw_rules):
            if not isinstance(raw_rule, Mapping):
                raise ContractValidationError(f"rules[{index}] must be a mapping")
            rules.append(ContractRule.from_mapping(raw_rule, index=index))
        identifiers = [rule.id for rule in rules]
        duplicate_ids = sorted({identifier for identifier in identifiers if identifiers.count(identifier) > 1})
        if duplicate_ids:
            raise ContractValidationError(
                f"contract has duplicate rule ids: {', '.join(duplicate_ids)}"
            )
        canonical = {
            "schema_version": 1,
            "rules": [rule.to_dict() for rule in rules],
        }
        canonical_bytes = _canonical_json_bytes(canonical, context="contract")
        return cls(
            schema_version=1,
            rules=tuple(rules),
            canonical_sha256=hashlib.sha256(canonical_bytes).hexdigest(),
            source_sha256=(hashlib.sha256(source_bytes).hexdigest() if source_bytes is not None else None),
            source_path=source_path,
        )

    @classmethod
    def load(cls, path: str | Path) -> ContractSpec:
        source = Path(path)
        raw = source.read_bytes()
        try:
            value = yaml.safe_load(raw)
        except yaml.YAMLError as exc:
            raise ContractValidationError(f"Invalid contract YAML: {exc}") from exc
        if not isinstance(value, Mapping):
            raise ContractValidationError("contract YAML root must be a mapping")
        return cls.from_mapping(value, source_bytes=raw, source_path=str(source))

    def to_dict(self) -> dict[str, Any]:
        return {
            **self.identity_dict(),
            "source_path": self.source_path,
        }

    def identity_dict(self) -> dict[str, Any]:
        """Return the path-independent contract identity used in certificates."""

        return {
            "schema_version": self.schema_version,
            "canonical_sha256": self.canonical_sha256,
            "source_sha256": self.source_sha256,
            "rules": [rule.to_dict() for rule in self.rules],
        }


@dataclass(frozen=True)
class EvidenceEvent:
    """One trusted harness-emitted artifact or evidence event."""

    kind: EventKind
    revision_before: str
    revision_after: str | None = None
    effects: frozenset[EffectKind] = frozenset()
    scope: EvidenceScope = EvidenceScope()
    formula_scope: EvidenceScope = EvidenceScope()
    predicates: frozenset[str] = frozenset()
    render_id: str | None = None
    render_manifest_sha256: str | None = None
    related_render_id: str | None = None
    related_render_manifest_sha256: str | None = None
    page_id: str | None = None
    page_sha256: str | None = None
    metadata: Mapping[str, Any] = field(default_factory=dict)

    def __post_init__(self) -> None:
        if not isinstance(self.kind, EventKind):
            raise TypeError("event kind must be EventKind")
        if not isinstance(self.effects, frozenset) or not all(
            isinstance(item, EffectKind) for item in self.effects
        ):
            raise TypeError("event effects must be a frozenset of EffectKind values")
        if not isinstance(self.scope, EvidenceScope) or not isinstance(
            self.formula_scope, EvidenceScope
        ):
            raise TypeError("event scope and formula_scope must be EvidenceScope values")
        if not isinstance(self.predicates, frozenset) or not all(
            isinstance(item, str) for item in self.predicates
        ):
            raise TypeError("event predicates must be a frozenset of strings")
        if not isinstance(self.metadata, Mapping):
            raise TypeError("event metadata must be a mapping")
        if not self.revision_before:
            raise ValueError("revision_before must not be empty")
        _validate_sha256(self.revision_before, context="revision_before")
        if self.revision_after is not None:
            _validate_sha256(self.revision_after, context="revision_after")
        revision_events = {
            EventKind.MUTATION_COMMITTED,
            EventKind.ARTIFACT_REWRITTEN,
            EventKind.WORKBOOK_RECALCULATED,
        }
        if self.kind in revision_events and not self.revision_after:
            raise ValueError(f"{self.kind.value} requires revision_after")
        if self.kind not in revision_events and self.revision_after is not None:
            raise ValueError(f"{self.kind.value} must not publish a new artifact revision")
        if self.kind is EventKind.WORKBOOK_RENDERED:
            if not isinstance(self.render_id, str):
                raise TypeError("workbook.rendered render_id must be a string")
            if not self.render_id or not self.render_manifest_sha256:
                raise ValueError(
                    "workbook.rendered requires render_id and render_manifest_sha256"
                )
            _validate_sha256(
                self.render_manifest_sha256,
                context="render_manifest_sha256",
            )
        elif self.render_id is not None or self.render_manifest_sha256 is not None:
            raise ValueError(
                f"{self.kind.value} must not define render_id or render_manifest_sha256"
            )
        if self.kind is EventKind.RENDERED_PAGE_VIEWED:
            if not all(
                isinstance(item, str)
                for item in (
                    self.related_render_id,
                    self.related_render_manifest_sha256,
                    self.page_id,
                    self.page_sha256,
                )
            ):
                raise TypeError("rendered_page.viewed identity fields must be strings")
            if not all(
                (
                    self.related_render_id,
                    self.related_render_manifest_sha256,
                    self.page_id,
                    self.page_sha256,
                )
            ):
                raise ValueError(
                    "rendered_page.viewed requires render identity, manifest, page id, and page hash"
                )
            assert self.related_render_manifest_sha256 is not None
            assert self.page_sha256 is not None
            _validate_sha256(
                self.related_render_manifest_sha256,
                context="related_render_manifest_sha256",
            )
            _validate_sha256(self.page_sha256, context="page_sha256")
            if self.scope.wildcard or self.scope.sheets:
                raise ValueError(
                    "rendered_page.viewed cannot promote one page to workbook or worksheet scope"
                )
        elif any(
            item is not None
            for item in (
                self.related_render_id,
                self.related_render_manifest_sha256,
                self.page_id,
                self.page_sha256,
            )
        ):
            raise ValueError(f"{self.kind.value} must not define rendered-page identity")
        unknown_predicates = self.predicates - _ALLOWED_PREDICATES
        if unknown_predicates:
            raise ValueError(f"Unsupported event predicates: {sorted(unknown_predicates)}")
        if self.kind is not EventKind.RANGE_INSPECTED and self.predicates:
            raise ValueError(f"{self.kind.value} must not define inspection predicates")
        if self.kind is EventKind.MUTATION_COMMITTED and self.changed and not self.effects:
            raise ValueError("changed mutation events require at least one typed effect")
        if self.kind is EventKind.MUTATION_COMMITTED and self.changed and self.scope.empty:
            raise ValueError("changed mutation events require an explicit non-empty scope")
        if self.kind is EventKind.MUTATION_ROLLED_BACK and self.scope.empty:
            raise ValueError("rolled-back mutation events require an explicit failure scope")
        if self.kind is EventKind.RANGE_INSPECTED and self.scope.empty:
            raise ValueError("range.inspected requires a non-empty inspected scope")
        scoped_events = {
            EventKind.MUTATION_COMMITTED,
            EventKind.MUTATION_ROLLED_BACK,
            EventKind.RANGE_INSPECTED,
            EventKind.RENDERED_PAGE_VIEWED,
        }
        if self.kind not in scoped_events and not self.scope.empty:
            raise ValueError(f"{self.kind.value} must not define an evidence scope")
        if self.kind is not EventKind.MUTATION_COMMITTED and self.effects:
            raise ValueError(f"{self.kind.value} must not define mutation effects")
        if EffectKind.FORMULA in self.effects:
            if self.formula_scope.empty:
                raise ValueError("formula mutations require a non-empty formula_scope")
            if not self.scope.covers(self.formula_scope):
                raise ValueError("mutation scope must cover formula_scope")
        elif not self.formula_scope.empty:
            raise ValueError("formula_scope requires a formula mutation effect")
        canonical_metadata = _canonical_json_bytes(self.metadata, context="event metadata")
        normalized_metadata = json.loads(canonical_metadata.decode("ascii"))
        if self.kind is EventKind.WORKBOOK_RECALCULATED:
            if not normalized_metadata:
                raise ValueError(
                    "workbook.recalculated requires portable recalculation metadata"
                )
            if set(normalized_metadata) != {"producer_tool", "calculation"}:
                raise ValueError(
                    "workbook.recalculated requires portable recalculation metadata only"
                )
            producer = normalized_metadata["producer_tool"]
            calculation = normalized_metadata["calculation"]
            if producer != "recalculate_and_read":
                raise ValueError(
                    "workbook.recalculated producer_tool must be recalculate_and_read"
                )
            if not isinstance(calculation, dict) or set(calculation) != set(
                _RECALCULATION_METADATA_KEYS
            ):
                raise ValueError(
                    "workbook.recalculated requires portable recalculation metadata only"
                )
            for key in ("backend", "version"):
                _validate_portable_label(
                    calculation[key],
                    context=f"recalculation {key}",
                )
            _validate_sha256(
                calculation["source_sha256"],
                context="recalculation source_sha256",
            )
            _validate_sha256(
                calculation["output_sha256"],
                context="recalculation output_sha256",
            )
            if calculation["source_sha256"] != self.revision_before:
                raise ValueError("recalculation source_sha256 must match revision_before")
            if calculation["output_sha256"] != (
                self.revision_after or self.revision_before
            ):
                raise ValueError("recalculation output_sha256 must match revision_after")
            if calculation["atomic_replace"] is not True:
                raise ValueError("recalculation atomic_replace must be true")
        elif self.kind is EventKind.WORKBOOK_RENDERED:
            _validate_render_metadata(normalized_metadata)
        elif self.kind is EventKind.RENDERED_PAGE_VIEWED:
            _validate_view_metadata(self, normalized_metadata)
        object.__setattr__(self, "metadata", normalized_metadata)

    @property
    def changed(self) -> bool:
        return bool(self.revision_after and self.revision_after != self.revision_before)


@dataclass(frozen=True)
class Witness:
    event_id: int
    kind: EventKind
    revision_sha256: str
    scope: EvidenceScope
    predicates: frozenset[str]
    render_id: str | None
    render_manifest_sha256: str | None
    page_id: str | None
    page_sha256: str | None
    page_file_sha256: str | None
    page_pixel_sha256: str | None
    pixel_sha256_algorithm: str | None
    width: int | None
    height: int | None
    image_mode: str | None

    def to_dict(self) -> dict[str, Any]:
        return {
            "event_id": self.event_id,
            "kind": self.kind.value,
            "revision_sha256": self.revision_sha256,
            "scope": self.scope.to_dict(),
            "predicates": sorted(self.predicates),
            "render_id": self.render_id,
            "render_manifest_sha256": self.render_manifest_sha256,
            "page_id": self.page_id,
            "page_sha256": self.page_sha256,
            "page_file_sha256": self.page_file_sha256,
            "page_pixel_sha256": self.page_pixel_sha256,
            "pixel_sha256_algorithm": self.pixel_sha256_algorithm,
            "width": self.width,
            "height": self.height,
            "image_mode": self.image_mode,
        }


@dataclass
class Obligation:
    id: str
    rule: ContractRule
    trigger_event_id: int
    trigger_revision_sha256: str
    required_scope: EvidenceScope
    next_step: int = 0
    witnesses: list[Witness] = field(default_factory=list)
    render_id: str | None = None
    render_manifest_sha256: str | None = None
    step_coverage: EvidenceScope = EvidenceScope()
    required_page_ids: tuple[str, ...] = ()
    viewed_page_ids: tuple[str, ...] = ()
    visual_coverage_policy: str | None = None
    render_pages: dict[str, dict[str, Any]] = field(default_factory=dict, repr=False)

    @property
    def complete(self) -> bool:
        return self.next_step == len(self.rule.requirements)

    @property
    def requirement(self) -> Requirement | None:
        if self.complete:
            return None
        return self.rule.requirements[self.next_step]

    def invalidate(self, revision_sha256: str) -> None:
        self.trigger_revision_sha256 = revision_sha256
        self.next_step = 0
        self.witnesses.clear()
        self.render_id = None
        self.render_manifest_sha256 = None
        self.step_coverage = EvidenceScope()
        self.required_page_ids = ()
        self.viewed_page_ids = ()
        self.visual_coverage_policy = None
        self.render_pages.clear()

    def to_dict(self) -> dict[str, Any]:
        requirement = self.requirement
        return {
            "id": self.id,
            "rule_id": self.rule.id,
            "trigger_event_id": self.trigger_event_id,
            "revision_sha256": self.trigger_revision_sha256,
            "required_scope": self.required_scope.to_dict(),
            "complete": self.complete,
            "next_step": self.next_step,
            "next_requirement": requirement.to_dict() if requirement is not None else None,
            "step_coverage": self.step_coverage.to_dict(),
            "required_page_ids": list(self.required_page_ids),
            "viewed_page_ids": list(self.viewed_page_ids),
            "visual_coverage_policy": self.visual_coverage_policy,
            "witnesses": [item.to_dict() for item in self.witnesses],
        }


@dataclass(frozen=True)
class SubmissionDecision:
    allowed: bool
    contract_satisfied: bool
    artifact_changed: bool
    mode: ContractMode
    reasons: tuple[str, ...]
    pending: tuple[dict[str, Any], ...]
    certificate: dict[str, Any] | None

    @property
    def enforcement_active(self) -> bool:
        return self.mode is ContractMode.ENFORCE

    @property
    def would_block(self) -> bool:
        return not self.contract_satisfied

    def to_dict(self) -> dict[str, Any]:
        return {
            "allowed": self.allowed,
            "contract_satisfied": self.contract_satisfied,
            "artifact_changed": self.artifact_changed,
            "mode": self.mode.value,
            "enforcement_active": self.enforcement_active,
            "would_block": self.would_block,
            "reasons": list(self.reasons),
            "pending": [dict(item) for item in self.pending],
            "certificate": self.certificate,
        }


class EvidenceContractMonitor:
    """Deterministic monitor over trusted workbook/tool events."""

    def __init__(
        self,
        spec: ContractSpec,
        initial_revision_sha256: str,
        *,
        mode: ContractMode = ContractMode.ENFORCE,
    ) -> None:
        if not initial_revision_sha256:
            raise ValueError("initial_revision_sha256 must not be empty")
        _validate_sha256(initial_revision_sha256, context="initial_revision_sha256")
        if not isinstance(spec, ContractSpec):
            raise TypeError("spec must be a ContractSpec")
        try:
            normalized_mode = ContractMode(mode)
        except (TypeError, ValueError) as exc:
            raise ValueError("mode must be 'shadow' or 'enforce'") from exc
        self.spec = spec
        self.mode = normalized_mode
        self.initial_revision_sha256 = initial_revision_sha256
        self.current_revision_sha256 = initial_revision_sha256
        self.revision_index = 0
        self._event_index = 0
        self._obligation_index = 0
        self._obligations: list[Obligation] = []
        self._events: list[dict[str, Any]] = []
        self._event_chain_sha256 = _EVENT_CHAIN_GENESIS_SHA256
        self._failures: list[dict[str, Any]] = []
        self._failure_scopes: dict[int, EvidenceScope] = {}
        self._has_committed_mutation = False

    @property
    def obligations(self) -> tuple[Obligation, ...]:
        return tuple(self._obligations)

    @property
    def pending(self) -> tuple[Obligation, ...]:
        return tuple(item for item in self._obligations if not item.complete)

    @property
    def failures(self) -> tuple[dict[str, Any], ...]:
        return tuple(dict(item) for item in self._failures)

    def _event_matches_trigger(self, event: EvidenceEvent, trigger: TriggerKind) -> bool:
        if trigger is TriggerKind.MUTATION_COMMITTED:
            return event.kind is EventKind.MUTATION_COMMITTED and event.changed
        if trigger is TriggerKind.FORMULA_CHANGED:
            return (
                event.kind is EventKind.MUTATION_COMMITTED
                and event.changed
                and EffectKind.FORMULA in event.effects
            )
        if trigger is TriggerKind.VISUAL_CHANGED:
            return (
                event.kind is EventKind.MUTATION_COMMITTED
                and event.changed
                and bool(event.effects & _VISUAL_EFFECTS)
            )
        raise AssertionError(f"Unhandled trigger: {trigger}")

    def _scope_for_rule(self, event: EvidenceEvent, rule: ContractRule) -> EvidenceScope:
        policies = {item.scope for item in rule.requirements if item.scope is not ScopePolicy.NONE}
        if not policies:
            return EvidenceScope()
        if len(policies) != 1:
            raise ContractStateError(
                f"Rule {rule.id!r} combines incompatible scope policies: {sorted(item.value for item in policies)}"
            )
        policy = next(iter(policies))
        if policy is ScopePolicy.CHANGED_CELLS_PLUS_BOUNDARY:
            if event.scope.empty:
                raise ContractStateError(
                    f"Rule {rule.id!r} cannot bind an empty mutation scope"
                )
            return event.scope.expand(1)
        if policy is ScopePolicy.CHANGED_FORMULA_CELLS_PLUS_BOUNDARY:
            source = event.formula_scope if not event.formula_scope.empty else event.scope
            if source.empty:
                raise ContractStateError(
                    f"Rule {rule.id!r} cannot bind an empty formula scope"
                )
            return source.expand(1)
        if policy is ScopePolicy.CHANGED_VISUAL_SCOPE:
            if event.scope.empty:
                raise ContractStateError(
                    f"Rule {rule.id!r} cannot bind an empty visual scope"
                )
            return event.scope
        raise AssertionError(f"Unhandled scope policy: {policy}")

    def _invalidate_for_revision(self, revision_sha256: str) -> None:
        for obligation in self._obligations:
            obligation.invalidate(revision_sha256)

    def _append_obligation(
        self,
        rule: ContractRule,
        *,
        event_id: int,
        event: EvidenceEvent,
    ) -> None:
        self._obligation_index += 1
        self._obligations.append(
            Obligation(
                id=f"{rule.id}:{self._obligation_index}",
                rule=rule,
                trigger_event_id=event_id,
                trigger_revision_sha256=self.current_revision_sha256,
                required_scope=self._scope_for_rule(event, rule),
            )
        )

    @staticmethod
    def _required_visual_pages(
        required_scope: EvidenceScope,
        render_metadata: Mapping[str, Any],
    ) -> tuple[tuple[str, ...], str, dict[str, dict[str, Any]]]:
        raw_pages = render_metadata["pages"]
        assert isinstance(raw_pages, list) and raw_pages
        pages = {
            str(page["page_id"]): {
                **dict(page),
                "_render_mode": render_metadata["mode"],
            }
            for page in raw_pages
            if isinstance(page, Mapping)
        }
        all_page_ids = tuple(sorted(pages))

        # Exact cell mappings may select a strict subset, but only when their
        # authenticated union covers every changed range.
        if not required_scope.wildcard and not required_scope.sheets:
            mapped: list[tuple[str, EvidenceScope]] = []
            for page_id, page in pages.items():
                raw_scope = page.get("cell_scope")
                if isinstance(raw_scope, Mapping):
                    page_scope = EvidenceScope.from_dict(raw_scope)
                    if page_scope.intersects(required_scope):
                        mapped.append((page_id, page_scope))
            if mapped:
                coverage = EvidenceScope()
                for _, page_scope in mapped:
                    coverage = coverage.merged(page_scope)
                if coverage.covers(required_scope):
                    return (
                        tuple(sorted(page_id for page_id, _ in mapped)),
                        "authenticated_cell_mapping",
                        pages,
                    )

        affected_sheets = {
            *required_scope.sheets,
            *(item.sheet for item in required_scope.ranges),
        }
        if render_metadata["mode"] == "per_sheet" and affected_sheets:
            represented_sheets = {
                str(page["sheet"])
                for page in pages.values()
                if isinstance(page.get("sheet"), str)
            }
            selected = tuple(
                sorted(
                    page_id
                    for page_id, page in pages.items()
                    if page.get("sheet") in affected_sheets
                )
            )
            if affected_sheets <= represented_sheets and selected:
                return selected, "all_pages_for_affected_sheets", pages
            missing = tuple(
                f"<missing-sheet>:{sheet}"
                for sheet in sorted(affected_sheets - represented_sheets)
            )
            if missing:
                # Keep the obligation visibly unsatisfied. Viewing unrelated
                # rendered pages cannot prove a sheet omitted by the render.
                return (
                    tuple(sorted((*selected, *missing))),
                    "missing_affected_sheet_pages",
                    pages,
                )

        return all_page_ids, "all_rendered_pages", pages

    @staticmethod
    def _view_matches_bound_page(
        obligation: Obligation,
        event: EvidenceEvent,
    ) -> bool:
        if event.page_id is None or event.page_id not in obligation.render_pages:
            return False
        page = obligation.render_pages[event.page_id]
        metadata = event.metadata
        expected_scope = page.get("cell_scope")
        return bool(
            event.page_sha256 == page.get("file_sha256")
            and metadata.get("page_file_sha256") == page.get("file_sha256")
            and metadata.get("page_index") == page.get("page_index")
            and metadata.get("width") == page.get("width")
            and metadata.get("height") == page.get("height")
            and metadata.get("sheet") == page.get("sheet")
            and metadata.get("sheet_page") == page.get("sheet_page")
            and metadata.get("render_mode") == page.get("_render_mode")
            and metadata.get("cell_scope") == expected_scope
            and (
                event.scope.empty
                if expected_scope is None
                else event.scope == EvidenceScope.from_dict(expected_scope)
            )
        )

    def _requirement_matches(
        self,
        obligation: Obligation,
        requirement: Requirement,
        event: EvidenceEvent,
        *,
        check_scope: bool = True,
    ) -> bool:
        if event.kind is not requirement.event:
            return False
        observed_revision = event.revision_after or event.revision_before
        if observed_revision != self.current_revision_sha256:
            return False
        if requirement.artifact is ArtifactConstraint.SAME_RENDER:
            if (
                not obligation.render_id
                or not obligation.render_manifest_sha256
                or event.related_render_id != obligation.render_id
                or event.related_render_manifest_sha256
                != obligation.render_manifest_sha256
            ):
                return False
        if requirement.artifact is ArtifactConstraint.RECALCULATED_REVISION:
            if not obligation.witnesses or not any(
                witness.kind is EventKind.WORKBOOK_RECALCULATED
                for witness in obligation.witnesses
            ):
                return False
        if (
            check_scope
            and requirement.scope is not ScopePolicy.NONE
            and event.kind is not EventKind.RENDERED_PAGE_VIEWED
            and not event.scope.covers(obligation.required_scope)
        ):
            return False
        if not requirement.predicates.issubset(event.predicates):
            return False
        return True

    def _apply_evidence(self, event_id: int, event: EvidenceEvent) -> None:
        observed_revision = event.revision_after or event.revision_before
        for obligation in self._obligations:
            if obligation.complete:
                continue
            requirement = obligation.requirement
            assert requirement is not None
            if not self._requirement_matches(
                obligation,
                requirement,
                event,
                check_scope=False,
            ):
                first = obligation.rule.requirements[0]
                if obligation.next_step and self._requirement_matches(
                    obligation,
                    first,
                    event,
                    check_scope=False,
                ):
                    obligation.next_step = 0
                    obligation.witnesses.clear()
                    obligation.render_id = None
                    obligation.render_manifest_sha256 = None
                    obligation.step_coverage = EvidenceScope()
                    obligation.required_page_ids = ()
                    obligation.viewed_page_ids = ()
                    obligation.visual_coverage_policy = None
                    obligation.render_pages.clear()
                    requirement = first
                else:
                    continue
            if event.kind is EventKind.RENDERED_PAGE_VIEWED:
                if (
                    event.page_id not in obligation.required_page_ids
                    or event.page_id in obligation.viewed_page_ids
                    or not self._view_matches_bound_page(obligation, event)
                ):
                    continue
            witness_render_id = event.render_id or event.related_render_id
            witness_render_manifest = (
                event.render_manifest_sha256 or event.related_render_manifest_sha256
            )
            obligation.witnesses.append(
                Witness(
                    event_id=event_id,
                    kind=event.kind,
                    revision_sha256=observed_revision,
                    scope=event.scope,
                    predicates=event.predicates,
                    render_id=witness_render_id,
                    render_manifest_sha256=witness_render_manifest,
                    page_id=event.page_id,
                    page_sha256=event.page_sha256,
                    page_file_sha256=(
                        event.metadata.get("page_file_sha256")
                        if event.kind is EventKind.RENDERED_PAGE_VIEWED
                        else None
                    ),
                    page_pixel_sha256=(
                        event.metadata.get("page_pixel_sha256")
                        if event.kind is EventKind.RENDERED_PAGE_VIEWED
                        else None
                    ),
                    pixel_sha256_algorithm=(
                        event.metadata.get("pixel_sha256_algorithm")
                        if event.kind is EventKind.RENDERED_PAGE_VIEWED
                        else None
                    ),
                    width=(
                        event.metadata.get("width")
                        if event.kind is EventKind.RENDERED_PAGE_VIEWED
                        else None
                    ),
                    height=(
                        event.metadata.get("height")
                        if event.kind is EventKind.RENDERED_PAGE_VIEWED
                        else None
                    ),
                    image_mode=(
                        event.metadata.get("image_mode")
                        if event.kind is EventKind.RENDERED_PAGE_VIEWED
                        else None
                    ),
                )
            )
            if event.kind is EventKind.RENDERED_PAGE_VIEWED:
                assert event.page_id is not None
                obligation.viewed_page_ids = tuple(
                    sorted({*obligation.viewed_page_ids, event.page_id})
                )
                if not set(obligation.required_page_ids) <= set(
                    obligation.viewed_page_ids
                ):
                    continue
            elif requirement.scope is not ScopePolicy.NONE:
                obligation.step_coverage = obligation.step_coverage.merged(event.scope)
                if not obligation.step_coverage.covers(obligation.required_scope):
                    continue
            if event.kind is EventKind.WORKBOOK_RENDERED:
                obligation.render_id = event.render_id
                obligation.render_manifest_sha256 = event.render_manifest_sha256
                (
                    obligation.required_page_ids,
                    obligation.visual_coverage_policy,
                    obligation.render_pages,
                ) = self._required_visual_pages(obligation.required_scope, event.metadata)
            obligation.next_step += 1
            obligation.step_coverage = EvidenceScope()

    def observe(self, event: EvidenceEvent) -> dict[str, Any]:
        """Advance the monitor and return bounded, model-facing state."""

        if not isinstance(event, EvidenceEvent):
            raise TypeError("event must be an EvidenceEvent")
        if event.revision_before != self.current_revision_sha256:
            raise ContractStateError(
                "Event revision does not match current artifact: "
                f"expected {self.current_revision_sha256}, got {event.revision_before}"
            )
        triggered_rules = [
            rule
            for rule in self.spec.rules
            if self._event_matches_trigger(event, rule.trigger)
        ]
        # Validate every derived scope before changing monitor state.
        for rule in triggered_rules:
            self._scope_for_rule(event, rule)
        self._event_index += 1
        event_id = self._event_index
        revision_changed = bool(
            event.revision_after and event.revision_after != self.current_revision_sha256
        )
        if revision_changed:
            self.current_revision_sha256 = str(event.revision_after)
            self.revision_index += 1
            self._invalidate_for_revision(self.current_revision_sha256)

        if event.kind is EventKind.MUTATION_ROLLED_BACK:
            failure = {
                "event_id": event_id,
                "kind": event.kind.value,
                "revision_sha256": self.current_revision_sha256,
                "scope": event.scope.to_dict(),
            }
            self._failures.append(failure)
            self._failure_scopes[event_id] = event.scope
        elif event.kind is EventKind.MUTATION_COMMITTED and event.changed:
            self._has_committed_mutation = True
            self._failures = [
                failure
                for failure in self._failures
                if not event.scope.covers(self._failure_scopes[int(failure["event_id"])])
            ]
            active_failure_ids = {int(item["event_id"]) for item in self._failures}
            self._failure_scopes = {
                failure_id: scope
                for failure_id, scope in self._failure_scopes.items()
                if failure_id in active_failure_ids
            }

        for rule in triggered_rules:
            self._append_obligation(rule, event_id=event_id, event=event)

        self._apply_evidence(event_id, event)
        event_record = {
            "event_id": event_id,
            "kind": event.kind.value,
            "revision_before": event.revision_before,
            "revision_after": event.revision_after,
            "effects": sorted(item.value for item in event.effects),
            "scope": event.scope.to_dict(),
            "formula_scope": event.formula_scope.to_dict(),
            "predicates": sorted(event.predicates),
            "render_id": event.render_id,
            "render_manifest_sha256": event.render_manifest_sha256,
            "related_render_id": event.related_render_id,
            "related_render_manifest_sha256": event.related_render_manifest_sha256,
            "page_id": event.page_id,
            "page_sha256": event.page_sha256,
            "metadata": dict(event.metadata),
        }
        event_bytes = _canonical_json_bytes(event_record, context="evidence event")
        self._event_chain_sha256 = hashlib.sha256(
            bytes.fromhex(self._event_chain_sha256) + event_bytes
        ).hexdigest()
        self._events.append(
            {**event_record, "event_chain_sha256": self._event_chain_sha256}
        )
        return self.status()

    def status(self) -> dict[str, Any]:
        pending = [item.to_dict() for item in self.pending]
        artifact_changed = bool(
            self._has_committed_mutation
            and self.current_revision_sha256 != self.initial_revision_sha256
        )
        evidence_satisfied = not pending and not self._failures
        contract_satisfied = artifact_changed and evidence_satisfied
        submission_allowed = contract_satisfied or self.mode is ContractMode.SHADOW
        return {
            "mode": self.mode.value,
            "enforcement_active": self.mode is ContractMode.ENFORCE,
            "contract_sha256": self.spec.canonical_sha256,
            "initial_revision_sha256": self.initial_revision_sha256,
            "current_revision_sha256": self.current_revision_sha256,
            "revision_index": self.revision_index,
            "event_count": self._event_index,
            "event_chain_sha256": self._event_chain_sha256,
            "triggered_obligations": len(self._obligations),
            "has_committed_mutation": self._has_committed_mutation,
            "pending_obligations": pending,
            "unresolved_failures": [dict(item) for item in self._failures],
            "artifact_changed": artifact_changed,
            "evidence_satisfied": evidence_satisfied,
            "contract_satisfied": contract_satisfied,
            "would_block": not contract_satisfied,
            "submission_allowed": submission_allowed,
            # Retained for compatibility; this means contract-satisfied, not
            # mode-dependent permission to submit.
            "submission_ready": contract_satisfied,
        }

    def compact_status(self, *, max_pending: int = 8) -> dict[str, Any]:
        if max_pending < 1:
            raise ValueError("max_pending must be positive")
        full = self.status()
        pending = self.pending
        summaries: list[dict[str, Any]] = []
        for obligation in pending[:max_pending]:
            requirement = obligation.requirement
            assert requirement is not None
            summaries.append(
                {
                    "id": obligation.id,
                    "rule_id": obligation.rule.id,
                    "next_event": requirement.event.value,
                    "required_scope": obligation.required_scope.to_dict(),
                    "required_predicates": sorted(requirement.predicates),
                    "visual_coverage_policy": obligation.visual_coverage_policy,
                    "required_page_ids": list(obligation.required_page_ids),
                    "viewed_page_ids": list(obligation.viewed_page_ids),
                }
            )
        return {
            "mode": self.mode.value,
            "enforcement_active": full["enforcement_active"],
            "contract_sha256": self.spec.canonical_sha256,
            "artifact_revision_sha256": self.current_revision_sha256,
            "artifact_changed": full["artifact_changed"],
            "evidence_satisfied": full["evidence_satisfied"],
            "contract_satisfied": full["contract_satisfied"],
            "would_block": full["would_block"],
            "submission_allowed": full["submission_allowed"],
            "submission_ready": full["submission_ready"],
            "pending_count": len(pending),
            "pending_returned": len(summaries),
            "pending_truncated": len(summaries) < len(pending),
            "pending": summaries,
            "unresolved_failure_count": len(self._failures),
            "next_required_event": (
                self.next_required_event().value
                if self.next_required_event() is not None
                else None
            ),
            "minimum_evidence_calls": self.minimum_evidence_calls(),
        }

    def next_required_event(self) -> EventKind | None:
        for obligation in self.pending:
            requirement = obligation.requirement
            if requirement is not None:
                return requirement.event
        return None

    def minimum_evidence_calls(self) -> int:
        """Return a conservative lower bound for budget-aware routing."""

        if not self.pending:
            return 0
        def remaining(item: Obligation) -> int:
            requirement = item.requirement
            base = len(item.rule.requirements) - item.next_step
            if (
                requirement is not None
                and requirement.event is EventKind.RENDERED_PAGE_VIEWED
                and item.required_page_ids
            ):
                outstanding = len(
                    set(item.required_page_ids) - set(item.viewed_page_ids)
                )
                return outstanding + max(0, base - 1)
            return base

        # One event can satisfy multiple obligations, so use the longest remaining sequence.
        return max(remaining(item) for item in self.pending)

    def certificate(self) -> dict[str, Any]:
        if self.pending or self._failures:
            raise ContractStateError("Cannot issue a certificate with unresolved obligations")
        artifact_changed = bool(
            self._has_committed_mutation
            and self.current_revision_sha256 != self.initial_revision_sha256
        )
        if not artifact_changed:
            raise ContractStateError("Cannot issue a certificate for an unchanged artifact")
        payload = {
            "schema_version": "spreadsheet-evidence-certificate-v1",
            "certificate_digest_algorithm": _CERTIFICATE_DIGEST_ALGORITHM,
            "contract": self.spec.identity_dict(),
            "initial_revision_sha256": self.initial_revision_sha256,
            "accepted_revision_sha256": self.current_revision_sha256,
            "revision_index": self.revision_index,
            "event_count": self._event_index,
            "event_chain_algorithm": _EVENT_CHAIN_ALGORITHM,
            "event_chain_genesis_sha256": _EVENT_CHAIN_GENESIS_SHA256,
            "event_chain_sha256": self._event_chain_sha256,
            "events": [dict(item) for item in self._events],
            "obligations": [item.to_dict() for item in self._obligations],
        }
        digest_payload = _canonical_json_bytes(payload, context="evidence certificate")
        return {**payload, "certificate_sha256": hashlib.sha256(digest_payload).hexdigest()}

    def submission_decision(self) -> SubmissionDecision:
        artifact_changed = bool(
            self._has_committed_mutation
            and self.current_revision_sha256 != self.initial_revision_sha256
        )
        evidence_satisfied = not self.pending and not self._failures
        reasons: list[str] = []
        if not artifact_changed:
            reasons.append("artifact_unchanged")
        if self.pending:
            reasons.append("pending_evidence_obligations")
        if self._failures:
            reasons.append("unresolved_mutation_failure")
        contract_satisfied = artifact_changed and evidence_satisfied
        allowed = contract_satisfied or self.mode is ContractMode.SHADOW
        certificate = self.certificate() if contract_satisfied else None
        return SubmissionDecision(
            allowed=allowed,
            contract_satisfied=contract_satisfied,
            artifact_changed=artifact_changed,
            mode=self.mode,
            reasons=tuple(reasons),
            pending=tuple(item.to_dict() for item in self.pending),
            certificate=certificate,
        )


def _certificate_contract_spec(value: Any) -> ContractSpec:
    if not isinstance(value, Mapping):
        raise ContractValidationError("certificate contract must be a mapping")
    _require_keys(
        value,
        required=frozenset(
            {"schema_version", "canonical_sha256", "source_sha256", "rules"}
        ),
        optional=frozenset(),
        context="certificate contract",
    )
    source_sha256 = value["source_sha256"]
    if source_sha256 is not None:
        try:
            _validate_sha256(source_sha256, context="certificate contract source_sha256")
        except ValueError as exc:
            raise ContractValidationError(str(exc)) from exc
    try:
        _validate_sha256(
            value["canonical_sha256"],
            context="certificate contract canonical_sha256",
        )
    except ValueError as exc:
        raise ContractValidationError(str(exc)) from exc
    raw_rules = value["rules"]
    if not isinstance(raw_rules, list) or not raw_rules:
        raise ContractValidationError("certificate contract rules must be a non-empty list")
    input_rules: list[dict[str, Any]] = []
    for index, raw_rule in enumerate(raw_rules):
        if not isinstance(raw_rule, Mapping):
            raise ContractValidationError(
                f"certificate contract rules[{index}] must be a mapping"
            )
        _require_keys(
            raw_rule,
            required=frozenset({"id", "trigger", "requirements"}),
            optional=frozenset(),
            context=f"certificate contract rules[{index}]",
        )
        requirements = raw_rule["requirements"]
        if not isinstance(requirements, list) or not requirements:
            raise ContractValidationError(
                f"certificate contract rules[{index}].requirements must be non-empty"
            )
        input_rules.append(
            {
                "id": raw_rule["id"],
                "trigger": raw_rule["trigger"],
                "require_sequence": requirements,
            }
        )
    parsed = ContractSpec.from_mapping(
        {"schema_version": value["schema_version"], "rules": input_rules}
    )
    if parsed.canonical_sha256 != value["canonical_sha256"]:
        raise ContractValidationError("certificate contract canonical_sha256 mismatch")
    return ContractSpec(
        schema_version=parsed.schema_version,
        rules=parsed.rules,
        canonical_sha256=parsed.canonical_sha256,
        source_sha256=source_sha256,
    )


def _certificate_event(value: Any, *, index: int) -> EvidenceEvent:
    context = f"certificate events[{index}]"
    if not isinstance(value, Mapping):
        raise ContractValidationError(f"{context} must be a mapping")
    _require_keys(
        value,
        required=frozenset(
            {
                "event_id",
                "kind",
                "revision_before",
                "revision_after",
                "effects",
                "scope",
                "formula_scope",
                "predicates",
                "render_id",
                "render_manifest_sha256",
                "related_render_id",
                "related_render_manifest_sha256",
                "page_id",
                "page_sha256",
                "metadata",
                "event_chain_sha256",
            }
        ),
        optional=frozenset(),
        context=context,
    )
    if type(value["event_id"]) is not int or value["event_id"] != index + 1:
        raise ContractValidationError(f"{context}.event_id must be contiguous and 1-based")
    raw_effects = value["effects"]
    raw_predicates = value["predicates"]
    if not isinstance(raw_effects, list) or not all(
        isinstance(item, str) for item in raw_effects
    ):
        raise ContractValidationError(f"{context}.effects must be a string list")
    if len(raw_effects) != len(set(raw_effects)):
        raise ContractValidationError(f"{context}.effects must not contain duplicates")
    if not isinstance(raw_predicates, list) or not all(
        isinstance(item, str) for item in raw_predicates
    ):
        raise ContractValidationError(f"{context}.predicates must be a string list")
    if len(raw_predicates) != len(set(raw_predicates)):
        raise ContractValidationError(f"{context}.predicates must not contain duplicates")
    try:
        event = EvidenceEvent(
            kind=EventKind(value["kind"]),
            revision_before=value["revision_before"],
            revision_after=value["revision_after"],
            effects=frozenset(EffectKind(item) for item in raw_effects),
            scope=EvidenceScope.from_dict(value["scope"]),
            formula_scope=EvidenceScope.from_dict(value["formula_scope"]),
            predicates=frozenset(raw_predicates),
            render_id=value["render_id"],
            render_manifest_sha256=value["render_manifest_sha256"],
            related_render_id=value["related_render_id"],
            related_render_manifest_sha256=value[
                "related_render_manifest_sha256"
            ],
            page_id=value["page_id"],
            page_sha256=value["page_sha256"],
            metadata=value["metadata"],
        )
        _validate_sha256(
            value["event_chain_sha256"],
            context=f"{context}.event_chain_sha256",
        )
    except (TypeError, ValueError, KeyError) as exc:
        raise ContractValidationError(f"{context} is invalid: {exc}") from exc
    return event


def audit_evidence_certificate(certificate: Mapping[str, Any]) -> dict[str, Any]:
    """Validate a certificate from embedded rules and events, without local run state."""

    try:
        normalized = json.loads(
            _canonical_json_bytes(certificate, context="evidence certificate").decode(
                "ascii"
            )
        )
    except (TypeError, ValueError) as exc:
        raise ContractValidationError(f"Invalid evidence certificate: {exc}") from exc
    _require_keys(
        normalized,
        required=frozenset(
            {
                "schema_version",
                "certificate_digest_algorithm",
                "contract",
                "initial_revision_sha256",
                "accepted_revision_sha256",
                "revision_index",
                "event_count",
                "event_chain_algorithm",
                "event_chain_genesis_sha256",
                "event_chain_sha256",
                "events",
                "obligations",
                "certificate_sha256",
            }
        ),
        optional=frozenset(),
        context="evidence certificate",
    )
    if normalized["schema_version"] != "spreadsheet-evidence-certificate-v1":
        raise ContractValidationError("Unsupported evidence certificate schema_version")
    if normalized["certificate_digest_algorithm"] != _CERTIFICATE_DIGEST_ALGORITHM:
        raise ContractValidationError("Unsupported certificate digest algorithm")
    if normalized["event_chain_algorithm"] != _EVENT_CHAIN_ALGORITHM:
        raise ContractValidationError("Unsupported event chain algorithm")
    if normalized["event_chain_genesis_sha256"] != _EVENT_CHAIN_GENESIS_SHA256:
        raise ContractValidationError("Invalid event chain genesis")
    for key in (
        "initial_revision_sha256",
        "accepted_revision_sha256",
        "event_chain_sha256",
        "certificate_sha256",
    ):
        try:
            _validate_sha256(normalized[key], context=f"certificate {key}")
        except ValueError as exc:
            raise ContractValidationError(str(exc)) from exc
    if type(normalized["revision_index"]) is not int or normalized["revision_index"] < 0:
        raise ContractValidationError("certificate revision_index must be non-negative")
    if type(normalized["event_count"]) is not int or normalized["event_count"] < 0:
        raise ContractValidationError("certificate event_count must be non-negative")
    if not isinstance(normalized["events"], list):
        raise ContractValidationError("certificate events must be a list")
    if normalized["event_count"] != len(normalized["events"]):
        raise ContractValidationError("certificate event_count does not match events")
    if not isinstance(normalized["obligations"], list):
        raise ContractValidationError("certificate obligations must be a list")

    digest_payload = {
        key: value for key, value in normalized.items() if key != "certificate_sha256"
    }
    expected_certificate_sha256 = hashlib.sha256(
        _canonical_json_bytes(digest_payload, context="evidence certificate payload")
    ).hexdigest()
    if normalized["certificate_sha256"] != expected_certificate_sha256:
        raise ContractValidationError("Evidence certificate digest mismatch")

    spec = _certificate_contract_spec(normalized["contract"])
    previous_chain = _EVENT_CHAIN_GENESIS_SHA256
    events: list[EvidenceEvent] = []
    for index, raw_event in enumerate(normalized["events"]):
        event = _certificate_event(raw_event, index=index)
        event_payload = {
            key: value
            for key, value in raw_event.items()
            if key != "event_chain_sha256"
        }
        expected_chain = hashlib.sha256(
            bytes.fromhex(previous_chain)
            + _canonical_json_bytes(event_payload, context=f"certificate events[{index}]")
        ).hexdigest()
        if raw_event["event_chain_sha256"] != expected_chain:
            raise ContractValidationError(
                f"Evidence event chain mismatch at event {index + 1}"
            )
        previous_chain = expected_chain
        events.append(event)
    if normalized["event_chain_sha256"] != previous_chain:
        raise ContractValidationError("Evidence event chain head mismatch")

    try:
        replay = EvidenceContractMonitor(spec, normalized["initial_revision_sha256"])
        for event in events:
            replay.observe(event)
        replayed_certificate = replay.certificate()
    except (ContractStateError, TypeError, ValueError) as exc:
        raise ContractValidationError(f"Certificate replay failed: {exc}") from exc
    if replayed_certificate != normalized:
        raise ContractValidationError(
            "Certificate replay does not reproduce its obligations or witnesses"
        )
    return {
        "valid": True,
        "certificate_sha256": normalized["certificate_sha256"],
        "accepted_revision_sha256": normalized["accepted_revision_sha256"],
        "revision_index": normalized["revision_index"],
        "event_count": normalized["event_count"],
        "obligation_count": len(normalized["obligations"]),
    }


__all__ = [
    "ArtifactConstraint",
    "ArtifactRef",
    "ArtifactTransition",
    "CellRange",
    "ContractMode",
    "ContractSpec",
    "ContractStateError",
    "ContractValidationError",
    "EffectKind",
    "EventKind",
    "EvidenceContractMonitor",
    "EvidenceEvent",
    "EvidenceScope",
    "PIXEL_SHA256_ALGORITHM",
    "ScopePolicy",
    "SubmissionDecision",
    "TriggerKind",
    "audit_evidence_certificate",
]
