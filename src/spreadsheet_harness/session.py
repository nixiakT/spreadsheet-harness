"""Transactional workbook workspace and spreadsheet operations."""

from __future__ import annotations

import hashlib
import json
import re
import shutil
import tempfile
import threading
import uuid
from collections.abc import Callable, Iterable, Iterator
from contextlib import contextmanager
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.styles import Border, PatternFill, Side
from openpyxl.styles.cell_style import StyleArray
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple, range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from .code_interpreter import validate_formula_transaction
from .completion_attempt import CompletionAttemptLedger, CompletionAttemptRecord
from .errors import ToolInputError, WorkbookValidationError
from .evidence_contract import ArtifactRef, ArtifactTransition, EvidenceScope
from .target_grounding import (
    AdvisoryLifecycleEvent,
    CommittedAdvisoryTargetAssessment,
    CommittedTargetAuthorization,
    PreparedAdvisoryTargetAssessment,
    PreparedTargetAuthorization,
    TargetGroundingError,
    TargetGroundingMode,
    TargetGroundingRejected,
    TargetGroundingStateMachine,
    is_target_grounding_protected_transition_kind,
)
from .trajectory import TrajectoryRecorder
from .workbook_diff import WorkbookEffectDiff, diff_workbooks

SUPPORTED_EDIT_FORMATS = {".xlsx", ".xlsm"}
_SAFE_NAME = re.compile(r"[^A-Za-z0-9_.-]+")
_FORMULA_RANGE_RE = re.compile(
    r"(?P<sheet>(?:'[^']+'|[A-Za-z_][A-Za-z0-9_ .]*)!)?"
    r"(?P<start>\$?[A-Za-z]{1,3}\$?\d+):(?P<end>\$?[A-Za-z]{1,3}\$?\d+)"
)
_CELL_REF_RE = re.compile(r"(?P<col_abs>\$?)(?P<col>[A-Za-z]{1,3})(?P<row_abs>\$?)(?P<row>\d+)\Z")


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _json_value(value: Any) -> Any:
    if value is None or isinstance(value, bool | int | float | str):
        return value
    if isinstance(value, datetime | date | time):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return str(value)


def _cell_ref_parts(ref: str) -> dict[str, Any] | None:
    match = _CELL_REF_RE.fullmatch(ref)
    if match is None:
        return None
    return {
        "column_absolute": bool(match.group("col_abs")),
        "row_absolute": bool(match.group("row_abs")),
    }


def _formula_sample_coordinates(
    source_cell: str,
    bounds: tuple[int, int, int, int],
) -> list[str]:
    min_col, min_row, max_col, max_row = bounds
    candidates = [
        source_cell.replace("$", ""),
        f"{get_column_letter(min_col)}{min_row}",
        f"{get_column_letter(min(min_col + 1, max_col))}{min_row}",
        f"{get_column_letter(min_col)}{min(min_row + 1, max_row)}",
        f"{get_column_letter(max_col)}{max_row}",
    ]
    return list(dict.fromkeys(candidates))


def _normalize_fill_target_range(
    source_cell: str,
    target_range: str,
) -> tuple[str, bool]:
    source = source_cell.replace("$", "")
    target = target_range.replace("$", "")
    if ":" in target:
        return target_range, False
    try:
        range_boundaries(target)
        range_boundaries(source)
    except (TypeError, ValueError):
        return target_range, False
    if target.upper() == source.upper():
        return target_range, False
    return f"{source}:{target}", True


def _fill_formula_warnings(
    source_formula: str,
    source_cell: str,
    bounds: tuple[int, int, int, int],
    samples: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    min_col, min_row, max_col, max_row = bounds
    fills_horizontally = max_col > min_col
    fills_vertically = max_row > min_row
    if not fills_horizontally and not fills_vertically:
        return []

    sample_cells = [
        str(sample["cell"])
        for sample in samples
        if sample.get("cell") != source_cell.replace("$", "")
    ]
    warnings: list[dict[str, Any]] = []
    for match in _FORMULA_RANGE_RE.finditer(source_formula):
        start = _cell_ref_parts(match.group("start"))
        end = _cell_ref_parts(match.group("end"))
        if start is None or end is None:
            continue
        issues: list[str] = []
        if fills_horizontally and not (start["column_absolute"] and end["column_absolute"]):
            issues.append("column endpoints are not both absolute")
        if fills_vertically and start["row_absolute"] != end["row_absolute"]:
            issues.append("mixed row anchors")
        if not issues:
            continue

        translated_examples: list[dict[str, str]] = []
        for destination in sample_cells:
            translated = Translator(
                "=" + match.group(0),
                origin=source_cell,
            ).translate_formula(destination)[1:]
            if translated != match.group(0):
                translated_examples.append({"cell": destination, "translated_range": translated})
            if len(translated_examples) >= 3:
                break
        if not translated_examples:
            continue
        warnings.append(
            {
                "type": "possible_expanding_or_drifting_range",
                "source_range": match.group(0),
                "issues": issues,
                "examples": translated_examples,
                "message": (
                    "This range changes during fill_formula. If the range should stay "
                    "fixed across the fill direction, lock both endpoints, e.g. use "
                    "$E6:$G6 instead of E6:G6 or $E6:G6, then refill and verify cached "
                    "values."
                ),
            }
        )
    return warnings


def _color_value(color: Any) -> str | None:
    if color is None:
        return None
    value = getattr(color, "rgb", None)
    if value and value not in {"00000000", "000000"}:
        return str(value)
    indexed = getattr(color, "indexed", None)
    if indexed is not None:
        return f"indexed:{indexed}"
    theme = getattr(color, "theme", None)
    if theme is not None:
        return f"theme:{theme}"
    return None


def _normalize_color(value: str) -> str:
    cleaned = value.strip().lstrip("#").upper()
    if not re.fullmatch(r"[0-9A-F]{6}|[0-9A-F]{8}", cleaned):
        raise ToolInputError(f"Invalid RGB/ARGB color: {value!r}")
    return cleaned if len(cleaned) == 8 else "FF" + cleaned


def _intersects(left: tuple[int, int, int, int], right: tuple[int, int, int, int]) -> bool:
    l_min_col, l_min_row, l_max_col, l_max_row = left
    r_min_col, r_min_row, r_max_col, r_max_row = right
    return not (
        l_max_col < r_min_col
        or r_max_col < l_min_col
        or l_max_row < r_min_row
        or r_max_row < l_min_row
    )


def _table_ref(table: Any) -> str:
    """Return an openpyxl table range across TableList API variants."""
    ref = getattr(table, "ref", table)
    if not isinstance(ref, str):
        raise ToolInputError(f"Workbook table has invalid ref: {ref!r}")
    return ref


@dataclass(frozen=True)
class SessionPaths:
    root: Path
    input: Path
    workbook: Path
    snapshots: Path
    artifacts: Path
    trajectory: Path


class WorkbookSession:
    """Own one isolated workbook copy and apply atomic, auditable mutations."""

    def __init__(
        self,
        paths: SessionPaths,
        run_id: str,
        *,
        recorder_secrets: tuple[str, ...] = (),
    ) -> None:
        self.paths = paths
        self.run_id = run_id
        self._write_lock = threading.RLock()
        self._snapshot_counter = 0
        self._artifact = ArtifactRef(0, _sha256(paths.workbook))
        self._artifact_bytes = paths.workbook.read_bytes()
        self._artifact_transitions: list[ArtifactTransition] = []
        self._completion_attempt_ledger: CompletionAttemptLedger | None = None
        self._target_grounding: TargetGroundingStateMachine | None = None
        self._target_grounding_mode = TargetGroundingMode.OFF
        self._target_grounding_initial_artifact: ArtifactRef | None = None
        self._target_grounding_initial_transition_count: int | None = None
        self.recorder = TrajectoryRecorder(
            paths.trajectory,
            run_id,
            secrets=recorder_secrets,
        )

    @classmethod
    def create(
        cls,
        source: str | Path,
        run_dir: str | Path,
        *,
        run_id: str | None = None,
        recorder_secrets: tuple[str, ...] = (),
    ) -> WorkbookSession:
        source_path = Path(source).expanduser().resolve(strict=True)
        if source_path.suffix.lower() not in SUPPORTED_EDIT_FORMATS:
            raise ToolInputError(
                f"Editing requires .xlsx or .xlsm, got {source_path.suffix}. "
                "Normalize legacy/ODS/CSV input with preprocess first."
            )
        root = Path(run_dir).expanduser().resolve()
        if root.exists() and any(root.iterdir()):
            raise FileExistsError(f"Run directory is not empty: {root}")
        root.mkdir(parents=True, exist_ok=True)
        input_dir = root / "input"
        artifacts = root / "artifacts"
        snapshots = root / "snapshots"
        for directory in (input_dir, artifacts, snapshots):
            directory.mkdir(mode=0o700)
        safe_name = _SAFE_NAME.sub("_", source_path.name)
        input_copy = input_dir / safe_name
        workbook_copy = artifacts / f"output{source_path.suffix.lower()}"
        shutil.copy2(source_path, input_copy)
        shutil.copy2(source_path, workbook_copy)
        paths = SessionPaths(
            root=root,
            input=input_copy,
            workbook=workbook_copy,
            snapshots=snapshots,
            artifacts=artifacts,
            trajectory=root / "trajectory.jsonl",
        )
        session = cls(
            paths,
            run_id or root.name or uuid.uuid4().hex,
            recorder_secrets=recorder_secrets,
        )
        session._validate(workbook_copy)
        session.recorder.record(
            "session.created",
            {
                "source": str(source_path),
                "input_copy": str(input_copy),
                "workbook": str(workbook_copy),
            },
        )
        return session

    @property
    def workbook_path(self) -> Path:
        return self.paths.workbook

    @property
    def workspace(self) -> Path:
        return self.paths.root

    def _assert_artifact_sync_locked(self) -> None:
        observed_sha256 = _sha256(self.workbook_path)
        if observed_sha256 != self._artifact.sha256:
            raise WorkbookValidationError(
                "Managed workbook bytes changed outside a recorded artifact transition: "
                f"revision {self._artifact.revision} expected {self._artifact.sha256}, "
                f"observed {observed_sha256}"
            )

    def artifact_ref(self) -> ArtifactRef:
        with self._write_lock:
            self._assert_artifact_sync_locked()
            return self._artifact

    def enable_completion_attempt_capture(self) -> CompletionAttemptLedger:
        """Enable one task-wide attempt ledger without changing default runs."""

        with self._write_lock:
            self._assert_artifact_sync_locked()
            if self._completion_attempt_ledger is None:
                self._completion_attempt_ledger = CompletionAttemptLedger(self.workspace)
            return self._completion_attempt_ledger

    def capture_completion_attempt(
        self,
        *,
        stage: str,
        turn: int,
        response_id: str | None,
        call_id: str,
    ) -> CompletionAttemptRecord:
        """Atomically bind a submit attempt to the current revision and bytes."""

        with self._write_lock:
            self._assert_artifact_sync_locked()
            if self._completion_attempt_ledger is None:
                raise WorkbookValidationError(
                    "Completion-attempt capture was not explicitly enabled"
                )
            return self._completion_attempt_ledger.capture(
                self.workbook_path,
                self._artifact,
                stage=stage,
                turn=turn,
                response_id=response_id,
                call_id=call_id,
            )

    @property
    def artifact_transitions(self) -> tuple[ArtifactTransition, ...]:
        with self._write_lock:
            return tuple(self._artifact_transitions)

    @property
    def target_grounding_enabled(self) -> bool:
        """Compatibility alias for :attr:`target_grounding_enforced`."""

        return self.target_grounding_enforced

    @property
    def target_grounding_enforced(self) -> bool:
        """Whether target assessment can reject a staged publication."""

        with self._write_lock:
            return self._target_grounding_mode is TargetGroundingMode.ENFORCE

    @property
    def target_grounding_active(self) -> bool:
        """Whether either advisory observation or enforcement is active."""

        with self._write_lock:
            return self._target_grounding is not None

    @property
    def target_grounding_mode(self) -> TargetGroundingMode:
        with self._write_lock:
            return self._target_grounding_mode

    @property
    def target_grounding_initial_artifact(self) -> ArtifactRef | None:
        with self._write_lock:
            return self._target_grounding_initial_artifact

    @property
    def target_grounding_initial_transition_count(self) -> int | None:
        with self._write_lock:
            return self._target_grounding_initial_transition_count

    @property
    def committed_target_authorizations(
        self,
    ) -> tuple[CommittedTargetAuthorization, ...]:
        with self._write_lock:
            if self._target_grounding is None:
                return ()
            return self._target_grounding.committed_authorizations

    @property
    def committed_advisory_target_assessments(
        self,
    ) -> tuple[CommittedAdvisoryTargetAssessment, ...]:
        with self._write_lock:
            if self._target_grounding is None:
                return ()
            return self._target_grounding.committed_advisory_assessments

    @property
    def advisory_target_lifecycle_events(self) -> tuple[AdvisoryLifecycleEvent, ...]:
        with self._write_lock:
            if self._target_grounding is None:
                return ()
            return self._target_grounding.advisory_lifecycle_events

    @property
    def advisory_target_lifecycle_genesis_sha256(self) -> str | None:
        with self._write_lock:
            if self._target_grounding is None:
                return None
            return self._target_grounding.advisory_lifecycle_genesis_sha256

    @property
    def advisory_target_lifecycle_final_counters(self) -> dict[str, int] | None:
        with self._write_lock:
            if self._target_grounding is None:
                return None
            return dict(self._target_grounding.advisory_lifecycle_final_counters)

    def enable_target_grounding(
        self,
        mode: TargetGroundingMode = TargetGroundingMode.ENFORCE,
    ) -> None:
        if not isinstance(mode, TargetGroundingMode):
            raise TypeError("mode must be a TargetGroundingMode")
        if mode is TargetGroundingMode.OFF:
            raise ValueError("use the default session state to keep target grounding off")
        with self._write_lock:
            self._assert_artifact_sync_locked()
            if self._target_grounding is None:
                transition_count = len(self._artifact_transitions)
                gate = TargetGroundingStateMachine(
                    self._artifact,
                    mode=mode,
                    initial_transition_count=transition_count,
                )
                self.recorder.record(
                    "target_grounding.enabled",
                    {
                        "mode": mode.value,
                        "decision": "enabled",
                        "artifact": self._artifact.to_dict(),
                        "initial_transition_count": transition_count,
                    },
                )
                self._target_grounding = gate
                self._target_grounding_mode = mode
                self._target_grounding_initial_artifact = self._artifact
                self._target_grounding_initial_transition_count = transition_count
            elif self._target_grounding.mode is not mode:
                raise WorkbookValidationError(
                    "Target-grounding mode cannot change after its ledger is initialized"
                )
            elif self._target_grounding.current_artifact != self._artifact:
                raise WorkbookValidationError(
                    "Target-grounding ledger is not synchronized with the managed artifact"
                )

    def record_target_observation(
        self,
        *,
        artifact: ArtifactRef,
        scope: EvidenceScope,
    ) -> dict[str, Any]:
        with self._write_lock:
            if self._target_grounding is None:
                raise TargetGroundingError("target grounding is not enabled")
            self._assert_artifact_sync_locked()
            observation = self._target_grounding.record_trusted_observation(
                artifact=artifact,
                scope=scope,
            )
            document = observation.to_dict()
            self.recorder.record(
                "target_grounding.observation",
                {
                    **document,
                    "mode": self._target_grounding_mode.value,
                    "decision": "recorded",
                },
            )
            return document

    def declare_edit_target(
        self,
        *,
        target_scope: EvidenceScope,
        observation_ids: tuple[int, ...] | list[int],
    ) -> dict[str, Any]:
        with self._write_lock:
            if self._target_grounding is None:
                raise TargetGroundingError("target grounding is not enabled")
            self._assert_artifact_sync_locked()
            declaration = self._target_grounding.declare_target(
                artifact=self._artifact,
                target_scope=target_scope,
                observation_ids=observation_ids,
            )
            document = declaration.to_dict()
            self.recorder.record(
                "target_grounding.declaration",
                {
                    **document,
                    "mode": self._target_grounding_mode.value,
                    "decision": "recorded",
                },
            )
            return document

    def _planned_transition_locked(
        self,
        *,
        operation: str,
        kind: str,
        before_sha256: str,
        after_sha256: str,
    ) -> ArtifactTransition | None:
        if self._artifact.sha256 != before_sha256:
            raise WorkbookValidationError(
                "Artifact transition source does not match the managed revision"
            )
        if after_sha256 == before_sha256:
            return None
        before = self._artifact
        return ArtifactTransition(
            transition_id=len(self._artifact_transitions) + 1,
            operation=operation,
            kind=kind,
            before=before,
            after=ArtifactRef(before.revision + 1, after_sha256),
        )

    def _publish_artifact_locked(
        self,
        *,
        operation: str,
        kind: str,
        before_sha256: str,
        after_sha256: str,
        target_prepared: (
            PreparedTargetAuthorization | PreparedAdvisoryTargetAssessment | None
        ) = None,
    ) -> ArtifactTransition | None:
        transition = self._planned_transition_locked(
            operation=operation,
            kind=kind,
            before_sha256=before_sha256,
            after_sha256=after_sha256,
        )
        published_bytes: bytes | None = None
        if transition is not None:
            published_bytes = self.workbook_path.read_bytes()
            if hashlib.sha256(published_bytes).hexdigest() != transition.after.sha256:
                raise WorkbookValidationError(
                    "Published workbook bytes changed before artifact ledger commit"
                )
        gate = self._target_grounding
        committed_authorization: CommittedTargetAuthorization | None = None
        committed_advisory: CommittedAdvisoryTargetAssessment | None = None
        if gate is not None:
            protected_transition = is_target_grounding_protected_transition_kind(kind)
            if protected_transition is not (target_prepared is not None):
                raise WorkbookValidationError(
                    "Protected artifact publications require exactly one target assessment"
                )
            if isinstance(target_prepared, PreparedTargetAuthorization):
                committed_authorization = gate.preview_committed_authorization(
                    target_prepared,
                    transition,
                )
            elif isinstance(target_prepared, PreparedAdvisoryTargetAssessment):
                committed_advisory = gate.preview_committed_advisory_assessment(
                    target_prepared,
                    transition,
                )
        if transition is not None:
            # Persist the ledger entry before advancing in-memory lineage. If the
            # recorder raises, callers can restore staged bytes while all ledgers
            # still point at the original artifact.
            transition_document = transition.to_dict()
            if committed_authorization is not None:
                transition_document["target_grounding_commit_json"] = (
                    committed_authorization.canonical_json()
                )
                transition_document.update(
                    {
                        "target_grounding_mode": TargetGroundingMode.ENFORCE.value,
                        "target_grounding_decision": (
                            committed_authorization.provenance.decision.value
                        ),
                    }
                )
            elif committed_advisory is not None:
                transition_document.update(
                    {
                        "target_grounding_advisory_commit_json": (
                            committed_advisory.canonical_json()
                        ),
                        "target_grounding_mode": TargetGroundingMode.ADVISORY.value,
                        "target_grounding_decision": ("published_after_advisory_assessment"),
                        "target_grounding_would_reject": (
                            committed_advisory.assessment.would_reject
                        ),
                    }
                )
            self.recorder.record("artifact.transition", transition_document)
            self._artifact = transition.after
            self._artifact_transitions.append(transition)
        elif committed_authorization is not None:
            self.recorder.record(
                "target_grounding.authorization.committed",
                {
                    "mode": TargetGroundingMode.ENFORCE.value,
                    "decision": committed_authorization.provenance.decision.value,
                    "target_grounding_commit_json": committed_authorization.canonical_json(),
                },
            )
        elif committed_advisory is not None:
            self.recorder.record(
                "target_grounding.advisory_assessment.committed",
                {
                    "mode": TargetGroundingMode.ADVISORY.value,
                    "decision": "published_after_advisory_assessment",
                    "would_reject": committed_advisory.assessment.would_reject,
                    "target_grounding_advisory_commit_json": (committed_advisory.canonical_json()),
                },
            )
        if gate is not None:
            if isinstance(target_prepared, PreparedTargetAuthorization):
                gate.commit_prepared(
                    target_prepared,
                    transition,
                    committed_authorization=committed_authorization,
                )
            elif isinstance(target_prepared, PreparedAdvisoryTargetAssessment):
                gate.commit_advisory_assessment(
                    target_prepared,
                    transition,
                    committed_assessment=committed_advisory,
                )
            elif transition is not None:
                gate.record_artifact_transition(transition)
        if published_bytes is not None:
            self._artifact_bytes = published_bytes
        return transition

    def reconcile_external_artifact(
        self,
        before: ArtifactRef,
        *,
        operation: str,
        kind: str = "external_mutation",
        declaration_id: int | None = None,
    ) -> ArtifactTransition | None:
        """Register bytes written by a sandboxed tool that bypasses session helpers."""

        with self._write_lock:
            if before != self._artifact:
                raise WorkbookValidationError(
                    "External mutation source does not match the managed artifact revision"
                )
            target_prepared: (
                PreparedTargetAuthorization | PreparedAdvisoryTargetAssessment | None
            ) = None
            try:
                self._validate(self.workbook_path)
                after_sha256 = _sha256(self.workbook_path)
                if self._target_grounding is not None and (
                    is_target_grounding_protected_transition_kind(kind)
                ):
                    if hashlib.sha256(self._artifact_bytes).hexdigest() != before.sha256:
                        raise WorkbookValidationError(
                            "Cached source bytes do not match the external mutation source"
                        )
                    with tempfile.TemporaryDirectory(
                        prefix="sheet-grounding-reconcile-"
                    ) as raw_root:
                        cached_before = Path(raw_root) / f"before{self.workbook_path.suffix}"
                        cached_before.write_bytes(self._artifact_bytes)
                        workbook_diff = diff_workbooks(
                            cached_before,
                            self.workbook_path,
                        )
                    target_prepared = self._prepare_target_authorization_locked(
                        declaration_id=declaration_id,
                        diff=workbook_diff,
                        after_sha256=after_sha256,
                    )
                    planned_transition = self._planned_transition_locked(
                        operation=operation,
                        kind=kind,
                        before_sha256=before.sha256,
                        after_sha256=after_sha256,
                    )
                    self._validate_prepared_target_transition_locked(
                        target_prepared,
                        planned_transition,
                    )
                return self._publish_artifact_locked(
                    operation=operation,
                    kind=kind,
                    before_sha256=before.sha256,
                    after_sha256=after_sha256,
                    target_prepared=target_prepared,
                )
            except Exception:
                if self._artifact == before:
                    try:
                        observed_sha256 = _sha256(self.workbook_path)
                    except OSError:
                        observed_sha256 = None
                    if observed_sha256 != before.sha256:
                        self._restore_cached_artifact_locked()
                    if target_prepared is not None:
                        try:
                            self._abort_prepared_target_locked(target_prepared)
                        except TargetGroundingError:
                            pass
                raise

    def run_staged_external_mutation(
        self,
        *,
        operation: str,
        declaration_id: int | None,
        runner: Callable[[Path], dict[str, Any]],
    ) -> dict[str, Any]:
        """Run opaque code on an isolated copy and publish only an authorized diff."""

        with self._write_lock:
            if self._target_grounding is None:
                raise TargetGroundingError("target grounding is not enabled")
            self._assert_artifact_sync_locked()
            artifact_before = self._artifact
            self._snapshot_counter += 1
            snapshot = (
                self.paths.snapshots
                / f"{self._snapshot_counter:04d}_{operation}{self.workbook_path.suffix}"
            )
            shutil.copy2(self.workbook_path, snapshot)
            target_prepared: (
                PreparedTargetAuthorization | PreparedAdvisoryTargetAssessment | None
            ) = None
            keep_snapshot = False
            self.recorder.record(
                "workbook.mutation.started",
                {
                    "operation": operation,
                    "arguments": {"opaque_staged_execution": True},
                    "snapshot": snapshot,
                },
            )
            try:
                with tempfile.TemporaryDirectory(prefix="sheet-grounding-") as root:
                    staging_root = Path(root)
                    staged = staging_root / f"staged{self.workbook_path.suffix}"
                    shutil.copy2(self.workbook_path, staged)
                    result = runner(staged)
                    if not isinstance(result, dict):
                        raise WorkbookValidationError(
                            "Opaque mutation runner must return a result mapping"
                        )

                    managed_sha256 = _sha256(self.workbook_path)
                    if managed_sha256 != artifact_before.sha256:
                        self._restore_managed_artifact_locked(snapshot)
                        snapshot.unlink(missing_ok=True)
                        return {
                            **result,
                            "ok": False,
                            "error": (
                                "Opaque code modified the managed workbook outside its staging "
                                "artifact; the source bytes were restored and publication denied."
                            ),
                            "type": "StagingBoundaryViolation",
                            "workbook_sha256_before": artifact_before.sha256,
                            "workbook_sha256_rejected": managed_sha256,
                            "workbook_sha256_after": artifact_before.sha256,
                            "workbook_changed": False,
                            "workbook_rolled_back": True,
                            "mutation_published": False,
                            "artifact_revision_before": artifact_before.revision,
                            "artifact_revision_after": artifact_before.revision,
                            "artifact_transition_id": None,
                        }

                    if result.get("ok") is not True:
                        snapshot.unlink(missing_ok=True)
                        return {
                            **result,
                            "workbook_sha256_before": artifact_before.sha256,
                            "workbook_sha256_after": artifact_before.sha256,
                            "workbook_changed": False,
                            "mutation_published": False,
                            "artifact_revision_before": artifact_before.revision,
                            "artifact_revision_after": artifact_before.revision,
                            "artifact_transition_id": None,
                        }

                    self._validate(staged)
                    invalid_references, formula_text = validate_formula_transaction(
                        snapshot,
                        staged,
                    )
                    if invalid_references or formula_text:
                        raise ToolInputError(
                            "Opaque mutation introduced formula validation failures after its "
                            "staged execution completed"
                        )
                    workbook_diff = diff_workbooks(snapshot, staged)
                    after_sha256 = _sha256(staged)
                    try:
                        target_prepared = self._prepare_target_authorization_locked(
                            declaration_id=declaration_id,
                            diff=workbook_diff,
                            after_sha256=after_sha256,
                        )
                    except TargetGroundingRejected as rejected:
                        rejection = self._target_rejection_result_locked(
                            artifact_before=artifact_before,
                            rejected_sha256=after_sha256,
                            diff=workbook_diff,
                            rejected=rejected,
                        )
                        snapshot.unlink(missing_ok=True)
                        rejection.update(
                            {
                                key: value
                                for key, value in result.items()
                                if key
                                in {
                                    "stdout",
                                    "stderr",
                                    "truncated",
                                    "sandbox",
                                    "bubblewrap_error",
                                    "script",
                                    "managed_mutation_attempted",
                                    "helper_module",
                                }
                            }
                        )
                        self.recorder.record(
                            "target_grounding.rejected",
                            {
                                "mode": TargetGroundingMode.ENFORCE.value,
                                "decision": rejected.decision.value,
                                "operation": operation,
                                "result": rejection,
                            },
                        )
                        return rejection

                    publish_temporary = self.workbook_path.with_name(
                        f".{self.workbook_path.stem}.publish-{uuid.uuid4().hex}"
                        f"{self.workbook_path.suffix}"
                    )
                    try:
                        shutil.copy2(staged, publish_temporary)
                        if _sha256(publish_temporary) != after_sha256:
                            raise WorkbookValidationError(
                                "Staged workbook changed while preparing atomic publication"
                            )
                        planned_transition = self._planned_transition_locked(
                            operation=operation,
                            kind="external_mutation",
                            before_sha256=artifact_before.sha256,
                            after_sha256=after_sha256,
                        )
                        self._validate_prepared_target_transition_locked(
                            target_prepared,
                            planned_transition,
                        )
                        publish_temporary.replace(self.workbook_path)
                        transition = self._publish_artifact_locked(
                            operation=operation,
                            kind="external_mutation",
                            before_sha256=artifact_before.sha256,
                            after_sha256=after_sha256,
                            target_prepared=target_prepared,
                        )
                    finally:
                        publish_temporary.unlink(missing_ok=True)

                    keep_snapshot = bool(transition is not None and workbook_diff.semantic_changed)
                    published = {
                        **result,
                        "workbook_sha256_before": artifact_before.sha256,
                        "workbook_sha256_after": after_sha256,
                        "workbook_changed": after_sha256 != artifact_before.sha256,
                        "mutation_published": True,
                        "artifact_revision_before": artifact_before.revision,
                        "artifact_revision_after": self._artifact.revision,
                        "artifact_transition_id": (
                            transition.transition_id if transition is not None else None
                        ),
                        "workbook_effects": workbook_diff.to_dict(),
                    }
                    target_diagnostic = self._target_model_diagnostic(target_prepared)
                    if target_diagnostic is not None:
                        published["target_grounding"] = target_diagnostic
                    self.recorder.record(
                        "workbook.mutation.committed",
                        {
                            "operation": operation,
                            "snapshot": snapshot,
                            "result": published,
                        },
                    )
                    return published
            except Exception as exc:
                if self._artifact == artifact_before:
                    if _sha256(self.workbook_path) != artifact_before.sha256:
                        self._restore_managed_artifact_locked(snapshot)
                    if target_prepared is not None:
                        try:
                            self._abort_prepared_target_locked(target_prepared)
                        except TargetGroundingError:
                            pass
                try:
                    self.recorder.record(
                        "workbook.mutation.rolled_back",
                        {"operation": operation, "snapshot": snapshot, "error": str(exc)},
                    )
                except Exception:
                    pass
                raise
            finally:
                if not keep_snapshot:
                    snapshot.unlink(missing_ok=True)

    @contextmanager
    def read_artifact(self) -> Iterator[ArtifactRef]:
        """Hold the session lock while producing evidence from stable workbook bytes."""

        with self._write_lock:
            self._assert_artifact_sync_locked()
            artifact = self._artifact
            yield artifact
            self._assert_artifact_sync_locked()

    def _load(self, *, data_only: bool = False):
        return load_workbook(
            self.workbook_path,
            data_only=data_only,
            keep_vba=self.workbook_path.suffix.lower() == ".xlsm",
            keep_links=True,
        )

    def _validate(self, path: Path) -> None:
        try:
            workbook = load_workbook(
                path,
                read_only=True,
                data_only=False,
                keep_vba=path.suffix.lower() == ".xlsm",
                keep_links=True,
            )
            if not workbook.sheetnames:
                raise WorkbookValidationError("Workbook contains no worksheets")
            workbook.close()
        except WorkbookValidationError:
            raise
        except Exception as exc:
            raise WorkbookValidationError(f"Workbook validation failed: {exc}") from exc

    def _restore_managed_artifact_locked(self, snapshot: Path) -> None:
        restore = self.workbook_path.with_name(
            f".{self.workbook_path.stem}.restore-{uuid.uuid4().hex}{self.workbook_path.suffix}"
        )
        try:
            shutil.copy2(snapshot, restore)
            restore.replace(self.workbook_path)
        finally:
            restore.unlink(missing_ok=True)

    def _restore_cached_artifact_locked(self) -> None:
        if hashlib.sha256(self._artifact_bytes).hexdigest() != self._artifact.sha256:
            raise WorkbookValidationError(
                "Cached artifact bytes do not match the current artifact ledger"
            )
        restore = self.workbook_path.with_name(
            f".{self.workbook_path.stem}.cached-restore-{uuid.uuid4().hex}"
            f"{self.workbook_path.suffix}"
        )
        try:
            restore.write_bytes(self._artifact_bytes)
            restore.replace(self.workbook_path)
        finally:
            restore.unlink(missing_ok=True)

    def _sheet(self, workbook: Any, name: str) -> Worksheet:
        if name not in workbook.sheetnames:
            raise ToolInputError(f"Unknown sheet {name!r}; available: {workbook.sheetnames}")
        return workbook[name]

    def _bounds(
        self,
        range_ref: str,
        *,
        max_cells: int = 10_000,
    ) -> tuple[int, int, int, int]:
        try:
            bounds = range_boundaries(range_ref.replace("$", ""))
        except (TypeError, ValueError) as exc:
            raise ToolInputError(f"Invalid A1 range: {range_ref!r}") from exc
        min_col, min_row, max_col, max_row = bounds
        if not all(isinstance(item, int) and item >= 1 for item in bounds):
            raise ToolInputError(f"Range must be bounded: {range_ref!r}")
        count = (max_col - min_col + 1) * (max_row - min_row + 1)
        if count > max_cells:
            raise ToolInputError(f"Range contains {count} cells; limit is {max_cells}")
        return min_col, min_row, max_col, max_row

    def _prepare_target_authorization_locked(
        self,
        *,
        declaration_id: int | None,
        diff: WorkbookEffectDiff,
        after_sha256: str,
    ) -> PreparedTargetAuthorization | PreparedAdvisoryTargetAssessment | None:
        gate = self._target_grounding
        if gate is None:
            return None
        staged_artifact = ArtifactRef(
            self._artifact.revision + int(after_sha256 != self._artifact.sha256),
            after_sha256,
        )
        if gate.mode is TargetGroundingMode.ADVISORY:
            return gate.prepare_advisory_staged_diff(
                declaration_id,
                diff,
                staged_artifact=staged_artifact,
            )
        return gate.prepare_staged_diff(
            declaration_id,
            diff,
            staged_artifact=staged_artifact,
        )

    def _validate_prepared_target_transition_locked(
        self,
        prepared: PreparedTargetAuthorization | PreparedAdvisoryTargetAssessment,
        transition: ArtifactTransition | None,
    ) -> None:
        gate = self._target_grounding
        if gate is None:
            raise TargetGroundingError("target grounding is not enabled")
        if isinstance(prepared, PreparedTargetAuthorization):
            gate.validate_prepared_transition(prepared, transition)
        else:
            gate.validate_prepared_advisory_transition(prepared, transition)

    def _abort_prepared_target_locked(
        self,
        prepared: PreparedTargetAuthorization | PreparedAdvisoryTargetAssessment,
    ) -> None:
        gate = self._target_grounding
        if gate is None:
            raise TargetGroundingError("target grounding is not enabled")
        if isinstance(prepared, PreparedTargetAuthorization):
            gate.abort_prepared(prepared)
        else:
            gate.abort_prepared_advisory_assessment(prepared)

    @staticmethod
    def _target_model_diagnostic(
        prepared: PreparedTargetAuthorization | PreparedAdvisoryTargetAssessment,
    ) -> dict[str, Any] | None:
        if isinstance(prepared, PreparedTargetAuthorization):
            return prepared.record.model_diagnostic()
        return prepared.assessment.model_diagnostic()

    def _target_rejection_result_locked(
        self,
        *,
        artifact_before: ArtifactRef,
        rejected_sha256: str,
        diff: WorkbookEffectDiff,
        rejected: TargetGroundingRejected,
    ) -> dict[str, Any]:
        return {
            "ok": False,
            "error": str(rejected),
            "type": (
                type(rejected).__name__
                if rejected.model_diagnostic.get("declaration_status") == "valid"
                else TargetGroundingError.__name__
            ),
            "workbook_sha256_before": artifact_before.sha256,
            "workbook_sha256_rejected": rejected_sha256,
            "workbook_sha256_after": artifact_before.sha256,
            "workbook_changed": False,
            "workbook_rolled_back": True,
            "mutation_published": False,
            "artifact_revision_before": artifact_before.revision,
            "artifact_revision_after": artifact_before.revision,
            "artifact_transition_id": None,
            "workbook_effects": diff.to_dict(),
            "target_grounding": rejected.model_diagnostic,
            "message": (
                "The staged edit was rejected before publication; inspect the exact target, "
                "declare a newly grounded finite scope, and retry."
            ),
        }

    def _mutate(
        self,
        operation: str,
        arguments: dict[str, Any],
        callback: Callable[[Any], Any],
        *,
        declaration_id: int | None = None,
    ) -> Any:
        with self._write_lock:
            self._assert_artifact_sync_locked()
            artifact_before = self._artifact
            before_sha256 = _sha256(self.workbook_path)
            self._snapshot_counter += 1
            snapshot = (
                self.paths.snapshots
                / f"{self._snapshot_counter:04d}_{operation}{self.workbook_path.suffix}"
            )
            shutil.copy2(self.workbook_path, snapshot)
            temporary = self.workbook_path.with_name(
                f".{self.workbook_path.stem}.{uuid.uuid4().hex}.tmp{self.workbook_path.suffix}"
            )
            self.recorder.record(
                "workbook.mutation.started",
                {"operation": operation, "arguments": arguments, "snapshot": snapshot},
            )
            workbook = None
            target_prepared: (
                PreparedTargetAuthorization | PreparedAdvisoryTargetAssessment | None
            ) = None
            keep_snapshot = False
            try:
                workbook = self._load(data_only=False)
                result = callback(workbook)
                calculation = getattr(workbook, "calculation", None)
                if calculation is not None:
                    calculation.fullCalcOnLoad = True
                    calculation.forceFullCalc = True
                    calculation.calcMode = "auto"
                workbook.save(temporary)
                workbook.close()
                workbook = None
                self._validate(temporary)
                invalid_references, formula_text = validate_formula_transaction(
                    snapshot,
                    temporary,
                )
                if invalid_references or formula_text:
                    issue_locations = sorted(
                        {(sheet, cell) for sheet, cell, *_ in invalid_references}
                        | {(sheet, cell) for sheet, cell, _ in formula_text}
                    )
                    locations = ", ".join(f"{sheet}!{cell}" for sheet, cell in issue_locations[:8])
                    raise ToolInputError(
                        "Mutation introduced invalid or high-confidence formula-like text at "
                        f"{locations}. Excel formulas must be strings beginning with '='; "
                        "correct every reported formula issue and retry the complete edit."
                    )
                workbook_diff = diff_workbooks(snapshot, temporary)
                after_sha256 = _sha256(temporary)
                try:
                    target_prepared = self._prepare_target_authorization_locked(
                        declaration_id=declaration_id,
                        diff=workbook_diff,
                        after_sha256=after_sha256,
                    )
                except TargetGroundingRejected as rejected:
                    rejection = self._target_rejection_result_locked(
                        artifact_before=artifact_before,
                        rejected_sha256=after_sha256,
                        diff=workbook_diff,
                        rejected=rejected,
                    )
                    snapshot.unlink(missing_ok=True)
                    self.recorder.record(
                        "target_grounding.rejected",
                        {
                            "mode": TargetGroundingMode.ENFORCE.value,
                            "decision": rejected.decision.value,
                            "operation": operation,
                            "result": rejection,
                        },
                    )
                    return rejection
                planned_transition = self._planned_transition_locked(
                    operation=operation,
                    kind="mutation",
                    before_sha256=before_sha256,
                    after_sha256=after_sha256,
                )
                if self._target_grounding is not None and target_prepared is not None:
                    self._validate_prepared_target_transition_locked(
                        target_prepared,
                        planned_transition,
                    )
                temporary.replace(self.workbook_path)
                transition = self._publish_artifact_locked(
                    operation=operation,
                    kind="mutation",
                    before_sha256=before_sha256,
                    after_sha256=after_sha256,
                    target_prepared=target_prepared,
                )
                keep_snapshot = bool(transition is not None and workbook_diff.semantic_changed)
                if isinstance(result, dict):
                    result = {
                        **result,
                        "workbook_sha256_before": before_sha256,
                        "workbook_sha256_after": after_sha256,
                        "workbook_changed": before_sha256 != after_sha256,
                        "artifact_revision_before": artifact_before.revision,
                        "artifact_revision_after": self._artifact.revision,
                        "artifact_transition_id": (
                            transition.transition_id if transition is not None else None
                        ),
                        "workbook_effects": workbook_diff.to_dict(),
                        "message": (
                            "Workbook changed. If the target range has been verified, finish now; "
                            "otherwise run one narrow verification or correction."
                            if before_sha256 != after_sha256
                            else "Workbook did not change; revise the mutation before submitting."
                        ),
                    }
                    if target_prepared is not None:
                        target_diagnostic = self._target_model_diagnostic(target_prepared)
                        if target_diagnostic is not None:
                            result["target_grounding"] = target_diagnostic
                self.recorder.record(
                    "workbook.mutation.committed",
                    {"operation": operation, "snapshot": snapshot, "result": result},
                )
                return result
            except Exception as exc:
                if self._artifact == artifact_before:
                    observed_sha256 = _sha256(self.workbook_path)
                    if observed_sha256 != artifact_before.sha256:
                        self._restore_managed_artifact_locked(snapshot)
                    if target_prepared is not None and self._target_grounding is not None:
                        try:
                            self._abort_prepared_target_locked(target_prepared)
                        except TargetGroundingError:
                            pass
                try:
                    self.recorder.record(
                        "workbook.mutation.rolled_back",
                        {"operation": operation, "snapshot": snapshot, "error": str(exc)},
                    )
                except Exception:
                    pass
                raise
            finally:
                if workbook is not None:
                    workbook.close()
                temporary.unlink(missing_ok=True)
                if not keep_snapshot:
                    snapshot.unlink(missing_ok=True)

    def list_sheets(self) -> dict[str, Any]:
        with self.read_artifact() as artifact:
            result = self._list_sheets_unlocked()
            return {
                **result,
                "artifact_revision": artifact.revision,
                "artifact_sha256": artifact.sha256,
            }

    def _list_sheets_unlocked(self) -> dict[str, Any]:
        workbook = self._load(data_only=False)
        try:
            sheets = []
            for index, worksheet in enumerate(workbook.worksheets):
                dimension = worksheet.calculate_dimension()
                sheets.append(
                    {
                        "index": index,
                        "name": worksheet.title,
                        "state": worksheet.sheet_state,
                        "dimension": dimension,
                        "max_row": worksheet.max_row,
                        "max_column": worksheet.max_column,
                        "merged_ranges": len(worksheet.merged_cells.ranges),
                        "tables": list(worksheet.tables.keys()),
                    }
                )
            return {"ok": True, "sheets": sheets, "active": workbook.active.title}
        finally:
            workbook.close()

    def inspect_range(
        self,
        sheet: str,
        range_ref: str,
        *,
        include_styles: bool = True,
        max_cells: int = 500,
    ) -> dict[str, Any]:
        with self.read_artifact() as artifact:
            result = self._inspect_range_unlocked(
                sheet,
                range_ref,
                include_styles=include_styles,
                max_cells=max_cells,
            )
            return {
                **result,
                "artifact_revision": artifact.revision,
                "artifact_sha256": artifact.sha256,
            }

    def _inspect_range_unlocked(
        self,
        sheet: str,
        range_ref: str,
        *,
        include_styles: bool,
        max_cells: int,
    ) -> dict[str, Any]:
        bounds = self._bounds(range_ref, max_cells=max_cells)
        min_col, min_row, max_col, max_row = bounds
        formula_book = self._load(data_only=False)
        value_book = self._load(data_only=True)
        try:
            formula_sheet = self._sheet(formula_book, sheet)
            value_sheet = self._sheet(value_book, sheet)
            matrix: list[list[Any]] = []
            cells: list[dict[str, Any]] = []
            for row in range(min_row, max_row + 1):
                matrix_row: list[Any] = []
                for column in range(min_col, max_col + 1):
                    cell = formula_sheet.cell(row, column)
                    cached = value_sheet.cell(row, column).value
                    raw = cell.value
                    display = raw if isinstance(raw, str) and raw.startswith("=") else cached
                    if display is None:
                        display = raw
                    matrix_row.append(_json_value(display))
                    if raw is not None or cached is not None or cell.has_style:
                        item: dict[str, Any] = {
                            "coordinate": cell.coordinate,
                            "value": _json_value(cached if cached is not None else raw),
                            "formula": raw
                            if isinstance(raw, str) and raw.startswith("=")
                            else None,
                            "data_type": cell.data_type,
                        }
                        if include_styles:
                            item["style"] = {
                                "style_id": cell.style_id,
                                "number_format": cell.number_format,
                                "font": {
                                    "bold": bool(cell.font.bold),
                                    "italic": bool(cell.font.italic),
                                    "color": _color_value(cell.font.color),
                                },
                                "fill": _color_value(cell.fill.fgColor),
                                "alignment": {
                                    "horizontal": cell.alignment.horizontal,
                                    "vertical": cell.alignment.vertical,
                                    "wrap_text": cell.alignment.wrap_text,
                                },
                            }
                        cells.append(item)
                matrix.append(matrix_row)

            merged = [
                str(item)
                for item in formula_sheet.merged_cells.ranges
                if _intersects(bounds, range_boundaries(str(item)))
            ]
            tables = []
            for name in formula_sheet.tables.keys():
                ref = _table_ref(formula_sheet.tables[name])
                if _intersects(bounds, range_boundaries(ref)):
                    tables.append({"name": name, "ref": ref})
            return {
                "ok": True,
                "sheet": sheet,
                "range": f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}",
                "matrix": matrix,
                "cells": cells,
                "merged_ranges": merged,
                "tables": tables,
            }
        finally:
            formula_book.close()
            value_book.close()

    def find_cells(
        self,
        query: str,
        *,
        sheet: str | None = None,
        use_regex: bool = False,
        match_case: bool = False,
        search_formulas: bool = True,
        max_results: int = 200,
        max_scanned_cells: int = 250_000,
    ) -> dict[str, Any]:
        if not query:
            raise ToolInputError("query must not be empty")
        flags = 0 if match_case else re.IGNORECASE
        pattern = re.compile(query if use_regex else re.escape(query), flags)
        workbook = self._load(data_only=not search_formulas)
        scanned = 0
        matches: list[dict[str, Any]] = []
        truncated = False
        try:
            worksheets: Iterable[Worksheet]
            worksheets = [self._sheet(workbook, sheet)] if sheet else workbook.worksheets
            for worksheet in worksheets:
                for row in worksheet.iter_rows():
                    for cell in row:
                        scanned += 1
                        if scanned > max_scanned_cells:
                            truncated = True
                            break
                        value = cell.value
                        if value is not None and pattern.search(str(value)):
                            matches.append(
                                {
                                    "sheet": worksheet.title,
                                    "coordinate": cell.coordinate,
                                    "value": _json_value(value),
                                }
                            )
                            if len(matches) >= max_results:
                                truncated = True
                                break
                    if truncated:
                        break
                if truncated:
                    break
            return {
                "ok": True,
                "query": query,
                "matches": matches,
                "scanned_cells": scanned,
                "truncated": truncated,
            }
        finally:
            workbook.close()

    def write_range(
        self,
        sheet: str,
        start_cell: str,
        values: list[list[Any]],
        *,
        declaration_id: int | None = None,
    ) -> dict[str, Any]:
        if (
            not values
            or not isinstance(values, list)
            or any(not isinstance(row, list) for row in values)
        ):
            raise ToolInputError("values must be a non-empty two-dimensional list")
        width = max((len(row) for row in values), default=0)
        if width == 0 or any(len(row) != width for row in values):
            raise ToolInputError("values must be rectangular and contain at least one column")
        if len(values) * width > 10_000:
            raise ToolInputError("write_range is limited to 10,000 cells")
        try:
            start_row, start_column = coordinate_to_tuple(start_cell.replace("$", ""))
        except (TypeError, ValueError) as exc:
            raise ToolInputError(f"Invalid start_cell: {start_cell!r}") from exc

        def apply(workbook: Any) -> dict[str, Any]:
            worksheet = self._sheet(workbook, sheet)
            for row_offset, row_values in enumerate(values):
                for column_offset, value in enumerate(row_values):
                    worksheet.cell(
                        start_row + row_offset, start_column + column_offset
                    ).value = value
            end = f"{get_column_letter(start_column + width - 1)}{start_row + len(values) - 1}"
            return {
                "ok": True,
                "sheet": sheet,
                "range": f"{start_cell}:{end}",
                "cells_written": len(values) * width,
            }

        return self._mutate(
            "write_range",
            {"sheet": sheet, "start_cell": start_cell, "values": values},
            apply,
            declaration_id=declaration_id,
        )

    def fill_formula(
        self,
        sheet: str,
        source_cell: str,
        target_range: str,
        *,
        declaration_id: int | None = None,
    ) -> dict[str, Any]:
        normalized_target_range, expanded_from_endpoint = _normalize_fill_target_range(
            source_cell, target_range
        )
        bounds = self._bounds(normalized_target_range)

        def apply(workbook: Any) -> dict[str, Any]:
            worksheet = self._sheet(workbook, sheet)
            formula = worksheet[source_cell].value
            if not isinstance(formula, str) or not formula.startswith("="):
                if (
                    isinstance(formula, str)
                    and formula == formula.strip()
                    and _FORMULA_RANGE_RE.search(formula) is not None
                    and re.match(r"[A-Za-z][A-Za-z0-9_.]*\(", formula) is not None
                ):
                    raise ToolInputError(
                        f"Source {sheet}!{source_cell} contains formula-like text without a "
                        "leading '='; assign an Excel formula string beginning with '=' before "
                        "calling fill_formula"
                    )
                raise ToolInputError(f"Source {sheet}!{source_cell} does not contain a formula")
            min_col, min_row, max_col, max_row = bounds
            count = 0
            samples: list[dict[str, Any]] = []
            sample_coordinates = set(_formula_sample_coordinates(source_cell, bounds))
            for row in range(min_row, max_row + 1):
                for column in range(min_col, max_col + 1):
                    destination = f"{get_column_letter(column)}{row}"
                    translated = Translator(formula, origin=source_cell).translate_formula(
                        destination
                    )
                    worksheet[destination] = translated
                    if destination in sample_coordinates:
                        samples.append({"cell": destination, "formula": translated})
                    count += 1
            warnings = _fill_formula_warnings(
                formula,
                source_cell.replace("$", ""),
                bounds,
                samples,
            )
            return {
                "ok": True,
                "sheet": sheet,
                "range": normalized_target_range,
                "requested_range": target_range,
                "target_range_expanded_from_endpoint": expanded_from_endpoint,
                "cells_filled": count,
                "source_formula": formula,
                "sample_formulas": samples,
                "warnings": warnings,
            }

        return self._mutate(
            "fill_formula",
            {
                "sheet": sheet,
                "source_cell": source_cell,
                "target_range": target_range,
                "normalized_target_range": normalized_target_range,
            },
            apply,
            declaration_id=declaration_id,
        )

    def clear_range(
        self,
        sheet: str,
        range_ref: str,
        *,
        contents: bool = True,
        formats: bool = False,
        declaration_id: int | None = None,
    ) -> dict[str, Any]:
        if not contents and not formats:
            raise ToolInputError("At least one of contents or formats must be true")
        bounds = self._bounds(range_ref)

        def apply(workbook: Any) -> dict[str, Any]:
            worksheet = self._sheet(workbook, sheet)
            min_col, min_row, max_col, max_row = bounds
            count = 0
            for row in range(min_row, max_row + 1):
                for column in range(min_col, max_col + 1):
                    cell = worksheet.cell(row, column)
                    if contents:
                        cell.value = None
                        cell.comment = None
                        cell.hyperlink = None
                    if formats:
                        cell._style = StyleArray()
                    count += 1
            return {"ok": True, "sheet": sheet, "range": range_ref, "cells_cleared": count}

        return self._mutate(
            "clear_range",
            {"sheet": sheet, "range_ref": range_ref, "contents": contents, "formats": formats},
            apply,
            declaration_id=declaration_id,
        )

    def format_range(
        self,
        sheet: str,
        range_ref: str,
        format_spec: dict[str, Any],
        *,
        declaration_id: int | None = None,
    ) -> dict[str, Any]:
        if not format_spec:
            raise ToolInputError("format_spec must not be empty")
        bounds = self._bounds(range_ref)
        supported = {
            "number_format",
            "font",
            "fill_color",
            "alignment",
            "border",
            "protection",
            "row_height",
            "column_width",
        }
        unknown = set(format_spec) - supported
        if unknown:
            raise ToolInputError(f"Unsupported format keys: {sorted(unknown)}")

        def apply(workbook: Any) -> dict[str, Any]:
            worksheet = self._sheet(workbook, sheet)
            min_col, min_row, max_col, max_row = bounds
            for row in range(min_row, max_row + 1):
                if "row_height" in format_spec:
                    worksheet.row_dimensions[row].height = float(format_spec["row_height"])
                for column in range(min_col, max_col + 1):
                    if "column_width" in format_spec:
                        worksheet.column_dimensions[get_column_letter(column)].width = float(
                            format_spec["column_width"]
                        )
                    cell = worksheet.cell(row, column)
                    if "number_format" in format_spec:
                        cell.number_format = str(format_spec["number_format"])
                    if "fill_color" in format_spec:
                        cell.fill = PatternFill(
                            fill_type="solid", fgColor=_normalize_color(format_spec["fill_color"])
                        )
                    if "font" in format_spec:
                        font_spec = dict(format_spec["font"])
                        if "color" in font_spec:
                            font_spec["color"] = _normalize_color(font_spec["color"])
                        base = copy(cell.font)
                        for key, value in font_spec.items():
                            if not hasattr(base, key):
                                raise ToolInputError(f"Unsupported font property: {key}")
                            setattr(base, key, value)
                        cell.font = base
                    if "alignment" in format_spec:
                        alignment = copy(cell.alignment)
                        for key, value in dict(format_spec["alignment"]).items():
                            if not hasattr(alignment, key):
                                raise ToolInputError(f"Unsupported alignment property: {key}")
                            setattr(alignment, key, value)
                        cell.alignment = alignment
                    if "protection" in format_spec:
                        protection = copy(cell.protection)
                        for key, value in dict(format_spec["protection"]).items():
                            if not hasattr(protection, key):
                                raise ToolInputError(f"Unsupported protection property: {key}")
                            setattr(protection, key, value)
                        cell.protection = protection
                    if "border" in format_spec:
                        border_spec = dict(format_spec["border"])
                        style = border_spec.get("style", "thin")
                        color = _normalize_color(border_spec.get("color", "000000"))
                        side = Side(style=style, color=color)
                        sides = border_spec.get("sides", ["left", "right", "top", "bottom"])
                        current = copy(cell.border)
                        values = {
                            name: getattr(current, name)
                            for name in ("left", "right", "top", "bottom")
                        }
                        for name in sides:
                            if name not in values:
                                raise ToolInputError(f"Unsupported border side: {name}")
                            values[name] = side
                        cell.border = Border(**values)
            count = (max_col - min_col + 1) * (max_row - min_row + 1)
            return {"ok": True, "sheet": sheet, "range": range_ref, "cells_formatted": count}

        return self._mutate(
            "format_range",
            {"sheet": sheet, "range_ref": range_ref, "format_spec": format_spec},
            apply,
            declaration_id=declaration_id,
        )

    def delete_rows(
        self,
        sheet: str,
        start: int,
        amount: int = 1,
        *,
        declaration_id: int | None = None,
    ) -> dict[str, Any]:
        if start < 1 or amount < 1 or amount > 10_000:
            raise ToolInputError("start and amount must be positive; amount limit is 10,000")

        def apply(workbook: Any) -> dict[str, Any]:
            self._sheet(workbook, sheet).delete_rows(start, amount)
            return {"ok": True, "sheet": sheet, "start": start, "rows_deleted": amount}

        return self._mutate(
            "delete_rows",
            {"sheet": sheet, "start": start, "amount": amount},
            apply,
            declaration_id=declaration_id,
        )

    def delete_columns(
        self,
        sheet: str,
        start: int,
        amount: int = 1,
        *,
        declaration_id: int | None = None,
    ) -> dict[str, Any]:
        if start < 1 or amount < 1 or amount > 1_000:
            raise ToolInputError("start and amount must be positive; amount limit is 1,000")

        def apply(workbook: Any) -> dict[str, Any]:
            self._sheet(workbook, sheet).delete_cols(start, amount)
            return {"ok": True, "sheet": sheet, "start": start, "columns_deleted": amount}

        return self._mutate(
            "delete_columns",
            {"sheet": sheet, "start": start, "amount": amount},
            apply,
            declaration_id=declaration_id,
        )

    def manage_sheet(
        self,
        action: str,
        name: str,
        *,
        new_name: str | None = None,
        source: str | None = None,
        index: int | None = None,
        declaration_id: int | None = None,
    ) -> dict[str, Any]:
        if action not in {"create", "rename", "delete", "copy"}:
            raise ToolInputError("action must be create, rename, delete, or copy")

        def apply(workbook: Any) -> dict[str, Any]:
            if action == "create":
                if name in workbook.sheetnames:
                    raise ToolInputError(f"Sheet already exists: {name}")
                workbook.create_sheet(name, index)
            elif action == "rename":
                if not new_name:
                    raise ToolInputError("new_name is required for rename")
                self._sheet(workbook, name).title = new_name
            elif action == "delete":
                if len(workbook.sheetnames) == 1:
                    raise ToolInputError("Cannot delete the only worksheet")
                workbook.remove(self._sheet(workbook, name))
            else:
                origin = self._sheet(workbook, source or name)
                copied = workbook.copy_worksheet(origin)
                copied.title = new_name or f"{origin.title} Copy"
            return {"ok": True, "action": action, "sheets": workbook.sheetnames}

        return self._mutate(
            "manage_sheet",
            {
                "action": action,
                "name": name,
                "new_name": new_name,
                "source": source,
                "index": index,
            },
            apply,
            declaration_id=declaration_id,
        )

    def undo_last(self, *, declaration_id: int | None = None) -> dict[str, Any]:
        with self._write_lock:
            self._assert_artifact_sync_locked()
            artifact_before = self._artifact
            snapshots = sorted(self.paths.snapshots.glob(f"*{self.workbook_path.suffix}"))
            if not snapshots:
                raise ToolInputError("No snapshot is available")
            snapshot = snapshots[-1]
            temporary = self.workbook_path.with_name(
                f".{self.workbook_path.stem}.undo-{uuid.uuid4().hex}{self.workbook_path.suffix}"
            )
            recovery = self.workbook_path.with_name(
                f".{self.workbook_path.stem}.undo-recovery-{uuid.uuid4().hex}"
                f"{self.workbook_path.suffix}"
            )
            shutil.copy2(snapshot, temporary)
            shutil.copy2(self.workbook_path, recovery)
            target_prepared: (
                PreparedTargetAuthorization | PreparedAdvisoryTargetAssessment | None
            ) = None
            try:
                self._validate(temporary)
                after_sha256 = _sha256(temporary)
                workbook_diff = diff_workbooks(self.workbook_path, temporary)
                try:
                    target_prepared = self._prepare_target_authorization_locked(
                        declaration_id=declaration_id,
                        diff=workbook_diff,
                        after_sha256=after_sha256,
                    )
                except TargetGroundingRejected as rejected:
                    rejection = self._target_rejection_result_locked(
                        artifact_before=artifact_before,
                        rejected_sha256=after_sha256,
                        diff=workbook_diff,
                        rejected=rejected,
                    )
                    self.recorder.record(
                        "target_grounding.rejected",
                        {
                            "mode": TargetGroundingMode.ENFORCE.value,
                            "decision": rejected.decision.value,
                            "operation": "undo_last",
                            "result": rejection,
                        },
                    )
                    return rejection
                planned_transition = self._planned_transition_locked(
                    operation="undo_last",
                    kind="undo",
                    before_sha256=artifact_before.sha256,
                    after_sha256=after_sha256,
                )
                if self._target_grounding is not None and target_prepared is not None:
                    self._validate_prepared_target_transition_locked(
                        target_prepared,
                        planned_transition,
                    )
                temporary.replace(self.workbook_path)
                transition = self._publish_artifact_locked(
                    operation="undo_last",
                    kind="undo",
                    before_sha256=artifact_before.sha256,
                    after_sha256=after_sha256,
                    target_prepared=target_prepared,
                )
                snapshot.unlink()
                self.recorder.record("workbook.undo", {"snapshot": snapshot})
                result = {
                    "ok": True,
                    "restored_snapshot": snapshot.name,
                    "workbook_sha256_before": artifact_before.sha256,
                    "workbook_sha256_after": after_sha256,
                    "workbook_changed": artifact_before.sha256 != after_sha256,
                    "artifact_revision_before": artifact_before.revision,
                    "artifact_revision_after": self._artifact.revision,
                    "artifact_transition_id": (
                        transition.transition_id if transition is not None else None
                    ),
                    "workbook_effects": workbook_diff.to_dict(),
                }
                if target_prepared is not None:
                    target_diagnostic = self._target_model_diagnostic(target_prepared)
                    if target_diagnostic is not None:
                        result["target_grounding"] = target_diagnostic
                return result
            except Exception:
                if self._artifact == artifact_before:
                    if _sha256(self.workbook_path) != artifact_before.sha256:
                        self._restore_managed_artifact_locked(recovery)
                    if target_prepared is not None and self._target_grounding is not None:
                        try:
                            self._abort_prepared_target_locked(target_prepared)
                        except TargetGroundingError:
                            pass
                raise
            finally:
                temporary.unlink(missing_ok=True)
                recovery.unlink(missing_ok=True)

    def recalculate(self, *, timeout_seconds: float = 120.0) -> dict[str, Any]:
        """Recalculate with LibreOffice while preserving a pre-operation snapshot."""

        from .render import recalculate_workbook

        if isinstance(timeout_seconds, bool) or not isinstance(timeout_seconds, int | float):
            raise ToolInputError("timeout_seconds must be numeric")
        if timeout_seconds <= 0:
            raise ToolInputError("timeout_seconds must be positive")

        with self._write_lock:
            self._assert_artifact_sync_locked()
            artifact_before = self._artifact
            self._snapshot_counter += 1
            snapshot = (
                self.paths.snapshots
                / f"{self._snapshot_counter:04d}_recalculate{self.workbook_path.suffix}"
            )
            shutil.copy2(self.workbook_path, snapshot)
            self.recorder.record(
                "workbook.mutation.started",
                {"operation": "recalculate", "arguments": {}, "snapshot": snapshot},
            )
            keep_snapshot = False
            try:
                metadata = recalculate_workbook(
                    self.workbook_path,
                    self.workbook_path,
                    timeout_seconds=float(timeout_seconds),
                )
                self._validate(self.workbook_path)
                workbook_effects = diff_workbooks(snapshot, self.workbook_path).to_dict()
                after_sha256 = _sha256(self.workbook_path)
                transition = self._publish_artifact_locked(
                    operation="recalculate",
                    kind="derived_recalculation",
                    before_sha256=artifact_before.sha256,
                    after_sha256=after_sha256,
                )
                keep_snapshot = transition is not None
                metadata = {
                    **metadata,
                    "workbook_sha256_before": artifact_before.sha256,
                    "workbook_sha256_after": after_sha256,
                    "workbook_changed": artifact_before.sha256 != after_sha256,
                    "artifact_revision_before": artifact_before.revision,
                    "artifact_revision_after": self._artifact.revision,
                    "artifact_transition_id": (
                        transition.transition_id if transition is not None else None
                    ),
                    "workbook_effects": workbook_effects,
                }
                self.recorder.record(
                    "workbook.mutation.committed",
                    {"operation": "recalculate", "snapshot": snapshot, "result": metadata},
                )
                return metadata
            except Exception as exc:
                # The renderer publishes atomically, but restore explicitly in case a
                # platform-specific replace succeeded immediately before validation.
                if self._artifact == artifact_before:
                    self._restore_managed_artifact_locked(snapshot)
                try:
                    self.recorder.record(
                        "workbook.mutation.rolled_back",
                        {
                            "operation": "recalculate",
                            "snapshot": snapshot,
                            "error": str(exc),
                        },
                    )
                except Exception:
                    pass
                raise
            finally:
                if not keep_snapshot:
                    snapshot.unlink(missing_ok=True)

    def recalculate_for_finalization(
        self,
        *,
        timeout_seconds: float = 120.0,
    ) -> dict[str, Any]:
        """Return a byte-identical no-op only after an independent formula scan.

        Formula-bearing and ambiguous packages fail before any snapshot, trajectory
        event, artifact transition, or LibreOffice process is created. A future
        cache-preserving formula path can extend this boundary transactionally.
        """

        from .ooxml_formula_scan import (
            OOXML_FORMULA_SCAN_SCHEMA_VERSION,
            OOXML_NO_FORMULA_BACKEND,
            OOXML_NO_FORMULA_PROFILE,
            OOXMLFormulaScanError,
            OOXMLFormulaScanLease,
        )

        if isinstance(timeout_seconds, bool) or not isinstance(timeout_seconds, int | float):
            raise ToolInputError("timeout_seconds must be numeric")
        if timeout_seconds <= 0:
            raise ToolInputError("timeout_seconds must be positive")

        with self._write_lock:
            self._assert_artifact_sync_locked()
            artifact = self._artifact
            if hashlib.sha256(self._artifact_bytes).hexdigest() != artifact.sha256:
                raise WorkbookValidationError(
                    "Cached artifact bytes do not match the finalization candidate"
                )
            try:
                with OOXMLFormulaScanLease.open(self.workbook_path) as scan_lease:
                    scan = scan_lease.scan
                    if (
                        scan.package_sha256 != artifact.sha256
                        or scan_lease.snapshot_bytes != self._artifact_bytes
                    ):
                        raise WorkbookValidationError(
                            "Formula scan is not bound to the managed artifact revision"
                        )
                    if scan.has_formulas:
                        raise WorkbookValidationError(
                            "Formula-bearing finalization requires a cache-preserving "
                            "transactional recalculation backend"
                        )
                    result = {
                        "backend": OOXML_NO_FORMULA_BACKEND,
                        "version": OOXML_FORMULA_SCAN_SCHEMA_VERSION,
                        "profile": OOXML_NO_FORMULA_PROFILE,
                        "format": scan.workbook_format,
                        "source_sha256": artifact.sha256,
                        "output_sha256": artifact.sha256,
                        "atomic_replace": False,
                        "publication": "verified_no_write",
                        "workbook_sha256_before": artifact.sha256,
                        "workbook_sha256_after": artifact.sha256,
                        "workbook_changed": False,
                        "artifact_revision_before": artifact.revision,
                        "artifact_revision_after": artifact.revision,
                        "artifact_transition_id": None,
                        "formula_scan": scan.to_dict(),
                        "workbook_effects": {
                            "schema_version": "workbook-effect-diff-v1",
                            "semantic_changed": False,
                            "complete": True,
                            "effects": [],
                            "scope": EvidenceScope().to_dict(),
                            "formula_scope": EvidenceScope().to_dict(),
                            "changed_cell_count": 0,
                            "scanned_cell_count": scan.scanned_cell_count,
                            "reasons": [],
                        },
                    }
                    scan_lease.verify_binding(checkpoint="session_finalization")
                    return result
            except OOXMLFormulaScanError as exc:
                raise WorkbookValidationError(
                    "Finalization cannot certify a safe no-formula recalculation"
                ) from exc

    def write_manifest(self, values: dict[str, Any]) -> Path:
        path = self.paths.root / "run.json"
        temporary = path.with_suffix(".json.tmp")
        temporary.write_text(
            json.dumps(values, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
        )
        temporary.replace(path)
        return path
