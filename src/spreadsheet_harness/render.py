"""Safe LibreOffice rendering for spreadsheet workbooks.

The source workbook is never passed to LibreOffice directly.  Every conversion
uses a private copy and a fresh ``UserInstallation`` directory so concurrent
renders cannot share a LibreOffice profile or modify the caller's workbook.
"""

from __future__ import annotations

import hashlib
import json
import os
import posixpath
import re
import shutil
import subprocess
import tempfile
import zipfile
import zlib
from collections.abc import Iterator, Sequence
from contextlib import contextmanager
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from urllib.parse import unquote
from xml.etree import ElementTree

from openpyxl import load_workbook

from .errors import RecalculationIntegrityError, RenderError, ScoringInfrastructureError

SUPPORTED_SPREADSHEET_EXTENSIONS = frozenset({".xlsx", ".xlsm", ".ods", ".xls", ".csv"})
PNG_SIGNATURE = b"\x89PNG\r\n\x1a\n"
_RECALCULATION_FORMATS = {
    ".xlsx": "xlsx:Calc MS Excel 2007 XML",
    ".xlsm": "xlsm:Calc MS Excel 2007 VBA XML",
    ".ods": "ods:calc8",
    ".xls": "xls:MS Excel 97",
    ".csv": "csv:Text - txt - csv (StarCalc)",
}
RECALCULATION_SHEET_INTEGRITY_POLICY = "exact-ordered-sheet-kind-name-visibility-v2"
_SHEET_INVENTORY_FORMATS = frozenset({".xlsx", ".xlsm"})
_WORKBOOK_XML_PART = "xl/workbook.xml"
_WORKBOOK_RELATIONSHIPS_PART = "xl/_rels/workbook.xml.rels"
_CONTENT_TYPES_PART = "[Content_Types].xml"
_OOXML_INVENTORY_PART_MAX_BYTES = 8 * 1024 * 1024
_TRANSITIONAL_SPREADSHEETML_NAMESPACE = (
    "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
)
_STRICT_SPREADSHEETML_NAMESPACE = "http://purl.oclc.org/ooxml/spreadsheetml/main"
_RELATIONSHIP_NAMESPACES = {
    _TRANSITIONAL_SPREADSHEETML_NAMESPACE: (
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    ),
    _STRICT_SPREADSHEETML_NAMESPACE: (
        "http://purl.oclc.org/ooxml/officeDocument/relationships"
    ),
}
_PACKAGE_RELATIONSHIP_NAMESPACES = frozenset(
    {
        "http://schemas.openxmlformats.org/package/2006/relationships",
        "http://purl.oclc.org/ooxml/package/relationships",
    }
)
_CONTENT_TYPE_NAMESPACES = frozenset(
    {
        "http://schemas.openxmlformats.org/package/2006/content-types",
        "http://purl.oclc.org/ooxml/package/content-types",
    }
)
_SHEET_RELATIONSHIP_KINDS = frozenset(
    {"worksheet", "chartsheet", "dialogsheet", "macrosheet", "intlMacrosheet"}
)
_SHEET_RELATIONSHIP_TYPE_PREFIXES = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/",
    "http://purl.oclc.org/ooxml/officeDocument/relationships/",
)
_POSITIVE_INTEGER = re.compile(r"[1-9][0-9]*\Z")


@dataclass(frozen=True)
class RenderPage:
    """One rasterized PDF page."""

    index: int
    path: Path
    sha256: str
    width: int
    height: int
    sheet: str | None = None
    sheet_page: int | None = None

    def to_dict(self, *, relative_to: Path | None = None) -> dict[str, Any]:
        path = self.path
        if relative_to is not None:
            try:
                path = path.relative_to(relative_to)
            except ValueError:
                pass
        return {
            "index": self.index,
            "page": self.sheet_page if self.sheet_page is not None else self.index,
            "path": path.as_posix(),
            "image_path": str(self.path.resolve()),
            "sha256": self.sha256,
            "width": self.width,
            "height": self.height,
            "sheet": self.sheet,
            "sheet_page": self.sheet_page,
        }


@dataclass(frozen=True)
class RenderResult:
    """Published render artifacts and their reproducibility metadata."""

    source: Path
    output_dir: Path
    manifest_path: Path
    backend: str
    version: dict[str, str]
    source_sha256: str
    mode: str
    dpi: int
    pages: tuple[RenderPage, ...]

    def to_dict(self) -> dict[str, Any]:
        return {
            "schema_version": 1,
            "source": {
                "name": self.source.name,
                "format": self.source.suffix.lower().lstrip("."),
                "sha256": self.source_sha256,
            },
            "backend": self.backend,
            "version": dict(self.version),
            "hash": self.source_sha256,
            "mode": self.mode,
            "dpi": self.dpi,
            "manifest_path": str(self.manifest_path.resolve()),
            "page_count": len(self.pages),
            "pages": [page.to_dict(relative_to=self.output_dir) for page in self.pages],
        }


@dataclass(frozen=True)
class _RasterizedPage:
    path: Path
    width: int
    height: int
    sheet: str | None
    sheet_page: int | None


def sha256_file(path: str | Path) -> str:
    """Return the SHA-256 digest of *path* without loading it all in memory."""

    digest = hashlib.sha256()
    with Path(path).open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _sheet_inventory_sha256(sheets: list[dict[str, Any]]) -> str:
    encoded = json.dumps(
        sheets,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _read_unique_inventory_part(
    package: zipfile.ZipFile,
    part_name: str,
) -> bytes:
    matches = [member for member in package.infolist() if member.filename == part_name]
    if len(matches) != 1:
        raise RenderError(
            f"OOXML package must contain exactly one {part_name}; found {len(matches)}"
        )
    member = matches[0]
    if member.is_dir() or member.flag_bits & 0x1:
        raise RenderError(f"OOXML part {part_name} is not a readable regular ZIP member")
    if (
        member.file_size <= 0
        or member.file_size > _OOXML_INVENTORY_PART_MAX_BYTES
    ):
        raise RenderError(
            f"OOXML part {part_name} size is outside the accepted bound of "
            f"1..{_OOXML_INVENTORY_PART_MAX_BYTES} bytes"
        )
    with package.open(member) as handle:
        raw = handle.read(_OOXML_INVENTORY_PART_MAX_BYTES + 1)
    if (
        len(raw) != member.file_size
        or len(raw) > _OOXML_INVENTORY_PART_MAX_BYTES
    ):
        raise RenderError(f"OOXML part {part_name} size does not match its ZIP metadata")
    return raw


def _read_sheet_inventory_parts(workbook_path: Path) -> tuple[bytes, bytes]:
    """Read the unique bounded inventory parts without extracting ZIP members."""

    try:
        with zipfile.ZipFile(workbook_path) as package:
            workbook_xml = _read_unique_inventory_part(package, _WORKBOOK_XML_PART)
            relationships_xml = _read_unique_inventory_part(
                package,
                _WORKBOOK_RELATIONSHIPS_PART,
            )
            return workbook_xml, relationships_xml
    except RenderError:
        raise
    except (
        OSError,
        RuntimeError,
        EOFError,
        ValueError,
        NotImplementedError,
        zipfile.BadZipFile,
        zipfile.LargeZipFile,
        zlib.error,
    ) as exc:
        raise RenderError(
            f"Could not read OOXML workbook package: {type(exc).__name__}: {exc}"
        ) from exc


def _xml_namespace(tag: Any) -> str | None:
    if not isinstance(tag, str) or not tag.startswith("{") or "}" not in tag:
        return None
    return tag[1 : tag.index("}")]


def _xml_local_name(tag: Any) -> str | None:
    if not isinstance(tag, str):
        return None
    return tag.rsplit("}", 1)[-1] if tag.startswith("{") else tag


def _parse_inventory_xml(xml: bytes, *, label: str) -> ElementTree.Element:
    # ElementTree does not fetch external resources, and rejecting DTD/entity
    # declarations also prevents internal entity-expansion payloads.
    declaration_scan = xml.replace(b"\x00", b"").upper()
    if b"<!DOCTYPE" in declaration_scan or b"<!ENTITY" in declaration_scan:
        raise RenderError(f"OOXML {label} XML must not contain DTD or entity declarations")
    try:
        return ElementTree.fromstring(xml)
    except (ElementTree.ParseError, LookupError, ValueError) as exc:
        raise RenderError(f"OOXML {label} XML is malformed: {exc}") from exc


def _parse_workbook_relationships(
    relationships_xml: bytes,
) -> dict[str, dict[str, str]]:
    root = _parse_inventory_xml(relationships_xml, label="workbook relationships")
    namespace = _xml_namespace(root.tag)
    if (
        namespace not in _PACKAGE_RELATIONSHIP_NAMESPACES
        or root.tag != f"{{{namespace}}}Relationships"
    ):
        raise RenderError("OOXML workbook relationships use an unsupported namespace")

    relationship_tag = f"{{{namespace}}}Relationship"
    relationships: dict[str, dict[str, str]] = {}
    for relationship in root:
        if relationship.tag != relationship_tag or list(relationship):
            raise RenderError("OOXML workbook relationships contain an invalid child")
        relationship_id = relationship.attrib.get("Id")
        relationship_type = relationship.attrib.get("Type")
        target = relationship.attrib.get("Target")
        target_mode = relationship.attrib.get("TargetMode", "Internal")
        if not isinstance(relationship_id, str) or not relationship_id:
            raise RenderError("OOXML workbook relationship is missing a non-empty Id")
        if relationship_id in relationships:
            raise RenderError("OOXML workbook relationships contain duplicate Id values")
        if not isinstance(relationship_type, str) or not relationship_type:
            raise RenderError("OOXML workbook relationship is missing a non-empty Type")
        if not isinstance(target, str) or not target:
            raise RenderError("OOXML workbook relationship is missing a non-empty Target")
        if target_mode not in {"Internal", "External"}:
            raise RenderError("OOXML workbook relationship has an invalid TargetMode")
        relationships[relationship_id] = {
            "type": relationship_type,
            "target": target,
            "target_mode": target_mode,
        }
    return relationships


def _sheet_kind(relationship: dict[str, str]) -> str:
    if relationship["target_mode"] != "Internal":
        raise RenderError("OOXML sheet relationship must target an internal package part")
    relationship_type = relationship["type"]
    for prefix in _SHEET_RELATIONSHIP_TYPE_PREFIXES:
        if relationship_type.startswith(prefix):
            kind = relationship_type[len(prefix) :]
            if kind in _SHEET_RELATIONSHIP_KINDS:
                return kind
            break
    raise RenderError("OOXML sheet relationship has an unsupported sheet type")


def _parse_sheet_inventory(
    workbook_xml: bytes,
    relationships_xml: bytes,
) -> list[dict[str, Any]]:
    root = _parse_inventory_xml(workbook_xml, label="workbook")
    relationships = _parse_workbook_relationships(relationships_xml)

    namespace = _xml_namespace(root.tag)
    if (
        namespace not in _RELATIONSHIP_NAMESPACES
        or root.tag != f"{{{namespace}}}workbook"
    ):
        raise RenderError("OOXML workbook root uses an unsupported SpreadsheetML namespace")

    sheets_tag = f"{{{namespace}}}sheets"
    sheet_tag = f"{{{namespace}}}sheet"
    sheets_nodes = [child for child in root if child.tag == sheets_tag]
    if len(sheets_nodes) != 1:
        raise RenderError(
            f"OOXML workbook must contain exactly one sheets element; found {len(sheets_nodes)}"
        )
    if any(
        _xml_local_name(child.tag) == "sheets" and child.tag != sheets_tag
        for child in root
    ):
        raise RenderError("OOXML workbook contains a sheets element in the wrong namespace")

    relationship_attribute = f"{{{_RELATIONSHIP_NAMESPACES[namespace]}}}id"
    sheets: list[dict[str, Any]] = []
    names: set[str] = set()
    sheet_ids: set[str] = set()
    relationship_ids: set[str] = set()
    for index, sheet in enumerate(sheets_nodes[0]):
        if sheet.tag != sheet_tag:
            raise RenderError("OOXML sheets element contains a non-sheet child")
        if list(sheet):
            raise RenderError("OOXML sheet records must not contain child elements")

        name = sheet.attrib.get("name")
        sheet_id = sheet.attrib.get("sheetId")
        relationship_id = sheet.attrib.get(relationship_attribute)
        visibility = sheet.attrib.get("state", "visible")
        if not isinstance(name, str) or not name:
            raise RenderError("OOXML sheet record is missing a non-empty name")
        normalized_name = name.casefold()
        if normalized_name in names:
            raise RenderError("OOXML workbook contains duplicate sheet names")
        if not isinstance(sheet_id, str) or _POSITIVE_INTEGER.fullmatch(sheet_id) is None:
            raise RenderError("OOXML sheet record has an invalid or missing sheetId")
        if sheet_id in sheet_ids:
            raise RenderError("OOXML workbook contains duplicate sheetId values")
        if not isinstance(relationship_id, str) or not relationship_id:
            raise RenderError(
                "OOXML sheet record has an invalid or missing relationship identifier"
            )
        if relationship_id in relationship_ids:
            raise RenderError("OOXML workbook contains duplicate sheet relationships")
        relationship = relationships.get(relationship_id)
        if relationship is None:
            raise RenderError("OOXML sheet references a missing workbook relationship")
        kind = _sheet_kind(relationship)
        if visibility not in {"visible", "hidden", "veryHidden"}:
            raise RenderError("OOXML sheet record has an invalid visibility state")

        names.add(normalized_name)
        sheet_ids.add(sheet_id)
        relationship_ids.add(relationship_id)
        sheets.append(
            {
                "index": index,
                "kind": kind,
                "name": name,
                "visibility": visibility,
            }
        )

    if not sheets:
        raise RenderError("OOXML workbook contains no sheets")
    if not any(sheet["visibility"] == "visible" for sheet in sheets):
        raise RenderError("OOXML workbook contains no visible sheet")
    return sheets


def sheet_inventory_identity(path: str | Path) -> dict[str, Any]:
    """Return ordered sheet kinds, names, and visibility bound to the file hash."""

    workbook_path = Path(path).expanduser().resolve()
    suffix = workbook_path.suffix.lower()
    if suffix not in _SHEET_INVENTORY_FORMATS:
        supported = ", ".join(sorted(_SHEET_INVENTORY_FORMATS))
        raise RenderError(
            f"Sheet inventory integrity requires {supported}; got {suffix!r}"
        )
    workbook_sha256 = sha256_file(workbook_path)
    try:
        workbook_xml, relationships_xml = _read_sheet_inventory_parts(workbook_path)
        sheets = _parse_sheet_inventory(workbook_xml, relationships_xml)
    except RenderError:
        raise
    except Exception as exc:
        raise RenderError(f"Could not read workbook sheet inventory: {exc}") from exc
    if sha256_file(workbook_path) != workbook_sha256:
        raise RenderError("Workbook changed while its sheet inventory was being read")
    return {
        "schema_version": 2,
        "workbook_sha256": workbook_sha256,
        "inventory_sha256": _sheet_inventory_sha256(sheets),
        "sheets": sheets,
    }


def _relationship_target_part(source_part: str, target: str) -> str:
    decoded = unquote(target)
    if not decoded or "\\" in decoded or "\x00" in decoded:
        raise RenderError("OOXML sheet relationship target is not a valid package path")
    if decoded.startswith("/"):
        resolved = posixpath.normpath(decoded.lstrip("/"))
    else:
        resolved = posixpath.normpath(
            posixpath.join(posixpath.dirname(source_part), decoded)
        )
    if resolved in {"", ".", ".."} or resolved.startswith("../"):
        raise RenderError("OOXML sheet relationship target escapes the package")
    return resolved


def _part_relationships_name(part_name: str) -> str:
    directory, filename = posixpath.split(part_name)
    return posixpath.join(directory, "_rels", f"{filename}.rels")


def _worksheet_only_package_parts(
    workbook_xml: bytes,
    relationships_xml: bytes,
    content_types_xml: bytes,
    sheets: list[dict[str, Any]],
) -> tuple[bytes, bytes, bytes, set[str]]:
    workbook_root = _parse_inventory_xml(workbook_xml, label="workbook")
    workbook_namespace = _xml_namespace(workbook_root.tag)
    if workbook_namespace not in _RELATIONSHIP_NAMESPACES:
        raise RenderError("OOXML workbook root uses an unsupported SpreadsheetML namespace")
    sheets_node = workbook_root.find(f"{{{workbook_namespace}}}sheets")
    if sheets_node is None or len(sheets_node) != len(sheets):
        raise RenderError("OOXML workbook sheet inventory changed while building scorer view")
    relationship_attribute = (
        f"{{{_RELATIONSHIP_NAMESPACES[workbook_namespace]}}}id"
    )
    removed_relationship_ids = {
        node.attrib.get(relationship_attribute)
        for node, sheet in zip(list(sheets_node), sheets, strict=True)
        if sheet["kind"] != "worksheet"
    }
    if None in removed_relationship_ids:
        raise RenderError("OOXML non-worksheet sheet is missing its relationship identifier")
    for node, sheet in zip(list(sheets_node), sheets, strict=True):
        if sheet["kind"] != "worksheet":
            sheets_node.remove(node)

    relationships_root = _parse_inventory_xml(
        relationships_xml,
        label="workbook relationships",
    )
    removed_parts: set[str] = set()
    removed_relationship_count = 0
    for relationship in list(relationships_root):
        if relationship.attrib.get("Id") not in removed_relationship_ids:
            continue
        removed_relationship_count += 1
        removed_part = _relationship_target_part(
            _WORKBOOK_XML_PART,
            relationship.attrib.get("Target", ""),
        )
        removed_parts.add(removed_part)
        removed_parts.add(_part_relationships_name(removed_part))
        relationships_root.remove(relationship)
    if removed_relationship_count != len(removed_relationship_ids):
        raise RenderError("OOXML non-worksheet relationship set changed during scorer view")

    content_types_root = _parse_inventory_xml(content_types_xml, label="content types")
    content_types_namespace = _xml_namespace(content_types_root.tag)
    if (
        content_types_namespace not in _CONTENT_TYPE_NAMESPACES
        or content_types_root.tag != f"{{{content_types_namespace}}}Types"
    ):
        raise RenderError("OOXML content types use an unsupported namespace")
    override_tag = f"{{{content_types_namespace}}}Override"
    for child in list(content_types_root):
        part_name = child.attrib.get("PartName") if child.tag == override_tag else None
        if isinstance(part_name, str) and part_name.lstrip("/") in removed_parts:
            content_types_root.remove(child)

    return (
        ElementTree.tostring(workbook_root, encoding="utf-8", xml_declaration=True),
        ElementTree.tostring(
            relationships_root,
            encoding="utf-8",
            xml_declaration=True,
        ),
        ElementTree.tostring(
            content_types_root,
            encoding="utf-8",
            xml_declaration=True,
        ),
        removed_parts,
    )


@contextmanager
def openpyxl_worksheet_view(
    path: str | Path,
) -> Iterator[tuple[Path, dict[str, Any]]]:
    """Yield an immutable OOXML view containing only worksheet sheet records."""

    workbook_path = Path(path).expanduser().resolve()
    identity = sheet_inventory_identity(workbook_path)
    sheets = identity["sheets"]
    if all(sheet["kind"] == "worksheet" for sheet in sheets):
        yield workbook_path, identity
        return
    if not any(sheet["kind"] == "worksheet" for sheet in sheets):
        raise ScoringInfrastructureError(
            "The workbook contains no worksheet that the cell scorer can read"
        )

    try:
        with zipfile.ZipFile(workbook_path) as source_package:
            workbook_xml = _read_unique_inventory_part(
                source_package,
                _WORKBOOK_XML_PART,
            )
            relationships_xml = _read_unique_inventory_part(
                source_package,
                _WORKBOOK_RELATIONSHIPS_PART,
            )
            content_types_xml = _read_unique_inventory_part(
                source_package,
                _CONTENT_TYPES_PART,
            )
            (
                worksheet_workbook_xml,
                worksheet_relationships_xml,
                worksheet_content_types_xml,
                removed_parts,
            ) = _worksheet_only_package_parts(
                workbook_xml,
                relationships_xml,
                content_types_xml,
                sheets,
            )
            member_counts: dict[str, int] = {}
            for member in source_package.infolist():
                member_counts[member.filename] = member_counts.get(member.filename, 0) + 1
            for part in removed_parts:
                if part.endswith(".rels"):
                    continue
                matches = member_counts.get(part, 0)
                if matches != 1:
                    raise RenderError(
                        "OOXML non-worksheet relationship target must resolve to exactly "
                        f"one package part; {part} resolved to {matches}"
                    )

            with tempfile.TemporaryDirectory(prefix="sheet-harness-worksheet-view-") as raw:
                view_path = Path(raw) / f"workbook{workbook_path.suffix.lower()}"
                replacements = {
                    _WORKBOOK_XML_PART: worksheet_workbook_xml,
                    _WORKBOOK_RELATIONSHIPS_PART: worksheet_relationships_xml,
                    _CONTENT_TYPES_PART: worksheet_content_types_xml,
                }
                with zipfile.ZipFile(view_path, "w") as view_package:
                    for member in source_package.infolist():
                        if member.filename in removed_parts:
                            continue
                        data = replacements.get(member.filename)
                        if data is None:
                            with source_package.open(member) as source:
                                data = source.read()
                        view_package.writestr(member, data)
                if sha256_file(workbook_path) != identity["workbook_sha256"]:
                    raise RenderError(
                        "Workbook changed while its worksheet-only scorer view was built"
                    )
                yield view_path, identity
    except (RenderError, ScoringInfrastructureError):
        raise
    except (
        OSError,
        RuntimeError,
        EOFError,
        ValueError,
        NotImplementedError,
        zipfile.BadZipFile,
        zipfile.LargeZipFile,
        zlib.error,
    ) as exc:
        raise RenderError(
            f"Could not build OOXML worksheet-only scorer view: {type(exc).__name__}: {exc}"
        ) from exc


def _publish_recalculation_failure_artifact(
    converted: Path,
    destination: Path,
    output_sha256: str,
) -> Path:
    evidence_path = destination.with_name(
        f"{destination.stem}.recalculation-integrity-failure-"
        f"{output_sha256}{destination.suffix}"
    )
    if evidence_path.exists():
        if evidence_path.is_symlink() or not evidence_path.is_file():
            raise RenderError(
                f"Recalculation failure evidence path is not a regular file: {evidence_path}"
            )
        if sha256_file(evidence_path) != output_sha256:
            raise RenderError(
                f"Recalculation failure evidence path has conflicting content: {evidence_path}"
            )
        return evidence_path

    descriptor, raw_temporary = tempfile.mkstemp(
        dir=destination.parent,
        prefix=f".{destination.stem}.recalculation-integrity-failure-",
        suffix=destination.suffix,
    )
    os.close(descriptor)
    temporary = Path(raw_temporary)
    try:
        shutil.copy2(converted, temporary)
        _validate_recalculated_file(temporary)
        if sha256_file(temporary) != output_sha256:
            raise RenderError("Recalculation failure evidence changed while being published")
        temporary.replace(evidence_path)
    finally:
        temporary.unlink(missing_ok=True)
    return evidence_path


def find_libreoffice(explicit: str | Path | None = None) -> str | None:
    """Locate a LibreOffice/soffice executable, returning ``None`` if absent."""

    if explicit is not None:
        value = str(explicit)
        resolved = shutil.which(value)
        if resolved:
            return resolved
        candidate = Path(value).expanduser()
        if candidate.is_file():
            return str(candidate.resolve())
        return None

    for name in ("libreoffice", "soffice"):
        resolved = shutil.which(name)
        if resolved:
            return resolved

    # Helpful for local development on macOS; Linux normally resolves via PATH.
    macos_binary = Path("/Applications/LibreOffice.app/Contents/MacOS/soffice")
    if macos_binary.is_file():
        return str(macos_binary)
    return None


@contextmanager
def isolated_user_profile() -> Iterator[tuple[Path, str]]:
    """Yield a fresh LibreOffice profile directory and its required file URI."""

    with tempfile.TemporaryDirectory(prefix="spreadsheet-lo-profile-") as raw_profile:
        profile = Path(raw_profile).resolve()
        yield profile, profile.as_uri()


def libreoffice_command(
    binary: str,
    source: Path,
    output_dir: Path,
    target_format: str,
    profile_uri: str,
) -> list[str]:
    """Build a non-interactive conversion command with an isolated profile."""

    return [
        binary,
        "--headless",
        "--nologo",
        "--nodefault",
        "--nolockcheck",
        "--nofirststartwizard",
        f"-env:UserInstallation={profile_uri}",
        "--convert-to",
        target_format,
        "--outdir",
        str(output_dir),
        str(source),
    ]


def libreoffice_version(binary: str, *, timeout_seconds: float = 30.0) -> str:
    """Return a concise LibreOffice version string."""

    try:
        completed = subprocess.run(
            [binary, "--version"],
            check=False,
            capture_output=True,
            text=True,
            timeout=timeout_seconds,
        )
    except (OSError, subprocess.SubprocessError):
        return "unknown"
    version = (completed.stdout or completed.stderr).strip()
    return version or "unknown"


def _converted_candidates(output_dir: Path, stem: str, suffix: str) -> list[Path]:
    expected = output_dir / f"{stem}{suffix}"
    if expected.is_file():
        return [expected]
    return sorted(
        path
        for path in output_dir.iterdir()
        if path.is_file() and path.stem.casefold() == stem.casefold() and path.suffix == suffix
    )


def _convert_with_libreoffice(
    source_copy: Path,
    output_dir: Path,
    *,
    target_format: str,
    binary: str,
    timeout_seconds: float,
) -> Path:
    """Convert a disposable source copy with a fresh LibreOffice profile."""

    output_dir.mkdir(parents=True, exist_ok=True)
    normalized_format = target_format.split(":", 1)[0].lower().lstrip(".")
    suffix = f".{normalized_format}"
    with isolated_user_profile() as (_, profile_uri):
        command = libreoffice_command(
            binary,
            source_copy,
            output_dir,
            target_format,
            profile_uri,
        )
        try:
            completed = subprocess.run(
                command,
                check=False,
                capture_output=True,
                text=True,
                timeout=timeout_seconds,
            )
        except subprocess.TimeoutExpired as exc:
            raise RenderError(
                f"LibreOffice timed out after {timeout_seconds:g}s converting {source_copy.name}"
            ) from exc
        except OSError as exc:
            raise RenderError(f"Could not start LibreOffice: {exc}") from exc

    candidates = _converted_candidates(output_dir, source_copy.stem, suffix)
    if completed.returncode != 0 or not candidates:
        details = (completed.stderr or completed.stdout).strip()
        message = f"LibreOffice failed to convert {source_copy.name} to {normalized_format}"
        if details:
            message += f": {details[-1000:]}"
        raise RenderError(message)
    return candidates[0]


def convert_spreadsheet_copy(
    source: str | Path,
    output_dir: str | Path,
    *,
    target_format: str,
    libreoffice_binary: str | Path | None = None,
    timeout_seconds: float = 120.0,
) -> Path:
    """Convert a spreadsheet without ever handing LibreOffice the original file.

    This helper is also used by preprocessing for legacy XLS and ODS inputs.  It
    deliberately refuses to replace an existing destination artifact.
    """

    source_path = _validate_source(source)
    binary = find_libreoffice(libreoffice_binary)
    if binary is None:
        raise RenderError("LibreOffice executable was not found")
    destination = Path(output_dir).expanduser().resolve()
    destination.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(prefix="spreadsheet-convert-") as raw_work:
        work = Path(raw_work)
        private_source = work / source_path.name
        shutil.copy2(source_path, private_source)
        converted = _convert_with_libreoffice(
            private_source,
            work / "converted",
            target_format=target_format,
            binary=binary,
            timeout_seconds=timeout_seconds,
        )
        published = destination / converted.name
        if published.exists():
            raise RenderError(f"Refusing to overwrite conversion artifact: {published}")
        if published.resolve() == source_path:
            raise RenderError("Conversion output would overwrite the source workbook")
        shutil.copy2(converted, published)
    return published


def _validate_recalculated_file(path: Path) -> None:
    if not path.is_file() or path.stat().st_size == 0:
        raise RenderError("LibreOffice produced an empty recalculated workbook")
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return
    try:
        sheet_inventory_identity(path)
    except Exception as exc:
        raise RenderError(f"Recalculated workbook validation failed: {exc}") from exc


def recalculate_workbook(
    source: str | Path,
    destination: str | Path,
    *,
    libreoffice_binary: str | Path | None = None,
    timeout_seconds: float = 120.0,
) -> dict[str, Any]:
    """Recalculate a private copy and atomically publish the verified result.

    ``source`` and ``destination`` may be the same path for an isolated session
    working copy.  LibreOffice still receives only a temporary copy; the
    destination is replaced only after conversion and validation succeed.
    """

    source_path = _validate_source(source)
    destination_path = Path(destination).expanduser().resolve()
    destination_format = destination_path.suffix.lower()
    target_format = _RECALCULATION_FORMATS.get(destination_format)
    if target_format is None:
        supported = ", ".join(sorted(_RECALCULATION_FORMATS))
        raise RenderError(
            f"Unsupported recalculation destination {destination_format!r}; expected {supported}"
        )
    if timeout_seconds <= 0:
        raise RenderError("timeout_seconds must be positive")
    binary = find_libreoffice(libreoffice_binary)
    if binary is None:
        raise RenderError("LibreOffice executable was not found")

    inventory_enforced = bool(
        source_path.suffix.lower() in _SHEET_INVENTORY_FORMATS
        and destination_format in _SHEET_INVENTORY_FORMATS
    )
    source_identity = (
        sheet_inventory_identity(source_path) if inventory_enforced else None
    )
    source_hash = (
        str(source_identity["workbook_sha256"])
        if source_identity is not None
        else sha256_file(source_path)
    )
    destination_path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(prefix="spreadsheet-recalculate-") as raw_work:
        work = Path(raw_work)
        private_source = work / "source" / source_path.name
        private_source.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(source_path, private_source)
        converted = _convert_with_libreoffice(
            private_source,
            work / "converted",
            target_format=target_format,
            binary=binary,
            timeout_seconds=timeout_seconds,
        )
        _validate_recalculated_file(converted)
        output_identity = (
            sheet_inventory_identity(converted) if inventory_enforced else None
        )
        output_hash = (
            str(output_identity["workbook_sha256"])
            if output_identity is not None
            else sha256_file(converted)
        )
        inventory_integrity = (
            {
                "schema_version": 2,
                "policy": RECALCULATION_SHEET_INTEGRITY_POLICY,
                "enforced": True,
                "matched": source_identity["sheets"] == output_identity["sheets"],
                "pre": source_identity,
                "post": output_identity,
            }
            if source_identity is not None and output_identity is not None
            else {
                "schema_version": 2,
                "policy": RECALCULATION_SHEET_INTEGRITY_POLICY,
                "enforced": False,
                "matched": None,
                "pre": None,
                "post": None,
                "reason": "source-or-destination-is-not-ooxml",
            }
        )
        metadata = {
            "backend": "libreoffice-headless",
            "version": libreoffice_version(binary),
            "profile": "isolated-per-invocation",
            "source_path": str(source_path),
            "destination_path": str(destination_path),
            "source_sha256": source_hash,
            "output_sha256": output_hash,
            "format": destination_format.lstrip("."),
            "sheet_inventory_integrity": inventory_integrity,
        }
        if inventory_integrity["enforced"] and not inventory_integrity["matched"]:
            metadata.update(
                {
                    "atomic_replace": False,
                    "published": False,
                    "failure_artifact_sha256": output_hash,
                }
            )
            try:
                failure_artifact = _publish_recalculation_failure_artifact(
                    converted,
                    destination_path,
                    output_hash,
                )
            except Exception as exc:
                metadata.update(
                    {
                        "failure_artifact_path": None,
                        "failure_artifact_error_type": type(exc).__name__,
                    }
                )
                raise RecalculationIntegrityError(
                    "Recalculation changed sheet identity and its post-recalculation "
                    "artifact could not be preserved",
                    evidence=metadata,
                ) from exc
            metadata["failure_artifact_path"] = str(failure_artifact)
            raise RecalculationIntegrityError(
                "Recalculation changed sheet kind, order, name, or visibility",
                evidence=metadata,
            )

        descriptor, raw_temporary = tempfile.mkstemp(
            dir=destination_path.parent,
            prefix=f".{destination_path.stem}.recalculated-",
            suffix=destination_path.suffix,
        )
        os.close(descriptor)
        temporary_destination = Path(raw_temporary)
        try:
            shutil.copy2(converted, temporary_destination)
            _validate_recalculated_file(temporary_destination)
            if sha256_file(temporary_destination) != output_hash:
                raise RenderError("Recalculated workbook changed before atomic publication")
            temporary_destination.replace(destination_path)
        finally:
            temporary_destination.unlink(missing_ok=True)

    metadata.update({"atomic_replace": True, "published": True})
    return metadata


def _validate_source(source: str | Path) -> Path:
    source_path = Path(source).expanduser().resolve()
    if not source_path.is_file():
        raise RenderError(f"Spreadsheet does not exist or is not a file: {source_path}")
    if source_path.suffix.lower() not in SUPPORTED_SPREADSHEET_EXTENSIONS:
        supported = ", ".join(sorted(SUPPORTED_SPREADSHEET_EXTENSIONS))
        raise RenderError(
            f"Unsupported spreadsheet format {source_path.suffix!r}; expected {supported}"
        )
    return source_path


def _make_single_sheet_copy(source_copy: Path, destination: Path, sheet_name: str) -> None:
    """Create a disposable OOXML copy with only *sheet_name* visible."""

    keep_vba = source_copy.suffix.lower() == ".xlsm"
    workbook = load_workbook(source_copy, read_only=False, keep_vba=keep_vba, data_only=False)
    try:
        if sheet_name not in workbook.sheetnames:
            raise RenderError(f"Worksheet disappeared while rendering: {sheet_name}")
        target_index = workbook.sheetnames.index(sheet_name)
        for worksheet in workbook.worksheets:
            worksheet.sheet_state = "visible" if worksheet.title == sheet_name else "hidden"
        workbook.active = target_index
        if workbook.views:
            workbook.views[0].activeTab = target_index
        workbook.save(destination)
    finally:
        workbook.close()


def _sheet_names(source_copy: Path) -> list[str]:
    keep_vba = source_copy.suffix.lower() == ".xlsm"
    workbook = load_workbook(source_copy, read_only=True, keep_vba=keep_vba, data_only=False)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def _pymupdf_module() -> Any:
    try:
        import pymupdf

        return pymupdf
    except ImportError:
        try:
            import fitz

            return fitz
        except ImportError as exc:  # pragma: no cover - declared project dependency
            raise RenderError("PyMuPDF is required to rasterize LibreOffice PDFs") from exc


def pymupdf_version() -> str:
    module = _pymupdf_module()
    for attribute in ("VersionBind", "__version__"):
        value = getattr(module, attribute, None)
        if value:
            return str(value)
    version_tuple = getattr(module, "version", None)
    if version_tuple:
        return str(version_tuple[0])
    return "unknown"


def _rasterize_pdf(
    pdf_path: Path,
    output_dir: Path,
    *,
    dpi: int,
    filename_prefix: str,
    sheet: str | None,
) -> list[_RasterizedPage]:
    module = _pymupdf_module()
    output_dir.mkdir(parents=True, exist_ok=True)
    try:
        document = module.open(str(pdf_path))
    except Exception as exc:
        raise RenderError(f"PyMuPDF could not open {pdf_path.name}: {exc}") from exc

    rendered: list[_RasterizedPage] = []
    try:
        if document.page_count < 1:
            raise RenderError(f"LibreOffice produced an empty PDF for {pdf_path.name}")
        scale = dpi / 72.0
        matrix = module.Matrix(scale, scale)
        for page_index in range(document.page_count):
            page = document.load_page(page_index)
            pixmap = page.get_pixmap(matrix=matrix, alpha=False)
            png_path = output_dir / f"{filename_prefix}-{page_index + 1:04d}.png"
            pixmap.save(str(png_path))
            rendered.append(
                _RasterizedPage(
                    path=png_path,
                    width=int(pixmap.width),
                    height=int(pixmap.height),
                    sheet=sheet,
                    sheet_page=page_index + 1 if sheet is not None else None,
                )
            )
    except RenderError:
        raise
    except Exception as exc:
        raise RenderError(f"PyMuPDF failed to rasterize {pdf_path.name}: {exc}") from exc
    finally:
        document.close()
    return rendered


def _render_per_sheet(
    source_copy: Path,
    work_dir: Path,
    *,
    binary: str,
    dpi: int,
    timeout_seconds: float,
) -> list[_RasterizedPage]:
    sheets = _sheet_names(source_copy)
    if not sheets:
        raise RenderError("Workbook contains no worksheets")

    rendered: list[_RasterizedPage] = []
    for sheet_index, sheet_name in enumerate(sheets, start=1):
        sheet_dir = work_dir / f"sheet-{sheet_index:04d}"
        sheet_dir.mkdir(parents=True, exist_ok=True)
        private_book = sheet_dir / f"workbook{source_copy.suffix.lower()}"
        _make_single_sheet_copy(source_copy, private_book, sheet_name)
        pdf_path = _convert_with_libreoffice(
            private_book,
            sheet_dir / "pdf",
            target_format="pdf",
            binary=binary,
            timeout_seconds=timeout_seconds,
        )
        rendered.extend(
            _rasterize_pdf(
                pdf_path,
                work_dir / "png",
                dpi=dpi,
                filename_prefix=f"sheet-{sheet_index:04d}-page",
                sheet=sheet_name,
            )
        )
    return rendered


def _render_whole_workbook(
    source_copy: Path,
    work_dir: Path,
    *,
    binary: str,
    dpi: int,
    timeout_seconds: float,
) -> list[_RasterizedPage]:
    pdf_path = _convert_with_libreoffice(
        source_copy,
        work_dir / "pdf",
        target_format="pdf",
        binary=binary,
        timeout_seconds=timeout_seconds,
    )
    return _rasterize_pdf(
        pdf_path,
        work_dir / "png",
        dpi=dpi,
        filename_prefix="workbook-page",
        sheet=None,
    )


def _publish_pages(
    temporary_pages: Sequence[_RasterizedPage], output_dir: Path
) -> tuple[RenderPage, ...]:
    output_dir.mkdir(parents=True, exist_ok=True)
    destinations = [output_dir / page.path.name for page in temporary_pages]
    collisions = [path for path in destinations if path.exists()]
    if collisions:
        raise RenderError(f"Refusing to overwrite render artifact: {collisions[0]}")

    published: list[RenderPage] = []
    for index, (temporary, destination) in enumerate(
        zip(temporary_pages, destinations, strict=True), start=1
    ):
        shutil.copy2(temporary.path, destination)
        published.append(
            RenderPage(
                index=index,
                path=destination,
                sha256=sha256_file(destination),
                width=temporary.width,
                height=temporary.height,
                sheet=temporary.sheet,
                sheet_page=temporary.sheet_page,
            )
        )
    return tuple(published)


def _write_manifest(path: Path, data: dict[str, Any]) -> None:
    if path.exists():
        raise RenderError(f"Refusing to overwrite render manifest: {path}")
    temporary = path.with_name(f".{path.name}.tmp")
    temporary.write_text(
        json.dumps(data, ensure_ascii=False, indent=2, sort_keys=False) + "\n",
        encoding="utf-8",
    )
    temporary.replace(path)


def render_workbook(
    source: str | Path,
    output_dir: str | Path,
    *,
    dpi: int = 144,
    per_sheet: bool = True,
    libreoffice_binary: str | Path | None = None,
    timeout_seconds: float = 120.0,
) -> RenderResult:
    """Render a spreadsheet to PNG pages through LibreOffice and PyMuPDF.

    OOXML workbooks are rendered one sheet at a time when possible by hiding all
    other sheets in disposable copies.  If that strategy fails, the function
    retries with a whole-workbook PDF.  ODS, XLS, and CSV inputs use the whole
    workbook path directly because modifying their sheet visibility would
    require changing the original format.
    """

    source_path = _validate_source(source)
    if dpi <= 0:
        raise RenderError("dpi must be a positive integer")
    if timeout_seconds <= 0:
        raise RenderError("timeout_seconds must be positive")
    binary = find_libreoffice(libreoffice_binary)
    if binary is None:
        raise RenderError("LibreOffice executable was not found")

    destination = Path(output_dir).expanduser().resolve()
    if destination == source_path or source_path in destination.parents:
        # A directory can never equal a regular source file, but retaining this
        # guard makes the no-overwrite invariant explicit for unusual paths.
        if destination == source_path:
            raise RenderError("Render output directory cannot be the source workbook")
    source_hash = sha256_file(source_path)
    fallback_reason: str | None = None

    with tempfile.TemporaryDirectory(prefix="spreadsheet-render-") as raw_work:
        work = Path(raw_work)
        source_copy = work / "source" / source_path.name
        source_copy.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(source_path, source_copy)

        mode = "whole_workbook"
        temporary_pages: list[_RasterizedPage]
        if per_sheet and source_path.suffix.lower() in {".xlsx", ".xlsm"}:
            try:
                temporary_pages = _render_per_sheet(
                    source_copy,
                    work / "per-sheet",
                    binary=binary,
                    dpi=dpi,
                    timeout_seconds=timeout_seconds,
                )
                mode = "per_sheet"
            except Exception as exc:
                fallback_reason = f"{type(exc).__name__}: {exc}"
                temporary_pages = _render_whole_workbook(
                    source_copy,
                    work / "whole-workbook",
                    binary=binary,
                    dpi=dpi,
                    timeout_seconds=timeout_seconds,
                )
        else:
            temporary_pages = _render_whole_workbook(
                source_copy,
                work / "whole-workbook",
                binary=binary,
                dpi=dpi,
                timeout_seconds=timeout_seconds,
            )

        if not temporary_pages:
            raise RenderError("Rendering produced no PNG pages")
        pages = _publish_pages(temporary_pages, destination)

    versions = {
        "libreoffice": libreoffice_version(binary),
        "pymupdf": pymupdf_version(),
    }
    manifest_path = destination / "render-manifest.json"
    result = RenderResult(
        source=source_path,
        output_dir=destination,
        manifest_path=manifest_path,
        backend="libreoffice-headless+pymupdf",
        version=versions,
        source_sha256=source_hash,
        mode=mode,
        dpi=int(dpi),
        pages=pages,
    )
    manifest = result.to_dict()
    if fallback_reason is not None:
        manifest["fallback"] = {"from": "per_sheet", "reason": fallback_reason}
    _write_manifest(manifest_path, manifest)
    return result


def read_png(path: str | Path) -> bytes:
    """Return the exact PNG bytes for a view-image tool, without re-encoding."""

    png_path = Path(path).expanduser().resolve()
    try:
        data = png_path.read_bytes()
    except OSError as exc:
        raise RenderError(f"Could not read PNG {png_path}: {exc}") from exc
    if not data.startswith(PNG_SIGNATURE):
        raise RenderError(f"Not a PNG file: {png_path}")
    return data


# Small, discoverable aliases for callers that prefer verb-style APIs.
render = render_workbook
load_png = read_png


__all__ = [
    "RenderPage",
    "RenderResult",
    "RECALCULATION_SHEET_INTEGRITY_POLICY",
    "SUPPORTED_SPREADSHEET_EXTENSIONS",
    "convert_spreadsheet_copy",
    "find_libreoffice",
    "isolated_user_profile",
    "libreoffice_command",
    "libreoffice_version",
    "load_png",
    "pymupdf_version",
    "read_png",
    "recalculate_workbook",
    "render",
    "render_workbook",
    "sheet_inventory_identity",
    "sha256_file",
]
