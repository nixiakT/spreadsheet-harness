"""Revision-bound finalization and read-only scoring copies.

The deliverable certificate is an integrity certificate, not a proof that the
workbook satisfies the user's semantic intent.  It binds trusted agent evidence,
session artifact transitions, post-processing, final-revision witnesses, and the
exact bytes presented to the benchmark scorer.  Records are bound with canonical
SHA-256 digests, not private-key signatures or hostile-storage authentication.
"""

from __future__ import annotations

import ctypes
import errno
import hashlib
import io
import json
import os
import re
import stat
import sys
import tempfile
from collections.abc import Callable, Mapping, Sequence
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Protocol

from openpyxl import load_workbook

from .completion_attempt import CompletionAttemptRecord, audit_completion_attempt
from .evidence_contract import (
    ArtifactRef,
    ArtifactTransition,
    ContractMode,
    ContractSpec,
    EffectKind,
    EventKind,
    EvidenceContractMonitor,
    EvidenceEvent,
    EvidenceScope,
)
from .target_grounding import (
    AdvisoryLifecycleEvent,
    CommittedAdvisoryTargetAssessment,
    CommittedTargetAuthorization,
    TargetGroundingError,
    TargetGroundingMode,
    validate_committed_advisory_assessment_chain,
    validate_committed_authorization_chain,
)

LEGACY_DELIVERABLE_CERTIFICATE_SCHEMA_VERSION = "spreadsheet-deliverable-certificate-v2"
DELIVERABLE_CERTIFICATE_SCHEMA_VERSION = "spreadsheet-deliverable-certificate-v3"
NO_FORMULA_ATTESTATION_SCHEMA_VERSION = "ooxml-no-formula-finalization-witness-v1"
TARGET_GROUNDING_CERTIFICATE_SCHEMA_VERSION = "target-grounding-commit-chain-v1"
ADVISORY_TARGET_GROUNDING_CERTIFICATE_SCHEMA_VERSION = "target-grounding-observer-ledger-v2"
COMPARISON_RESULT_SCHEMA_VERSION = "spreadsheet-comparison-result-v28"
SCORING_COPY_RELATIVE_PATH = "scoring-output.xlsx"
FINAL_RENDER_RELATIVE_DIR = "postprocess/final-render"
_AUTHORIZATION_CHAIN_GENESIS_SHA256 = "0" * 64
_ACCEPTED_CANDIDATE = "accepted_candidate"
_AUDITED_NONCOMPLETION = "audited_noncompletion"

_CALC_ERROR_VALUES = frozenset(
    {
        "#NULL!",
        "#DIV/0!",
        "#VALUE!",
        "#REF!",
        "#NAME?",
        "#NUM!",
        "#N/A",
        "#GETTING_DATA",
        "#SPILL!",
        "#CALC!",
        "#FIELD!",
        "#BLOCKED!",
        "#UNKNOWN!",
        "#CONNECT!",
    }
)
_VISUAL_EFFECTS = frozenset(
    {
        EffectKind.STYLE.value,
        EffectKind.STRUCTURE.value,
        EffectKind.VISUAL.value,
        EffectKind.UNKNOWN.value,
    }
)
_ABSOLUTE_PATH_TOKEN = re.compile(r"(?:^|[\s:=])(?:/[A-Za-z0-9_.-]|[A-Za-z]:[\\/])")


class DeliverableValidationError(ValueError):
    """Raised when a final artifact cannot be certified fail-closed."""


class _DeliverableSession(Protocol):
    @property
    def workbook_path(self) -> Path: ...

    @property
    def workspace(self) -> Path: ...

    def artifact_ref(self) -> ArtifactRef: ...

    @property
    def artifact_transitions(self) -> tuple[ArtifactTransition, ...]: ...

    @property
    def target_grounding_enabled(self) -> bool: ...

    @property
    def target_grounding_active(self) -> bool: ...

    @property
    def target_grounding_enforced(self) -> bool: ...

    @property
    def target_grounding_mode(self) -> TargetGroundingMode: ...

    @property
    def target_grounding_initial_artifact(self) -> ArtifactRef | None: ...

    @property
    def target_grounding_initial_transition_count(self) -> int | None: ...

    @property
    def committed_target_authorizations(
        self,
    ) -> tuple[CommittedTargetAuthorization, ...]: ...

    @property
    def committed_advisory_target_assessments(
        self,
    ) -> tuple[CommittedAdvisoryTargetAssessment, ...]: ...

    @property
    def advisory_target_lifecycle_events(
        self,
    ) -> tuple[AdvisoryLifecycleEvent, ...]: ...

    @property
    def advisory_target_lifecycle_genesis_sha256(self) -> str | None: ...

    @property
    def advisory_target_lifecycle_final_counters(self) -> dict[str, int] | None: ...

    def recalculate(self, *, timeout_seconds: float = 120.0) -> dict[str, Any]: ...

    def recalculate_for_finalization(self, *, timeout_seconds: float = 120.0) -> dict[str, Any]: ...


@dataclass(frozen=True)
class DeliverableBundle:
    """One finalized workbook and its byte-identical scoring copy."""

    candidate_artifact: ArtifactRef
    final_artifact: ArtifactRef
    scoring_copy: Path
    recalculation: dict[str, Any]
    certificate: dict[str, Any]


@dataclass
class _ScoringCopyPublication:
    sha256: str
    identity: tuple[int, int]
    source_path: Path
    destination_path: Path
    source_identity: tuple[int, int, int, int, int, int, int]
    source_parent_identity: tuple[int, int]
    workspace_identity: tuple[int, int]
    source_descriptor: int | None
    source_parent_descriptor: int | None
    scoring_descriptor: int | None
    workspace_descriptor: int | None

    def _require_open(self) -> tuple[int, int, int, int]:
        descriptors = (
            self.source_descriptor,
            self.source_parent_descriptor,
            self.scoring_descriptor,
            self.workspace_descriptor,
        )
        if any(descriptor is None for descriptor in descriptors):
            raise DeliverableValidationError("Scoring publication lease is closed")
        return tuple(descriptors)  # type: ignore[return-value]

    def verify(self) -> None:
        (
            source_descriptor,
            source_parent_descriptor,
            scoring_descriptor,
            workspace_descriptor,
        ) = self._require_open()
        try:
            _validate_real_directory_path(
                self.source_path.parent,
                expected_identity=self.source_parent_identity,
                label="scoring source parent",
            )
            _validate_real_directory_path(
                self.destination_path.parent,
                expected_identity=self.workspace_identity,
                label="scoring workspace",
            )
            source_descriptor_metadata = os.fstat(source_descriptor)
            source_name_metadata = os.stat(
                self.source_path.name,
                dir_fd=source_parent_descriptor,
                follow_symlinks=False,
            )
            source_path_metadata = self.source_path.lstat()
            if not all(
                _stable_file_identity(metadata) == self.source_identity
                for metadata in (
                    source_descriptor_metadata,
                    source_name_metadata,
                    source_path_metadata,
                )
            ):
                raise DeliverableValidationError(
                    "Scoring source changed identity during finalization"
                )
            if (
                not stat.S_ISREG(source_descriptor_metadata.st_mode)
                or source_descriptor_metadata.st_nlink != 1
                or _sha256_descriptor(
                    source_descriptor,
                    expected_size=source_descriptor_metadata.st_size,
                )
                != self.sha256
            ):
                raise DeliverableValidationError(
                    "Scoring source changed contents during finalization"
                )
            if any(
                _stable_file_identity(metadata) != self.source_identity
                for metadata in (
                    os.fstat(source_descriptor),
                    os.stat(
                        self.source_path.name,
                        dir_fd=source_parent_descriptor,
                        follow_symlinks=False,
                    ),
                    self.source_path.lstat(),
                )
            ):
                raise DeliverableValidationError(
                    "Scoring source changed identity during finalization"
                )

            workspace_descriptor_metadata = os.fstat(workspace_descriptor)
            workspace_path_metadata = self.destination_path.parent.lstat()
            if not stat.S_ISDIR(workspace_descriptor_metadata.st_mode) or any(
                (metadata.st_dev, metadata.st_ino) != self.workspace_identity
                for metadata in (
                    workspace_descriptor_metadata,
                    workspace_path_metadata,
                )
            ):
                raise DeliverableValidationError(
                    "Scoring workspace changed identity during finalization"
                )

            scoring_descriptor_metadata = os.fstat(scoring_descriptor)
            scoring_stable_identity = _stable_file_identity(scoring_descriptor_metadata)
            scoring_name_metadata = os.stat(
                self.destination_path.name,
                dir_fd=workspace_descriptor,
                follow_symlinks=False,
            )
            scoring_path_metadata = self.destination_path.lstat()
            if (
                not stat.S_ISREG(scoring_descriptor_metadata.st_mode)
                or (
                    scoring_descriptor_metadata.st_dev,
                    scoring_descriptor_metadata.st_ino,
                )
                != self.identity
                or scoring_descriptor_metadata.st_nlink != 1
                or stat.S_IMODE(scoring_descriptor_metadata.st_mode) != 0o400
                or scoring_descriptor_metadata.st_size != source_descriptor_metadata.st_size
                or any(
                    _stable_file_identity(metadata) != scoring_stable_identity
                    for metadata in (
                        scoring_descriptor_metadata,
                        scoring_name_metadata,
                        scoring_path_metadata,
                    )
                )
                or _sha256_descriptor(
                    scoring_descriptor,
                    expected_size=scoring_descriptor_metadata.st_size,
                )
                != self.sha256
            ):
                raise DeliverableValidationError(
                    "Scoring replica changed identity or contents during finalization"
                )
            if any(
                _stable_file_identity(metadata) != scoring_stable_identity
                for metadata in (
                    os.fstat(scoring_descriptor),
                    os.stat(
                        self.destination_path.name,
                        dir_fd=workspace_descriptor,
                        follow_symlinks=False,
                    ),
                    self.destination_path.lstat(),
                )
            ):
                raise DeliverableValidationError(
                    "Scoring replica changed identity or contents during finalization"
                )
            _validate_real_directory_path(
                self.destination_path.parent,
                expected_identity=self.workspace_identity,
                label="scoring workspace",
            )
            _validate_real_directory_path(
                self.source_path.parent,
                expected_identity=self.source_parent_identity,
                label="scoring source parent",
            )
            final_workspace_metadata = os.fstat(workspace_descriptor)
            if (
                not stat.S_ISDIR(final_workspace_metadata.st_mode)
                or (
                    final_workspace_metadata.st_dev,
                    final_workspace_metadata.st_ino,
                )
                != self.workspace_identity
            ):
                raise DeliverableValidationError(
                    "Scoring workspace changed identity during finalization"
                )
        except DeliverableValidationError:
            raise
        except OSError as exc:
            raise DeliverableValidationError(
                "Failed to verify the live scoring publication lease"
            ) from exc

    def close(self) -> None:
        failure: OSError | None = None
        for attribute in (
            "scoring_descriptor",
            "source_descriptor",
            "source_parent_descriptor",
            "workspace_descriptor",
        ):
            descriptor = getattr(self, attribute)
            setattr(self, attribute, None)
            if descriptor is None:
                continue
            try:
                os.close(descriptor)
            except OSError as exc:
                failure = failure or exc
        if failure is not None:
            raise DeliverableValidationError(
                "Failed to close the scoring publication lease"
            ) from failure

    def commit(self) -> None:
        self.verify()
        self.close()

    def rollback(self) -> None:
        """Abandon the lease without mutating a published namespace.

        Once the anonymous inode is linked, retaining that immutable name is the
        only race-safe failure behavior; a later finalization attempt adopts it.
        """

        self.close()


@dataclass(frozen=True)
class DeliverableAudit:
    """Structured result from independently checking a deliverable certificate."""

    valid: bool
    reasons: tuple[str, ...]
    final_artifact: ArtifactRef | None = None
    scoring_copy: Path | None = None


def _canonical_json_bytes(value: Any, *, label: str) -> bytes:
    try:
        rendered = json.dumps(
            value,
            ensure_ascii=True,
            sort_keys=True,
            separators=(",", ":"),
            allow_nan=False,
        )
    except (TypeError, ValueError) as exc:
        raise DeliverableValidationError(f"{label} is not canonical JSON data: {exc}") from exc
    return rendered.encode("ascii")


def _canonical_sha256(value: Any, *, label: str) -> str:
    return hashlib.sha256(_canonical_json_bytes(value, label=label)).hexdigest()


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _valid_sha256(value: Any) -> bool:
    return bool(
        isinstance(value, str)
        and len(value) == 64
        and all(character in "0123456789abcdef" for character in value)
    )


def _contains_absolute_path(value: str) -> bool:
    return Path(value).is_absolute() or bool(_ABSOLUTE_PATH_TOKEN.search(value))


def _artifact_ref(value: Any, *, label: str) -> ArtifactRef:
    if not isinstance(value, Mapping) or set(value) != {"revision", "sha256"}:
        raise DeliverableValidationError(f"{label} must be an exact artifact reference")
    revision = value.get("revision")
    sha256 = value.get("sha256")
    if isinstance(revision, bool) or not isinstance(revision, int):
        raise DeliverableValidationError(f"{label}.revision must be an integer")
    if not isinstance(sha256, str):
        raise DeliverableValidationError(f"{label}.sha256 must be a string")
    try:
        return ArtifactRef(revision, sha256)
    except ValueError as exc:
        raise DeliverableValidationError(f"Invalid {label}: {exc}") from exc


def _relative_certificate_path(value: Any, *, label: str) -> Path:
    if not isinstance(value, str) or not value:
        raise DeliverableValidationError(f"{label} must be a non-empty relative path")
    path = Path(value)
    if path.is_absolute() or ".." in path.parts or path.as_posix() != value:
        raise DeliverableValidationError(f"{label} must be a normalized relative POSIX path")
    return path


def _path_inside(root: Path, relative: Path, *, label: str) -> Path:
    resolved_root = root.resolve(strict=True)
    candidate = resolved_root / relative
    resolved_candidate = candidate.resolve(strict=False)
    if resolved_candidate == resolved_root or resolved_root not in resolved_candidate.parents:
        raise DeliverableValidationError(f"{label} escapes the isolated run directory")
    current = resolved_root
    for component in relative.parts:
        current /= component
        try:
            metadata = current.lstat()
        except FileNotFoundError:
            continue
        except OSError as exc:
            raise DeliverableValidationError(f"{label} path could not be inspected") from exc
        if stat.S_ISLNK(metadata.st_mode):
            raise DeliverableValidationError(f"{label} path contains a symbolic link")
    return candidate


def _reject_absolute_certificate_paths(value: Any, *, label: str) -> None:
    if isinstance(value, Mapping):
        for raw_key, item in value.items():
            key = str(raw_key)
            if (
                isinstance(item, str)
                and (key == "path" or key.endswith("_path"))
                and _contains_absolute_path(item)
            ):
                raise DeliverableValidationError(f"{label} contains an absolute path")
            _reject_absolute_certificate_paths(item, label=label)
    elif isinstance(value, list | tuple):
        for item in value:
            _reject_absolute_certificate_paths(item, label=label)


def _scope_from_dict(value: Any, *, label: str) -> EvidenceScope:
    if not isinstance(value, Mapping):
        raise DeliverableValidationError(f"{label} must be an evidence scope")
    try:
        return EvidenceScope.from_dict(value)
    except (TypeError, ValueError) as exc:
        raise DeliverableValidationError(f"Invalid {label}: {exc}") from exc


def _contract_spec_from_identity(value: Any) -> ContractSpec:
    if not isinstance(value, Mapping) or set(value) != {
        "schema_version",
        "canonical_sha256",
        "source_sha256",
        "rules",
    }:
        raise DeliverableValidationError("Evidence certificate contract identity is invalid")
    raw_rules = value.get("rules")
    if not isinstance(raw_rules, list):
        raise DeliverableValidationError("Evidence certificate contract rules must be a list")
    source_rules: list[dict[str, Any]] = []
    for index, raw_rule in enumerate(raw_rules):
        if not isinstance(raw_rule, Mapping) or set(raw_rule) != {
            "id",
            "trigger",
            "requirements",
        }:
            raise DeliverableValidationError(f"Evidence contract rule {index} is invalid")
        requirements = raw_rule.get("requirements")
        if not isinstance(requirements, list) or not requirements:
            raise DeliverableValidationError(
                f"Evidence contract rule {index} has invalid requirements"
            )
        normalized_rule: dict[str, Any] = {
            "id": raw_rule.get("id"),
            "trigger": raw_rule.get("trigger"),
        }
        key = "require" if len(requirements) == 1 else "require_sequence"
        normalized_rule[key] = requirements[0] if len(requirements) == 1 else requirements
        source_rules.append(normalized_rule)
    try:
        parsed = ContractSpec.from_mapping(
            {"schema_version": value.get("schema_version"), "rules": source_rules}
        )
    except (TypeError, ValueError) as exc:
        raise DeliverableValidationError(f"Evidence contract identity is invalid: {exc}") from exc
    if parsed.canonical_sha256 != value.get("canonical_sha256"):
        raise DeliverableValidationError("Evidence contract canonical SHA-256 does not match")
    source_sha256 = value.get("source_sha256")
    if source_sha256 is not None and not _valid_sha256(source_sha256):
        raise DeliverableValidationError("Evidence contract source SHA-256 is invalid")
    return ContractSpec(
        schema_version=parsed.schema_version,
        rules=parsed.rules,
        canonical_sha256=parsed.canonical_sha256,
        source_sha256=source_sha256,
        source_path=None,
    )


def _evidence_event_from_record(value: Any) -> EvidenceEvent:
    expected = {
        "event_id",
        "kind",
        "revision_before",
        "revision_after",
        "effects",
        "scope",
        "formula_scope",
        "predicates",
        "render_id",
        "render_manifest_sha256",
        "related_render_id",
        "related_render_manifest_sha256",
        "page_id",
        "page_sha256",
        "metadata",
        "event_chain_sha256",
    }
    if not isinstance(value, Mapping) or set(value) != expected:
        raise DeliverableValidationError("Evidence certificate event record is invalid")
    raw_effects = value.get("effects")
    raw_predicates = value.get("predicates")
    if not isinstance(raw_effects, list) or not all(isinstance(item, str) for item in raw_effects):
        raise DeliverableValidationError("Evidence certificate effects are invalid")
    if not isinstance(raw_predicates, list) or not all(
        isinstance(item, str) for item in raw_predicates
    ):
        raise DeliverableValidationError("Evidence certificate predicates are invalid")
    try:
        kind = EventKind(str(value.get("kind")))
        effects = frozenset(EffectKind(item) for item in raw_effects)
    except ValueError as exc:
        raise DeliverableValidationError(
            f"Evidence certificate event enum is invalid: {exc}"
        ) from exc
    metadata = value.get("metadata")
    if not isinstance(metadata, Mapping):
        raise DeliverableValidationError("Evidence certificate event metadata is invalid")
    try:
        return EvidenceEvent(
            kind=kind,
            revision_before=str(value.get("revision_before")),
            revision_after=(
                str(value["revision_after"]) if value.get("revision_after") is not None else None
            ),
            effects=effects,
            scope=_scope_from_dict(value.get("scope"), label="event scope"),
            formula_scope=_scope_from_dict(value.get("formula_scope"), label="event formula scope"),
            predicates=frozenset(raw_predicates),
            render_id=value.get("render_id"),
            render_manifest_sha256=value.get("render_manifest_sha256"),
            related_render_id=value.get("related_render_id"),
            related_render_manifest_sha256=value.get("related_render_manifest_sha256"),
            page_id=value.get("page_id"),
            page_sha256=value.get("page_sha256"),
            metadata=dict(metadata),
        )
    except (TypeError, ValueError) as exc:
        raise DeliverableValidationError(f"Evidence certificate event is invalid: {exc}") from exc


def validate_evidence_certificate(value: Any) -> dict[str, Any]:
    """Replay a candidate certificate through the contract monitor exactly."""

    if not isinstance(value, Mapping):
        raise DeliverableValidationError("Candidate evidence certificate is missing")
    certificate = json.loads(json.dumps(value))
    expected = {
        "schema_version",
        "certificate_digest_algorithm",
        "contract",
        "initial_revision_sha256",
        "accepted_revision_sha256",
        "revision_index",
        "event_count",
        "event_chain_algorithm",
        "event_chain_genesis_sha256",
        "event_chain_sha256",
        "events",
        "obligations",
        "certificate_sha256",
    }
    if set(certificate) != expected:
        raise DeliverableValidationError("Candidate evidence certificate fields are invalid")
    if certificate.get("schema_version") != "spreadsheet-evidence-certificate-v1":
        raise DeliverableValidationError("Candidate evidence certificate schema is unsupported")
    stored_digest = certificate.get("certificate_sha256")
    payload = {key: certificate[key] for key in certificate if key != "certificate_sha256"}
    if (
        not _valid_sha256(stored_digest)
        or _canonical_sha256(payload, label="candidate evidence certificate") != stored_digest
    ):
        raise DeliverableValidationError("Candidate evidence certificate digest does not match")
    spec = _contract_spec_from_identity(certificate.get("contract"))
    initial = certificate.get("initial_revision_sha256")
    if not isinstance(initial, str):
        raise DeliverableValidationError("Candidate evidence initial revision is invalid")
    try:
        monitor = EvidenceContractMonitor(spec, initial, mode=ContractMode.ENFORCE)
    except ValueError as exc:
        raise DeliverableValidationError(
            f"Candidate evidence initial revision is invalid: {exc}"
        ) from exc
    events = certificate.get("events")
    if not isinstance(events, list):
        raise DeliverableValidationError("Candidate evidence events must be a list")
    for index, record in enumerate(events, start=1):
        if not isinstance(record, Mapping) or record.get("event_id") != index:
            raise DeliverableValidationError("Candidate evidence event ids are not contiguous")
        event = _evidence_event_from_record(record)
        try:
            monitor.observe(event)
        except (RuntimeError, ValueError) as exc:
            raise DeliverableValidationError(
                f"Candidate evidence event replay failed at event {index}: {exc}"
            ) from exc
    try:
        replayed = monitor.certificate()
    except RuntimeError as exc:
        raise DeliverableValidationError(
            f"Candidate evidence obligations are not satisfied: {exc}"
        ) from exc
    if replayed != certificate:
        raise DeliverableValidationError(
            "Candidate evidence event chain, scope, witnesses, or obligations do not match"
        )
    return certificate


def _completion_attempt_records(
    agent_evidence: Mapping[str, Any],
    *,
    workspace: Path,
    require_nonempty: bool,
) -> tuple[CompletionAttemptRecord, ...]:
    raw_attempts = agent_evidence.get("completion_attempts")
    if not isinstance(raw_attempts, list) or (require_nonempty and not raw_attempts):
        raise DeliverableValidationError("v28 requires completion-attempt records")
    try:
        attempts = tuple(CompletionAttemptRecord.from_dict(item) for item in raw_attempts)
    except (TypeError, ValueError) as exc:
        raise DeliverableValidationError("v28 completion-attempt records are invalid") from exc
    attempt_ids = [item.attempt_id for item in attempts]
    if attempt_ids != list(range(1, len(attempts) + 1)) or len(
        {item.call_id for item in attempts}
    ) != len(attempts):
        raise DeliverableValidationError("v28 completion-attempt sequence is not canonical")
    for raw_attempt, expected_attempt in zip(raw_attempts, attempts, strict=True):
        audit = audit_completion_attempt(workspace, raw_attempt)
        if not audit.valid or audit.record != expected_attempt:
            raise DeliverableValidationError("v28 completion-attempt snapshot failed fresh audit")
    return attempts


def _candidate_submission(
    agent_evidence: Mapping[str, Any],
    certificate: Mapping[str, Any],
    *,
    workspace: Path,
) -> dict[str, Any]:
    terminal_response = agent_evidence.get("terminal_response")
    response_id = agent_evidence.get("response_id")
    terminal_submissions = agent_evidence.get("terminal_submissions")
    if (
        agent_evidence.get("terminal_tool") != "submit_result"
        or agent_evidence.get("observed_terminal_tool") != "submit_result"
        or isinstance(terminal_submissions, bool)
        or not isinstance(terminal_submissions, int)
        or terminal_submissions < 1
        or not isinstance(response_id, str)
        or not response_id
    ):
        raise DeliverableValidationError("v28 requires an observed terminal submission")
    if (
        not isinstance(terminal_response, Mapping)
        or set(terminal_response)
        != {
            "status",
            "response_id",
            "acknowledgement",
            "completion_attempt_id",
        }
        or terminal_response.get("status") != "accepted"
        or terminal_response.get("response_id") != response_id
        or not isinstance(terminal_response.get("acknowledgement"), Mapping)
    ):
        raise DeliverableValidationError("v28 requires an accepted terminal response")
    raw_attempts = agent_evidence.get("completion_attempts")
    attempts = _completion_attempt_records(
        agent_evidence,
        workspace=workspace,
        require_nonempty=True,
    )
    assert isinstance(raw_attempts, list)

    accepted_id = terminal_response.get("completion_attempt_id")
    accepted_attempt = attempts[-1]
    if (
        type(accepted_id) is not int
        or accepted_id != accepted_attempt.attempt_id
        or accepted_attempt.response_id != response_id
        or accepted_attempt.artifact.sha256 != certificate.get("accepted_revision_sha256")
        or accepted_attempt.artifact.revision != certificate.get("revision_index")
    ):
        raise DeliverableValidationError(
            "v28 accepted terminal response is not bound to its completion attempt"
        )

    actual_call = agent_evidence
    actual_stage = agent_evidence.get("stage")
    stages = agent_evidence.get("stages")
    if stages is not None:
        if not isinstance(stages, list) or not stages:
            raise DeliverableValidationError("v28 terminal stage evidence is invalid")
        final_stage = stages[-1]
        if not isinstance(final_stage, Mapping) or not isinstance(
            final_stage.get("agent"), Mapping
        ):
            raise DeliverableValidationError("v28 terminal stage evidence is invalid")
        actual_call = final_stage["agent"]
        actual_stage = final_stage.get("name")
        if (
            actual_call.get("completion_attempts") != raw_attempts
            or actual_call.get("terminal_response") != terminal_response
        ):
            raise DeliverableValidationError(
                "v28 aggregated result does not preserve the accepted terminal call"
            )
    if (
        actual_call.get("terminal_tool") != "submit_result"
        or actual_call.get("observed_terminal_tool") != "submit_result"
        or actual_call.get("response_id") != response_id
        or actual_call.get("terminal_submissions") != len(attempts)
        or actual_call.get("turns") != accepted_attempt.turn
        or actual_stage != accepted_attempt.stage
    ):
        raise DeliverableValidationError(
            "v28 completion attempt is not the actual accepted terminal call"
        )
    final_text = agent_evidence.get("final_text")
    if not isinstance(final_text, str):
        raise DeliverableValidationError("Agent final_text is missing")
    return {
        "terminal_tool": agent_evidence.get("terminal_tool"),
        "observed_terminal_tool": agent_evidence.get("observed_terminal_tool"),
        "terminal_submissions": terminal_submissions,
        "response_id": response_id,
        "completion_attempt_id": accepted_attempt.attempt_id,
        "completion_attempt_record_sha256": accepted_attempt.record_sha256,
        "completion_attempt_count": len(attempts),
        "terminal_response_sha256": _canonical_sha256(terminal_response, label="terminal response"),
        "final_text_sha256": hashlib.sha256(final_text.encode("utf-8")).hexdigest(),
        "final_text_chars": len(final_text),
        "evidence_certificate_sha256": certificate.get("certificate_sha256"),
    }


def _candidate_certificate(
    agent_evidence: Any,
    *,
    workspace: Path,
) -> tuple[dict[str, Any], dict[str, Any]]:
    if not isinstance(agent_evidence, Mapping):
        raise DeliverableValidationError("v28 requires agent evidence")
    contract = agent_evidence.get("evidence_contract")
    if not isinstance(contract, Mapping):
        raise DeliverableValidationError("v28 agent evidence contract report is missing")
    decision = contract.get("decision")
    if (
        not isinstance(decision, Mapping)
        or decision.get("contract_satisfied") is not True
        or decision.get("allowed") is not True
    ):
        raise DeliverableValidationError("v28 candidate evidence contract was not satisfied")
    certificate = validate_evidence_certificate(decision.get("certificate"))
    status = contract.get("status")
    if (
        not isinstance(status, Mapping)
        or status.get("submission_ready") is not True
        or status.get("initial_revision_sha256") != certificate.get("initial_revision_sha256")
        or status.get("current_revision_sha256") != certificate.get("accepted_revision_sha256")
        or status.get("revision_index") != certificate.get("revision_index")
        or status.get("event_count") != certificate.get("event_count")
        or status.get("event_chain_sha256") != certificate.get("event_chain_sha256")
    ):
        raise DeliverableValidationError("Candidate evidence status does not bind the certificate")
    return certificate, _candidate_submission(
        agent_evidence,
        certificate,
        workspace=workspace,
    )


def _audited_noncompletion(
    agent_evidence: Mapping[str, Any],
    *,
    workspace: Path,
) -> dict[str, Any]:
    terminal_response = agent_evidence.get("terminal_response")
    if isinstance(terminal_response, Mapping) and terminal_response.get("status") == "accepted":
        raise DeliverableValidationError(
            "An accepted terminal response cannot be recorded as a noncompletion"
        )
    if terminal_response is not None and not isinstance(terminal_response, Mapping):
        raise DeliverableValidationError("Noncompletion terminal response is invalid")
    response_id = agent_evidence.get("response_id")
    if response_id is not None and (not isinstance(response_id, str) or not response_id):
        raise DeliverableValidationError("Noncompletion response_id is invalid")
    terminal_submissions = agent_evidence.get("terminal_submissions")
    if (
        isinstance(terminal_submissions, bool)
        or not isinstance(terminal_submissions, int)
        or terminal_submissions < 0
    ):
        raise DeliverableValidationError("Noncompletion terminal submission count is invalid")
    for field in ("terminal_tool", "observed_terminal_tool"):
        value = agent_evidence.get(field)
        if value is not None and (not isinstance(value, str) or not value):
            raise DeliverableValidationError(f"Noncompletion {field} is invalid")
    final_text = agent_evidence.get("final_text")
    if not isinstance(final_text, str):
        raise DeliverableValidationError("Agent final_text is missing")
    contract = agent_evidence.get("evidence_contract")
    if contract is not None and not isinstance(contract, Mapping):
        raise DeliverableValidationError("Noncompletion evidence contract is invalid")
    attempts = _completion_attempt_records(
        agent_evidence,
        workspace=workspace,
        require_nonempty=False,
    )
    raw_attempts = agent_evidence.get("completion_attempts")
    stages = agent_evidence.get("stages")
    if stages is not None:
        if not isinstance(stages, list) or not stages:
            raise DeliverableValidationError("v28 terminal stage evidence is invalid")
        final_stage = stages[-1]
        if not isinstance(final_stage, Mapping) or not isinstance(
            final_stage.get("agent"), Mapping
        ):
            raise DeliverableValidationError("v28 terminal stage evidence is invalid")
        final_agent = final_stage["agent"]
        final_attempts = final_agent.get("completion_attempts")
        attempts_preserved = final_attempts == raw_attempts or (
            raw_attempts == [] and final_attempts is None
        )
        if not attempts_preserved or final_agent.get("terminal_response") != terminal_response:
            raise DeliverableValidationError(
                "v28 aggregated result does not preserve noncompletion evidence"
            )
    return {
        "outcome": _AUDITED_NONCOMPLETION,
        "accepted_submission": False,
        "accepted_evidence_certificate": False,
        "terminal_tool": agent_evidence.get("terminal_tool"),
        "observed_terminal_tool": agent_evidence.get("observed_terminal_tool"),
        "terminal_submissions": terminal_submissions,
        "response_id": response_id,
        "terminal_response_sha256": (
            _canonical_sha256(terminal_response, label="terminal response")
            if terminal_response is not None
            else None
        ),
        "completion_attempt_count": len(attempts),
        "completion_attempt_record_sha256s": [item.record_sha256 for item in attempts],
        "evidence_contract_sha256": (
            _canonical_sha256(contract, label="noncompletion evidence contract")
            if contract is not None
            else None
        ),
        "final_text_sha256": hashlib.sha256(final_text.encode("utf-8")).hexdigest(),
        "final_text_chars": len(final_text),
        "agent_evidence_sha256": _canonical_sha256(
            agent_evidence,
            label="noncompletion agent evidence",
        ),
        "assurance": {
            "binds_termination_evidence": True,
            "claims_accepted_submission": False,
            "claims_task_correctness": False,
        },
    }


def _candidate_outcome(
    agent_evidence: Any,
    *,
    workspace: Path,
    candidate: ArtifactRef,
) -> tuple[str, dict[str, Any] | None, dict[str, Any]]:
    if not isinstance(agent_evidence, Mapping):
        raise DeliverableValidationError("v28 requires agent evidence")
    terminal_response = agent_evidence.get("terminal_response")
    if isinstance(terminal_response, Mapping) and terminal_response.get("status") == "accepted":
        certificate, submission = _candidate_certificate(
            agent_evidence,
            workspace=workspace,
        )
        if (
            certificate.get("accepted_revision_sha256") != candidate.sha256
            or certificate.get("revision_index") != candidate.revision
        ):
            raise DeliverableValidationError(
                "Candidate evidence certificate is stale for the managed artifact revision"
            )
        return _ACCEPTED_CANDIDATE, certificate, submission
    return (
        _AUDITED_NONCOMPLETION,
        None,
        _audited_noncompletion(agent_evidence, workspace=workspace),
    )


def _sanitized_workbook_effects(
    value: Any,
    *,
    allow_incomplete: bool,
) -> dict[str, Any]:
    expected = {
        "schema_version",
        "semantic_changed",
        "complete",
        "effects",
        "scope",
        "formula_scope",
        "changed_cell_count",
        "scanned_cell_count",
        "reasons",
    }
    if not isinstance(value, Mapping) or set(value) != expected:
        raise DeliverableValidationError("Recalculation workbook effects schema is invalid")
    if value.get("schema_version") != "workbook-effect-diff-v1":
        raise DeliverableValidationError("Recalculation workbook effects version is invalid")
    if type(value.get("semantic_changed")) is not bool or type(value.get("complete")) is not bool:
        raise DeliverableValidationError("Recalculation workbook effect flags are invalid")
    effects = value.get("effects")
    if (
        not isinstance(effects, list)
        or not all(isinstance(item, str) for item in effects)
        or len(effects) != len(set(effects))
        or any(item not in {effect.value for effect in EffectKind} for item in effects)
    ):
        raise DeliverableValidationError("Recalculation workbook effect kinds are invalid")
    scope = _scope_from_dict(value.get("scope"), label="recalculation effect scope")
    formula_scope = _scope_from_dict(
        value.get("formula_scope"), label="recalculation formula scope"
    )
    for field in ("changed_cell_count", "scanned_cell_count"):
        item = value.get(field)
        if isinstance(item, bool) or not isinstance(item, int) or item < 0:
            raise DeliverableValidationError(f"Recalculation workbook effects {field} is invalid")
    reasons = value.get("reasons")
    if not isinstance(reasons, list) or not all(isinstance(item, str) for item in reasons):
        raise DeliverableValidationError("Recalculation workbook effect reasons are invalid")
    if any(_contains_absolute_path(item) for item in reasons):
        raise DeliverableValidationError(
            "Recalculation workbook effect reasons contain an absolute path"
        )
    if value["complete"] is not True:
        if not allow_incomplete:
            raise DeliverableValidationError("Recalculation workbook effects are incomplete")
        if (
            value["semantic_changed"] is not True
            or effects != [EffectKind.UNKNOWN.value]
            or scope != EvidenceScope.workbook()
            or not formula_scope.empty
            or value["changed_cell_count"] != 0
            or not reasons
        ):
            raise DeliverableValidationError(
                "Incomplete recalculation effects must be a fail-closed unknown footprint"
            )
        return json.loads(
            _canonical_json_bytes(
                value,
                label="incomplete recalculation workbook effects",
            ).decode("ascii")
        )
    if value["semantic_changed"]:
        if not effects or scope.empty:
            raise DeliverableValidationError(
                "Semantic recalculation effects require typed, non-empty scope"
            )
    elif effects or not scope.empty or not formula_scope.empty or value["changed_cell_count"]:
        raise DeliverableValidationError(
            "No-op recalculation cannot claim semantic effects or changed scope"
        )
    if EffectKind.FORMULA.value in effects:
        if formula_scope.empty or not scope.covers(formula_scope):
            raise DeliverableValidationError(
                "Formula recalculation effects require covered formula scope"
            )
    elif not formula_scope.empty:
        raise DeliverableValidationError("Recalculation formula scope requires a formula effect")
    normalized = json.loads(
        _canonical_json_bytes(value, label="recalculation workbook effects").decode("ascii")
    )
    return normalized


def _sanitized_formula_scan(value: Any, *, artifact: ArtifactRef) -> dict[str, Any]:
    from .ooxml_formula_scan import (
        OOXML_FORMULA_POLICY_VERSION,
        OOXML_FORMULA_SCAN_SCHEMA_VERSION,
    )

    expected = {
        "schema_version",
        "formula_policy_version",
        "package_sha256",
        "workbook_format",
        "xml_part_count",
        "worksheet_count",
        "scanned_cell_count",
        "formula_marker_count",
        "formula_kinds",
    }
    if not isinstance(value, Mapping) or set(value) != expected:
        raise DeliverableValidationError("Formula scan attestation schema is invalid")
    if (
        value.get("schema_version") != OOXML_FORMULA_SCAN_SCHEMA_VERSION
        or value.get("formula_policy_version") != OOXML_FORMULA_POLICY_VERSION
        or value.get("package_sha256") != artifact.sha256
        or value.get("workbook_format") not in {"xlsx", "xlsm"}
    ):
        raise DeliverableValidationError("Formula scan attestation binding is invalid")
    for field in (
        "xml_part_count",
        "worksheet_count",
        "scanned_cell_count",
        "formula_marker_count",
    ):
        item = value.get(field)
        if isinstance(item, bool) or not isinstance(item, int) or item < 0:
            raise DeliverableValidationError("Formula scan attestation counts are invalid")
    kinds = value.get("formula_kinds")
    if (
        not isinstance(kinds, list)
        or not all(isinstance(item, str) and item for item in kinds)
        or kinds != sorted(set(kinds))
    ):
        raise DeliverableValidationError("Formula scan attestation kinds are invalid")
    if value.get("formula_marker_count") != 0 or kinds:
        raise DeliverableValidationError("Formula scan attestation is not formula-free")
    return json.loads(
        _canonical_json_bytes(value, label="formula scan attestation").decode("ascii")
    )


def _sanitize_recalculation(
    value: Any,
    *,
    before: ArtifactRef,
    after: ArtifactRef,
    candidate_outcome: str,
) -> dict[str, Any]:
    if not isinstance(value, Mapping):
        raise DeliverableValidationError("Session recalculation metadata is missing")
    required = {
        "backend",
        "version",
        "source_sha256",
        "output_sha256",
        "atomic_replace",
        "artifact_revision_before",
        "artifact_revision_after",
        "artifact_transition_id",
        "workbook_changed",
        "workbook_effects",
    }
    if not required.issubset(value):
        raise DeliverableValidationError("Session recalculation metadata is incomplete")
    backend = value.get("backend")
    version = value.get("version")
    profile = value.get("profile")
    workbook_format = value.get("format")
    from .ooxml_formula_scan import (
        OOXML_FORMULA_SCAN_SCHEMA_VERSION,
        OOXML_NO_FORMULA_BACKEND,
        OOXML_NO_FORMULA_PROFILE,
    )

    verified_no_write = backend == OOXML_NO_FORMULA_BACKEND
    if not isinstance(backend, str) or not backend.strip():
        raise DeliverableValidationError("Recalculation backend must be a non-empty string")
    if not isinstance(version, str) or not version.strip():
        raise DeliverableValidationError("Recalculation version must be a non-empty string")
    if not isinstance(profile, str) or not profile.strip():
        raise DeliverableValidationError("Recalculation profile must be a non-empty string")
    if any(_contains_absolute_path(item) for item in (backend, version, profile)):
        raise DeliverableValidationError("Recalculation metadata contains an absolute path")
    if workbook_format not in {"xlsx", "xlsm"}:
        raise DeliverableValidationError("Recalculation format is unsupported")
    if (
        value.get("source_sha256") != before.sha256
        or value.get("output_sha256") != after.sha256
        or value.get("artifact_revision_before") != before.revision
        or value.get("artifact_revision_after") != after.revision
        or value.get("workbook_changed") is not (before != after)
    ):
        raise DeliverableValidationError("Session recalculation metadata does not match lineage")
    transition_id = value.get("artifact_transition_id")
    if before == after:
        if transition_id is not None:
            raise DeliverableValidationError("Unchanged recalculation cannot publish a transition")
    elif isinstance(transition_id, bool) or not isinstance(transition_id, int):
        raise DeliverableValidationError("Changed recalculation requires a transition id")
    if candidate_outcome not in {_ACCEPTED_CANDIDATE, _AUDITED_NONCOMPLETION}:
        raise DeliverableValidationError("Candidate outcome is invalid for recalculation")
    workbook_effects = _sanitized_workbook_effects(
        value.get("workbook_effects"),
        allow_incomplete=candidate_outcome == _AUDITED_NONCOMPLETION,
    )
    formula_scan: dict[str, Any] | None = None
    if verified_no_write:
        if (
            version != OOXML_FORMULA_SCAN_SCHEMA_VERSION
            or profile != OOXML_NO_FORMULA_PROFILE
            or value.get("atomic_replace") is not False
            or value.get("publication") != "verified_no_write"
            or before != after
            or transition_id is not None
        ):
            raise DeliverableValidationError(
                "Verified no-write recalculation metadata is inconsistent"
            )
        formula_scan = _sanitized_formula_scan(value.get("formula_scan"), artifact=after)
        strict_noop = {
            "schema_version": "workbook-effect-diff-v1",
            "semantic_changed": False,
            "complete": True,
            "effects": [],
            "scope": EvidenceScope().to_dict(),
            "formula_scope": EvidenceScope().to_dict(),
            "changed_cell_count": 0,
            "scanned_cell_count": formula_scan["scanned_cell_count"],
            "reasons": [],
        }
        if workbook_effects != strict_noop:
            raise DeliverableValidationError(
                "Verified no-write recalculation effects are not a strict no-op"
            )
    elif value.get("atomic_replace") is not True:
        raise DeliverableValidationError("Legacy recalculation must attest atomic replacement")
    if workbook_effects["semantic_changed"] is True and candidate_outcome == _ACCEPTED_CANDIDATE:
        raise DeliverableValidationError("Postprocess recalculation changed workbook semantics")
    sanitized = {
        "backend": backend,
        "version": version,
        "profile": profile,
        "format": workbook_format,
        "source_sha256": value.get("source_sha256"),
        "output_sha256": value.get("output_sha256"),
        "atomic_replace": value.get("atomic_replace"),
        "artifact_revision_before": value.get("artifact_revision_before"),
        "artifact_revision_after": value.get("artifact_revision_after"),
        "artifact_transition_id": transition_id,
        "workbook_changed": value.get("workbook_changed"),
        "workbook_effects": workbook_effects,
    }
    if verified_no_write:
        sanitized["publication"] = "verified_no_write"
        sanitized["formula_scan"] = formula_scan
    _canonical_json_bytes(sanitized, label="sanitized recalculation metadata")
    return sanitized


def _no_formula_attestation(
    recalculation: Mapping[str, Any],
    *,
    artifact: ArtifactRef,
) -> dict[str, Any]:
    from .ooxml_formula_scan import (
        OOXML_FORMULA_SCAN_SCHEMA_VERSION,
        OOXML_NO_FORMULA_BACKEND,
        OOXML_NO_FORMULA_PROFILE,
    )

    formula_scan = _sanitized_formula_scan(
        recalculation.get("formula_scan"),
        artifact=artifact,
    )
    payload = {
        "schema_version": NO_FORMULA_ATTESTATION_SCHEMA_VERSION,
        "artifact": artifact.to_dict(),
        "backend": OOXML_NO_FORMULA_BACKEND,
        "version": OOXML_FORMULA_SCAN_SCHEMA_VERSION,
        "profile": OOXML_NO_FORMULA_PROFILE,
        "publication": "verified_no_write",
        "formula_scan": formula_scan,
        "claims": {
            "formula_free": True,
            "byte_identical": True,
            "managed_artifact_write": False,
            "artifact_transition": False,
            "offline_rescan_required": True,
        },
    }
    return {
        **payload,
        "witness_sha256": _canonical_sha256(
            payload,
            label="no-formula finalization witness",
        ),
    }


def _scan_final_revision_snapshot(
    payload: bytes,
    artifact: ArtifactRef,
    *,
    workbook_format: str,
) -> dict[str, Any]:
    formulas = cached = None
    try:
        formulas = load_workbook(
            io.BytesIO(payload),
            data_only=False,
            read_only=True,
            keep_vba=workbook_format == "xlsm",
            keep_links=True,
        )
        cached = load_workbook(
            io.BytesIO(payload),
            data_only=True,
            read_only=True,
            keep_vba=workbook_format == "xlsm",
            keep_links=True,
        )
        if not formulas.sheetnames or formulas.sheetnames != cached.sheetnames:
            raise DeliverableValidationError("Final workbook sheet manifest is invalid")
        formula_cells = 0
        nonempty_cells = 0
        blank_formula_caches = 0
        calc_errors: list[str] = []
        sheet_manifest: list[dict[str, Any]] = []
        for sheet_name in formulas.sheetnames:
            formula_sheet = formulas[sheet_name]
            cached_sheet = cached[sheet_name]
            sheet_manifest.append(
                {
                    "name": sheet_name,
                    "max_row": formula_sheet.max_row,
                    "max_column": formula_sheet.max_column,
                }
            )
            for row in formula_sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        nonempty_cells += 1
                    is_formula = cell.data_type == "f" or (
                        isinstance(cell.value, str) and cell.value.startswith("=")
                    )
                    if not is_formula:
                        continue
                    formula_cells += 1
                    cached_value = cached_sheet[cell.coordinate].value
                    if cached_value is None:
                        blank_formula_caches += 1
                    if isinstance(cached_value, str) and cached_value.upper() in _CALC_ERROR_VALUES:
                        calc_errors.append(f"{sheet_name}!{cell.coordinate}")
        if calc_errors:
            raise DeliverableValidationError(
                f"Final revision contains {len(calc_errors)} cached calculation errors"
            )
        witness = {
            "schema_version": "spreadsheet-final-revision-witness-v1",
            "artifact": artifact.to_dict(),
            "scope": EvidenceScope.workbook().to_dict(),
            "predicates": ["workbook_reopened", "no_calc_error"],
            "sheet_manifest_sha256": _canonical_sha256(
                sheet_manifest, label="final sheet manifest"
            ),
            "sheet_count": len(sheet_manifest),
            "scanned_nonempty_cells": nonempty_cells,
            "formula_cell_count": formula_cells,
            "blank_formula_cache_count": blank_formula_caches,
            "calculation_error_count": 0,
        }
        return {
            **witness,
            "witness_sha256": _canonical_sha256(witness, label="final revision witness"),
        }
    except DeliverableValidationError:
        raise
    except Exception as exc:
        raise DeliverableValidationError(
            f"Final revision could not be reopened and inspected: {type(exc).__name__}: {exc}"
        ) from exc
    finally:
        if formulas is not None:
            formulas.close()
        if cached is not None:
            cached.close()


def _required_candidate_views(certificate: Mapping[str, Any]) -> list[dict[str, Any]]:
    """Return only page views that actually discharged certificate obligations."""

    events = certificate.get("events")
    obligations = certificate.get("obligations")
    if not isinstance(events, list) or not isinstance(obligations, list):
        raise DeliverableValidationError("Candidate visual evidence records are invalid")
    events_by_id = {
        event.get("event_id"): event
        for event in events
        if isinstance(event, Mapping)
        and isinstance(event.get("event_id"), int)
        and not isinstance(event.get("event_id"), bool)
    }
    required_event_ids: list[int] = []
    for obligation in obligations:
        if not isinstance(obligation, Mapping):
            raise DeliverableValidationError("Candidate visual obligation is invalid")
        witnesses = obligation.get("witnesses")
        if not isinstance(witnesses, list):
            raise DeliverableValidationError("Candidate visual witnesses are invalid")
        for witness in witnesses:
            if (
                isinstance(witness, Mapping)
                and witness.get("kind") == EventKind.RENDERED_PAGE_VIEWED.value
                and isinstance(witness.get("event_id"), int)
                and not isinstance(witness.get("event_id"), bool)
            ):
                required_event_ids.append(int(witness["event_id"]))

    views: list[dict[str, Any]] = []
    seen: set[tuple[str, str, str]] = set()
    for event_id in required_event_ids:
        event = events_by_id.get(event_id)
        if (
            not isinstance(event, Mapping)
            or event.get("kind") != EventKind.RENDERED_PAGE_VIEWED.value
        ):
            raise DeliverableValidationError("Candidate required page view event is missing")
        render_id = event.get("related_render_id")
        manifest_sha256 = event.get("related_render_manifest_sha256")
        page_id = event.get("page_id")
        page_sha256 = event.get("page_sha256")
        metadata = event.get("metadata")
        if (
            not isinstance(render_id, str)
            or not render_id
            or not _valid_sha256(manifest_sha256)
            or not isinstance(page_id, str)
            or not page_id
            or not _valid_sha256(page_sha256)
            or not isinstance(metadata, Mapping)
        ):
            raise DeliverableValidationError("Candidate required page view identity is invalid")
        render_event = next(
            (
                item
                for item in events
                if isinstance(item, Mapping)
                and item.get("kind") == EventKind.WORKBOOK_RENDERED.value
                and item.get("render_id") == render_id
                and item.get("render_manifest_sha256") == manifest_sha256
                and int(item.get("event_id", 0)) < event_id
            ),
            None,
        )
        render_metadata = (
            render_event.get("metadata") if isinstance(render_event, Mapping) else None
        )
        dpi = render_metadata.get("dpi") if isinstance(render_metadata, Mapping) else None
        if isinstance(dpi, bool) or not isinstance(dpi, int) or dpi <= 0:
            raise DeliverableValidationError("Candidate render DPI is unavailable")
        sheet = metadata.get("sheet")
        sheet_page = metadata.get("sheet_page")
        if sheet is not None and (not isinstance(sheet, str) or not sheet):
            raise DeliverableValidationError("Candidate viewed sheet identity is invalid")
        if sheet_page is not None and (
            isinstance(sheet_page, bool) or not isinstance(sheet_page, int) or sheet_page < 1
        ):
            raise DeliverableValidationError("Candidate viewed sheet page is invalid")
        key = (render_id, page_id, str(page_sha256))
        if key in seen:
            continue
        seen.add(key)
        views.append(
            {
                "event_id": event_id,
                "render_id": render_id,
                "render_manifest_sha256": manifest_sha256,
                "page_id": page_id,
                "page_sha256": page_sha256,
                "sheet": sheet,
                "sheet_page": sheet_page,
                "dpi": dpi,
            }
        )
    return views


def _candidate_has_visual_evidence(certificate: Mapping[str, Any]) -> bool:
    events = certificate.get("events")
    visual_mutation = bool(
        isinstance(events, list)
        and any(
            isinstance(event, Mapping)
            and isinstance(event.get("effects"), list)
            and bool(set(event["effects"]) & _VISUAL_EFFECTS)
            for event in events
        )
    )
    required_views = _required_candidate_views(certificate)
    if visual_mutation and not required_views:
        raise DeliverableValidationError(
            "Candidate visual mutation lacks a required viewed-page witness"
        )
    return visual_mutation


def _render_final_revision(
    workbook_path: Path,
    output_dir: Path,
    *,
    workspace: Path,
    artifact: ArtifactRef,
    timeout_seconds: float,
    dpi: int,
) -> dict[str, Any]:
    from .render import PNG_SIGNATURE, render_workbook

    if output_dir.exists() and any(output_dir.iterdir()):
        raise DeliverableValidationError("Final render directory is not empty")
    result = render_workbook(
        workbook_path,
        output_dir,
        per_sheet=True,
        dpi=dpi,
        timeout_seconds=timeout_seconds,
    )
    if result.source_sha256 != artifact.sha256 or not result.pages:
        raise DeliverableValidationError("Final render is not bound to the final artifact")
    pages: list[dict[str, Any]] = []
    for page in result.pages:
        try:
            relative = page.path.resolve(strict=True).relative_to(workspace.resolve(strict=True))
            data = page.path.read_bytes()
        except (OSError, ValueError) as exc:
            raise DeliverableValidationError("Final render page escapes the run directory") from exc
        if not data.startswith(PNG_SIGNATURE) or hashlib.sha256(data).hexdigest() != page.sha256:
            raise DeliverableValidationError("Final render page hash or format is invalid")
        pages.append(
            {
                "index": page.index,
                "relative_path": relative.as_posix(),
                "sha256": page.sha256,
                "width": page.width,
                "height": page.height,
                "sheet": page.sheet,
                "sheet_page": page.sheet_page,
            }
        )
    portable_manifest = {
        "schema_version": "spreadsheet-portable-render-manifest-v1",
        "artifact": artifact.to_dict(),
        "backend": result.backend,
        "version": dict(result.version),
        "mode": result.mode,
        "dpi": result.dpi,
        "page_count": len(pages),
        "pages": pages,
    }
    portable_manifest_path = output_dir / "portable-render-manifest.json"
    portable_manifest_path.write_bytes(
        json.dumps(
            portable_manifest,
            ensure_ascii=True,
            sort_keys=True,
            separators=(",", ":"),
            allow_nan=False,
        ).encode("ascii")
        + b"\n"
    )
    manifest_relative = portable_manifest_path.resolve(strict=True).relative_to(
        workspace.resolve(strict=True)
    )
    record = {
        **portable_manifest,
        "schema_version": "spreadsheet-final-render-witness-v1",
        "manifest_relative_path": manifest_relative.as_posix(),
        "manifest_sha256": _sha256(portable_manifest_path),
    }
    return {
        **record,
        "witness_sha256": _canonical_sha256(record, label="final render witness"),
    }


def _final_page_for_candidate_view(
    render: Mapping[str, Any],
    view: Mapping[str, Any],
) -> Mapping[str, Any]:
    pages = render.get("pages")
    if not isinstance(pages, list):
        raise DeliverableValidationError("Final render page manifest is invalid")
    sheet = view.get("sheet")
    sheet_page = view.get("sheet_page")
    if isinstance(sheet, str) and isinstance(sheet_page, int):
        matches = [
            page
            for page in pages
            if isinstance(page, Mapping)
            and page.get("sheet") == sheet
            and page.get("sheet_page") == sheet_page
        ]
    else:
        raw_page_id = str(view.get("page_id"))
        candidate_name = raw_page_id.split(":", 1)[1] if ":" in raw_page_id else ""
        matches = [
            page
            for page in pages
            if isinstance(page, Mapping)
            and Path(str(page.get("relative_path", ""))).name == candidate_name
        ]
    if len(matches) != 1:
        raise DeliverableValidationError(
            "A required candidate page has no unique final-render counterpart"
        )
    return matches[0]


def _visual_equivalence_witness(
    certificate: Mapping[str, Any],
    *,
    session: _DeliverableSession,
    final: ArtifactRef,
    timeout_seconds: float,
) -> dict[str, Any]:
    required_views = _required_candidate_views(certificate)
    if not required_views:
        raise DeliverableValidationError("Visual equivalence requires candidate page views")
    groups: list[dict[str, Any]] = []
    render_keys = sorted(
        {
            (
                str(view["render_id"]),
                str(view["render_manifest_sha256"]),
                int(view["dpi"]),
            )
            for view in required_views
        }
    )
    mapped_page_count = 0
    for group_index, (render_id, manifest_sha256, dpi) in enumerate(render_keys, start=1):
        final_render = _render_final_revision(
            session.workbook_path,
            session.workspace / FINAL_RENDER_RELATIVE_DIR / f"render-{group_index:03d}",
            workspace=session.workspace,
            artifact=final,
            timeout_seconds=timeout_seconds,
            dpi=dpi,
        )
        page_equivalences: list[dict[str, Any]] = []
        group_views = [
            view
            for view in required_views
            if view["render_id"] == render_id
            and view["render_manifest_sha256"] == manifest_sha256
            and view["dpi"] == dpi
        ]
        for view in group_views:
            final_page = _final_page_for_candidate_view(final_render, view)
            final_page_sha256 = final_page.get("sha256")
            pixel_identical = final_page_sha256 == view.get("page_sha256")
            if not pixel_identical:
                raise DeliverableValidationError(
                    "A required viewed page changed after postprocess recalculation"
                )
            page_equivalences.append(
                {
                    "candidate_view_event_id": view["event_id"],
                    "page_id": view["page_id"],
                    "sheet": view["sheet"],
                    "sheet_page": view["sheet_page"],
                    "candidate_page_sha256": view["page_sha256"],
                    "final_page_relative_path": final_page["relative_path"],
                    "final_page_sha256": final_page_sha256,
                    "pixel_identical": True,
                }
            )
        mapped_page_count += len(page_equivalences)
        groups.append(
            {
                "candidate_render_id": render_id,
                "candidate_render_manifest_sha256": manifest_sha256,
                "dpi": dpi,
                "final_render": final_render,
                "pages": page_equivalences,
            }
        )
    if mapped_page_count != len(required_views):
        raise DeliverableValidationError("Not every required viewed page was mapped")
    payload = {
        "schema_version": "spreadsheet-render-equivalence-witness-v1",
        "final_artifact": final.to_dict(),
        "policy": "carry-view-only-through-pixel-identical-final-render-v1",
        "candidate_required_page_count": len(required_views),
        "mapped_page_count": mapped_page_count,
        "all_required_pages_pixel_identical": True,
        "render_groups": groups,
    }
    return {
        **payload,
        "witness_sha256": _canonical_sha256(payload, label="visual equivalence witness"),
    }


def _link_anonymous_file_noreplace_at(
    source_descriptor: int,
    destination_parent_descriptor: int,
    destination_name: str,
) -> None:
    """Publish one held anonymous inode without consulting a source pathname."""

    libc = ctypes.CDLL(None, use_errno=True)
    linkat = getattr(libc, "linkat", None)
    if linkat is None:
        raise DeliverableValidationError("Anonymous scoring-file publication is unavailable")
    linkat.argtypes = [
        ctypes.c_int,
        ctypes.c_char_p,
        ctypes.c_int,
        ctypes.c_char_p,
        ctypes.c_int,
    ]
    linkat.restype = ctypes.c_int
    destination_bytes = os.fsencode(destination_name)
    if (
        linkat(
            source_descriptor,
            b"",
            destination_parent_descriptor,
            destination_bytes,
            0x1000,  # AT_EMPTY_PATH
        )
        == 0
    ):
        return
    error_number = ctypes.get_errno()
    if error_number in {errno.ENOENT, errno.EPERM}:
        # Unprivileged Linux may require the kernel-owned procfs descriptor
        # symlink instead of AT_EMPTY_PATH. It still names only the held inode.
        proc_descriptor_path = os.fsencode(f"/proc/self/fd/{source_descriptor}")
        if (
            linkat(
                -100,  # AT_FDCWD
                proc_descriptor_path,
                destination_parent_descriptor,
                destination_bytes,
                0x400,  # AT_SYMLINK_FOLLOW
            )
            == 0
        ):
            return
        error_number = ctypes.get_errno()
    if error_number == errno.EEXIST:
        raise DeliverableValidationError("Scoring copy already exists")
    raise OSError(error_number, os.strerror(error_number), destination_name)


def _stable_file_identity(
    metadata: os.stat_result,
) -> tuple[int, int, int, int, int, int, int]:
    return (
        metadata.st_dev,
        metadata.st_ino,
        metadata.st_mode,
        metadata.st_nlink,
        metadata.st_size,
        metadata.st_mtime_ns,
        metadata.st_ctime_ns,
    )


def _validate_real_directory_path(
    path: Path,
    *,
    expected_identity: tuple[int, int],
    label: str,
) -> None:
    absolute = path.absolute()
    for candidate in (absolute, *absolute.parents):
        metadata = candidate.lstat()
        if not stat.S_ISDIR(metadata.st_mode) or stat.S_ISLNK(metadata.st_mode):
            raise DeliverableValidationError(f"{label} path contains a symbolic component")
    metadata = absolute.lstat()
    if (metadata.st_dev, metadata.st_ino) != expected_identity:
        raise DeliverableValidationError(f"{label} changed identity")


def _close_all_descriptors(*descriptors: int | None) -> OSError | None:
    failure: OSError | None = None
    for descriptor in descriptors:
        if descriptor is None:
            continue
        try:
            os.close(descriptor)
        except OSError as exc:
            failure = failure or exc
    return failure


def _sha256_descriptor(descriptor: int, *, expected_size: int) -> str:
    """Hash a descriptor without changing its shared file offset."""
    digest = hashlib.sha256()
    offset = 0
    while offset < expected_size:
        chunk = os.pread(descriptor, min(1024 * 1024, expected_size - offset), offset)
        if not chunk:
            break
        digest.update(chunk)
        offset += len(chunk)
    if offset != expected_size or os.pread(descriptor, 1, offset):
        raise DeliverableValidationError("Scoring descriptor size changed while hashing")
    return digest.hexdigest()


def _open_bound_scoring_source(
    source: Path,
    *,
    directory_flags: int,
) -> tuple[
    int,
    int,
    tuple[int, int, int, int, int, int, int],
    tuple[int, int],
    str,
]:
    parent_descriptor: int | None = None
    source_descriptor: int | None = None
    try:
        parent_descriptor = os.open(source.parent, directory_flags)
        parent_metadata = os.fstat(parent_descriptor)
        path_parent_metadata = source.parent.lstat()
        if not stat.S_ISDIR(parent_metadata.st_mode) or (
            parent_metadata.st_dev,
            parent_metadata.st_ino,
        ) != (path_parent_metadata.st_dev, path_parent_metadata.st_ino):
            raise DeliverableValidationError("Scoring source parent changed identity")
        parent_identity = (parent_metadata.st_dev, parent_metadata.st_ino)
        _validate_real_directory_path(
            source.parent,
            expected_identity=parent_identity,
            label="scoring source parent",
        )
        name_metadata = os.stat(
            source.name,
            dir_fd=parent_descriptor,
            follow_symlinks=False,
        )
        path_metadata = source.lstat()
        source_flags = os.O_RDONLY | getattr(os, "O_CLOEXEC", 0) | getattr(os, "O_NOFOLLOW", 0)
        source_descriptor = os.open(
            source.name,
            source_flags,
            dir_fd=parent_descriptor,
        )
        descriptor_metadata = os.fstat(source_descriptor)
        identity = _stable_file_identity(descriptor_metadata)
        if (
            not stat.S_ISREG(descriptor_metadata.st_mode)
            or descriptor_metadata.st_nlink != 1
            or _stable_file_identity(name_metadata) != identity
            or _stable_file_identity(path_metadata) != identity
        ):
            raise DeliverableValidationError("Scoring source must be a stable one-link file")
        digest = _sha256_descriptor(
            source_descriptor,
            expected_size=descriptor_metadata.st_size,
        )
        if any(
            _stable_file_identity(metadata) != identity
            for metadata in (
                os.fstat(source_descriptor),
                os.stat(
                    source.name,
                    dir_fd=parent_descriptor,
                    follow_symlinks=False,
                ),
                source.lstat(),
            )
        ):
            raise DeliverableValidationError("Scoring source changed while it was hashed")
        return parent_descriptor, source_descriptor, identity, parent_identity, digest
    except BaseException as exc:
        close_error = _close_all_descriptors(source_descriptor, parent_descriptor)
        if close_error is not None:
            raise DeliverableValidationError(
                f"Failed to close scoring-source descriptors: {close_error}"
            ) from exc
        raise


def _atomic_scoring_copy(
    source: Path,
    destination: Path,
) -> _ScoringCopyPublication:
    directory_flags = (
        os.O_RDONLY
        | getattr(os, "O_DIRECTORY", 0)
        | getattr(os, "O_CLOEXEC", 0)
        | getattr(os, "O_NOFOLLOW", 0)
    )
    source_descriptor: int | None = None
    source_parent_descriptor: int | None = None
    workspace_descriptor: int | None = None
    scoring_descriptor: int | None = None
    scoring_identity: tuple[int, int] | None = None
    source_identity: tuple[int, int, int, int, int, int, int] | None = None
    source_parent_identity: tuple[int, int] | None = None
    source_sha256: str | None = None
    try:
        if destination.name in {"", ".", ".."}:
            raise DeliverableValidationError("Scoring copy name is invalid")
        workspace = destination.parent
        workspace_descriptor = os.open(workspace, directory_flags)
        workspace_metadata = os.fstat(workspace_descriptor)
        workspace_path_metadata = workspace.lstat()
        if not stat.S_ISDIR(workspace_metadata.st_mode) or (
            workspace_metadata.st_dev,
            workspace_metadata.st_ino,
        ) != (
            workspace_path_metadata.st_dev,
            workspace_path_metadata.st_ino,
        ):
            raise OSError("scoring workspace is not a directory")
        workspace_identity = (
            workspace_metadata.st_dev,
            workspace_metadata.st_ino,
        )
        _validate_real_directory_path(
            workspace,
            expected_identity=workspace_identity,
            label="scoring workspace",
        )
        try:
            os.stat(
                destination.name,
                dir_fd=workspace_descriptor,
                follow_symlinks=False,
            )
        except FileNotFoundError:
            pass
        else:
            raise DeliverableValidationError("Scoring copy already exists")
        (
            source_parent_descriptor,
            source_descriptor,
            source_identity,
            source_parent_identity,
            source_sha256,
        ) = _open_bound_scoring_source(source, directory_flags=directory_flags)
        assert source_descriptor is not None
        assert source_parent_descriptor is not None
        assert source_identity is not None
        assert source_parent_identity is not None
        assert source_sha256 is not None
        anonymous_flag = getattr(os, "O_TMPFILE", 0)
        if not anonymous_flag:
            raise DeliverableValidationError("Anonymous scoring-file creation is unavailable")
        scoring_flags = os.O_RDWR | anonymous_flag | getattr(os, "O_CLOEXEC", 0)
        scoring_descriptor = os.open(
            ".",
            scoring_flags,
            0o600,
            dir_fd=workspace_descriptor,
        )
        created_scoring_metadata = os.fstat(scoring_descriptor)
        scoring_identity = (
            created_scoring_metadata.st_dev,
            created_scoring_metadata.st_ino,
        )
        if (
            not stat.S_ISREG(created_scoring_metadata.st_mode)
            or created_scoring_metadata.st_nlink != 0
            or created_scoring_metadata.st_dev != workspace_metadata.st_dev
            or created_scoring_metadata.st_uid != os.geteuid()
        ):
            raise DeliverableValidationError(
                "Scoring copy was not created as a private anonymous regular file"
            )
        digest = hashlib.sha256()
        copied = 0
        while True:
            chunk = os.read(source_descriptor, 1024 * 1024)
            if not chunk:
                break
            digest.update(chunk)
            copied += len(chunk)
            offset = 0
            while offset < len(chunk):
                written = os.write(scoring_descriptor, chunk[offset:])
                if written <= 0:
                    raise OSError("Scoring copy write made no progress")
                offset += written
        final_source_metadata = os.fstat(source_descriptor)
        if (
            _stable_file_identity(final_source_metadata) != source_identity
            or _stable_file_identity(
                os.stat(
                    source.name,
                    dir_fd=source_parent_descriptor,
                    follow_symlinks=False,
                )
            )
            != source_identity
            or _stable_file_identity(source.lstat()) != source_identity
            or copied != final_source_metadata.st_size
            or digest.hexdigest() != source_sha256
        ):
            raise DeliverableValidationError("Scoring source changed while it was copied")
        os.fchmod(scoring_descriptor, 0o400)
        os.fsync(scoring_descriptor)
        prepared_scoring_metadata = os.fstat(scoring_descriptor)
        if (
            not stat.S_ISREG(prepared_scoring_metadata.st_mode)
            or (prepared_scoring_metadata.st_dev, prepared_scoring_metadata.st_ino)
            != scoring_identity
            or prepared_scoring_metadata.st_nlink != 0
            or prepared_scoring_metadata.st_size != copied
            or stat.S_IMODE(prepared_scoring_metadata.st_mode) != 0o400
        ):
            raise DeliverableValidationError("Prepared scoring copy is not a read-only file")
        if _sha256_descriptor(scoring_descriptor, expected_size=copied) != source_sha256:
            raise DeliverableValidationError(
                "Prepared scoring copy is not byte-identical to its source"
            )
        os.fsync(workspace_descriptor)
        prepared_workspace_metadata = os.fstat(workspace_descriptor)
        if (
            not stat.S_ISDIR(prepared_workspace_metadata.st_mode)
            or (
                prepared_workspace_metadata.st_dev,
                prepared_workspace_metadata.st_ino,
            )
            != workspace_identity
            or _stable_file_identity(os.fstat(scoring_descriptor))
            != _stable_file_identity(prepared_scoring_metadata)
            or _stable_file_identity(
                os.stat(
                    source.name,
                    dir_fd=source_parent_descriptor,
                    follow_symlinks=False,
                )
            )
            != source_identity
            or _stable_file_identity(source.lstat()) != source_identity
        ):
            raise DeliverableValidationError("Prepared scoring publication changed identity")
        _validate_real_directory_path(
            workspace,
            expected_identity=workspace_identity,
            label="scoring workspace",
        )
        try:
            os.stat(
                destination.name,
                dir_fd=workspace_descriptor,
                follow_symlinks=False,
            )
        except FileNotFoundError:
            pass
        else:
            raise DeliverableValidationError("Scoring copy already exists")
        publication = _ScoringCopyPublication(
            sha256=source_sha256,
            identity=scoring_identity,
            source_path=source,
            destination_path=destination,
            source_identity=source_identity,
            source_parent_identity=source_parent_identity,
            workspace_identity=workspace_identity,
            source_descriptor=source_descriptor,
            source_parent_descriptor=source_parent_descriptor,
            scoring_descriptor=scoring_descriptor,
            workspace_descriptor=workspace_descriptor,
        )
        _link_anonymous_file_noreplace_at(
            scoring_descriptor,
            workspace_descriptor,
            destination.name,
        )
        publication.verify()
        # A failed durability sync has an ambiguous commit outcome. Keep the
        # exact immutable publication in place so a retry can adopt it.
        os.fsync(workspace_descriptor)
        source_descriptor = None
        source_parent_descriptor = None
        scoring_descriptor = None
        workspace_descriptor = None
        return publication
    except BaseException as exc:
        if isinstance(exc, DeliverableValidationError):
            raise
        if not isinstance(exc, Exception):
            raise
        raise DeliverableValidationError("Failed to publish the scoring copy") from exc
    finally:
        active_error = sys.exc_info()[1]
        close_error = _close_all_descriptors(
            scoring_descriptor,
            source_descriptor,
            source_parent_descriptor,
            workspace_descriptor,
        )
        if close_error is not None:
            message = f"Failed to close scoring-publication descriptors: {close_error}"
            if active_error is not None:
                raise DeliverableValidationError(message) from active_error
            raise DeliverableValidationError(message) from close_error


def _existing_scoring_copy(
    source: Path,
    destination: Path,
) -> _ScoringCopyPublication:
    directory_flags = (
        os.O_RDONLY
        | getattr(os, "O_DIRECTORY", 0)
        | getattr(os, "O_CLOEXEC", 0)
        | getattr(os, "O_NOFOLLOW", 0)
    )
    source_descriptor: int | None = None
    source_parent_descriptor: int | None = None
    scoring_descriptor: int | None = None
    workspace_descriptor: int | None = None
    try:
        (
            source_parent_descriptor,
            source_descriptor,
            source_identity,
            source_parent_identity,
            source_sha256,
        ) = _open_bound_scoring_source(source, directory_flags=directory_flags)
        workspace_descriptor = os.open(destination.parent, directory_flags)
        workspace_metadata = os.fstat(workspace_descriptor)
        workspace_path_metadata = destination.parent.lstat()
        if not stat.S_ISDIR(workspace_metadata.st_mode) or (
            workspace_metadata.st_dev,
            workspace_metadata.st_ino,
        ) != (workspace_path_metadata.st_dev, workspace_path_metadata.st_ino):
            raise DeliverableValidationError("Scoring workspace changed identity")
        workspace_identity = (workspace_metadata.st_dev, workspace_metadata.st_ino)
        _validate_real_directory_path(
            destination.parent,
            expected_identity=workspace_identity,
            label="scoring workspace",
        )
        scoring_flags = os.O_RDONLY | getattr(os, "O_CLOEXEC", 0) | getattr(os, "O_NOFOLLOW", 0)
        scoring_descriptor = os.open(
            destination.name,
            scoring_flags,
            dir_fd=workspace_descriptor,
        )
        scoring_metadata = os.fstat(scoring_descriptor)
        scoring_name_metadata = os.stat(
            destination.name,
            dir_fd=workspace_descriptor,
            follow_symlinks=False,
        )
        scoring_path_metadata = destination.lstat()
        scoring_identity = (scoring_metadata.st_dev, scoring_metadata.st_ino)
        if (
            not stat.S_ISREG(scoring_metadata.st_mode)
            or scoring_metadata.st_nlink != 1
            or stat.S_IMODE(scoring_metadata.st_mode) != 0o400
            or (scoring_name_metadata.st_dev, scoring_name_metadata.st_ino) != scoring_identity
            or (scoring_path_metadata.st_dev, scoring_path_metadata.st_ino) != scoring_identity
            or scoring_metadata.st_size != source_identity[4]
            or _sha256_descriptor(
                scoring_descriptor,
                expected_size=scoring_metadata.st_size,
            )
            != source_sha256
        ):
            raise DeliverableValidationError(
                "Published scoring copy is not an exact read-only regular-file replica"
            )
        publication = _ScoringCopyPublication(
            sha256=source_sha256,
            identity=scoring_identity,
            source_path=source,
            destination_path=destination,
            source_identity=source_identity,
            source_parent_identity=source_parent_identity,
            workspace_identity=workspace_identity,
            source_descriptor=source_descriptor,
            source_parent_descriptor=source_parent_descriptor,
            scoring_descriptor=scoring_descriptor,
            workspace_descriptor=workspace_descriptor,
        )
        publication.verify()
        source_descriptor = None
        source_parent_descriptor = None
        scoring_descriptor = None
        workspace_descriptor = None
        return publication
    except BaseException as exc:
        if isinstance(exc, DeliverableValidationError):
            raise
        if not isinstance(exc, Exception):
            raise
        raise DeliverableValidationError("Failed to bind the existing scoring copy") from exc
    finally:
        active_error = sys.exc_info()[1]
        close_error = _close_all_descriptors(
            scoring_descriptor,
            source_descriptor,
            source_parent_descriptor,
            workspace_descriptor,
        )
        if close_error is not None:
            message = f"Failed to close existing-scoring descriptors: {close_error}"
            if active_error is not None:
                raise DeliverableValidationError(message) from active_error
            raise DeliverableValidationError(message) from close_error


def _record_secondary_exception(error: BaseException, note: str) -> None:
    """Attach diagnostics without allowing note handling to replace the error."""

    try:
        add_note = getattr(error, "add_note", None)
        if callable(add_note):
            add_note(note)
            return
        notes = list(getattr(error, "__notes__", ()))
        notes.append(note)
        error.__notes__ = notes
    except BaseException:
        pass


@dataclass
class _BoundScoringInputDirectory:
    """A temporary directory whose cleanup never trusts its pathname alone."""

    path: Path
    identity: tuple[int, int]
    parent_identity: tuple[int, int]
    descriptor: int | None
    parent_descriptor: int | None

    @classmethod
    def create(cls) -> _BoundScoringInputDirectory:
        path = Path(os.path.abspath(tempfile.mkdtemp(prefix="sheetledger-scoring-input-")))
        directory_flags = (
            os.O_RDONLY
            | getattr(os, "O_DIRECTORY", 0)
            | getattr(os, "O_CLOEXEC", 0)
            | getattr(os, "O_NOFOLLOW", 0)
        )
        parent_descriptor: int | None = None
        descriptor: int | None = None
        try:
            parent_descriptor = os.open(path.parent, directory_flags)
            parent_metadata = os.fstat(parent_descriptor)
            parent_path_metadata = path.parent.lstat()
            parent_identity = (parent_metadata.st_dev, parent_metadata.st_ino)
            if (
                not stat.S_ISDIR(parent_metadata.st_mode)
                or (parent_path_metadata.st_dev, parent_path_metadata.st_ino) != parent_identity
            ):
                raise DeliverableValidationError("Temporary scoring parent changed identity")
            _validate_real_directory_path(
                path.parent,
                expected_identity=parent_identity,
                label="temporary scoring parent",
            )
            name_metadata = os.stat(
                path.name,
                dir_fd=parent_descriptor,
                follow_symlinks=False,
            )
            descriptor = os.open(path.name, directory_flags, dir_fd=parent_descriptor)
            directory_metadata = os.fstat(descriptor)
            path_metadata = path.lstat()
            identity = (directory_metadata.st_dev, directory_metadata.st_ino)
            if (
                not stat.S_ISDIR(directory_metadata.st_mode)
                or stat.S_ISLNK(directory_metadata.st_mode)
                or stat.S_IMODE(directory_metadata.st_mode) != 0o700
                or directory_metadata.st_uid != os.geteuid()
                or any(
                    (metadata.st_dev, metadata.st_ino) != identity
                    for metadata in (name_metadata, path_metadata)
                )
            ):
                raise DeliverableValidationError(
                    "Temporary scoring directory could not be bound safely"
                )
            return cls(
                path=path,
                identity=identity,
                parent_identity=parent_identity,
                descriptor=descriptor,
                parent_descriptor=parent_descriptor,
            )
        except BaseException as exc:
            close_error = _close_all_descriptors(descriptor, parent_descriptor)
            if close_error is not None:
                _record_secondary_exception(
                    exc,
                    "Temporary scoring directory descriptor cleanup also failed: "
                    f"{type(close_error).__name__}: {close_error}",
                )
            _record_secondary_exception(
                exc,
                "Temporary scoring directory creation failed before a cleanup capability "
                "was established; any unbound residue was left untouched",
            )
            raise

    def make_non_writable(self) -> None:
        descriptor = self.descriptor
        parent_descriptor = self.parent_descriptor
        if descriptor is None or parent_descriptor is None:
            raise DeliverableValidationError("Temporary scoring directory lease is closed")
        try:
            metadata = os.fstat(descriptor)
            if (
                not stat.S_ISDIR(metadata.st_mode)
                or (metadata.st_dev, metadata.st_ino) != self.identity
            ):
                raise DeliverableValidationError("Temporary scoring directory changed identity")
            os.fchmod(descriptor, 0o500)
            metadata = os.fstat(descriptor)
            named_metadata = os.stat(
                self.path.name,
                dir_fd=parent_descriptor,
                follow_symlinks=False,
            )
            if (
                stat.S_IMODE(metadata.st_mode) != 0o500
                or (metadata.st_dev, metadata.st_ino) != self.identity
                or (named_metadata.st_dev, named_metadata.st_ino) != self.identity
            ):
                raise DeliverableValidationError(
                    "Temporary scoring directory could not be made non-writable"
                )
        except DeliverableValidationError:
            raise
        except OSError as exc:
            raise DeliverableValidationError(
                "Temporary scoring directory could not be protected"
            ) from exc

    def cleanup(
        self,
        *,
        snapshot_name: str,
        snapshot_identity: tuple[int, int] | None,
    ) -> list[BaseException]:
        """Remove only entries proven to belong to this held directory capability."""

        descriptor = self.descriptor
        parent_descriptor = self.parent_descriptor
        self.descriptor = None
        self.parent_descriptor = None
        failures: list[BaseException] = []
        try:
            directory_valid = False
            if descriptor is not None:
                try:
                    metadata = os.fstat(descriptor)
                    directory_valid = bool(
                        stat.S_ISDIR(metadata.st_mode)
                        and (metadata.st_dev, metadata.st_ino) == self.identity
                    )
                    if not directory_valid:
                        failures.append(
                            DeliverableValidationError(
                                "Held temporary scoring directory changed identity"
                            )
                        )
                except BaseException as exc:
                    failures.append(exc)
                if directory_valid:
                    try:
                        os.fchmod(descriptor, 0o700)
                    except BaseException as exc:
                        failures.append(exc)
                    if snapshot_identity is not None:
                        try:
                            snapshot_metadata = os.stat(
                                snapshot_name,
                                dir_fd=descriptor,
                                follow_symlinks=False,
                            )
                        except FileNotFoundError:
                            pass
                        except BaseException as exc:
                            failures.append(exc)
                        else:
                            if (
                                not stat.S_ISREG(snapshot_metadata.st_mode)
                                or (snapshot_metadata.st_dev, snapshot_metadata.st_ino)
                                != snapshot_identity
                            ):
                                failures.append(
                                    DeliverableValidationError(
                                        "Temporary scoring snapshot name was replaced; "
                                        "the replacement was left untouched"
                                    )
                                )
                            else:
                                try:
                                    os.unlink(snapshot_name, dir_fd=descriptor)
                                except BaseException as exc:
                                    failures.append(exc)

            parent_valid = False
            if parent_descriptor is not None:
                try:
                    parent_metadata = os.fstat(parent_descriptor)
                    parent_valid = bool(
                        stat.S_ISDIR(parent_metadata.st_mode)
                        and (parent_metadata.st_dev, parent_metadata.st_ino) == self.parent_identity
                    )
                    if not parent_valid:
                        failures.append(
                            DeliverableValidationError(
                                "Held temporary scoring parent changed identity"
                            )
                        )
                except BaseException as exc:
                    failures.append(exc)
            if directory_valid and parent_valid:
                try:
                    named_metadata = os.stat(
                        self.path.name,
                        dir_fd=parent_descriptor,
                        follow_symlinks=False,
                    )
                except FileNotFoundError:
                    try:
                        if os.fstat(descriptor).st_nlink != 0:
                            failures.append(
                                DeliverableValidationError(
                                    "Original temporary scoring directory was renamed; "
                                    "the capability-bound empty residue was retained"
                                )
                            )
                    except BaseException as exc:
                        failures.append(exc)
                except BaseException as exc:
                    failures.append(exc)
                else:
                    if (
                        not stat.S_ISDIR(named_metadata.st_mode)
                        or (named_metadata.st_dev, named_metadata.st_ino) != self.identity
                    ):
                        failures.append(
                            DeliverableValidationError(
                                "Temporary scoring directory name was replaced; the "
                                "replacement was left untouched and the original residue "
                                "was retained"
                            )
                        )
                    else:
                        try:
                            rebound_metadata = os.stat(
                                self.path.name,
                                dir_fd=parent_descriptor,
                                follow_symlinks=False,
                            )
                            if (
                                not stat.S_ISDIR(rebound_metadata.st_mode)
                                or (rebound_metadata.st_dev, rebound_metadata.st_ino)
                                != self.identity
                            ):
                                raise DeliverableValidationError(
                                    "Temporary scoring directory changed before removal"
                                )
                            os.rmdir(self.path.name, dir_fd=parent_descriptor)
                        except BaseException as exc:
                            failures.append(exc)
        finally:
            close_error = _close_all_descriptors(descriptor, parent_descriptor)
            if close_error is not None:
                failures.append(close_error)
        return failures


def _run_with_bound_scoring_input(
    source: Path,
    expected_sha256: str,
    scorer: Callable[[Path], Any],
) -> Any:
    """Score a private snapshot while holding and rebinding the published source."""

    if not _valid_sha256(expected_sha256):
        raise DeliverableValidationError("Expected scoring SHA-256 is invalid")
    suffix = source.suffix.casefold()
    if suffix not in {".xlsx", ".xlsm"}:
        raise DeliverableValidationError("Scoring input format is unsupported")

    temporary = _BoundScoringInputDirectory.create()
    snapshot = temporary.path / f"input{suffix}"
    publication: _ScoringCopyPublication | None = None
    try:
        publication = _atomic_scoring_copy(source, snapshot)
        if publication.workspace_identity != temporary.identity:
            raise DeliverableValidationError(
                "Scoring snapshot was published outside its held directory"
            )
        if publication.sha256 != expected_sha256:
            raise DeliverableValidationError("Scoring copy changed before scoring")
        if stat.S_IMODE(publication.source_identity[2]) & 0o222:
            raise DeliverableValidationError(
                "Scoring copy must be a read-only regular non-symbolic file before scoring"
            )
        publication.verify()

        temporary.make_non_writable()
        publication.verify()

        scorer_error: BaseException | None = None
        scorer_traceback = None
        result: Any = None
        try:
            result = scorer(snapshot)
        except BaseException as exc:
            scorer_error = exc
            scorer_traceback = exc.__traceback__
        try:
            _verify_rebound_scoring_input(publication)
        except DeliverableValidationError as binding_error:
            if scorer_error is not None:
                _record_secondary_exception(
                    scorer_error,
                    "Scoring input verification also failed: "
                    f"{type(binding_error).__name__}: {binding_error}",
                )
                raise scorer_error.with_traceback(scorer_traceback) from binding_error
            raise DeliverableValidationError(
                "Scoring input changed during scoring"
            ) from binding_error
        if scorer_error is not None:
            raise scorer_error.with_traceback(scorer_traceback)
        return result
    finally:
        active_error = sys.exc_info()[1]
        try:
            cleanup_failures = temporary.cleanup(
                snapshot_name=snapshot.name,
                snapshot_identity=publication.identity if publication is not None else None,
            )
        except BaseException as exc:
            cleanup_failures = [exc]
        if publication is not None:
            try:
                publication.close()
            except BaseException as exc:
                cleanup_failures.append(exc)
        if cleanup_failures:
            if active_error is not None:
                for cleanup_error in cleanup_failures:
                    _record_secondary_exception(
                        active_error,
                        "Bound scoring input cleanup also failed: "
                        f"{type(cleanup_error).__name__}: {cleanup_error}",
                    )
            else:
                cleanup_error = DeliverableValidationError(
                    "Failed to clean up the bound scoring input"
                )
                for secondary_error in cleanup_failures[1:]:
                    _record_secondary_exception(
                        cleanup_error,
                        "Additional cleanup failure: "
                        f"{type(secondary_error).__name__}: {secondary_error}",
                    )
                raise cleanup_error from cleanup_failures[0]


def _verify_rebound_scoring_input(publication: _ScoringCopyPublication) -> None:
    """Verify scoring bytes without treating a harmless source rename as mutation."""

    (
        source_descriptor,
        source_parent_descriptor,
        scoring_descriptor,
        workspace_descriptor,
    ) = publication._require_open()
    try:
        _validate_real_directory_path(
            publication.source_path.parent,
            expected_identity=publication.source_parent_identity,
            label="scoring source parent",
        )
        _validate_real_directory_path(
            publication.destination_path.parent,
            expected_identity=publication.workspace_identity,
            label="scoring workspace",
        )

        source_metadata = os.fstat(source_descriptor)
        source_stable_identity = _stable_file_identity(source_metadata)
        source_identity = (source_metadata.st_dev, source_metadata.st_ino)
        expected_source_identity = publication.source_identity[:2]
        source_name_metadata = os.stat(
            publication.source_path.name,
            dir_fd=source_parent_descriptor,
            follow_symlinks=False,
        )
        source_path_metadata = publication.source_path.lstat()
        if (
            not stat.S_ISREG(source_metadata.st_mode)
            or source_metadata.st_nlink != 1
            or stat.S_IMODE(source_metadata.st_mode) & 0o222
            or source_identity != expected_source_identity
            or any(
                _stable_file_identity(metadata) != source_stable_identity
                for metadata in (source_name_metadata, source_path_metadata)
            )
            or source_metadata.st_size != publication.source_identity[4]
            or _sha256_descriptor(
                source_descriptor,
                expected_size=source_metadata.st_size,
            )
            != publication.sha256
        ):
            raise DeliverableValidationError("Scoring source changed during scoring")

        scoring_metadata = os.fstat(scoring_descriptor)
        scoring_stable_identity = _stable_file_identity(scoring_metadata)
        scoring_name_metadata = os.stat(
            publication.destination_path.name,
            dir_fd=workspace_descriptor,
            follow_symlinks=False,
        )
        scoring_path_metadata = publication.destination_path.lstat()
        if (
            not stat.S_ISREG(scoring_metadata.st_mode)
            or scoring_metadata.st_nlink != 1
            or stat.S_IMODE(scoring_metadata.st_mode) != 0o400
            or (scoring_metadata.st_dev, scoring_metadata.st_ino) != publication.identity
            or any(
                _stable_file_identity(metadata) != scoring_stable_identity
                for metadata in (scoring_name_metadata, scoring_path_metadata)
            )
            or scoring_metadata.st_size != publication.source_identity[4]
            or _sha256_descriptor(
                scoring_descriptor,
                expected_size=scoring_metadata.st_size,
            )
            != publication.sha256
        ):
            raise DeliverableValidationError("Private scoring snapshot changed during scoring")

        # Rebind the names after hashing so a concurrent replacement cannot
        # survive by racing only the first set of metadata checks.
        if any(
            _stable_file_identity(metadata) != expected_stable_identity
            for metadata, expected_stable_identity in (
                (os.fstat(source_descriptor), source_stable_identity),
                (
                    os.stat(
                        publication.source_path.name,
                        dir_fd=source_parent_descriptor,
                        follow_symlinks=False,
                    ),
                    source_stable_identity,
                ),
                (publication.source_path.lstat(), source_stable_identity),
                (os.fstat(scoring_descriptor), scoring_stable_identity),
                (
                    os.stat(
                        publication.destination_path.name,
                        dir_fd=workspace_descriptor,
                        follow_symlinks=False,
                    ),
                    scoring_stable_identity,
                ),
                (publication.destination_path.lstat(), scoring_stable_identity),
            )
        ):
            raise DeliverableValidationError("Scoring input changed while it was verified")
    except DeliverableValidationError:
        raise
    except OSError as exc:
        raise DeliverableValidationError("Failed to verify the bound scoring input") from exc


def _transition_lineage(
    transitions: Sequence[ArtifactTransition],
    *,
    candidate: ArtifactRef,
    candidate_transition_count: int,
    final: ArtifactRef,
) -> dict[str, Any]:
    if candidate_transition_count < 0 or candidate_transition_count > len(transitions):
        raise DeliverableValidationError("Candidate transition count is invalid")
    if transitions:
        initial = transitions[0].before
        current = initial
        for index, transition in enumerate(transitions, start=1):
            if transition.transition_id != index or transition.before != current:
                raise DeliverableValidationError("Session artifact transitions are not contiguous")
            current = transition.after
        if current != final:
            raise DeliverableValidationError(
                "Session artifact transitions do not reach final artifact"
            )
        candidate_state = (
            initial
            if candidate_transition_count == 0
            else transitions[candidate_transition_count - 1].after
        )
    else:
        initial = candidate
        candidate_state = candidate
        current = candidate
    if candidate_state != candidate or current != final:
        raise DeliverableValidationError("Candidate/final artifacts do not match session lineage")
    return {
        "initial_artifact": initial.to_dict(),
        "candidate_artifact": candidate.to_dict(),
        "final_artifact": final.to_dict(),
        "candidate_transition_count": candidate_transition_count,
        "transition_count": len(transitions),
        "transitions": [transition.to_dict() for transition in transitions],
    }


def _target_grounding_commit_chain(
    session: _DeliverableSession,
) -> dict[str, Any]:
    raw_mode = getattr(session, "target_grounding_mode", None)
    if isinstance(raw_mode, TargetGroundingMode):
        mode = raw_mode
    elif isinstance(raw_mode, str):
        try:
            mode = TargetGroundingMode(raw_mode)
        except ValueError as exc:
            raise DeliverableValidationError("Target-grounding session mode is invalid") from exc
    else:
        mode = (
            TargetGroundingMode.ENFORCE
            if bool(getattr(session, "target_grounding_enabled", False))
            else TargetGroundingMode.OFF
        )
    active = bool(
        getattr(
            session,
            "target_grounding_active",
            mode is not TargetGroundingMode.OFF,
        )
    )
    enforced = bool(
        getattr(
            session,
            "target_grounding_enforced",
            getattr(session, "target_grounding_enabled", False),
        )
    )
    if active is not (mode is not TargetGroundingMode.OFF) or enforced is not (
        mode is TargetGroundingMode.ENFORCE
    ):
        raise DeliverableValidationError(
            "Target-grounding session predicates disagree with its mode"
        )

    authorizations = session.committed_target_authorizations
    advisory_assessments = tuple(getattr(session, "committed_advisory_target_assessments", ()))
    if mode is TargetGroundingMode.ADVISORY:
        if authorizations:
            raise DeliverableValidationError(
                "Advisory target grounding cannot expose enforced authorizations"
            )
        initial_artifact = session.target_grounding_initial_artifact
        initial_transition_count = session.target_grounding_initial_transition_count
        if initial_artifact is None or initial_transition_count is None:
            raise DeliverableValidationError(
                "Advisory target grounding is missing its initial lineage binding"
            )
        documents = [record.to_dict() for record in advisory_assessments]
        lifecycle_events = tuple(session.advisory_target_lifecycle_events)
        lifecycle_documents = [event.to_dict() for event in lifecycle_events]
        lifecycle_genesis_sha256 = session.advisory_target_lifecycle_genesis_sha256
        lifecycle_final_counters = session.advisory_target_lifecycle_final_counters
        if (
            not _valid_sha256(lifecycle_genesis_sha256)
            or not isinstance(lifecycle_final_counters, dict)
            or lifecycle_final_counters.get("pending_preparation_count") != 0
        ):
            raise DeliverableValidationError("Advisory target-grounding lifecycle is not finalized")
        chain_head = (
            advisory_assessments[-1].canonical_sha256
            if advisory_assessments
            else _AUTHORIZATION_CHAIN_GENESIS_SHA256
        )
        return {
            "schema_version": ADVISORY_TARGET_GROUNDING_CERTIFICATE_SCHEMA_VERSION,
            "mode": TargetGroundingMode.ADVISORY.value,
            "active": True,
            "enforced": False,
            "initial_artifact": initial_artifact.to_dict(),
            "initial_transition_count": initial_transition_count,
            "assessment_count": len(advisory_assessments),
            "assessment_chain_head_sha256": chain_head,
            "assessments": documents,
            "lifecycle_replay_complete": True,
            "lifecycle_genesis_sha256": lifecycle_genesis_sha256,
            "lifecycle_event_count": len(lifecycle_events),
            "lifecycle_chain_head_sha256": (
                lifecycle_events[-1].canonical_sha256
                if lifecycle_events
                else lifecycle_genesis_sha256
            ),
            "lifecycle_final_counters": lifecycle_final_counters,
            "lifecycle_events": lifecycle_documents,
        }

    if advisory_assessments:
        raise DeliverableValidationError(
            "Off/enforced target grounding cannot expose advisory assessments"
        )
    enabled = mode is TargetGroundingMode.ENFORCE
    if enabled:
        initial_artifact = session.target_grounding_initial_artifact
        initial_transition_count = session.target_grounding_initial_transition_count
        if initial_artifact is None or initial_transition_count is None:
            raise DeliverableValidationError(
                "Enabled target grounding is missing its initial lineage binding"
            )
    else:
        initial_artifact = None
        initial_transition_count = None
        if authorizations:
            raise DeliverableValidationError(
                "Disabled target grounding cannot expose committed authorizations"
            )
    documents = [record.to_dict() for record in authorizations]
    chain_head = (
        authorizations[-1].canonical_sha256
        if authorizations
        else _AUTHORIZATION_CHAIN_GENESIS_SHA256
    )
    return {
        "schema_version": TARGET_GROUNDING_CERTIFICATE_SCHEMA_VERSION,
        "enabled": enabled,
        "initial_artifact": (initial_artifact.to_dict() if initial_artifact is not None else None),
        "initial_transition_count": initial_transition_count,
        "authorization_count": len(authorizations),
        "authorization_chain_head_sha256": chain_head,
        "authorizations": documents,
    }


def _audit_target_grounding_commit_chain(
    value: Any,
    *,
    transitions: tuple[ArtifactTransition, ...],
) -> None:
    if (
        isinstance(value, Mapping)
        and value.get("schema_version") == ADVISORY_TARGET_GROUNDING_CERTIFICATE_SCHEMA_VERSION
    ):
        expected_advisory_fields = {
            "schema_version",
            "mode",
            "active",
            "enforced",
            "initial_artifact",
            "initial_transition_count",
            "assessment_count",
            "assessment_chain_head_sha256",
            "assessments",
            "lifecycle_replay_complete",
            "lifecycle_genesis_sha256",
            "lifecycle_event_count",
            "lifecycle_chain_head_sha256",
            "lifecycle_final_counters",
            "lifecycle_events",
        }
        assessments = value.get("assessments")
        assessment_count = value.get("assessment_count")
        chain_head = value.get("assessment_chain_head_sha256")
        lifecycle_events = value.get("lifecycle_events")
        lifecycle_event_count = value.get("lifecycle_event_count")
        lifecycle_genesis_sha256 = value.get("lifecycle_genesis_sha256")
        lifecycle_chain_head_sha256 = value.get("lifecycle_chain_head_sha256")
        lifecycle_final_counters = value.get("lifecycle_final_counters")
        if (
            set(value) != expected_advisory_fields
            or value.get("mode") != TargetGroundingMode.ADVISORY.value
            or value.get("active") is not True
            or value.get("enforced") is not False
            or not isinstance(assessments, list)
            or type(assessment_count) is not int
            or assessment_count != len(assessments)
            or not _valid_sha256(chain_head)
            or value.get("lifecycle_replay_complete") is not True
            or not isinstance(lifecycle_events, list)
            or type(lifecycle_event_count) is not int
            or lifecycle_event_count != len(lifecycle_events)
            or not _valid_sha256(lifecycle_genesis_sha256)
            or not _valid_sha256(lifecycle_chain_head_sha256)
            or not isinstance(lifecycle_final_counters, Mapping)
        ):
            raise DeliverableValidationError("Advisory target-grounding certificate is invalid")
        initial_artifact = _artifact_ref(
            value.get("initial_artifact"),
            label="advisory target-grounding initial artifact",
        )
        initial_transition_count = value.get("initial_transition_count")
        if type(initial_transition_count) is not int:
            raise DeliverableValidationError(
                "Advisory target-grounding initial transition count is invalid"
            )
        try:
            records = validate_committed_advisory_assessment_chain(
                assessments,
                lifecycle_events=lifecycle_events,
                lifecycle_genesis_sha256=lifecycle_genesis_sha256,
                lifecycle_final_counters=lifecycle_final_counters,
                transitions=transitions,
                initial_artifact=initial_artifact,
                initial_transition_count=initial_transition_count,
            )
        except (TargetGroundingError, TypeError, ValueError) as exc:
            raise DeliverableValidationError(
                "Advisory target-grounding assessments do not replay"
            ) from exc
        expected_head = (
            records[-1].canonical_sha256 if records else _AUTHORIZATION_CHAIN_GENESIS_SHA256
        )
        if chain_head != expected_head:
            raise DeliverableValidationError("Advisory target-grounding chain head does not match")
        expected_lifecycle_head = (
            lifecycle_events[-1].get("event_sha256")
            if lifecycle_events
            else lifecycle_genesis_sha256
        )
        if lifecycle_chain_head_sha256 != expected_lifecycle_head:
            raise DeliverableValidationError(
                "Advisory target-grounding lifecycle head does not match"
            )
        return

    expected_fields = {
        "schema_version",
        "enabled",
        "initial_artifact",
        "initial_transition_count",
        "authorization_count",
        "authorization_chain_head_sha256",
        "authorizations",
    }
    if (
        not isinstance(value, Mapping)
        or set(value) != expected_fields
        or value.get("schema_version") != TARGET_GROUNDING_CERTIFICATE_SCHEMA_VERSION
        or type(value.get("enabled")) is not bool
    ):
        raise DeliverableValidationError("Target-grounding commit-chain certificate is invalid")
    authorizations = value.get("authorizations")
    authorization_count = value.get("authorization_count")
    chain_head = value.get("authorization_chain_head_sha256")
    if (
        not isinstance(authorizations, list)
        or type(authorization_count) is not int
        or authorization_count != len(authorizations)
        or not _valid_sha256(chain_head)
    ):
        raise DeliverableValidationError(
            "Target-grounding authorization count or chain head is invalid"
        )
    if value["enabled"] is False:
        if (
            value.get("initial_artifact") is not None
            or value.get("initial_transition_count") is not None
            or authorizations
            or chain_head != _AUTHORIZATION_CHAIN_GENESIS_SHA256
        ):
            raise DeliverableValidationError(
                "Disabled target grounding has non-empty authorization state"
            )
        return

    initial_artifact = _artifact_ref(
        value.get("initial_artifact"), label="target-grounding initial artifact"
    )
    initial_transition_count = value.get("initial_transition_count")
    if type(initial_transition_count) is not int:
        raise DeliverableValidationError("Target-grounding initial transition count is invalid")
    try:
        records = validate_committed_authorization_chain(
            authorizations,
            transitions=transitions,
            initial_artifact=initial_artifact,
            initial_transition_count=initial_transition_count,
        )
    except (TargetGroundingError, TypeError, ValueError) as exc:
        raise DeliverableValidationError(
            "Target-grounding committed authorizations do not replay"
        ) from exc
    expected_head = records[-1].canonical_sha256 if records else _AUTHORIZATION_CHAIN_GENESIS_SHA256
    if chain_head != expected_head:
        raise DeliverableValidationError("Target-grounding authorization chain head does not match")


def finalize_deliverable(
    session: _DeliverableSession,
    agent_evidence: Mapping[str, Any],
    *,
    recalculation_callback: Callable[[], Mapping[str, Any]] | None = None,
    recalculation_timeout_seconds: float | None = None,
    render_timeout_seconds: float = 120.0,
) -> DeliverableBundle:
    """Finalize one candidate and record immutable final/scoring-copy integrity."""

    if render_timeout_seconds <= 0:
        raise DeliverableValidationError("render_timeout_seconds must be positive")
    effective_recalculation_timeout = (
        render_timeout_seconds
        if recalculation_timeout_seconds is None
        else recalculation_timeout_seconds
    )
    if (
        isinstance(effective_recalculation_timeout, bool)
        or not isinstance(effective_recalculation_timeout, int | float)
        or effective_recalculation_timeout <= 0
    ):
        raise DeliverableValidationError("recalculation_timeout_seconds must be positive")
    if recalculation_callback is not None:
        raise DeliverableValidationError(
            "Finalization recalculation callbacks are disabled; use the session-owned "
            "transactional finalization backend"
        )
    candidate = session.artifact_ref()
    candidate_outcome, certificate, submission = _candidate_outcome(
        agent_evidence,
        workspace=session.workspace,
        candidate=candidate,
    )
    candidate_transition_count = len(session.artifact_transitions)
    raw_recalculation = session.recalculate_for_finalization(
        timeout_seconds=float(effective_recalculation_timeout)
    )
    final = session.artifact_ref()
    recalculation = _sanitize_recalculation(
        raw_recalculation,
        before=candidate,
        after=final,
        candidate_outcome=candidate_outcome,
    )
    transitions = session.artifact_transitions
    lineage = _transition_lineage(
        transitions,
        candidate=candidate,
        candidate_transition_count=candidate_transition_count,
        final=final,
    )
    if len(transitions) - candidate_transition_count not in {0, 1}:
        raise DeliverableValidationError("Postprocess published unexpected artifact transitions")
    if len(transitions) > candidate_transition_count:
        postprocess_transition = transitions[-1]
        if (
            postprocess_transition.operation != "recalculate"
            or postprocess_transition.kind != "derived_recalculation"
            or postprocess_transition.before != candidate
            or postprocess_transition.after != final
        ):
            raise DeliverableValidationError(
                "Postprocess transition is not a session recalculation"
            )

    no_formula_attestation = _no_formula_attestation(
        recalculation,
        artifact=final,
    )
    from .ooxml_formula_scan import OOXMLFormulaScanError, OOXMLFormulaScanLease

    try:
        with OOXMLFormulaScanLease.open(session.workbook_path) as final_scan_lease:
            if (
                final_scan_lease.scan.package_sha256 != final.sha256
                or final_scan_lease.scan.has_formulas
                or final_scan_lease.scan.to_dict() != recalculation.get("formula_scan")
            ):
                raise DeliverableValidationError(
                    "Final workbook does not reproduce its no-formula scan"
                )
            final_witness = _scan_final_revision_snapshot(
                final_scan_lease.snapshot_bytes,
                final,
                workbook_format=final_scan_lease.scan.workbook_format,
            )
            final_scan_lease.verify_binding(checkpoint="deliverable_finalization")
    except OOXMLFormulaScanError as exc:
        raise DeliverableValidationError(
            "Final workbook could not be held for descriptor-bound inspection"
        ) from exc
    visual_was_required = bool(
        certificate is not None and _candidate_has_visual_evidence(certificate)
    )
    visual_equivalence: dict[str, Any] | None = None
    if candidate != final and visual_was_required:
        assert certificate is not None
        visual_equivalence = _visual_equivalence_witness(
            certificate,
            session=session,
            final=final,
            timeout_seconds=render_timeout_seconds,
        )

    scoring_copy = _path_inside(
        session.workspace,
        Path(SCORING_COPY_RELATIVE_PATH),
        label="scoring copy",
    )
    evidence_policy = {
        "accepted_candidate_evidence": candidate_outcome == _ACCEPTED_CANDIDATE,
        "candidate_evidence_carried_forward": (
            candidate_outcome == _ACCEPTED_CANDIDATE and candidate == final
        ),
        "changed_recalculation_invalidates_candidate_evidence": True,
        "fresh_final_revision_readback": True,
        "candidate_visual_evidence_present": visual_was_required,
        "pixel_equivalence_required_for_visual_carry": (candidate != final and visual_was_required),
        "pixel_equivalence_observed": visual_equivalence is not None,
        "unviewed_final_render_never_counts_as_viewed": True,
    }
    target_grounding = _target_grounding_commit_chain(session)
    _audit_target_grounding_commit_chain(
        target_grounding,
        transitions=transitions,
    )
    payload = {
        "schema_version": DELIVERABLE_CERTIFICATE_SCHEMA_VERSION,
        "candidate": {
            "outcome": candidate_outcome,
            "artifact": candidate.to_dict(),
            "submission": submission,
            "evidence_certificate": certificate,
        },
        "lineage": lineage,
        "target_grounding": target_grounding,
        "postprocess": {
            "operation": "session.recalculate",
            "timeout_seconds": float(effective_recalculation_timeout),
            "recalculation": recalculation,
        },
        "recalculation_attestation": no_formula_attestation,
        "evidence_policy": evidence_policy,
        "final_artifact": final.to_dict(),
        "final_revision_witness": final_witness,
        "visual_equivalence_witness": visual_equivalence,
        "scoring_copy": {
            "relative_path": SCORING_COPY_RELATIVE_PATH,
            "source_artifact": final.to_dict(),
            "artifact_role": "same_revision_replica",
            "creates_artifact_transition": False,
            "sha256": final.sha256,
            "byte_identical": True,
            "read_only": True,
        },
    }
    _reject_absolute_certificate_paths(payload, label="deliverable certificate")
    deliverable_certificate = {
        **payload,
        "certificate_sha256": _canonical_sha256(payload, label="deliverable certificate"),
    }
    bundle = DeliverableBundle(
        candidate_artifact=candidate,
        final_artifact=final,
        scoring_copy=scoring_copy,
        recalculation=recalculation,
        certificate=deliverable_certificate,
    )
    recorder = getattr(session, "recorder", None)
    transition_count_before_copy = len(session.artifact_transitions)
    publication: _ScoringCopyPublication | None = None
    trajectory_durable = False
    transaction_handle: Any = None

    def publish_scoring_copy(*, recover_existing: bool = False) -> None:
        nonlocal publication
        publication = (
            _existing_scoring_copy(session.workbook_path, scoring_copy)
            if recover_existing
            else _atomic_scoring_copy(session.workbook_path, scoring_copy)
        )
        if len(session.artifact_transitions) != transition_count_before_copy:
            raise DeliverableValidationError(
                "Creating the scoring replica must not publish an artifact transition"
            )
        if publication.sha256 != final.sha256:
            raise DeliverableValidationError("Scoring copy SHA-256 differs from final artifact")
        publication.verify()

    finalization_event = {
        "schema_version": DELIVERABLE_CERTIFICATE_SCHEMA_VERSION,
        "candidate_outcome": candidate_outcome,
        # The append-only trajectory observes finalization but cannot atomically
        # authorize a separate filesystem publication. Authorization is derived
        # only by a fresh descriptor-bound audit of both records.
        "accepted_deliverable": False,
        "record_role": "observer_only_fresh_audit_required",
        "candidate_artifact": candidate.to_dict(),
        "final_artifact": final.to_dict(),
        "scoring_copy_relative_path": SCORING_COPY_RELATIVE_PATH,
        "certificate_sha256": deliverable_certificate["certificate_sha256"],
    }
    try:
        if recorder is not None and callable(getattr(recorder, "record", None)):
            transaction_factory = getattr(recorder, "transaction", None)
            if not callable(transaction_factory):
                raise DeliverableValidationError(
                    "Finalization recorder does not support transactional publication"
                )
            with transaction_factory() as transaction:
                transaction_handle = transaction
                event_count, matching_event_count = transaction.event_counts(
                    "observer.finalization_recorded",
                    finalization_event,
                )
                scoring_copy_exists = scoring_copy.exists() or scoring_copy.is_symlink()
                if event_count == 1 and matching_event_count == 1:
                    publish_scoring_copy(recover_existing=True)
                    assert publication is not None
                    publication.verify()
                    transaction.commit_read_only()
                elif event_count == 0 and matching_event_count == 0:
                    publish_scoring_copy(recover_existing=scoring_copy_exists)
                    assert publication is not None
                    transaction.record("observer.finalization_recorded", finalization_event)
                    publication.verify()
                    transaction.commit()
                else:
                    raise DeliverableValidationError(
                        "Trajectory contains a conflicting finalization record"
                    )
            trajectory_durable = True
            assert publication is not None
            publication.commit()
        else:
            publish_scoring_copy()
            assert publication is not None
            publication.commit()
        return bundle
    except BaseException as exc:
        trajectory_durable = trajectory_durable or bool(
            getattr(transaction_handle, "durable", False)
        )
        if publication is not None:
            try:
                if trajectory_durable:
                    publication.close()
                else:
                    publication.rollback()
            except DeliverableValidationError as cleanup_error:
                raise cleanup_error from exc
        raise


def score_read_only(
    bundle: DeliverableBundle,
    scorer: Callable[[Path], Any],
) -> Any:
    """Score immutable bytes while the published name remains descriptor-bound."""

    return _run_with_bound_scoring_input(
        bundle.scoring_copy,
        bundle.final_artifact.sha256,
        scorer,
    )


def _transition_from_dict(value: Any) -> ArtifactTransition:
    if not isinstance(value, Mapping) or set(value) != {
        "transition_id",
        "operation",
        "kind",
        "before",
        "after",
    }:
        raise DeliverableValidationError("Deliverable transition fields are invalid")
    transition_id = value.get("transition_id")
    operation = value.get("operation")
    kind = value.get("kind")
    if (
        isinstance(transition_id, bool)
        or not isinstance(transition_id, int)
        or not isinstance(operation, str)
        or not isinstance(kind, str)
    ):
        raise DeliverableValidationError("Deliverable transition values are invalid")
    try:
        return ArtifactTransition(
            transition_id=transition_id,
            operation=operation,
            kind=kind,
            before=_artifact_ref(value.get("before"), label="transition.before"),
            after=_artifact_ref(value.get("after"), label="transition.after"),
        )
    except ValueError as exc:
        raise DeliverableValidationError(f"Deliverable transition is invalid: {exc}") from exc


def _audit_render_witness(value: Any, *, root: Path, final: ArtifactRef) -> None:
    from .render import PNG_SIGNATURE

    if not isinstance(value, Mapping):
        raise DeliverableValidationError("Fresh final render witness is missing")
    payload = {key: item for key, item in value.items() if key != "witness_sha256"}
    if (
        value.get("schema_version") != "spreadsheet-final-render-witness-v1"
        or not _valid_sha256(value.get("witness_sha256"))
        or _canonical_sha256(payload, label="final render witness") != value.get("witness_sha256")
    ):
        raise DeliverableValidationError("Final render witness digest is invalid")
    if _artifact_ref(value.get("artifact"), label="render artifact") != final:
        raise DeliverableValidationError("Final render witness is stale")
    manifest_relative = _relative_certificate_path(
        value.get("manifest_relative_path"), label="render manifest path"
    )
    manifest = _path_inside(root, manifest_relative, label="render manifest")
    if not manifest.is_file() or _sha256(manifest) != value.get("manifest_sha256"):
        raise DeliverableValidationError("Final render manifest hash does not match")
    pages = value.get("pages")
    if not isinstance(pages, list) or len(pages) != value.get("page_count") or not pages:
        raise DeliverableValidationError("Final render page manifest is invalid")
    portable_manifest = {
        "schema_version": "spreadsheet-portable-render-manifest-v1",
        "artifact": value.get("artifact"),
        "backend": value.get("backend"),
        "version": value.get("version"),
        "mode": value.get("mode"),
        "dpi": value.get("dpi"),
        "page_count": value.get("page_count"),
        "pages": pages,
    }
    try:
        stored_manifest = json.loads(manifest.read_text(encoding="ascii"))
    except (OSError, UnicodeError, json.JSONDecodeError) as exc:
        raise DeliverableValidationError("Portable final render manifest is invalid") from exc
    if stored_manifest != portable_manifest:
        raise DeliverableValidationError("Portable final render manifest does not reproduce")
    _reject_absolute_certificate_paths(
        stored_manifest,
        label="portable final render manifest",
    )
    for index, page in enumerate(pages, start=1):
        if not isinstance(page, Mapping) or page.get("index") != index:
            raise DeliverableValidationError("Final render page ids are not contiguous")
        relative = _relative_certificate_path(page.get("relative_path"), label="render page path")
        path = _path_inside(root, relative, label="render page")
        data = path.read_bytes()
        if not data.startswith(PNG_SIGNATURE) or hashlib.sha256(data).hexdigest() != page.get(
            "sha256"
        ):
            raise DeliverableValidationError("Final render page hash or format does not match")


def _audit_visual_equivalence_witness(
    value: Any,
    *,
    certificate: Mapping[str, Any],
    root: Path,
    final: ArtifactRef,
) -> None:
    if not isinstance(value, Mapping):
        raise DeliverableValidationError("Visual render-equivalence witness is missing")
    payload = {key: item for key, item in value.items() if key != "witness_sha256"}
    if (
        value.get("schema_version") != "spreadsheet-render-equivalence-witness-v1"
        or not _valid_sha256(value.get("witness_sha256"))
        or _canonical_sha256(payload, label="visual equivalence witness")
        != value.get("witness_sha256")
        or _artifact_ref(value.get("final_artifact"), label="visual final artifact") != final
        or value.get("policy") != "carry-view-only-through-pixel-identical-final-render-v1"
        or value.get("all_required_pages_pixel_identical") is not True
    ):
        raise DeliverableValidationError("Visual render-equivalence witness is invalid")
    required_views = _required_candidate_views(certificate)
    groups = value.get("render_groups")
    if (
        not isinstance(groups, list)
        or value.get("candidate_required_page_count") != len(required_views)
        or value.get("mapped_page_count") != len(required_views)
    ):
        raise DeliverableValidationError("Visual equivalence page counts are invalid")
    expected_keys = sorted(
        {
            (
                str(view["render_id"]),
                str(view["render_manifest_sha256"]),
                int(view["dpi"]),
            )
            for view in required_views
        }
    )
    observed_keys: list[tuple[str, str, int]] = []
    observed_view_ids: list[int] = []
    for group in groups:
        if not isinstance(group, Mapping) or set(group) != {
            "candidate_render_id",
            "candidate_render_manifest_sha256",
            "dpi",
            "final_render",
            "pages",
        }:
            raise DeliverableValidationError("Visual equivalence render group is invalid")
        render_id = group.get("candidate_render_id")
        manifest_sha256 = group.get("candidate_render_manifest_sha256")
        dpi = group.get("dpi")
        if (
            not isinstance(render_id, str)
            or not _valid_sha256(manifest_sha256)
            or isinstance(dpi, bool)
            or not isinstance(dpi, int)
            or dpi <= 0
        ):
            raise DeliverableValidationError("Visual equivalence render identity is invalid")
        observed_keys.append((render_id, str(manifest_sha256), dpi))
        final_render = group.get("final_render")
        _audit_render_witness(final_render, root=root, final=final)
        group_views = [
            view
            for view in required_views
            if view["render_id"] == render_id
            and view["render_manifest_sha256"] == manifest_sha256
            and view["dpi"] == dpi
        ]
        raw_pages = group.get("pages")
        if not isinstance(raw_pages, list) or len(raw_pages) != len(group_views):
            raise DeliverableValidationError("Visual equivalence group pages are invalid")
        expected_pages: list[dict[str, Any]] = []
        assert isinstance(final_render, Mapping)
        for view in group_views:
            final_page = _final_page_for_candidate_view(final_render, view)
            if final_page.get("sha256") != view["page_sha256"]:
                raise DeliverableValidationError(
                    "Final page is not pixel-identical to the viewed candidate page"
                )
            expected_pages.append(
                {
                    "candidate_view_event_id": view["event_id"],
                    "page_id": view["page_id"],
                    "sheet": view["sheet"],
                    "sheet_page": view["sheet_page"],
                    "candidate_page_sha256": view["page_sha256"],
                    "final_page_relative_path": final_page["relative_path"],
                    "final_page_sha256": final_page["sha256"],
                    "pixel_identical": True,
                }
            )
            observed_view_ids.append(int(view["event_id"]))
        if raw_pages != expected_pages:
            raise DeliverableValidationError("Visual equivalence page mapping differs")
    if sorted(observed_keys) != expected_keys or sorted(observed_view_ids) != sorted(
        int(view["event_id"]) for view in required_views
    ):
        raise DeliverableValidationError("Visual equivalence omits or duplicates required views")


def _audit_deliverable_certificate(
    certificate: Any,
    *,
    agent_evidence: Mapping[str, Any],
    run_root: Path,
    output_workbook: Path,
) -> tuple[ArtifactRef, Path]:
    if not isinstance(certificate, Mapping):
        raise DeliverableValidationError("Deliverable certificate is missing")
    document = json.loads(json.dumps(certificate))
    schema_version = document.get("schema_version")
    expected = {
        "schema_version",
        "candidate",
        "lineage",
        "target_grounding",
        "postprocess",
        "evidence_policy",
        "final_artifact",
        "final_revision_witness",
        "visual_equivalence_witness",
        "scoring_copy",
        "certificate_sha256",
    }
    if schema_version == DELIVERABLE_CERTIFICATE_SCHEMA_VERSION:
        expected.add("recalculation_attestation")
    elif schema_version != LEGACY_DELIVERABLE_CERTIFICATE_SCHEMA_VERSION:
        raise DeliverableValidationError("Deliverable certificate schema is unsupported")
    if set(document) != expected:
        raise DeliverableValidationError("Deliverable certificate schema or fields are invalid")
    payload = {key: document[key] for key in document if key != "certificate_sha256"}
    _reject_absolute_certificate_paths(payload, label="deliverable certificate")
    if not _valid_sha256(document.get("certificate_sha256")) or _canonical_sha256(
        payload, label="deliverable certificate"
    ) != document.get("certificate_sha256"):
        raise DeliverableValidationError("Deliverable certificate digest does not match")

    candidate_block = document.get("candidate")
    if not isinstance(candidate_block, Mapping) or set(candidate_block) != {
        "outcome",
        "artifact",
        "submission",
        "evidence_certificate",
    }:
        raise DeliverableValidationError("Deliverable candidate block is invalid")
    candidate = _artifact_ref(candidate_block.get("artifact"), label="candidate artifact")
    candidate_outcome, evidence_certificate, submission = _candidate_outcome(
        agent_evidence,
        workspace=run_root,
        candidate=candidate,
    )
    if candidate_outcome != candidate_block.get("outcome"):
        raise DeliverableValidationError("Deliverable candidate outcome differs from result row")
    if evidence_certificate != candidate_block.get("evidence_certificate"):
        raise DeliverableValidationError("Deliverable candidate evidence differs from result row")
    if submission != candidate_block.get("submission"):
        raise DeliverableValidationError("Deliverable candidate submission differs from result row")

    final = _artifact_ref(document.get("final_artifact"), label="final artifact")
    lineage = document.get("lineage")
    if not isinstance(lineage, Mapping) or set(lineage) != {
        "initial_artifact",
        "candidate_artifact",
        "final_artifact",
        "candidate_transition_count",
        "transition_count",
        "transitions",
    }:
        raise DeliverableValidationError("Deliverable lineage fields are invalid")
    transitions_value = lineage.get("transitions")
    if not isinstance(transitions_value, list):
        raise DeliverableValidationError("Deliverable lineage transitions are invalid")
    transitions = tuple(_transition_from_dict(item) for item in transitions_value)
    candidate_transition_count = lineage.get("candidate_transition_count")
    if isinstance(candidate_transition_count, bool) or not isinstance(
        candidate_transition_count, int
    ):
        raise DeliverableValidationError("Candidate transition count is invalid")
    expected_lineage = _transition_lineage(
        transitions,
        candidate=candidate,
        candidate_transition_count=candidate_transition_count,
        final=final,
    )
    if lineage != expected_lineage:
        raise DeliverableValidationError("Deliverable lineage does not replay exactly")
    _audit_target_grounding_commit_chain(
        document.get("target_grounding"),
        transitions=transitions,
    )

    postprocess = document.get("postprocess")
    if (
        not isinstance(postprocess, Mapping)
        or set(postprocess)
        != {
            "operation",
            "timeout_seconds",
            "recalculation",
        }
        or postprocess.get("operation") != "session.recalculate"
    ):
        raise DeliverableValidationError("Deliverable postprocess record is invalid")
    timeout_seconds = postprocess.get("timeout_seconds")
    if (
        isinstance(timeout_seconds, bool)
        or not isinstance(timeout_seconds, int | float)
        or timeout_seconds <= 0
    ):
        raise DeliverableValidationError("Deliverable recalculation timeout is invalid")
    recalculation = _sanitize_recalculation(
        postprocess.get("recalculation"),
        before=candidate,
        after=final,
        candidate_outcome=candidate_outcome,
    )
    if recalculation != postprocess.get("recalculation"):
        raise DeliverableValidationError("Deliverable recalculation contains untrusted fields")
    from .ooxml_formula_scan import OOXML_NO_FORMULA_BACKEND

    if (
        schema_version == LEGACY_DELIVERABLE_CERTIFICATE_SCHEMA_VERSION
        and recalculation.get("backend") == OOXML_NO_FORMULA_BACKEND
    ):
        raise DeliverableValidationError(
            "Verified no-formula certificates cannot be downgraded to the legacy schema"
        )
    if schema_version == DELIVERABLE_CERTIFICATE_SCHEMA_VERSION:
        expected_attestation = _no_formula_attestation(
            recalculation,
            artifact=final,
        )
        if document.get("recalculation_attestation") != expected_attestation:
            raise DeliverableValidationError(
                "Deliverable no-formula finalization witness is invalid"
            )
    postprocess_count = len(transitions) - candidate_transition_count
    if postprocess_count not in {0, 1}:
        raise DeliverableValidationError("Deliverable postprocess transition count is invalid")
    if postprocess_count == 1:
        transition = transitions[-1]
        if (
            transition.operation != "recalculate"
            or transition.kind != "derived_recalculation"
            or transition.before != candidate
            or transition.after != final
            or recalculation.get("artifact_transition_id") != transition.transition_id
        ):
            raise DeliverableValidationError("Deliverable recalculation transition is invalid")
    elif candidate != final:
        raise DeliverableValidationError("Changed deliverable is missing recalculation transition")

    from .ooxml_formula_scan import OOXMLFormulaScanError, OOXMLFormulaScanLease

    try:
        with OOXMLFormulaScanLease.open(output_workbook) as audit_scan_lease:
            independent_scan = audit_scan_lease.scan
            if independent_scan.package_sha256 != final.sha256:
                raise DeliverableValidationError("Final output workbook SHA-256 does not match")
            if schema_version == DELIVERABLE_CERTIFICATE_SCHEMA_VERSION and (
                independent_scan.has_formulas
                or independent_scan.to_dict() != recalculation.get("formula_scan")
            ):
                raise DeliverableValidationError(
                    "Final output does not reproduce the no-formula attestation"
                )
            expected_witness = _scan_final_revision_snapshot(
                audit_scan_lease.snapshot_bytes,
                final,
                workbook_format=independent_scan.workbook_format,
            )
            audit_scan_lease.verify_binding(checkpoint="certificate_audit")
    except OOXMLFormulaScanError as exc:
        raise DeliverableValidationError(
            "Final output does not reproduce the no-formula attestation"
        ) from exc
    if document.get("final_revision_witness") != expected_witness:
        raise DeliverableValidationError("Final revision witness does not reproduce")

    policy = document.get("evidence_policy")
    candidate_has_visual_evidence = bool(
        evidence_certificate is not None and _candidate_has_visual_evidence(evidence_certificate)
    )
    expected_policy = {
        "accepted_candidate_evidence": candidate_outcome == _ACCEPTED_CANDIDATE,
        "candidate_evidence_carried_forward": (
            candidate_outcome == _ACCEPTED_CANDIDATE and candidate == final
        ),
        "changed_recalculation_invalidates_candidate_evidence": True,
        "fresh_final_revision_readback": True,
        "candidate_visual_evidence_present": candidate_has_visual_evidence,
        "pixel_equivalence_required_for_visual_carry": candidate != final
        and candidate_has_visual_evidence,
        "pixel_equivalence_observed": (document.get("visual_equivalence_witness") is not None),
        "unviewed_final_render_never_counts_as_viewed": True,
    }
    if policy != expected_policy:
        raise DeliverableValidationError("Deliverable evidence carry-forward policy is invalid")
    if expected_policy["pixel_equivalence_required_for_visual_carry"]:
        assert evidence_certificate is not None
        _audit_visual_equivalence_witness(
            document.get("visual_equivalence_witness"),
            certificate=evidence_certificate,
            root=run_root,
            final=final,
        )
    elif document.get("visual_equivalence_witness") is not None:
        raise DeliverableValidationError("Unexpected visual render-equivalence witness is present")

    scoring = document.get("scoring_copy")
    if not isinstance(scoring, Mapping) or set(scoring) != {
        "relative_path",
        "source_artifact",
        "artifact_role",
        "creates_artifact_transition",
        "sha256",
        "byte_identical",
        "read_only",
    }:
        raise DeliverableValidationError("Scoring copy certificate fields are invalid")
    relative = _relative_certificate_path(scoring.get("relative_path"), label="scoring copy")
    if relative != Path(SCORING_COPY_RELATIVE_PATH):
        raise DeliverableValidationError("Scoring copy path is not the fixed publication name")
    scoring_path = _path_inside(run_root, relative, label="scoring copy")
    try:
        metadata = scoring_path.lstat()
    except OSError as exc:
        raise DeliverableValidationError("Scoring copy is missing") from exc
    if (
        not stat.S_ISREG(metadata.st_mode)
        or stat.S_ISLNK(metadata.st_mode)
        or metadata.st_mode & 0o222
    ):
        raise DeliverableValidationError(
            "Scoring copy is not a read-only regular non-symbolic file"
        )
    if (
        _artifact_ref(scoring.get("source_artifact"), label="scoring source") != final
        or scoring.get("artifact_role") != "same_revision_replica"
        or scoring.get("creates_artifact_transition") is not False
        or scoring.get("sha256") != final.sha256
        or scoring.get("byte_identical") is not True
        or scoring.get("read_only") is not True
        or _sha256(scoring_path) != final.sha256
    ):
        raise DeliverableValidationError("Scoring copy is not byte-identical to final artifact")
    return final, scoring_path


def audit_deliverable_certificate(
    certificate: Any,
    *,
    agent_evidence: Mapping[str, Any],
    run_root: Path,
    output_workbook: Path,
) -> DeliverableAudit:
    """Fail-closed artifact audit with stable, non-sensitive reason codes."""

    try:
        final, scoring_copy = _audit_deliverable_certificate(
            certificate,
            agent_evidence=agent_evidence,
            run_root=run_root,
            output_workbook=output_workbook,
        )
    except (DeliverableValidationError, OSError, ValueError) as exc:
        return DeliverableAudit(
            valid=False,
            reasons=(f"deliverable_lineage_invalid:{type(exc).__name__}",),
        )
    return DeliverableAudit(
        valid=True,
        reasons=(),
        final_artifact=final,
        scoring_copy=scoring_copy,
    )


__all__ = [
    "ADVISORY_TARGET_GROUNDING_CERTIFICATE_SCHEMA_VERSION",
    "COMPARISON_RESULT_SCHEMA_VERSION",
    "DELIVERABLE_CERTIFICATE_SCHEMA_VERSION",
    "LEGACY_DELIVERABLE_CERTIFICATE_SCHEMA_VERSION",
    "NO_FORMULA_ATTESTATION_SCHEMA_VERSION",
    "DeliverableAudit",
    "DeliverableBundle",
    "DeliverableValidationError",
    "SCORING_COPY_RELATIVE_PATH",
    "audit_deliverable_certificate",
    "finalize_deliverable",
    "score_read_only",
    "validate_evidence_certificate",
]
