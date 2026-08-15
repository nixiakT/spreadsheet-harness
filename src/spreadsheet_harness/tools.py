"""Model-facing spreadsheet tool registry."""

from __future__ import annotations

import json
import re
import uuid
from collections.abc import Callable
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

from .code_interpreter import LocalCodeInterpreter
from .errors import (
    CodeIsolationError,
    HarnessError,
    RecalculationIntegrityError,
    ToolInputError,
)
from .session import WorkbookSession

_INSPECT_MAX_CELLS = 500
_LATEX_MAX_CELLS = 500
_LATEX_MAX_CHARS = 65_536
_LATEX_MAX_CELL_CHARS = 512
_STYLE_MAX_GROUPS = 64
_STYLE_MAX_SAMPLE_CELLS = 24
_STYLE_MAX_TEXT_CHARS = 160
_INSPECT_DEFAULT_INCLUDE_STYLES = False
_CALCULATION_VALIDATION_MAX_CELLS = 500
_CALCULATION_ERROR_COORDINATE_LIMIT = 32
_SPREADSHEET_ERROR_VALUES = frozenset(
    {
        "#BLOCKED!",
        "#BUSY!",
        "#CALC!",
        "#CONNECT!",
        "#DIV/0!",
        "#FIELD!",
        "#GETTING_DATA",
        "#N/A",
        "#NAME?",
        "#NULL!",
        "#NUM!",
        "#PYTHON!",
        "#REF!",
        "#SPILL!",
        "#UNKNOWN!",
        "#VALUE!",
    }
)
_LIBREOFFICE_ERROR_VALUE = re.compile(r"Err:\s*\d{3}\Z", re.IGNORECASE)


def _bounded_inspection_range(range_ref: str) -> tuple[str, dict[str, Any]]:
    try:
        bounds = range_boundaries(range_ref.replace("$", ""))
    except (AttributeError, TypeError, ValueError) as exc:
        raise ToolInputError(f"Invalid A1 range: {range_ref!r}") from exc
    min_col, min_row, max_col, max_row = bounds
    if not all(isinstance(item, int) and item >= 1 for item in bounds):
        raise ToolInputError(f"Range must be bounded: {range_ref!r}")
    width = max_col - min_col + 1
    height = max_row - min_row + 1
    if width <= 0 or height <= 0:
        raise ToolInputError(f"Invalid A1 range: {range_ref!r}")

    requested_cell_count = width * height
    requested_range = (
        f"{get_column_letter(min_col)}{min_row}:"
        f"{get_column_letter(max_col)}{max_row}"
    )
    if requested_cell_count <= _INSPECT_MAX_CELLS:
        return requested_range, {
            "requested_range": requested_range,
            "requested_cell_count": requested_cell_count,
            "returned_cell_count": requested_cell_count,
            "truncated": False,
        }

    # Preserve all requested columns when they fit, then return as many complete
    # top rows as the limit allows. Very wide ranges return the first row prefix.
    returned_width = min(width, _INSPECT_MAX_CELLS)
    returned_height = min(height, max(1, _INSPECT_MAX_CELLS // returned_width))
    returned_max_col = min_col + returned_width - 1
    returned_max_row = min_row + returned_height - 1
    returned_range = (
        f"{get_column_letter(min_col)}{min_row}:"
        f"{get_column_letter(returned_max_col)}{returned_max_row}"
    )
    returned_cell_count = returned_width * returned_height
    return returned_range, {
        "requested_range": requested_range,
        "requested_cell_count": requested_cell_count,
        "returned_cell_count": returned_cell_count,
        "truncated": True,
        "omitted_cell_count": requested_cell_count - returned_cell_count,
        "policy": "top_left_rectangle_preserve_columns_when_possible",
    }


def _spreadsheet_error_value(cell: dict[str, Any]) -> str | None:
    value = cell.get("value")
    if not isinstance(value, str):
        return None
    candidate = value.strip()
    normalized = candidate.upper()
    is_error_typed = (
        cell.get("data_type") == "e" or cell.get("cached_data_type") == "e"
    )
    has_formula = cell.get("formula") is not None
    if normalized in _SPREADSHEET_ERROR_VALUES and is_error_typed:
        return normalized
    if is_error_typed and normalized.startswith("#"):
        return normalized
    if _LIBREOFFICE_ERROR_VALUE.fullmatch(candidate) and (is_error_typed or has_formula):
        return candidate
    return None


def _calculation_error_summary(inspection: dict[str, Any]) -> dict[str, Any]:
    errors: list[tuple[str, str]] = []
    for cell in inspection.get("cells", []):
        if not isinstance(cell, dict):
            continue
        error = _spreadsheet_error_value(cell)
        coordinate = cell.get("coordinate")
        if error is not None and isinstance(coordinate, str):
            errors.append((coordinate, error))

    counts: dict[str, int] = {}
    for _, error in errors:
        counts[error] = counts.get(error, 0) + 1
    sampled = errors[:_CALCULATION_ERROR_COORDINATE_LIMIT]
    return {
        "sheet": inspection.get("sheet"),
        "range": inspection.get("range"),
        "count": len(errors),
        "by_error": dict(sorted(counts.items())),
        "coordinates": [
            {"coordinate": coordinate, "error": error}
            for coordinate, error in sampled
        ],
        "coordinate_limit": _CALCULATION_ERROR_COORDINATE_LIMIT,
        "coordinates_truncated": len(sampled) < len(errors),
    }


def _bounded_text(value: Any, limit: int) -> str | None:
    if value is None:
        return None
    text = str(value)
    if len(text) <= limit:
        return text
    return text[: max(0, limit - 1)] + "…"


def _latex_escape(value: Any, *, max_chars: int) -> tuple[str, bool]:
    """Escape one scalar without ever cutting through a LaTeX escape sequence."""
    if value is None:
        return "", False
    replacements = {
        "\\": r"\textbackslash{}",
        "&": r"\&",
        "%": r"\%",
        "$": r"\$",
        "#": r"\#",
        "_": r"\_",
        "{": r"\{",
        "}": r"\}",
        "~": r"\textasciitilde{}",
        "^": r"\textasciicircum{}",
        "\n": r"\newline{}",
        "\t": " ",
    }
    output: list[str] = []
    length = 0
    truncated = False
    for character in str(value).replace("\r\n", "\n").replace("\r", "\n"):
        token = replacements.get(character, character if character.isprintable() else " ")
        if length + len(token) > max_chars:
            truncated = True
            break
        output.append(token)
        length += len(token)
    if truncated:
        marker = r"\ldots{}"
        while output and length + len(marker) > max_chars:
            removed = output.pop()
            length -= len(removed)
        if len(marker) <= max_chars:
            output.append(marker)
    return "".join(output), truncated


def _style_summary(cells: list[dict[str, Any]]) -> dict[str, Any]:
    groups: dict[int, dict[str, Any]] = {}
    for cell in cells:
        style = cell.get("style")
        if not isinstance(style, dict):
            continue
        style_id = int(style.get("style_id", 0))
        group = groups.get(style_id)
        if group is None:
            font = style.get("font") if isinstance(style.get("font"), dict) else {}
            alignment = (
                style.get("alignment") if isinstance(style.get("alignment"), dict) else {}
            )
            group = {
                "style_id": style_id,
                "cell_count": 0,
                "sample_cells": [],
                "sample_cells_truncated": False,
                "number_format": _bounded_text(style.get("number_format"), _STYLE_MAX_TEXT_CHARS),
                "font": {
                    "bold": bool(font.get("bold")),
                    "italic": bool(font.get("italic")),
                    "color": _bounded_text(font.get("color"), _STYLE_MAX_TEXT_CHARS),
                },
                "fill": _bounded_text(style.get("fill"), _STYLE_MAX_TEXT_CHARS),
                "alignment": {
                    "horizontal": _bounded_text(
                        alignment.get("horizontal"), _STYLE_MAX_TEXT_CHARS
                    ),
                    "vertical": _bounded_text(
                        alignment.get("vertical"), _STYLE_MAX_TEXT_CHARS
                    ),
                    "wrap_text": alignment.get("wrap_text"),
                },
            }
            groups[style_id] = group
        group["cell_count"] += 1
        samples = group["sample_cells"]
        if len(samples) < _STYLE_MAX_SAMPLE_CELLS:
            samples.append(str(cell.get("coordinate", "")))
        else:
            group["sample_cells_truncated"] = True

    all_groups = list(groups.values())
    returned = all_groups[:_STYLE_MAX_GROUPS]
    return {
        "distinct_style_count": len(all_groups),
        "styles_returned": len(returned),
        "truncated": len(returned) < len(all_groups),
        "styles": returned,
    }


def _matrix_to_latex(matrix: list[list[Any]]) -> tuple[str, int, int]:
    row_count = len(matrix)
    column_count = max((len(row) for row in matrix), default=0)
    cell_count = row_count * column_count
    column_spec = "l" * column_count
    prefix = rf"\begin{{tabular}}{{{column_spec}}}" + "\n"
    suffix = "\n" + r"\end{tabular}"
    separator_chars = row_count * len(" \\\\") + max(0, cell_count - row_count) * len(" & ")
    newline_chars = max(0, row_count - 1)
    fixed_chars = len(prefix) + len(suffix) + separator_chars + newline_chars
    available = max(0, _LATEX_MAX_CHARS - fixed_chars)
    per_cell_limit = min(
        _LATEX_MAX_CELL_CHARS,
        available // cell_count if cell_count else _LATEX_MAX_CELL_CHARS,
    )

    rendered_rows: list[str] = []
    truncated_cells = 0
    for row in matrix:
        padded = list(row) + [None] * (column_count - len(row))
        rendered: list[str] = []
        for value in padded:
            escaped, truncated = _latex_escape(value, max_chars=per_cell_limit)
            rendered.append(escaped)
            truncated_cells += int(truncated)
        rendered_rows.append(" & ".join(rendered) + " \\\\")
    latex = prefix + "\n".join(rendered_rows) + suffix
    if len(latex) > _LATEX_MAX_CHARS:  # Defensive guard if structure changes later.
        raise RuntimeError("Internal LaTeX output limit was exceeded")
    return latex, truncated_cells, per_cell_limit


@dataclass
class ToolOutcome:
    data: dict[str, Any]
    image_path: Path | None = None

    def as_output(self) -> str:
        return json.dumps(self.data, ensure_ascii=False, default=str)


def _object_schema(properties: dict[str, Any], required: list[str]) -> dict[str, Any]:
    return {
        "type": "object",
        "properties": properties,
        "required": required,
        "additionalProperties": False,
    }


class SpreadsheetToolRegistry:
    def __init__(
        self,
        session: WorkbookSession,
        *,
        enable_code: bool = True,
        allowed_tools: set[str] | None = None,
        require_code_isolation: bool = False,
        redaction_secrets: tuple[str, ...] = (),
    ) -> None:
        self.session = session
        self._allowed_tools = frozenset(allowed_tools) if allowed_tools is not None else None
        self.interpreter = (
            LocalCodeInterpreter(
                session.workspace,
                session.workbook_path,
                require_isolation=require_code_isolation,
                secrets=redaction_secrets,
            )
            if enable_code
            else None
        )
        self._last_render: dict[str, Any] | None = None
        self._handlers: dict[str, Callable[[dict[str, Any]], ToolOutcome]] = {
            "list_sheets": self._list_sheets,
            "inspect_range": self._inspect_range,
            "range_to_latex": self._range_to_latex,
            "find_cells": self._find_cells,
            "write_range": self._write_range,
            "fill_formula": self._fill_formula,
            "format_range": self._format_range,
            "clear_range": self._clear_range,
            "delete_rows": self._delete_rows,
            "delete_columns": self._delete_columns,
            "manage_sheet": self._manage_sheet,
            "recalculate_and_read": self._recalculate_and_read,
            "render_workbook": self._render_workbook,
            "view_image": self._view_image,
            "undo_last": self._undo_last,
        }
        if self.interpreter:
            self._handlers["code_interpreter"] = self._code_interpreter

    @property
    def schemas(self) -> list[dict[str, Any]]:
        sheet = {"type": "string", "description": "Exact worksheet name."}
        a1 = {"type": "string", "description": "Bounded A1 range such as A1:F20."}
        inspect_a1 = {
            "type": "string",
            "description": (
                "Bounded A1 range such as A1:F20. Runtime output is limited to "
                f"{_INSPECT_MAX_CELLS} cells; "
                "larger requests return a deterministic truncated top-left rectangle with "
                "the requested and returned ranges reported."
            ),
        }
        calculation_a1 = {
            "type": "string",
            "description": (
                "Bounded A1 validation target such as A1:F20, with a maximum of "
                f"{_CALCULATION_VALIDATION_MAX_CELLS} cells. Larger targets are rejected "
                "before recalculation; split them into fully inspected ranges."
            ),
        }
        schemas = [
            self._schema(
                "list_sheets",
                "List worksheets, visibility, used dimensions, merges and tables.",
                _object_schema({}, []),
            ),
            self._schema(
                "inspect_range",
                (
                    "Read formulas, cached values, merges and tables in a bounded range, with "
                    f"a runtime maximum of {_INSPECT_MAX_CELLS} returned cells. Larger "
                    "requests succeed with "
                    "explicit deterministic truncation metadata instead of failing. "
                    "Set include_styles=true only for formatting/layout tasks because style "
                    "details are verbose."
                ),
                _object_schema(
                    {
                        "sheet": sheet,
                        "range_ref": inspect_a1,
                        "include_styles": {
                            "type": "boolean",
                            "default": _INSPECT_DEFAULT_INCLUDE_STYLES,
                        },
                    },
                    ["sheet", "range_ref"],
                ),
            ),
            self._schema(
                "range_to_latex",
                "Convert up to 500 cells to a bounded, safely escaped LaTeX tabular and summarize merges and styles.",
                _object_schema({"sheet": sheet, "range_ref": a1}, ["sheet", "range_ref"]),
            ),
            self._schema(
                "find_cells",
                "Search cell values or formulas across one sheet or the workbook.",
                _object_schema(
                    {
                        "query": {"type": "string"},
                        "sheet": {"type": ["string", "null"]},
                        "use_regex": {"type": "boolean", "default": False},
                        "match_case": {"type": "boolean", "default": False},
                        "search_formulas": {"type": "boolean", "default": True},
                    },
                    ["query"],
                ),
            ),
            self._schema(
                "write_range",
                "Write a rectangular 2-D array starting at one cell. Strings beginning with '=' are formulas.",
                _object_schema(
                    {
                        "sheet": sheet,
                        "start_cell": {"type": "string", "description": "Top-left cell, e.g. C4."},
                        "values": {
                            "type": "array",
                            "items": {"type": "array", "items": {}},
                            "minItems": 1,
                        },
                    },
                    ["sheet", "start_cell", "values"],
                ),
            ),
            self._schema(
                "fill_formula",
                (
                    "Translate one source formula into every cell of a target range, preserving "
                    "relative references. Returns sample translated formulas and warnings for "
                    "ranges that may drift during fill. If target_range is a single cell "
                    "different from source_cell, it is treated as the endpoint of source_cell:target_range."
                ),
                _object_schema(
                    {"sheet": sheet, "source_cell": {"type": "string"}, "target_range": a1},
                    ["sheet", "source_cell", "target_range"],
                ),
            ),
            self._schema(
                "format_range",
                "Apply number/font/fill/alignment/border/size formatting without changing values.",
                _object_schema(
                    {"sheet": sheet, "range_ref": a1, "format_spec": {"type": "object"}},
                    ["sheet", "range_ref", "format_spec"],
                ),
            ),
            self._schema(
                "clear_range",
                "Clear cell contents and optionally formatting in a bounded range.",
                _object_schema(
                    {
                        "sheet": sheet,
                        "range_ref": a1,
                        "contents": {"type": "boolean", "default": True},
                        "formats": {"type": "boolean", "default": False},
                    },
                    ["sheet", "range_ref"],
                ),
            ),
            self._schema(
                "delete_rows",
                "Delete worksheet rows. This shifts cells and can affect references.",
                _object_schema(
                    {
                        "sheet": sheet,
                        "start": {"type": "integer", "minimum": 1},
                        "amount": {"type": "integer", "minimum": 1, "default": 1},
                    },
                    ["sheet", "start"],
                ),
            ),
            self._schema(
                "delete_columns",
                "Delete worksheet columns. This shifts cells and can affect references.",
                _object_schema(
                    {
                        "sheet": sheet,
                        "start": {"type": "integer", "minimum": 1},
                        "amount": {"type": "integer", "minimum": 1, "default": 1},
                    },
                    ["sheet", "start"],
                ),
            ),
            self._schema(
                "manage_sheet",
                "Create, rename, delete or copy a worksheet.",
                _object_schema(
                    {
                        "action": {
                            "type": "string",
                            "enum": ["create", "rename", "delete", "copy"],
                        },
                        "name": {"type": "string"},
                        "new_name": {"type": ["string", "null"]},
                        "source": {"type": ["string", "null"]},
                        "index": {"type": ["integer", "null"], "minimum": 0},
                    },
                    ["action", "name"],
                ),
            ),
            self._schema(
                "recalculate_and_read",
                (
                    "Recalculate a safe workspace copy with LibreOffice, replace the working "
                    "copy atomically, then inspect a range. ok=true reports transport/tool "
                    "success; calculation_valid separately reports whether the inspected "
                    "target contains spreadsheet error values. Validation targets are limited "
                    f"to {_CALCULATION_VALIDATION_MAX_CELLS} cells and oversized requests fail "
                    "before recalculation."
                ),
                _object_schema(
                    {"sheet": sheet, "range_ref": calculation_a1},
                    ["sheet", "range_ref"],
                ),
            ),
            self._schema(
                "render_workbook",
                "Render the current workbook to per-sheet PNG pages with headless LibreOffice.",
                _object_schema(
                    {"dpi": {"type": "integer", "minimum": 72, "maximum": 240, "default": 144}}, []
                ),
            ),
            self._schema(
                "view_image",
                "Return one original PNG page to the vision-capable model. Call render_workbook first.",
                _object_schema(
                    {
                        "image_path": {
                            "type": "string",
                            "description": "Path returned by render_workbook.",
                        }
                    },
                    ["image_path"],
                ),
            ),
            self._schema(
                "undo_last",
                "Restore the snapshot immediately before the latest committed mutation.",
                _object_schema({}, []),
            ),
        ]
        if self.interpreter:
            schemas.append(
                self._schema(
                    "code_interpreter",
                    (
                        "Run trusted Python in the task workspace for analysis and direct "
                        "workbook edits. The preloaded sheet_harness.load_workbook() and "
                        "sheet_harness.save_workbook(wb) no-path calls load and save the managed "
                        "SHEET_WORKBOOK; never spell or guess its path. "
                        "sheet_harness.workbook_overview(wb) returns a list of dictionaries, "
                        "one per worksheet. Openpyxl compatibility shims are also preloaded. "
                        "Every call starts a fresh Python process: variables, "
                        "imports, and workbook objects do not persist across calls. Make each "
                        "editing or recovery script self-contained: import, load, re-read the "
                        "request and inspected workbook state, edit, save, close, reopen, verify "
                        "the requested change and nearby cells, then print compact verification."
                    ),
                    _object_schema(
                        {
                            "code": {"type": "string"},
                            "timeout_seconds": {
                                "type": "integer",
                                "minimum": 1,
                                "maximum": 60,
                                "default": 30,
                            },
                        },
                        ["code"],
                    ),
                )
            )
        if self._allowed_tools is None:
            return schemas
        return [schema for schema in schemas if schema["name"] in self._allowed_tools]

    def _schema(self, name: str, description: str, parameters: dict[str, Any]) -> dict[str, Any]:
        return {
            "type": "function",
            "name": name,
            "description": description,
            "parameters": parameters,
            "strict": False,
        }

    def invoke(self, name: str, arguments: dict[str, Any]) -> ToolOutcome:
        if self._allowed_tools is not None and name not in self._allowed_tools:
            return ToolOutcome(
                {"ok": False, "error": f"Unknown tool: {name}", "type": "UnknownTool"}
            )
        handler = self._handlers.get(name)
        if handler is None:
            return ToolOutcome(
                {"ok": False, "error": f"Unknown tool: {name}", "type": "UnknownTool"}
            )
        self.session.recorder.record("tool.called", {"name": name, "arguments": arguments})
        try:
            outcome = handler(arguments)
        except RecalculationIntegrityError as exc:
            # Sheet-identity drift is a no-score infrastructure condition. It
            # must reach the runner rather than becoming model-visible output.
            self.session.recorder.record(
                "tool.failed",
                {
                    "name": name,
                    "error_type": type(exc).__name__,
                    "failure_category": "recalculation_infrastructure",
                    "recalculation": exc.evidence,
                },
            )
            raise
        except CodeIsolationError:
            # Required comparison isolation is a harness invariant, not a
            # recoverable model tool error. Never let the agent continue.
            raise
        except (HarnessError, TypeError, ValueError, KeyError) as exc:
            outcome = ToolOutcome({"ok": False, "error": str(exc), "type": type(exc).__name__})
        except Exception as exc:  # keep a malformed tool call from killing the agent loop
            outcome = ToolOutcome({"ok": False, "error": str(exc), "type": type(exc).__name__})
        self.session.recorder.record("tool.returned", {"name": name, "result": outcome.data})
        return outcome

    def _list_sheets(self, _: dict[str, Any]) -> ToolOutcome:
        return ToolOutcome(self.session.list_sheets())

    def _inspect_range(self, args: dict[str, Any]) -> ToolOutcome:
        returned_range, bounds = _bounded_inspection_range(args["range_ref"])
        inspection = self.session.inspect_range(
            args["sheet"],
            returned_range,
            include_styles=args.get(
                "include_styles", _INSPECT_DEFAULT_INCLUDE_STYLES
            ),
            max_cells=_INSPECT_MAX_CELLS,
        )
        result = {
            "ok": inspection["ok"],
            "sheet": inspection["sheet"],
            "requested_range": bounds["requested_range"],
            "returned_range": inspection["range"],
            "range": inspection["range"],
            "requested_cell_count": bounds["requested_cell_count"],
            "returned_cell_count": bounds["returned_cell_count"],
            "cell_count": bounds["returned_cell_count"],
            "truncated": bounds["truncated"],
            "limits": {"max_cells": _INSPECT_MAX_CELLS},
        }
        if bounds["truncated"]:
            result["truncation"] = {
                "policy": bounds["policy"],
                "omitted_cell_count": bounds["omitted_cell_count"],
                "message": (
                    f"The requested range exceeded {_INSPECT_MAX_CELLS} cells; this response "
                    "contains the reported deterministic top-left rectangle."
                ),
            }
        result.update(
            {
                key: value
                for key, value in inspection.items()
                if key not in result
            }
        )
        return ToolOutcome(result)

    def _range_to_latex(self, args: dict[str, Any]) -> ToolOutcome:
        inspection = self.session.inspect_range(
            args["sheet"], args["range_ref"], include_styles=True, max_cells=_LATEX_MAX_CELLS
        )
        matrix = inspection["matrix"]
        latex, truncated_cells, per_cell_limit = _matrix_to_latex(matrix)
        row_count = len(matrix)
        column_count = max((len(row) for row in matrix), default=0)
        return ToolOutcome(
            {
                "ok": True,
                "sheet": inspection["sheet"],
                "range": inspection["range"],
                "rows": row_count,
                "columns": column_count,
                "cell_count": row_count * column_count,
                "latex": latex,
                "latex_truncated": truncated_cells > 0,
                "truncated_cell_count": truncated_cells,
                "merged_ranges": inspection["merged_ranges"],
                "style_summary": _style_summary(inspection["cells"]),
                "limits": {
                    "max_cells": _LATEX_MAX_CELLS,
                    "max_latex_chars": _LATEX_MAX_CHARS,
                    "max_cell_latex_chars": per_cell_limit,
                },
            }
        )

    def _find_cells(self, args: dict[str, Any]) -> ToolOutcome:
        return ToolOutcome(
            self.session.find_cells(
                args["query"],
                sheet=args.get("sheet"),
                use_regex=args.get("use_regex", False),
                match_case=args.get("match_case", False),
                search_formulas=args.get("search_formulas", True),
            )
        )

    def _write_range(self, args: dict[str, Any]) -> ToolOutcome:
        return ToolOutcome(
            self.session.write_range(args["sheet"], args["start_cell"], args["values"])
        )

    def _fill_formula(self, args: dict[str, Any]) -> ToolOutcome:
        return ToolOutcome(
            self.session.fill_formula(args["sheet"], args["source_cell"], args["target_range"])
        )

    def _format_range(self, args: dict[str, Any]) -> ToolOutcome:
        return ToolOutcome(
            self.session.format_range(args["sheet"], args["range_ref"], args["format_spec"])
        )

    def _clear_range(self, args: dict[str, Any]) -> ToolOutcome:
        return ToolOutcome(
            self.session.clear_range(
                args["sheet"],
                args["range_ref"],
                contents=args.get("contents", True),
                formats=args.get("formats", False),
            )
        )

    def _delete_rows(self, args: dict[str, Any]) -> ToolOutcome:
        return ToolOutcome(
            self.session.delete_rows(args["sheet"], args["start"], args.get("amount", 1))
        )

    def _delete_columns(self, args: dict[str, Any]) -> ToolOutcome:
        return ToolOutcome(
            self.session.delete_columns(args["sheet"], args["start"], args.get("amount", 1))
        )

    def _manage_sheet(self, args: dict[str, Any]) -> ToolOutcome:
        return ToolOutcome(
            self.session.manage_sheet(
                args["action"],
                args["name"],
                new_name=args.get("new_name"),
                source=args.get("source"),
                index=args.get("index"),
            )
        )

    def _render_workbook(self, args: dict[str, Any]) -> ToolOutcome:
        from .render import render_workbook

        output_dir = self.session.paths.artifacts / "render" / f"render-{uuid.uuid4().hex[:12]}"
        result = render_workbook(self.session.workbook_path, output_dir, dpi=args.get("dpi", 144))
        self._last_render = result.to_dict()
        return ToolOutcome({"ok": True, **self._last_render})

    def _view_image(self, args: dict[str, Any]) -> ToolOutcome:
        candidate = Path(args["image_path"])
        if not candidate.is_absolute():
            candidate = (self.session.workspace / candidate).resolve()
        else:
            candidate = candidate.resolve()
        render_root = (self.session.paths.artifacts / "render").resolve()
        if render_root not in candidate.parents or candidate.suffix.lower() != ".png":
            raise ToolInputError(
                "view_image only accepts PNGs produced in this run's render directory"
            )
        if self._last_render is None:
            raise ToolInputError("Call render_workbook before view_image")
        latest_pages = self._last_render.get("pages")
        latest_paths = {
            Path(page["image_path"]).resolve()
            for page in latest_pages
            if isinstance(page, dict) and isinstance(page.get("image_path"), str)
        } if isinstance(latest_pages, list) else set()
        if candidate not in latest_paths:
            raise ToolInputError(
                "view_image only accepts a page returned by the most recent render_workbook call"
            )
        if not candidate.is_file():
            raise ToolInputError(f"Rendered image does not exist: {candidate}")
        from PIL import Image

        with Image.open(candidate) as image:
            width, height = image.size
            image.verify()
        return ToolOutcome(
            {
                "ok": True,
                "image_path": str(candidate),
                "width": width,
                "height": height,
                "message": "The original PNG is attached in the next multimodal input item.",
            },
            image_path=candidate,
        )

    def _recalculate_and_read(self, args: dict[str, Any]) -> ToolOutcome:
        normalized_range, request = _bounded_inspection_range(args["range_ref"])
        if request["requested_cell_count"] > _CALCULATION_VALIDATION_MAX_CELLS:
            raise ToolInputError(
                "recalculate_and_read validates at most "
                f"{_CALCULATION_VALIDATION_MAX_CELLS} cells; requested "
                f"{request['requested_range']} contains "
                f"{request['requested_cell_count']} cells. Split the target into smaller "
                "ranges; no recalculation was performed."
            )
        metadata = self.session.recalculate()
        inspected = self.session.inspect_range(
            args["sheet"],
            normalized_range,
            max_cells=_CALCULATION_VALIDATION_MAX_CELLS,
        )
        calculation_errors = _calculation_error_summary(inspected)
        return ToolOutcome(
            {
                "ok": True,
                "calculation_valid": calculation_errors["count"] == 0,
                "calculation_errors": calculation_errors,
                "limits": {"max_cells": _CALCULATION_VALIDATION_MAX_CELLS},
                "calculation": metadata,
                "inspection": inspected,
            }
        )

    def _undo_last(self, _: dict[str, Any]) -> ToolOutcome:
        return ToolOutcome(self.session.undo_last())

    def _code_interpreter(self, args: dict[str, Any]) -> ToolOutcome:
        if not self.interpreter:
            raise ToolInputError("code_interpreter is disabled")
        return ToolOutcome(
            self.interpreter.run(args["code"], timeout_seconds=args.get("timeout_seconds"))
        )
