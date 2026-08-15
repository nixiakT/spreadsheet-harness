"""Read-only integrity audit for completed SpreadsheetBench comparisons."""

from __future__ import annotations

import hashlib
import json
import os
import stat
import subprocess
from collections import Counter
from collections.abc import Iterable
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .agent import MODEL_RESPONSE_TRUNCATED_TERMINAL
from .arms import COMPARISON_FORCED_TOOL_PREFIX_POLICY, comparison_stage_turn_caps
from .benchmark import (
    SpreadsheetTask,
    _source_fingerprint,
    compare_workbooks,
    compare_workbooks_chartsheet_safe,
    verify_trace2skill_split_provenance,
)
from .comparison import (
    COMPARISON_CONFIGURATION_POLICIES,
    COMPARISON_MANIFEST_SCHEMA_VERSION,
    COMPARISON_PROTOCOL_VERSION,
    CONTINUATION_SOURCE_FILENAME,
    INFLIGHT_FILENAME,
    INTERRUPTED_SEALS_FILENAME,
    LEGACY_COMPARISON_CONFIGURATION_POLICIES,
    LEGACY_COMPARISON_MANIFEST_SCHEMA_VERSION,
    LEGACY_COMPARISON_PROTOCOL_VERSION,
    LEGACY_PILOT_MANIFEST_SHA256,
    PROVIDER_INFRASTRUCTURE_CATEGORIES,
    PROVIDER_INFRASTRUCTURE_FAILURE_STAGE,
    PROVIDER_REQUEST_AUDIT_SCHEMA_VERSION,
    RUN_SPEC_COPY_FILENAME,
    TERMINAL_SUBMISSION_TRUNCATED_OBSERVED,
    V24_COMPARISON_CONFIGURATION_POLICIES,
    V24_COMPARISON_MANIFEST_SCHEMA_VERSION,
    V24_COMPARISON_PROTOCOL_VERSION,
    V25_COMPARISON_CONFIGURATION_POLICIES,
    V25_COMPARISON_MANIFEST_SCHEMA_VERSION,
    V25_COMPARISON_PROTOCOL_VERSION,
    V26_COMPARISON_CONFIGURATION_POLICIES,
    V26_COMPARISON_MANIFEST_SCHEMA_VERSION,
    V26_COMPARISON_PROTOCOL_VERSION,
    V27_COMPARISON_CONFIGURATION_POLICIES,
    V27_COMPARISON_MANIFEST_SCHEMA_VERSION,
    V27_COMPARISON_PROTOCOL_VERSION,
    V28_COMPARISON_CONFIGURATION_POLICIES,
    V28_COMPARISON_MANIFEST_SCHEMA_VERSION,
    V28_COMPARISON_PROTOCOL_VERSION,
    _allowed_observed_terminals_policy,
    _request_attempt_audit,
    _stage_allowed_tools_policy,
    manifest_execution_contract,
    parse_pilot_run_spec_bytes,
    protected_run_spec_split_ids,
    resolve_run_spec_anchor,
    verify_pilot_run_spec_contract,
    verify_pilot_run_spec_provenance,
)
from .errors import (
    AGENT_TOOL_RECALCULATION_FAILURE_STAGE,
    POSTPROCESS_RECALCULATION_FAILURE_STAGE,
    RECALCULATION_VALIDATION_TOOL,
    HarnessError,
    RenderError,
    ScoringInfrastructureError,
)
from .render import (
    RECALCULATION_SHEET_INTEGRITY_POLICY,
    openpyxl_worksheet_view,
    sheet_inventory_identity,
)
from .trajectory import read_trajectory


@dataclass(frozen=True)
class _AuditProtocolContract:
    protocol_version: str
    manifest_schema_version: int
    configuration_policies: dict[str, Any]
    allowed_model_failure_reasons: frozenset[str]
    require_v24_outcome_fields: bool
    strict_current_source: bool
    allow_budget_exhaustion_evidence: bool = False
    allow_final_response_token_overage: bool = False
    require_exact_agent_evidence: bool = False
    require_truncated_terminal_evidence: bool = False
    require_accepted_terminal_evidence: bool = False
    require_recalculation_integrity: bool = False
    allow_generic_response_truncation: bool = False
    allow_provider_infrastructure_no_score: bool = False


_V23_AUDIT_CONTRACT = _AuditProtocolContract(
    protocol_version=LEGACY_COMPARISON_PROTOCOL_VERSION,
    manifest_schema_version=LEGACY_COMPARISON_MANIFEST_SCHEMA_VERSION,
    configuration_policies=LEGACY_COMPARISON_CONFIGURATION_POLICIES,
    allowed_model_failure_reasons=frozenset(),
    require_v24_outcome_fields=False,
    strict_current_source=False,
)
_V24_AUDIT_CONTRACT = _AuditProtocolContract(
    protocol_version=V24_COMPARISON_PROTOCOL_VERSION,
    manifest_schema_version=V24_COMPARISON_MANIFEST_SCHEMA_VERSION,
    configuration_policies=V24_COMPARISON_CONFIGURATION_POLICIES,
    allowed_model_failure_reasons=frozenset(
        {"edit_recovery_exhausted", "workbook_unchanged"}
    ),
    require_v24_outcome_fields=True,
    strict_current_source=False,
)
_V25_AUDIT_CONTRACT = _AuditProtocolContract(
    protocol_version=V25_COMPARISON_PROTOCOL_VERSION,
    manifest_schema_version=V25_COMPARISON_MANIFEST_SCHEMA_VERSION,
    configuration_policies=V25_COMPARISON_CONFIGURATION_POLICIES,
    allowed_model_failure_reasons=frozenset(
        {
            "budget_exhausted",
            "edit_recovery_exhausted",
            "terminal_submission_invalid",
            "workbook_unchanged",
        }
    ),
    require_v24_outcome_fields=True,
    strict_current_source=False,
    allow_budget_exhaustion_evidence=True,
    allow_final_response_token_overage=True,
    require_exact_agent_evidence=True,
)
_V26_AUDIT_CONTRACT = _AuditProtocolContract(
    protocol_version=V26_COMPARISON_PROTOCOL_VERSION,
    manifest_schema_version=V26_COMPARISON_MANIFEST_SCHEMA_VERSION,
    configuration_policies=V26_COMPARISON_CONFIGURATION_POLICIES,
    allowed_model_failure_reasons=frozenset(
        {
            "budget_exhausted",
            "edit_recovery_exhausted",
            "terminal_submission_invalid",
            "terminal_submission_truncated",
            "workbook_unchanged",
        }
    ),
    require_v24_outcome_fields=True,
    strict_current_source=False,
    allow_budget_exhaustion_evidence=True,
    allow_final_response_token_overage=True,
    require_exact_agent_evidence=True,
    require_truncated_terminal_evidence=True,
    require_accepted_terminal_evidence=True,
)
_V27_AUDIT_CONTRACT = _AuditProtocolContract(
    protocol_version=V27_COMPARISON_PROTOCOL_VERSION,
    manifest_schema_version=V27_COMPARISON_MANIFEST_SCHEMA_VERSION,
    configuration_policies=V27_COMPARISON_CONFIGURATION_POLICIES,
    allowed_model_failure_reasons=_V26_AUDIT_CONTRACT.allowed_model_failure_reasons,
    require_v24_outcome_fields=True,
    strict_current_source=False,
    allow_budget_exhaustion_evidence=True,
    allow_final_response_token_overage=True,
    require_exact_agent_evidence=True,
    require_truncated_terminal_evidence=True,
    require_accepted_terminal_evidence=True,
)
_V28_AUDIT_CONTRACT = _AuditProtocolContract(
    protocol_version=V28_COMPARISON_PROTOCOL_VERSION,
    manifest_schema_version=V28_COMPARISON_MANIFEST_SCHEMA_VERSION,
    configuration_policies=V28_COMPARISON_CONFIGURATION_POLICIES,
    allowed_model_failure_reasons=_V27_AUDIT_CONTRACT.allowed_model_failure_reasons,
    require_v24_outcome_fields=True,
    strict_current_source=False,
    allow_budget_exhaustion_evidence=True,
    allow_final_response_token_overage=True,
    require_exact_agent_evidence=True,
    require_truncated_terminal_evidence=True,
    require_accepted_terminal_evidence=True,
    require_recalculation_integrity=True,
)
_V29_AUDIT_CONTRACT = _AuditProtocolContract(
    protocol_version=COMPARISON_PROTOCOL_VERSION,
    manifest_schema_version=COMPARISON_MANIFEST_SCHEMA_VERSION,
    configuration_policies=COMPARISON_CONFIGURATION_POLICIES,
    allowed_model_failure_reasons=(
        _V28_AUDIT_CONTRACT.allowed_model_failure_reasons
        | {"model_response_truncated"}
    ),
    require_v24_outcome_fields=True,
    strict_current_source=True,
    allow_budget_exhaustion_evidence=True,
    allow_final_response_token_overage=True,
    require_exact_agent_evidence=True,
    require_truncated_terminal_evidence=True,
    require_accepted_terminal_evidence=True,
    require_recalculation_integrity=True,
    allow_generic_response_truncation=True,
    allow_provider_infrastructure_no_score=True,
)


def _select_audit_contract(
    manifest: dict[str, Any],
    manifest_sha256: str | None,
    reasons: list[str],
) -> _AuditProtocolContract | None:
    identity = (
        manifest.get("comparison_protocol_version"),
        manifest.get("schema_version"),
    )
    if identity == (
        _V29_AUDIT_CONTRACT.protocol_version,
        _V29_AUDIT_CONTRACT.manifest_schema_version,
    ):
        return _V29_AUDIT_CONTRACT
    if identity == (
        _V28_AUDIT_CONTRACT.protocol_version,
        _V28_AUDIT_CONTRACT.manifest_schema_version,
    ):
        return _V28_AUDIT_CONTRACT
    if identity == (
        _V27_AUDIT_CONTRACT.protocol_version,
        _V27_AUDIT_CONTRACT.manifest_schema_version,
    ):
        return _V27_AUDIT_CONTRACT
    if identity == (
        _V26_AUDIT_CONTRACT.protocol_version,
        _V26_AUDIT_CONTRACT.manifest_schema_version,
    ):
        return _V26_AUDIT_CONTRACT
    if identity == (
        _V25_AUDIT_CONTRACT.protocol_version,
        _V25_AUDIT_CONTRACT.manifest_schema_version,
    ):
        return _V25_AUDIT_CONTRACT
    if identity == (
        _V24_AUDIT_CONTRACT.protocol_version,
        _V24_AUDIT_CONTRACT.manifest_schema_version,
    ):
        return _V24_AUDIT_CONTRACT
    if identity == (
        _V23_AUDIT_CONTRACT.protocol_version,
        _V23_AUDIT_CONTRACT.manifest_schema_version,
    ):
        if manifest_sha256 != LEGACY_PILOT_MANIFEST_SHA256:
            _add_reason(reasons, "legacy_comparison_manifest_sha256_mismatch")
        return _V23_AUDIT_CONTRACT
    _add_reason(reasons, "comparison_manifest_protocol_schema_unsupported")
    if manifest.get("schema_version") not in {
        _V23_AUDIT_CONTRACT.manifest_schema_version,
        _V24_AUDIT_CONTRACT.manifest_schema_version,
        _V25_AUDIT_CONTRACT.manifest_schema_version,
        _V26_AUDIT_CONTRACT.manifest_schema_version,
        _V27_AUDIT_CONTRACT.manifest_schema_version,
        _V28_AUDIT_CONTRACT.manifest_schema_version,
        _V29_AUDIT_CONTRACT.manifest_schema_version,
    }:
        _add_reason(reasons, "comparison_manifest_schema_mismatch")
    if manifest.get("comparison_protocol_version") not in {
        _V23_AUDIT_CONTRACT.protocol_version,
        _V24_AUDIT_CONTRACT.protocol_version,
        _V25_AUDIT_CONTRACT.protocol_version,
        _V26_AUDIT_CONTRACT.protocol_version,
        _V27_AUDIT_CONTRACT.protocol_version,
        _V28_AUDIT_CONTRACT.protocol_version,
        _V29_AUDIT_CONTRACT.protocol_version,
    }:
        _add_reason(reasons, "comparison_manifest_protocol_mismatch")
    return None


def _add_reason(reasons: list[str], reason: str) -> None:
    if reason not in reasons:
        reasons.append(reason)


def _non_negative_int(value: Any) -> int | None:
    if isinstance(value, bool) or not isinstance(value, int) or value < 0:
        return None
    return value


def _usage_triplet(value: Any) -> dict[str, int] | None:
    if not isinstance(value, dict):
        return None
    usage = {
        field: _non_negative_int(value.get(field))
        for field in ("input_tokens", "output_tokens", "total_tokens")
    }
    if any(tokens is None for tokens in usage.values()):
        return None
    normalized = {field: int(tokens) for field, tokens in usage.items()}
    if normalized["input_tokens"] + normalized["output_tokens"] != normalized[
        "total_tokens"
    ]:
        return None
    return normalized


def _v25_budget_exhaustion(
    row: dict[str, Any],
    contract: _AuditProtocolContract | None,
) -> tuple[bool, dict[str, Any] | None]:
    budget = row.get("budget")
    termination = budget.get("termination") if isinstance(budget, dict) else None
    limit = budget.get("limit") if isinstance(budget, dict) else None
    used = budget.get("used") if isinstance(budget, dict) else None
    reason = termination.get("reason") if isinstance(termination, dict) else None
    limit_field = (
        "model_calls"
        if reason == "max_model_calls"
        else "total_tokens"
        if reason == "max_total_tokens"
        else None
    )
    ceiling = limit.get(limit_field) if isinstance(limit, dict) and limit_field else None
    consumed = used.get(limit_field) if isinstance(used, dict) and limit_field else None
    exhausted = bool(
        isinstance(ceiling, int)
        and not isinstance(ceiling, bool)
        and isinstance(consumed, int)
        and not isinstance(consumed, bool)
        and consumed >= ceiling
    )
    valid = bool(
        contract is not None
        and contract.allow_budget_exhaustion_evidence
        and row.get("status") == "completed"
        and row.get("outcome_kind") == "model_execution_failure"
        and row.get("passed") is False
        and row.get("model_failure_reason") == "budget_exhausted"
        and row.get("error_category") == "model_execution_failure"
        and row.get("error_type") == "AgentExecutionFailure"
        and row.get("error_retryable") is False
        and isinstance(termination, dict)
        and reason in {"max_model_calls", "max_total_tokens"}
        and exhausted
        and isinstance(termination.get("stage"), str)
        and termination["stage"]
    )
    return valid, termination if isinstance(termination, dict) else None


_TERMINAL_RESPONSE_TIMING_FIELDS = frozenset(
    {
        "attempts",
        "elapsed_seconds",
        "first_event_seconds",
        "headers_seconds",
        "terminal_seconds",
        "terminal_event",
        "status_code",
        "sse_events",
        "logical_request_id",
        "client_request_id",
        "request_payload_sha256",
        "response_headers",
        "delivery_state",
        "pacing_wait_seconds_total",
        "attempt_history",
    }
)


def _valid_sha256(value: Any) -> bool:
    return bool(
        isinstance(value, str)
        and len(value) == 64
        and all(character in "0123456789abcdef" for character in value)
    )


def _valid_sheet_inventory_identity(value: Any) -> bool:
    if not isinstance(value, dict) or set(value) != {
        "schema_version",
        "workbook_sha256",
        "inventory_sha256",
        "sheets",
    }:
        return False
    sheets = value.get("sheets")
    if (
        value.get("schema_version") != 2
        or not _valid_sha256(value.get("workbook_sha256"))
        or not _valid_sha256(value.get("inventory_sha256"))
        or not isinstance(sheets, list)
        or not sheets
    ):
        return False
    normalized_names: set[str] = set()
    visible_sheet = False
    for index, sheet in enumerate(sheets):
        if (
            not isinstance(sheet, dict)
            or set(sheet) != {"index", "kind", "name", "visibility"}
            or sheet.get("index") != index
            or sheet.get("kind")
            not in {"worksheet", "chartsheet", "dialogsheet", "macrosheet", "intlMacrosheet"}
            or not isinstance(sheet.get("name"), str)
            or not sheet["name"]
            or sheet.get("visibility") not in {"visible", "hidden", "veryHidden"}
        ):
            return False
        normalized_name = sheet["name"].casefold()
        if normalized_name in normalized_names:
            return False
        normalized_names.add(normalized_name)
        visible_sheet = visible_sheet or sheet["visibility"] == "visible"
    if not visible_sheet:
        return False
    encoded = json.dumps(
        sheets,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")
    return value["inventory_sha256"] == hashlib.sha256(encoded).hexdigest()


def _v26_output_limit_attempt_observed(stages: list[dict[str, Any]]) -> bool:
    for stage in stages:
        stage_agent = stage.get("agent")
        timings = (
            stage_agent.get("request_timings")
            if isinstance(stage_agent, dict)
            else None
        )
        if not isinstance(timings, list):
            continue
        for timing in timings:
            history = timing.get("attempt_history") if isinstance(timing, dict) else None
            if isinstance(history, list) and any(
                isinstance(attempt, dict)
                and attempt.get("error_type") == "ProviderOutputLimitError"
                for attempt in history
            ):
                return True
    return False


def _v26_truncated_terminal_evidence_valid(
    row: dict[str, Any],
    agent: dict[str, Any],
    stages: list[dict[str, Any]],
    *,
    budget_exhaustion: bool,
) -> bool:
    """Bind a known-false truncation to its committed, unexecuted final response."""

    budget = row.get("budget")
    termination = budget.get("termination") if isinstance(budget, dict) else None
    budget_precedence = bool(
        row.get("model_failure_reason") == "budget_exhausted"
        and budget_exhaustion
        and isinstance(termination, dict)
        and termination.get("reason") == "max_total_tokens"
    )
    if (
        row.get("model_failure_reason") != "terminal_submission_truncated"
        and not budget_precedence
    ):
        return False
    if row.get("api_protocol") != "chat-completions" or row.get("provider_error"):
        return False
    if not stages or any("terminal_response" in stage for stage in stages):
        return False

    final_stage = stages[-1]
    final_agent = final_stage.get("agent")
    if not isinstance(final_agent, dict):
        return False
    if any(
        stage.get("observed_terminal_tool")
        == TERMINAL_SUBMISSION_TRUNCATED_OBSERVED
        or (
            isinstance(stage.get("agent"), dict)
            and (
                stage["agent"].get("observed_terminal_tool")
                == TERMINAL_SUBMISSION_TRUNCATED_OBSERVED
                or (
                    isinstance(stage["agent"].get("terminal_response"), dict)
                    and stage["agent"]["terminal_response"].get("status")
                    == "truncated"
                )
            )
        )
        for stage in stages[:-1]
    ):
        return False

    terminal_response = final_agent.get("terminal_response")
    if not isinstance(terminal_response, dict) or set(terminal_response) != {
        "status",
        "finish_reason",
        "response_id",
        "usage",
        "timing",
        "discarded_message",
    }:
        return False
    response_id = terminal_response.get("response_id")
    if response_id is not None and (
        not isinstance(response_id, str) or not response_id
    ):
        return False
    if (
        terminal_response.get("status") != "truncated"
        or terminal_response.get("finish_reason") != "length"
        or terminal_response != agent.get("terminal_response")
        or response_id != final_agent.get("response_id")
        or response_id != agent.get("response_id")
    ):
        return False

    expected_observed_terminal = (
        "budget_exhausted"
        if budget_precedence
        else TERMINAL_SUBMISSION_TRUNCATED_OBSERVED
    )
    if (
        final_stage.get("terminal_tool") != "submit_result"
        or final_stage.get("observed_terminal_tool")
        != expected_observed_terminal
        or final_agent.get("terminal_tool") != "submit_result"
        or final_agent.get("observed_terminal_tool")
        != expected_observed_terminal
        or agent.get("terminal_tool") != "submit_result"
        or agent.get("observed_terminal_tool")
        != expected_observed_terminal
        or final_stage.get("post_prefix_tool_choice") != "auto"
        or final_agent.get("post_prefix_tool_choice") != "auto"
        or agent.get("post_prefix_tool_choice") != "auto"
        or final_agent.get("terminal_submissions") != 0
        or final_agent.get("function_calls_total") != final_agent.get("tool_calls")
    ):
        return False
    final_trace = final_agent.get("tool_trace")
    if not isinstance(final_trace, list) or any(
        isinstance(item, dict) and item.get("name") == "submit_result"
        for item in final_trace
    ):
        return False
    final_text = final_agent.get("final_text")
    if (
        not isinstance(final_text, str)
        or not final_text
        or agent.get("final_text") != final_text
    ):
        return False

    forced_prefix = final_stage.get("forced_tool_prefix")
    observed_prefix = final_stage.get("observed_forced_tool_prefix")
    turns = _non_negative_int(final_agent.get("turns"))
    max_turns = _non_negative_int(final_stage.get("max_turns"))
    final_budget = final_agent.get("budget")
    final_limit = final_budget.get("limit") if isinstance(final_budget, dict) else None
    final_used = final_budget.get("used") if isinstance(final_budget, dict) else None
    final_turn_basis = bool(turns is not None and max_turns is not None and turns == max_turns)
    last_call_basis = bool(
        isinstance(final_limit, dict)
        and isinstance(final_used, dict)
        and _non_negative_int(final_limit.get("model_calls")) is not None
        and final_used.get("model_calls") == final_limit.get("model_calls")
    )
    if (
        not isinstance(forced_prefix, list)
        or observed_prefix != forced_prefix
        or final_agent.get("forced_tool_prefix") != forced_prefix
        or final_agent.get("observed_forced_tool_prefix") != forced_prefix
        or turns is None
        or turns <= len(forced_prefix)
        or not (final_turn_basis or last_call_basis)
    ):
        return False

    stage_timings = final_agent.get("request_timings")
    aggregate_timings = agent.get("request_timings")
    if (
        not isinstance(stage_timings, list)
        or not stage_timings
        or not isinstance(aggregate_timings, list)
        or not aggregate_timings
    ):
        return False
    final_timing = stage_timings[-1]
    if not isinstance(final_timing, dict):
        return False
    if (
        final_timing.get("turn") != turns
        or final_timing.get("stage") != final_stage.get("name")
        or aggregate_timings[-1] != final_timing
    ):
        return False
    response_timing = terminal_response.get("timing")
    if (
        not isinstance(response_timing, dict)
        or set(response_timing) != _TERMINAL_RESPONSE_TIMING_FIELDS
        or response_timing
        != {field: final_timing.get(field) for field in _TERMINAL_RESPONSE_TIMING_FIELDS}
        or response_timing.get("terminal_event") != "chat.completion"
        or response_timing.get("status_code") != 200
        or response_timing.get("delivery_state") != "terminal_seen"
    ):
        return False
    attempts = _non_negative_int(response_timing.get("attempts"))
    attempt_history = response_timing.get("attempt_history")
    if (
        attempts is None
        or attempts < 1
        or not isinstance(attempt_history, list)
        or len(attempt_history) != attempts
        or not isinstance(attempt_history[-1], dict)
    ):
        return False
    final_attempt = attempt_history[-1]
    if any(
        final_attempt.get(field) != expected
        for field, expected in {
            "attempt": attempts,
            "outcome": "error",
            "error_type": "ProviderOutputLimitError",
            "phase": "response_body",
            "status_code": 200,
            "terminal_event": "chat.completion",
            "retryable": False,
            "safe_to_retry": False,
            "automatic_retry_scheduled": False,
            "automatic_retry_suppressed_reason": "delivery_not_known_safe",
            "delivery_state": "terminal_seen",
            "api_protocol": "chat-completions",
            "endpoint": "/chat/completions",
        }.items()
    ):
        return False

    response_usage = terminal_response.get("usage")
    timing_usage = {
        field: final_timing.get(field)
        for field in ("input_tokens", "output_tokens", "total_tokens")
    }
    if (
        not isinstance(response_usage, dict)
        or set(response_usage) != set(timing_usage)
        or _usage_triplet(response_usage) is None
        or response_usage != timing_usage
    ):
        return False

    discarded = terminal_response.get("discarded_message")
    if not isinstance(discarded, dict) or set(discarded) != {
        "sha256",
        "serialized_chars",
        "serialized_bytes",
        "top_level_field_count",
        "content_item_count",
        "tool_call_count",
    }:
        return False
    counts = {
        field: _non_negative_int(discarded.get(field))
        for field in (
            "serialized_chars",
            "serialized_bytes",
            "top_level_field_count",
            "content_item_count",
            "tool_call_count",
        )
    }
    discarded_valid = bool(
        _valid_sha256(discarded.get("sha256"))
        and all(value is not None for value in counts.values())
        and int(counts["serialized_chars"] or 0) > 0
        and int(counts["serialized_bytes"] or 0)
        >= int(counts["serialized_chars"] or 0)
        and int(counts["top_level_field_count"] or 0) > 0
    )
    if not discarded_valid:
        return False
    if budget_precedence:
        limit = budget.get("limit") if isinstance(budget, dict) else None
        used = budget.get("used") if isinstance(budget, dict) else None
        ceiling = limit.get("total_tokens") if isinstance(limit, dict) else None
        consumed = used.get("total_tokens") if isinstance(used, dict) else None
        aggregate_token_timings = [
            timing.get("total_tokens")
            for timing in aggregate_timings
            if isinstance(timing, dict)
        ]
        if not (
            _non_negative_int(ceiling) is not None
            and _non_negative_int(consumed) is not None
            and len(aggregate_token_timings) == len(aggregate_timings)
            and all(_non_negative_int(value) is not None for value in aggregate_token_timings)
            and sum(aggregate_token_timings) == consumed
            and sum(aggregate_token_timings[:-1]) < ceiling
            and consumed > ceiling
            and aggregate_token_timings[-1] == response_usage["total_tokens"]
        ):
            return False
    return True


def _discarded_output_metadata_valid(value: Any) -> bool:
    if not isinstance(value, dict) or set(value) != {
        "sha256",
        "serialized_chars",
        "serialized_bytes",
        "top_level_field_count",
        "content_item_count",
        "tool_call_count",
    }:
        return False
    counts = {
        field: _non_negative_int(value.get(field))
        for field in (
            "serialized_chars",
            "serialized_bytes",
            "top_level_field_count",
            "content_item_count",
            "tool_call_count",
        )
    }
    return bool(
        _valid_sha256(value.get("sha256"))
        and all(item is not None for item in counts.values())
        and int(counts["serialized_chars"] or 0) > 0
        and int(counts["serialized_bytes"] or 0)
        >= int(counts["serialized_chars"] or 0)
        and int(counts["top_level_field_count"] or 0) > 0
    )


def _v29_generic_output_limit_evidence_valid(
    row: dict[str, Any],
    agent: dict[str, Any],
    stages: list[dict[str, Any]],
    *,
    budget_exhaustion: bool,
) -> bool:
    """Bind a delivered output limit to a committed, unexecuted partial response."""

    budget = row.get("budget")
    termination = budget.get("termination") if isinstance(budget, dict) else None
    budget_precedence = bool(
        row.get("model_failure_reason") == "budget_exhausted"
        and budget_exhaustion
        and isinstance(termination, dict)
        and termination.get("reason") == "max_total_tokens"
    )
    terminal_submission = (
        row.get("model_failure_reason") == "terminal_submission_truncated"
    )
    if (
        row.get("model_failure_reason") != "model_response_truncated"
        and not terminal_submission
        and not budget_precedence
    ):
        return False
    api_protocol = row.get("api_protocol")
    if api_protocol not in {"chat-completions", "responses"} or row.get(
        "provider_error"
    ):
        return False
    if not stages or any("terminal_response" in stage for stage in stages):
        return False

    final_stage = stages[-1]
    final_agent = final_stage.get("agent")
    if not isinstance(final_agent, dict):
        return False
    if any(
        stage.get("observed_terminal_tool")
        in {
            MODEL_RESPONSE_TRUNCATED_TERMINAL,
            TERMINAL_SUBMISSION_TRUNCATED_OBSERVED,
        }
        or (
            isinstance(stage.get("agent"), dict)
            and isinstance(stage["agent"].get("terminal_response"), dict)
            and stage["agent"]["terminal_response"].get("status") == "truncated"
        )
        for stage in stages[:-1]
    ):
        return False

    terminal_response = final_agent.get("terminal_response")
    if not isinstance(terminal_response, dict) or set(terminal_response) != {
        "status",
        "finish_reason",
        "response_id",
        "usage",
        "timing",
        "discarded_message",
    }:
        return False
    response_id = terminal_response.get("response_id")
    if not isinstance(response_id, str) or not response_id:
        return False
    if (
        terminal_response.get("status") != "truncated"
        or terminal_response.get("finish_reason") != "length"
        or terminal_response != agent.get("terminal_response")
        or response_id != final_agent.get("response_id")
        or response_id != agent.get("response_id")
    ):
        return False

    expected_observed_terminal = (
        "budget_exhausted"
        if budget_precedence
        else TERMINAL_SUBMISSION_TRUNCATED_OBSERVED
        if terminal_submission
        else MODEL_RESPONSE_TRUNCATED_TERMINAL
    )
    terminal_tool = final_stage.get("terminal_tool")
    if (
        (terminal_submission and terminal_tool != "submit_result")
        or
        final_stage.get("observed_terminal_tool")
        != expected_observed_terminal
        or final_agent.get("observed_terminal_tool")
        != expected_observed_terminal
        or agent.get("observed_terminal_tool") != expected_observed_terminal
        or final_agent.get("terminal_tool") != terminal_tool
        or agent.get("terminal_tool") != terminal_tool
        or final_stage.get("post_prefix_tool_choice") != "auto"
        or final_agent.get("post_prefix_tool_choice") != "auto"
        or agent.get("post_prefix_tool_choice") != "auto"
        or final_agent.get("terminal_submissions") != 0
        or final_agent.get("function_calls_total") != final_agent.get("tool_calls")
    ):
        return False
    final_trace = final_agent.get("tool_trace")
    if not isinstance(final_trace, list) or any(
        isinstance(item, dict) and item.get("name") == "submit_result"
        for item in final_trace
    ):
        return False
    final_text = final_agent.get("final_text")
    if (
        not isinstance(final_text, str)
        or not final_text
        or agent.get("final_text") != final_text
        or (
            not budget_precedence
            and not terminal_submission
            and final_text
            != "Model response was truncated by the provider output limit."
        )
    ):
        return False

    turns = _non_negative_int(final_agent.get("turns"))
    max_turns = _non_negative_int(final_stage.get("max_turns"))
    if turns is None or turns < 1 or max_turns is None or turns > max_turns:
        return False
    if terminal_submission:
        forced_prefix = final_stage.get("forced_tool_prefix")
        observed_prefix = final_stage.get("observed_forced_tool_prefix")
        final_budget = final_agent.get("budget")
        final_limit = (
            final_budget.get("limit") if isinstance(final_budget, dict) else None
        )
        final_used = (
            final_budget.get("used") if isinstance(final_budget, dict) else None
        )
        final_turn_basis = turns == max_turns
        last_call_basis = bool(
            isinstance(final_limit, dict)
            and isinstance(final_used, dict)
            and _non_negative_int(final_limit.get("model_calls")) is not None
            and final_used.get("model_calls") == final_limit.get("model_calls")
        )
        if (
            not isinstance(forced_prefix, list)
            or observed_prefix != forced_prefix
            or final_agent.get("forced_tool_prefix") != forced_prefix
            or final_agent.get("observed_forced_tool_prefix") != forced_prefix
            or not (final_turn_basis or last_call_basis)
        ):
            return False
    stage_timings = final_agent.get("request_timings")
    aggregate_timings = agent.get("request_timings")
    if (
        not isinstance(stage_timings, list)
        or not stage_timings
        or not isinstance(aggregate_timings, list)
        or not aggregate_timings
    ):
        return False
    final_timing = stage_timings[-1]
    if not isinstance(final_timing, dict) or (
        final_timing.get("turn") != turns
        or final_timing.get("stage") != final_stage.get("name")
        or aggregate_timings[-1] != final_timing
    ):
        return False

    response_timing = terminal_response.get("timing")
    expected_terminal_event = (
        "chat.completion"
        if api_protocol == "chat-completions"
        else "response.incomplete"
    )
    expected_phase = (
        "response_body"
        if api_protocol == "chat-completions"
        else "response_stream"
    )
    expected_endpoint = (
        "/chat/completions"
        if api_protocol == "chat-completions"
        else "/responses"
    )
    if (
        not isinstance(response_timing, dict)
        or set(response_timing) != _TERMINAL_RESPONSE_TIMING_FIELDS
        or response_timing
        != {
            field: final_timing.get(field)
            for field in _TERMINAL_RESPONSE_TIMING_FIELDS
        }
        or response_timing.get("terminal_event") != expected_terminal_event
        or response_timing.get("status_code") != 200
        or response_timing.get("delivery_state") != "terminal_seen"
    ):
        return False
    attempts = _non_negative_int(response_timing.get("attempts"))
    attempt_history = response_timing.get("attempt_history")
    if (
        attempts is None
        or attempts < 1
        or not isinstance(attempt_history, list)
        or len(attempt_history) != attempts
        or not isinstance(attempt_history[-1], dict)
    ):
        return False
    final_attempt = attempt_history[-1]
    if any(
        final_attempt.get(field) != expected
        for field, expected in {
            "attempt": attempts,
            "outcome": "error",
            "error_type": "ProviderOutputLimitError",
            "phase": expected_phase,
            "status_code": 200,
            "terminal_event": expected_terminal_event,
            "retryable": False,
            "safe_to_retry": False,
            "safe_retry_reason": None,
            "automatic_retry_scheduled": False,
            "automatic_retry_suppressed_reason": "delivery_not_known_safe",
            "delivery_state": "terminal_seen",
            "api_protocol": api_protocol,
            "endpoint": expected_endpoint,
        }.items()
    ):
        return False

    response_usage = terminal_response.get("usage")
    timing_usage = {
        field: final_timing.get(field)
        for field in ("input_tokens", "output_tokens", "total_tokens")
    }
    if (
        not isinstance(response_usage, dict)
        or set(response_usage) != set(timing_usage)
        or _usage_triplet(response_usage) is None
        or response_usage != timing_usage
        or not _discarded_output_metadata_valid(
            terminal_response.get("discarded_message")
        )
    ):
        return False

    if budget_precedence:
        limit = budget.get("limit") if isinstance(budget, dict) else None
        used = budget.get("used") if isinstance(budget, dict) else None
        ceiling = limit.get("total_tokens") if isinstance(limit, dict) else None
        consumed = used.get("total_tokens") if isinstance(used, dict) else None
        aggregate_token_timings = [
            timing.get("total_tokens")
            for timing in aggregate_timings
            if isinstance(timing, dict)
        ]
        if not (
            _non_negative_int(ceiling) is not None
            and _non_negative_int(consumed) is not None
            and len(aggregate_token_timings) == len(aggregate_timings)
            and all(
                _non_negative_int(value) is not None
                for value in aggregate_token_timings
            )
            and sum(aggregate_token_timings) == consumed
            and sum(aggregate_token_timings[:-1]) < ceiling
            and consumed > ceiling
            and aggregate_token_timings[-1] == response_usage["total_tokens"]
        ):
            return False
    return True


_V26_PAPER_EVIDENCE_STAGES = frozenset(
    {"extract", "vision_verify", "latex_verify"}
)


def _v26_stage_terminal_response_mode(arm: str, stage: str) -> str | None:
    if arm == "paper":
        if stage in _V26_PAPER_EVIDENCE_STAGES:
            return "evidence_result"
        if stage == "reconcile":
            return "assistant_text"
        if stage == "solve":
            return "empty_ack"
        return None
    return "empty_ack" if stage == "solve" else None


def _v26_accepted_terminal_response_valid(
    terminal_response: Any,
    result: dict[str, Any],
    *,
    acknowledgement_mode: str,
) -> bool:
    if not isinstance(terminal_response, dict) or set(terminal_response) != {
        "status",
        "response_id",
        "acknowledgement",
    }:
        return False
    response_id = terminal_response.get("response_id")
    if response_id is not None and (
        not isinstance(response_id, str) or not response_id
    ):
        return False
    final_text = result.get("final_text")
    acknowledgement = terminal_response.get("acknowledgement")
    if (
        terminal_response.get("status") != "accepted"
        or response_id != result.get("response_id")
        or not isinstance(final_text, str)
        or not final_text
        or not isinstance(acknowledgement, dict)
    ):
        return False
    if acknowledgement_mode == "empty_ack":
        return acknowledgement == {} and final_text == "Spreadsheet task completed."
    if acknowledgement_mode != "evidence_result":
        return False
    return bool(
        set(acknowledgement) == {"mode", "result_chars", "result_sha256"}
        and acknowledgement.get("mode") == "evidence_result"
        and _non_negative_int(acknowledgement.get("result_chars")) is not None
        and acknowledgement.get("result_chars") == len(final_text)
        and acknowledgement.get("result_sha256")
        == hashlib.sha256(final_text.encode("utf-8")).hexdigest()
    )


def _v26_completed_stage_terminal_response_valid(
    stage: dict[str, Any],
    *,
    arm: str,
) -> bool:
    stage_agent = stage.get("agent")
    name = stage.get("name")
    if not isinstance(stage_agent, dict) or not isinstance(name, str):
        return False
    acknowledgement_mode = _v26_stage_terminal_response_mode(arm, name)
    if acknowledgement_mode is None:
        return False
    terminal_response = stage_agent.get("terminal_response")
    if acknowledgement_mode == "assistant_text":
        return bool(
            terminal_response is None
            and stage.get("terminal_tool") == "assistant_text"
            and stage.get("observed_terminal_tool") == "assistant_text"
            and stage_agent.get("terminal_tool") == "assistant_text"
            and stage_agent.get("observed_terminal_tool") == "assistant_text"
            and stage_agent.get("terminal_submissions") == 0
        )
    return bool(
        stage.get("terminal_tool") == "submit_result"
        and stage.get("observed_terminal_tool") == "submit_result"
        and stage_agent.get("terminal_tool") == "submit_result"
        and stage_agent.get("observed_terminal_tool") == "submit_result"
        and stage_agent.get("terminal_submissions") == 1
        and _v26_accepted_terminal_response_valid(
            terminal_response,
            stage_agent,
            acknowledgement_mode=acknowledgement_mode,
        )
    )


def _v26_model_failure_terminal_response_valid(
    agent: dict[str, Any],
    stages: list[dict[str, Any]],
    *,
    arm: str,
) -> bool:
    """Validate completed prior stages and reject a successful final acknowledgement."""

    if not stages or any("terminal_response" in stage for stage in stages):
        return False
    final_agent = stages[-1].get("agent")
    if not isinstance(final_agent, dict):
        return False
    aggregate_response = agent.get("terminal_response")
    final_response = final_agent.get("terminal_response")
    if not (
        (aggregate_response is None and final_response is None)
        or (
            isinstance(aggregate_response, dict)
            and isinstance(final_response, dict)
            and aggregate_response.get("status") == "truncated"
            and final_response.get("status") == "truncated"
        )
    ):
        return False
    return all(
        _v26_completed_stage_terminal_response_valid(stage, arm=arm)
        for stage in stages[:-1]
    )


def _v26_success_terminal_evidence_valid(
    row: dict[str, Any],
    agent: dict[str, Any],
    stages: list[dict[str, Any]],
    *,
    arm: str,
) -> bool:
    if (
        row.get("status") != "completed"
        or row.get("outcome_kind") != "scored"
        or not stages
        or any("terminal_response" in stage for stage in stages)
        or not all(
            _v26_completed_stage_terminal_response_valid(stage, arm=arm)
            for stage in stages
        )
    ):
        return False

    final_stage = stages[-1]
    final_agent = final_stage.get("agent")
    if not isinstance(final_agent, dict):
        return False
    terminal_response = final_agent.get("terminal_response")
    final_name = final_stage.get("name")
    final_mode = (
        _v26_stage_terminal_response_mode(arm, final_name)
        if isinstance(final_name, str)
        else None
    )
    return bool(
        final_mode in {"empty_ack", "evidence_result"}
        and _v26_accepted_terminal_response_valid(
            terminal_response,
            final_agent,
            acknowledgement_mode=str(final_mode),
        )
        and agent.get("terminal_response") == terminal_response
        and agent.get("final_text") == final_agent.get("final_text")
        and agent.get("response_id") == final_agent.get("response_id")
        and agent.get("terminal_tool") == "submit_result"
        and agent.get("observed_terminal_tool") == "submit_result"
        and _non_negative_int(agent.get("terminal_submissions")) is not None
        and agent.get("terminal_submissions") >= 1
    )


def _agent_tool_recalculation_failure_stage(row: dict[str, Any]) -> str | None:
    stage = row.get("agent_failure_stage")
    if (
        row.get("outcome_kind") == "infrastructure_failure"
        and row.get("error_category") == "recalculation_infrastructure"
        and row.get("infrastructure_failure_stage")
        == AGENT_TOOL_RECALCULATION_FAILURE_STAGE
        and row.get("infrastructure_failure_tool")
        == RECALCULATION_VALIDATION_TOOL
        and isinstance(stage, str)
        and stage
    ):
        return stage
    return None


def _agent_tool_recalculation_terminal_evidence_valid(
    row: dict[str, Any],
    agent: dict[str, Any],
    stages: list[dict[str, Any]],
    *,
    arm: str,
) -> bool:
    """Validate a tool-stage interruption without inventing a terminal call."""

    failure_stage = _agent_tool_recalculation_failure_stage(row)
    if failure_stage is None or not stages:
        return False
    final_stage = stages[-1]
    final_agent = final_stage.get("agent")
    wrapper_trace = final_stage.get("tool_trace")
    nested_trace = final_agent.get("tool_trace") if isinstance(final_agent, dict) else None
    aggregate_trace = agent.get("tool_trace")
    if not (
        final_stage.get("name") == failure_stage
        and isinstance(final_agent, dict)
        and isinstance(wrapper_trace, list)
        and wrapper_trace
        and nested_trace == wrapper_trace
        and isinstance(aggregate_trace, list)
        and aggregate_trace
    ):
        return False
    failed_call = wrapper_trace[-1]
    aggregate_failed_call = aggregate_trace[-1]
    expected_failure = {
        "name": RECALCULATION_VALIDATION_TOOL,
        "ok": False,
        "error_type": "RecalculationIntegrityError",
        "failure_category": "recalculation_infrastructure",
    }
    if not (
        isinstance(failed_call, dict)
        and all(failed_call.get(key) == value for key, value in expected_failure.items())
        and isinstance(aggregate_failed_call, dict)
        and aggregate_failed_call.get("stage") == failure_stage
        and all(
            aggregate_failed_call.get(key) == value
            for key, value in expected_failure.items()
        )
    ):
        return False
    return bool(
        all(
            _v26_completed_stage_terminal_response_valid(stage, arm=arm)
            for stage in stages[:-1]
        )
        and final_stage.get("terminal_tool") == "submit_result"
        and final_stage.get("observed_terminal_tool") is None
        and final_agent.get("terminal_tool") == "submit_result"
        and final_agent.get("observed_terminal_tool") is None
        and final_agent.get("terminal_submissions") == 0
        and final_agent.get("terminal_response") is None
        and agent.get("terminal_tool") == "submit_result"
        and agent.get("observed_terminal_tool") is None
        and agent.get("terminal_submissions") == 0
        and agent.get("terminal_response") is None
        and agent.get("final_text") == final_agent.get("final_text")
        and agent.get("response_id") == final_agent.get("response_id")
    )


def _text_sha256(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def _file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    flags = os.O_RDONLY | getattr(os, "O_CLOEXEC", 0) | getattr(os, "O_NOFOLLOW", 0)
    descriptor = os.open(path, flags)
    try:
        metadata = os.fstat(descriptor)
        if not stat.S_ISREG(metadata.st_mode):
            raise OSError(f"not a regular file: {path}")
        while chunk := os.read(descriptor, 1024 * 1024):
            digest.update(chunk)
    finally:
        os.close(descriptor)
    return digest.hexdigest()


def _regular_file_bytes_for_audit(path: Path) -> bytes:
    if path.is_symlink():
        raise OSError(f"symlink is not allowed: {path}")
    flags = os.O_RDONLY | getattr(os, "O_CLOEXEC", 0) | getattr(os, "O_NOFOLLOW", 0)
    descriptor = os.open(path, flags)
    try:
        metadata = os.fstat(descriptor)
        if not stat.S_ISREG(metadata.st_mode):
            raise OSError(f"not a regular file: {path}")
        chunks: list[bytes] = []
        while chunk := os.read(descriptor, 1024 * 1024):
            chunks.append(chunk)
        return b"".join(chunks)
    finally:
        os.close(descriptor)


def _reject_duplicate_json_keys_for_audit(
    pairs: list[tuple[str, Any]],
) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for key, value in pairs:
        if key in result:
            raise ValueError(f"duplicate JSON key: {key}")
        result[key] = value
    return result


def _load_interrupted_seals(
    path: Path,
    *,
    manifest: dict[str, Any],
    manifest_sha256: str | None,
    expected_keys: set[tuple[str, str]],
    reasons: list[str],
    protocol_version: str | None,
) -> tuple[dict[tuple[str, str], dict[str, Any]], str | None]:
    """Load exact non-replay seals without treating unknown outcomes as failures."""

    if not path.exists():
        return {}, None
    seals_sha256: str | None = None
    try:
        raw = _regular_file_bytes_for_audit(path)
        seals_sha256 = hashlib.sha256(raw).hexdigest()
        document = json.loads(
            raw.decode("utf-8"),
            object_pairs_hook=_reject_duplicate_json_keys_for_audit,
        )
        if not isinstance(document, dict) or set(document) != {
            "schema_version",
            "seals",
        }:
            raise ValueError("invalid interrupted seals document fields")
        seals = document.get("seals")
        if document.get("schema_version") != 1 or not isinstance(seals, list):
            raise ValueError("invalid interrupted seals document")
        required_fields = {
            "schema_version",
            "task_id",
            "arm",
            "comparison_protocol_version",
            "comparison_manifest_sha256",
            "split_provenance",
            "run_spec_provenance",
            "status",
            "passed",
            "outcome_observed",
            "score_available",
            "usage_observed",
            "replay_permitted",
            "error_retryable",
            "error_category",
            "sealed_at",
            "sealed_from_inflight_marker_sha256",
        }
        by_key: dict[tuple[str, str], dict[str, Any]] = {}
        for seal in seals:
            if not isinstance(seal, dict) or set(seal) != required_fields:
                raise ValueError("invalid interrupted seal fields")
            task_id = seal.get("task_id")
            arm = seal.get("arm")
            key = (task_id, arm)
            marker_sha256 = seal.get("sealed_from_inflight_marker_sha256")
            sealed_at = seal.get("sealed_at")
            try:
                parsed_sealed_at = datetime.fromisoformat(sealed_at)
            except (TypeError, ValueError):
                raise ValueError("invalid interrupted seal timestamp") from None
            if (
                not isinstance(task_id, str)
                or not isinstance(arm, str)
                or key not in expected_keys
                or key in by_key
                or seal.get("schema_version") != 1
                or seal.get("comparison_protocol_version")
                != protocol_version
                or seal.get("comparison_manifest_sha256") != manifest_sha256
                or seal.get("split_provenance") != manifest.get("split_provenance")
                or seal.get("run_spec_provenance")
                != manifest.get("run_spec_provenance")
                or seal.get("status") != "interrupted"
                or seal.get("passed") is not None
                or seal.get("outcome_observed") is not False
                or seal.get("score_available") is not False
                or seal.get("usage_observed") is not False
                or seal.get("replay_permitted") is not False
                or seal.get("error_retryable") is not False
                or seal.get("error_category") != "interrupted_unknown_outcome"
                or parsed_sealed_at.tzinfo is None
                or parsed_sealed_at.utcoffset() is None
                or not isinstance(marker_sha256, str)
                or len(marker_sha256) != 64
                or any(
                    character not in "0123456789abcdef"
                    for character in marker_sha256
                )
            ):
                raise ValueError("invalid interrupted seal")
            by_key[key] = seal
        return by_key, seals_sha256
    except (OSError, UnicodeError, json.JSONDecodeError, ValueError):
        _add_reason(reasons, "interrupted_arm_task_seals_invalid")
        return {}, seals_sha256


def _scoring_metadata_sha256(task: SpreadsheetTask) -> str:
    encoded = json.dumps(
        {
            "answer_position": task.answer_position,
            "answer_sheet": task.answer_sheet,
        },
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )
    return _text_sha256(encoded)


def _absolute_path(value: Any, *, base: Path) -> Path | None:
    if not isinstance(value, str) or not value.strip():
        return None
    candidate = Path(value).expanduser()
    if not candidate.is_absolute():
        candidate = base / candidate
    return Path(os.path.abspath(candidate))


def _has_symlink(root: Path, target: Path) -> bool:
    """Check every result-owned component without following the target first."""

    try:
        relative = target.relative_to(root)
    except ValueError:
        return False
    current = root
    if current.is_symlink():
        return True
    for part in relative.parts:
        current = current / part
        if current.is_symlink():
            return True
    return False


def _load_manifest(path: Path, reasons: list[str]) -> dict[str, Any]:
    try:
        raw = _regular_file_bytes_for_audit(path)
        value = json.loads(
            raw.decode("utf-8"),
            object_pairs_hook=_reject_duplicate_json_keys_for_audit,
        )
    except FileNotFoundError:
        _add_reason(reasons, "comparison_manifest_missing")
        return {}
    except (OSError, UnicodeError, json.JSONDecodeError, ValueError):
        _add_reason(reasons, "comparison_manifest_invalid")
        return {}
    if not isinstance(value, dict):
        _add_reason(reasons, "comparison_manifest_invalid")
        return {}
    return value


def _load_result_rows(path: Path, reasons: list[str]) -> list[dict[str, Any]]:
    try:
        raw = _regular_file_bytes_for_audit(path)
    except FileNotFoundError:
        _add_reason(reasons, "results_file_missing")
        return []
    except (OSError, UnicodeError):
        _add_reason(reasons, "results_file_unreadable")
        return []

    if raw and not raw.endswith(b"\n"):
        _add_reason(reasons, "results_file_non_terminated")
    try:
        lines = raw.decode("utf-8").splitlines()
    except UnicodeError:
        _add_reason(reasons, "results_file_unreadable")
        return []

    rows: list[dict[str, Any]] = []
    for line_number, line in enumerate(lines, start=1):
        if not line.strip():
            continue
        try:
            row = json.loads(
                line,
                object_pairs_hook=_reject_duplicate_json_keys_for_audit,
            )
        except (json.JSONDecodeError, ValueError):
            _add_reason(reasons, f"invalid_jsonl_line:{line_number}")
            continue
        if not isinstance(row, dict):
            _add_reason(reasons, f"non_object_jsonl_line:{line_number}")
            continue
        rows.append(row)
    return rows


def _manifest_task_reasons(
    manifest: dict[str, Any], tasks: list[SpreadsheetTask], reasons: list[str]
) -> dict[str, list[str]]:
    task_reasons = {task.task_id: [] for task in tasks}
    expected_ids = [task.task_id for task in tasks]
    if manifest.get("task_count") != len(tasks):
        _add_reason(reasons, "manifest_task_count_mismatch")
    if manifest.get("task_ids") != expected_ids:
        _add_reason(reasons, "manifest_task_ids_mismatch")

    raw_entries = manifest.get("tasks")
    if not isinstance(raw_entries, list):
        _add_reason(reasons, "manifest_tasks_invalid")
        raw_entries = []
    entries: dict[str, list[dict[str, Any]]] = {}
    for raw_entry in raw_entries:
        if not isinstance(raw_entry, dict) or raw_entry.get("task_id") is None:
            _add_reason(reasons, "manifest_task_entry_invalid")
            continue
        entries.setdefault(str(raw_entry["task_id"]), []).append(raw_entry)

    expected_id_set = set(expected_ids)
    for task_id in sorted(set(entries) - expected_id_set):
        _add_reason(reasons, f"manifest_unknown_task:{task_id}")

    for task in tasks:
        matching = entries.get(task.task_id, [])
        if len(matching) != 1:
            reason = "manifest_task_missing" if not matching else "manifest_task_duplicate"
            task_reasons[task.task_id].append(reason)
            _add_reason(reasons, f"{reason}:{task.task_id}")
            continue
        entry = matching[0]
        expected_hashes: dict[str, str | None] = {
            "instruction_sha256": _text_sha256(task.instruction),
            "scoring_metadata_sha256": _scoring_metadata_sha256(task),
            "input_sha256": None,
            "golden_sha256": None,
        }
        for field, path in (
            ("input_sha256", task.input_path),
            ("golden_sha256", task.golden_path),
        ):
            try:
                expected_hashes[field] = _file_sha256(Path(path))
            except OSError:
                code = f"task_{field.removesuffix('_sha256')}_unreadable"
                task_reasons[task.task_id].append(code)
                _add_reason(reasons, f"{code}:{task.task_id}")
        for field, expected in expected_hashes.items():
            if expected is not None and entry.get(field) != expected:
                code = f"manifest_task_hash_mismatch:{field}"
                task_reasons[task.task_id].append(code)
                _add_reason(reasons, f"{code}:{task.task_id}")
    return task_reasons


def _expected_artifact_hash(row: dict[str, Any]) -> tuple[str | None, bool]:
    values: list[str] = []
    direct = row.get("output_sha256")
    if isinstance(direct, str):
        values.append(direct)
    recalculation = row.get("recalculation")
    if isinstance(recalculation, dict) and isinstance(recalculation.get("output_sha256"), str):
        values.append(str(recalculation["output_sha256"]))
    unique = set(values)
    return (values[0] if len(unique) == 1 else None, len(unique) > 1)


def _valid_source_fingerprint(value: Any) -> bool:
    if not isinstance(value, dict):
        return False
    aggregate = value.get("sha256")
    files = value.get("files")
    if (
        not isinstance(aggregate, str)
        or len(aggregate) != 64
        or not isinstance(files, list)
        or not files
    ):
        return False
    combined = hashlib.sha256()
    seen: set[str] = set()
    for entry in files:
        if not isinstance(entry, dict):
            return False
        path = entry.get("path")
        digest = entry.get("sha256")
        if (
            not isinstance(path, str)
            or not path
            or path in seen
            or not isinstance(digest, str)
            or len(digest) != 64
            or any(character not in "0123456789abcdef" for character in digest)
        ):
            return False
        seen.add(path)
        combined.update(path.encode("utf-8"))
        combined.update(b"\0")
        combined.update(digest.encode("ascii"))
        combined.update(b"\n")
    return combined.hexdigest() == aggregate


def _valid_git_object_id(value: Any) -> bool:
    return (
        isinstance(value, str)
        and len(value) == 40
        and all(character in "0123456789abcdef" for character in value)
    )


def _current_repository_git_identity() -> dict[str, str] | None:
    root = Path(__file__).resolve().parents[2]

    def git(*arguments: str) -> str:
        completed = subprocess.run(
            ["git", "-C", str(root), *arguments],
            check=False,
            capture_output=True,
            text=True,
            timeout=15,
            env={**os.environ, "GIT_TERMINAL_PROMPT": "0"},
        )
        if completed.returncode != 0:
            raise OSError("git identity lookup failed")
        return completed.stdout.strip()

    try:
        return {
            "git_commit": git("rev-parse", "--verify", "HEAD^{commit}"),
            "git_tree": git("rev-parse", "--verify", "HEAD^{tree}"),
            "remote_tracking_commit": git(
                "rev-parse", "--verify", "refs/remotes/origin/main^{commit}"
            ),
        }
    except (OSError, subprocess.SubprocessError):
        return None


def _load_continuation_source(
    path: Path,
    *,
    manifest_sha256: str | None,
    required: bool,
    reasons: list[str],
    strict_current_source: bool = True,
    expected_repository_source: dict[str, Any] | None = None,
    bind_repository_source: bool = False,
) -> tuple[dict[str, Any] | None, str | None]:
    if not path.exists():
        if required:
            _add_reason(reasons, "continuation_source_missing")
        return None, None
    record_sha256: str | None = None
    try:
        raw = _regular_file_bytes_for_audit(path)
        file_sha256 = hashlib.sha256(raw).hexdigest()
        record = json.loads(
            raw.decode("utf-8"),
            object_pairs_hook=_reject_duplicate_json_keys_for_audit,
        )
        if not isinstance(record, dict) or set(record) != {
            "schema_version",
            "comparison_manifest_sha256",
            "repository_source",
            "record_sha256",
        }:
            raise ValueError("invalid continuation source fields")
        repository_source = record.get("repository_source")
        if not isinstance(repository_source, dict) or set(repository_source) != {
            "schema_version",
            "git_commit",
            "git_tree",
            "remote_tracking_ref",
            "remote_tracking_commit",
            "remote_name",
            "remote_ref",
            "remote_observed_commit",
            "source_fingerprint",
        }:
            raise ValueError("invalid repository source fields")
        unsigned = {
            "schema_version": record.get("schema_version"),
            "comparison_manifest_sha256": record.get(
                "comparison_manifest_sha256"
            ),
            "repository_source": repository_source,
        }
        record_sha256 = _text_sha256(
            json.dumps(
                unsigned,
                ensure_ascii=False,
                sort_keys=True,
                separators=(",", ":"),
            )
        )
        if (
            record.get("schema_version") != 1
            or record.get("comparison_manifest_sha256") != manifest_sha256
            or record.get("record_sha256") != record_sha256
            or repository_source.get("schema_version") != 1
            or not _valid_git_object_id(repository_source.get("git_commit"))
            or not _valid_git_object_id(repository_source.get("git_tree"))
            or repository_source.get("remote_tracking_ref")
            != "refs/remotes/origin/main"
            or repository_source.get("remote_tracking_commit")
            != repository_source.get("git_commit")
            or repository_source.get("remote_name") != "origin"
            or repository_source.get("remote_ref") != "refs/heads/main"
            or repository_source.get("remote_observed_commit")
            != repository_source.get("git_commit")
            or not _valid_source_fingerprint(
                repository_source.get("source_fingerprint")
            )
            or (
                strict_current_source
                and repository_source.get("source_fingerprint")
                != _source_fingerprint()
            )
            or (bind_repository_source and repository_source != expected_repository_source)
        ):
            raise ValueError("invalid continuation source")
        return record, file_sha256
    except (OSError, UnicodeError, json.JSONDecodeError, ValueError):
        _add_reason(reasons, "continuation_source_invalid")
        return None, record_sha256


def _audit_manifest_contract(
    manifest: dict[str, Any],
    reasons: list[str],
    *,
    results_root: Path,
    manifest_sha256: str | None,
    contract: _AuditProtocolContract | None,
) -> None:
    configuration = manifest.get("configuration")
    required_configuration = {
        "model",
        "api_protocol",
        "requested_reasoning_effort",
        "reasoning_effort",
        "request_interval_seconds",
        "litellm_timeout_seconds",
        "generation",
        "max_model_calls",
        "max_turns_per_arm",
        "max_total_tokens",
        "max_output_tokens_per_call",
        "task_timeout_seconds",
        "recalculate",
    }
    if not isinstance(configuration, dict) or not required_configuration.issubset(
        configuration
    ):
        _add_reason(reasons, "comparison_manifest_configuration_invalid")
    elif contract is not None and any(
        configuration.get(field) != expected
        for field, expected in contract.configuration_policies.items()
    ):
        _add_reason(reasons, "comparison_manifest_policy_mismatch")
    if not _valid_source_fingerprint(manifest.get("harness_source")):
        _add_reason(reasons, "comparison_manifest_source_fingerprint_invalid")
    elif (
        contract is not None
        and contract.strict_current_source
        and manifest.get("harness_source") != _source_fingerprint()
    ):
        _add_reason(reasons, "comparison_manifest_source_checkout_mismatch")
    runtime = manifest.get("runtime")
    if not isinstance(runtime, dict) or not runtime.get("python"):
        _add_reason(reasons, "comparison_manifest_runtime_invalid")
    split_provenance = manifest.get("split_provenance")
    if split_provenance is not None:
        task_ids = manifest.get("task_ids")
        required_split_fields = {
            "manifest_id",
            "schema_version",
            "manifest_sha256",
            "task_count",
            "task_ids_sha256",
            "dataset_json_sha256",
        }
        if (
            not isinstance(split_provenance, dict)
            or set(split_provenance) != required_split_fields
            or not isinstance(task_ids, list)
            or split_provenance.get("task_count") != len(task_ids)
            or split_provenance.get("task_ids_sha256")
            != _text_sha256("".join(f"{task_id}\n" for task_id in task_ids))
            or split_provenance.get("dataset_json_sha256")
            != manifest.get("dataset_manifest_sha256")
            or any(
                not isinstance(split_provenance.get(field), str)
                or not split_provenance[field]
                for field in ("schema_version", "manifest_sha256", "task_ids_sha256")
            )
            or any(
                len(str(split_provenance.get(field))) != 64
                or any(
                    character not in "0123456789abcdef"
                    for character in str(split_provenance.get(field))
                )
                for field in (
                    "manifest_sha256",
                    "task_ids_sha256",
                    "dataset_json_sha256",
                )
            )
            or not verify_trace2skill_split_provenance(split_provenance)
        ):
            _add_reason(reasons, "comparison_manifest_split_provenance_invalid")
    run_spec_provenance = manifest.get("run_spec_provenance")
    split_manifest_id = (
        split_provenance.get("manifest_id")
        if isinstance(split_provenance, dict)
        and isinstance(split_provenance.get("manifest_id"), str)
        else None
    )
    if (
        isinstance(split_provenance, dict)
        and split_manifest_id in protected_run_spec_split_ids()
        and run_spec_provenance is None
    ):
        _add_reason(reasons, "comparison_manifest_run_spec_missing_for_pilot")
    if run_spec_provenance is not None:
        if not verify_pilot_run_spec_provenance(run_spec_provenance):
            _add_reason(reasons, "comparison_manifest_run_spec_provenance_invalid")
        try:
            raw_spec = _regular_file_bytes_for_audit(
                results_root / RUN_SPEC_COPY_FILENAME
            )
            document, provenance = parse_pilot_run_spec_bytes(raw_spec)
            if provenance != run_spec_provenance:
                _add_reason(reasons, "comparison_manifest_run_spec_copy_mismatch")
            anchor = resolve_run_spec_anchor(provenance)
            if contract is None or (
                anchor.comparison_protocol_version != contract.protocol_version
                or anchor.comparison_manifest_schema_version
                != contract.manifest_schema_version
            ):
                _add_reason(reasons, "comparison_manifest_run_spec_version_mismatch")
            verify_pilot_run_spec_contract(
                document, manifest_execution_contract(manifest)
            )
        except (HarnessError, OSError):
            _add_reason(reasons, "comparison_manifest_run_spec_contract_invalid")
    for field in (
        "stage_turn_caps",
        "forced_tool_prefix_routing",
        "stage_allowed_tools",
        "allowed_observed_terminals",
    ):
        value = manifest.get(field)
        if not isinstance(value, dict) or set(value) != set(manifest.get("arms") or []):
            _add_reason(reasons, f"comparison_manifest_{field}_invalid")
    arms = tuple(str(arm) for arm in (manifest.get("arms") or []))
    max_turns = configuration.get("max_turns_per_arm") if isinstance(configuration, dict) else None
    try:
        expected_caps = comparison_stage_turn_caps(max_turns, arms)
    except (TypeError, ValueError):
        _add_reason(reasons, "comparison_manifest_turn_caps_invalid")
    else:
        if manifest.get("stage_turn_caps") != expected_caps:
            _add_reason(reasons, "comparison_manifest_turn_caps_mismatch")
        expected_prefixes = {
            arm: {
                stage: list(prefix)
                for stage, prefix in COMPARISON_FORCED_TOOL_PREFIX_POLICY[arm].items()
            }
            for arm in arms
        }
        if manifest.get("forced_tool_prefix_routing") != expected_prefixes:
            _add_reason(reasons, "comparison_manifest_forced_routing_mismatch")
        if manifest.get("stage_allowed_tools") != _stage_allowed_tools_policy(
            arms,
            protocol_version=(
                contract.protocol_version
                if contract is not None
                else COMPARISON_PROTOCOL_VERSION
            ),
        ):
            _add_reason(reasons, "comparison_manifest_stage_tools_mismatch")
        if manifest.get("allowed_observed_terminals") != (
            _allowed_observed_terminals_policy(
                expected_caps,
                protocol_version=(
                    contract.protocol_version
                    if contract is not None
                    else COMPARISON_PROTOCOL_VERSION
                ),
            )
        ):
            _add_reason(reasons, "comparison_manifest_terminal_policy_mismatch")


def _audit_row_contract(
    record: dict[str, Any],
    row: dict[str, Any],
    task: SpreadsheetTask,
    arm: str,
    manifest: dict[str, Any],
    manifest_sha256: str | None,
    contract: _AuditProtocolContract | None,
) -> None:
    reasons: list[str] = record["reasons"]
    configuration = manifest.get("configuration")
    if not isinstance(configuration, dict):
        _add_reason(reasons, "manifest_configuration_unavailable")
        return
    if row.get("comparison_manifest_sha256") != manifest_sha256:
        _add_reason(reasons, "manifest_sha256_binding_mismatch")
    if row.get("split_provenance") != manifest.get("split_provenance"):
        _add_reason(reasons, "row_manifest_mismatch:split_provenance")
    if row.get("run_spec_provenance") != manifest.get("run_spec_provenance"):
        _add_reason(reasons, "row_manifest_mismatch:run_spec_provenance")
    expected_fields = {
        "model": configuration.get("model"),
        "api_protocol": configuration.get("api_protocol"),
        "requested_reasoning_effort": configuration.get("requested_reasoning_effort"),
        "reasoning_effort": configuration.get("reasoning_effort"),
        "request_interval_seconds": configuration.get("request_interval_seconds"),
        "litellm_timeout_seconds": configuration.get("litellm_timeout_seconds"),
        "generation": configuration.get("generation"),
        "max_model_calls": configuration.get("max_model_calls"),
        "max_turns_per_arm": configuration.get("max_turns_per_arm"),
        "stage_turn_caps": (manifest.get("stage_turn_caps") or {}).get(arm),
        "calculation_backend": (
            "libreoffice" if configuration.get("recalculate") is True else "not_recalculated"
        ),
        "instruction_type": task.instruction_type,
    }
    for field, expected in expected_fields.items():
        if row.get(field) != expected:
            _add_reason(reasons, f"row_manifest_mismatch:{field}")
    budget = row.get("budget")
    limit = budget.get("limit") if isinstance(budget, dict) else None
    used = budget.get("used") if isinstance(budget, dict) else None
    expected_limit = {
        "model_calls": configuration.get("max_model_calls"),
        "total_tokens": configuration.get("max_total_tokens"),
        "elapsed_seconds": configuration.get("task_timeout_seconds"),
    }
    if limit != expected_limit:
        _add_reason(reasons, "budget_limit_mismatch")
    if not isinstance(used, dict):
        _add_reason(reasons, "budget_used_invalid")
    else:
        budget_exhaustion, budget_termination = _v25_budget_exhaustion(row, contract)
        for field in ("model_calls", "total_tokens"):
            value = used.get(field)
            ceiling = expected_limit[field]
            single_response_token_overage = False
            if (
                field == "total_tokens"
                and budget_exhaustion
                and contract is not None
                and contract.allow_final_response_token_overage
                and isinstance(value, int)
                and not isinstance(value, bool)
                and isinstance(ceiling, int)
                and not isinstance(ceiling, bool)
                and value > ceiling
                and (budget_termination or {}).get("reason") == "max_total_tokens"
            ):
                timings = (row.get("agent") or {}).get("request_timings")
                timing_tokens = (
                    [timing.get("total_tokens") for timing in timings]
                    if isinstance(timings, list)
                    and timings
                    and all(isinstance(timing, dict) for timing in timings)
                    else []
                )
                single_response_token_overage = bool(
                    timing_tokens
                    and all(
                        isinstance(tokens, int)
                        and not isinstance(tokens, bool)
                        and tokens >= 0
                        for tokens in timing_tokens
                    )
                    and sum(timing_tokens) == value
                    and sum(timing_tokens[:-1]) < ceiling
                    and sum(timing_tokens[:-1]) + timing_tokens[-1] > ceiling
                )
            if (
                isinstance(value, bool)
                or not isinstance(value, int)
                or value < 0
                or not isinstance(ceiling, int)
                or (value > ceiling and not single_response_token_overage)
            ):
                _add_reason(reasons, f"budget_used_invalid:{field}")


def _audit_completed_agent(
    record: dict[str, Any],
    row: dict[str, Any],
    arm: str,
    manifest: dict[str, Any],
    contract: _AuditProtocolContract | None,
) -> None:
    reasons: list[str] = record["reasons"]
    agent = row.get("agent")
    if not isinstance(agent, dict):
        _add_reason(reasons, "agent_evidence_missing")
        return
    if agent.get("arm") != arm:
        _add_reason(reasons, "agent_arm_mismatch")
    stages = agent.get("stages")
    expected_caps = (manifest.get("stage_turn_caps") or {}).get(arm)
    expected_prefixes = (manifest.get("forced_tool_prefix_routing") or {}).get(arm)
    expected_tools = (manifest.get("stage_allowed_tools") or {}).get(arm)
    allowed_terminals = (manifest.get("allowed_observed_terminals") or {}).get(arm)
    budget_exhaustion, budget_termination = _v25_budget_exhaustion(row, contract)
    agent_tool_failure_stage = _agent_tool_recalculation_failure_stage(row)
    exact_evidence = bool(contract and contract.require_exact_agent_evidence)
    generic_response_truncation = bool(
        contract
        and contract.allow_generic_response_truncation
        and row.get("model_failure_reason") == "model_response_truncated"
    )
    if not isinstance(stages, list) or not isinstance(expected_caps, dict):
        _add_reason(reasons, "agent_stages_invalid")
        return
    expected_names = list(expected_caps)
    observed_names = [
        stage.get("name") if isinstance(stage, dict) else None for stage in stages
    ]
    budget_truncated_paper_stages = bool(
        budget_exhaustion
        and arm == "paper"
        and observed_names
        and observed_names == expected_names[: len(observed_names)]
        and observed_names[-1] == (budget_termination or {}).get("stage")
    )
    infrastructure_truncated_stages = bool(
        agent_tool_failure_stage
        and observed_names
        and observed_names == expected_names[: len(observed_names)]
        and observed_names[-1] == agent_tool_failure_stage
    )
    generic_truncated_stages = bool(
        generic_response_truncation
        and observed_names
        and observed_names == expected_names[: len(observed_names)]
    )
    if budget_exhaustion and (
        not observed_names
        or observed_names[-1] != (budget_termination or {}).get("stage")
    ):
        _add_reason(reasons, "agent_budget_termination_stage_mismatch")
    if observed_names != expected_names and not (
        budget_truncated_paper_stages
        or infrastructure_truncated_stages
        or generic_truncated_stages
    ):
        _add_reason(reasons, "agent_stage_order_mismatch")
        return
    timing_count = 0
    expected_aggregate_timings: list[dict[str, Any]] = []
    expected_aggregate_tool_trace: list[dict[str, Any]] = []
    aggregate_turns = 0
    aggregate_tool_calls = 0
    aggregate_terminal_submissions = 0
    aggregate_usage = {"input_tokens": 0, "output_tokens": 0, "total_tokens": 0}
    cumulative_model_calls = 0
    cumulative_total_tokens = 0
    for stage in stages:
        assert isinstance(stage, dict)
        name = str(stage["name"])
        budget_failure_stage = bool(
            budget_exhaustion
            and stage is stages[-1]
            and name == (budget_termination or {}).get("stage")
        )
        infrastructure_failure_stage = bool(
            agent_tool_failure_stage
            and stage is stages[-1]
            and name == agent_tool_failure_stage
        )
        generic_failure_stage = bool(
            generic_response_truncation and stage is stages[-1]
        )
        if stage.get("max_turns") != expected_caps.get(name):
            _add_reason(reasons, f"agent_stage_turn_cap_mismatch:{name}")
        if stage.get("allowed_tools") != (expected_tools or {}).get(name):
            _add_reason(reasons, f"agent_stage_tools_mismatch:{name}")
        prefix = (expected_prefixes or {}).get(name)
        if stage.get("forced_tool_prefix") != prefix:
            _add_reason(reasons, f"agent_forced_prefix_mismatch:{name}")
        observed_prefix = stage.get("observed_forced_tool_prefix")
        observed_is_expected_prefix = bool(
            isinstance(prefix, list)
            and isinstance(observed_prefix, list)
            and observed_prefix == prefix[: len(observed_prefix)]
        )
        if observed_prefix != prefix and not (
            (budget_failure_stage or generic_failure_stage)
            and observed_is_expected_prefix
        ):
            _add_reason(reasons, f"agent_observed_prefix_mismatch:{name}")
        if prefix:
            expected_observed_first = (
                prefix[0] if isinstance(observed_prefix, list) and observed_prefix else None
            )
            if stage.get("first_tool_choice") != prefix[0] or (
                stage.get("observed_first_tool")
                != (
                    expected_observed_first
                    if budget_failure_stage or generic_failure_stage
                    else prefix[0]
                )
            ):
                _add_reason(reasons, f"agent_first_tool_mismatch:{name}")
        expected_terminal = (
            "assistant_text"
            if arm == "paper" and name == "reconcile"
            else (manifest.get("post_prefix_routing") or {}).get("terminal_tool")
        )
        budget_reconcile_terminal = bool(
            budget_failure_stage and arm == "paper" and name == "reconcile"
        )
        generic_reconcile_terminal = bool(
            generic_failure_stage and arm == "paper" and name == "reconcile"
        )
        if infrastructure_failure_stage:
            terminal_valid = stage.get("observed_terminal_tool") is None
        elif generic_failure_stage:
            terminal_valid = (
                stage.get("observed_terminal_tool")
                == MODEL_RESPONSE_TRUNCATED_TERMINAL
            )
        elif budget_failure_stage:
            terminal_valid = stage.get("observed_terminal_tool") == "budget_exhausted"
        else:
            observed_terminal = stage.get("observed_terminal_tool")
            terminal_valid = (
                observed_terminal != "budget_exhausted"
                and observed_terminal in (allowed_terminals or {}).get(name, [])
            )
        if not terminal_valid:
            _add_reason(reasons, f"agent_observed_terminal_invalid:{name}")
        stage_agent = stage.get("agent")
        if not isinstance(stage_agent, dict):
            _add_reason(reasons, f"agent_stage_evidence_missing:{name}")
            continue
        turns = stage_agent.get("turns")
        zero_turn_budget_failure = bool(budget_failure_stage and turns == 0)
        if (budget_reconcile_terminal or generic_reconcile_terminal) and exact_evidence:
            expected_budget_terminal = "assistant_text" if zero_turn_budget_failure else None
            terminal_matches = stage.get("terminal_tool") == expected_budget_terminal
        else:
            terminal_matches = stage.get("terminal_tool") == expected_terminal
        if not terminal_matches:
            _add_reason(reasons, f"agent_terminal_tool_mismatch:{name}")
        if (
            isinstance(turns, bool)
            or not isinstance(turns, int)
            or (turns < 1 and not zero_turn_budget_failure)
            or turns > expected_caps[name]
        ):
            _add_reason(reasons, f"agent_stage_turns_invalid:{name}")
        stage_timings = stage_agent.get("request_timings")
        if not isinstance(stage_timings, list) or len(stage_timings) != turns:
            _add_reason(reasons, f"agent_stage_request_count_mismatch:{name}")
        else:
            timing_count += len(stage_timings)
            timings_are_dicts = all(isinstance(timing, dict) for timing in stage_timings)
            if exact_evidence and timings_are_dicts:
                expected_aggregate_timings.extend(
                    {"stage": name, **timing} for timing in stage_timings
                )
            if any(
                not isinstance(timing, dict)
                or ("stage" in timing and timing.get("stage") != name)
                for timing in stage_timings
            ):
                _add_reason(reasons, f"agent_stage_request_stage_mismatch:{name}")
        if exact_evidence:
            wrapper_trace = stage.get("tool_trace")
            nested_trace = stage_agent.get("tool_trace")
            if not isinstance(wrapper_trace, list) or not all(
                isinstance(item, dict) for item in wrapper_trace
            ):
                _add_reason(reasons, f"agent_stage_tool_trace_invalid:{name}")
                wrapper_trace = []
            if nested_trace != wrapper_trace:
                _add_reason(reasons, f"agent_stage_tool_trace_mismatch:{name}")
            if stage.get("tool_name_trace") != [
                str(item.get("name", "")) for item in wrapper_trace
            ]:
                _add_reason(reasons, f"agent_stage_tool_names_mismatch:{name}")
            trace_names = [str(item.get("name", "")) for item in wrapper_trace]
            successful_trace_names = [
                str(item.get("name", ""))
                for item in wrapper_trace
                if item.get("ok") is True
            ]
            if isinstance(observed_prefix, list):
                missing_in_response_budget_tool = bool(
                    budget_failure_stage
                    and not zero_turn_budget_failure
                    and len(observed_prefix) == len(trace_names) + 1
                    and trace_names == observed_prefix[:-1]
                )
                if trace_names[: len(observed_prefix)] != observed_prefix and not (
                    missing_in_response_budget_tool
                ):
                    _add_reason(reasons, f"agent_stage_prefix_trace_mismatch:{name}")
                requires_successful_prefix = arm == "paper" and name in {
                    "vision_verify",
                    "latex_verify",
                }
                if requires_successful_prefix and successful_trace_names[
                    : len(observed_prefix)
                ] != observed_prefix:
                    _add_reason(reasons, f"agent_stage_prefix_success_mismatch:{name}")
            allowed_stage_tools = (expected_tools or {}).get(name)
            if allowed_stage_tools != "all" and any(
                item.get("name") not in (allowed_stage_tools or [])
                for item in wrapper_trace
            ):
                _add_reason(reasons, f"agent_stage_tool_not_allowed:{name}")
            expected_aggregate_tool_trace.extend(
                {"stage": name, **item} for item in wrapper_trace
            )

            stage_tool_calls = _non_negative_int(stage_agent.get("tool_calls"))
            stage_terminal_submissions = _non_negative_int(
                stage_agent.get("terminal_submissions")
            )
            if stage_tool_calls != len(wrapper_trace):
                _add_reason(reasons, f"agent_stage_tool_count_mismatch:{name}")
            if stage_terminal_submissions is None:
                _add_reason(reasons, f"agent_stage_terminal_count_invalid:{name}")
            expected_terminal_submissions = int(
                stage.get("observed_terminal_tool") == "submit_result"
            )
            if (
                arm == "paper"
                and name != "solve"
                and not budget_failure_stage
                and stage.get("observed_terminal_tool") == "assistant_text"
            ):
                expected_terminal_submissions = 0
            if stage_terminal_submissions != expected_terminal_submissions:
                _add_reason(reasons, f"agent_stage_terminal_count_mismatch:{name}")
            expected_function_calls = (
                None
                if stage_tool_calls is None or stage_terminal_submissions is None
                else stage_tool_calls + stage_terminal_submissions
            )
            if stage_agent.get("function_calls_total") != expected_function_calls:
                _add_reason(reasons, f"agent_stage_function_count_mismatch:{name}")
            if isinstance(turns, int) and not isinstance(turns, bool):
                aggregate_turns += turns
            if stage_tool_calls is not None:
                aggregate_tool_calls += stage_tool_calls
            if stage_terminal_submissions is not None:
                aggregate_terminal_submissions += stage_terminal_submissions

            stage_usage = _usage_triplet(stage_agent.get("usage"))
            if stage_usage is None:
                _add_reason(reasons, f"agent_stage_usage_invalid:{name}")
            else:
                for field, value in stage_usage.items():
                    aggregate_usage[field] += value
                timing_usage = {"input_tokens": 0, "output_tokens": 0, "total_tokens": 0}
                timing_usage_valid = isinstance(stage_timings, list)
                if isinstance(stage_timings, list):
                    for timing in stage_timings:
                        timing_tokens = _usage_triplet(timing)
                        if timing_tokens is None:
                            timing_usage_valid = False
                            break
                        for field, value in timing_tokens.items():
                            timing_usage[field] += value
                if not timing_usage_valid or timing_usage != stage_usage:
                    _add_reason(reasons, f"agent_stage_timing_usage_mismatch:{name}")
                cumulative_total_tokens += stage_usage["total_tokens"]
            if isinstance(stage_timings, list):
                cumulative_model_calls += len(stage_timings)

            stage_budget = stage_agent.get("budget")
            stage_limit = (
                stage_budget.get("limit") if isinstance(stage_budget, dict) else None
            )
            stage_used = (
                stage_budget.get("used") if isinstance(stage_budget, dict) else None
            )
            row_budget = row.get("budget") or {}
            stage_budget_valid = bool(
                isinstance(stage_budget, dict)
                and stage_limit == row_budget.get("limit")
                and isinstance(stage_used, dict)
                and all(
                    _non_negative_int(stage_used.get(field)) is not None
                    for field in ("model_calls", "total_tokens")
                )
                and stage_used.get("model_calls") == cumulative_model_calls
                and stage_used.get("total_tokens") == cumulative_total_tokens
                and (
                    stage_budget.get("termination") is None
                    or (
                        budget_failure_stage
                        and stage_budget.get("termination") == budget_termination
                    )
                )
            )
            if not stage_budget_valid:
                _add_reason(reasons, f"agent_stage_budget_mismatch:{name}")
        elif zero_turn_budget_failure:
            stage_budget = stage_agent.get("budget")
            stage_used = (
                stage_budget.get("used") if isinstance(stage_budget, dict) else None
            )
            row_used = ((row.get("budget") or {}).get("used") or {})
            if (
                not isinstance(stage_budget, dict)
                or stage_budget.get("termination") != budget_termination
                or not isinstance(stage_used, dict)
                or any(
                    stage_used.get(field) != row_used.get(field)
                    for field in ("model_calls", "total_tokens")
                )
            ):
                _add_reason(reasons, f"agent_stage_budget_mismatch:{name}")
    budget_calls = ((row.get("budget") or {}).get("used") or {}).get("model_calls")
    aggregate_timings = agent.get("request_timings")
    if (
        not isinstance(aggregate_timings, list)
        or len(aggregate_timings) != timing_count
        or len(aggregate_timings) != budget_calls
    ):
        _add_reason(reasons, "agent_request_count_mismatch")
    if exact_evidence and aggregate_timings != expected_aggregate_timings:
        _add_reason(reasons, "agent_request_timings_mismatch")
    if not bool(_request_attempt_audit(row)["exact"]):
        _add_reason(reasons, "request_attempt_audit_inexact")
    expected_endpoint = (
        "/responses" if row.get("api_protocol") == "responses" else "/chat/completions"
    )
    if isinstance(aggregate_timings, list):
        for timing in aggregate_timings:
            if not isinstance(timing, dict):
                continue
            attempts = timing.get("attempts")
            history = timing.get("attempt_history")
            if isinstance(history, list) and isinstance(attempts, int):
                for attempt in history:
                    if not isinstance(attempt, dict):
                        _add_reason(reasons, "request_attempt_history_invalid")
                        break
                    if attempt.get("api_protocol") != row.get("api_protocol"):
                        _add_reason(reasons, "request_attempt_api_protocol_mismatch")
                    if attempt.get("endpoint") != expected_endpoint:
                        _add_reason(reasons, "request_attempt_endpoint_mismatch")
    usage = agent.get("usage")
    used = (row.get("budget") or {}).get("used") or {}
    if not isinstance(usage, dict) or usage.get("total_tokens") != used.get("total_tokens"):
        _add_reason(reasons, "agent_budget_token_mismatch")
    if exact_evidence:
        if _usage_triplet(usage) != aggregate_usage:
            _add_reason(reasons, "agent_stage_usage_mismatch")
        if agent.get("turns") != aggregate_turns:
            _add_reason(reasons, "agent_turn_count_mismatch")
        if agent.get("tool_calls") != aggregate_tool_calls:
            _add_reason(reasons, "agent_tool_count_mismatch")
        if agent.get("terminal_submissions") != aggregate_terminal_submissions:
            _add_reason(reasons, "agent_terminal_count_mismatch")
        if agent.get("function_calls_total") != (
            aggregate_tool_calls + aggregate_terminal_submissions
        ):
            _add_reason(reasons, "agent_function_count_mismatch")
        if agent.get("tool_trace") != expected_aggregate_tool_trace:
            _add_reason(reasons, "agent_tool_trace_mismatch")
    agent_budget = agent.get("budget")
    agent_limit = agent_budget.get("limit") if isinstance(agent_budget, dict) else None
    agent_used = agent_budget.get("used") if isinstance(agent_budget, dict) else None
    if agent_limit != (row.get("budget") or {}).get("limit"):
        _add_reason(reasons, "agent_budget_limit_mismatch")
    if not isinstance(agent_used, dict) or any(
        agent_used.get(field) != used.get(field)
        for field in ("model_calls", "total_tokens")
    ):
        _add_reason(reasons, "agent_budget_usage_mismatch")
    if (
        not isinstance(agent_budget, dict)
        or agent_budget.get("termination")
        != (row.get("budget") or {}).get("termination")
    ):
        _add_reason(reasons, "agent_budget_termination_mismatch")
    if exact_evidence and any(
        expected != used.get(field)
        for field, expected in (
            ("model_calls", cumulative_model_calls),
            ("total_tokens", cumulative_total_tokens),
        )
    ):
        _add_reason(reasons, "agent_final_stage_budget_mismatch")
    if contract is not None and contract.require_truncated_terminal_evidence:
        truncation_claimed = bool(
            row.get("model_failure_reason")
            in {"model_response_truncated", "terminal_submission_truncated"}
            or agent.get("observed_terminal_tool")
            in {
                MODEL_RESPONSE_TRUNCATED_TERMINAL,
                TERMINAL_SUBMISSION_TRUNCATED_OBSERVED,
            }
            or (
                isinstance(agent.get("terminal_response"), dict)
                and agent["terminal_response"].get("status") == "truncated"
            )
            or any(
                stage.get("observed_terminal_tool")
                in {
                    MODEL_RESPONSE_TRUNCATED_TERMINAL,
                    TERMINAL_SUBMISSION_TRUNCATED_OBSERVED,
                }
                or (
                    isinstance(stage.get("agent"), dict)
                    and isinstance(stage["agent"].get("terminal_response"), dict)
                    and stage["agent"]["terminal_response"].get("status")
                    == "truncated"
                )
                for stage in stages
            )
            or _v26_output_limit_attempt_observed(stages)
        )
        if truncation_claimed:
            if contract.allow_generic_response_truncation:
                truncation_valid = _v29_generic_output_limit_evidence_valid(
                    row,
                    agent,
                    stages,
                    budget_exhaustion=budget_exhaustion,
                )
                invalid_reason = "output_limit_response_evidence_invalid"
            else:
                truncation_valid = _v26_truncated_terminal_evidence_valid(
                    row,
                    agent,
                    stages,
                    budget_exhaustion=budget_exhaustion,
                )
                invalid_reason = "terminal_submission_truncated_evidence_invalid"
            if not truncation_valid:
                _add_reason(reasons, invalid_reason)
    if (
        contract is not None
        and contract.require_accepted_terminal_evidence
        and row.get("status") in {"completed", "error"}
    ):
        outcome_kind = row.get("outcome_kind")
        if outcome_kind == "scored":
            terminal_evidence_valid = _v26_success_terminal_evidence_valid(
                row, agent, stages, arm=arm
            )
        elif outcome_kind == "infrastructure_failure":
            if agent_tool_failure_stage:
                terminal_evidence_valid = (
                    _agent_tool_recalculation_terminal_evidence_valid(
                        row,
                        agent,
                        stages,
                        arm=arm,
                    )
                )
            else:
                terminal_evidence_valid = _v26_success_terminal_evidence_valid(
                    {**row, "status": "completed", "outcome_kind": "scored"},
                    agent,
                    stages,
                    arm=arm,
                )
        elif outcome_kind == "model_execution_failure":
            terminal_evidence_valid = _v26_model_failure_terminal_response_valid(
                agent, stages, arm=arm
            )
        else:
            terminal_evidence_valid = True
        if not terminal_evidence_valid:
            _add_reason(reasons, "accepted_terminal_response_evidence_invalid")


def _audit_recalculation_integrity_evidence(
    reasons: list[str],
    row: dict[str, Any],
    output: Path,
    *,
    failure: bool,
) -> None:
    recalculation = row.get("recalculation")
    if not isinstance(recalculation, dict):
        _add_reason(reasons, "recalculation_evidence_missing")
        return
    integrity = recalculation.get("sheet_inventory_integrity")
    if (
        recalculation.get("backend") != "libreoffice-headless"
        or recalculation.get("profile") != "isolated-per-invocation"
        or recalculation.get("source_path") != str(output)
        or recalculation.get("destination_path") != str(output)
        or recalculation.get("format") != output.suffix.lower().lstrip(".")
        or not isinstance(recalculation.get("version"), str)
        or not recalculation["version"]
        or not isinstance(integrity, dict)
        or set(integrity)
        != {"schema_version", "policy", "enforced", "matched", "pre", "post"}
        or integrity.get("schema_version") != 2
        or integrity.get("policy") != RECALCULATION_SHEET_INTEGRITY_POLICY
        or integrity.get("enforced") is not True
    ):
        _add_reason(reasons, "recalculation_evidence_invalid")
        return

    pre = integrity.get("pre")
    post = integrity.get("post")
    pre_valid = _valid_sheet_inventory_identity(pre)
    post_valid = _valid_sheet_inventory_identity(post)
    if not pre_valid:
        _add_reason(reasons, "recalculation_pre_identity_invalid")
    if not post_valid:
        _add_reason(reasons, "recalculation_post_identity_invalid")
    if not pre_valid or not post_valid or not isinstance(pre, dict) or not isinstance(post, dict):
        return
    if (
        recalculation.get("source_sha256") != pre["workbook_sha256"]
        or recalculation.get("output_sha256") != post["workbook_sha256"]
    ):
        _add_reason(reasons, "recalculation_identity_hash_binding_mismatch")

    expected_match = pre["sheets"] == post["sheets"]
    if integrity.get("matched") is not expected_match:
        _add_reason(reasons, "recalculation_inventory_match_claim_invalid")
    if failure:
        if (
            expected_match
            or recalculation.get("atomic_replace") is not False
            or recalculation.get("published") is not False
        ):
            _add_reason(reasons, "recalculation_failure_evidence_invalid")
    elif (
        not expected_match
        or recalculation.get("atomic_replace") is not True
        or recalculation.get("published") is not True
    ):
        _add_reason(reasons, "recalculation_success_evidence_invalid")

    try:
        actual_identity = sheet_inventory_identity(output)
    except Exception:
        _add_reason(reasons, "recalculation_output_identity_unreadable")
        return
    expected_output_identity = pre if failure else post
    if actual_identity != expected_output_identity:
        _add_reason(reasons, "recalculation_output_identity_mismatch")
    if row.get("output_sha256") != actual_identity["workbook_sha256"]:
        _add_reason(reasons, "recalculation_row_output_hash_mismatch")

    if not failure:
        return
    failure_path = _absolute_path(recalculation.get("failure_artifact_path"), base=output.parent)
    failure_sha256 = recalculation.get("failure_artifact_sha256")
    expected_failure_path = output.with_name(
        f"{output.stem}.recalculation-integrity-failure-"
        f"{post['workbook_sha256']}{output.suffix}"
    )
    if (
        failure_path is None
        or failure_path != expected_failure_path
        or failure_sha256 != post["workbook_sha256"]
    ):
        _add_reason(reasons, "recalculation_failure_artifact_binding_invalid")
        return
    try:
        failure_metadata = failure_path.lstat()
        if not stat.S_ISREG(failure_metadata.st_mode):
            raise OSError("not a regular file")
        if _file_sha256(failure_path) != failure_sha256:
            _add_reason(reasons, "recalculation_failure_artifact_hash_mismatch")
            return
        failure_identity = sheet_inventory_identity(failure_path)
    except OSError:
        _add_reason(reasons, "recalculation_failure_artifact_unreadable")
        return
    except Exception:
        _add_reason(reasons, "recalculation_failure_artifact_identity_unreadable")
        return
    if failure_identity != post:
        _add_reason(reasons, "recalculation_failure_artifact_identity_mismatch")


def _infrastructure_agent_audit_row(
    record: dict[str, Any],
    row: dict[str, Any],
    contract: _AuditProtocolContract | None,
) -> dict[str, Any]:
    reasons: list[str] = record["reasons"]
    prior_failure = row.get("prior_model_execution_failure")
    agent_audit_row = row
    if prior_failure is not None:
        record["prior_model_execution_failure"] = prior_failure
        if (
            not isinstance(prior_failure, dict)
            or set(prior_failure)
            != {"error", "error_type", "model_failure_reason"}
            or not isinstance(prior_failure.get("error"), str)
            or not prior_failure["error"]
            or prior_failure.get("error_type") != "AgentExecutionFailure"
            or contract is None
            or prior_failure.get("model_failure_reason")
            not in contract.allowed_model_failure_reasons
        ):
            _add_reason(reasons, "prior_model_execution_failure_invalid")
        else:
            # The recalculation failure remains the row's primary taxonomy, while
            # the terminal evidence must be interpreted as the earlier model failure.
            agent_audit_row = {
                **row,
                "status": "completed",
                "outcome_kind": "model_execution_failure",
                "error": prior_failure["error"],
                "error_type": prior_failure["error_type"],
                "error_retryable": False,
                "error_category": "model_execution_failure",
                "model_failure_reason": prior_failure["model_failure_reason"],
            }
    return agent_audit_row


_PROVIDER_PHASES = frozenset(
    {
        "connect",
        "pool",
        "read",
        "response_body",
        "response_headers",
        "response_protocol",
        "response_stream",
        "total",
        "transport",
        "write",
    }
)
_PROVIDER_DELIVERY_STATES = frozenset(
    {"pre_send", "headers_seen", "terminal_seen", "ambiguous_post_send"}
)


def _provider_attempt_history_valid(
    provider_error: dict[str, Any],
    row: dict[str, Any],
    manifest: dict[str, Any],
) -> bool:
    attempts = _non_negative_int(provider_error.get("attempts"))
    history = provider_error.get("attempt_history")
    configuration = manifest.get("configuration")
    request_retries = (
        _non_negative_int(configuration.get("request_retries"))
        if isinstance(configuration, dict)
        else None
    )
    safe_reasons = (
        configuration.get("safe_automatic_retry_reasons")
        if isinstance(configuration, dict)
        else None
    )
    expected_endpoint = (
        "/responses"
        if row.get("api_protocol") == "responses"
        else "/chat/completions"
    )
    if (
        attempts is None
        or request_retries is None
        or attempts > request_retries + 1
        or not isinstance(history, list)
        or len(history) != attempts
    ):
        return False
    for index, attempt in enumerate(history, start=1):
        if (
            not isinstance(attempt, dict)
            or attempt.get("attempt") != index
            or attempt.get("outcome") != "error"
            or not isinstance(attempt.get("retryable"), bool)
            or not isinstance(attempt.get("safe_to_retry"), bool)
            or not isinstance(attempt.get("automatic_retry_scheduled"), bool)
            or attempt.get("delivery_state") not in _PROVIDER_DELIVERY_STATES
            or attempt.get("api_protocol") != row.get("api_protocol")
            or attempt.get("endpoint") != expected_endpoint
        ):
            return False
        safe_to_retry = attempt["safe_to_retry"]
        safe_reason = attempt.get("safe_retry_reason")
        if (
            safe_to_retry
            and (
                not isinstance(safe_reason, str)
                or not safe_reason
                or safe_reason not in (safe_reasons or [])
            )
        ) or (not safe_to_retry and safe_reason is not None):
            return False
        if attempt.get("delivery_state") == "ambiguous_post_send" and (
            safe_to_retry or safe_reason is not None
        ):
            return False
        retry_scheduled = attempt["automatic_retry_scheduled"]
        if retry_scheduled != (index < attempts):
            return False
        if retry_scheduled and (not safe_to_retry or index > request_retries):
            return False
    return True


def _audit_provider_infrastructure_row(
    record: dict[str, Any],
    row: dict[str, Any],
    task: SpreadsheetTask,
    arm: str,
    root: Path,
    manifest: dict[str, Any],
    contract: _AuditProtocolContract | None,
) -> None:
    reasons: list[str] = record["reasons"]
    provider_error = row.get("provider_error")
    provider_request_audit = row.get("provider_request_audit")
    configuration = manifest.get("configuration")
    error_category = row.get("error_category")
    record.update(
        {
            "outcome_kind": row.get("outcome_kind"),
            "error_category": error_category,
            "error_type": row.get("error_type"),
            "error": row.get("error"),
            "infrastructure_failure_stage": row.get(
                "infrastructure_failure_stage"
            ),
            "provider_failure_reason": row.get("provider_failure_reason"),
            "score_available": row.get("score_available"),
            "provider_request_audit": provider_request_audit,
        }
    )
    taxonomy_valid = bool(
        contract is not None
        and contract.allow_provider_infrastructure_no_score
        and row.get("status") == "error"
        and row.get("outcome_kind") == "infrastructure_failure"
        and row.get("passed") is False
        and row.get("score_available") is False
        and row.get("infrastructure_failure_stage")
        == PROVIDER_INFRASTRUCTURE_FAILURE_STAGE
        and error_category in PROVIDER_INFRASTRUCTURE_CATEGORIES
        and row.get("provider_failure_reason") == error_category
        and row.get("replay_permitted") is False
        and isinstance(row.get("error"), str)
        and bool(row["error"])
        and isinstance(row.get("error_type"), str)
        and row["error_type"].endswith("ProviderError")
        and isinstance(provider_error, dict)
        and "output_limit" not in provider_error
        and row.get("error_retryable") is provider_error.get("safe_to_retry")
    )
    if not taxonomy_valid:
        _add_reason(reasons, "provider_infrastructure_taxonomy_invalid")
    if any(field in row for field in ("comparison", "artifact_score_passed")):
        _add_reason(reasons, "provider_infrastructure_has_score_evidence")

    provider_evidence_valid = isinstance(provider_error, dict)
    if provider_evidence_valid:
        status_code = provider_error.get("status_code")
        retry_after = provider_error.get("retry_after_seconds")
        elapsed = provider_error.get("elapsed_seconds")
        retryable = provider_error.get("retryable")
        global_fatal = provider_error.get("global_fatal")
        safe_to_retry = provider_error.get("safe_to_retry")
        safe_reason = provider_error.get("safe_retry_reason")
        delivery_state = provider_error.get("delivery_state")
        expected_category = (
            "provider_transient"
            if retryable is True
            else "provider_fatal"
            if global_fatal is True
            else "provider_task"
        )
        safe_reasons = (
            configuration.get("safe_automatic_retry_reasons")
            if isinstance(configuration, dict)
            else None
        )
        provider_evidence_valid = bool(
            isinstance(provider_error.get("message"), str)
            and provider_error["message"]
            and isinstance(retryable, bool)
            and isinstance(global_fatal, bool)
            and isinstance(safe_to_retry, bool)
            and provider_error.get("phase") in _PROVIDER_PHASES
            and delivery_state in _PROVIDER_DELIVERY_STATES
            and (
                status_code is None
                or (
                    isinstance(status_code, int)
                    and not isinstance(status_code, bool)
                    and 100 <= status_code <= 599
                )
            )
            and (
                retry_after is None
                or (
                    isinstance(retry_after, int | float)
                    and not isinstance(retry_after, bool)
                    and retry_after >= 0
                )
            )
            and isinstance(elapsed, int | float)
            and not isinstance(elapsed, bool)
            and elapsed >= 0
            and expected_category == error_category
            and not (global_fatal and retryable)
            and (
                safe_to_retry
                and isinstance(safe_reason, str)
                and safe_reason in (safe_reasons or [])
                or not safe_to_retry
                and safe_reason is None
            )
            and not (
                delivery_state == "ambiguous_post_send" and safe_to_retry
            )
            and _provider_attempt_history_valid(
                provider_error,
                row,
                manifest,
            )
        )
    if not provider_evidence_valid:
        _add_reason(reasons, "provider_infrastructure_evidence_invalid")

    request_audit_valid = bool(
        isinstance(provider_request_audit, dict)
        and set(provider_request_audit)
        == {
            "schema_version",
            "request_retries",
            "successful_model_calls",
            "failed_attempts",
            "known_http_attempts",
            "exact",
        }
        and provider_request_audit.get("schema_version")
        == PROVIDER_REQUEST_AUDIT_SCHEMA_VERSION
        and isinstance(configuration, dict)
        and provider_request_audit.get("request_retries")
        == configuration.get("request_retries")
        and provider_request_audit.get("exact") is True
        and _request_attempt_audit(row)["exact"] is True
    )
    if not request_audit_valid:
        _add_reason(reasons, "provider_request_attempt_audit_invalid")

    run_dir = _absolute_path(row.get("run_dir"), base=root)
    output = _absolute_path(row.get("output_workbook"), base=root)
    record["run_dir"] = str(run_dir) if run_dir is not None else None
    record["output_workbook"] = str(output) if output is not None else None
    if run_dir is None:
        _add_reason(reasons, "run_dir_missing")
    if output is None:
        _add_reason(reasons, "output_path_missing")
    if run_dir is None or output is None:
        return
    try:
        resolved_root = root.resolve(strict=True)
    except OSError:
        _add_reason(reasons, "results_dir_unreadable")
        return
    expected_parent = resolved_root / "runs" / task.task_id
    resolved_run = run_dir.resolve(strict=False)
    resolved_output = output.resolve(strict=False)
    if resolved_run.parent != expected_parent or not (
        resolved_run.name == arm or resolved_run.name.startswith(f"{arm}-")
    ):
        _add_reason(reasons, "run_dir_outside_expected_arm")
    expected_output = resolved_run / "artifacts" / f"output{task.input_path.suffix.lower()}"
    if resolved_output != expected_output:
        _add_reason(reasons, "output_path_not_managed_artifact")
    trajectory = resolved_run / "trajectory.jsonl"
    if (
        _has_symlink(root, run_dir)
        or _has_symlink(root, output)
        or _has_symlink(root, trajectory)
    ):
        _add_reason(reasons, "provider_trajectory_path_contains_symlink")
        return
    try:
        output_metadata = output.lstat()
        if not stat.S_ISREG(output_metadata.st_mode):
            raise OSError("not a regular file")
        output_sha256 = _file_sha256(output)
    except OSError:
        _add_reason(reasons, "artifact_unreadable")
        return
    record["output_sha256"] = output_sha256
    if (
        not _valid_sha256(row.get("output_sha256"))
        or row.get("output_sha256") != output_sha256
    ):
        _add_reason(reasons, "artifact_hash_mismatch")
    try:
        metadata = trajectory.lstat()
        if not stat.S_ISREG(metadata.st_mode):
            raise OSError("not a regular file")
        trajectory_rows = read_trajectory(trajectory)
        record["trajectory_sha256"] = _file_sha256(trajectory)
    except (OSError, ValueError):
        _add_reason(reasons, "provider_trajectory_invalid")
        return
    not_evaluated = [
        item
        for item in trajectory_rows
        if item.get("event") == "benchmark.not_evaluated"
    ]
    expected_event_fields = {
        "arm": arm,
        "status": "error",
        "error_category": error_category,
        "outcome_kind": "infrastructure_failure",
        "score_available": False,
        "infrastructure_failure_stage": PROVIDER_INFRASTRUCTURE_FAILURE_STAGE,
        "provider_failure_reason": row.get("provider_failure_reason"),
        "provider_request_audit": provider_request_audit,
    }
    if (
        len(not_evaluated) != 1
        or not isinstance(not_evaluated[0].get("payload"), dict)
        or any(
            not_evaluated[0]["payload"].get(field) != expected
            for field, expected in expected_event_fields.items()
        )
    ):
        _add_reason(reasons, "provider_trajectory_evidence_invalid")

    requested = [
        item for item in trajectory_rows if item.get("event") == "model.requested"
    ]
    responded = [
        item for item in trajectory_rows if item.get("event") == "model.responded"
    ]
    failed = [item for item in trajectory_rows if item.get("event") == "model.failed"]
    successful_calls = (
        _non_negative_int(provider_request_audit.get("successful_model_calls"))
        if isinstance(provider_request_audit, dict)
        else None
    )
    budget_used = (row.get("budget") or {}).get("used") or {}
    response_total_tokens = 0
    response_evidence_valid = True
    expected_endpoint = (
        "/responses"
        if row.get("api_protocol") == "responses"
        else "/chat/completions"
    )
    for event in responded:
        payload = event.get("payload") if isinstance(event, dict) else None
        usage = _usage_triplet(payload.get("usage")) if isinstance(payload, dict) else None
        timing = payload.get("timing") if isinstance(payload, dict) else None
        attempts = (
            _non_negative_int(timing.get("attempts"))
            if isinstance(timing, dict)
            else None
        )
        history = timing.get("attempt_history") if isinstance(timing, dict) else None
        if (
            usage is None
            or attempts is None
            or attempts < 1
            or not isinstance(history, list)
            or len(history) != attempts
            or any(
                not isinstance(attempt, dict)
                or attempt.get("attempt") != index
                or attempt.get("api_protocol") != row.get("api_protocol")
                or attempt.get("endpoint") != expected_endpoint
                or (
                    index == attempts
                    and attempt.get("outcome") != "success"
                )
                for index, attempt in enumerate(history, start=1)
            )
        ):
            response_evidence_valid = False
            continue
        response_total_tokens += usage["total_tokens"]
    failed_payload = (
        failed[0].get("payload")
        if len(failed) == 1 and isinstance(failed[0], dict)
        else None
    )
    failed_provider_error = (
        failed_payload.get("provider_error")
        if isinstance(failed_payload, dict)
        else None
    )
    if isinstance(failed_provider_error, dict) and isinstance(provider_error, dict):
        failed_provider_error = dict(failed_provider_error)
        failed_provider_error["retryable"] = bool(
            failed_provider_error.get("retryable")
        )
    trajectory_accounting_valid = bool(
        response_evidence_valid
        and successful_calls is not None
        and len(responded) == successful_calls
        and len(requested) == successful_calls + 1
        and len(failed) == 1
        and failed_provider_error == provider_error
        and budget_used.get("model_calls") == successful_calls
        and budget_used.get("total_tokens") == response_total_tokens
    )
    if not trajectory_accounting_valid:
        _add_reason(reasons, "provider_trajectory_accounting_invalid")
    try:
        if _file_sha256(output) != output_sha256:
            _add_reason(reasons, "artifact_changed_during_audit")
    except OSError:
        _add_reason(reasons, "artifact_unreadable_after_audit")


def _audit_recalculation_infrastructure_row(
    record: dict[str, Any],
    row: dict[str, Any],
    task: SpreadsheetTask,
    arm: str,
    root: Path,
    manifest: dict[str, Any],
    contract: _AuditProtocolContract | None,
) -> None:
    reasons: list[str] = record["reasons"]
    record.update(
        {
            "outcome_kind": row.get("outcome_kind"),
            "error_category": row.get("error_category"),
            "error_type": row.get("error_type"),
            "error": row.get("error"),
            "infrastructure_failure_stage": row.get("infrastructure_failure_stage"),
            "agent_failure_stage": row.get("agent_failure_stage"),
            "infrastructure_failure_tool": row.get(
                "infrastructure_failure_tool"
            ),
            "recalculation_failure_reason": row.get("recalculation_failure_reason"),
            "score_available": row.get("score_available"),
        }
    )
    failure_stage = row.get("infrastructure_failure_stage")
    agent_tool_failure = failure_stage == AGENT_TOOL_RECALCULATION_FAILURE_STAGE
    postprocess_failure = failure_stage == POSTPROCESS_RECALCULATION_FAILURE_STAGE
    if (
        contract is None
        or not contract.require_recalculation_integrity
        or row.get("status") != "error"
        or row.get("outcome_kind") != "infrastructure_failure"
        or row.get("passed") is not False
        or row.get("score_available") is not False
        or row.get("error_category") != "recalculation_infrastructure"
        or not (agent_tool_failure or postprocess_failure)
        or row.get("recalculation_failure_reason") != "sheet_inventory_changed"
        or row.get("error_type") != "RecalculationIntegrityError"
        or row.get("error_retryable") is not False
        or not isinstance(row.get("error"), str)
        or not row["error"]
    ):
        _add_reason(reasons, "recalculation_infrastructure_taxonomy_invalid")
    if agent_tool_failure:
        expected_stages = (manifest.get("stage_turn_caps") or {}).get(arm)
        if (
            row.get("infrastructure_failure_tool")
            != RECALCULATION_VALIDATION_TOOL
            or not isinstance(row.get("agent_failure_stage"), str)
            or row.get("agent_failure_stage") not in (expected_stages or {})
            or row.get("prior_model_execution_failure") is not None
        ):
            _add_reason(reasons, "recalculation_agent_tool_taxonomy_invalid")
    elif any(
        field in row
        for field in ("agent_failure_stage", "infrastructure_failure_tool")
    ):
        _add_reason(reasons, "recalculation_postprocess_taxonomy_invalid")
    if any(field in row for field in ("comparison", "artifact_score_passed")):
        _add_reason(reasons, "recalculation_failure_has_score_evidence")

    agent_audit_row = _infrastructure_agent_audit_row(
        record,
        row,
        contract,
    )
    _audit_completed_agent(record, agent_audit_row, arm, manifest, contract)
    run_dir = _absolute_path(row.get("run_dir"), base=root)
    output = _absolute_path(row.get("output_workbook"), base=root)
    record["run_dir"] = str(run_dir) if run_dir is not None else None
    record["output_workbook"] = str(output) if output is not None else None
    if run_dir is None:
        _add_reason(reasons, "run_dir_missing")
    if output is None:
        _add_reason(reasons, "output_path_missing")
    if run_dir is None or output is None:
        return

    try:
        resolved_root = root.resolve(strict=True)
    except OSError:
        _add_reason(reasons, "results_dir_unreadable")
        return
    expected_parent = resolved_root / "runs" / task.task_id
    resolved_run = run_dir.resolve(strict=False)
    resolved_output = output.resolve(strict=False)
    if resolved_run.parent != expected_parent or not (
        resolved_run.name == arm or resolved_run.name.startswith(f"{arm}-")
    ):
        _add_reason(reasons, "run_dir_outside_expected_arm")
    if resolved_output != resolved_run / "artifacts" / "output.xlsx":
        _add_reason(reasons, "output_path_not_managed_artifact")
    if _has_symlink(root, run_dir) or _has_symlink(root, output):
        _add_reason(reasons, "artifact_path_contains_symlink")
    try:
        metadata = output.lstat()
        if not stat.S_ISREG(metadata.st_mode):
            raise OSError("not a regular file")
        output_sha256 = _file_sha256(output)
    except OSError:
        _add_reason(reasons, "artifact_unreadable")
        return
    record["output_sha256"] = output_sha256
    if not _valid_sha256(row.get("output_sha256")) or row.get("output_sha256") != output_sha256:
        _add_reason(reasons, "artifact_hash_mismatch")
    _audit_recalculation_integrity_evidence(reasons, row, output, failure=True)
    try:
        if _file_sha256(output) != output_sha256:
            _add_reason(reasons, "artifact_changed_during_audit")
    except OSError:
        _add_reason(reasons, "artifact_unreadable_after_audit")


def _audit_scoring_infrastructure_row(
    record: dict[str, Any],
    row: dict[str, Any],
    task: SpreadsheetTask,
    arm: str,
    root: Path,
    manifest: dict[str, Any],
    contract: _AuditProtocolContract | None,
) -> None:
    reasons: list[str] = record["reasons"]
    record.update(
        {
            "outcome_kind": row.get("outcome_kind"),
            "error_category": row.get("error_category"),
            "error_type": row.get("error_type"),
            "error": row.get("error"),
            "infrastructure_failure_stage": row.get("infrastructure_failure_stage"),
            "scoring_failure_reason": row.get("scoring_failure_reason"),
            "score_available": row.get("score_available"),
        }
    )
    if (
        contract is None
        or not contract.require_recalculation_integrity
        or row.get("status") != "error"
        or row.get("outcome_kind") != "infrastructure_failure"
        or row.get("passed") is not False
        or row.get("score_available") is not False
        or row.get("error_category") != "scoring_infrastructure"
        or row.get("infrastructure_failure_stage") != "scoring"
        or row.get("scoring_failure_reason") != "worksheet_scorer_unsupported"
        or row.get("error_type") != "ScoringInfrastructureError"
        or row.get("error_retryable") is not False
        or not isinstance(row.get("error"), str)
        or not row["error"]
    ):
        _add_reason(reasons, "scoring_infrastructure_taxonomy_invalid")
    if any(field in row for field in ("comparison", "artifact_score_passed")):
        _add_reason(reasons, "scoring_infrastructure_has_score_evidence")

    agent_audit_row = _infrastructure_agent_audit_row(record, row, contract)
    _audit_completed_agent(record, agent_audit_row, arm, manifest, contract)
    run_dir = _absolute_path(row.get("run_dir"), base=root)
    output = _absolute_path(row.get("output_workbook"), base=root)
    record["run_dir"] = str(run_dir) if run_dir is not None else None
    record["output_workbook"] = str(output) if output is not None else None
    if run_dir is None:
        _add_reason(reasons, "run_dir_missing")
    if output is None:
        _add_reason(reasons, "output_path_missing")
    if run_dir is None or output is None:
        return

    try:
        resolved_root = root.resolve(strict=True)
    except OSError:
        _add_reason(reasons, "results_dir_unreadable")
        return
    expected_parent = resolved_root / "runs" / task.task_id
    resolved_run = run_dir.resolve(strict=False)
    resolved_output = output.resolve(strict=False)
    if resolved_run.parent != expected_parent or not (
        resolved_run.name == arm or resolved_run.name.startswith(f"{arm}-")
    ):
        _add_reason(reasons, "run_dir_outside_expected_arm")
    if resolved_output != resolved_run / "artifacts" / "output.xlsx":
        _add_reason(reasons, "output_path_not_managed_artifact")
    if _has_symlink(root, run_dir) or _has_symlink(root, output):
        _add_reason(reasons, "artifact_path_contains_symlink")

    try:
        metadata = output.lstat()
        if not stat.S_ISREG(metadata.st_mode):
            raise OSError("not a regular file")
        output_sha256 = _file_sha256(output)
    except OSError:
        _add_reason(reasons, "artifact_unreadable")
        return
    record["output_sha256"] = output_sha256
    if (
        not _valid_sha256(row.get("output_sha256"))
        or row.get("output_sha256") != output_sha256
    ):
        _add_reason(reasons, "artifact_hash_mismatch")

    try:
        identity = sheet_inventory_identity(output)
    except Exception as exc:
        record["reopen_error_type"] = type(exc).__name__
        _add_reason(reasons, "artifact_reopen_failed")
        return
    record["artifact_validation_method"] = "ooxml-sheet-inventory-v2"
    record["sheet_inventory"] = identity
    record["sheet_names"] = [sheet["name"] for sheet in identity["sheets"]]

    configuration = manifest.get("configuration")
    recalculated = (
        isinstance(configuration, dict)
        and configuration.get("recalculate") is True
    )
    if recalculated:
        _audit_recalculation_integrity_evidence(reasons, row, output, failure=False)
    elif row.get("recalculation") is not None:
        _add_reason(reasons, "unexpected_recalculation_evidence")

    try:
        compare_workbooks_chartsheet_safe(
            task.golden_path,
            output,
            task.answer_position,
            answer_sheet=task.answer_sheet,
        )
    except ScoringInfrastructureError as exc:
        record["scorer_infrastructure_reproduced"] = True
        record["fresh_score_error_type"] = type(exc).__name__
    except Exception as exc:
        record["fresh_score_error_type"] = type(exc).__name__
        _add_reason(reasons, "scoring_infrastructure_evidence_invalid")
    else:
        record["scorer_infrastructure_reproduced"] = False
        _add_reason(reasons, "scoring_infrastructure_not_reproduced")

    try:
        if _file_sha256(output) != output_sha256:
            _add_reason(reasons, "artifact_changed_during_audit")
    except OSError:
        _add_reason(reasons, "artifact_unreadable_after_audit")


def _audit_completed_row(
    record: dict[str, Any],
    row: dict[str, Any],
    task: SpreadsheetTask,
    arm: str,
    root: Path,
    manifest: dict[str, Any],
    manifest_sha256: str | None,
    contract: _AuditProtocolContract | None,
) -> None:
    reasons: list[str] = record["reasons"]
    _audit_row_contract(
        record,
        row,
        task,
        arm,
        manifest,
        manifest_sha256,
        contract,
    )
    infrastructure_claim = bool(
        row.get("outcome_kind") == "infrastructure_failure"
        or row.get("error_category")
        in {"recalculation_infrastructure", "scoring_infrastructure"}
    )
    if infrastructure_claim:
        if (
            row.get("infrastructure_failure_stage")
            == PROVIDER_INFRASTRUCTURE_FAILURE_STAGE
        ):
            _audit_provider_infrastructure_row(
                record,
                row,
                task,
                arm,
                root,
                manifest,
                contract,
            )
        elif row.get("error_category") == "scoring_infrastructure":
            _audit_scoring_infrastructure_row(
                record,
                row,
                task,
                arm,
                root,
                manifest,
                contract,
            )
        elif row.get("error_category") == "recalculation_infrastructure":
            _audit_recalculation_infrastructure_row(
                record,
                row,
                task,
                arm,
                root,
                manifest,
                contract,
            )
        else:
            _add_reason(reasons, "infrastructure_failure_category_invalid")
        return
    if row.get("status") != "completed":
        _add_reason(reasons, "status_not_completed")
        return
    require_v24_fields = bool(contract and contract.require_v24_outcome_fields)
    outcome_kind = row.get("outcome_kind") if require_v24_fields else "scored"
    model_execution_failure = outcome_kind == "model_execution_failure"
    record["outcome_kind"] = outcome_kind
    record["error_category"] = row.get("error_category")
    record["model_failure_reason"] = row.get("model_failure_reason")
    record["error_type"] = row.get("error_type")
    record["error"] = row.get("error")
    if require_v24_fields and outcome_kind not in {"scored", "model_execution_failure"}:
        _add_reason(reasons, "outcome_kind_invalid")
    if model_execution_failure:
        if row.get("passed") is not False:
            _add_reason(reasons, "model_execution_failure_not_failed")
        if row.get("error_category") != "model_execution_failure":
            _add_reason(reasons, "model_execution_failure_category_invalid")
        if (
            contract is None
            or row.get("model_failure_reason")
            not in contract.allowed_model_failure_reasons
        ):
            _add_reason(reasons, "model_execution_failure_reason_invalid")
        if row.get("error_type") != "AgentExecutionFailure":
            _add_reason(reasons, "model_execution_failure_type_invalid")
        if not isinstance(row.get("error"), str) or not row["error"]:
            _add_reason(reasons, "model_execution_failure_error_missing")
        if row.get("error_retryable") is not False:
            _add_reason(reasons, "model_execution_failure_retryable_invalid")
    elif require_v24_fields and any(
        field in row
        for field in (
            "model_failure_reason",
            "error_category",
            "error_type",
            "error_retryable",
        )
    ):
        _add_reason(reasons, "scored_outcome_has_failure_metadata")
    _audit_completed_agent(record, row, arm, manifest, contract)

    run_dir = _absolute_path(row.get("run_dir"), base=root)
    output = _absolute_path(row.get("output_workbook"), base=root)
    record["run_dir"] = str(run_dir) if run_dir is not None else None
    record["output_workbook"] = str(output) if output is not None else None
    if run_dir is None:
        _add_reason(reasons, "run_dir_missing")
    if output is None:
        _add_reason(reasons, "output_path_missing")
    if run_dir is None or output is None:
        return

    try:
        resolved_root = root.resolve(strict=True)
    except OSError:
        _add_reason(reasons, "results_dir_unreadable")
        return
    expected_parent = resolved_root / "runs" / task.task_id
    resolved_run = run_dir.resolve(strict=False)
    resolved_output = output.resolve(strict=False)
    if resolved_run.parent != expected_parent or not (
        resolved_run.name == arm or resolved_run.name.startswith(f"{arm}-")
    ):
        _add_reason(reasons, "run_dir_outside_expected_arm")
    expected_output = resolved_run / "artifacts" / "output.xlsx"
    if resolved_output != expected_output:
        _add_reason(reasons, "output_path_not_managed_artifact")
    if _has_symlink(root, run_dir) or _has_symlink(root, output):
        _add_reason(reasons, "artifact_path_contains_symlink")
    if reasons:
        return

    try:
        metadata = output.lstat()
    except OSError:
        _add_reason(reasons, "artifact_missing")
        return
    if not stat.S_ISREG(metadata.st_mode):
        _add_reason(reasons, "artifact_not_regular_file")
        return

    try:
        output_sha256_before = _file_sha256(output)
    except OSError:
        _add_reason(reasons, "artifact_unreadable")
        return
    record["output_sha256"] = output_sha256_before
    expected_sha256, conflicting_hashes = _expected_artifact_hash(row)
    record["expected_output_sha256"] = expected_sha256
    if conflicting_hashes:
        _add_reason(reasons, "stored_artifact_hashes_conflict")
    elif expected_sha256 is None:
        _add_reason(reasons, "stored_artifact_hash_missing")
    elif not all(character in "0123456789abcdef" for character in expected_sha256) or len(
        expected_sha256
    ) != 64:
        _add_reason(reasons, "stored_artifact_hash_invalid")
    elif output_sha256_before != expected_sha256:
        _add_reason(reasons, "artifact_hash_mismatch")

    v28_artifact_policy = bool(
        contract is not None and contract.require_recalculation_integrity
    )
    if v28_artifact_policy:
        try:
            with openpyxl_worksheet_view(output) as (worksheet_view, identity):
                record["artifact_validation_method"] = (
                    "ooxml-inventory-plus-worksheet-only-openpyxl-view-v1"
                )
                record["sheet_inventory"] = identity
                record["sheet_names"] = [sheet["name"] for sheet in identity["sheets"]]
                expected_worksheet_names = [
                    sheet["name"]
                    for sheet in identity["sheets"]
                    if sheet["kind"] == "worksheet"
                ]
                workbook = load_workbook(
                    worksheet_view,
                    read_only=True,
                    data_only=False,
                )
                try:
                    record["worksheet_names"] = list(workbook.sheetnames)
                finally:
                    workbook.close()
                if record["worksheet_names"] != expected_worksheet_names:
                    raise ScoringInfrastructureError(
                        "The worksheet-only scorer view did not preserve worksheet names"
                    )
        except RenderError as exc:
            record["reopen_error_type"] = type(exc).__name__
            _add_reason(reasons, "artifact_reopen_failed")
            return
        except Exception as exc:
            record.update(
                {
                    "score_available": False,
                    "scorer_infrastructure_stage": "artifact_reopen",
                    "scorer_infrastructure_error_type": type(exc).__name__,
                }
            )
            _add_reason(reasons, "scorer_infrastructure_no_score")
            return
    else:
        try:
            workbook = load_workbook(output, read_only=True, data_only=False)
            try:
                record["sheet_names"] = list(workbook.sheetnames)
            finally:
                workbook.close()
        except Exception as exc:
            record["reopen_error_type"] = type(exc).__name__
            _add_reason(reasons, "artifact_reopen_failed")
            return

    if contract is not None and contract.require_recalculation_integrity:
        configuration = manifest.get("configuration")
        recalculated = isinstance(configuration, dict) and configuration.get("recalculate") is True
        if recalculated:
            _audit_recalculation_integrity_evidence(reasons, row, output, failure=False)
        elif row.get("recalculation") is not None:
            _add_reason(reasons, "unexpected_recalculation_evidence")

    try:
        scorer = (
            compare_workbooks_chartsheet_safe
            if v28_artifact_policy
            else compare_workbooks
        )
        fresh = scorer(
            task.golden_path,
            output,
            task.answer_position,
            answer_sheet=task.answer_sheet,
        )
    except Exception as exc:
        if v28_artifact_policy:
            record.update(
                {
                    "score_available": False,
                    "scorer_infrastructure_stage": "fresh_score",
                    "scorer_infrastructure_error_type": type(exc).__name__,
                }
            )
            _add_reason(reasons, "scorer_infrastructure_no_score")
        else:
            record["fresh_score_error_type"] = type(exc).__name__
            _add_reason(reasons, "fresh_score_failed")
        return
    fresh_dict = fresh.to_dict()
    record["fresh_comparison"] = fresh_dict
    if require_v24_fields and row.get("artifact_score_passed") is not fresh.passed:
        _add_reason(reasons, "stored_artifact_score_passed_mismatch")
    stored_passed = row.get("passed")
    if not isinstance(stored_passed, bool):
        _add_reason(reasons, "stored_passed_not_boolean")
    elif model_execution_failure:
        record["outcome_passed"] = False
    elif stored_passed != fresh.passed:
        _add_reason(reasons, "stored_passed_mismatch")
    else:
        record["outcome_passed"] = stored_passed
    if row.get("comparison") != fresh_dict:
        _add_reason(reasons, "stored_comparison_mismatch")

    try:
        output_sha256_after = _file_sha256(output)
    except OSError:
        _add_reason(reasons, "artifact_unreadable_after_scoring")
        return
    if output_sha256_after != output_sha256_before:
        _add_reason(reasons, "artifact_changed_during_audit")


def audit_comparison(
    results_dir: str | Path,
    tasks: Iterable[SpreadsheetTask],
    arms: Iterable[str] | None = None,
) -> dict[str, Any]:
    """Audit a comparison directory without modifying its journal or artifacts.

    Integrity is deliberately fail-closed. A selected task/arm must have either
    one freshly verified scored row, one audited recalculation infrastructure
    failure, or one exact non-replay interruption seal. Infrastructure failures
    and valid seals preserve journal integrity but invalidate score inference.
    """

    root = Path(os.path.abspath(Path(results_dir).expanduser()))
    manifest_path = root / "comparison-manifest.json"
    results_path = root / "results.jsonl"
    reasons: list[str] = []
    manifest = _load_manifest(manifest_path, reasons)
    try:
        manifest_sha256 = _file_sha256(manifest_path)
    except OSError:
        manifest_sha256 = None
    contract = _select_audit_contract(manifest, manifest_sha256, reasons)
    _audit_manifest_contract(
        manifest,
        reasons,
        results_root=root,
        manifest_sha256=manifest_sha256,
        contract=contract,
    )
    if (root / INFLIGHT_FILENAME).exists():
        _add_reason(reasons, "ambiguous_inflight_arm_task")
    task_list = list(tasks)
    task_counts = Counter(task.task_id for task in task_list)
    duplicate_task_ids = sorted(task_id for task_id, count in task_counts.items() if count > 1)
    for task_id in duplicate_task_ids:
        _add_reason(reasons, f"duplicate_input_task:{task_id}")
    unique_tasks: list[SpreadsheetTask] = []
    seen_tasks: set[str] = set()
    for task in task_list:
        if task.task_id not in seen_tasks:
            seen_tasks.add(task.task_id)
            unique_tasks.append(task)

    raw_arms = list(arms) if arms is not None else manifest.get("arms")
    if not isinstance(raw_arms, list | tuple) or not raw_arms:
        selected_arms: tuple[str, ...] = ()
        _add_reason(reasons, "audit_arms_invalid")
    else:
        selected_arms = tuple(str(arm) for arm in raw_arms)
        if len(selected_arms) != len(set(selected_arms)):
            _add_reason(reasons, "audit_arms_duplicate")
            selected_arms = tuple(dict.fromkeys(selected_arms))
    if manifest and manifest.get("arms") != list(selected_arms):
        _add_reason(reasons, "manifest_arms_mismatch")

    split_provenance = manifest.get("split_provenance")
    raw_split_manifest_id = (
        split_provenance.get("manifest_id")
        if isinstance(split_provenance, dict)
        else None
    )
    registered_run = isinstance(raw_split_manifest_id, str) and (
        raw_split_manifest_id in protected_run_spec_split_ids()
    )
    strict_registered_source = bool(
        registered_run and contract is not None and contract.strict_current_source
    )
    manifest_repository_source = manifest.get("repository_source")
    if strict_registered_source:
        current_identity = _current_repository_git_identity()
        if not isinstance(manifest_repository_source, dict):
            _add_reason(reasons, "comparison_manifest_repository_source_invalid")
        elif current_identity is None:
            _add_reason(reasons, "current_repository_git_identity_unavailable")
        elif any(
            manifest_repository_source.get(field) != value
            for field, value in current_identity.items()
        ):
            _add_reason(reasons, "comparison_manifest_repository_checkout_mismatch")
    continuation_source, continuation_source_file_sha256 = _load_continuation_source(
        root / CONTINUATION_SOURCE_FILENAME,
        manifest_sha256=manifest_sha256,
        required=registered_run,
        reasons=reasons,
        strict_current_source=bool(contract is None or contract.strict_current_source),
        expected_repository_source=manifest_repository_source,
        bind_repository_source=bool(
            contract is not None
            and contract.strict_current_source
            and (registered_run or manifest_repository_source is not None)
        ),
    )

    if not root.is_dir():
        _add_reason(reasons, "results_dir_missing")
    if root.is_symlink():
        _add_reason(reasons, "results_dir_is_symlink")

    task_manifest_reasons = _manifest_task_reasons(manifest, unique_tasks, reasons)
    raw_rows = _load_result_rows(results_path, reasons)
    rows_by_key: dict[tuple[str, str], list[dict[str, Any]]] = {}
    expected_keys = {
        (task.task_id, arm) for task in unique_tasks for arm in selected_arms
    }
    interrupted_seals, interrupted_seals_sha256 = _load_interrupted_seals(
        root / INTERRUPTED_SEALS_FILENAME,
        manifest=manifest,
        manifest_sha256=manifest_sha256,
        expected_keys=expected_keys,
        reasons=reasons,
        protocol_version=contract.protocol_version if contract is not None else None,
    )
    if (
        not results_path.exists()
        and set(interrupted_seals) == expected_keys
        and "results_file_missing" in reasons
    ):
        reasons.remove("results_file_missing")
    interrupted_keys = sorted(f"{task_id}::{arm}" for task_id, arm in interrupted_seals)
    for key in interrupted_keys:
        _add_reason(reasons, f"interrupted_unknown_outcome:{key}")
    for row_number, row in enumerate(raw_rows, start=1):
        if (
            contract is None
            or row.get("comparison_protocol_version") != contract.protocol_version
        ):
            _add_reason(reasons, f"result_protocol_mismatch:{row_number}")
        if row.get("task_id") is None or row.get("arm") is None:
            _add_reason(reasons, f"result_identity_missing:{row_number}")
            continue
        key = (str(row["task_id"]), str(row["arm"]))
        rows_by_key.setdefault(key, []).append(row)
        if key not in expected_keys:
            _add_reason(reasons, f"unexpected_result_row:{key[0]}::{key[1]}")
        if key in interrupted_seals:
            _add_reason(reasons, f"result_row_conflicts_with_interrupted_seal:{key[0]}::{key[1]}")
        row_continuation_source = row.get("continuation_source")
        if continuation_source is not None:
            legacy_row_without_continuation = (
                manifest_sha256 == LEGACY_PILOT_MANIFEST_SHA256
                and row_continuation_source is None
            )
            if (
                row_continuation_source != continuation_source
                and not legacy_row_without_continuation
            ):
                _add_reason(reasons, f"result_continuation_source_mismatch:{row_number}")
        elif row_continuation_source is not None:
            _add_reason(reasons, f"result_continuation_source_without_record:{row_number}")

    audited_rows: list[dict[str, Any]] = []
    for task in unique_tasks:
        for arm in selected_arms:
            key = (task.task_id, arm)
            candidates = rows_by_key.get(key, [])
            record: dict[str, Any] = {
                "task_id": task.task_id,
                "arm": arm,
                "audit_valid": False,
                "journal_integrity_valid": False,
                "outcome_observed": key not in interrupted_seals,
                "reasons": list(task_manifest_reasons.get(task.task_id, [])),
            }
            if key in interrupted_seals and not candidates:
                record["status"] = "interrupted"
                record["error_category"] = "interrupted_unknown_outcome"
                record["seal"] = interrupted_seals[key]
                record["journal_integrity_valid"] = not record["reasons"]
                _add_reason(record["reasons"], "interrupted_unknown_outcome")
            elif not candidates:
                _add_reason(record["reasons"], "missing_result_row")
            else:
                if len(candidates) > 1:
                    _add_reason(record["reasons"], "duplicate_result_rows")
                _audit_completed_row(
                    record,
                    candidates[0],
                    task,
                    arm,
                    root,
                    manifest,
                    manifest_sha256,
                    contract,
                )
            record["audit_valid"] = not record["reasons"]
            if key not in interrupted_seals:
                record["journal_integrity_valid"] = record["audit_valid"]
                for reason in record["reasons"]:
                    _add_reason(reasons, f"{task.task_id}::{arm}:{reason}")
            audited_rows.append(record)

    results_sha256 = None
    try:
        results_sha256 = _file_sha256(results_path)
    except OSError:
        pass
    valid_rows = sum(bool(row["audit_valid"]) for row in audited_rows)
    journal_integrity_valid = all(
        row["journal_integrity_valid"] for row in audited_rows
    ) and not any(
        not reason.startswith("interrupted_unknown_outcome:") for reason in reasons
    )
    recalculation_infrastructure_rows = sum(
        row.get("outcome_kind") == "infrastructure_failure"
        and row.get("error_category") == "recalculation_infrastructure"
        for row in audited_rows
        if row.get("outcome_observed") is True
    )
    scoring_infrastructure_rows = sum(
        row.get("outcome_kind") == "infrastructure_failure"
        and row.get("error_category") == "scoring_infrastructure"
        for row in audited_rows
        if row.get("outcome_observed") is True
    )
    provider_infrastructure_rows = sum(
        row.get("outcome_kind") == "infrastructure_failure"
        and row.get("infrastructure_failure_stage")
        == PROVIDER_INFRASTRUCTURE_FAILURE_STAGE
        and row.get("error_category") in PROVIDER_INFRASTRUCTURE_CATEGORIES
        for row in audited_rows
        if row.get("outcome_observed") is True
    )
    provider_infrastructure_reasons = Counter(
        str(row.get("provider_failure_reason") or "unspecified")
        for row in audited_rows
        if row.get("outcome_observed") is True
        and row.get("outcome_kind") == "infrastructure_failure"
        and row.get("infrastructure_failure_stage")
        == PROVIDER_INFRASTRUCTURE_FAILURE_STAGE
    )
    study_complete = bool(
        journal_integrity_valid
        and not interrupted_seals
        and not recalculation_infrastructure_rows
        and not scoring_infrastructure_rows
        and not provider_infrastructure_rows
    )
    inference_invalid_reasons: list[str] = []
    if not journal_integrity_valid:
        inference_invalid_reasons.append("comparison_audit_failed")
    else:
        if interrupted_seals:
            inference_invalid_reasons.append("interrupted_unknown_outcome")
        if recalculation_infrastructure_rows:
            inference_invalid_reasons.append("recalculation_infrastructure_failure")
        if scoring_infrastructure_rows:
            inference_invalid_reasons.append("scoring_infrastructure_failure")
        if provider_infrastructure_rows:
            inference_invalid_reasons.append("provider_infrastructure_failure")
    known_passed_rows = sum(
        row.get("outcome_passed") is True
        for row in audited_rows
        if row.get("outcome_observed") is True
    )
    known_failed_rows = sum(
        row.get("outcome_passed") is False
        for row in audited_rows
        if row.get("outcome_observed") is True
    )
    report = {
        "schema_version": 3,
        "audit_valid": journal_integrity_valid,
        "journal_integrity_valid": journal_integrity_valid,
        "study_complete": study_complete,
        "inference_valid": study_complete,
        "inference_invalid_reasons": inference_invalid_reasons,
        "reasons": reasons,
        "results_dir": str(root),
        "manifest_sha256": manifest_sha256,
        "results_sha256": results_sha256,
        "interrupted_seals_sha256": interrupted_seals_sha256,
        "continuation_source": continuation_source,
        "continuation_source_file_sha256": continuation_source_file_sha256,
        "split_provenance": manifest.get("split_provenance"),
        "run_spec_provenance": manifest.get("run_spec_provenance"),
        "task_count": len(unique_tasks),
        "arms": list(selected_arms),
        "expected_rows": len(expected_keys),
        "observed_rows": len(raw_rows),
        "valid_rows": valid_rows,
        "interrupted_arm_tasks": len(interrupted_keys),
        "interrupted_arm_task_keys": interrupted_keys,
        "known_passed_rows": known_passed_rows,
        "known_failed_rows": known_failed_rows,
        "known_model_execution_failure_rows": sum(
            row.get("outcome_kind") == "model_execution_failure"
            for row in audited_rows
            if row.get("outcome_observed") is True
        ),
        "known_recalculation_infrastructure_failure_rows": (
            recalculation_infrastructure_rows
        ),
        "known_scoring_infrastructure_failure_rows": scoring_infrastructure_rows,
        "known_provider_infrastructure_failure_rows": provider_infrastructure_rows,
        "provider_infrastructure_failure_reasons": dict(
            sorted(provider_infrastructure_reasons.items())
        ),
        "rows": audited_rows,
    }
    if not study_complete:
        report.update(
            {
                "mcnemar_exact_p": None,
                "stratified_bootstrap_95": None,
                "holm_adjusted_p": None,
            }
        )
    return report
