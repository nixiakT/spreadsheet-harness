"""Pinned SpreadsheetBench Verified adapter and clean-room value scorer."""

from __future__ import annotations

import fcntl
import hashlib
import importlib.metadata
import json
import math
import multiprocessing
import os
import platform
import re
import shutil
import tarfile
import urllib.request
import uuid
from collections import Counter, deque
from collections.abc import Iterable, Iterator
from concurrent.futures import FIRST_COMPLETED, Future, ProcessPoolExecutor, wait
from contextlib import ExitStack, contextmanager
from dataclasses import asdict, dataclass
from datetime import date, datetime, time, timezone
from pathlib import Path
from statistics import fmean
from time import monotonic
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from .agent import CONTEXT_POLICY, SpreadsheetAgent
from .config import ProviderConfig
from .errors import (
    AgentTimeoutError,
    HarnessError,
    ProviderError,
    RecalculationIntegrityError,
    RenderError,
    ScoringInfrastructureError,
)
from .pacing import PACING_POLICY, RelayPacer
from .render import openpyxl_worksheet_view
from .session import WorkbookSession
from .skills import SkillRegistry
from .tools import SpreadsheetToolRegistry

VERIFIED_REVISION = "ab0b742b0fc95b946f212d80ac7771b5531272e4"
VERIFIED_URL = (
    "https://huggingface.co/datasets/KAKA22/SpreadsheetBench/resolve/"
    f"{VERIFIED_REVISION}/spreadsheetbench_verified_400.tar.gz"
)
VERIFIED_SHA256 = "10ef893dd29cb13ab97143ea787e68cdc9574a13873ab9a54e50b31dc03fc949"
VERIFIED_DATASET_JSON_SHA256 = (
    "bcecaa89a005bd4e3bbe98da150a86e8062c27f262e575d5e47bd9861b3525e7"
)
TRACE2SKILL_HELDOUT_START = 200
TRACE2SKILL_HELDOUT_STOP = 400
TRACE2SKILL_HELDOUT_EXCLUDED_INDICES = (337, 338)
TRACE2SKILL_HELDOUT_TASK_COUNT = 198
TRACE2SKILL_HELDOUT_TASK_IDS_SHA256 = (
    "445ceec8e033601a054babf7997e340cf21d1c1d2d54a4aa421a8ba29b189582"
)
TRACE2SKILL_SPLIT_SCHEMA_VERSION = "spreadsheetbench-trace2skill-heldout-v1"
TRACE2SKILL_HELDOUT_MANIFEST_ID = "qwen35-trace2skill-heldout-v1"
TRACE2SKILL_DERIVATIVE_SPLIT_SCHEMA_VERSION = (
    "spreadsheetbench-trace2skill-derivative-v2"
)
TRACE2SKILL_PARENT_MANIFEST_FILENAME = "qwen35-trace2skill-heldout-v1.json"
TRACE2SKILL_PARENT_MANIFEST_SHA256 = (
    "e64a44f8f1a73816f215f99a38a276f2eda732a84c837f96c4079835fa8b627c"
)
TRACE2SKILL_LOCAL_UNATTEMPTED_MANIFEST_ID = (
    "qwen35-trace2skill-local-unattempted-v2"
)
TRACE2SKILL_LOCAL_UNATTEMPTED_MANIFEST_FILENAME = (
    f"{TRACE2SKILL_LOCAL_UNATTEMPTED_MANIFEST_ID}.json"
)
TRACE2SKILL_LOCAL_UNATTEMPTED_MANIFEST_SHA256 = (
    "aa12a17a65e8e60cc7678257e63d5a58f5760935ee3df1d27135b982b4de09cd"
)
TRACE2SKILL_LOCAL_UNATTEMPTED_TASK_COUNT = 143
TRACE2SKILL_LOCAL_UNATTEMPTED_TASK_IDS_SHA256 = (
    "7b76ebca59be0e97964108b5e2d0552ea6a9c0f11eb51d15c10552b82efd3386"
)
TRACE2SKILL_LOCAL_ATTEMPTED_TASK_COUNT = 51
TRACE2SKILL_LOCAL_ATTEMPTED_TASK_IDS_SHA256 = (
    "87101f52ad49e0badd47fdc73976b56cc3a29e665641e2600a69d96398a5eb2d"
)
TRACE2SKILL_LOCAL_PROTOCOL_LISTED_TASK_IDS = ("53994", "58949", "56915", "58032")
TRACE2SKILL_LOCAL_PROTOCOL_LISTED_TASK_IDS_SHA256 = (
    "7a67543e3c648ae2de672f3fddd4326e0dc5a925109eda2780a5638d491cb39d"
)
TRACE2SKILL_LOCAL_EXPOSURE_TASK_COUNT = 55
TRACE2SKILL_LOCAL_EXPOSURE_TASK_IDS_SHA256 = (
    "716101fcb8f87cf80b3c2e3d170ed43e2c9e2c0891068728a026044e26767ff3"
)
TRACE2SKILL_PILOT_MANIFEST_ID = "qwen35-trace2skill-local-unattempted-pilot16-v2"
TRACE2SKILL_PILOT_MANIFEST_FILENAME = f"{TRACE2SKILL_PILOT_MANIFEST_ID}.json"
TRACE2SKILL_PILOT_MANIFEST_SHA256 = (
    "e21c4fa091ce3c1dd23e797d98ee1bcae45eb54e1c8b2be6c3da47545241c327"
)
TRACE2SKILL_PILOT_TASK_IDS = (
    "33157",
    "35747",
    "37229",
    "46121",
    "53383",
    "53449",
    "54474",
    "56419",
    "56599",
    "58723",
    "55977",
    "56563",
    "57117",
    "57262",
    "57612",
    "59511",
)
TRACE2SKILL_PILOT_TASK_IDS_SHA256 = (
    "f25f8b75ac231f81e23e812097d3060d3e1f16597b0a5196cdeee9a833b91b82"
)
TRACE2SKILL_RESERVE_TASK_COUNT = 127
TRACE2SKILL_RESERVE_TASK_IDS_SHA256 = (
    "71c14c013bb98a1fe8d0219a5be4a784fc9aa13dcbce2419d2ec963c7457d6b7"
)
TRACE2SKILL_POSTOPT_MANIFEST_ID = "qwen35-trace2skill-local-postopt16-v1"
TRACE2SKILL_POSTOPT_MANIFEST_FILENAME = f"{TRACE2SKILL_POSTOPT_MANIFEST_ID}.json"
TRACE2SKILL_POSTOPT_MANIFEST_SHA256 = (
    "de82b9a5f17aaaf66e112f4d38938abbe9651ceab1a784ba815c82d171569c1b"
)
TRACE2SKILL_POSTOPT_SELECTION_SEED = "20260812"
TRACE2SKILL_POSTOPT_TASK_IDS = (
    "36191",
    "37456",
    "39190",
    "45944",
    "50631",
    "51354",
    "52050",
    "52532",
    "58147",
    "42902",
    "43657",
    "55260",
    "55421",
    "55976",
    "59160",
    "59358",
)
TRACE2SKILL_POSTOPT_TASK_IDS_SHA256 = (
    "a4a485d5543710352a20be947d3ac3dc251ca8fbaa32b9a0dfe571d0506b6f7a"
)
TRACE2SKILL_POSTOPT_REMAINING_TASK_COUNT = 111
TRACE2SKILL_POSTOPT_REMAINING_TASK_IDS_SHA256 = (
    "2b62d8104fe5fd65abe4fccea90f392af4fba8479a290b4fa518a9feded38a59"
)
TRACE2SKILL_POSTOPT_LAST_INCLUDED_RANK = (
    "50631",
    "2542bdd11a794d58e4cc127f40f551b09be67e611a96620be712c5b093d3f8ff",
)
TRACE2SKILL_POSTOPT_FIRST_EXCLUDED_RANK = (
    "43436",
    "2551c80e876bac09b8ecf694187f37f684af6935594de8e1b2ddca96f5623ba9",
)
TRACE2SKILL_CONFIRM_MANIFEST_ID = "qwen35-trace2skill-local-confirm16-v1"
TRACE2SKILL_CONFIRM_MANIFEST_FILENAME = f"{TRACE2SKILL_CONFIRM_MANIFEST_ID}.json"
TRACE2SKILL_CONFIRM_MANIFEST_SHA256 = (
    "c5b878de7fef5367f1e2e771f413c6724e5d4ea0c9079e9c0e99fe6feab3dc22"
)
TRACE2SKILL_CONFIRM_TASK_IDS = (
    "35739",
    "36277",
    "40959",
    "43436",
    "44266",
    "45063",
    "50683",
    "51556",
    "54717",
    "55049",
    "57232",
    "59196",
    "56427",
    "57989",
    "58904",
    "59884",
)
TRACE2SKILL_CONFIRM_TASK_IDS_SHA256 = (
    "41fef0069fb4b5c7c0e14f5ce06e8dcb504685c33c00fe620675e5669250ee11"
)
TRACE2SKILL_CONFIRM_REMAINING_TASK_COUNT = 95
TRACE2SKILL_CONFIRM_REMAINING_TASK_IDS_SHA256 = (
    "ab2d825f7dba9f2706325251bd55eaf1d433043e9c5b1614677239a6bb9b20aa"
)
TRACE2SKILL_CONFIRM_ORIGINAL_RANKS = (17, 32)
TRACE2SKILL_CONFIRM_FIRST_INCLUDED_RANK = TRACE2SKILL_POSTOPT_FIRST_EXCLUDED_RANK
TRACE2SKILL_CONFIRM_LAST_INCLUDED_RANK = (
    "59196",
    "3977e0ddd36e969835b8c8e83162dc87315dc132394a785a2c488941b3e48943",
)
TRACE2SKILL_CONFIRM_FIRST_EXCLUDED_RANK = (
    "45372",
    "3a161c5074c3900586d74620ff8711c42574073b15f27a87788cdc4081905f63",
)
TRACE2SKILL_V26_CONFIRM_MANIFEST_ID = (
    "qwen35-trace2skill-local-v26-confirm16-v1"
)
TRACE2SKILL_V26_CONFIRM_MANIFEST_FILENAME = (
    f"{TRACE2SKILL_V26_CONFIRM_MANIFEST_ID}.json"
)
TRACE2SKILL_V26_CONFIRM_MANIFEST_SHA256 = (
    "5471aadfa319948fd60e97048f5f07aa39418195770c589085d7da63ed27cb61"
)
TRACE2SKILL_V26_CONFIRM_TASK_IDS = (
    "34210",
    "37462",
    "37554",
    "44628",
    "45372",
    "50051",
    "50521",
    "52541",
    "54085",
    "54513",
    "55817",
    "57033",
    "57445",
    "57558",
    "57693",
    "59639",
)
TRACE2SKILL_V26_CONFIRM_TASK_IDS_SHA256 = (
    "f735283a19d2d464f46b10387764cc600598bb15f00a767ff4df17d154629d27"
)
TRACE2SKILL_V26_CONFIRM_REMAINING_TASK_COUNT = 79
TRACE2SKILL_V26_CONFIRM_REMAINING_TASK_IDS_SHA256 = (
    "40e4491074477ddb2bd11a0e4dc7e5513447b1e1efb90b2a169b4026fc839e7b"
)
TRACE2SKILL_V26_CONFIRM_ORIGINAL_RANKS = (33, 48)
TRACE2SKILL_V26_CONFIRM_FIRST_INCLUDED_RANK = TRACE2SKILL_CONFIRM_FIRST_EXCLUDED_RANK
TRACE2SKILL_V26_CONFIRM_LAST_INCLUDED_RANK = (
    "57033",
    "59f32fab788822c5986f82ec555c88de2446f8fbf76bf898892859a225f927d6",
)
TRACE2SKILL_V26_CONFIRM_FIRST_EXCLUDED_RANK = (
    "32789",
    "5b6dd2234d3cb84ba55fd9900a170e275b6428fd7ccccc1d6dc9e68fec02ad73",
)
TRACE2SKILL_V27_RESERVE_MANIFEST_ID = "qwen35-trace2skill-local-v27-reserve79-v1"
TRACE2SKILL_V27_RESERVE_MANIFEST_FILENAME = (
    f"{TRACE2SKILL_V27_RESERVE_MANIFEST_ID}.json"
)
TRACE2SKILL_V27_RESERVE_MANIFEST_SHA256 = (
    "28e7f5ecc4549077a8c966d2704c46fca6bc36dbad53cb2692acaf51e536105b"
)
TRACE2SKILL_V27_RESERVE_TASK_IDS = (
    "32612",
    "32789",
    "32902",
    "35742",
    "36097",
    "36764",
    "37086",
    "37378",
    "40757",
    "40892",
    "41265",
    "41348",
    "41420",
    "41589",
    "41978",
    "42181",
    "42216",
    "42354",
    "42515",
    "42930",
    "43589",
    "44296",
    "45738",
    "46897",
    "48527",
    "49857",
    "49945",
    "50250",
    "50486",
    "50811",
    "50971",
    "51249",
    "51680",
    "52220",
    "52233",
    "52305",
    "52964",
    "53117",
    "53161",
    "53647",
    "54274",
    "54638",
    "54667",
    "54925",
    "55060",
    "55085",
    "55427",
    "55965",
    "55979",
    "56378",
    "56920",
    "56921",
    "57113",
    "57590",
    "57743",
    "58484",
    "58701",
    "59595",
    "43213",
    "44017",
    "44913",
    "45707",
    "45937",
    "46646",
    "47827",
    "55468",
    "55708",
    "56786",
    "56953",
    "57354",
    "58109",
    "58499",
    "58687",
    "58942",
    "59129",
    "59224",
    "59734",
    "59794",
    "59902",
)
TRACE2SKILL_V27_RESERVE_TASK_COUNT = 79
TRACE2SKILL_V27_RESERVE_TASK_IDS_SHA256 = (
    "40e4491074477ddb2bd11a0e4dc7e5513447b1e1efb90b2a169b4026fc839e7b"
)
TRACE2SKILL_V27_REMAINING_TASK_COUNT = 0
TRACE2SKILL_V27_REMAINING_TASK_IDS_SHA256 = (
    "e3b0c44298fc1c149afbf4c8996fb92427ae41e4649b934ca495991b7852b855"
)

PROTECTED_EVALUATION_COHORTS: dict[str, tuple[str, ...]] = {
    TRACE2SKILL_PILOT_MANIFEST_ID: TRACE2SKILL_PILOT_TASK_IDS,
    TRACE2SKILL_POSTOPT_MANIFEST_ID: TRACE2SKILL_POSTOPT_TASK_IDS,
    TRACE2SKILL_CONFIRM_MANIFEST_ID: TRACE2SKILL_CONFIRM_TASK_IDS,
    TRACE2SKILL_V26_CONFIRM_MANIFEST_ID: TRACE2SKILL_V26_CONFIRM_TASK_IDS,
    TRACE2SKILL_V27_RESERVE_MANIFEST_ID: TRACE2SKILL_V27_RESERVE_TASK_IDS,
}


def require_evaluation_task_authorization(
    task_ids: Iterable[str],
    *,
    authorized_manifest_id: str | None = None,
) -> None:
    """Prevent frozen or quarantined tasks from being sampled outside their run spec."""

    selected = tuple(str(task_id) for task_id in task_ids)
    protected = {
        task_id
        for cohort in PROTECTED_EVALUATION_COHORTS.values()
        for task_id in cohort
    }
    conflicts = sorted(set(selected) & protected)
    authorized = PROTECTED_EVALUATION_COHORTS.get(str(authorized_manifest_id))
    if conflicts and (authorized is None or selected != authorized):
        raise HarnessError(
            "Protected evaluation task IDs may run only as their exact frozen cohort "
            "under its launchable registered run spec: " + ", ".join(conflicts)
        )
TRACE2SKILL_LOCAL_SCAN_REVISION = "7af635617e8f78de34cd3cdbff9fec7e373f8ba5"
TRACE2SKILL_LOCAL_SCAN_CUTOFF_UTC = "2026-08-13T15:50:24Z"
TRACE2SKILL_LOCAL_EXPOSURE_EVIDENCE_FILENAME = (
    "qwen35-trace2skill-local-exposure-evidence-v1.json"
)
TRACE2SKILL_LOCAL_EXPOSURE_EVIDENCE_SCHEMA_VERSION = (
    "spreadsheetbench-local-exposure-evidence-v1"
)
TRACE2SKILL_LOCAL_EXPOSURE_EVIDENCE_SHA256 = (
    "063dd66299cfb34a59d634d396f2d8df31b0980b29de315a33c467fb3569521e"
)
TRACE2SKILL_PILOT_FIRST_COMMITTED_REVISION = (
    "ef45aed8bcf5cccfe3e13b63c9df457926fd76d1"
)
BENCHMARK_MANIFEST_SCHEMA_VERSION = 2
BENCHMARK_PROTOCOL_VERSION = "agent_per_workbook_v2"
_PROCESS_PACERS: dict[str, RelayPacer] = {}


def _process_pacer(scope_id: str, interval_seconds: float) -> RelayPacer:
    """Reuse one pacer across every task handled by a benchmark worker process."""

    pacer = _PROCESS_PACERS.get(scope_id)
    if pacer is None:
        pacer = RelayPacer(interval_seconds)
        _PROCESS_PACERS[scope_id] = pacer
    elif pacer.interval_seconds != interval_seconds:
        raise RuntimeError("benchmark pacing scope was reused with a different interval")
    return pacer


@dataclass(frozen=True)
class SpreadsheetTask:
    task_id: str
    instruction: str
    input_path: Path
    golden_path: Path
    instruction_type: str
    answer_position: str
    answer_sheet: str | None
    protocol: str = "agent_per_workbook"
    excluded: bool = False


def _ordered_task_ids_sha256(task_ids: Iterable[str]) -> str:
    return _text_sha256("".join(f"{task_id}\n" for task_id in task_ids))


@dataclass(frozen=True)
class Comparison:
    passed: bool
    checked_cells: int
    differences: tuple[dict[str, Any], ...]

    def to_dict(self) -> dict[str, Any]:
        return {
            "passed": self.passed,
            "checked_cells": self.checked_cells,
            "differences": list(self.differences),
        }


def comparison_evidence(comparison: Comparison) -> dict[str, Any]:
    """Return evaluator evidence without copying golden values into trajectories."""

    categories: Counter[str] = Counter()
    for difference in comparison.differences:
        reasons = difference.get("reasons")
        if isinstance(reasons, list):
            categories.update(str(reason) for reason in reasons)
        elif difference.get("error"):
            categories["metadata_or_structure"] += 1
        else:
            categories["other"] += 1
    return {
        "checked_cells": comparison.checked_cells,
        "difference_count": len(comparison.differences),
        "difference_categories": dict(sorted(categories.items())),
    }


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _text_sha256(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def _source_fingerprint() -> dict[str, Any]:
    package_root = Path(__file__).resolve().parent
    repository_root = package_root.parents[1]
    files = sorted(package_root.glob("*.py"))
    pyproject = repository_root / "pyproject.toml"
    if pyproject.is_file():
        files.append(pyproject)
    entries: list[dict[str, str]] = []
    combined = hashlib.sha256()
    for path in files:
        relative = str(path.relative_to(repository_root))
        digest = _sha256(path)
        entries.append({"path": relative, "sha256": digest})
        combined.update(relative.encode("utf-8"))
        combined.update(b"\0")
        combined.update(digest.encode("ascii"))
        combined.update(b"\n")
    return {"sha256": combined.hexdigest(), "files": entries}


def _run_spec_source_fingerprint() -> dict[str, Any]:
    """Fingerprint executable source without creating a run-spec hash cycle."""

    package_root = Path(__file__).resolve().parent
    repository_root = package_root.parents[1]
    files = sorted(package_root.glob("*.py"))
    pyproject = repository_root / "pyproject.toml"
    if pyproject.is_file():
        files.append(pyproject)
    anchor_pattern = re.compile(
        rb'(?m)^(\s+sha256=)"[0-9a-f]{64}",(\s*)$'
    )
    entries: list[dict[str, str]] = []
    combined = hashlib.sha256()
    for path in files:
        relative = str(path.relative_to(repository_root))
        content = path.read_bytes()
        if relative == "src/spreadsheet_harness/comparison.py":
            content = anchor_pattern.sub(
                rb'\1"<NORMALIZED_RUN_SPEC_SHA256>",\2',
                content,
            )
        digest = hashlib.sha256(content).hexdigest()
        entries.append({"path": relative, "sha256": digest})
        combined.update(relative.encode("utf-8"))
        combined.update(b"\0")
        combined.update(digest.encode("ascii"))
        combined.update(b"\n")
    return {
        "schema_version": 1,
        "policy": "python-package-pyproject-normalized-run-spec-anchor-sha-v1",
        "sha256": combined.hexdigest(),
        "file_count": len(entries),
    }


def _runtime_fingerprint() -> dict[str, Any]:
    dependencies: dict[str, str | None] = {}
    for distribution in ("httpx", "openpyxl", "pandas", "Pillow", "PyMuPDF", "PyYAML"):
        try:
            dependencies[distribution] = importlib.metadata.version(distribution)
        except importlib.metadata.PackageNotFoundError:
            dependencies[distribution] = None
    from .render import find_libreoffice, libreoffice_version

    libreoffice = find_libreoffice()
    return {
        "python": platform.python_version(),
        "dependencies": dependencies,
        "libreoffice": libreoffice_version(libreoffice) if libreoffice else None,
    }


def _atomic_write_json(path: Path, value: dict[str, Any]) -> None:
    """Durably replace a JSON file without exposing a partially written document."""

    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp")
    try:
        with temporary.open("w", encoding="utf-8") as handle:
            os.chmod(temporary, 0o600)
            json.dump(value, handle, indent=2, ensure_ascii=False, default=str)
            handle.write("\n")
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
        directory = os.open(path.parent, os.O_RDONLY)
        try:
            os.fsync(directory)
        finally:
            os.close(directory)
    finally:
        temporary.unlink(missing_ok=True)


def _valid_jsonl_rows(path: Path) -> tuple[list[dict[str, Any]], int]:
    if not path.is_file():
        return [], 0
    rows: list[dict[str, Any]] = []
    invalid = 0
    for raw_line in path.read_bytes().splitlines():
        try:
            line = raw_line.decode("utf-8")
        except UnicodeDecodeError:
            invalid += 1
            continue
        if not line.strip():
            continue
        try:
            row = json.loads(line)
        except json.JSONDecodeError:
            invalid += 1
            continue
        if isinstance(row, dict):
            rows.append(row)
        else:
            invalid += 1
    return rows, invalid


def _repair_jsonl(path: Path) -> int:
    """Rewrite a damaged journal with every intact row before appending again."""

    if not path.is_file():
        return 0
    raw = path.read_bytes()
    rows, invalid = _valid_jsonl_rows(path)
    if invalid == 0 and (not raw or raw.endswith(b"\n")):
        return 0
    temporary = path.with_name(f".{path.name}.{uuid.uuid4().hex}.repair")
    try:
        with temporary.open("wb") as handle:
            os.chmod(temporary, 0o600)
            for row in rows:
                handle.write(
                    (json.dumps(row, ensure_ascii=False, default=str) + "\n").encode("utf-8")
                )
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
        directory = os.open(path.parent, os.O_RDONLY)
        try:
            os.fsync(directory)
        finally:
            os.close(directory)
    finally:
        temporary.unlink(missing_ok=True)
    return invalid


def _safe_extract(archive: Path, destination: Path) -> None:
    destination = destination.resolve()
    with tarfile.open(archive, "r:gz") as tar:
        members = tar.getmembers()
        for member in members:
            target = (destination / member.name).resolve()
            if target != destination and destination not in target.parents:
                raise ValueError(f"Unsafe path in dataset archive: {member.name}")
            if member.issym() or member.islnk():
                raise ValueError(f"Links are not accepted in dataset archive: {member.name}")
        tar.extractall(destination, members=members)  # noqa: S202 - validated above


def download_verified(destination: str | Path) -> Path:
    """Download and checksum the pinned Verified 400 archive."""

    destination_path = Path(destination).expanduser().resolve()
    root = destination_path / "spreadsheetbench_verified_400"
    dataset_json = root / "dataset.json"
    if dataset_json.is_file():
        return root
    destination_path.mkdir(parents=True, exist_ok=True)
    archive = destination_path / "spreadsheetbench_verified_400.tar.gz"
    partial = archive.with_name(archive.name + ".part")
    if not archive.is_file() or _sha256(archive) != VERIFIED_SHA256:
        partial.unlink(missing_ok=True)
        request = urllib.request.Request(VERIFIED_URL, headers={"User-Agent": "sheet-harness/0.1"})
        with urllib.request.urlopen(request, timeout=60) as response, partial.open("wb") as output:
            shutil.copyfileobj(response, output, length=1024 * 1024)
        if _sha256(partial) != VERIFIED_SHA256:
            partial.unlink(missing_ok=True)
            raise ValueError("SpreadsheetBench Verified archive checksum mismatch")
        partial.replace(archive)
    _safe_extract(archive, destination_path)
    if not dataset_json.is_file():
        raise FileNotFoundError(f"Archive did not contain expected dataset file: {dataset_json}")
    return root


def _find_pair(task_dir: Path, task_id: str) -> tuple[Path, Path]:
    candidates = [
        (task_dir / f"1_{task_id}_init.xlsx", task_dir / f"1_{task_id}_golden.xlsx"),
        (task_dir / "initial.xlsx", task_dir / "golden.xlsx"),
    ]
    for initial, golden in candidates:
        if initial.is_file() and golden.is_file():
            return initial, golden
    initial_matches = sorted(task_dir.glob("*_init.xlsx"))
    golden_matches = sorted(task_dir.glob("*_golden.xlsx"))
    if len(initial_matches) == 1 and len(golden_matches) == 1:
        return initial_matches[0], golden_matches[0]
    raise FileNotFoundError(f"Could not resolve init/golden pair in {task_dir}")


def load_verified_tasks(
    dataset_root: str | Path,
    *,
    include_excluded: bool = False,
    original_index_start: int | None = None,
    original_index_stop: int | None = None,
) -> list[SpreadsheetTask]:
    root = Path(dataset_root).expanduser().resolve(strict=True)
    rows = json.loads((root / "dataset.json").read_text(encoding="utf-8"))
    if not isinstance(rows, list):
        raise ValueError("Verified dataset.json must be a JSON array")
    if original_index_start is not None and original_index_start < 0:
        raise ValueError("original_index_start must be non-negative")
    if original_index_stop is not None and original_index_stop < 0:
        raise ValueError("original_index_stop must be non-negative")
    start = 0 if original_index_start is None else original_index_start
    stop = len(rows) if original_index_stop is None else original_index_stop
    if stop < start:
        raise ValueError("original_index_stop must be at least original_index_start")
    tasks: list[SpreadsheetTask] = []
    for original_index, row in enumerate(rows):
        if not start <= original_index < stop:
            continue
        excluded = bool(row.get("exclude"))
        if excluded and not include_excluded:
            continue
        task_id = str(row["id"])
        raw_path = Path(str(row.get("spreadsheet_path", f"spreadsheet/{task_id}")))
        task_dir = raw_path if raw_path.is_absolute() else root / raw_path
        if not task_dir.is_dir():
            task_dir = root / "spreadsheet" / task_id
        initial, golden = _find_pair(task_dir, task_id)
        tasks.append(
            SpreadsheetTask(
                task_id=task_id,
                instruction=str(row["instruction"]),
                input_path=initial,
                golden_path=golden,
                instruction_type=str(row.get("instruction_type", "")),
                answer_position=str(row["answer_position"]),
                answer_sheet=str(row["answer_sheet"]) if row.get("answer_sheet") else None,
                excluded=excluded,
            )
        )
    return tasks


def trace2skill_heldout_manifest(dataset_root: str | Path) -> dict[str, Any]:
    """Build a frozen split manifest from raw row indices before exclusions."""

    root = Path(dataset_root).expanduser().resolve(strict=True)
    dataset_path = root / "dataset.json"
    raw = dataset_path.read_bytes()
    dataset_hash = hashlib.sha256(raw).hexdigest()
    if dataset_hash != VERIFIED_DATASET_JSON_SHA256:
        raise ValueError(
            "Pinned SpreadsheetBench dataset.json checksum changed: "
            f"expected {VERIFIED_DATASET_JSON_SHA256}, got {dataset_hash}"
        )
    rows = json.loads(raw)
    if not isinstance(rows, list):
        raise ValueError("Verified dataset.json must be a JSON array")
    if len(rows) < TRACE2SKILL_HELDOUT_STOP:
        raise ValueError(
            f"Trace2Skill held-out split requires at least {TRACE2SKILL_HELDOUT_STOP} raw rows"
        )
    selected_rows = list(
        enumerate(
            rows[TRACE2SKILL_HELDOUT_START:TRACE2SKILL_HELDOUT_STOP],
            start=TRACE2SKILL_HELDOUT_START,
        )
    )
    excluded = [
        {
            "original_index": index,
            "task_id": str(row["id"]),
            "reason_sha256": _text_sha256(str(row.get("exclude", ""))),
        }
        for index, row in selected_rows
        if row.get("exclude")
    ]
    task_ids = [str(row["id"]) for _, row in selected_rows if not row.get("exclude")]
    split_hash = _ordered_task_ids_sha256(task_ids)
    if tuple(item["original_index"] for item in excluded) != TRACE2SKILL_HELDOUT_EXCLUDED_INDICES:
        raise ValueError(
            "Trace2Skill held-out exclusions changed; expected raw indices "
            f"{list(TRACE2SKILL_HELDOUT_EXCLUDED_INDICES)}, got "
            f"{[item['original_index'] for item in excluded]}"
        )
    if len(task_ids) != TRACE2SKILL_HELDOUT_TASK_COUNT:
        raise ValueError(
            f"Trace2Skill held-out split expected {TRACE2SKILL_HELDOUT_TASK_COUNT} usable tasks, "
            f"got {len(task_ids)}"
        )
    if split_hash != TRACE2SKILL_HELDOUT_TASK_IDS_SHA256:
        raise ValueError(
            "Trace2Skill held-out task IDs/order changed: "
            f"expected {TRACE2SKILL_HELDOUT_TASK_IDS_SHA256}, got {split_hash}"
        )
    tasks = load_verified_tasks(
        root,
        original_index_start=TRACE2SKILL_HELDOUT_START,
        original_index_stop=TRACE2SKILL_HELDOUT_STOP,
    )
    loaded_ids = [task.task_id for task in tasks]
    if loaded_ids != task_ids:
        raise ValueError("Loaded held-out tasks do not match raw-row split selection")
    return {
        "schema_version": TRACE2SKILL_SPLIT_SCHEMA_VERSION,
        "dataset": {
            "revision": VERIFIED_REVISION,
            "archive_sha256": VERIFIED_SHA256,
            "dataset_json_sha256": dataset_hash,
            "raw_row_count": len(rows),
        },
        "selection": {
            "original_index_start_inclusive": TRACE2SKILL_HELDOUT_START,
            "original_index_stop_exclusive": TRACE2SKILL_HELDOUT_STOP,
            "apply_before_exclusion_filter": True,
            "raw_rows": TRACE2SKILL_HELDOUT_STOP - TRACE2SKILL_HELDOUT_START,
            "usable_tasks": len(task_ids),
            "excluded_tasks": excluded,
        },
        "task_ids": task_ids,
        "task_ids_sha256": split_hash,
    }


def verify_trace2skill_heldout_manifest(
    dataset_root: str | Path, manifest_path: str | Path
) -> dict[str, Any]:
    """Read-only verification of a frozen held-out manifest against a dataset."""

    path, frozen, manifest_hash = _read_split_manifest(manifest_path)
    if frozen.get("schema_version") != TRACE2SKILL_SPLIT_SCHEMA_VERSION:
        raise ValueError("Expected a frozen Trace2Skill held-out v1 manifest")
    return _verify_trace2skill_heldout_document(
        dataset_root, path, frozen, manifest_hash
    )


def _reject_duplicate_json_keys(pairs: list[tuple[str, Any]]) -> dict[str, Any]:
    value: dict[str, Any] = {}
    for key, item in pairs:
        if key in value:
            raise ValueError(f"Duplicate JSON object key in split manifest: {key}")
        value[key] = item
    return value


def _read_split_manifest(
    manifest_path: str | Path,
) -> tuple[Path, dict[str, Any], str]:
    path = Path(manifest_path).expanduser().resolve(strict=True)
    raw = path.read_bytes()
    try:
        text = raw.decode("utf-8")
    except UnicodeDecodeError as exc:
        raise ValueError(f"Split manifest is not valid UTF-8: {path}") from exc
    try:
        frozen = json.loads(text, object_pairs_hook=_reject_duplicate_json_keys)
    except json.JSONDecodeError as exc:
        raise ValueError(f"Invalid split manifest JSON: {path}") from exc
    if not isinstance(frozen, dict):
        raise ValueError("Split manifest must be a JSON object")
    return path, frozen, hashlib.sha256(raw).hexdigest()


def _manifest_mismatch_fields(
    frozen: dict[str, Any], expected: dict[str, Any]
) -> list[str]:
    return sorted(
        key
        for key in set(frozen) | set(expected)
        if frozen.get(key) != expected.get(key)
    )


def _verify_trace2skill_heldout_document(
    dataset_root: str | Path,
    path: Path,
    frozen: dict[str, Any],
    manifest_hash: str,
) -> dict[str, Any]:
    if manifest_hash != TRACE2SKILL_PARENT_MANIFEST_SHA256:
        raise ValueError("Frozen held-out manifest checksum does not match its code anchor")
    expected = trace2skill_heldout_manifest(dataset_root)
    if frozen != expected:
        mismatches = _manifest_mismatch_fields(frozen, expected)
        raise ValueError(
            "Frozen Trace2Skill held-out manifest does not match dataset; fields: "
            + ", ".join(mismatches)
        )
    return {
        "valid": True,
        "manifest": str(path),
        "manifest_sha256": manifest_hash,
        "manifest_id": TRACE2SKILL_HELDOUT_MANIFEST_ID,
        "schema_version": expected["schema_version"],
        "usable_tasks": expected["selection"]["usable_tasks"],
        "task_ids": list(expected["task_ids"]),
        "task_ids_sha256": expected["task_ids_sha256"],
        "dataset_json_sha256": expected["dataset"]["dataset_json_sha256"],
    }


def trace2skill_local_unattempted_manifest(dataset_root: str | Path) -> dict[str, Any]:
    """Build the frozen local-exposure difference recorded at the v2 cutoff."""

    parent = trace2skill_heldout_manifest(dataset_root)
    parent_ids = [str(task_id) for task_id in parent["task_ids"]]
    attempted_ids = parent_ids[:TRACE2SKILL_LOCAL_ATTEMPTED_TASK_COUNT]
    if _ordered_task_ids_sha256(attempted_ids) != TRACE2SKILL_LOCAL_ATTEMPTED_TASK_IDS_SHA256:
        raise ValueError("Frozen local attempted/run task anchor no longer matches the parent")
    protocol_listed_ids = list(TRACE2SKILL_LOCAL_PROTOCOL_LISTED_TASK_IDS)
    if any(task_id not in parent_ids for task_id in protocol_listed_ids):
        raise ValueError("Frozen protocol-listed task anchor is not a parent subset")
    if _ordered_task_ids_sha256(protocol_listed_ids) != (
        TRACE2SKILL_LOCAL_PROTOCOL_LISTED_TASK_IDS_SHA256
    ):
        raise ValueError("Frozen protocol-listed task anchor hash changed")
    if set(attempted_ids) & set(protocol_listed_ids):
        raise ValueError("Frozen local exposure categories must be disjoint")
    exposed = set(attempted_ids) | set(protocol_listed_ids)
    exposed_ids = [task_id for task_id in parent_ids if task_id in exposed]
    task_ids = [task_id for task_id in parent_ids if task_id not in exposed]
    if len(exposed_ids) != TRACE2SKILL_LOCAL_EXPOSURE_TASK_COUNT or (
        _ordered_task_ids_sha256(exposed_ids) != TRACE2SKILL_LOCAL_EXPOSURE_TASK_IDS_SHA256
    ):
        raise ValueError("Frozen local exposure union no longer matches the parent")
    if len(task_ids) != TRACE2SKILL_LOCAL_UNATTEMPTED_TASK_COUNT or (
        _ordered_task_ids_sha256(task_ids)
        != TRACE2SKILL_LOCAL_UNATTEMPTED_TASK_IDS_SHA256
    ):
        raise ValueError("Frozen local-unattempted difference no longer matches the parent")
    return {
        "schema_version": TRACE2SKILL_DERIVATIVE_SPLIT_SCHEMA_VERSION,
        "manifest_id": TRACE2SKILL_LOCAL_UNATTEMPTED_MANIFEST_ID,
        "dataset": dict(parent["dataset"]),
        "parent": {
            "relative_path": TRACE2SKILL_PARENT_MANIFEST_FILENAME,
            "schema_version": TRACE2SKILL_SPLIT_SCHEMA_VERSION,
            "manifest_sha256": TRACE2SKILL_PARENT_MANIFEST_SHA256,
            "task_count": TRACE2SKILL_HELDOUT_TASK_COUNT,
            "task_ids_sha256": TRACE2SKILL_HELDOUT_TASK_IDS_SHA256,
        },
        "derivation": {
            "operation": "ordered_parent_difference",
            "classification_label": (
                "locally_unattempted_and_not_substantively_selected_as_of_freeze"
            ),
            "freeze_scope": "local repository artifacts under the recorded scan policy",
            "repository_revision_scanned": TRACE2SKILL_LOCAL_SCAN_REVISION,
            "artifact_scan_cutoff_utc": TRACE2SKILL_LOCAL_SCAN_CUTOFF_UTC,
            "scan_policy_version": "substantive-local-exposure-v1",
            "evidence_snapshot": {
                "relative_path": TRACE2SKILL_LOCAL_EXPOSURE_EVIDENCE_FILENAME,
                "schema_version": TRACE2SKILL_LOCAL_EXPOSURE_EVIDENCE_SCHEMA_VERSION,
                "sha256": TRACE2SKILL_LOCAL_EXPOSURE_EVIDENCE_SHA256,
                "claim_level": "committed_local_attestation",
                "raw_source_artifacts_available_in_fresh_clone": False,
            },
            "limitations": [
                "Administrative enumeration in full-split manifests is not treated as substantive selection.",
                "This label does not mean globally unseen or training-uncontaminated.",
                "The evidence is a frozen local attestation and is not recomputed from mutable result directories.",
            ],
            "evidence": {
                "attempted_or_run": {
                    "definition": (
                        "task ID had a local result row or task-specific run directory by cutoff"
                    ),
                    "task_ids": attempted_ids,
                    "task_count": len(attempted_ids),
                    "task_ids_sha256": _ordered_task_ids_sha256(attempted_ids),
                },
                "protocol_listed": {
                    "definition": (
                        "task ID was named in a task-specific protocol list without a local run"
                    ),
                    "task_ids": protocol_listed_ids,
                    "task_count": len(protocol_listed_ids),
                    "task_ids_sha256": _ordered_task_ids_sha256(protocol_listed_ids),
                },
                "union": {
                    "ordering": "parent_manifest_order",
                    "task_ids": exposed_ids,
                    "task_count": len(exposed_ids),
                    "task_ids_sha256": _ordered_task_ids_sha256(exposed_ids),
                },
            },
        },
        "task_ids": task_ids,
        "task_count": len(task_ids),
        "task_ids_sha256": _ordered_task_ids_sha256(task_ids),
    }


def trace2skill_local_unattempted_pilot_manifest(
    dataset_root: str | Path,
) -> dict[str, Any]:
    """Build the fixed exploratory pilot selected from the v2 local pool."""

    pool = trace2skill_local_unattempted_manifest(dataset_root)
    pool_ids = [str(task_id) for task_id in pool["task_ids"]]
    pilot_ids = list(TRACE2SKILL_PILOT_TASK_IDS)
    if pilot_ids != [task_id for task_id in pool_ids if task_id in set(pilot_ids)]:
        raise ValueError("Frozen pilot is not an ordered subset of the local-unattempted pool")
    if _ordered_task_ids_sha256(pilot_ids) != TRACE2SKILL_PILOT_TASK_IDS_SHA256:
        raise ValueError("Frozen pilot task anchor hash changed")
    reserve_ids = [task_id for task_id in pool_ids if task_id not in set(pilot_ids)]
    if len(reserve_ids) != TRACE2SKILL_RESERVE_TASK_COUNT or (
        _ordered_task_ids_sha256(reserve_ids) != TRACE2SKILL_RESERVE_TASK_IDS_SHA256
    ):
        raise ValueError("Frozen post-pilot reserve no longer matches the parent pool")
    return {
        "schema_version": TRACE2SKILL_DERIVATIVE_SPLIT_SCHEMA_VERSION,
        "manifest_id": TRACE2SKILL_PILOT_MANIFEST_ID,
        "dataset": dict(pool["dataset"]),
        "parent": {
            "relative_path": TRACE2SKILL_LOCAL_UNATTEMPTED_MANIFEST_FILENAME,
            "schema_version": TRACE2SKILL_DERIVATIVE_SPLIT_SCHEMA_VERSION,
            "manifest_sha256": TRACE2SKILL_LOCAL_UNATTEMPTED_MANIFEST_SHA256,
            "task_count": TRACE2SKILL_LOCAL_UNATTEMPTED_TASK_COUNT,
            "task_ids_sha256": TRACE2SKILL_LOCAL_UNATTEMPTED_TASK_IDS_SHA256,
        },
        "derivation": {
            "operation": "ordered_explicit_subset",
            "purpose": "exploratory_development_pilot",
            "selection": "explicit_fixed_order_before_inference",
            "selection_based_on_repository_revision": TRACE2SKILL_LOCAL_SCAN_REVISION,
            "first_committed_in_revision": TRACE2SKILL_PILOT_FIRST_COMMITTED_REVISION,
            "freeze_time_utc": TRACE2SKILL_LOCAL_SCAN_CUTOFF_UTC,
            "state_rules": [
                "Every pilot task enters development/quarantine when this manifest is frozen.",
                "A failed, timed-out, or partial request does not authorize a replacement from reserve.",
                "Observed pilot results may guide optimization only on quarantined development tasks.",
            ],
        },
        "selection": {
            "ordering": "parent_manifest_order",
            "task_count": len(pilot_ids),
            "task_ids_sha256": _ordered_task_ids_sha256(pilot_ids),
        },
        "reserve": {
            "classification_label": "locally_attested_not_yet_selected_reserve",
            "ordering": "parent_manifest_order_minus_pilot",
            "task_count": len(reserve_ids),
            "task_ids_sha256": _ordered_task_ids_sha256(reserve_ids),
            "limitations": [
                "Reserve status is local and does not imply global novelty or training isolation.",
                "Reserve status does not prove that tasks were never enumerated or preprocessed.",
                "Any later confirmatory claim requires separate justification and a new frozen manifest.",
            ],
        },
        "task_ids": pilot_ids,
        "task_count": len(pilot_ids),
        "task_ids_sha256": _ordered_task_ids_sha256(pilot_ids),
    }


def _trace2skill_postopt_rank_key(task_id: str) -> str:
    return _text_sha256(f"{TRACE2SKILL_POSTOPT_SELECTION_SEED}:{task_id}")


def trace2skill_local_postopt_manifest(
    dataset_root: str | Path,
) -> dict[str, Any]:
    """Build the deterministic post-optimization split outside the old pilot."""

    pool = trace2skill_local_unattempted_manifest(dataset_root)
    pool_ids = [str(task_id) for task_id in pool["task_ids"]]
    prior_pilot_ids = list(TRACE2SKILL_PILOT_TASK_IDS)
    if prior_pilot_ids != [
        task_id for task_id in pool_ids if task_id in set(prior_pilot_ids)
    ]:
        raise ValueError("Frozen prior pilot is not an ordered subset of the local pool")
    if _ordered_task_ids_sha256(prior_pilot_ids) != TRACE2SKILL_PILOT_TASK_IDS_SHA256:
        raise ValueError("Frozen prior pilot task anchor hash changed")
    candidates = [task_id for task_id in pool_ids if task_id not in set(prior_pilot_ids)]
    if len(candidates) != TRACE2SKILL_RESERVE_TASK_COUNT or (
        _ordered_task_ids_sha256(candidates) != TRACE2SKILL_RESERVE_TASK_IDS_SHA256
    ):
        raise ValueError("Frozen post-optimization candidate reserve changed")
    ranked = sorted(candidates, key=lambda task_id: (_trace2skill_postopt_rank_key(task_id), task_id))
    selected_set = set(ranked[: len(TRACE2SKILL_POSTOPT_TASK_IDS)])
    selected_ids = [task_id for task_id in pool_ids if task_id in selected_set]
    if tuple(selected_ids) != TRACE2SKILL_POSTOPT_TASK_IDS or (
        _ordered_task_ids_sha256(selected_ids) != TRACE2SKILL_POSTOPT_TASK_IDS_SHA256
    ):
        raise ValueError("Frozen post-optimization selection anchor changed")
    remaining_ids = [
        task_id
        for task_id in pool_ids
        if task_id not in set(prior_pilot_ids) and task_id not in selected_set
    ]
    if len(remaining_ids) != TRACE2SKILL_POSTOPT_REMAINING_TASK_COUNT or (
        _ordered_task_ids_sha256(remaining_ids)
        != TRACE2SKILL_POSTOPT_REMAINING_TASK_IDS_SHA256
    ):
        raise ValueError("Frozen post-optimization remaining reserve changed")
    rank_boundary = (
        (ranked[len(selected_ids) - 1], _trace2skill_postopt_rank_key(ranked[len(selected_ids) - 1])),
        (ranked[len(selected_ids)], _trace2skill_postopt_rank_key(ranked[len(selected_ids)])),
    )
    if rank_boundary != (
        TRACE2SKILL_POSTOPT_LAST_INCLUDED_RANK,
        TRACE2SKILL_POSTOPT_FIRST_EXCLUDED_RANK,
    ):
        raise ValueError("Frozen post-optimization rank boundary changed")
    return {
        "schema_version": TRACE2SKILL_DERIVATIVE_SPLIT_SCHEMA_VERSION,
        "manifest_id": TRACE2SKILL_POSTOPT_MANIFEST_ID,
        "dataset": dict(pool["dataset"]),
        "parent": {
            "relative_path": TRACE2SKILL_LOCAL_UNATTEMPTED_MANIFEST_FILENAME,
            "schema_version": TRACE2SKILL_DERIVATIVE_SPLIT_SCHEMA_VERSION,
            "manifest_sha256": TRACE2SKILL_LOCAL_UNATTEMPTED_MANIFEST_SHA256,
            "task_count": TRACE2SKILL_LOCAL_UNATTEMPTED_TASK_COUNT,
            "task_ids_sha256": TRACE2SKILL_LOCAL_UNATTEMPTED_TASK_IDS_SHA256,
        },
        "prior_development_pilot": {
            "relative_path": TRACE2SKILL_PILOT_MANIFEST_FILENAME,
            "schema_version": TRACE2SKILL_DERIVATIVE_SPLIT_SCHEMA_VERSION,
            "manifest_sha256": TRACE2SKILL_PILOT_MANIFEST_SHA256,
            "task_count": len(TRACE2SKILL_PILOT_TASK_IDS),
            "task_ids_sha256": TRACE2SKILL_PILOT_TASK_IDS_SHA256,
            "exclusion": "exact_task_id_set_before_ranking",
        },
        "derivation": {
            "operation": "deterministic_hash_ranked_subset_after_exact_exclusion",
            "purpose": "post_optimization_evaluation",
            "candidate_ordering": "parent_manifest_order_minus_prior_development_pilot",
            "rank_key": "sha256_utf8_seed_colon_task_id",
            "rank_seed": TRACE2SKILL_POSTOPT_SELECTION_SEED,
            "rank_ordering": "ascending_digest_hex_then_task_id",
            "selection_count": len(selected_ids),
            "output_ordering": "parent_manifest_order",
            "last_included_rank": {
                "task_id": TRACE2SKILL_POSTOPT_LAST_INCLUDED_RANK[0],
                "sha256": TRACE2SKILL_POSTOPT_LAST_INCLUDED_RANK[1],
            },
            "first_excluded_rank": {
                "task_id": TRACE2SKILL_POSTOPT_FIRST_EXCLUDED_RANK[0],
                "sha256": TRACE2SKILL_POSTOPT_FIRST_EXCLUDED_RANK[1],
            },
        },
        "candidate_pool": {
            "ordering": "parent_manifest_order_minus_prior_development_pilot",
            "task_count": len(candidates),
            "task_ids_sha256": _ordered_task_ids_sha256(candidates),
        },
        "selection": {
            "ordering": "parent_manifest_order",
            "task_count": len(selected_ids),
            "task_ids_sha256": _ordered_task_ids_sha256(selected_ids),
        },
        "remaining_reserve": {
            "ordering": "parent_manifest_order_minus_prior_pilot_and_selection",
            "task_count": len(remaining_ids),
            "task_ids_sha256": _ordered_task_ids_sha256(remaining_ids),
        },
        "task_ids": selected_ids,
        "task_count": len(selected_ids),
        "task_ids_sha256": _ordered_task_ids_sha256(selected_ids),
    }


def trace2skill_local_confirmation_manifest(
    dataset_root: str | Path,
) -> dict[str, Any]:
    """Continue the frozen hash ranking outside both quarantined cohorts."""

    pool = trace2skill_local_unattempted_manifest(dataset_root)
    pool_ids = [str(task_id) for task_id in pool["task_ids"]]
    pilot_ids = list(TRACE2SKILL_PILOT_TASK_IDS)
    postopt_ids = list(TRACE2SKILL_POSTOPT_TASK_IDS)
    pilot_set = set(pilot_ids)
    postopt_set = set(postopt_ids)
    for label, cohort_ids, cohort_hash in (
        ("pilot", pilot_ids, TRACE2SKILL_PILOT_TASK_IDS_SHA256),
        ("post-optimization evaluation", postopt_ids, TRACE2SKILL_POSTOPT_TASK_IDS_SHA256),
    ):
        if cohort_ids != [task_id for task_id in pool_ids if task_id in set(cohort_ids)]:
            raise ValueError(f"Frozen prior {label} is not an ordered subset of the local pool")
        if _ordered_task_ids_sha256(cohort_ids) != cohort_hash:
            raise ValueError(f"Frozen prior {label} task anchor hash changed")
    if pilot_set & postopt_set:
        raise ValueError("Frozen prior evaluation cohorts overlap")

    candidates = [
        task_id
        for task_id in pool_ids
        if task_id not in pilot_set and task_id not in postopt_set
    ]
    if len(candidates) != TRACE2SKILL_POSTOPT_REMAINING_TASK_COUNT or (
        _ordered_task_ids_sha256(candidates)
        != TRACE2SKILL_POSTOPT_REMAINING_TASK_IDS_SHA256
    ):
        raise ValueError("Frozen confirmation candidate reserve changed")

    original_candidates = [task_id for task_id in pool_ids if task_id not in pilot_set]
    original_ranked = sorted(
        original_candidates,
        key=lambda task_id: (_trace2skill_postopt_rank_key(task_id), task_id),
    )
    ranked = sorted(
        candidates,
        key=lambda task_id: (_trace2skill_postopt_rank_key(task_id), task_id),
    )
    if ranked != original_ranked[len(postopt_ids) :] or set(original_ranked[:16]) != postopt_set:
        raise ValueError("Frozen confirmation ranking does not continue the prior selection")

    selected_set = set(ranked[: len(TRACE2SKILL_CONFIRM_TASK_IDS)])
    selected_ids = [task_id for task_id in pool_ids if task_id in selected_set]
    if tuple(selected_ids) != TRACE2SKILL_CONFIRM_TASK_IDS or (
        _ordered_task_ids_sha256(selected_ids) != TRACE2SKILL_CONFIRM_TASK_IDS_SHA256
    ):
        raise ValueError("Frozen confirmation selection anchor changed")
    remaining_ids = [
        task_id
        for task_id in pool_ids
        if task_id not in pilot_set
        and task_id not in postopt_set
        and task_id not in selected_set
    ]
    if len(remaining_ids) != TRACE2SKILL_CONFIRM_REMAINING_TASK_COUNT or (
        _ordered_task_ids_sha256(remaining_ids)
        != TRACE2SKILL_CONFIRM_REMAINING_TASK_IDS_SHA256
    ):
        raise ValueError("Frozen confirmation remaining reserve changed")
    rank_boundary = (
        (ranked[0], _trace2skill_postopt_rank_key(ranked[0])),
        (
            ranked[len(selected_ids) - 1],
            _trace2skill_postopt_rank_key(ranked[len(selected_ids) - 1]),
        ),
        (ranked[len(selected_ids)], _trace2skill_postopt_rank_key(ranked[len(selected_ids)])),
    )
    if rank_boundary != (
        TRACE2SKILL_CONFIRM_FIRST_INCLUDED_RANK,
        TRACE2SKILL_CONFIRM_LAST_INCLUDED_RANK,
        TRACE2SKILL_CONFIRM_FIRST_EXCLUDED_RANK,
    ):
        raise ValueError("Frozen confirmation rank boundary changed")

    return {
        "schema_version": TRACE2SKILL_DERIVATIVE_SPLIT_SCHEMA_VERSION,
        "manifest_id": TRACE2SKILL_CONFIRM_MANIFEST_ID,
        "dataset": dict(pool["dataset"]),
        "parent": {
            "relative_path": TRACE2SKILL_LOCAL_UNATTEMPTED_MANIFEST_FILENAME,
            "schema_version": TRACE2SKILL_DERIVATIVE_SPLIT_SCHEMA_VERSION,
            "manifest_sha256": TRACE2SKILL_LOCAL_UNATTEMPTED_MANIFEST_SHA256,
            "task_count": TRACE2SKILL_LOCAL_UNATTEMPTED_TASK_COUNT,
            "task_ids_sha256": TRACE2SKILL_LOCAL_UNATTEMPTED_TASK_IDS_SHA256,
        },
        "prior_quarantined_cohorts": [
            {
                "relative_path": TRACE2SKILL_PILOT_MANIFEST_FILENAME,
                "schema_version": TRACE2SKILL_DERIVATIVE_SPLIT_SCHEMA_VERSION,
                "manifest_sha256": TRACE2SKILL_PILOT_MANIFEST_SHA256,
                "task_count": len(TRACE2SKILL_PILOT_TASK_IDS),
                "task_ids_sha256": TRACE2SKILL_PILOT_TASK_IDS_SHA256,
            },
            {
                "relative_path": TRACE2SKILL_POSTOPT_MANIFEST_FILENAME,
                "schema_version": TRACE2SKILL_DERIVATIVE_SPLIT_SCHEMA_VERSION,
                "manifest_sha256": TRACE2SKILL_POSTOPT_MANIFEST_SHA256,
                "task_count": len(TRACE2SKILL_POSTOPT_TASK_IDS),
                "task_ids_sha256": TRACE2SKILL_POSTOPT_TASK_IDS_SHA256,
            },
        ],
        "derivation": {
            "operation": "continue_frozen_hash_ranking_after_exact_quarantine_exclusion",
            "purpose": "post_optimization_confirmation",
            "candidate_ordering": (
                "parent_manifest_order_minus_prior_pilot_and_postopt_evaluation"
            ),
            "rank_key": "sha256_utf8_seed_colon_task_id",
            "rank_seed": TRACE2SKILL_POSTOPT_SELECTION_SEED,
            "rank_ordering": "ascending_digest_hex_then_task_id",
            "selection_count": len(selected_ids),
            "equivalent_original_candidate_ranks_inclusive": list(
                TRACE2SKILL_CONFIRM_ORIGINAL_RANKS
            ),
            "output_ordering": "parent_manifest_order",
            "last_included_rank": {
                "task_id": TRACE2SKILL_CONFIRM_LAST_INCLUDED_RANK[0],
                "sha256": TRACE2SKILL_CONFIRM_LAST_INCLUDED_RANK[1],
            },
            "first_excluded_rank": {
                "task_id": TRACE2SKILL_CONFIRM_FIRST_EXCLUDED_RANK[0],
                "sha256": TRACE2SKILL_CONFIRM_FIRST_EXCLUDED_RANK[1],
            },
        },
        "candidate_pool": {
            "ordering": "parent_manifest_order_minus_prior_pilot_and_postopt_selection",
            "task_count": len(candidates),
            "task_ids_sha256": _ordered_task_ids_sha256(candidates),
        },
        "selection": {
            "ordering": "parent_manifest_order",
            "task_count": len(selected_ids),
            "task_ids_sha256": _ordered_task_ids_sha256(selected_ids),
        },
        "remaining_reserve": {
            "ordering": "parent_manifest_order_minus_all_three_frozen_cohorts",
            "task_count": len(remaining_ids),
            "task_ids_sha256": _ordered_task_ids_sha256(remaining_ids),
        },
        "task_ids": selected_ids,
        "task_count": len(selected_ids),
        "task_ids_sha256": _ordered_task_ids_sha256(selected_ids),
    }


def trace2skill_local_v26_confirmation_manifest(
    dataset_root: str | Path,
) -> dict[str, Any]:
    """Continue the frozen ranking outside all three observed cohorts."""

    pool = trace2skill_local_unattempted_manifest(dataset_root)
    pool_ids = [str(task_id) for task_id in pool["task_ids"]]
    prior_cohorts = (
        (
            "pilot",
            list(TRACE2SKILL_PILOT_TASK_IDS),
            TRACE2SKILL_PILOT_TASK_IDS_SHA256,
        ),
        (
            "post-optimization evaluation",
            list(TRACE2SKILL_POSTOPT_TASK_IDS),
            TRACE2SKILL_POSTOPT_TASK_IDS_SHA256,
        ),
        (
            "v25 confirmation",
            list(TRACE2SKILL_CONFIRM_TASK_IDS),
            TRACE2SKILL_CONFIRM_TASK_IDS_SHA256,
        ),
    )
    prior_ids: set[str] = set()
    for label, cohort_ids, cohort_hash in prior_cohorts:
        cohort_set = set(cohort_ids)
        if cohort_ids != [task_id for task_id in pool_ids if task_id in cohort_set]:
            raise ValueError(f"Frozen prior {label} is not an ordered subset of the local pool")
        if _ordered_task_ids_sha256(cohort_ids) != cohort_hash:
            raise ValueError(f"Frozen prior {label} task anchor hash changed")
        if prior_ids & cohort_set:
            raise ValueError("Frozen prior evaluation cohorts overlap")
        prior_ids.update(cohort_set)

    candidates = [task_id for task_id in pool_ids if task_id not in prior_ids]
    if len(candidates) != TRACE2SKILL_CONFIRM_REMAINING_TASK_COUNT or (
        _ordered_task_ids_sha256(candidates)
        != TRACE2SKILL_CONFIRM_REMAINING_TASK_IDS_SHA256
    ):
        raise ValueError("Frozen v26 confirmation candidate reserve changed")

    original_candidates = [
        task_id for task_id in pool_ids if task_id not in set(TRACE2SKILL_PILOT_TASK_IDS)
    ]
    original_ranked = sorted(
        original_candidates,
        key=lambda task_id: (_trace2skill_postopt_rank_key(task_id), task_id),
    )
    ranked = sorted(
        candidates,
        key=lambda task_id: (_trace2skill_postopt_rank_key(task_id), task_id),
    )
    prior_ranked_ids = set(TRACE2SKILL_POSTOPT_TASK_IDS) | set(
        TRACE2SKILL_CONFIRM_TASK_IDS
    )
    prior_ranked_count = len(prior_ranked_ids)
    if (
        ranked != original_ranked[prior_ranked_count:]
        or set(original_ranked[:prior_ranked_count]) != prior_ranked_ids
    ):
        raise ValueError("Frozen v26 confirmation ranking does not continue prior selections")

    selected_set = set(ranked[: len(TRACE2SKILL_V26_CONFIRM_TASK_IDS)])
    selected_ids = [task_id for task_id in pool_ids if task_id in selected_set]
    if tuple(selected_ids) != TRACE2SKILL_V26_CONFIRM_TASK_IDS or (
        _ordered_task_ids_sha256(selected_ids)
        != TRACE2SKILL_V26_CONFIRM_TASK_IDS_SHA256
    ):
        raise ValueError("Frozen v26 confirmation selection anchor changed")
    remaining_ids = [
        task_id
        for task_id in pool_ids
        if task_id not in prior_ids and task_id not in selected_set
    ]
    if len(remaining_ids) != TRACE2SKILL_V26_CONFIRM_REMAINING_TASK_COUNT or (
        _ordered_task_ids_sha256(remaining_ids)
        != TRACE2SKILL_V26_CONFIRM_REMAINING_TASK_IDS_SHA256
    ):
        raise ValueError("Frozen v26 confirmation remaining reserve changed")
    rank_boundary = (
        (ranked[0], _trace2skill_postopt_rank_key(ranked[0])),
        (
            ranked[len(selected_ids) - 1],
            _trace2skill_postopt_rank_key(ranked[len(selected_ids) - 1]),
        ),
        (
            ranked[len(selected_ids)],
            _trace2skill_postopt_rank_key(ranked[len(selected_ids)]),
        ),
    )
    if rank_boundary != (
        TRACE2SKILL_V26_CONFIRM_FIRST_INCLUDED_RANK,
        TRACE2SKILL_V26_CONFIRM_LAST_INCLUDED_RANK,
        TRACE2SKILL_V26_CONFIRM_FIRST_EXCLUDED_RANK,
    ):
        raise ValueError("Frozen v26 confirmation rank boundary changed")

    prior_references = (
        (
            TRACE2SKILL_PILOT_MANIFEST_FILENAME,
            TRACE2SKILL_PILOT_MANIFEST_SHA256,
            TRACE2SKILL_PILOT_TASK_IDS,
            TRACE2SKILL_PILOT_TASK_IDS_SHA256,
        ),
        (
            TRACE2SKILL_POSTOPT_MANIFEST_FILENAME,
            TRACE2SKILL_POSTOPT_MANIFEST_SHA256,
            TRACE2SKILL_POSTOPT_TASK_IDS,
            TRACE2SKILL_POSTOPT_TASK_IDS_SHA256,
        ),
        (
            TRACE2SKILL_CONFIRM_MANIFEST_FILENAME,
            TRACE2SKILL_CONFIRM_MANIFEST_SHA256,
            TRACE2SKILL_CONFIRM_TASK_IDS,
            TRACE2SKILL_CONFIRM_TASK_IDS_SHA256,
        ),
    )
    return {
        "schema_version": TRACE2SKILL_DERIVATIVE_SPLIT_SCHEMA_VERSION,
        "manifest_id": TRACE2SKILL_V26_CONFIRM_MANIFEST_ID,
        "dataset": dict(pool["dataset"]),
        "parent": {
            "relative_path": TRACE2SKILL_LOCAL_UNATTEMPTED_MANIFEST_FILENAME,
            "schema_version": TRACE2SKILL_DERIVATIVE_SPLIT_SCHEMA_VERSION,
            "manifest_sha256": TRACE2SKILL_LOCAL_UNATTEMPTED_MANIFEST_SHA256,
            "task_count": TRACE2SKILL_LOCAL_UNATTEMPTED_TASK_COUNT,
            "task_ids_sha256": TRACE2SKILL_LOCAL_UNATTEMPTED_TASK_IDS_SHA256,
        },
        "prior_quarantined_cohorts": [
            {
                "relative_path": filename,
                "schema_version": TRACE2SKILL_DERIVATIVE_SPLIT_SCHEMA_VERSION,
                "manifest_sha256": manifest_hash,
                "task_count": len(task_ids),
                "task_ids_sha256": task_ids_hash,
            }
            for filename, manifest_hash, task_ids, task_ids_hash in prior_references
        ],
        "derivation": {
            "operation": "continue_frozen_hash_ranking_after_exact_quarantine_exclusion",
            "purpose": "v26_post_optimization_confirmation",
            "candidate_ordering": (
                "parent_manifest_order_minus_prior_pilot_postopt_and_v25_confirmation"
            ),
            "rank_key": "sha256_utf8_seed_colon_task_id",
            "rank_seed": TRACE2SKILL_POSTOPT_SELECTION_SEED,
            "rank_ordering": "ascending_digest_hex_then_task_id",
            "selection_count": len(selected_ids),
            "equivalent_original_candidate_ranks_inclusive": list(
                TRACE2SKILL_V26_CONFIRM_ORIGINAL_RANKS
            ),
            "output_ordering": "parent_manifest_order",
            "first_included_rank": {
                "task_id": TRACE2SKILL_V26_CONFIRM_FIRST_INCLUDED_RANK[0],
                "sha256": TRACE2SKILL_V26_CONFIRM_FIRST_INCLUDED_RANK[1],
            },
            "last_included_rank": {
                "task_id": TRACE2SKILL_V26_CONFIRM_LAST_INCLUDED_RANK[0],
                "sha256": TRACE2SKILL_V26_CONFIRM_LAST_INCLUDED_RANK[1],
            },
            "first_excluded_rank": {
                "task_id": TRACE2SKILL_V26_CONFIRM_FIRST_EXCLUDED_RANK[0],
                "sha256": TRACE2SKILL_V26_CONFIRM_FIRST_EXCLUDED_RANK[1],
            },
        },
        "candidate_pool": {
            "ordering": (
                "parent_manifest_order_minus_prior_pilot_postopt_and_v25_confirmation"
            ),
            "task_count": len(candidates),
            "task_ids_sha256": _ordered_task_ids_sha256(candidates),
        },
        "selection": {
            "ordering": "parent_manifest_order",
            "task_count": len(selected_ids),
            "task_ids_sha256": _ordered_task_ids_sha256(selected_ids),
        },
        "remaining_reserve": {
            "ordering": "parent_manifest_order_minus_all_four_frozen_cohorts",
            "task_count": len(remaining_ids),
            "task_ids_sha256": _ordered_task_ids_sha256(remaining_ids),
        },
        "task_ids": selected_ids,
        "task_count": len(selected_ids),
        "task_ids_sha256": _ordered_task_ids_sha256(selected_ids),
    }


def trace2skill_local_v27_reserve_manifest(
    dataset_root: str | Path,
) -> dict[str, Any]:
    """Select the entire fresh reserve outside all four frozen 16-task cohorts."""

    pool = trace2skill_local_unattempted_manifest(dataset_root)
    pool_ids = [str(task_id) for task_id in pool["task_ids"]]
    if (
        TRACE2SKILL_V27_RESERVE_TASK_COUNT
        != TRACE2SKILL_V26_CONFIRM_REMAINING_TASK_COUNT
        or TRACE2SKILL_V27_RESERVE_TASK_IDS_SHA256
        != TRACE2SKILL_V26_CONFIRM_REMAINING_TASK_IDS_SHA256
    ):
        raise ValueError("Frozen v27 selection no longer matches the v26 remaining reserve")
    prior_cohorts = (
        (
            "pilot",
            list(TRACE2SKILL_PILOT_TASK_IDS),
            TRACE2SKILL_PILOT_TASK_IDS_SHA256,
        ),
        (
            "post-optimization evaluation",
            list(TRACE2SKILL_POSTOPT_TASK_IDS),
            TRACE2SKILL_POSTOPT_TASK_IDS_SHA256,
        ),
        (
            "v25 confirmation",
            list(TRACE2SKILL_CONFIRM_TASK_IDS),
            TRACE2SKILL_CONFIRM_TASK_IDS_SHA256,
        ),
        (
            "v26 confirmation",
            list(TRACE2SKILL_V26_CONFIRM_TASK_IDS),
            TRACE2SKILL_V26_CONFIRM_TASK_IDS_SHA256,
        ),
    )
    prior_ids: set[str] = set()
    for label, cohort_ids, cohort_hash in prior_cohorts:
        cohort_set = set(cohort_ids)
        if cohort_ids != [task_id for task_id in pool_ids if task_id in cohort_set]:
            raise ValueError(f"Frozen prior {label} is not an ordered subset of the local pool")
        if _ordered_task_ids_sha256(cohort_ids) != cohort_hash:
            raise ValueError(f"Frozen prior {label} task anchor hash changed")
        if prior_ids & cohort_set:
            raise ValueError("Frozen prior evaluation cohorts overlap")
        prior_ids.update(cohort_set)

    candidates = [task_id for task_id in pool_ids if task_id not in prior_ids]
    if len(candidates) != TRACE2SKILL_V27_RESERVE_TASK_COUNT or (
        _ordered_task_ids_sha256(candidates)
        != TRACE2SKILL_V27_RESERVE_TASK_IDS_SHA256
    ):
        raise ValueError("Frozen v27 exhaustive candidate reserve changed")
    if tuple(candidates) != TRACE2SKILL_V27_RESERVE_TASK_IDS:
        raise ValueError("Frozen v27 exhaustive selection anchor changed")

    selected_ids = list(TRACE2SKILL_V27_RESERVE_TASK_IDS)
    if len(selected_ids) != TRACE2SKILL_V27_RESERVE_TASK_COUNT or (
        _ordered_task_ids_sha256(selected_ids)
        != TRACE2SKILL_V27_RESERVE_TASK_IDS_SHA256
    ):
        raise ValueError("Frozen v27 exhaustive task anchor changed")
    if selected_ids != candidates:
        raise ValueError("Frozen v27 selection must exhaust the candidate reserve")

    selected_set = set(selected_ids)
    remaining_ids = [
        task_id
        for task_id in pool_ids
        if task_id not in prior_ids and task_id not in selected_set
    ]
    if len(remaining_ids) != TRACE2SKILL_V27_REMAINING_TASK_COUNT or (
        _ordered_task_ids_sha256(remaining_ids)
        != TRACE2SKILL_V27_REMAINING_TASK_IDS_SHA256
    ):
        raise ValueError("Frozen v27 remaining reserve must be empty")
    if prior_ids & selected_set or prior_ids | selected_set != set(pool_ids):
        raise ValueError("Frozen v27 cohorts must be disjoint and exhaust the local pool")

    prior_references = (
        (
            TRACE2SKILL_PILOT_MANIFEST_FILENAME,
            TRACE2SKILL_PILOT_MANIFEST_SHA256,
            TRACE2SKILL_PILOT_TASK_IDS,
            TRACE2SKILL_PILOT_TASK_IDS_SHA256,
        ),
        (
            TRACE2SKILL_POSTOPT_MANIFEST_FILENAME,
            TRACE2SKILL_POSTOPT_MANIFEST_SHA256,
            TRACE2SKILL_POSTOPT_TASK_IDS,
            TRACE2SKILL_POSTOPT_TASK_IDS_SHA256,
        ),
        (
            TRACE2SKILL_CONFIRM_MANIFEST_FILENAME,
            TRACE2SKILL_CONFIRM_MANIFEST_SHA256,
            TRACE2SKILL_CONFIRM_TASK_IDS,
            TRACE2SKILL_CONFIRM_TASK_IDS_SHA256,
        ),
        (
            TRACE2SKILL_V26_CONFIRM_MANIFEST_FILENAME,
            TRACE2SKILL_V26_CONFIRM_MANIFEST_SHA256,
            TRACE2SKILL_V26_CONFIRM_TASK_IDS,
            TRACE2SKILL_V26_CONFIRM_TASK_IDS_SHA256,
        ),
    )
    return {
        "schema_version": TRACE2SKILL_DERIVATIVE_SPLIT_SCHEMA_VERSION,
        "manifest_id": TRACE2SKILL_V27_RESERVE_MANIFEST_ID,
        "dataset": dict(pool["dataset"]),
        "parent": {
            "relative_path": TRACE2SKILL_LOCAL_UNATTEMPTED_MANIFEST_FILENAME,
            "schema_version": TRACE2SKILL_DERIVATIVE_SPLIT_SCHEMA_VERSION,
            "manifest_sha256": TRACE2SKILL_LOCAL_UNATTEMPTED_MANIFEST_SHA256,
            "task_count": TRACE2SKILL_LOCAL_UNATTEMPTED_TASK_COUNT,
            "task_ids_sha256": TRACE2SKILL_LOCAL_UNATTEMPTED_TASK_IDS_SHA256,
        },
        "prior_quarantined_cohorts": [
            {
                "relative_path": filename,
                "schema_version": TRACE2SKILL_DERIVATIVE_SPLIT_SCHEMA_VERSION,
                "manifest_sha256": manifest_hash,
                "task_count": len(task_ids),
                "task_ids_sha256": task_ids_hash,
            }
            for filename, manifest_hash, task_ids, task_ids_hash in prior_references
        ],
        "derivation": {
            "operation": (
                "exhaustive_ordered_parent_difference_after_exact_quarantine_exclusion"
            ),
            "purpose": "v27_exhaustive_fresh_reserve_evaluation",
            "candidate_ordering": (
                "parent_manifest_order_minus_all_four_prior_frozen_cohorts"
            ),
            "selection_rule": "select_every_remaining_candidate_exactly_once",
            "selection_count": TRACE2SKILL_V27_RESERVE_TASK_COUNT,
            "output_ordering": "parent_manifest_order",
            "remaining_reserve_count": TRACE2SKILL_V27_REMAINING_TASK_COUNT,
        },
        "candidate_pool": {
            "ordering": "parent_manifest_order_minus_all_four_prior_frozen_cohorts",
            "task_count": len(candidates),
            "task_ids_sha256": _ordered_task_ids_sha256(candidates),
        },
        "selection": {
            "ordering": "parent_manifest_order",
            "task_count": len(selected_ids),
            "task_ids_sha256": _ordered_task_ids_sha256(selected_ids),
        },
        "remaining_reserve": {
            "ordering": "parent_manifest_order_minus_all_five_frozen_cohorts",
            "task_count": len(remaining_ids),
            "task_ids_sha256": _ordered_task_ids_sha256(remaining_ids),
        },
        "task_ids": selected_ids,
        "task_count": len(selected_ids),
        "task_ids_sha256": _ordered_task_ids_sha256(selected_ids),
    }


def _resolve_derivative_sibling(
    path: Path, expected_filename: str, *, artifact_label: str
) -> Path:
    sibling = path.parent / expected_filename
    resolved_directory = path.parent.resolve(strict=True)
    if sibling.is_symlink():
        raise ValueError(f"Derivative split {artifact_label} must be a sibling regular file")
    try:
        resolved_sibling = sibling.resolve(strict=True)
    except OSError as exc:
        raise ValueError(
            f"Derivative split {artifact_label} must be a sibling regular file"
        ) from exc
    if (
        resolved_sibling.parent != resolved_directory
        or resolved_sibling.name != expected_filename
        or not resolved_sibling.is_file()
    ):
        raise ValueError(f"Derivative split {artifact_label} must be a sibling regular file")
    return resolved_sibling


def _resolve_derivative_parent(path: Path, expected_filename: str) -> Path:
    return _resolve_derivative_sibling(
        path, expected_filename, artifact_label="parent"
    )


def _read_local_exposure_evidence(path: Path) -> tuple[dict[str, Any], str]:
    raw = path.read_bytes()
    evidence_hash = hashlib.sha256(raw).hexdigest()
    if evidence_hash != TRACE2SKILL_LOCAL_EXPOSURE_EVIDENCE_SHA256:
        raise ValueError("Frozen local exposure evidence checksum does not match its code anchor")
    try:
        text = raw.decode("utf-8")
    except UnicodeDecodeError as exc:
        raise ValueError("Frozen local exposure evidence is not valid UTF-8") from exc
    try:
        evidence = json.loads(text, object_pairs_hook=_reject_duplicate_json_keys)
    except json.JSONDecodeError as exc:
        raise ValueError("Invalid frozen local exposure evidence JSON") from exc
    if not isinstance(evidence, dict):
        raise ValueError("Frozen local exposure evidence must be a JSON object")
    return evidence, evidence_hash


def _verify_local_exposure_evidence(path: Path, pool: dict[str, Any]) -> None:
    derivation = pool.get("derivation")
    if not isinstance(derivation, dict):
        raise ValueError("Local pool derivation must be a JSON object")
    reference = derivation.get("evidence_snapshot")
    if not isinstance(reference, dict):
        raise ValueError("Local pool evidence snapshot reference must be a JSON object")
    relative_path = reference.get("relative_path")
    if (
        relative_path != TRACE2SKILL_LOCAL_EXPOSURE_EVIDENCE_FILENAME
        or Path(str(relative_path)).is_absolute()
    ):
        raise ValueError("Local exposure evidence path does not match its frozen sibling")
    evidence_path = _resolve_derivative_sibling(
        path,
        TRACE2SKILL_LOCAL_EXPOSURE_EVIDENCE_FILENAME,
        artifact_label="local exposure evidence",
    )
    evidence, evidence_hash = _read_local_exposure_evidence(evidence_path)
    if (
        reference.get("schema_version")
        != TRACE2SKILL_LOCAL_EXPOSURE_EVIDENCE_SCHEMA_VERSION
        or reference.get("sha256") != evidence_hash
        or reference.get("claim_level") != "committed_local_attestation"
        or reference.get("raw_source_artifacts_available_in_fresh_clone") is not False
    ):
        raise ValueError("Local exposure evidence reference does not match its anchors")

    observation = evidence.get("observation")
    parent = evidence.get("parent")
    scan_policy = evidence.get("scan_policy")
    if (
        evidence.get("schema_version")
        != TRACE2SKILL_LOCAL_EXPOSURE_EVIDENCE_SCHEMA_VERSION
        or evidence.get("claim_level") != "committed_local_attestation"
        or evidence.get("raw_source_artifacts_available_in_fresh_clone") is not False
        or not isinstance(observation, dict)
        or observation.get("cutoff_utc") != TRACE2SKILL_LOCAL_SCAN_CUTOFF_UTC
        or observation.get("source_repository_commit")
        != TRACE2SKILL_LOCAL_SCAN_REVISION
        or not isinstance(parent, dict)
        or parent.get("relative_path") != TRACE2SKILL_PARENT_MANIFEST_FILENAME
        or parent.get("sha256") != TRACE2SKILL_PARENT_MANIFEST_SHA256
        or parent.get("task_count") != TRACE2SKILL_HELDOUT_TASK_COUNT
        or parent.get("task_ids_sha256") != TRACE2SKILL_HELDOUT_TASK_IDS_SHA256
        or not isinstance(scan_policy, dict)
        or scan_policy.get("version") != "substantive-local-exposure-v1"
    ):
        raise ValueError("Frozen local exposure evidence metadata does not match its anchors")

    pool_evidence = derivation.get("evidence")
    derived = evidence.get("derived")
    if not isinstance(pool_evidence, dict) or not isinstance(derived, dict):
        raise ValueError("Frozen local exposure evidence derivation must be a JSON object")
    expected_sections = {
        "attempted_or_run": pool_evidence.get("attempted_or_run"),
        "protocol_listed_not_attempted": pool_evidence.get("protocol_listed"),
        "exposed_union_parent_order": pool_evidence.get("union"),
        "local_pool_parent_difference": pool,
    }
    for section_name, expected_section in expected_sections.items():
        section = derived.get(section_name)
        if not isinstance(section, dict) or not isinstance(expected_section, dict):
            raise ValueError(
                f"Frozen local exposure evidence section is invalid: {section_name}"
            )
        task_ids = section.get("task_ids")
        if (
            not isinstance(task_ids, list)
            or not all(isinstance(task_id, str) for task_id in task_ids)
            or section.get("task_count") != len(task_ids)
            or section.get("task_ids_sha256") != _ordered_task_ids_sha256(task_ids)
            or task_ids != expected_section.get("task_ids")
            or section.get("task_count") != expected_section.get("task_count")
            or section.get("task_ids_sha256")
            != expected_section.get("task_ids_sha256")
        ):
            raise ValueError(
                "Frozen local exposure evidence does not match the pool derivation: "
                f"{section_name}"
            )


def _verify_trace2skill_derivative_document(
    dataset_root: str | Path,
    path: Path,
    frozen: dict[str, Any],
    manifest_hash: str,
) -> dict[str, Any]:
    manifest_id = frozen.get("manifest_id")
    if manifest_id == TRACE2SKILL_LOCAL_UNATTEMPTED_MANIFEST_ID:
        expected_parent_filename = TRACE2SKILL_PARENT_MANIFEST_FILENAME
        expected_parent_hash = TRACE2SKILL_PARENT_MANIFEST_SHA256
        expected_parent_schema = TRACE2SKILL_SPLIT_SCHEMA_VERSION
        expected_parent_count = TRACE2SKILL_HELDOUT_TASK_COUNT
        expected_parent_ids_hash = TRACE2SKILL_HELDOUT_TASK_IDS_SHA256
        expected = trace2skill_local_unattempted_manifest(dataset_root)
    elif manifest_id == TRACE2SKILL_PILOT_MANIFEST_ID:
        expected_parent_filename = TRACE2SKILL_LOCAL_UNATTEMPTED_MANIFEST_FILENAME
        expected_parent_hash = TRACE2SKILL_LOCAL_UNATTEMPTED_MANIFEST_SHA256
        expected_parent_schema = TRACE2SKILL_DERIVATIVE_SPLIT_SCHEMA_VERSION
        expected_parent_count = TRACE2SKILL_LOCAL_UNATTEMPTED_TASK_COUNT
        expected_parent_ids_hash = TRACE2SKILL_LOCAL_UNATTEMPTED_TASK_IDS_SHA256
        expected = trace2skill_local_unattempted_pilot_manifest(dataset_root)
    elif manifest_id == TRACE2SKILL_POSTOPT_MANIFEST_ID:
        expected_parent_filename = TRACE2SKILL_LOCAL_UNATTEMPTED_MANIFEST_FILENAME
        expected_parent_hash = TRACE2SKILL_LOCAL_UNATTEMPTED_MANIFEST_SHA256
        expected_parent_schema = TRACE2SKILL_DERIVATIVE_SPLIT_SCHEMA_VERSION
        expected_parent_count = TRACE2SKILL_LOCAL_UNATTEMPTED_TASK_COUNT
        expected_parent_ids_hash = TRACE2SKILL_LOCAL_UNATTEMPTED_TASK_IDS_SHA256
        expected = trace2skill_local_postopt_manifest(dataset_root)
    elif manifest_id == TRACE2SKILL_CONFIRM_MANIFEST_ID:
        expected_parent_filename = TRACE2SKILL_LOCAL_UNATTEMPTED_MANIFEST_FILENAME
        expected_parent_hash = TRACE2SKILL_LOCAL_UNATTEMPTED_MANIFEST_SHA256
        expected_parent_schema = TRACE2SKILL_DERIVATIVE_SPLIT_SCHEMA_VERSION
        expected_parent_count = TRACE2SKILL_LOCAL_UNATTEMPTED_TASK_COUNT
        expected_parent_ids_hash = TRACE2SKILL_LOCAL_UNATTEMPTED_TASK_IDS_SHA256
        expected = trace2skill_local_confirmation_manifest(dataset_root)
    elif manifest_id == TRACE2SKILL_V26_CONFIRM_MANIFEST_ID:
        expected_parent_filename = TRACE2SKILL_LOCAL_UNATTEMPTED_MANIFEST_FILENAME
        expected_parent_hash = TRACE2SKILL_LOCAL_UNATTEMPTED_MANIFEST_SHA256
        expected_parent_schema = TRACE2SKILL_DERIVATIVE_SPLIT_SCHEMA_VERSION
        expected_parent_count = TRACE2SKILL_LOCAL_UNATTEMPTED_TASK_COUNT
        expected_parent_ids_hash = TRACE2SKILL_LOCAL_UNATTEMPTED_TASK_IDS_SHA256
        expected = trace2skill_local_v26_confirmation_manifest(dataset_root)
    elif manifest_id == TRACE2SKILL_V27_RESERVE_MANIFEST_ID:
        expected_parent_filename = TRACE2SKILL_LOCAL_UNATTEMPTED_MANIFEST_FILENAME
        expected_parent_hash = TRACE2SKILL_LOCAL_UNATTEMPTED_MANIFEST_SHA256
        expected_parent_schema = TRACE2SKILL_DERIVATIVE_SPLIT_SCHEMA_VERSION
        expected_parent_count = TRACE2SKILL_LOCAL_UNATTEMPTED_TASK_COUNT
        expected_parent_ids_hash = TRACE2SKILL_LOCAL_UNATTEMPTED_TASK_IDS_SHA256
        expected = trace2skill_local_v27_reserve_manifest(dataset_root)
    else:
        raise ValueError(f"Unsupported derivative split manifest_id: {manifest_id!r}")
    parent_reference = frozen.get("parent")
    if not isinstance(parent_reference, dict):
        raise ValueError("Derivative split parent reference must be a JSON object")
    relative_path = parent_reference.get("relative_path")
    if relative_path != expected_parent_filename or Path(str(relative_path)).is_absolute():
        raise ValueError("Derivative split parent path does not match its frozen sibling")
    parent_path = _resolve_derivative_parent(path, expected_parent_filename)
    parent_report = load_and_verify_trace2skill_split_manifest(dataset_root, parent_path)
    if (
        parent_report["manifest_sha256"] != expected_parent_hash
        or parent_report["schema_version"] != expected_parent_schema
        or parent_report["usable_tasks"] != expected_parent_count
        or parent_report["task_ids_sha256"] != expected_parent_ids_hash
    ):
        raise ValueError("Derivative split parent manifest checksum mismatch or invalid report")
    if manifest_id == TRACE2SKILL_POSTOPT_MANIFEST_ID:
        prior_reference = frozen.get("prior_development_pilot")
        if not isinstance(prior_reference, dict):
            raise ValueError("Post-optimization split prior pilot reference must be a JSON object")
        prior_path = prior_reference.get("relative_path")
        if (
            prior_path != TRACE2SKILL_PILOT_MANIFEST_FILENAME
            or Path(str(prior_path)).is_absolute()
        ):
            raise ValueError("Post-optimization split prior pilot path is invalid")
        prior_pilot_path = _resolve_derivative_sibling(
            path,
            TRACE2SKILL_PILOT_MANIFEST_FILENAME,
            artifact_label="prior development pilot",
        )
        prior_report = load_and_verify_trace2skill_split_manifest(
            dataset_root, prior_pilot_path
        )
        if (
            prior_report["manifest_sha256"] != TRACE2SKILL_PILOT_MANIFEST_SHA256
            or prior_report["schema_version"]
            != TRACE2SKILL_DERIVATIVE_SPLIT_SCHEMA_VERSION
            or prior_report["usable_tasks"] != len(TRACE2SKILL_PILOT_TASK_IDS)
            or prior_report["task_ids_sha256"] != TRACE2SKILL_PILOT_TASK_IDS_SHA256
        ):
            raise ValueError("Post-optimization split prior pilot sibling is invalid")
    if manifest_id in {
        TRACE2SKILL_CONFIRM_MANIFEST_ID,
        TRACE2SKILL_V26_CONFIRM_MANIFEST_ID,
        TRACE2SKILL_V27_RESERVE_MANIFEST_ID,
    }:
        prior_references = frozen.get("prior_quarantined_cohorts")
        prior_anchors: tuple[tuple[str, str, int, str, str], ...] = (
            (
                TRACE2SKILL_PILOT_MANIFEST_FILENAME,
                TRACE2SKILL_PILOT_MANIFEST_SHA256,
                len(TRACE2SKILL_PILOT_TASK_IDS),
                TRACE2SKILL_PILOT_TASK_IDS_SHA256,
                "prior development pilot",
            ),
            (
                TRACE2SKILL_POSTOPT_MANIFEST_FILENAME,
                TRACE2SKILL_POSTOPT_MANIFEST_SHA256,
                len(TRACE2SKILL_POSTOPT_TASK_IDS),
                TRACE2SKILL_POSTOPT_TASK_IDS_SHA256,
                "prior post-optimization evaluation",
            ),
        )
        if manifest_id in {
            TRACE2SKILL_V26_CONFIRM_MANIFEST_ID,
            TRACE2SKILL_V27_RESERVE_MANIFEST_ID,
        }:
            prior_anchors += (
                (
                    TRACE2SKILL_CONFIRM_MANIFEST_FILENAME,
                    TRACE2SKILL_CONFIRM_MANIFEST_SHA256,
                    len(TRACE2SKILL_CONFIRM_TASK_IDS),
                    TRACE2SKILL_CONFIRM_TASK_IDS_SHA256,
                    "prior v25 confirmation",
                ),
            )
        if manifest_id == TRACE2SKILL_V27_RESERVE_MANIFEST_ID:
            prior_anchors += (
                (
                    TRACE2SKILL_V26_CONFIRM_MANIFEST_FILENAME,
                    TRACE2SKILL_V26_CONFIRM_MANIFEST_SHA256,
                    len(TRACE2SKILL_V26_CONFIRM_TASK_IDS),
                    TRACE2SKILL_V26_CONFIRM_TASK_IDS_SHA256,
                    "prior v26 confirmation",
                ),
            )
        if not isinstance(prior_references, list) or len(prior_references) != len(
            prior_anchors
        ):
            raise ValueError(
                "Confirmation split prior cohort references must be an ordered list"
            )
        for reference, (
            expected_filename,
            expected_hash,
            expected_count,
            expected_ids_hash,
            label,
        ) in zip(prior_references, prior_anchors, strict=True):
            if not isinstance(reference, dict):
                raise ValueError(f"Confirmation split {label} reference is invalid")
            relative_path = reference.get("relative_path")
            if (
                relative_path != expected_filename
                or Path(str(relative_path)).is_absolute()
            ):
                raise ValueError(f"Confirmation split {label} path is invalid")
            sibling_path = _resolve_derivative_sibling(
                path,
                expected_filename,
                artifact_label=label,
            )
            sibling_report = load_and_verify_trace2skill_split_manifest(
                dataset_root, sibling_path
            )
            if (
                sibling_report["manifest_sha256"] != expected_hash
                or sibling_report["schema_version"]
                != TRACE2SKILL_DERIVATIVE_SPLIT_SCHEMA_VERSION
                or sibling_report["usable_tasks"] != expected_count
                or sibling_report["task_ids_sha256"] != expected_ids_hash
            ):
                raise ValueError(f"Confirmation split {label} sibling is invalid")
    if frozen != expected:
        mismatches = _manifest_mismatch_fields(frozen, expected)
        raise ValueError(
            "Frozen Trace2Skill derivative manifest does not match its anchors; fields: "
            + ", ".join(mismatches)
        )
    if manifest_id == TRACE2SKILL_LOCAL_UNATTEMPTED_MANIFEST_ID:
        _verify_local_exposure_evidence(path, frozen)
    if (
        manifest_id == TRACE2SKILL_LOCAL_UNATTEMPTED_MANIFEST_ID
        and manifest_hash != TRACE2SKILL_LOCAL_UNATTEMPTED_MANIFEST_SHA256
    ):
        raise ValueError("Frozen local pool manifest checksum does not match its code anchor")
    if (
        manifest_id == TRACE2SKILL_PILOT_MANIFEST_ID
        and manifest_hash != TRACE2SKILL_PILOT_MANIFEST_SHA256
    ):
        raise ValueError("Frozen pilot manifest checksum does not match its code anchor")
    if (
        manifest_id == TRACE2SKILL_POSTOPT_MANIFEST_ID
        and manifest_hash != TRACE2SKILL_POSTOPT_MANIFEST_SHA256
    ):
        raise ValueError("Frozen post-optimization manifest checksum does not match its code anchor")
    if (
        manifest_id == TRACE2SKILL_CONFIRM_MANIFEST_ID
        and manifest_hash != TRACE2SKILL_CONFIRM_MANIFEST_SHA256
    ):
        raise ValueError("Frozen confirmation manifest checksum does not match its code anchor")
    if (
        manifest_id == TRACE2SKILL_V26_CONFIRM_MANIFEST_ID
        and manifest_hash != TRACE2SKILL_V26_CONFIRM_MANIFEST_SHA256
    ):
        raise ValueError("Frozen v26 confirmation manifest checksum does not match its code anchor")
    if (
        manifest_id == TRACE2SKILL_V27_RESERVE_MANIFEST_ID
        and manifest_hash != TRACE2SKILL_V27_RESERVE_MANIFEST_SHA256
    ):
        raise ValueError("Frozen v27 reserve manifest checksum does not match its code anchor")
    return {
        "valid": True,
        "manifest": str(path),
        "manifest_sha256": manifest_hash,
        "manifest_id": manifest_id,
        "schema_version": expected["schema_version"],
        "usable_tasks": expected["task_count"],
        "task_ids": list(expected["task_ids"]),
        "task_ids_sha256": expected["task_ids_sha256"],
        "dataset_json_sha256": expected["dataset"]["dataset_json_sha256"],
    }


def verify_trace2skill_derivative_manifest(
    dataset_root: str | Path, manifest_path: str | Path
) -> dict[str, Any]:
    """Read-only verification of a recognized v2 derivative manifest."""

    path, frozen, manifest_hash = _read_split_manifest(manifest_path)
    if frozen.get("schema_version") != TRACE2SKILL_DERIVATIVE_SPLIT_SCHEMA_VERSION:
        raise ValueError("Expected a frozen Trace2Skill derivative v2 manifest")
    return _verify_trace2skill_derivative_document(
        dataset_root, path, frozen, manifest_hash
    )


def load_and_verify_trace2skill_split_manifest(
    dataset_root: str | Path, manifest_path: str | Path
) -> dict[str, Any]:
    """Verify a supported split exactly once and return its frozen task order."""

    path, frozen, manifest_hash = _read_split_manifest(manifest_path)
    schema = frozen.get("schema_version")
    if schema == TRACE2SKILL_SPLIT_SCHEMA_VERSION:
        return _verify_trace2skill_heldout_document(
            dataset_root, path, frozen, manifest_hash
        )
    if schema == TRACE2SKILL_DERIVATIVE_SPLIT_SCHEMA_VERSION:
        return _verify_trace2skill_derivative_document(
            dataset_root, path, frozen, manifest_hash
        )
    raise ValueError(f"Unsupported split manifest schema_version: {schema!r}")


def _trace2skill_split_provenance_anchors() -> dict[str, dict[str, Any]]:
    return {
        TRACE2SKILL_HELDOUT_MANIFEST_ID: {
            "schema_version": TRACE2SKILL_SPLIT_SCHEMA_VERSION,
            "manifest_sha256": TRACE2SKILL_PARENT_MANIFEST_SHA256,
            "task_count": TRACE2SKILL_HELDOUT_TASK_COUNT,
            "task_ids_sha256": TRACE2SKILL_HELDOUT_TASK_IDS_SHA256,
            "dataset_json_sha256": VERIFIED_DATASET_JSON_SHA256,
        },
        TRACE2SKILL_LOCAL_UNATTEMPTED_MANIFEST_ID: {
            "schema_version": TRACE2SKILL_DERIVATIVE_SPLIT_SCHEMA_VERSION,
            "manifest_sha256": TRACE2SKILL_LOCAL_UNATTEMPTED_MANIFEST_SHA256,
            "task_count": TRACE2SKILL_LOCAL_UNATTEMPTED_TASK_COUNT,
            "task_ids_sha256": TRACE2SKILL_LOCAL_UNATTEMPTED_TASK_IDS_SHA256,
            "dataset_json_sha256": VERIFIED_DATASET_JSON_SHA256,
        },
        TRACE2SKILL_PILOT_MANIFEST_ID: {
            "schema_version": TRACE2SKILL_DERIVATIVE_SPLIT_SCHEMA_VERSION,
            "manifest_sha256": TRACE2SKILL_PILOT_MANIFEST_SHA256,
            "task_count": len(TRACE2SKILL_PILOT_TASK_IDS),
            "task_ids_sha256": TRACE2SKILL_PILOT_TASK_IDS_SHA256,
            "dataset_json_sha256": VERIFIED_DATASET_JSON_SHA256,
        },
        TRACE2SKILL_POSTOPT_MANIFEST_ID: {
            "schema_version": TRACE2SKILL_DERIVATIVE_SPLIT_SCHEMA_VERSION,
            "manifest_sha256": TRACE2SKILL_POSTOPT_MANIFEST_SHA256,
            "task_count": len(TRACE2SKILL_POSTOPT_TASK_IDS),
            "task_ids_sha256": TRACE2SKILL_POSTOPT_TASK_IDS_SHA256,
            "dataset_json_sha256": VERIFIED_DATASET_JSON_SHA256,
        },
        TRACE2SKILL_CONFIRM_MANIFEST_ID: {
            "schema_version": TRACE2SKILL_DERIVATIVE_SPLIT_SCHEMA_VERSION,
            "manifest_sha256": TRACE2SKILL_CONFIRM_MANIFEST_SHA256,
            "task_count": len(TRACE2SKILL_CONFIRM_TASK_IDS),
            "task_ids_sha256": TRACE2SKILL_CONFIRM_TASK_IDS_SHA256,
            "dataset_json_sha256": VERIFIED_DATASET_JSON_SHA256,
        },
        TRACE2SKILL_V26_CONFIRM_MANIFEST_ID: {
            "schema_version": TRACE2SKILL_DERIVATIVE_SPLIT_SCHEMA_VERSION,
            "manifest_sha256": TRACE2SKILL_V26_CONFIRM_MANIFEST_SHA256,
            "task_count": len(TRACE2SKILL_V26_CONFIRM_TASK_IDS),
            "task_ids_sha256": TRACE2SKILL_V26_CONFIRM_TASK_IDS_SHA256,
            "dataset_json_sha256": VERIFIED_DATASET_JSON_SHA256,
        },
        TRACE2SKILL_V27_RESERVE_MANIFEST_ID: {
            "schema_version": TRACE2SKILL_DERIVATIVE_SPLIT_SCHEMA_VERSION,
            "manifest_sha256": TRACE2SKILL_V27_RESERVE_MANIFEST_SHA256,
            "task_count": TRACE2SKILL_V27_RESERVE_TASK_COUNT,
            "task_ids_sha256": TRACE2SKILL_V27_RESERVE_TASK_IDS_SHA256,
            "dataset_json_sha256": VERIFIED_DATASET_JSON_SHA256,
        },
    }


def verify_trace2skill_split_provenance(provenance: Any) -> bool:
    """Validate comparison provenance against immutable known split anchors."""

    if not isinstance(provenance, dict):
        return False
    required = {
        "manifest_id",
        "schema_version",
        "manifest_sha256",
        "task_count",
        "task_ids_sha256",
        "dataset_json_sha256",
    }
    if set(provenance) != required:
        return False
    anchors = _trace2skill_split_provenance_anchors()
    manifest_id = provenance.get("manifest_id")
    if not isinstance(manifest_id, str) or not isinstance(
        provenance.get("task_count"), int
    ) or isinstance(provenance.get("task_count"), bool):
        return False
    if any(
        not isinstance(provenance.get(field), str)
        for field in (
            "schema_version",
            "manifest_sha256",
            "task_ids_sha256",
            "dataset_json_sha256",
        )
    ):
        return False
    expected = anchors.get(manifest_id)
    return expected is not None and all(
        provenance.get(field) == value for field, value in expected.items()
    )


def trace2skill_split_provenance(report: Any) -> dict[str, Any]:
    """Extract and revalidate canonical provenance from a verified split report."""

    if not isinstance(report, dict) or report.get("valid") is not True:
        raise ValueError("Split provenance requires a successful verifier report")
    try:
        provenance = {
            "manifest_id": report["manifest_id"],
            "schema_version": report["schema_version"],
            "manifest_sha256": report["manifest_sha256"],
            "task_count": report["usable_tasks"],
            "task_ids_sha256": report["task_ids_sha256"],
            "dataset_json_sha256": report["dataset_json_sha256"],
        }
    except KeyError as exc:
        raise ValueError("Verified split report is missing provenance fields") from exc
    if not verify_trace2skill_split_provenance(provenance):
        raise ValueError("Verified split report does not match a known provenance anchor")
    return provenance


def _transform_official(value: Any) -> Any:
    """Match the published SpreadsheetBench value normalization."""

    if isinstance(value, bool):
        return value
    if isinstance(value, int | float):
        return round(float(value), 2)
    if isinstance(value, time):
        return str(value)[:-3]
    if isinstance(value, datetime):
        origin = datetime(1899, 12, 30)
        delta = value - origin
        return round(delta.days + delta.seconds / 86400.0, 0)
    if isinstance(value, date):
        origin_date = date(1899, 12, 30)
        return float((value - origin_date).days)
    if isinstance(value, str):
        try:
            return round(float(value), 2)
        except ValueError:
            return value
    return value


def _equal_values(expected: Any, actual: Any) -> bool:
    expected = _transform_official(expected)
    actual = _transform_official(actual)
    if expected in (None, "") and actual in (None, ""):
        return True
    if isinstance(expected, float) and isinstance(actual, float):
        if math.isnan(expected) and math.isnan(actual):
            return True
    return type(expected) is type(actual) and expected == actual


_RANGE_PATTERN = r"(?:\$?[A-Za-z]+\$?\d+(?::(?:\$?[A-Za-z]*\$?\d+))?|\$?[A-Za-z]+:\$?[A-Za-z]+)"


def _answer_sheet_names(answer_sheet: str | None, sheetnames: list[str]) -> list[str]:
    if not answer_sheet:
        return []
    cleaned = answer_sheet.strip().strip('"')
    if cleaned.strip("'") in sheetnames:
        return [cleaned.strip("'")]
    requested = [piece.strip().strip("'") for piece in cleaned.split(",")]
    return [name for name in requested if name in sheetnames]


def _split_answer_positions(
    answer_position: str,
    *,
    sheetnames: list[str],
    answer_sheet: str | None,
) -> list[tuple[str, str]]:
    """Parse dataset ranges using known sheet names to survive malformed quoting."""

    answer_position = re.sub(r"('!)'(?=\s*\$?[A-Za-z])", r"\1", answer_position)
    answer_position = re.sub(r"!'(?=\s*\$?[A-Za-z])", "'!", answer_position)
    positions: list[tuple[str, str]] = []
    masked = list(answer_position)
    alternatives = "|".join(re.escape(name) for name in sorted(sheetnames, key=len, reverse=True))
    if alternatives:
        explicit = re.compile(
            rf"'?\s*(?P<sheet>{alternatives})'?\s*!\s*(?P<range>{_RANGE_PATTERN})",
            re.IGNORECASE,
        )
        canonical = {name.casefold(): name for name in sheetnames}
        for match in explicit.finditer(answer_position):
            sheet = canonical[match.group("sheet").casefold()]
            positions.append((sheet, match.group("range")))
            for index in range(match.start(), match.end()):
                masked[index] = " "

    remaining = "".join(masked)
    unqualified = [match.group(0) for match in re.finditer(_RANGE_PATTERN, remaining)]
    targets = _answer_sheet_names(answer_sheet, sheetnames) or [sheetnames[0]]
    for cell_range in unqualified:
        positions.extend((sheet, cell_range) for sheet in targets)
    return positions


def _coordinates(range_ref: str, *, max_row_hint: int, max_column_hint: int) -> Iterable[str]:
    cleaned = range_ref.replace("$", "")
    malformed_end = re.fullmatch(r"([A-Za-z]+)(\d+):(\d+)", cleaned)
    if malformed_end:
        column, start_row, end_row = malformed_end.groups()
        cleaned = f"{column}{start_row}:{column}{end_row}"
    if ":" not in cleaned and any(char.isdigit() for char in cleaned):
        yield cleaned
        return
    min_col, min_row, max_col, max_row = range_boundaries(cleaned)
    from openpyxl.utils import get_column_letter

    if min_col is None and max_col is None:
        min_col, max_col = 1, max_column_hint
    else:
        min_col = min_col or max_col or 1
        max_col = max_col or min_col
    if min_row is None and max_row is None:
        min_row, max_row = 1, max_row_hint
    else:
        min_row = min_row or max_row or 1
        max_row = max_row or min_row
    for column in range(min_col, max_col + 1):
        for row in range(min_row, max_row + 1):
            yield f"{get_column_letter(column)}{row}"


def compare_workbooks(
    golden_path: str | Path,
    candidate_path: str | Path,
    answer_position: str,
    *,
    answer_sheet: str | None = None,
    compare_styles: bool = False,
    max_differences: int = 100,
) -> Comparison:
    """Compare answer cells using published value semantics plus optional styles."""

    golden_file = Path(golden_path)
    candidate_file = Path(candidate_path)
    if not candidate_file.is_file():
        return Comparison(False, 0, ({"error": "candidate file does not exist"},))
    expected_book = load_workbook(golden_file, data_only=True, read_only=False)
    actual_book = load_workbook(candidate_file, data_only=True, read_only=False)
    differences: list[dict[str, Any]] = []
    checked = 0
    try:
        positions = _split_answer_positions(
            answer_position,
            sheetnames=expected_book.sheetnames,
            answer_sheet=answer_sheet,
        )
        if not positions:
            differences.append(
                {"range": answer_position, "error": "answer range could not be parsed"}
            )
        for sheet, range_ref in positions:
            if sheet not in expected_book.sheetnames or sheet not in actual_book.sheetnames:
                differences.append(
                    {"sheet": sheet, "range": range_ref, "error": "worksheet missing"}
                )
                continue
            expected_sheet = expected_book[sheet]
            actual_sheet = actual_book[sheet]
            for coordinate in _coordinates(
                range_ref,
                max_row_hint=max(expected_sheet.max_row, actual_sheet.max_row),
                max_column_hint=max(expected_sheet.max_column, actual_sheet.max_column),
            ):
                checked += 1
                expected_cell = expected_sheet[coordinate]
                actual_cell = actual_sheet[coordinate]
                reasons: list[str] = []
                if not _equal_values(expected_cell.value, actual_cell.value):
                    reasons.append("value")
                if compare_styles:
                    if expected_cell.number_format != actual_cell.number_format:
                        reasons.append("number_format")
                    if expected_cell._style != actual_cell._style:
                        reasons.append("style")
                if reasons and len(differences) < max_differences:
                    differences.append(
                        {
                            "sheet": sheet,
                            "cell": coordinate,
                            "expected": repr(expected_cell.value),
                            "actual": repr(actual_cell.value),
                            "reasons": reasons,
                        }
                    )
    finally:
        expected_book.close()
        actual_book.close()
    return Comparison(not differences, checked, tuple(differences))


def compare_workbooks_chartsheet_safe(
    golden_path: str | Path,
    candidate_path: str | Path,
    answer_position: str,
    *,
    answer_sheet: str | None = None,
    compare_styles: bool = False,
    max_differences: int = 100,
) -> Comparison:
    """Run the legacy cell scorer through immutable worksheet-only OOXML views."""

    candidate_file = Path(candidate_path)
    if not candidate_file.is_file():
        return compare_workbooks(
            golden_path,
            candidate_path,
            answer_position,
            answer_sheet=answer_sheet,
            compare_styles=compare_styles,
            max_differences=max_differences,
        )

    try:
        with ExitStack() as stack:
            scoring_paths: list[Path] = []
            for raw_path in (golden_path, candidate_path):
                path = Path(raw_path)
                if path.suffix.lower() in {".xlsx", ".xlsm"}:
                    view_path, _ = stack.enter_context(openpyxl_worksheet_view(path))
                    scoring_paths.append(view_path)
                else:
                    scoring_paths.append(path)
            return compare_workbooks(
                scoring_paths[0],
                scoring_paths[1],
                answer_position,
                answer_sheet=answer_sheet,
                compare_styles=compare_styles,
                max_differences=max_differences,
            )
    except (RenderError, ScoringInfrastructureError):
        raise
    except Exception as exc:
        raise ScoringInfrastructureError(
            "The worksheet scorer could not consume a validated workbook view: "
            f"{type(exc).__name__}: {exc}"
        ) from exc


class VerifiedBenchmarkRunner:
    """Run isolated agents with bounded concurrency and resumable durable results."""

    def __init__(
        self,
        config: ProviderConfig,
        output_dir: Path,
        *,
        skill_registry: SkillRegistry | None = None,
        max_turns: int = 30,
        max_output_tokens: int = 16_000,
        enable_code: bool = True,
        recalculate: bool = True,
        workers: int = 1,
        task_timeout_seconds: float | None = 7200.0,
        task_retries: int = 1,
        circuit_breaker_threshold: int = 3,
    ) -> None:
        if workers < 1 or workers > 16:
            raise ValueError("workers must be between 1 and 16")
        if config.request_interval_seconds > 0 and workers != 1:
            raise ValueError(
                "Non-zero request pacing requires benchmark workers=1 because pacing is "
                "process-local"
            )
        if max_turns < 1:
            raise ValueError("max_turns must be positive")
        if max_output_tokens < 1:
            raise ValueError("max_output_tokens must be positive")
        if task_timeout_seconds is not None and task_timeout_seconds <= 0:
            raise ValueError("task_timeout_seconds must be positive")
        if task_retries < 0 or task_retries > 5:
            raise ValueError("task_retries must be between 0 and 5")
        if circuit_breaker_threshold < 1:
            raise ValueError("circuit_breaker_threshold must be positive")
        self.config = config
        self.output_dir = output_dir.resolve()
        self.skill_registry = skill_registry.freeze() if skill_registry is not None else None
        self.max_turns = max_turns
        self.max_output_tokens = max_output_tokens
        self.enable_code = enable_code
        self.recalculate = recalculate
        self.workers = workers
        self.task_timeout_seconds = task_timeout_seconds
        self.task_retries = task_retries
        self.circuit_breaker_threshold = circuit_breaker_threshold
        self._pacing_scope_id = uuid.uuid4().hex
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.results_path = self.output_dir / "results.jsonl"
        self.manifest_path = self.output_dir / "benchmark-manifest.json"
        self.lock_path = self.output_dir / ".benchmark.lock"
        self.recovered_invalid_rows = 0

    @contextmanager
    def _exclusive_lock(self) -> Iterator[None]:
        descriptor = os.open(self.lock_path, os.O_CREAT | os.O_RDWR, 0o600)
        try:
            try:
                fcntl.flock(descriptor, fcntl.LOCK_EX | fcntl.LOCK_NB)
            except BlockingIOError as exc:
                raise HarnessError(
                    f"Another benchmark process is already using {self.output_dir}"
                ) from exc
            yield
        finally:
            fcntl.flock(descriptor, fcntl.LOCK_UN)
            os.close(descriptor)

    def _manifest(self, tasks: list[SpreadsheetTask]) -> dict[str, Any]:
        skills = []
        if self.skill_registry is not None:
            skills = [
                {"name": skill.name, "sha256": skill.sha256}
                for skill in self.skill_registry.discover()
            ]
        return {
            "schema_version": BENCHMARK_MANIFEST_SCHEMA_VERSION,
            "benchmark_protocol_version": BENCHMARK_PROTOCOL_VERSION,
            "dataset_revision": f"KAKA22/SpreadsheetBench@{VERIFIED_REVISION}",
            "dataset_archive_sha256": VERIFIED_SHA256,
            "protocol": "agent_per_workbook",
            "task_count": len(tasks),
            "task_ids": [task.task_id for task in tasks],
            "tasks": [
                {
                    "task_id": task.task_id,
                    "instruction_sha256": _text_sha256(task.instruction),
                    "instruction_type": task.instruction_type,
                    "answer_position": task.answer_position,
                    "answer_sheet": task.answer_sheet,
                    "input_sha256": _sha256(task.input_path),
                    "golden_sha256": _sha256(task.golden_path),
                }
                for task in tasks
            ],
            "harness_source": _source_fingerprint(),
            "runtime": _runtime_fingerprint(),
            "configuration": {
                "model": self.config.model,
                "api_protocol": self.config.api_protocol,
                "requested_reasoning_effort": (
                    self.config.requested_reasoning_effort or self.config.reasoning_effort
                ),
                "reasoning_effort": self.config.reasoning_effort,
                "provider_base_url": self.config.base_url,
                "request_timeout_seconds": self.config.timeout_seconds,
                "litellm_timeout_seconds": self.config.litellm_timeout_seconds,
                "request_retries": self.config.max_retries,
                "request_interval_seconds": self.config.request_interval_seconds,
                "request_pacing_policy": PACING_POLICY,
                "request_pacing_scope": "single_worker_process",
                "request_pacing_retries_included": True,
                "request_pacing_first_attempt_immediate": True,
                "store_responses": self.config.store_responses,
                "generation": self.config.generation_dict(),
                "max_turns": self.max_turns,
                "max_output_tokens": self.max_output_tokens,
                "context_policy": CONTEXT_POLICY,
                "workers": self.workers,
                "task_timeout_seconds": self.task_timeout_seconds,
                "task_retries": self.task_retries,
                "circuit_breaker_threshold": self.circuit_breaker_threshold,
                "enable_code": self.enable_code,
                "recalculate": self.recalculate,
                "skills": skills,
            },
        }

    def _prepare_manifest(self, tasks: list[SpreadsheetTask]) -> None:
        expected = self._manifest(tasks)
        if self.manifest_path.is_file():
            try:
                actual = json.loads(self.manifest_path.read_text(encoding="utf-8"))
            except json.JSONDecodeError as exc:
                raise HarnessError(f"Invalid benchmark manifest: {self.manifest_path}") from exc
            if actual != expected:
                raise HarnessError(
                    "Refusing to resume with a different provider, task set, or run config"
                )
            return
        if self.results_path.is_file() and self.results_path.stat().st_size:
            raise HarnessError(
                f"Results exist without a compatibility manifest: {self.results_path}"
            )
        _atomic_write_json(self.manifest_path, expected)

    def _rows(self) -> list[dict[str, Any]]:
        return _valid_jsonl_rows(self.results_path)[0]

    def _latest_rows(self) -> dict[str, dict[str, Any]]:
        latest: dict[str, dict[str, Any]] = {}
        for row in self._rows():
            if row.get("task_id") is not None:
                latest[str(row["task_id"])] = row
        return latest

    def _attempt_counts(self) -> Counter[str]:
        return Counter(str(row.get("task_id")) for row in self._rows() if row.get("task_id"))

    def _final_ids(self) -> set[str]:
        attempts = self._attempt_counts()
        final: set[str] = set()
        for task_id, row in self._latest_rows().items():
            if row.get("status") == "completed":
                final.add(task_id)
            elif row.get("error_retryable") is not True:
                final.add(task_id)
            elif attempts[task_id] >= self.task_retries + 1:
                final.add(task_id)
        return final

    def _append_result(self, row: dict[str, Any]) -> None:
        encoded = (json.dumps(row, ensure_ascii=False, default=str) + "\n").encode("utf-8")
        descriptor = os.open(
            self.results_path,
            os.O_CREAT | os.O_WRONLY | os.O_APPEND,
            0o600,
        )
        try:
            offset = 0
            while offset < len(encoded):
                offset += os.write(descriptor, encoded[offset:])
            os.fsync(descriptor)
        finally:
            os.close(descriptor)

    def _task_directory(self, task_id: str, attempt: int) -> Path:
        name = task_id if attempt == 1 else f"{task_id}-retry-{attempt}"
        path = self.output_dir / "runs" / name
        if path.exists():
            path = path.with_name(f"{name}-{uuid.uuid4().hex[:8]}")
        return path

    def _run_task(self, task: SpreadsheetTask, attempt: int) -> dict[str, Any]:
        started_at = datetime.now(timezone.utc)
        started_clock = monotonic()
        task_dir = self._task_directory(task.task_id, attempt)
        row: dict[str, Any] = {
            "task_id": task.task_id,
            "task_attempt": attempt,
            "instruction_type": task.instruction_type,
            "answer_position": task.answer_position,
            "protocol": task.protocol,
            "benchmark_protocol_version": BENCHMARK_PROTOCOL_VERSION,
            "model": self.config.model,
            "api_protocol": self.config.api_protocol,
            "requested_reasoning_effort": (
                self.config.requested_reasoning_effort or self.config.reasoning_effort
            ),
            "reasoning_effort": self.config.reasoning_effort,
            "provider_base_url": self.config.base_url,
            "request_timeout_seconds": self.config.timeout_seconds,
            "litellm_timeout_seconds": self.config.litellm_timeout_seconds,
            "request_retries": self.config.max_retries,
            "request_interval_seconds": self.config.request_interval_seconds,
            "generation": self.config.generation_dict(),
            "request_pacing_scope": "single_worker_process",
            "max_turns": self.max_turns,
            "max_output_tokens": self.max_output_tokens,
            "task_timeout_seconds": self.task_timeout_seconds,
            "calculation_backend": "libreoffice" if self.recalculate else "not_recalculated",
            "run_dir": str(task_dir),
            "started_at": started_at.isoformat(),
        }
        session: WorkbookSession | None = None
        agent_result: Any | None = None
        recalc_metadata: dict[str, Any] | None = None
        try:
            session = WorkbookSession.create(
                task.input_path,
                task_dir,
                run_id=task.task_id,
                recorder_secrets=(self.config.api_key,),
            )
            tools = SpreadsheetToolRegistry(
                session,
                enable_code=self.enable_code,
                redaction_secrets=(self.config.api_key,),
            )
            pacer = _process_pacer(
                self._pacing_scope_id, self.config.request_interval_seconds
            )
            session.recorder.record(
                "benchmark.configured",
                {
                    "schema_version": BENCHMARK_MANIFEST_SCHEMA_VERSION,
                    "benchmark_protocol_version": BENCHMARK_PROTOCOL_VERSION,
                    "request_interval_seconds": self.config.request_interval_seconds,
                    "litellm_timeout_seconds": self.config.litellm_timeout_seconds,
                    "api_protocol": self.config.api_protocol,
                    "generation": self.config.generation_dict(),
                    "request_pacing_policy": PACING_POLICY,
                    "request_pacing_scope": "single_worker_process",
                    "max_turns": self.max_turns,
                    "max_output_tokens_per_call": self.max_output_tokens,
                    "task_timeout_seconds": self.task_timeout_seconds,
                },
            )
            agent = SpreadsheetAgent(
                self.config,
                tools,
                skills=self.skill_registry,
                max_turns=self.max_turns,
                max_output_tokens=self.max_output_tokens,
                max_elapsed_seconds=self.task_timeout_seconds,
                pacer=pacer,
            )
            agent_result = agent.run(task.instruction)
            if self.recalculate:
                from .render import recalculate_workbook

                recalc_metadata = recalculate_workbook(session.workbook_path, session.workbook_path)
            comparison = compare_workbooks_chartsheet_safe(
                task.golden_path,
                session.workbook_path,
                task.answer_position,
                answer_sheet=task.answer_sheet,
            )
            row.update(
                {
                    "status": "completed",
                    "passed": comparison.passed,
                    "comparison": comparison.to_dict(),
                    "agent": agent_result.to_dict(),
                    "recalculation": recalc_metadata,
                    "output_workbook": str(session.workbook_path),
                    "output_sha256": _sha256(session.workbook_path),
                }
            )
            session.recorder.record(
                "benchmark.evaluated",
                {
                    "task_id": task.task_id,
                    "passed": comparison.passed,
                    "status": "completed",
                    "scorer": "cleanroom-corrected-value-v1",
                    "style_checked": False,
                    "calculation_backend": row["calculation_backend"],
                    **comparison_evidence(comparison),
                    "scoring_metadata_sha256": _text_sha256(
                        json.dumps(
                            {
                                "answer_position": task.answer_position,
                                "answer_sheet": task.answer_sheet,
                            },
                            ensure_ascii=False,
                            sort_keys=True,
                            separators=(",", ":"),
                        )
                    ),
                },
            )
        except Exception as exc:
            safe_error = str(exc).replace(self.config.api_key, "[REDACTED]")
            row.update(
                {
                    "status": "error",
                    "passed": False,
                    "error": safe_error,
                    "error_type": type(exc).__name__,
                    "error_retryable": False,
                    "error_category": "harness",
                }
            )
            if isinstance(exc, ProviderError):
                row.update(
                    {
                        "error_retryable": bool(exc.safe_to_retry),
                        "error_category": (
                            "provider_transient"
                            if exc.retryable
                            else "provider_fatal"
                            if exc.global_fatal
                            else "provider_task"
                        ),
                        "provider_error": exc.public_dict(secrets=(self.config.api_key,)),
                    }
                )
            elif isinstance(exc, AgentTimeoutError):
                row["error_category"] = "task_timeout"
            elif isinstance(exc, RecalculationIntegrityError):
                row.update(
                    {
                        "outcome_kind": "infrastructure_failure",
                        "score_available": False,
                        "error_category": "recalculation_infrastructure",
                        "infrastructure_failure_stage": "recalculation",
                        "recalculation_failure_reason": "sheet_inventory_changed",
                        "agent": agent_result.to_dict(),
                        "recalculation": exc.evidence,
                        "output_workbook": str(session.workbook_path),
                        "output_sha256": _sha256(session.workbook_path),
                    }
                )
            elif isinstance(exc, ScoringInfrastructureError):
                if agent_result is None or session is None:
                    raise HarnessError(
                        "Scoring infrastructure failure omitted auditable run evidence"
                    ) from exc
                row.update(
                    {
                        "outcome_kind": "infrastructure_failure",
                        "score_available": False,
                        "error_category": "scoring_infrastructure",
                        "infrastructure_failure_stage": "scoring",
                        "scoring_failure_reason": "worksheet_scorer_unsupported",
                        "agent": agent_result.to_dict(),
                        "recalculation": recalc_metadata,
                        "output_workbook": str(session.workbook_path),
                        "output_sha256": _sha256(session.workbook_path),
                    }
                )
        finally:
            row["finished_at"] = datetime.now(timezone.utc).isoformat()
            row["elapsed_seconds"] = round(monotonic() - started_clock, 3)
            if session is not None:
                if row.get("status") != "completed":
                    try:
                        session.recorder.record(
                            "benchmark.not_evaluated",
                            {
                                "task_id": task.task_id,
                                "status": str(row.get("status", "error")),
                                "scorer": "cleanroom-corrected-value-v1",
                                "style_checked": False,
                                "calculation_backend": row["calculation_backend"],
                                "error_category": row.get("error_category"),
                                "outcome_kind": row.get("outcome_kind"),
                                "score_available": row.get("score_available"),
                                "infrastructure_failure_stage": row.get(
                                    "infrastructure_failure_stage"
                                ),
                                "recalculation_failure_reason": row.get(
                                    "recalculation_failure_reason"
                                ),
                                "scoring_failure_reason": row.get(
                                    "scoring_failure_reason"
                                ),
                                "recalculation": row.get("recalculation"),
                            },
                        )
                    except Exception:
                        pass
                task_manifest = asdict(task)
                task_manifest["input_path"] = str(task.input_path)
                task_manifest["golden_path"] = str(task.golden_path)
                try:
                    session.write_manifest({"task": task_manifest, "result": row})
                except Exception:
                    pass
        return row

    def run(self, tasks: Iterable[SpreadsheetTask], *, resume: bool = True) -> dict[str, Any]:
        task_list = list(tasks)
        task_ids = [task.task_id for task in task_list]
        if len(task_ids) != len(set(task_ids)):
            raise ValueError("Benchmark task IDs must be unique")
        require_evaluation_task_authorization(task_ids)
        with self._exclusive_lock():
            self.recovered_invalid_rows = _repair_jsonl(self.results_path)
            self._prepare_manifest(task_list)
            if not resume and self.results_path.is_file() and self.results_path.stat().st_size:
                raise HarnessError("--no-resume requires an empty output directory")
            attempts = self._attempt_counts()
            final_ids = self._final_ids() if resume else set()
            pending: deque[tuple[SpreadsheetTask, int]] = deque(
                (task, attempts[task.task_id] + 1)
                for task in task_list
                if task.task_id not in final_ids
            )
            finalized = set(final_ids)
            exhausted_transient_tasks = 0
            circuit_breaker_tripped = False
            futures: dict[Future[dict[str, Any]], tuple[SpreadsheetTask, int]] = {}

            # Process isolation keeps LibreOffice and the local code interpreter
            # independent. In particular, code execution uses POSIX resource
            # limits in a child process, which is unsafe to fork from worker threads.
            with ProcessPoolExecutor(
                max_workers=self.workers,
                mp_context=multiprocessing.get_context("spawn"),
            ) as executor:
                while pending or futures:
                    while (
                        pending
                        and len(futures) < self.workers
                        and not circuit_breaker_tripped
                    ):
                        task, attempt = pending.popleft()
                        future = executor.submit(self._run_task, task, attempt)
                        futures[future] = (task, attempt)
                    if not futures:
                        break
                    done, _ = wait(futures, return_when=FIRST_COMPLETED)
                    for future in done:
                        task, attempt = futures.pop(future)
                        try:
                            row = future.result()
                        except Exception as exc:
                            safe_error = str(exc).replace(self.config.api_key, "[REDACTED]")
                            now = datetime.now(timezone.utc).isoformat()
                            row = {
                                "task_id": task.task_id,
                                "task_attempt": attempt,
                                "instruction_type": task.instruction_type,
                                "answer_position": task.answer_position,
                                "protocol": task.protocol,
                                "benchmark_protocol_version": BENCHMARK_PROTOCOL_VERSION,
                                "model": self.config.model,
                                "requested_reasoning_effort": (
                                    self.config.requested_reasoning_effort
                                    or self.config.reasoning_effort
                                ),
                                "reasoning_effort": self.config.reasoning_effort,
                                "provider_base_url": self.config.base_url,
                                "request_timeout_seconds": self.config.timeout_seconds,
                                "litellm_timeout_seconds": (
                                    self.config.litellm_timeout_seconds
                                ),
                                "request_retries": self.config.max_retries,
                                "request_interval_seconds": (
                                    self.config.request_interval_seconds
                                ),
                                "request_pacing_scope": "single_worker_process",
                                "max_turns": self.max_turns,
                                "max_output_tokens": self.max_output_tokens,
                                "task_timeout_seconds": self.task_timeout_seconds,
                                "calculation_backend": (
                                    "libreoffice"
                                    if self.recalculate
                                    else "not_recalculated"
                                ),
                                "status": "error",
                                "passed": False,
                                "error": f"Worker process failed: {safe_error}",
                                "error_type": type(exc).__name__,
                                "error_retryable": False,
                                "error_category": "worker_process",
                                "started_at": now,
                                "finished_at": now,
                                "elapsed_seconds": 0.0,
                            }
                        self._append_result(row)

                        category = row.get("error_category")
                        retry_eligible = (
                            row.get("status") == "error"
                            and row.get("error_retryable") is True
                            and attempt < self.task_retries + 1
                        )
                        if category == "provider_transient" and not retry_eligible:
                            exhausted_transient_tasks += 1
                        if category in {
                            "provider_fatal",
                            "recalculation_infrastructure",
                            "scoring_infrastructure",
                        } or (
                            exhausted_transient_tasks >= self.circuit_breaker_threshold
                        ):
                            circuit_breaker_tripped = True
                        should_retry = retry_eligible and not circuit_breaker_tripped
                        if should_retry:
                            pending.appendleft((task, attempt + 1))
                        else:
                            finalized.add(task.task_id)

                        progress = {
                            "event": "benchmark.task_finished",
                            "task_id": task.task_id,
                            "task_attempt": attempt,
                            "status": row.get("status"),
                            "passed": row.get("passed"),
                            "error_category": category,
                            "elapsed_seconds": row.get("elapsed_seconds"),
                            "finalized": len(finalized),
                            "expected": len(task_list),
                            "retry_eligible": retry_eligible,
                            "retry_queued": should_retry,
                            "circuit_breaker_tripped": circuit_breaker_tripped,
                        }
                        print(json.dumps(progress, ensure_ascii=False), flush=True)

            summary = summarize_results(
                self.results_path,
                expected_task_ids=task_ids,
                write_summary=False,
            )
            summary.update(
                {
                    "dataset_revision": f"KAKA22/SpreadsheetBench@{VERIFIED_REVISION}",
                    "selected_task_count": len(task_list),
                    "circuit_breaker_tripped": circuit_breaker_tripped,
                    "exhausted_transient_tasks": exhausted_transient_tasks,
                    "recovered_invalid_result_rows": self.recovered_invalid_rows,
                }
            )
            _atomic_write_json(self.output_dir / "summary.json", summary)
            return summary


def summarize_results(
    results_path: str | Path,
    *,
    expected_task_ids: Iterable[str] | None = None,
    write_summary: bool = True,
) -> dict[str, Any]:
    path = Path(results_path)
    comparison_manifest_path = path.with_name("comparison-manifest.json")
    if comparison_manifest_path.is_file():
        raise HarnessError(
            "Refusing to summarize comparison results as a single-arm benchmark: "
            f"{comparison_manifest_path}"
        )
    rows_from_disk, invalid_rows = _valid_jsonl_rows(path)
    latest: dict[str, dict[str, Any]] = {}
    anonymous = 0
    for row in rows_from_disk:
        task_id = row.get("task_id")
        if task_id is None:
            anonymous += 1
            task_id = f"__anonymous_{anonymous}"
        latest[str(task_id)] = row
    if expected_task_ids is None:
        manifest_path = path.with_name("benchmark-manifest.json")
        if manifest_path.is_file():
            try:
                manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
                expected_task_ids = [str(item) for item in manifest.get("task_ids", [])]
            except (json.JSONDecodeError, TypeError):
                expected_task_ids = None
    expected = list(dict.fromkeys(str(item) for item in (expected_task_ids or latest.keys())))
    rows = [latest[task_id] for task_id in expected if task_id in latest]
    completed = [row for row in rows if row.get("status") == "completed"]
    no_score_infrastructure = [
        row
        for row in rows
        if row.get("outcome_kind") == "infrastructure_failure"
        and row.get("score_available") is False
    ]
    recalculation_infrastructure = [
        row
        for row in no_score_infrastructure
        if row.get("error_category") == "recalculation_infrastructure"
    ]
    scoring_infrastructure = [
        row
        for row in no_score_infrastructure
        if row.get("error_category") == "scoring_infrastructure"
    ]
    scores = [1.0 if row.get("passed") else 0.0 for row in completed]
    passed = int(sum(scores))
    denominator = len(expected)
    attempted_score = passed / denominator if denominator else 0.0
    inference_valid = not no_score_infrastructure
    primary_score = attempted_score if inference_valid else None
    completed_score = fmean(scores) if scores else 0.0
    inference_invalid_reasons: list[str] = []
    if recalculation_infrastructure:
        inference_invalid_reasons.append("recalculation_infrastructure_failures")
    if scoring_infrastructure:
        inference_invalid_reasons.append("scoring_infrastructure_failures")
    if no_score_infrastructure and not inference_invalid_reasons:
        inference_invalid_reasons.append("unclassified_infrastructure_failures")
    error_categories = Counter(
        str(row.get("error_category") or "unspecified")
        for row in rows
        if row.get("status") != "completed"
    )
    known_usage = {"input_tokens": 0, "output_tokens": 0, "total_tokens": 0}
    successful_request_retries = 0
    for row in completed:
        agent = row.get("agent") or {}
        usage = agent.get("usage") or {}
        for key in known_usage:
            known_usage[key] += int(usage.get(key, 0) or 0)
        for timing in agent.get("request_timings") or []:
            successful_request_retries += max(int(timing.get("attempts", 1) or 1) - 1, 0)
    summary = {
        "protocol": "agent_per_workbook",
        "scorer": "cleanroom-corrected-value-v1",
        "style_checked": False,
        "calculation_backend": sorted({row.get("calculation_backend") for row in rows}),
        "attempted": len(rows),
        "expected": denominator,
        "missing": denominator - len(rows),
        "missing_task_ids": [task_id for task_id in expected if task_id not in latest],
        "completed": len(completed),
        "errors": len(rows) - len(completed),
        "error_categories": dict(sorted(error_categories.items())),
        "known_completed_usage_lower_bound": known_usage,
        "successful_request_retries": successful_request_retries,
        "invalid_result_rows_ignored": invalid_rows,
        "completion_rate": len(completed) / denominator if denominator else 0.0,
        "passed": passed,
        "verified_accuracy": primary_score,
        "completed_accuracy": completed_score if inference_valid else None,
        "soft": primary_score,
        "hard": primary_score,
        "inference_valid": inference_valid,
        "inference_invalid_reasons": inference_invalid_reasons,
        "no_score_infrastructure_failures": len(no_score_infrastructure),
        "no_score_infrastructure_task_ids": sorted(
            str(row.get("task_id")) for row in no_score_infrastructure
        ),
        "recalculation_infrastructure_failures": len(
            recalculation_infrastructure
        ),
        "scoring_infrastructure_failures": len(scoring_infrastructure),
        "known_scored_descriptive": {
            "tasks": len(completed),
            "passed": passed,
            "accuracy": completed_score if scores else None,
            "primary": False,
        },
    }
    if write_summary:
        _atomic_write_json(path.with_name("summary.json"), summary)
    return summary
