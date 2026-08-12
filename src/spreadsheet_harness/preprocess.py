"""Structured, text, and human-readable views of spreadsheet workbooks."""

from __future__ import annotations

import csv
import hashlib
import json
import math
import re
import tempfile
from collections import Counter
from collections.abc import Iterator, Mapping, Sequence
from contextlib import contextmanager
from copy import deepcopy
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl
import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles.numbers import is_date_format
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from .render import (
    SUPPORTED_SPREADSHEET_EXTENSIONS,
    convert_spreadsheet_copy,
    find_libreoffice,
    libreoffice_version,
    sha256_file,
)


class PreprocessError(RuntimeError):
    """Raised when a workbook cannot be converted into structured views."""


DETERMINISTIC_PROFILE_SCHEMA_VERSION = "deterministic-workbook-profile-v1"
DETERMINISTIC_PROFILE_BOUNDS: dict[str, int] = {
    "max_sheets": 12,
    "max_cells_per_sheet": 512,
    "max_regions_per_sheet": 4,
    "max_header_rows": 3,
    "max_columns_per_region": 10,
    "max_sample_rows_per_region": 4,
    "max_formula_clusters_per_sheet": 8,
    "max_provenance_cells_per_claim": 4,
    "max_scalar_chars": 128,
    "max_rendered_chars": 20_000,
}

_PROFILE_CELL_REFERENCE = re.compile(
    r"(?:(?:'[^']+'|[A-Za-z_][A-Za-z0-9_.]*)!)?\$?[A-Z]{1,3}\$?\d+"
)
_PROFILE_BRACKETED_UNIT = re.compile(r"(?:\(([^()]{1,16})\)|\[([^\[\]]{1,16})\])\s*$")


@dataclass(frozen=True)
class PreprocessResult:
    """Paths and in-memory data produced by :func:`preprocess_workbook`."""

    source: Path
    output_dir: Path
    json_path: Path
    yaml_path: Path
    markdown_path: Path
    source_sha256: str
    data: dict[str, Any]

    def to_dict(self) -> dict[str, Any]:
        return {
            "source": str(self.source),
            "source_sha256": self.source_sha256,
            "json_path": str(self.json_path),
            "yaml_path": str(self.yaml_path),
            "markdown_path": str(self.markdown_path),
        }


@dataclass
class _LoadedWorkbooks:
    formulas: Workbook
    values: Workbook | None
    backend: dict[str, Any]


def _validate_source(source: str | Path) -> Path:
    source_path = Path(source).expanduser().resolve()
    if not source_path.is_file():
        raise PreprocessError(f"Spreadsheet does not exist or is not a file: {source_path}")
    if source_path.suffix.lower() not in SUPPORTED_SPREADSHEET_EXTENSIONS:
        supported = ", ".join(sorted(SUPPORTED_SPREADSHEET_EXTENSIONS))
        raise PreprocessError(
            f"Unsupported spreadsheet format {source_path.suffix!r}; expected {supported}"
        )
    return source_path


def _safe_csv_sheet_name(stem: str) -> str:
    name = re.sub(r"[\\/*?:\[\]]", "_", stem).strip("'") or "CSV"
    return name[:31]


def _detect_csv_dialect(sample: str) -> csv.Dialect:
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        return csv.excel


def _load_csv(path: Path) -> Workbook:
    try:
        text = path.read_text(encoding="utf-8-sig")
    except UnicodeDecodeError:
        # Latin-1 is deliberately a lossless byte-to-text fallback.  The chosen
        # encoding is recorded in neither the source nor a rewritten CSV.
        text = path.read_text(encoding="latin-1")
    dialect = _detect_csv_dialect(text[:8192])
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = _safe_csv_sheet_name(path.stem)
    for row in csv.reader(text.splitlines(), dialect):
        worksheet.append(row)
    return workbook


@contextmanager
def _open_workbooks(
    source: Path,
    *,
    libreoffice_binary: str | Path | None,
    timeout_seconds: float,
) -> Iterator[_LoadedWorkbooks]:
    suffix = source.suffix.lower()
    formulas: Workbook | None = None
    values: Workbook | None = None
    temporary: tempfile.TemporaryDirectory[str] | None = None
    try:
        if suffix == ".csv":
            formulas = _load_csv(source)
            backend: dict[str, Any] = {
                "reader": "python-csv+openpyxl",
                "version": openpyxl.__version__,
                "conversion": None,
            }
        else:
            load_path = source
            conversion: dict[str, Any] | None = None
            if suffix in {".ods", ".xls"}:
                binary = find_libreoffice(libreoffice_binary)
                if binary is None:
                    raise PreprocessError(
                        f"LibreOffice is required to preprocess {suffix} workbooks"
                    )
                temporary = tempfile.TemporaryDirectory(prefix="spreadsheet-preprocess-")
                converted_dir = Path(temporary.name) / "converted"
                load_path = convert_spreadsheet_copy(
                    source,
                    converted_dir,
                    target_format="xlsx",
                    libreoffice_binary=binary,
                    timeout_seconds=timeout_seconds,
                )
                conversion = {
                    "backend": "libreoffice-headless",
                    "version": libreoffice_version(binary),
                    "from": suffix.lstrip("."),
                    "to": "xlsx",
                }

            keep_vba = suffix == ".xlsm"
            formulas = load_workbook(
                load_path,
                read_only=False,
                data_only=False,
                keep_vba=keep_vba,
                keep_links=True,
            )
            values = load_workbook(
                load_path,
                read_only=False,
                data_only=True,
                keep_vba=keep_vba,
                keep_links=True,
            )
            backend = {
                "reader": "openpyxl",
                "version": openpyxl.__version__,
                "conversion": conversion,
            }
        yield _LoadedWorkbooks(formulas=formulas, values=values, backend=backend)
    except PreprocessError:
        raise
    except Exception as exc:
        raise PreprocessError(f"Could not read {source.name}: {exc}") from exc
    finally:
        if values is not None:
            values.close()
        if formulas is not None:
            formulas.close()
        if temporary is not None:
            temporary.cleanup()


def _json_value(value: Any) -> Any:
    if value is None or isinstance(value, str | int | bool):
        return value
    if isinstance(value, float):
        return value if math.isfinite(value) else str(value)
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, datetime | date | time):
        return value.isoformat()
    if isinstance(value, timedelta):
        return value.total_seconds()
    if isinstance(value, bytes):
        return value.hex()
    if isinstance(value, list | tuple):
        return [_json_value(item) for item in value]
    if isinstance(value, dict):
        return {str(key): _json_value(item) for key, item in value.items()}
    return str(value)


def _color_summary(color: Any) -> dict[str, Any] | None:
    if color is None:
        return None
    fields = {
        "type": getattr(color, "type", None),
        "rgb": getattr(color, "rgb", None),
        "indexed": getattr(color, "indexed", None),
        "theme": getattr(color, "theme", None),
        "tint": getattr(color, "tint", None),
    }
    cleaned = {key: _json_value(value) for key, value in fields.items() if value is not None}
    return cleaned or None


def _side_summary(side: Any) -> dict[str, Any] | None:
    if side is None:
        return None
    style = getattr(side, "style", None)
    color = _color_summary(getattr(side, "color", None))
    if style is None and color is None:
        return None
    return {"style": style, "color": color}


def _style_summary(cell: Cell) -> dict[str, Any]:
    font = cell.font
    fill = cell.fill
    alignment = cell.alignment
    border = cell.border
    return {
        "style_id": int(cell.style_id),
        "number_format": cell.number_format,
        "font": {
            "name": font.name,
            "size": _json_value(font.sz),
            "bold": bool(font.bold),
            "italic": bool(font.italic),
            "underline": font.underline,
            "color": _color_summary(font.color),
        },
        "fill": {
            "type": fill.fill_type,
            "foreground": _color_summary(fill.fgColor),
            "background": _color_summary(fill.bgColor),
        },
        "alignment": {
            "horizontal": alignment.horizontal,
            "vertical": alignment.vertical,
            "wrap_text": alignment.wrap_text,
            "text_rotation": alignment.text_rotation,
            "shrink_to_fit": alignment.shrink_to_fit,
        },
        "border": {
            side: _side_summary(getattr(border, side))
            for side in ("left", "right", "top", "bottom")
        },
        "protection": {
            "locked": cell.protection.locked,
            "hidden": cell.protection.hidden,
        },
    }


def _worksheet_cells(worksheet: Worksheet) -> list[Cell]:
    # openpyxl has no public iterator over only materialized cells.  Iterating to
    # max_row/max_column can allocate millions of blank cells after a whole-column
    # format, so inspect the already-loaded cells and never mutate the worksheet.
    materialized = getattr(worksheet, "_cells", {})
    cells = [cell for cell in materialized.values() if isinstance(cell, Cell)]
    return sorted(cells, key=lambda cell: (cell.row, cell.column))


def _is_formula(cell: Cell) -> bool:
    return cell.data_type == "f" or (isinstance(cell.value, str) and cell.value.startswith("="))


def _is_meaningful(cell: Cell) -> bool:
    return cell.value is not None or _is_formula(cell)


def _cell_record(cell: Cell, value_worksheet: Worksheet | None) -> dict[str, Any]:
    formula = str(cell.value) if _is_formula(cell) else None
    if formula is not None and value_worksheet is not None:
        value = value_worksheet[cell.coordinate].value
    elif formula is not None:
        value = None
    else:
        value = cell.value
    return {
        "coordinate": cell.coordinate,
        "row": cell.row,
        "column": cell.column,
        "formula": formula,
        "value": _json_value(value),
        "data_type": cell.data_type,
        "style_id": int(cell.style_id),
        "number_format": cell.number_format,
    }


def _bounds_from_ref(reference: str) -> tuple[int, int, int, int]:
    try:
        return range_boundaries(reference)
    except (TypeError, ValueError) as exc:
        raise PreprocessError(f"Invalid cell range: {reference!r}") from exc


def _used_bounds(
    meaningful_cells: Sequence[Cell], merged_ranges: Sequence[str], table_ranges: Sequence[str]
) -> tuple[int, int, int, int] | None:
    columns: list[int] = [cell.column for cell in meaningful_cells]
    rows: list[int] = [cell.row for cell in meaningful_cells]
    for reference in [*merged_ranges, *table_ranges]:
        min_column, min_row, max_column, max_row = _bounds_from_ref(reference)
        columns.extend((min_column, max_column))
        rows.extend((min_row, max_row))
    if not rows or not columns:
        return None
    return min(columns), min(rows), max(columns), max(rows)


def _bounds_to_ref(bounds: tuple[int, int, int, int] | None) -> str | None:
    if bounds is None:
        return None
    min_column, min_row, max_column, max_row = bounds
    return f"{get_column_letter(min_column)}{min_row}:{get_column_letter(max_column)}{max_row}"


def _table_summaries(worksheet: Worksheet) -> list[dict[str, Any]]:
    summaries: list[dict[str, Any]] = []
    for table in worksheet.tables.values():
        table_style = table.tableStyleInfo
        summaries.append(
            {
                "name": table.name,
                "display_name": table.displayName,
                "range": table.ref,
                "totals_row_shown": table.totalsRowShown,
                "style": None
                if table_style is None
                else {
                    "name": table_style.name,
                    "show_first_column": table_style.showFirstColumn,
                    "show_last_column": table_style.showLastColumn,
                    "show_row_stripes": table_style.showRowStripes,
                    "show_column_stripes": table_style.showColumnStripes,
                },
            }
        )
    return summaries


def _style_summaries(cells: Sequence[Cell]) -> list[dict[str, Any]]:
    counts: Counter[int] = Counter()
    first_cell: dict[int, Cell] = {}
    samples: dict[int, list[str]] = {}
    for cell in cells:
        style_id = int(cell.style_id)
        counts[style_id] += 1
        first_cell.setdefault(style_id, cell)
        sample = samples.setdefault(style_id, [])
        if len(sample) < 5:
            sample.append(cell.coordinate)

    summaries: list[dict[str, Any]] = []
    for style_id in sorted(counts):
        summary = _style_summary(first_cell[style_id])
        summary["cell_count"] = counts[style_id]
        summary["sample_cells"] = samples[style_id]
        summaries.append(summary)
    return summaries


def _requested_ranges(
    sheet_name: str,
    used_range: str | None,
    requested: Mapping[str, Sequence[str] | str] | None,
) -> list[str]:
    if requested is None or sheet_name not in requested:
        return [used_range] if used_range is not None else []
    raw = requested[sheet_name]
    if isinstance(raw, str):
        ranges = [raw]
    else:
        ranges = list(raw)
    return [str(reference) for reference in ranges]


def _region_summary(
    reference: str,
    meaningful_cells: Sequence[Cell],
    value_worksheet: Worksheet | None,
    *,
    max_cells: int,
) -> dict[str, Any]:
    min_column, min_row, max_column, max_row = _bounds_from_ref(reference)
    matching = [
        cell
        for cell in meaningful_cells
        if min_row <= cell.row <= max_row and min_column <= cell.column <= max_column
    ]
    selected = matching[:max_cells]
    return {
        "range": reference,
        "min_row": min_row,
        "max_row": max_row,
        "min_column": min_column,
        "max_column": max_column,
        "row_count": max_row - min_row + 1,
        "column_count": max_column - min_column + 1,
        "nonempty_cell_count": len(matching),
        "truncated": len(selected) < len(matching),
        "cells": [_cell_record(cell, value_worksheet) for cell in selected],
    }


def _worksheet_view(
    worksheet: Worksheet,
    value_worksheet: Worksheet | None,
    *,
    sheet_index: int,
    requested_regions: Mapping[str, Sequence[str] | str] | None,
    max_cells_per_sheet: int,
) -> tuple[dict[str, Any], dict[str, Any]]:
    cells = _worksheet_cells(worksheet)
    meaningful_cells = [cell for cell in cells if _is_meaningful(cell)]
    merged_ranges = [str(reference) for reference in worksheet.merged_cells.ranges]
    tables = _table_summaries(worksheet)
    table_ranges = [str(table["range"]) for table in tables]
    used_bounds = _used_bounds(meaningful_cells, merged_ranges, table_ranges)
    used_range = _bounds_to_ref(used_bounds)
    formula_count = sum(_is_formula(cell) for cell in meaningful_cells)
    detailed_cells = meaningful_cells[:max_cells_per_sheet]
    ranges = _requested_ranges(worksheet.title, used_range, requested_regions)
    regions = [
        _region_summary(
            reference,
            meaningful_cells,
            value_worksheet,
            max_cells=max_cells_per_sheet,
        )
        for reference in ranges
    ]

    style_cells = [cell for cell in cells if cell.has_style or _is_meaningful(cell)]
    sheet_view = {
        "name": worksheet.title,
        "index": sheet_index,
        "state": worksheet.sheet_state,
        "used_region": used_range,
        "worksheet_dimension": worksheet.calculate_dimension(),
        "cells": [_cell_record(cell, value_worksheet) for cell in detailed_cells],
        "cells_truncated": len(detailed_cells) < len(meaningful_cells),
        "regions": regions,
        "merged_ranges": merged_ranges,
        "tables": tables,
        "styles": _style_summaries(style_cells),
    }
    inventory = {
        "name": worksheet.title,
        "index": sheet_index,
        "state": worksheet.sheet_state,
        "used_region": used_range,
        "max_row": worksheet.max_row,
        "max_column": worksheet.max_column,
        "nonempty_cell_count": len(meaningful_cells),
        "formula_count": formula_count,
        "value_count": len(meaningful_cells) - formula_count,
        "merged_range_count": len(merged_ranges),
        "table_count": len(tables),
        "style_count": len(sheet_view["styles"]),
        "cells_truncated": sheet_view["cells_truncated"],
    }
    return sheet_view, inventory


def build_preprocess_view(
    source: str | Path,
    *,
    regions: Mapping[str, Sequence[str] | str] | None = None,
    max_cells_per_sheet: int = 5000,
    libreoffice_binary: str | Path | None = None,
    timeout_seconds: float = 120.0,
) -> dict[str, Any]:
    """Build a serializable workbook view without writing output artifacts."""

    source_path = _validate_source(source)
    if max_cells_per_sheet <= 0:
        raise PreprocessError("max_cells_per_sheet must be positive")
    if timeout_seconds <= 0:
        raise PreprocessError("timeout_seconds must be positive")
    source_hash = sha256_file(source_path)

    with _open_workbooks(
        source_path,
        libreoffice_binary=libreoffice_binary,
        timeout_seconds=timeout_seconds,
    ) as loaded:
        unknown_sheets = set(regions or {}).difference(loaded.formulas.sheetnames)
        if unknown_sheets:
            names = ", ".join(sorted(unknown_sheets))
            raise PreprocessError(f"Requested regions refer to unknown sheets: {names}")

        sheet_views: list[dict[str, Any]] = []
        inventory: list[dict[str, Any]] = []
        for sheet_index, worksheet in enumerate(loaded.formulas.worksheets, start=1):
            value_worksheet = None
            if loaded.values is not None and worksheet.title in loaded.values.sheetnames:
                value_worksheet = loaded.values[worksheet.title]
            sheet_view, sheet_inventory = _worksheet_view(
                worksheet,
                value_worksheet,
                sheet_index=sheet_index,
                requested_regions=regions,
                max_cells_per_sheet=max_cells_per_sheet,
            )
            sheet_views.append(sheet_view)
            inventory.append(sheet_inventory)

        return {
            "schema_version": 1,
            "source": {
                "name": source_path.name,
                "format": source_path.suffix.lower().lstrip("."),
                "sha256": source_hash,
            },
            "backend": loaded.backend,
            "inventory": {
                "sheet_count": len(inventory),
                "sheets": inventory,
            },
            "sheets": sheet_views,
        }


def _profile_value_kind(cell: Mapping[str, Any]) -> str:
    if cell.get("formula") is not None:
        return "formula"
    value = cell.get("value")
    if value is None:
        return "blank"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, int | float):
        return "date_or_time" if is_date_format(str(cell.get("number_format") or "")) else "number"
    if isinstance(value, str):
        return "date_or_time" if is_date_format(str(cell.get("number_format") or "")) else "text"
    return "other"


def _bounded_profile_scalar(value: Any, *, max_chars: int) -> dict[str, Any]:
    normalized = _json_value(value)
    text = json.dumps(normalized, ensure_ascii=False, separators=(",", ":"))
    if len(text) <= max_chars:
        return {"value": normalized, "truncated": False}
    return {
        "value": text[:max_chars],
        "truncated": True,
        "original_chars": len(text),
        "sha256": hashlib.sha256(text.encode("utf-8")).hexdigest(),
    }


def _profile_header_rows(cells: Sequence[Mapping[str, Any]], *, maximum: int) -> int:
    by_row: dict[int, list[Mapping[str, Any]]] = {}
    for cell in cells:
        by_row.setdefault(int(cell["row"]), []).append(cell)
    if not by_row:
        return 0
    first_row = min(by_row)
    count = 0
    for row_number in range(first_row, first_row + maximum):
        row = by_row.get(row_number, [])
        if not row:
            break
        non_formula = [cell for cell in row if cell.get("formula") is None]
        text = sum(_profile_value_kind(cell) == "text" for cell in non_formula)
        numeric = sum(_profile_value_kind(cell) in {"number", "date_or_time"} for cell in non_formula)
        if text == 0 or numeric > text:
            break
        count += 1
    return count


def _profile_unit_hints(
    cells: Sequence[Mapping[str, Any]], *, header_rows: int, maximum: int
) -> list[dict[str, Any]]:
    if not cells:
        return []
    first_row = min(int(cell["row"]) for cell in cells)
    hints: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    for cell in cells:
        value = cell.get("value")
        if header_rows and int(cell["row"]) < first_row + header_rows and isinstance(value, str):
            match = _PROFILE_BRACKETED_UNIT.search(value.strip())
            token = next((item for item in match.groups() if item), None) if match else None
            if token is not None and (token, str(cell["coordinate"])) not in seen:
                seen.add((token, str(cell["coordinate"])))
                hints.append(
                    {
                        "unit": token,
                        "confidence": "heuristic",
                        "provenance": {"cell": cell["coordinate"], "method": "header-suffix"},
                    }
                )
        number_format = str(cell.get("number_format") or "")
        format_unit = (
            "percent"
            if "%" in number_format
            else "currency"
            if any(symbol in number_format for symbol in ("$", "€", "£", "¥"))
            else None
        )
        key = (str(format_unit), number_format)
        if format_unit is not None and key not in seen:
            seen.add(key)
            hints.append(
                {
                    "unit": format_unit,
                    "confidence": "format-derived",
                    "provenance": {
                        "cell": cell["coordinate"],
                        "number_format": number_format,
                        "method": "number-format",
                    },
                }
            )
        if len(hints) >= maximum:
            break
    return hints


def _contiguous_profile_regions(
    cells: Sequence[Mapping[str, Any]], *, maximum: int
) -> tuple[list[tuple[int, int, int, int]], bool]:
    occupied = {(int(cell["row"]), int(cell["column"])) for cell in cells}
    regions: list[tuple[int, int, int, int]] = []
    while occupied:
        pending = [min(occupied)]
        component: set[tuple[int, int]] = set()
        while pending:
            current = pending.pop()
            if current not in occupied:
                continue
            occupied.remove(current)
            component.add(current)
            row, column = current
            pending.extend(
                candidate
                for candidate in (
                    (row - 1, column),
                    (row + 1, column),
                    (row, column - 1),
                    (row, column + 1),
                )
                if candidate in occupied
            )
        rows = [row for row, _ in component]
        columns = [column for _, column in component]
        regions.append((min(columns), min(rows), max(columns), max(rows)))
    regions.sort(key=lambda item: (item[1], item[0], item[3], item[2]))
    return regions[:maximum], len(regions) > maximum


def _profile_formula_clusters(
    cells: Sequence[Mapping[str, Any]], *, maximum: int, provenance_limit: int
) -> tuple[list[dict[str, Any]], bool]:
    clusters: dict[tuple[str, ...], list[Mapping[str, Any]]] = {}
    for cell in cells:
        formula = cell.get("formula")
        if not isinstance(formula, str):
            continue
        references = tuple(
            dict.fromkeys(reference.replace("$", "") for reference in _PROFILE_CELL_REFERENCE.findall(formula))
        )
        clusters.setdefault(references, []).append(cell)
    ordered = sorted(
        clusters.items(),
        key=lambda item: (int(item[1][0]["row"]), int(item[1][0]["column"]), item[0]),
    )
    result: list[dict[str, Any]] = []
    for references, members in ordered[:maximum]:
        coordinates = [str(cell["coordinate"]) for cell in members]
        result.append(
            {
                "cells": coordinates[:provenance_limit],
                "cell_count": len(coordinates),
                "references": list(references[:provenance_limit]),
                "confidence": "high" if references else "medium",
                "provenance": {
                    "method": "openpyxl-formula-token-pattern",
                    "cells": coordinates[:provenance_limit],
                    "truncated": len(coordinates) > provenance_limit,
                },
            }
        )
    return result, len(ordered) > maximum


def _profile_sheet(
    sheet: Mapping[str, Any],
    inventory: Mapping[str, Any],
    *,
    bounds: Mapping[str, int],
) -> dict[str, Any]:
    cells = list(sheet.get("cells", []))
    region_bounds, regions_truncated = _contiguous_profile_regions(
        cells, maximum=bounds["max_regions_per_sheet"]
    )
    regions: list[dict[str, Any]] = []
    for min_column, min_row, max_column, max_row in region_bounds:
        members = [
            cell
            for cell in cells
            if min_row <= int(cell["row"]) <= max_row
            and min_column <= int(cell["column"]) <= max_column
        ]
        header_rows = _profile_header_rows(members, maximum=bounds["max_header_rows"])
        kind_counts = Counter(_profile_value_kind(cell) for cell in members)
        number_formats = Counter(
            str(cell.get("number_format"))
            for cell in members
            if cell.get("number_format") not in {None, "", "General"}
        )
        sample: list[dict[str, Any]] = []
        sample_limit = (
            bounds["max_columns_per_region"] * bounds["max_sample_rows_per_region"]
        )
        for cell in members[:sample_limit]:
            bounded = _bounded_profile_scalar(
                cell.get("formula") if cell.get("formula") is not None else cell.get("value"),
                max_chars=bounds["max_scalar_chars"],
            )
            sample.append(
                {
                    "cell": cell["coordinate"],
                    "kind": _profile_value_kind(cell),
                    **bounded,
                }
            )
        regions.append(
            {
                "range": _bounds_to_ref((min_column, min_row, max_column, max_row)),
                "header_rows": header_rows,
                "data_start_row": min_row + header_rows if header_rows else min_row,
                "row_count": max_row - min_row + 1,
                "column_count": max_column - min_column + 1,
                "type_counts": dict(sorted(kind_counts.items())),
                "number_formats": dict(sorted(number_formats.items())),
                "unit_hints": _profile_unit_hints(
                    members,
                    header_rows=header_rows,
                    maximum=bounds["max_provenance_cells_per_claim"],
                ),
                "sample": sample,
                "confidence": "medium" if header_rows else "high",
                "provenance": {
                    "method": "deterministic-four-neighbor-components",
                    "sheet": sheet["name"],
                    "range": _bounds_to_ref((min_column, min_row, max_column, max_row)),
                    "sample_cells": [item["cell"] for item in sample][
                        : bounds["max_provenance_cells_per_claim"]
                    ],
                },
            }
        )
    formula_clusters, formulas_truncated = _profile_formula_clusters(
        cells,
        maximum=bounds["max_formula_clusters_per_sheet"],
        provenance_limit=bounds["max_provenance_cells_per_claim"],
    )
    return {
        "name": sheet["name"],
        "state": sheet["state"],
        "used_region": sheet["used_region"],
        "counts": {
            "nonempty_cells": inventory["nonempty_cell_count"],
            "formulas": inventory["formula_count"],
            "merges": inventory["merged_range_count"],
            "tables": inventory["table_count"],
        },
        "regions": regions,
        "formula_clusters": formula_clusters,
        "merges": list(sheet.get("merged_ranges", []))[
            : bounds["max_provenance_cells_per_claim"]
        ],
        "tables": [
            {"name": table.get("name"), "range": table.get("range")}
            for table in list(sheet.get("tables", []))[
                : bounds["max_provenance_cells_per_claim"]
            ]
        ],
        "confidence": {
            "inventory": "high",
            "regions": "heuristic",
            "header_rows": "heuristic",
            "formula_dependencies": "syntactic",
        },
        "provenance": {
            "method": "openpyxl-read-only-profile",
            "sheet": sheet["name"],
            "range": sheet["used_region"],
        },
        "truncation": {
            "cells": bool(sheet.get("cells_truncated")),
            "regions": regions_truncated,
            "formula_clusters": formulas_truncated,
            "merges": len(sheet.get("merged_ranges", []))
            > bounds["max_provenance_cells_per_claim"],
            "tables": len(sheet.get("tables", []))
            > bounds["max_provenance_cells_per_claim"],
        },
    }


def build_deterministic_profile(
    source: str | Path,
    *,
    bounds: Mapping[str, int] | None = None,
    libreoffice_binary: str | Path | None = None,
    timeout_seconds: float = 120.0,
) -> dict[str, Any]:
    """Build task-independent, bounded structural evidence for an agent prompt."""

    effective = deepcopy(DETERMINISTIC_PROFILE_BOUNDS)
    if bounds is not None:
        unknown = set(bounds).difference(effective)
        if unknown:
            raise PreprocessError(f"Unknown deterministic profile bounds: {sorted(unknown)}")
        effective.update({key: int(value) for key, value in bounds.items()})
    if any(value <= 0 for value in effective.values()):
        raise PreprocessError("Deterministic profile bounds must be positive")
    view = build_preprocess_view(
        source,
        max_cells_per_sheet=effective["max_cells_per_sheet"],
        libreoffice_binary=libreoffice_binary,
        timeout_seconds=timeout_seconds,
    )
    sheet_limit = effective["max_sheets"]
    inventories = view["inventory"]["sheets"][:sheet_limit]
    sheets = view["sheets"][:sheet_limit]
    profile = {
        "schema_version": DETERMINISTIC_PROFILE_SCHEMA_VERSION,
        "bounds": effective,
        "source": {
            "format": view["source"]["format"],
            "sha256": view["source"]["sha256"],
        },
        "backend": view["backend"],
        "task_independent": True,
        "sheets": [
            _profile_sheet(sheet, inventory, bounds=effective)
            for sheet, inventory in zip(sheets, inventories, strict=True)
        ],
        "truncation": {
            "sheets": len(view["sheets"]) > sheet_limit,
            "rendered": False,
        },
    }
    profile["profile_sha256"] = hashlib.sha256(
        json.dumps(profile, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode(
            "utf-8"
        )
    ).hexdigest()
    return profile


def render_deterministic_profile(profile: Mapping[str, Any]) -> str:
    """Serialize profile evidence as bounded JSON with an auditable truncation marker."""

    maximum = int(profile["bounds"]["max_rendered_chars"])
    rendered = json.dumps(profile, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    rendered = rendered.replace("<", "\\u003c").replace(">", "\\u003e")
    if len(rendered) <= maximum:
        return rendered
    digest = hashlib.sha256(rendered.encode("utf-8")).hexdigest()
    compact = deepcopy(dict(profile))
    for sheet in compact.get("sheets", []):
        sheet["regions"] = [
            {key: value for key, value in region.items() if key != "sample"}
            for region in sheet.get("regions", [])
        ]
    compact["truncation"]["rendered"] = True
    compact["truncation"]["unabridged_chars"] = len(rendered)
    compact["truncation"]["unabridged_sha256"] = digest
    rendered = json.dumps(compact, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    rendered = rendered.replace("<", "\\u003c").replace(">", "\\u003e")
    if len(rendered) > maximum:
        raise PreprocessError(
            f"Deterministic profile metadata exceeds max_rendered_chars={maximum}"
        )
    return rendered


def _markdown_escape(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, dict | list):
        text = json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    else:
        text = str(value)
    return text.replace("|", "\\|").replace("\r", " ").replace("\n", "<br>")


def _display_cell(cell: Mapping[str, Any] | None) -> str:
    if cell is None:
        return ""
    formula = cell.get("formula")
    value = cell.get("value")
    if formula:
        rendered_formula = _markdown_escape(formula)
        if value is not None:
            return f"`{rendered_formula}` → {_markdown_escape(value)}"
        return f"`{rendered_formula}`"
    return _markdown_escape(value)


def _region_markdown(region: Mapping[str, Any]) -> list[str]:
    lines = [f"#### Region `{_markdown_escape(region['range'])}`", ""]
    min_column = int(region["min_column"])
    max_column = int(region["max_column"])
    min_row = int(region["min_row"])
    max_row = int(region["max_row"])
    visible_max_column = min(max_column, min_column + 19)
    visible_max_row = min(max_row, min_row + 29)
    cells = {str(cell["coordinate"]): cell for cell in region["cells"]}
    headers = [get_column_letter(column) for column in range(min_column, visible_max_column + 1)]
    lines.append("| Row | " + " | ".join(headers) + " |")
    lines.append("| ---: | " + " | ".join("---" for _ in headers) + " |")
    for row in range(min_row, visible_max_row + 1):
        row_values = [
            _display_cell(cells.get(f"{get_column_letter(column)}{row}"))
            for column in range(min_column, visible_max_column + 1)
        ]
        lines.append(f"| {row} | " + " | ".join(row_values) + " |")
    if visible_max_column < max_column or visible_max_row < max_row or region.get("truncated"):
        lines.extend(["", "_Region display truncated; use the JSON/YAML view for metadata._"])
    lines.append("")
    return lines


def render_markdown(data: Mapping[str, Any]) -> str:
    """Render the structured preprocessing payload as a concise Markdown view."""

    source = data["source"]
    inventory = data["inventory"]
    lines = [
        f"# Workbook view: {_markdown_escape(source['name'])}",
        "",
        f"- Format: `{_markdown_escape(source['format'])}`",
        f"- SHA-256: `{_markdown_escape(source['sha256'])}`",
        f"- Sheets: {inventory['sheet_count']}",
        "",
        "## Inventory",
        "",
        "| # | Sheet | State | Used region | Values | Formulas | Merged | Tables | Styles |",
        "| ---: | --- | --- | --- | ---: | ---: | ---: | ---: | ---: |",
    ]
    for sheet in inventory["sheets"]:
        lines.append(
            "| {index} | {name} | {state} | {region} | {values} | {formulas} | "
            "{merged} | {tables} | {styles} |".format(
                index=sheet["index"],
                name=_markdown_escape(sheet["name"]),
                state=_markdown_escape(sheet["state"]),
                region=_markdown_escape(sheet["used_region"]),
                values=sheet["value_count"],
                formulas=sheet["formula_count"],
                merged=sheet["merged_range_count"],
                tables=sheet["table_count"],
                styles=sheet["style_count"],
            )
        )

    for sheet in data["sheets"]:
        lines.extend(
            [
                "",
                f"## Sheet: {_markdown_escape(sheet['name'])}",
                "",
                f"- State: `{_markdown_escape(sheet['state'])}`",
                f"- Used region: `{_markdown_escape(sheet['used_region'])}`",
                f"- Worksheet dimension: `{_markdown_escape(sheet['worksheet_dimension'])}`",
                "",
                "### Formulas and values",
                "",
                "| Cell | Formula | Value | Type | Style | Number format |",
                "| --- | --- | --- | --- | ---: | --- |",
            ]
        )
        for cell in sheet["cells"]:
            lines.append(
                "| {coordinate} | {formula} | {value} | {data_type} | {style_id} | "
                "{number_format} |".format(
                    coordinate=_markdown_escape(cell["coordinate"]),
                    formula=_markdown_escape(cell["formula"]),
                    value=_markdown_escape(cell["value"]),
                    data_type=_markdown_escape(cell["data_type"]),
                    style_id=cell["style_id"],
                    number_format=_markdown_escape(cell["number_format"]),
                )
            )
        if sheet["cells_truncated"]:
            lines.extend(["", "_Cell listing truncated by `max_cells_per_sheet`._"])

        lines.extend(["", "### Merged ranges", ""])
        if sheet["merged_ranges"]:
            lines.append(", ".join(f"`{item}`" for item in sheet["merged_ranges"]))
        else:
            lines.append("None")

        lines.extend(["", "### Tables", ""])
        if sheet["tables"]:
            lines.extend(
                f"- `{_markdown_escape(table['name'])}`: `{_markdown_escape(table['range'])}`"
                for table in sheet["tables"]
            )
        else:
            lines.append("None")

        lines.extend(
            [
                "",
                "### Style summary",
                "",
                "| Style ID | Cells | Samples | Number format | Fill |",
                "| ---: | ---: | --- | --- | --- |",
            ]
        )
        for style in sheet["styles"]:
            lines.append(
                "| {style_id} | {count} | {samples} | {number_format} | {fill} |".format(
                    style_id=style["style_id"],
                    count=style["cell_count"],
                    samples=_markdown_escape(", ".join(style["sample_cells"])),
                    number_format=_markdown_escape(style["number_format"]),
                    fill=_markdown_escape(style["fill"]),
                )
            )

        lines.extend(["", "### Regions", ""])
        if not sheet["regions"]:
            lines.append("No populated region.")
        for region in sheet["regions"]:
            lines.extend(_region_markdown(region))

    return "\n".join(lines).rstrip() + "\n"


def _write_text_atomic(path: Path, content: str) -> None:
    temporary = path.with_name(f".{path.name}.tmp")
    temporary.write_text(content, encoding="utf-8")
    temporary.replace(path)


def preprocess_workbook(
    source: str | Path,
    output_dir: str | Path,
    *,
    regions: Mapping[str, Sequence[str] | str] | None = None,
    max_cells_per_sheet: int = 5000,
    libreoffice_binary: str | Path | None = None,
    timeout_seconds: float = 120.0,
) -> PreprocessResult:
    """Write equivalent JSON, YAML, and Markdown views for a spreadsheet."""

    source_path = _validate_source(source)
    source_hash = sha256_file(source_path)
    data = build_preprocess_view(
        source_path,
        regions=regions,
        max_cells_per_sheet=max_cells_per_sheet,
        libreoffice_binary=libreoffice_binary,
        timeout_seconds=timeout_seconds,
    )
    if sha256_file(source_path) != source_hash:
        raise PreprocessError("Source workbook changed while it was being preprocessed")

    destination = Path(output_dir).expanduser().resolve()
    destination.mkdir(parents=True, exist_ok=True)
    json_path = destination / "preprocess.json"
    yaml_path = destination / "preprocess.yaml"
    markdown_path = destination / "preprocess.md"
    for path in (json_path, yaml_path, markdown_path):
        if path.exists():
            raise PreprocessError(f"Refusing to overwrite preprocessing artifact: {path}")
        if path == source_path:
            raise PreprocessError("Preprocessing output would overwrite the source workbook")

    _write_text_atomic(
        json_path,
        json.dumps(data, ensure_ascii=False, indent=2, sort_keys=False) + "\n",
    )
    _write_text_atomic(
        yaml_path,
        yaml.safe_dump(data, allow_unicode=True, sort_keys=False),
    )
    _write_text_atomic(markdown_path, render_markdown(data))
    return PreprocessResult(
        source=source_path,
        output_dir=destination,
        json_path=json_path,
        yaml_path=yaml_path,
        markdown_path=markdown_path,
        source_sha256=source_hash,
        data=data,
    )


preprocess = preprocess_workbook


__all__ = [
    "DETERMINISTIC_PROFILE_BOUNDS",
    "DETERMINISTIC_PROFILE_SCHEMA_VERSION",
    "PreprocessError",
    "PreprocessResult",
    "build_deterministic_profile",
    "build_preprocess_view",
    "preprocess",
    "preprocess_workbook",
    "render_deterministic_profile",
    "render_markdown",
]
